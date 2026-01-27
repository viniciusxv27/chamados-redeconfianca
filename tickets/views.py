from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.utils import timezone
from django.db import models
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.conf import settings
from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from .models import Ticket, Category, TicketLog, TicketComment, Webhook, TicketView, TicketAssignment
from .serializers import TicketSerializer, CategorySerializer, TicketLogSerializer, TicketCommentSerializer, WebhookSerializer
from users.models import Sector, User
from core.middleware import log_action
import csv
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


@login_required
def tickets_list_view(request):
    user = request.user
    
    # Filtros de pesquisa
    search = request.GET.get('search', '')
    status_filter = request.GET.getlist('status')  # Mudado para getlist para m√∫ltiplos valores
    origem_filter = request.GET.get('origem', '')
    categoria_filter = request.GET.get('categoria', '')
    setor_filter = request.GET.get('setor', '')
    prioridade_filter = request.GET.get('prioridade', '')
    carteira_filter = request.GET.get('carteira', '')
    atribuidos_filter = request.GET.get('atribuidos', '')  # Novo filtro para chamados atribu√≠dos
    
    # Filtros avan√ßados para SUPERADMIN - definir logo no in√≠cio
    created_by_filter = request.GET.get('created_by', '')
    created_by_sector_filter = request.GET.get('created_by_sector', '')
    assigned_to_filter = request.GET.get('assigned_to', '')
    date_from_filter = request.GET.get('date_from', '')
    date_to_filter = request.GET.get('date_to', '')
    user_hierarchy_filter = request.GET.get('user_hierarchy', '')
    has_attachments_filter = request.GET.get('has_attachments', '')
    has_comments_filter = request.GET.get('has_comments', '')
    overdue_filter = request.GET.get('overdue', '')
    duplicates_filter = request.GET.get('duplicates', '')
    
    # Filtro base: TODOS os usu√°rios sempre veem seus pr√≥prios chamados
    base_filter = models.Q(created_by=user)
    
    # Filtrar tickets baseado na hierarquia do usu√°rio
    if user.can_view_all_tickets():
        # Admin v√™ todos os tickets (incluindo fechados)
        tickets = Ticket.objects.all()
    elif user.can_view_sector_tickets():
        # Supervisores veem: seus pr√≥prios tickets + TODOS os tickets dos setores (independente de atribui√ß√£o) + tickets atribu√≠dos
        user_sectors = list(user.sectors.all())
        if user.sector:
            user_sectors.append(user.sector)
        
        tickets = Ticket.objects.filter(
            base_filter |  # Sempre inclui pr√≥prios tickets
            models.Q(sector__in=user_sectors) |  # TODOS os tickets dos setores
            models.Q(assigned_to=user) |  # Tickets atribu√≠dos diretamente a mim
            models.Q(additional_assignments__user=user, additional_assignments__is_active=True)  # Atribui√ß√µes adicionais
        ).distinct()
    else:
        # Usu√°rios comuns veem: seus pr√≥prios tickets + tickets do setor (sem atribui√ß√£o espec√≠fica) + tickets atribu√≠dos
        # Excluindo tickets fechados
        tickets = Ticket.objects.filter(
            base_filter |  # Sempre inclui pr√≥prios tickets
            models.Q(sector=user.sector, assigned_to__isnull=True) |  # Tickets do setor SEM atribui√ß√£o espec√≠fica
            models.Q(assigned_to=user) |  # Tickets atribu√≠dos diretamente a mim
            models.Q(additional_assignments__user=user, additional_assignments__is_active=True)  # Atribui√ß√µes adicionais
        ).exclude(status='FECHADO').distinct()
    
    # Aplicar filtros adicionais
    
    # Filtro por origem
    if origem_filter == 'meus':
        tickets = tickets.filter(created_by=user)
    elif origem_filter == 'setor':
        # Tickets dos setores do usu√°rio (excluindo os pr√≥prios)
        user_sectors = list(user.sectors.all())
        if user.sector:
            user_sectors.append(user.sector)
        tickets = tickets.filter(sector__in=user_sectors).exclude(created_by=user)
    
    # Filtro por chamados atribu√≠dos (atribu√≠do a mim)
    if atribuidos_filter == 'sim':
        tickets = tickets.filter(
            models.Q(assigned_to=user) |
            models.Q(additional_assignments__user=user, additional_assignments__is_active=True)
        ).distinct()
    
    # Filtro por status - suporte para m√∫ltiplos valores
    if status_filter:
        if 'abertos' in status_filter:
            tickets = tickets.filter(status='ABERTO')
        elif 'nao_resolvidos' in status_filter:
            tickets = tickets.exclude(status__in=['RESOLVIDO', 'FECHADO'])
        else:
            # Filtrar por m√∫ltiplos status espec√≠ficos
            valid_statuses = [s for s in status_filter if s in ['ABERTO', 'EM_ANDAMENTO', 'RESOLVIDO', 'FECHADO']]
            if valid_statuses:
                tickets = tickets.filter(status__in=valid_statuses)
    
    # Filtro por categoria - SUPERADMINs podem filtrar por qualquer categoria
    if categoria_filter:
        if user.hierarchy == 'SUPERADMIN' or user.can_view_all_tickets():
            # SUPERADMIN pode filtrar por qualquer categoria
            tickets = tickets.filter(category_id=categoria_filter)
        else:
            # Outros usu√°rios s√≥ podem filtrar pelas categorias que t√™m acesso
            tickets = tickets.filter(category_id=categoria_filter)
    
    # Filtro por setor - SUPERADMINs podem filtrar por qualquer setor
    if setor_filter:
        if user.hierarchy == 'SUPERADMIN' or user.can_view_all_tickets():
            # SUPERADMIN pode filtrar por qualquer setor
            tickets = tickets.filter(sector_id=setor_filter)
        else:
            # Outros usu√°rios s√≥ podem filtrar pelos setores que t√™m acesso
            user_sectors = list(user.sectors.all())
            if user.sector:
                user_sectors.append(user.sector)
            sector_ids = [s.id for s in user_sectors] if user_sectors else []
            if int(setor_filter) in sector_ids:
                tickets = tickets.filter(sector_id=setor_filter)
    
    # Filtro por prioridade
    if prioridade_filter:
        tickets = tickets.filter(priority=prioridade_filter)
    
    # Filtro por carteira espec√≠fica (chamados direcionados PARA setores da carteira selecionada)
    if carteira_filter:
        from communications.models import CommunicationGroup
        try:
            # Buscar o grupo de carteira espec√≠fico pelo ID
            carteira_group = CommunicationGroup.objects.get(id=carteira_filter)
            
            # Obter usu√°rios deste grupo de carteira
            carteira_users = carteira_group.members.filter(is_active=True)
            
            # Obter setores onde os usu√°rios da carteira trabalham (corrigir relacionamentos)
            carteira_sectors = Sector.objects.filter(
                models.Q(users__in=carteira_users) |  # Setores ManyToMany
                models.Q(primary_users__in=carteira_users)  # Setor principal (ForeignKey)
            ).distinct()
            
            print(f"DEBUG: Filtro carteira ativo: Grupo '{carteira_group.name}' (ID: {carteira_filter})")
            print(f"DEBUG: Usu√°rios da carteira ({carteira_users.count()}): {[u.username for u in carteira_users]}")
            print(f"DEBUG: Setores da carteira ({carteira_sectors.count()}): {[s.name for s in carteira_sectors]}")
            
            # Contar tickets antes do filtro
            tickets_antes = tickets.count()
            print(f"DEBUG: Tickets antes do filtro carteira: {tickets_antes}")
            
            # Filtrar chamados que foram direcionados PARA os setores da carteira
            # OU que foram atribu√≠dos a usu√°rios da carteira
            tickets = tickets.filter(
                models.Q(sector__in=carteira_sectors) |  # Chamados para setores da carteira
                models.Q(assigned_to__in=carteira_users) |  # Chamados atribu√≠dos a usu√°rios da carteira
                models.Q(additional_assignments__user__in=carteira_users, additional_assignments__is_active=True)  # Atribui√ß√µes adicionais
            ).distinct()
            
            print(f"DEBUG: Tickets ap√≥s filtro carteira: {tickets.count()}")
            
            # Debug adicional: verificar se h√° tickets nos setores da carteira
            tickets_por_setor = tickets.filter(sector__in=carteira_sectors).count()
            tickets_atribuidos = tickets.filter(assigned_to__in=carteira_users).count()
            print(f"DEBUG: Tickets por setor da carteira: {tickets_por_setor}")
            print(f"DEBUG: Tickets atribu√≠dos a usu√°rios da carteira: {tickets_atribuidos}")
            
        except CommunicationGroup.DoesNotExist:
            print(f"DEBUG: Grupo de carteira com ID {carteira_filter} n√£o encontrado")
        except Exception as e:
            print(f"DEBUG: Erro no filtro carteira: {str(e)}")
    
    # Filtro por pesquisa
    if search:
        tickets = tickets.filter(
            models.Q(title__icontains=search) |
            models.Q(description__icontains=search) |
            models.Q(id__icontains=search)
        )
    
    # Aplicar ordena√ß√£o por data de atualiza√ß√£o (mais recente primeiro)
    tickets = tickets.order_by('-updated_at')
    
    # Configurar pagina√ß√£o - permite escolher quantidade por p√°gina
    per_page = request.GET.get('per_page', '25')
    try:
        per_page = int(per_page)
        if per_page not in [25, 50, 100, 200]:
            per_page = 25
    except (ValueError, TypeError):
        per_page = 25
    
    paginator = Paginator(tickets, per_page)
    page = request.GET.get('page')
    
    try:
        tickets_page = paginator.page(page)
    except PageNotAnInteger:
        # Se a p√°gina n√£o for um inteiro, mostrar a primeira p√°gina
        tickets_page = paginator.page(1)
    except EmptyPage:
        # Se a p√°gina estiver fora do range, mostrar a √∫ltima p√°gina
        tickets_page = paginator.page(paginator.num_pages)
    
    # Preservar par√¢metros de filtro para a pagina√ß√£o
    filter_params = {}
    if per_page != 25:
        filter_params['per_page'] = per_page
    if search:
        filter_params['search'] = search
    if status_filter:
        # Para m√∫ltiplos valores, precisamos trat√°-los de forma especial na URL
        for status in status_filter:
            filter_params.setdefault('status', []).append(status)
    if origem_filter:
        filter_params['origem'] = origem_filter
    if categoria_filter:
        filter_params['categoria'] = categoria_filter
    if setor_filter:
        filter_params['setor'] = setor_filter
    if prioridade_filter:
        filter_params['prioridade'] = prioridade_filter
    if carteira_filter:
        filter_params['carteira'] = carteira_filter
    if atribuidos_filter:
        filter_params['atribuidos'] = atribuidos_filter
    
    # Filtros de data dispon√≠veis para TODOS os usu√°rios
    if date_from_filter:
        filter_params['date_from'] = date_from_filter
    if date_to_filter:
        filter_params['date_to'] = date_to_filter
    
    # Filtro por setor solicitante - dispon√≠vel para TODOS os usu√°rios
    if created_by_sector_filter:
        filter_params['created_by_sector'] = created_by_sector_filter
    
    # Filtros avan√ßados apenas para SUPERADMIN
    if user.hierarchy == 'SUPERADMIN':
        if created_by_filter:
            filter_params['created_by'] = created_by_filter
        if assigned_to_filter:
            filter_params['assigned_to'] = assigned_to_filter
        if user_hierarchy_filter:
            filter_params['user_hierarchy'] = user_hierarchy_filter
        if has_attachments_filter:
            filter_params['has_attachments'] = has_attachments_filter
        if has_comments_filter:
            filter_params['has_comments'] = has_comments_filter
        if overdue_filter:
            filter_params['overdue'] = overdue_filter
        if duplicates_filter:
            filter_params['duplicates'] = duplicates_filter
    
    # Converter par√¢metros para query string
    # doseq=True permite m√∫ltiplos valores para o mesmo par√¢metro (ex: status=ABERTO&status=EM_ANDAMENTO)
    from urllib.parse import urlencode
    filter_query_string = urlencode(filter_params, doseq=True)
    
    # Obter categorias e setores do usu√°rio para os filtros
    user_sectors = list(user.sectors.all())
    if user.sector:
        user_sectors.append(user.sector)
    
    # Remover duplicatas
    user_sectors = list(set(user_sectors))
    
    # Filtro por data - dispon√≠vel para TODOS os usu√°rios
    if date_from_filter:
        from django.utils.dateparse import parse_date
        date_from = parse_date(date_from_filter)
        if date_from:
            tickets = tickets.filter(created_at__date__gte=date_from)
    
    if date_to_filter:
        from django.utils.dateparse import parse_date
        date_to = parse_date(date_to_filter)
        if date_to:
            tickets = tickets.filter(created_at__date__lte=date_to)
    
    # Filtro por setor do solicitante - dispon√≠vel para TODOS os usu√°rios
    if created_by_sector_filter:
        # Filtrar apenas pelo setor principal do criador do ticket
        tickets = tickets.filter(created_by__sector_id=created_by_sector_filter)
    
    # Aplicar filtros avan√ßados apenas para SUPERADMIN
    if user.hierarchy == 'SUPERADMIN':
        # Filtro por usu√°rio que criou
        if created_by_filter:
            tickets = tickets.filter(created_by_id=created_by_filter)
        
        # Filtro por respons√°vel
        if assigned_to_filter:
            if assigned_to_filter == 'unassigned':
                tickets = tickets.filter(assigned_to__isnull=True)
            else:
                tickets = tickets.filter(assigned_to_id=assigned_to_filter)
        
        # Filtro por hierarquia do usu√°rio
        if user_hierarchy_filter:
            tickets = tickets.filter(created_by__hierarchy=user_hierarchy_filter)
        
        # Filtro por anexos
        if has_attachments_filter == 'yes':
            tickets = tickets.filter(attachments__isnull=False).distinct()
        elif has_attachments_filter == 'no':
            tickets = tickets.filter(attachments__isnull=True)
        
        # Filtro por coment√°rios
        if has_comments_filter == 'yes':
            tickets = tickets.filter(comments__isnull=False).distinct()
        elif has_comments_filter == 'no':
            tickets = tickets.filter(comments__isnull=True)
        
        # Filtro por prazo (em atraso)
        if overdue_filter:
            now = timezone.now()
            if overdue_filter == 'yes':
                # Tickets em atraso (n√£o resolvidos e com data de vencimento passada)
                tickets = tickets.filter(
                    models.Q(due_date__lt=now) & 
                    ~models.Q(status__in=['RESOLVIDO', 'FECHADO'])
                )
            elif overdue_filter == 'no':
                # Tickets no prazo
                tickets = tickets.filter(
                    models.Q(due_date__gte=now) | 
                    models.Q(status__in=['RESOLVIDO', 'FECHADO'])
                )
        
        # Filtro por t√≠tulos duplicados
        if duplicates_filter == 'yes':
            # Encontrar t√≠tulos que aparecem mais de uma vez
            from django.db.models import Count
            duplicate_titles = Ticket.objects.values('title').annotate(
                title_count=Count('id')
            ).filter(title_count__gt=1).values_list('title', flat=True)
            tickets = tickets.filter(title__in=list(duplicate_titles))

    # Verificar se √© solicita√ß√£o de exporta√ß√£o (apenas para SUPERADMIN)
    export_format = request.GET.get('export', '')
    if export_format and user.hierarchy == 'SUPERADMIN':
        if export_format == 'csv':
            return export_tickets_csv(tickets)
        elif export_format == 'xlsx':
            return export_tickets_xlsx(tickets)

    # Obter categorias e setores baseado na hierarquia do usu√°rio
    try:
        # Todos os setores para filtro de setor solicitante (dispon√≠vel para todos)
        all_sectors_for_filter = Sector.objects.all().order_by('name')
        
        if user.hierarchy == 'SUPERADMIN':
            # SUPERADMIN pode ver todas as categorias e setores
            user_categories = Category.objects.all().order_by('sector__name', 'name')
            all_categories = Category.objects.all().order_by('sector__name', 'name')
            all_sectors = Sector.objects.all().order_by('name')
            all_users = User.objects.filter(is_active=True).order_by('first_name', 'last_name')
        else:
            # Usu√°rios normais veem apenas as categorias dos seus setores
            user_categories = Category.objects.filter(sector__in=user_sectors).order_by('sector__name', 'name')
            all_categories = user_categories  # Mesma coisa para usu√°rios normais
            all_sectors = user_sectors
            all_users = []  # Usu√°rios normais n√£o precisam desta lista
        
        # Obter grupos de carteira para todos os usu√°rios (busca case-insensitive)
        from communications.models import CommunicationGroup
        carteira_groups = CommunicationGroup.objects.filter(name__icontains='carteira').order_by('name')
        
    except Exception as e:
        # Em caso de erro, usar valores padr√£o vazios
        print(f"Erro ao carregar categorias e setores: {str(e)}")
        user_categories = []
        all_categories = []
        all_sectors = []
        all_users = []
        carteira_groups = []
        all_sectors_for_filter = []
    
    # Obter nomes para exibi√ß√£o dos filtros aplicados
    categoria_name = ''
    setor_name = ''
    created_by_name = ''
    assigned_to_name = ''
    created_by_sector_name = ''
    
    if categoria_filter:
        try:
            categoria_obj = Category.objects.get(id=categoria_filter)
            categoria_name = categoria_obj.name
        except Category.DoesNotExist:
            pass
    
    if setor_filter:
        try:
            setor_obj = Sector.objects.get(id=setor_filter)
            setor_name = setor_obj.name
        except Sector.DoesNotExist:
            pass
    
    # Nome do setor solicitante - dispon√≠vel para todos os usu√°rios
    if created_by_sector_filter:
        try:
            created_by_sector_obj = Sector.objects.get(id=created_by_sector_filter)
            created_by_sector_name = created_by_sector_obj.name
        except Sector.DoesNotExist:
            pass
    
    if created_by_filter and user.hierarchy == 'SUPERADMIN':
        try:
            created_by_obj = User.objects.get(id=created_by_filter)
            created_by_name = created_by_obj.get_full_name()
        except User.DoesNotExist:
            pass
    
    if assigned_to_filter and user.hierarchy == 'SUPERADMIN':
        if assigned_to_filter == 'unassigned':
            assigned_to_name = 'N√£o Atribu√≠do'
        else:
            try:
                assigned_to_obj = User.objects.get(id=assigned_to_filter)
                assigned_to_name = assigned_to_obj.get_full_name()
            except User.DoesNotExist:
                pass
    
    context = {
        'tickets': tickets_page,
        'user': user,
        'search': search,
        'status': status_filter,
        'origem': origem_filter,
        'categoria': categoria_filter,
        'setor': setor_filter,
        'prioridade': prioridade_filter,
        'carteira': carteira_filter,
        'atribuidos': atribuidos_filter,
        'categoria_name': categoria_name,
        'setor_name': setor_name,
        'user_categories': user_categories,
        'user_sectors': user_sectors,
        'carteira_groups': carteira_groups,
        'paginator': paginator,
        'page_obj': tickets_page,
        'per_page': per_page,
        # Filtro por setor solicitante - dispon√≠vel para TODOS
        'created_by_sector': created_by_sector_filter,
        'created_by_sector_name': created_by_sector_name,
        'all_sectors_for_filter': all_sectors_for_filter,
        # Filtros avan√ßados para SUPERADMIN
        'created_by': created_by_filter,
        'assigned_to': assigned_to_filter,
        'date_from': date_from_filter,
        'date_to': date_to_filter,
        'user_hierarchy': user_hierarchy_filter,
        'has_attachments': has_attachments_filter,
        'has_comments': has_comments_filter,
        'overdue': overdue_filter,
        'duplicates': duplicates_filter,
        'created_by_name': created_by_name,
        'assigned_to_name': assigned_to_name,
        # Dados completos para SUPERADMIN
        'all_categories': all_categories if user.hierarchy == 'SUPERADMIN' else user_categories,
        'all_sectors': all_sectors if user.hierarchy == 'SUPERADMIN' else user_sectors,
        'all_users': all_users if user.hierarchy == 'SUPERADMIN' else [],
        # Par√¢metros de filtro para preservar na pagina√ß√£o
        'filter_query_string': filter_query_string,
        'filter_params': filter_params,
        # Data atual para verificar tickets atrasados no template
        'now': timezone.now(),
    }
    return render(request, 'tickets/list.html', context)


@login_required
def tickets_history_view(request):
    """View para mostrar hist√≥rico de chamados conclu√≠dos"""
    user = request.user
    
    # Filtro base: TODOS os usu√°rios sempre veem seus pr√≥prios chamados fechados
    base_filter = models.Q(created_by=user, status='FECHADO')
    
    # Filtrar apenas tickets fechados baseado na hierarquia do usu√°rio
    if user.can_view_all_tickets():
        # Admin v√™ todos os tickets fechados
        tickets = Ticket.objects.filter(status='FECHADO')
    elif user.can_view_sector_tickets():
        # Supervisores veem: pr√≥prios tickets fechados + tickets fechados dos setores + atribu√≠dos fechados
        user_sectors = list(user.sectors.all())
        if user.sector:
            user_sectors.append(user.sector)
            
        tickets = Ticket.objects.filter(
            base_filter |  # Sempre inclui pr√≥prios tickets fechados
            models.Q(sector__in=user_sectors, status='FECHADO') |
            models.Q(assigned_to=user, status='FECHADO')
        ).distinct()
    else:
        # Usu√°rios comuns veem: pr√≥prios tickets fechados + atribu√≠dos fechados
        tickets = Ticket.objects.filter(
            base_filter |  # Sempre inclui pr√≥prios tickets fechados
            models.Q(assigned_to=user, status='FECHADO') |
            models.Q(additional_assignments__user=user, additional_assignments__is_active=True, status='FECHADO')
        ).distinct()
    
    # Aplicar ordena√ß√£o
    tickets = tickets.order_by('-closed_at')
    
    # Configurar pagina√ß√£o
    paginator = Paginator(tickets, 10)  # 15 tickets por p√°gina
    page = request.GET.get('page')
    
    try:
        tickets_page = paginator.page(page)
    except PageNotAnInteger:
        # Se a p√°gina n√£o for um inteiro, mostrar a primeira p√°gina
        tickets_page = paginator.page(1)
    except EmptyPage:
        # Se a p√°gina estiver fora do range, mostrar a √∫ltima p√°gina
        tickets_page = paginator.page(paginator.num_pages)
    
    context = {
        'tickets': tickets_page,
        'user': user,
        'is_history': True,
        'paginator': paginator,
        'page_obj': tickets_page,
    }
    return render(request, 'tickets/history.html', context)


@login_required
def ticket_delete_view(request, ticket_id):
    """View para exclus√£o de tickets - apenas SUPERADMIN"""
    user = request.user
    
    # Verificar se √© SUPERADMIN
    if user.hierarchy != 'SUPERADMIN':
        messages.error(request, 'Voc√™ n√£o tem permiss√£o para excluir chamados.')
        return redirect('tickets_list')
    
    ticket = get_object_or_404(Ticket, id=ticket_id)
    
    if request.method == 'POST':
        ticket_id_str = f"#{ticket.id}"
        ticket_title = ticket.title
        
        # Registrar log antes de excluir
        log_action(
            user, 
            'TICKET_DELETE', 
            f'Chamado exclu√≠do: {ticket_id_str} - {ticket_title}',
            request
        )
        
        # Excluir o ticket (isso tamb√©m excluir√° registros relacionados via CASCADE)
        ticket.delete()
        
        messages.success(request, f'Chamado {ticket_id_str} - "{ticket_title}" exclu√≠do com sucesso!')
        
        # Retornar para a p√°gina anterior se especificado, sen√£o para a lista
        next_url = request.POST.get('next') or request.GET.get('next')
        if next_url:
            return redirect(next_url)
        return redirect('tickets_list')
    
    # Para GET, redireciona para a lista (exclus√£o deve ser via POST)
    return redirect('tickets_list')


@login_required
def tickets_bulk_delete_view(request):
    """View para exclus√£o em lote de tickets - apenas SUPERADMIN"""
    user = request.user
    
    # Verificar se √© SUPERADMIN
    if user.hierarchy != 'SUPERADMIN':
        return JsonResponse({'success': False, 'error': 'Sem permiss√£o'}, status=403)
    
    if request.method == 'POST':
        import json
        try:
            data = json.loads(request.body)
            ticket_ids = data.get('ticket_ids', [])
            
            if not ticket_ids:
                return JsonResponse({'success': False, 'error': 'Nenhum ticket selecionado'}, status=400)
            
            # Buscar tickets
            tickets = Ticket.objects.filter(id__in=ticket_ids)
            count = tickets.count()
            
            if count == 0:
                return JsonResponse({'success': False, 'error': 'Nenhum ticket encontrado'}, status=404)
            
            # Registrar log
            ticket_ids_str = ', '.join([f'#{t.id}' for t in tickets])
            log_action(
                user, 
                'TICKET_BULK_DELETE', 
                f'{count} chamados exclu√≠dos em lote: {ticket_ids_str}',
                request
            )
            
            # Excluir tickets
            tickets.delete()
            
            return JsonResponse({
                'success': True, 
                'message': f'{count} chamado(s) exclu√≠do(s) com sucesso!',
                'deleted_count': count
            })
            
        except json.JSONDecodeError:
            return JsonResponse({'success': False, 'error': 'Dados inv√°lidos'}, status=400)
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)}, status=500)
    
    return JsonResponse({'success': False, 'error': 'M√©todo n√£o permitido'}, status=405)


@login_required
def ticket_edit_view(request, ticket_id):
    """View para edi√ß√£o completa de tickets - apenas SUPERADMIN"""
    user = request.user
    
    # Verificar se √© SUPERADMIN
    if user.hierarchy != 'SUPERADMIN':
        messages.error(request, 'Voc√™ n√£o tem permiss√£o para editar chamados.')
        return redirect('tickets_list')
    
    ticket = get_object_or_404(Ticket, id=ticket_id)
    
    if request.method == 'POST':
        # Capturar valores antigos para log
        old_values = {
            'title': ticket.title,
            'description': ticket.description,
            'sector': ticket.sector.name,
            'category': ticket.category.name if ticket.category else None,
            'status': ticket.status,
            'priority': ticket.priority,
            'assigned_to': ticket.assigned_to.get_full_name() if ticket.assigned_to else None,
            'store_location': ticket.store_location,
        }
        
        # Atualizar campos
        ticket.title = request.POST.get('title', ticket.title)
        ticket.description = request.POST.get('description', ticket.description)
        
        sector_id = request.POST.get('sector')
        if sector_id:
            ticket.sector = get_object_or_404(Sector, id=sector_id)
        
        category_id = request.POST.get('category', '').strip()
        if category_id:
            ticket.category = get_object_or_404(Category, id=category_id)
        else:
            ticket.category = None
        
        ticket.status = request.POST.get('status', ticket.status)
        ticket.priority = request.POST.get('priority', ticket.priority)
        
        assigned_to_id = request.POST.get('assigned_to', '').strip()
        if assigned_to_id:
            ticket.assigned_to = get_object_or_404(User, id=assigned_to_id)
        else:
            ticket.assigned_to = None
        
        ticket.store_location = request.POST.get('store_location', '').strip() or None
        ticket.responsible_person = request.POST.get('responsible_person', '').strip() or None
        ticket.phone = request.POST.get('phone', '').strip() or None
        ticket.solution = request.POST.get('solution', '').strip() or ''
        
        # Atualizar datas de status se necess√°rio
        if ticket.status in ['RESOLVIDO', 'FECHADO'] and not ticket.resolved_at:
            ticket.resolved_at = timezone.now()
        if ticket.status == 'FECHADO' and not ticket.closed_at:
            ticket.closed_at = timezone.now()
        
        ticket.save()
        
        # Criar log de altera√ß√£o
        changes = []
        new_values = {
            'title': ticket.title,
            'description': ticket.description,
            'sector': ticket.sector.name,
            'category': ticket.category.name if ticket.category else None,
            'status': ticket.status,
            'priority': ticket.priority,
            'assigned_to': ticket.assigned_to.get_full_name() if ticket.assigned_to else None,
            'store_location': ticket.store_location,
        }
        
        for key, old_val in old_values.items():
            new_val = new_values.get(key)
            if old_val != new_val:
                changes.append(f"{key}: {old_val} ‚Üí {new_val}")
        
        if changes:
            TicketLog.objects.create(
                ticket=ticket,
                user=user,
                old_status=old_values['status'],
                new_status=ticket.status,
                observation=f"Chamado editado por SUPERADMIN. Altera√ß√µes: {'; '.join(changes)}"
            )
            
            log_action(
                user, 
                'TICKET_EDIT', 
                f'Chamado #{ticket.id} editado: {"; ".join(changes)}',
                request
            )
        
        messages.success(request, f'Chamado #{ticket.id} atualizado com sucesso!')
        
        next_url = request.POST.get('next') or request.GET.get('next')
        if next_url:
            return redirect(next_url)
        return redirect('ticket_detail', ticket_id=ticket.id)
    
    # GET - Carregar formul√°rio de edi√ß√£o
    sectors = Sector.objects.all().order_by('name')
    categories = Category.objects.all().order_by('sector__name', 'name')
    users = User.objects.filter(is_active=True).order_by('first_name', 'last_name')
    
    context = {
        'ticket': ticket,
        'sectors': sectors,
        'categories': categories,
        'users': users,
        'status_choices': Ticket.STATUS_CHOICES,
        'priority_choices': Ticket.PRIORITY_CHOICES,
    }
    return render(request, 'tickets/edit.html', context)


@login_required
def ticket_detail_view(request, ticket_id):
    ticket = get_object_or_404(Ticket, id=ticket_id)
    user = request.user
    
    # Verificar permiss√£o para visualizar o ticket
    user_sectors = list(user.sectors.all())
    if user.sector:
        user_sectors.append(user.sector)
    
    can_view = (
        user.can_view_all_tickets() or 
        (user.can_view_sector_tickets() and ticket.sector in user_sectors) or
        ticket.created_by == user or
        user in ticket.get_all_assigned_users()
    )
    
    if not can_view:
        messages.error(request, 'Voc√™ n√£o tem permiss√£o para visualizar este chamado.')
        return redirect('tickets_list')
    
    # Processar upload de arquivos via POST
    if request.method == 'POST' and 'upload_files' in request.POST:
        # Verificar permiss√£o para adicionar arquivos
        can_upload = (
            user.can_view_all_tickets() or 
            (user.can_view_sector_tickets() and ticket.sector in user_sectors) or
            ticket.created_by == user or
            ticket.assigned_to == user or
            user in ticket.get_all_assigned_users()
        )
        
        if not can_upload:
            messages.error(request, 'Voc√™ n√£o tem permiss√£o para adicionar arquivos neste chamado.')
            return redirect('ticket_detail', ticket_id=ticket.id)
        
        # Processar arquivos anexados
        from .models import TicketAttachment
        attachments = request.FILES.getlist('new_attachments')
        
        if not attachments:
            messages.error(request, 'Nenhum arquivo foi selecionado.')
            return redirect('ticket_detail', ticket_id=ticket.id)
        
        uploaded_count = 0
        for attachment in attachments:
            # Verificar tamanho do arquivo (limite de 50MB por exemplo)
            if attachment.size > 50 * 1024 * 1024:  # 50MB
                messages.warning(request, f'Arquivo "{attachment.name}" √© muito grande (m√°ximo 50MB). Arquivo ignorado.')
                continue
                
            TicketAttachment.objects.create(
                ticket=ticket,
                file=attachment,
                original_filename=attachment.name,
                file_size=attachment.size,
                content_type=attachment.content_type,
                uploaded_by=user
            )
            uploaded_count += 1
        
        if uploaded_count > 0:
            # Adicionar coment√°rio informativo sobre os arquivos adicionados
            TicketComment.objects.create(
                ticket=ticket,
                user=user,
                comment=f'{uploaded_count} arquivo(s) adicionado(s) ao chamado.',
                comment_type='COMMENT'
            )
            
            messages.success(request, f'{uploaded_count} arquivo(s) adicionado(s) com sucesso!')
            log_action(
                user, 
                'TICKET_ATTACHMENT', 
                f'{uploaded_count} arquivo(s) adicionado(s) ao chamado #{ticket.id}',
                request
            )
        
        return redirect('ticket_detail', ticket_id=ticket.id)
    
    # Marcar como visualizado
    ticket.mark_as_viewed(user)
    
    # Verificar se pode assumir o chamado
    can_assume = ticket.can_assume(user)
    
    # Verificar se pode atribuir outros usu√°rios
    can_assign = (
        user.can_view_sector_tickets() or 
        user.can_view_all_tickets() or
        ticket.assigned_to == user
    )
    
    # Verificar se pode fazer upload de arquivos
    can_upload = (
        user.can_view_all_tickets() or 
        (user.can_view_sector_tickets() and ticket.sector in user_sectors) or
        ticket.created_by == user or
        ticket.assigned_to == user or
        user in ticket.get_all_assigned_users()
    )
    
    # Buscar usu√°rios para atribui√ß√£o (todos os setores) - sempre dispon√≠vel
    sector_users = User.objects.filter(is_active=True).exclude(id=user.id).order_by('sector__name', 'first_name')
    
    # Obter URL de retorno (para o bot√£o voltar manter filtros/busca/p√°gina)
    from urllib.parse import unquote
    next_url = request.GET.get('next', '')
    if next_url:
        next_url = unquote(next_url)
    else:
        next_url = '/tickets/'  # URL padr√£o se n√£o houver par√¢metro next
    
    context = {
        'ticket': ticket,
        'logs': ticket.logs.all(),
        'comments': ticket.comments.all(),
        'user': user,
        'can_assume': can_assume,
        'can_assign': can_assign,
        'can_upload': can_upload,
        'assigned_users': ticket.get_all_assigned_users(),
        'additional_assignments': ticket.additional_assignments.filter(is_active=True).select_related('user', 'assigned_by'),
        'sector_users': sector_users,
        'next_url': next_url,
    }
    return render(request, 'tickets/detail.html', context)


@login_required
def assume_ticket_view(request, ticket_id):
    """Assumir um chamado"""
    if request.method == 'POST':
        ticket = get_object_or_404(Ticket, id=ticket_id)
        comment = request.POST.get('comment', '')
        
        if ticket.assume_ticket(request.user, comment):
            messages.success(request, f'Chamado #{ticket.id} assumido com sucesso!')
            log_action(
                request.user, 
                'TICKET_ASSUME', 
                f'Chamado #{ticket.id} assumido',
                request
            )
        else:
            messages.error(request, 'N√£o foi poss√≠vel assumir este chamado.')
        
        return redirect('ticket_detail', ticket_id=ticket.id)
    
    return redirect('tickets_list')


@login_required
def update_priority_view(request, ticket_id):
    """Atualizar prioridade do chamado (apenas para o respons√°vel)"""
    if request.method == 'POST':
        ticket = get_object_or_404(Ticket, id=ticket_id)
        user = request.user
        
        # Verificar se o usu√°rio √© o respons√°vel ou auxiliar do chamado
        is_assigned = (ticket.assigned_to == user or user in ticket.get_all_assigned_users())
        
        if not is_assigned:
            messages.error(request, 'Apenas o respons√°vel pelo chamado pode alterar a prioridade.')
            return redirect('ticket_detail', ticket_id=ticket.id)
        
        new_priority = request.POST.get('priority', '').strip()
        
        # Validar prioridade
        valid_priorities = ['BAIXA', 'MEDIA', 'ALTA', 'CRITICA']
        if new_priority not in valid_priorities:
            messages.error(request, 'Prioridade inv√°lida.')
            return redirect('ticket_detail', ticket_id=ticket.id)
        
        old_priority = ticket.get_priority_display()
        old_priority_value = ticket.priority
        ticket.priority = new_priority
        ticket.save()
        
        # Registrar no log
        TicketComment.objects.create(
            ticket=ticket,
            user=user,
            comment=f'Prioridade alterada de {old_priority} para {ticket.get_priority_display()}',
            comment_type='PRIORITY_CHANGE'
        )
        
        # Notificar usu√°rios envolvidos sobre a mudan√ßa de prioridade
        try:
            from notifications.services import notification_service, NotificationType, NotificationChannel
            
            # Coletar usu√°rios envolvidos (criador + atribu√≠dos)
            involved_users = set()
            if ticket.created_by and ticket.created_by != user:
                involved_users.add(ticket.created_by)
            if ticket.assigned_to and ticket.assigned_to != user:
                involved_users.add(ticket.assigned_to)
            for aux_user in ticket.get_all_assigned_users():
                if aux_user != user:
                    involved_users.add(aux_user)
            
            if involved_users:
                # Definir √≠cone baseado na nova prioridade
                priority_icons = {
                    'BAIXA': 'üü¢',
                    'MEDIA': 'üü°', 
                    'ALTA': 'üü†',
                    'CRITICA': 'üî¥'
                }
                icon = priority_icons.get(new_priority, 'üìã')
                
                notification_service.send_notification(
                    recipients=list(involved_users),
                    title=f'{icon} Prioridade Alterada - Chamado #{ticket.id}',
                    message=f'{user.get_full_name() or user.username} alterou a prioridade de "{old_priority}" para "{ticket.get_priority_display()}"',
                    notification_type=NotificationType.TICKET_STATUS_CHANGED,
                    channels=[NotificationChannel.IN_APP, NotificationChannel.PUSH, NotificationChannel.ONESIGNAL],
                    action_url=f'/tickets/{ticket.id}/',
                    priority='ALTA' if new_priority in ['ALTA', 'CRITICA'] else 'NORMAL',
                    icon='fas fa-exclamation-triangle',
                    extra_data={
                        'ticket_id': ticket.id,
                        'old_priority': old_priority_value,
                        'new_priority': new_priority,
                        'changed_by': user.id
                    },
                    created_by=user
                )
        except Exception as e:
            import logging
            logger = logging.getLogger(__name__)
            logger.error(f'Erro ao enviar notifica√ß√£o de mudan√ßa de prioridade: {e}')
        
        messages.success(request, f'Prioridade alterada para {ticket.get_priority_display()}.')
        log_action(
            user, 
            'TICKET_PRIORITY_CHANGE', 
            f'Prioridade do chamado #{ticket.id} alterada para {ticket.get_priority_display()}',
            request
        )
        
        return redirect('ticket_detail', ticket_id=ticket.id)
    
    return redirect('tickets_list')


@login_required
def add_comment_view(request, ticket_id):
    """Adicionar coment√°rio ao chamado"""
    if request.method == 'POST':
        ticket = get_object_or_404(Ticket, id=ticket_id)
        comment_text = request.POST.get('comment')
        comment_type = request.POST.get('comment_type', 'COMMENT')
        # Suporte para m√∫ltiplos usu√°rios (select multiple)
        assigned_user_ids = request.POST.getlist('assigned_to')
        
        # Verificar permiss√£o para comentar
        can_comment = (
            request.user.can_view_all_tickets() or 
            (request.user.can_view_sector_tickets() and ticket.sector == request.user.sector) or
            ticket.created_by == request.user or
            request.user in ticket.get_all_assigned_users()
        )
        
        if not can_comment:
            messages.error(request, 'Voc√™ n√£o tem permiss√£o para comentar neste chamado.')
            return redirect('ticket_detail', ticket_id=ticket.id)
        
        if not comment_text:
            messages.error(request, 'Coment√°rio √© obrigat√≥rio.')
            return redirect('ticket_detail', ticket_id=ticket.id)
        
        # Criar coment√°rio principal
        comment = TicketComment.objects.create(
            ticket=ticket,
            user=request.user,
            comment=comment_text,
            comment_type=comment_type
        )
        
        # Se √© uma atribui√ß√£o, adicionar usu√°rios (suporte para m√∫ltiplos)
        if comment_type == 'ASSIGNMENT' and assigned_user_ids:
            from users.models import User
            assigned_users = User.objects.filter(id__in=assigned_user_ids)
            assigned_names = []
            
            for assigned_user in assigned_users:
                ticket.assign_additional_user(assigned_user, request.user, comment_text)
                assigned_names.append(assigned_user.full_name)
            
            # Atualizar o coment√°rio principal com a lista de usu√°rios atribu√≠dos
            if len(assigned_names) > 1:
                comment.comment = f"{comment_text}\n\nUsu√°rios atribu√≠dos: {', '.join(assigned_names)}"
                comment.save()
            elif len(assigned_names) == 1:
                # Para manter compatibilidade, atribui ao primeiro usu√°rio se for apenas um
                comment.assigned_to = assigned_users.first()
                comment.save()
        
        if comment_type == 'ASSIGNMENT' and assigned_user_ids:
            count = len(assigned_user_ids)
            if count > 1:
                messages.success(request, f'{count} usu√°rios atribu√≠dos com sucesso!')
            else:
                messages.success(request, 'Usu√°rio atribu√≠do com sucesso!')
        else:
            messages.success(request, 'Coment√°rio adicionado com sucesso!')
        
        log_action(
            request.user, 
            'TICKET_COMMENT', 
            f'Coment√°rio adicionado ao chamado #{ticket.id}',
            request
        )
        
        return redirect('ticket_detail', ticket_id=ticket.id)
    
    return redirect('tickets_list')


@login_required
def update_ticket_status_view(request, ticket_id):
    """Atualizar status do chamado"""
    if request.method == 'POST':
        ticket = get_object_or_404(Ticket, id=ticket_id)
        new_status = request.POST.get('status')
        observation = request.POST.get('observation', '')
        solution = request.POST.get('solution', '')
        
        # Verificar permiss√£o para atualizar
        can_update = (
            request.user.can_view_all_tickets() or 
            (request.user.can_view_sector_tickets() and ticket.sector == request.user.sector) or
            ticket.assigned_to == request.user or
            request.user in ticket.get_all_assigned_users() or
            ticket.created_by == request.user  # Criador pode aprovar/reprovar
        )
        
        if not can_update:
            messages.error(request, 'Voc√™ n√£o tem permiss√£o para atualizar este chamado.')
            return redirect('ticket_detail', ticket_id=ticket.id)
        
        if not new_status:
            messages.error(request, 'Status √© obrigat√≥rio.')
            return redirect('ticket_detail', ticket_id=ticket.id)
        
        old_status = ticket.status
        ticket.status = new_status
        
        if new_status == 'RESOLVIDO':
            ticket.resolved_at = timezone.now()
            if solution:
                ticket.solution = solution
            else:
                messages.error(request, 'Solu√ß√£o √© obrigat√≥ria para marcar como resolvido.')
                return redirect('ticket_detail', ticket_id=ticket.id)
            
            # Se a categoria n√£o requer aprova√ß√£o ou n√£o tem categoria, fechar direto
            if not ticket.category or not ticket.category.requires_approval:
                ticket.status = 'FECHADO'
                ticket.closed_at = timezone.now()
        elif new_status == 'FECHADO':
            ticket.closed_at = timezone.now()
        elif new_status == 'EM_ANDAMENTO' and old_status in ['RESOLVIDO', 'FECHADO', 'AGUARDANDO_APROVACAO']:
            # Reabertura do chamado - limpar campos de resolu√ß√£o se necess√°rio
            if old_status in ['RESOLVIDO', 'FECHADO']:
                ticket.resolved_at = None
                ticket.closed_at = None
        
        ticket.save()
        
        # Criar log
        TicketLog.objects.create(
            ticket=ticket,
            user=request.user,
            old_status=old_status,
            new_status=new_status,
            observation=observation
        )
        
        # Criar coment√°rio se houver observa√ß√£o
        if observation:
            comment_type = 'STATUS_CHANGE'
            if new_status == 'FECHADO' and old_status == 'RESOLVIDO':
                comment_type = 'COMMENT'
                observation = f"Solu√ß√£o aprovada pelo usu√°rio. {observation}"
            elif new_status == 'EM_ANDAMENTO' and old_status in ['RESOLVIDO', 'FECHADO', 'AGUARDANDO_APROVACAO']:
                comment_type = 'COMMENT'
                if old_status == 'RESOLVIDO':
                    observation = f"Solu√ß√£o reprovada pelo usu√°rio. Motivo: {observation}"
                else:
                    observation = f"Chamado reaberto. Motivo: {observation}"
                
            TicketComment.objects.create(
                ticket=ticket,
                user=request.user,
                comment=observation,
                comment_type=comment_type
            )
        
        # Mensagem de sucesso personalizada
        if ticket.status == 'FECHADO' and old_status == 'RESOLVIDO':
            # Fechamento direto sem aprova√ß√£o
            messages.success(request, f'Chamado #{ticket.id} resolvido e fechado automaticamente (categoria n√£o requer aprova√ß√£o do usu√°rio).')
        elif new_status == 'RESOLVIDO' and ticket.category and ticket.category.requires_approval:
            messages.success(request, f'Chamado #{ticket.id} marcado como resolvido. Aguardando aprova√ß√£o do usu√°rio.')
        elif new_status == 'FECHADO':
            messages.success(request, f'Chamado #{ticket.id} fechado com sucesso!')
        elif new_status == 'EM_ANDAMENTO' and old_status in ['RESOLVIDO', 'FECHADO', 'AGUARDANDO_APROVACAO']:
            messages.warning(request, f'Chamado #{ticket.id} foi reaberto e retornou para "Em Andamento".')
        else:
            messages.success(request, f'Status do chamado #{ticket.id} atualizado com sucesso!')
        
        log_action(
            request.user, 
            'TICKET_UPDATE', 
            f'Status do chamado #{ticket.id} alterado: {old_status} ‚Üí {new_status}',
            request
        )
        
        return redirect('ticket_detail', ticket_id=ticket.id)
    
    return redirect('tickets_list')


@login_required
def ticket_create_view(request):
    if request.method == 'POST':
        title = request.POST.get('title')
        description = request.POST.get('description')
        sector_id = request.POST.get('sector')
        category_id = request.POST.get('category', '').strip() or None
        specific_user_id = request.POST.get('specific_user', '').strip() or None
        requires_approval = request.POST.get('requires_approval') == 'on'
        approval_user_id = request.POST.get('approval_user', '').strip() or None
        
        # Validar que a descri√ß√£o n√£o est√° vazia
        if not description or description.strip() == '':
            messages.error(request, 'O campo Mensagem √© obrigat√≥rio.')
            sectors = Sector.objects.all()
            users = User.objects.filter(is_active=True).exclude(id=request.user.id).order_by('sector__name', 'first_name')
            return render(request, 'tickets/create.html', {
                'sectors': sectors,
                'users': users,
                'title': title,
                'sector_id': sector_id,
                'category_id': category_id,
            })
        
        # Validar categoria: obrigat√≥ria apenas se n√£o houver usu√°rio espec√≠fico
        if not specific_user_id and not category_id:
            messages.error(request, 'Categoria √© obrigat√≥ria quando o chamado √© para o setor inteiro.')
            sectors = Sector.objects.all()
            users = User.objects.filter(is_active=True).exclude(id=request.user.id).order_by('sector__name', 'first_name')
            return render(request, 'tickets/create.html', {
                'sectors': sectors,
                'users': users,
                'title': title,
                'sector_id': sector_id,
                'category_id': category_id,
            })
        
        sector = get_object_or_404(Sector, id=sector_id)
        category = get_object_or_404(Category, id=category_id) if category_id else None
        
        # Novos campos opcionais
        store_location = request.POST.get('store_location', '').strip() or None
        responsible_person = request.POST.get('responsible_person', '').strip() or None
        phone = request.POST.get('phone', '').strip() or None
        
        # Criar ticket
        ticket = Ticket.objects.create(
            title=title,
            description=description,
            sector=sector,
            category=category,
            created_by=request.user,
            requires_approval=requires_approval or (category.requires_approval if category else False),
            approval_user_id=approval_user_id if approval_user_id else None,
            solution_time_hours=int(request.POST.get('solution_time_hours', 24)),
            priority=request.POST.get('priority', 'MEDIA'),
            store_location=store_location,
            responsible_person=responsible_person,
            phone=phone,
            assigned_to_id=specific_user_id if specific_user_id else None
        )
        
        # Processar arquivos anexados
        from .models import TicketAttachment
        attachments = request.FILES.getlist('attachments')
        for attachment in attachments:
            TicketAttachment.objects.create(
                ticket=ticket,
                file=attachment,
                original_filename=attachment.name,
                file_size=attachment.size,
                content_type=attachment.content_type,
                uploaded_by=request.user
            )
        
        # Criar log inicial
        TicketLog.objects.create(
            ticket=ticket,
            user=request.user,
            new_status='ABERTO',
            observation='Chamado criado'
        )
        
        log_action(
            request.user, 
            'TICKET_CREATE', 
            f'Chamado criado: #{ticket.id} - {ticket.title}',
            request
        )
        
        if attachments:
            messages.success(request, f'Chamado #{ticket.id} criado com sucesso! {len(attachments)} arquivo(s) anexado(s).')
        else:
            messages.success(request, f'Chamado #{ticket.id} criado com sucesso!')
        return redirect('ticket_detail', ticket_id=ticket.id)
    
    sectors = Sector.objects.all()
    # Buscar todos os usu√°rios ativos para sele√ß√£o
    users = User.objects.filter(is_active=True).exclude(id=request.user.id).order_by('sector__name', 'first_name')
    context = {
        'sectors': sectors,
        'users': users,
    }
    return render(request, 'tickets/create.html', context)


def get_categories_by_sector(request):
    sector_id = request.GET.get('sector_id')
    if sector_id:
        categories = Category.objects.filter(sector_id=sector_id, is_active=True)
        data = [{'id': cat.id, 'name': cat.name, 'default_description': cat.default_description, 'default_solution_time_hours': cat.default_solution_time_hours} for cat in categories]
        return JsonResponse({'categories': data})
    return JsonResponse({'categories': []})


def get_users_by_sector(request):
    """API para retornar todos os usu√°rios de um setor espec√≠fico"""
    sector_id = request.GET.get('sector_id')
    if sector_id:
        users = User.objects.filter(sector_id=sector_id, is_active=True).order_by('first_name', 'last_name')
        data = [{'id': user.id, 'name': f'{user.first_name} {user.last_name}'} for user in users]
        return JsonResponse({'users': data})
    return JsonResponse({'users': []})


from django.utils import timezone
from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from .models import Ticket, Category, TicketLog, TicketComment, Webhook
from .serializers import TicketSerializer, CategorySerializer, TicketLogSerializer, TicketCommentSerializer, WebhookSerializer
from users.models import Sector
from core.middleware import log_action


@login_required
def tickets_list_view_duplicate(request):
    user = request.user
    
    # Filtro base: TODOS os usu√°rios sempre veem seus pr√≥prios chamados
    base_filter = models.Q(created_by=user)
    
    # Filtrar tickets baseado na hierarquia do usu√°rio
    if user.can_view_all_tickets():
        tickets = Ticket.objects.all()
    elif user.can_view_sector_tickets():
        # Ver tickets dos setores + pr√≥prios tickets + atribu√≠dos
        user_sectors = list(user.sectors.all())
        if user.sector:
            user_sectors.append(user.sector)
        tickets = Ticket.objects.filter(
            base_filter |  # Sempre inclui pr√≥prios tickets
            models.Q(sector__in=user_sectors) |
            models.Q(assigned_to=user)
        ).distinct()
    else:
        tickets = Ticket.objects.filter(
            base_filter |  # Sempre inclui pr√≥prios tickets
            models.Q(assigned_to=user) |
            models.Q(additional_assignments__user=user, additional_assignments__is_active=True)
        ).exclude(status='FECHADO').distinct()
    
    context = {
        'tickets': tickets.order_by('-updated_at'),
        'user': user,
    }
    return render(request, 'tickets/list.html', context)


class TicketViewSet(viewsets.ModelViewSet):
    queryset = Ticket.objects.all()
    serializer_class = TicketSerializer
    permission_classes = [IsAuthenticated]
    
    def get_queryset(self):
        user = self.request.user
        
        # Filtro base: TODOS os usu√°rios sempre veem seus pr√≥prios chamados
        base_filter = models.Q(created_by=user)
        
        if user.can_view_all_tickets():
            return Ticket.objects.all()
        elif user.can_view_sector_tickets():
            # Ver tickets dos setores + pr√≥prios tickets + atribu√≠dos
            user_sectors = list(user.sectors.all())
            if user.sector:
                user_sectors.append(user.sector)
            return Ticket.objects.filter(
                base_filter |  # Sempre inclui pr√≥prios tickets
                models.Q(sector__in=user_sectors) |
                models.Q(assigned_to=user)
            ).distinct()
        else:
            return Ticket.objects.filter(
                base_filter |  # Sempre inclui pr√≥prios tickets
                models.Q(assigned_to=user) |
                models.Q(additional_assignments__user=user, additional_assignments__is_active=True)
            ).exclude(status='FECHADO').distinct()
    
    @action(detail=True, methods=['post'])
    def update_status(self, request, pk=None):
        ticket = self.get_object()
        new_status = request.data.get('status')
        observation = request.data.get('observation', '')
        solution = request.data.get('solution', '')
        
        if not new_status:
            return Response(
                {'error': 'Status √© obrigat√≥rio'}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        old_status = ticket.status
        ticket.status = new_status
        
        if new_status == 'RESOLVIDO':
            ticket.resolved_at = timezone.now()
            ticket.solution = solution
            # Se a categoria n√£o requer aprova√ß√£o ou n√£o tem categoria, vai direto para fechado
            if not ticket.category or not ticket.category.requires_approval:
                ticket.status = 'FECHADO'
                ticket.closed_at = timezone.now()
        elif new_status == 'FECHADO':
            ticket.closed_at = timezone.now()
        elif new_status == 'EM_ANDAMENTO' and old_status in ['RESOLVIDO', 'FECHADO', 'AGUARDANDO_APROVACAO']:
            # Reabertura do chamado - limpar campos de resolu√ß√£o se necess√°rio
            if old_status in ['RESOLVIDO', 'FECHADO']:
                ticket.resolved_at = None
                ticket.closed_at = None
        
        ticket.save()
        
        # Criar log
        TicketLog.objects.create(
            ticket=ticket,
            user=request.user,
            old_status=old_status,
            new_status=new_status,
            observation=observation
        )
        
        log_action(
            request.user, 
            'TICKET_UPDATE', 
            f'Status do chamado #{ticket.id} alterado: {old_status} ‚Üí {new_status}',
            request
        )
        
        return Response({'message': 'Status atualizado com sucesso'})
    
    @action(detail=True, methods=['post'])
    def add_comment(self, request, pk=None):
        ticket = self.get_object()
        comment_text = request.data.get('comment')
        
        if not comment_text:
            return Response(
                {'error': 'Coment√°rio √© obrigat√≥rio'}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        comment = TicketComment.objects.create(
            ticket=ticket,
            user=request.user,
            comment=comment_text
        )
        
        serializer = TicketCommentSerializer(comment)
        return Response(serializer.data)


class CategoryViewSet(viewsets.ModelViewSet):
    queryset = Category.objects.all()
    serializer_class = CategorySerializer
    permission_classes = [IsAuthenticated]


# ViewSets p√∫blicos (sem autentica√ß√£o) para produ√ß√£o
class PublicTicketViewSet(viewsets.ReadOnlyModelViewSet):
    """ViewSet p√∫blico apenas para leitura de tickets"""
    queryset = Ticket.objects.all()
    serializer_class = TicketSerializer
    permission_classes = []  # Sem autentica√ß√£o
    
    def get_queryset(self):
        """Retorna apenas informa√ß√µes b√°sicas dos tickets"""
        return Ticket.objects.filter(
            status__in=['open', 'in_progress', 'waiting']
        ).select_related('category', 'sector', 'created_by')


class PublicCategoryViewSet(viewsets.ReadOnlyModelViewSet):
    """ViewSet p√∫blico apenas para leitura de categorias"""
    queryset = Category.objects.all()
    serializer_class = CategorySerializer
    permission_classes = []  # Sem autentica√ß√£o


@login_required
def manage_webhooks_view(request):
    """Gerenciar webhooks"""
    if not request.user.can_manage_users():
        messages.error(request, 'Voc√™ n√£o tem permiss√£o para acessar esta √°rea.')
        return redirect('dashboard')
    
    webhooks = Webhook.objects.all().select_related('category', 'sector')
    active_webhooks_count = webhooks.filter(is_active=True).count()
    inactive_webhooks_count = webhooks.filter(is_active=False).count()
    
    context = {
        'webhooks': webhooks,
        'active_webhooks_count': active_webhooks_count,
        'inactive_webhooks_count': inactive_webhooks_count,
        'user': request.user,
    }
    return render(request, 'admin/webhooks.html', context)


@login_required
def create_webhook_view(request):
    """Criar novo webhook"""
    if not request.user.can_manage_users():
        messages.error(request, 'Voc√™ n√£o tem permiss√£o para acessar esta √°rea.')
        return redirect('dashboard')
    
    if request.method == 'POST':
        name = request.POST.get('name')
        url = request.POST.get('url')
        events = request.POST.getlist('events')  # Pegar lista de eventos selecionados
        category_id = request.POST.get('category')
        sector_id = request.POST.get('sector')
        is_active = request.POST.get('is_active') == 'on'
        headers = request.POST.get('headers', '{}')
        
        try:
            # Validar se pelo menos um evento foi selecionado
            if not events:
                messages.error(request, 'Por favor, selecione pelo menos um evento.')
                context = {
                    'event_choices': Webhook.EVENT_CHOICES,
                    'categories': Category.objects.all(),
                    'sectors': Sector.objects.all(),
                    'user': request.user,
                }
                return render(request, 'admin/create_webhook.html', context)
            
            # Validar e parsear headers JSON
            import json
            try:
                headers_dict = json.loads(headers) if headers.strip() else {}
            except json.JSONDecodeError:
                headers_dict = {}
            
            category = get_object_or_404(Category, id=category_id) if category_id else None
            sector = get_object_or_404(Sector, id=sector_id) if sector_id else None
            
            # Criar um webhook para cada evento selecionado
            created_webhooks = []
            for event in events:
                webhook_name = f"{name} - {dict(Webhook.EVENT_CHOICES)[event]}"
                
                webhook = Webhook.objects.create(
                    name=webhook_name,
                    url=url,
                    event=event,
                    category=category,
                    sector=sector,
                    is_active=is_active,
                    headers=headers_dict
                )
                created_webhooks.append(webhook)
            
            log_action(
                request.user, 
                'WEBHOOK_CREATE', 
                f'Webhooks criados: {len(created_webhooks)} webhook(s) para {name}',
                request
            )
            
            if len(created_webhooks) == 1:
                messages.success(request, f'Webhook "{created_webhooks[0].name}" criado com sucesso!')
            else:
                messages.success(request, f'{len(created_webhooks)} webhooks criados com sucesso para "{name}"!')
            return redirect('manage_webhooks')
            
        except Exception as e:
            messages.error(request, f'Erro ao criar webhook: {str(e)}')
    
    context = {
        'event_choices': Webhook.EVENT_CHOICES,
        'categories': Category.objects.all(),
        'sectors': Sector.objects.all(),
        'user': request.user,
    }
    return render(request, 'admin/create_webhook.html', context)


@login_required
def edit_webhook_view(request, webhook_id):
    """Editar webhook existente"""
    if not request.user.can_manage_users():
        messages.error(request, 'Voc√™ n√£o tem permiss√£o para acessar esta √°rea.')
        return redirect('dashboard')
    
    webhook = get_object_or_404(Webhook, id=webhook_id)
    
    if request.method == 'POST':
        webhook.name = request.POST.get('name')
        webhook.url = request.POST.get('url')
        webhook.event = request.POST.get('event')
        
        category_id = request.POST.get('category')
        sector_id = request.POST.get('sector')
        
        webhook.category = get_object_or_404(Category, id=category_id) if category_id else None
        webhook.sector = get_object_or_404(Sector, id=sector_id) if sector_id else None
        webhook.is_active = request.POST.get('is_active') == 'on'
        
        headers = request.POST.get('headers', '{}')
        try:
            import json
            webhook.headers = json.loads(headers) if headers.strip() else {}
        except json.JSONDecodeError:
            webhook.headers = {}
        
        try:
            webhook.save()
            
            log_action(
                request.user,
                'WEBHOOK_UPDATE',
                f'Webhook atualizado: {webhook.name}',
                request
            )
            
            messages.success(request, f'Webhook "{webhook.name}" atualizado com sucesso!')
            return redirect('manage_webhooks')
            
        except Exception as e:
            messages.error(request, f'Erro ao atualizar webhook: {str(e)}')
    
    context = {
        'webhook': webhook,
        'event_choices': Webhook.EVENT_CHOICES,
        'categories': Category.objects.all(),
        'sectors': Sector.objects.all(),
        'user': request.user,
    }
    return render(request, 'admin/edit_webhook.html', context)


@login_required
def delete_webhook_view(request, webhook_id):
    """Excluir webhook"""
    if not request.user.can_manage_users():
        messages.error(request, 'Voc√™ n√£o tem permiss√£o para acessar esta √°rea.')
        return redirect('dashboard')
    
    webhook = get_object_or_404(Webhook, id=webhook_id)
    
    if request.method == 'POST':
        webhook_name = webhook.name
        webhook.delete()
        
        log_action(
            request.user,
            'WEBHOOK_DELETE',
            f'Webhook exclu√≠do: {webhook_name}',
            request
        )
        
        messages.success(request, f'Webhook "{webhook_name}" exclu√≠do com sucesso!')
        return redirect('manage_webhooks')
    
    context = {
        'webhook': webhook,
        'user': request.user,
    }
    return render(request, 'admin/delete_webhook.html', context)


class WebhookViewSet(viewsets.ModelViewSet):
    queryset = Webhook.objects.all()
    serializer_class = WebhookSerializer
    permission_classes = [IsAuthenticated]
    
    @action(detail=True, methods=['post'])
    def test(self, request, pk=None):
        webhook = self.get_object()
        
        # Criar payload de teste
        test_payload = {
            'event': webhook.event,
            'webhook_name': webhook.name,
            'test': True,
            'timestamp': timezone.now().isoformat(),
            'message': 'Este √© um teste do webhook'
        }
        
        try:
            import requests
            response = requests.post(webhook.url, json=test_payload, timeout=10)
            return Response({
                'message': 'Webhook testado com sucesso',
                'status_code': response.status_code,
                'response': response.text[:500]
            })
        except Exception as e:
            return Response({
                'error': f'Erro ao testar webhook: {str(e)}'
            }, status=status.HTTP_400_BAD_REQUEST)


@login_required
def ticket_create_fixed_view(request):
    """View corrigida para cria√ß√£o de tickets com todos os usu√°rios"""
    if request.method == 'POST':
        title = request.POST.get('title')
        description = request.POST.get('description')
        sector_id = request.POST.get('sector')
        category_id = request.POST.get('category', '').strip() or None
        specific_user_id = request.POST.get('specific_user', '').strip() or None
        requires_approval = request.POST.get('requires_approval') == 'on'
        approval_user_id = request.POST.get('approval_user', '').strip() or None
        copy_user_id = request.POST.get('copy', '').strip() or None
        
        # Validar que a descri√ß√£o n√£o est√° vazia
        if not description or description.strip() == '':
            messages.error(request, 'O campo Mensagem √© obrigat√≥rio.')
            sectors = Sector.objects.all()
            users = User.objects.filter(is_active=True).exclude(id=request.user.id).order_by('sector__name', 'first_name')
            return render(request, 'tickets/create.html', {
                'sectors': sectors,
                'users': users,
                'title': title,
                'sector_id': sector_id,
                'category_id': category_id,
            })
        
        # Validar categoria: obrigat√≥ria apenas se n√£o houver usu√°rio espec√≠fico
        if not specific_user_id and not category_id:
            messages.error(request, 'Categoria √© obrigat√≥ria quando o chamado √© para o setor inteiro.')
            sectors = Sector.objects.all()
            users = User.objects.filter(is_active=True).exclude(id=request.user.id).order_by('sector__name', 'first_name')
            return render(request, 'tickets/create.html', {
                'sectors': sectors,
                'users': users,
                'title': title,
                'sector_id': sector_id,
                'category_id': category_id,
            })
        
        sector = get_object_or_404(Sector, id=sector_id)
        category = get_object_or_404(Category, id=category_id) if category_id else None
        
        # Novos campos opcionais
        store_location = request.POST.get('store_location', '').strip() or None
        responsible_person = request.POST.get('responsible_person', '').strip() or None
        phone = request.POST.get('phone', '').strip() or None
        
        # Validar Loja/Local obrigat√≥rio para setor Manuten√ß√£o
        if sector.name.lower() in ['manuten√ß√£o', 'manutencao'] and not store_location:
            messages.error(request, 'O campo Loja/Local √© obrigat√≥rio para o setor de Manuten√ß√£o.')
            sectors = Sector.objects.all()
            users = User.objects.filter(is_active=True).exclude(id=request.user.id).order_by('sector__name', 'first_name')
            return render(request, 'tickets/create.html', {
                'sectors': sectors,
                'users': users,
                'title': title,
                'sector_id': sector_id,
                'category_id': category_id,
            })
        
        # Criar ticket
        ticket = Ticket.objects.create(
            title=title,
            description=description,
            sector=sector,
            category=category,
            created_by=request.user,
            requires_approval=requires_approval or (category.requires_approval if category else False),
            approval_user_id=approval_user_id if approval_user_id else None,
            assigned_to_id=specific_user_id if specific_user_id else None,
            solution_time_hours=int(request.POST.get('solution_time_hours', 24)),
            priority=request.POST.get('priority', 'MEDIA'),
            store_location=store_location,
            responsible_person=responsible_person,
            phone=phone
        )
        
        # Processar arquivos anexados
        from .models import TicketAttachment
        attachments = request.FILES.getlist('attachments')
        for attachment in attachments:
            TicketAttachment.objects.create(
                ticket=ticket,
                file=attachment,
                original_filename=attachment.name,
                file_size=attachment.size,
                content_type=attachment.content_type,
                uploaded_by=request.user
            )
        
        # Criar atribui√ß√£o em c√≥pia se um usu√°rio foi selecionado
        if copy_user_id:
            try:
                copy_user = User.objects.get(id=copy_user_id)
                TicketAssignment.objects.create(
                    ticket=ticket,
                    user=copy_user,
                    assigned_by=request.user,
                    is_active=True
                )
            except User.DoesNotExist:
                pass
        
        # Criar log inicial
        TicketLog.objects.create(
            ticket=ticket,
            user=request.user,
            new_status='ABERTO',
            observation='Chamado criado'
        )
        
        log_action(
            request.user, 
            'TICKET_CREATE', 
            f'Chamado criado: #{ticket.id} - {ticket.title}',
            request
        )
        
        if attachments:
            messages.success(request, f'Chamado #{ticket.id} criado com sucesso! {len(attachments)} arquivo(s) anexado(s).')
        else:
            messages.success(request, f'Chamado #{ticket.id} criado com sucesso!')
        return redirect('ticket_detail', ticket_id=ticket.id)
    
    sectors = Sector.objects.all()
    # Buscar todos os usu√°rios ativos para c√≥pia, exceto o usu√°rio atual
    users = User.objects.filter(is_active=True).exclude(id=request.user.id).order_by('sector__name', 'first_name')
    context = {
        'sectors': sectors,
        'users': users,
    }
    return render(request, 'tickets/create.html', context)


# ========================
# PURCHASE ORDER API VIEWS
# ========================

from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework import status
from .models import PurchaseOrderApproval, TicketComment
import json


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def approve_purchase_order(request, ticket_id, approval_id):
    """API para aprovar uma ordem de compra"""
    try:
        approval = PurchaseOrderApproval.objects.get(
            id=approval_id,
            ticket_id=ticket_id,
            approver=request.user,
            status='PENDING'
        )
        
        comment = request.data.get('comment', '')
        approval.approve(comment)
        
        # Adicionar coment√°rio no ticket
        TicketComment.objects.create(
            ticket=approval.ticket,
            user=request.user,
            comment=f"Ordem de compra aprovada (R$ {approval.amount:.2f}). {comment}".strip(),
            comment_type='COMMENT'
        )
        
        return Response({
            'message': 'Ordem de compra aprovada com sucesso',
            'status': 'approved',
            'next_step': approval.approval_step + 1 if approval.approval_step < 3 else 'completed'
        })
        
    except PurchaseOrderApproval.DoesNotExist:
        return Response({
            'error': 'Aprova√ß√£o n√£o encontrada ou voc√™ n√£o tem permiss√£o para aprov√°-la'
        }, status=status.HTTP_404_NOT_FOUND)
    except Exception as e:
        return Response({
            'error': f'Erro ao processar aprova√ß√£o: {str(e)}'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def reject_purchase_order(request, ticket_id, approval_id):
    """API para rejeitar uma ordem de compra"""
    try:
        approval = PurchaseOrderApproval.objects.get(
            id=approval_id,
            ticket_id=ticket_id,
            approver=request.user,
            status='PENDING'
        )
        
        comment = request.data.get('comment', 'Ordem rejeitada')
        approval.reject(comment)
        
        # Adicionar coment√°rio no ticket
        TicketComment.objects.create(
            ticket=approval.ticket,
            user=request.user,
            comment=f"Ordem de compra rejeitada (R$ {approval.amount:.2f}). Motivo: {comment}",
            comment_type='COMMENT'
        )
        
        # Atualizar status do ticket
        approval.ticket.status = 'REJEITADO'
        approval.ticket.save()
        
        return Response({
            'message': 'Ordem de compra rejeitada',
            'status': 'rejected',
            'reason': comment
        })
        
    except PurchaseOrderApproval.DoesNotExist:
        return Response({
            'error': 'Aprova√ß√£o n√£o encontrada ou voc√™ n√£o tem permiss√£o para rejeit√°-la'
        }, status=status.HTTP_404_NOT_FOUND)
    except Exception as e:
        return Response({
            'error': f'Erro ao processar rejei√ß√£o: {str(e)}'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def pending_approvals(request):
    """API para listar aprova√ß√µes pendentes do usu√°rio"""
    try:
        approvals = PurchaseOrderApproval.objects.filter(
            approver=request.user,
            status='PENDING'
        ).select_related('ticket', 'ticket__category', 'ticket__created_by')
        
        approvals_data = []
        for approval in approvals:
            approvals_data.append({
                'id': approval.id,
                'ticket_id': approval.ticket.id,
                'ticket_title': approval.ticket.title,
                'amount': float(approval.amount),
                'created_at': approval.created_at.isoformat(),
                'approval_step': approval.approval_step,
                'created_by': approval.ticket.created_by.full_name,
                'category': approval.ticket.category.name if approval.ticket.category else None,
            })
        
        return Response({
            'approvals': approvals_data,
            'count': len(approvals_data)
        })
        
    except Exception as e:
        return Response({
            'error': f'Erro ao carregar aprova√ß√µes: {str(e)}'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


def user_tickets_api(request, user_id):
    """
    API REST para buscar chamados de um usu√°rio espec√≠fico
    URL: /api/users/{user_id}/tickets/
    Retorna quantidade e t√≠tulos dos chamados do usu√°rio
    """
    try:
        # Buscar o usu√°rio
        user = get_object_or_404(User, id=user_id)
        
        # Buscar todos os tickets do usu√°rio
        tickets = Ticket.objects.filter(created_by=user).select_related('category', 'sector')
        
        # Preparar dados para resposta
        tickets_data = []
        for ticket in tickets:
            tickets_data.append({
                'id': ticket.id,
                'title': ticket.title,
                'status': ticket.status,
                'category': ticket.category.name if ticket.category else None,
                'sector': ticket.sector.name if ticket.sector else None,
                'created_at': ticket.created_at.isoformat(),
                'updated_at': ticket.updated_at.isoformat(),
            })
        
        # Contar por status
        status_counts = {
            'total': tickets.count(),
            'open': tickets.filter(status='OPEN').count(),
            'in_progress': tickets.filter(status='IN_PROGRESS').count(),
            'closed': tickets.filter(status='CLOSED').count(),
            'cancelled': tickets.filter(status='CANCELLED').count(),
        }
        
        return JsonResponse({
            'user': {
                'id': user.id,
                'username': user.username,
                'full_name': user.full_name,
                'email': user.email,
            },
            'tickets': {
                'count': status_counts['total'],
                'status_breakdown': status_counts,
                'data': tickets_data
            },
            'success': True
        })
        
    except User.DoesNotExist:
        return JsonResponse({
            'error': 'Usu√°rio n√£o encontrado',
            'success': False
        }, status=404)
        
    except Exception as e:
        return JsonResponse({
            'error': f'Erro interno do servidor: {str(e)}',
            'success': False
        }, status=500)


def export_tickets_csv(tickets):
    """Exporta tickets para CSV com descri√ß√£o completa"""
    # Criar resposta HTTP com tipo CSV
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = f'attachment; filename="chamados_completo_{timezone.now().strftime("%Y%m%d_%H%M%S")}.csv"'
    
    # Adicionar BOM para Excel reconhecer UTF-8
    response.write('\ufeff')
    
    writer = csv.writer(response)
    
    # Cabe√ßalhos expandidos
    writer.writerow([
        'ID',
        'T√≠tulo',
        'Status',
        'Prioridade',
        'Categoria',
        'Setor',
        'Solicitante',
        'Email Solicitante',
        'Respons√°vel',
        'Data Cria√ß√£o',
        'Data Atualiza√ß√£o',
        'Descri√ß√£o Completa',
        'Imagens Anexadas',
        'Arquivos Anexados'
    ])
    
    # Dados
    for ticket in tickets.select_related('created_by', 'assigned_to', 'category', 'sector').prefetch_related('attachments'):
        # Processar anexos
        attachments = ticket.attachments.all()
        images_list = []
        files_list = []
        
        if attachments.exists():
            for attachment in attachments:
                # Construir URL completo do anexo
                # Em produ√ß√£o com MinIO, MEDIA_URL j√° cont√©m o endpoint completo
                if hasattr(settings, 'USE_S3') and settings.USE_S3:
                    # MinIO: MEDIA_URL j√° inclui endpoint + bucket, precisa adicionar /media/
                    file_url = f"{settings.MEDIA_URL.rstrip('/')}/media/{attachment.file.name}"
                elif hasattr(settings, 'MEDIA_URL') and settings.MEDIA_URL.startswith('http'):
                    # URL absoluta (CDN ou similar)
                    file_url = f"{settings.MEDIA_URL.rstrip('/')}/{attachment.file.name}"
                elif hasattr(settings, 'MEDIA_URL') and hasattr(settings, 'SITE_URL'):
                    # URL relativa + dom√≠nio do site
                    file_url = f"{settings.SITE_URL.rstrip('/')}{settings.MEDIA_URL.rstrip('/')}/{attachment.file.name}"
                else:
                    # Fallback para m√©todos padr√£o
                    file_url = attachment.file.url if hasattr(attachment.file, 'url') else attachment.file.name
                
                # Verificar se √© imagem
                is_image = False
                if attachment.content_type:
                    is_image = attachment.content_type.startswith('image/')
                elif attachment.original_filename:
                    image_extensions = ['.jpg', '.jpeg', '.png', '.gif', '.bmp', '.webp', '.svg']
                    is_image = any(attachment.original_filename.lower().endswith(ext) for ext in image_extensions)
                
                attachment_info = f"{attachment.original_filename}: {file_url}"
                
                if is_image:
                    images_list.append(attachment_info)
                else:
                    files_list.append(attachment_info)
        
        images_text = '; '.join(images_list) if images_list else 'Nenhuma imagem'
        files_text = '; '.join(files_list) if files_list else 'Nenhum arquivo'
        
        writer.writerow([
            f'#{ticket.id:04d}',
            ticket.title,
            ticket.get_status_display(),
            ticket.get_priority_display() if hasattr(ticket, 'priority') else '',
            ticket.category.name if ticket.category else '',
            ticket.sector.name if ticket.sector else '',
            ticket.created_by.get_full_name() if ticket.created_by else '',
            ticket.created_by.email if ticket.created_by else '',
            ticket.assigned_to.get_full_name() if ticket.assigned_to else 'N√£o atribu√≠do',
            ticket.created_at.strftime('%d/%m/%Y %H:%M'),
            ticket.updated_at.strftime('%d/%m/%Y %H:%M'),
            ticket.description,  # DESCRI√á√ÉO COMPLETA SEM CORTE
            images_text,
            files_text
        ])
    
    return response


def export_tickets_xlsx(tickets):
    """Exporta tickets para Excel (XLSX) com descri√ß√£o completa e anexos"""
    from django.conf import settings
    
    # Criar workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Chamados"
    
    # Estilo do cabe√ßalho
    header_fill = PatternFill(start_color="1F4788", end_color="1F4788", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # Borda para c√©lulas
    thin_border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )
    
    # Cabe√ßalhos expandidos
    headers = [
        'ID',
        'T√≠tulo',
        'Status',
        'Prioridade',
        'Categoria',
        'Setor',
        'Solicitante',
        'Email Solicitante',
        'Respons√°vel',
        'Data Cria√ß√£o',
        'Data Atualiza√ß√£o',
        'Descri√ß√£o Completa',
        'Imagens Anexadas',
        'Arquivos Anexados'
    ]
    
    # Aplicar cabe√ßalhos
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Ajustar largura das colunas
    column_widths = {
        1: 8,   # ID
        2: 35,  # T√≠tulo
        3: 12,  # Status
        4: 12,  # Prioridade
        5: 20,  # Categoria
        6: 18,  # Setor
        7: 22,  # Solicitante
        8: 28,  # Email
        9: 22,  # Respons√°vel
        10: 16, # Data Cria√ß√£o
        11: 16, # Data Atualiza√ß√£o
        12: 60, # Descri√ß√£o Completa
        13: 50, # Imagens
        14: 50  # Arquivos
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width
    
    # Congelar primeira linha (cabe√ßalho)
    ws.freeze_panes = 'A2'
    
    # Estilos para dados
    link_font = Font(color="0563C1", underline="single")
    data_alignment = Alignment(vertical="top", wrap_text=True)
    
    # Dados
    for row_num, ticket in enumerate(tickets.select_related('created_by', 'assigned_to', 'category', 'sector').prefetch_related('attachments'), start=2):
        # Coluna 1: ID
        cell = ws.cell(row=row_num, column=1, value=f'#{ticket.id:04d}')
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
        
        # Coluna 2: T√≠tulo
        cell = ws.cell(row=row_num, column=2, value=ticket.title)
        cell.alignment = data_alignment
        cell.border = thin_border
        
        # Coluna 3: Status
        cell = ws.cell(row=row_num, column=3, value=ticket.get_status_display())
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
        
        # Coluna 4: Prioridade
        priority = ticket.get_priority_display() if hasattr(ticket, 'priority') else ''
        cell = ws.cell(row=row_num, column=4, value=priority)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
        
        # Coluna 5: Categoria
        cell = ws.cell(row=row_num, column=5, value=ticket.category.name if ticket.category else '')
        cell.alignment = data_alignment
        cell.border = thin_border
        
        # Coluna 6: Setor
        cell = ws.cell(row=row_num, column=6, value=ticket.sector.name if ticket.sector else '')
        cell.alignment = data_alignment
        cell.border = thin_border
        
        # Coluna 7: Solicitante
        cell = ws.cell(row=row_num, column=7, value=ticket.created_by.get_full_name() if ticket.created_by else '')
        cell.alignment = data_alignment
        cell.border = thin_border
        
        # Coluna 8: Email Solicitante
        cell = ws.cell(row=row_num, column=8, value=ticket.created_by.email if ticket.created_by else '')
        cell.alignment = data_alignment
        cell.border = thin_border
        
        # Coluna 9: Respons√°vel
        cell = ws.cell(row=row_num, column=9, value=ticket.assigned_to.get_full_name() if ticket.assigned_to else 'N√£o atribu√≠do')
        cell.alignment = data_alignment
        cell.border = thin_border
        
        # Coluna 10: Data Cria√ß√£o
        cell = ws.cell(row=row_num, column=10, value=ticket.created_at.strftime('%d/%m/%Y %H:%M'))
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
        
        # Coluna 11: Data Atualiza√ß√£o
        cell = ws.cell(row=row_num, column=11, value=ticket.updated_at.strftime('%d/%m/%Y %H:%M'))
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
        
        # Coluna 12: Descri√ß√£o Completa (SEM CORTE)
        cell = ws.cell(row=row_num, column=12, value=ticket.description)
        cell.alignment = data_alignment
        cell.border = thin_border
        
        # Colunas 13 e 14: Anexos (Imagens e Arquivos)
        attachments = ticket.attachments.all()
        
        if attachments.exists():
            # Separar imagens e outros arquivos
            images = []
            files = []
            
            for attachment in attachments:
                # Construir URL completo do anexo
                # Em produ√ß√£o com MinIO, MEDIA_URL j√° cont√©m o endpoint completo
                # Ex: https://minio.example.com/bucket/ ou http://localhost:9000/bucket/
                if hasattr(settings, 'USE_S3') and settings.USE_S3:
                    # MinIO: MEDIA_URL j√° inclui endpoint + bucket, precisa adicionar /media/
                    file_url = f"{settings.MEDIA_URL.rstrip('/')}/media/{attachment.file.name}"
                elif hasattr(settings, 'MEDIA_URL') and settings.MEDIA_URL.startswith('http'):
                    # URL absoluta (CDN ou similar)
                    file_url = f"{settings.MEDIA_URL.rstrip('/')}/{attachment.file.name}"
                elif hasattr(settings, 'MEDIA_URL') and hasattr(settings, 'SITE_URL'):
                    # URL relativa + dom√≠nio do site
                    file_url = f"{settings.SITE_URL.rstrip('/')}{settings.MEDIA_URL.rstrip('/')}/{attachment.file.name}"
                else:
                    # Fallback para m√©todos padr√£o
                    file_url = attachment.file.url if hasattr(attachment.file, 'url') else attachment.file.name
                
                # Verificar se √© imagem
                is_image = False
                if attachment.content_type:
                    is_image = attachment.content_type.startswith('image/')
                elif attachment.original_filename:
                    image_extensions = ['.jpg', '.jpeg', '.png', '.gif', '.bmp', '.webp', '.svg']
                    is_image = any(attachment.original_filename.lower().endswith(ext) for ext in image_extensions)
                
                attachment_info = f"{attachment.original_filename}\n{file_url}"
                
                if is_image:
                    images.append(attachment_info)
                else:
                    files.append(attachment_info)
            
            # Coluna 13: Imagens
            images_text = '\n\n'.join(images) if images else 'Nenhuma imagem'
            cell = ws.cell(row=row_num, column=13, value=images_text)
            cell.alignment = data_alignment
            cell.border = thin_border
            if images:
                cell.font = link_font
            
            # Coluna 14: Arquivos
            files_text = '\n\n'.join(files) if files else 'Nenhum arquivo'
            cell = ws.cell(row=row_num, column=14, value=files_text)
            cell.alignment = data_alignment
            cell.border = thin_border
            if files:
                cell.font = link_font
        else:
            # Sem anexos
            cell = ws.cell(row=row_num, column=13, value='Nenhuma imagem')
            cell.alignment = data_alignment
            cell.border = thin_border
            cell.font = Font(color="999999", italic=True)
            
            cell = ws.cell(row=row_num, column=14, value='Nenhum arquivo')
            cell.alignment = data_alignment
            cell.border = thin_border
            cell.font = Font(color="999999", italic=True)
        
        # Ajustar altura da linha para acomodar conte√∫do
        ws.row_dimensions[row_num].height = max(30, len(ticket.description) / 3)
    
    # Criar resposta HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="chamados_completo_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    
    # Salvar workbook na resposta
    wb.save(response)
    
    return response


@login_required
def tickets_export_view(request):
    """View dedicada para exporta√ß√£o de chamados com todos os filtros aplicados"""
    user = request.user
    
    # Coletar todos os filtros da URL (mesma l√≥gica da tickets_list_view)
    search = request.GET.get('search', '')
    status_filter = request.GET.getlist('status')
    origem_filter = request.GET.get('origem', '')
    categoria_filter = request.GET.get('categoria', '')
    setor_filter = request.GET.get('setor', '')
    prioridade_filter = request.GET.get('prioridade', '')
    carteira_filter = request.GET.get('carteira', '')
    atribuidos_filter = request.GET.get('atribuidos', '')
    
    # Filtros avan√ßados para SUPERADMIN
    created_by_filter = request.GET.get('created_by', '')
    created_by_sector_filter = request.GET.get('created_by_sector', '')
    assigned_to_filter = request.GET.get('assigned_to', '')
    date_from_filter = request.GET.get('date_from', '')
    date_to_filter = request.GET.get('date_to', '')
    user_hierarchy_filter = request.GET.get('user_hierarchy', '')
    has_attachments_filter = request.GET.get('has_attachments', '')
    has_comments_filter = request.GET.get('has_comments', '')
    overdue_filter = request.GET.get('overdue', '')
    
    # Filtro base
    base_filter = models.Q(created_by=user)
    
    # Filtrar tickets baseado na hierarquia do usu√°rio (mesma l√≥gica)
    if user.can_view_all_tickets():
        tickets = Ticket.objects.all()
    elif user.can_view_sector_tickets():
        user_sectors = list(user.sectors.all())
        if user.sector:
            user_sectors.append(user.sector)
        
        tickets = Ticket.objects.filter(
            base_filter |
            models.Q(sector__in=user_sectors) |
            models.Q(assigned_to=user) |
            models.Q(additional_assignments__user=user, additional_assignments__is_active=True)
        ).distinct()
    else:
        tickets = Ticket.objects.filter(
            base_filter |
            models.Q(assigned_to=user) |
            models.Q(additional_assignments__user=user, additional_assignments__is_active=True)
        ).exclude(status='FECHADO').distinct()
    
    # Aplicar filtros (copiado da tickets_list_view)
    
    # Filtro por origem
    if origem_filter == 'meus':
        tickets = tickets.filter(created_by=user)
    elif origem_filter == 'setor':
        user_sectors = list(user.sectors.all())
        if user.sector:
            user_sectors.append(user.sector)
        tickets = tickets.filter(sector__in=user_sectors).exclude(created_by=user)
    
    # Filtro por chamados atribu√≠dos
    if atribuidos_filter == 'sim':
        tickets = tickets.filter(
            models.Q(assigned_to=user) |
            models.Q(additional_assignments__user=user, additional_assignments__is_active=True)
        ).distinct()
    
    # Filtro por status
    if status_filter:
        if 'abertos' in status_filter:
            tickets = tickets.filter(status='ABERTO')
        elif 'nao_resolvidos' in status_filter:
            tickets = tickets.exclude(status__in=['RESOLVIDO', 'FECHADO'])
        else:
            valid_statuses = [s for s in status_filter if s in ['ABERTO', 'EM_ANDAMENTO', 'RESOLVIDO', 'FECHADO']]
            if valid_statuses:
                tickets = tickets.filter(status__in=valid_statuses)
    
    # Filtro por categoria
    if categoria_filter:
        tickets = tickets.filter(category_id=categoria_filter)
    
    # Filtro por setor
    if setor_filter:
        if user.hierarchy == 'SUPERADMIN' or user.can_view_all_tickets():
            tickets = tickets.filter(sector_id=setor_filter)
        else:
            user_sectors = list(user.sectors.all())
            if user.sector:
                user_sectors.append(user.sector)
            sector_ids = [s.id for s in user_sectors] if user_sectors else []
            if int(setor_filter) in sector_ids:
                tickets = tickets.filter(sector_id=setor_filter)
    
    # Filtro por prioridade
    if prioridade_filter:
        tickets = tickets.filter(priority=prioridade_filter)
    
    # Filtro por carteira
    if carteira_filter:
        from communications.models import CommunicationGroup
        try:
            carteira_group = CommunicationGroup.objects.get(id=carteira_filter)
            carteira_users = carteira_group.members.filter(is_active=True)
            carteira_sectors = Sector.objects.filter(
                models.Q(users__in=carteira_users) |
                models.Q(primary_users__in=carteira_users)
            ).distinct()
            
            tickets = tickets.filter(
                models.Q(sector__in=carteira_sectors) |
                models.Q(assigned_to__in=carteira_users) |
                models.Q(additional_assignments__user__in=carteira_users, additional_assignments__is_active=True)
            ).distinct()
        except CommunicationGroup.DoesNotExist:
            pass
    
    # Filtro por pesquisa
    if search:
        tickets = tickets.filter(
            models.Q(title__icontains=search) |
            models.Q(description__icontains=search) |
            models.Q(id__icontains=search)
        )
    
    # Aplicar filtros avan√ßados apenas para SUPERADMIN
    if user.hierarchy == 'SUPERADMIN':
        if created_by_filter:
            tickets = tickets.filter(created_by_id=created_by_filter)
        
        if created_by_sector_filter:
            tickets = tickets.filter(
                models.Q(created_by__sector_id=created_by_sector_filter) |
                models.Q(created_by__sectors__id=created_by_sector_filter)
            ).distinct()
        
        if assigned_to_filter:
            if assigned_to_filter == 'unassigned':
                tickets = tickets.filter(assigned_to__isnull=True)
            else:
                tickets = tickets.filter(assigned_to_id=assigned_to_filter)
        
        if date_from_filter:
            from django.utils.dateparse import parse_date
            date_from = parse_date(date_from_filter)
            if date_from:
                tickets = tickets.filter(created_at__date__gte=date_from)
        
        if date_to_filter:
            from django.utils.dateparse import parse_date
            date_to = parse_date(date_to_filter)
            if date_to:
                tickets = tickets.filter(created_at__date__lte=date_to)
        
        if user_hierarchy_filter:
            tickets = tickets.filter(created_by__hierarchy=user_hierarchy_filter)
        
        if has_attachments_filter == 'yes':
            tickets = tickets.filter(attachments__isnull=False).distinct()
        elif has_attachments_filter == 'no':
            tickets = tickets.filter(attachments__isnull=True)
        
        if has_comments_filter == 'yes':
            tickets = tickets.filter(comments__isnull=False).distinct()
        elif has_comments_filter == 'no':
            tickets = tickets.filter(comments__isnull=True)
        
        if overdue_filter:
            now = timezone.now()
            if overdue_filter == 'yes':
                tickets = tickets.filter(
                    models.Q(due_date__lt=now) & 
                    ~models.Q(status__in=['RESOLVIDO', 'FECHADO'])
                )
            elif overdue_filter == 'no':
                tickets = tickets.filter(
                    models.Q(due_date__gte=now) | 
                    models.Q(status__in=['RESOLVIDO', 'FECHADO'])
                )
    
    # Ordenar por data de atualiza√ß√£o
    tickets = tickets.order_by('-updated_at')
    
    # Verificar formato de exporta√ß√£o (padr√£o: Excel)
    export_format = request.GET.get('export', 'excel')
    
    if export_format == 'csv':
        return export_tickets_csv(tickets)
    else:
        return export_tickets_xlsx(tickets)
