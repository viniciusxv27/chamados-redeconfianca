from dataclasses import dataclass
from datetime import date, timedelta
import calendar
import logging
import unicodedata
from typing import Any, Dict, List, Optional, Tuple

import pandas as pd
from django.conf import settings
from django.core.cache import cache
from django.db import transaction
from django.utils import timezone
from openpyxl import load_workbook

from communications.models import CommunicationGroup
from users.models import User, Sector
from .models import SimulatorFactorSet, CoordinatorStoreAccess
from .sql_realizado import get_realized_sales_from_mysql


logger = logging.getLogger(__name__)


ROLE_CONSULTOR = 'consultor'
ROLE_GERENTE = 'gerente'
ROLE_COORDENADOR = 'coordenador'
ROLE_SUPERADMIN = 'superadmin'

# Grupo de comunicação SNIPER (configurado em /users/manage/groups/).
# Coordenadores neste grupo recebem 75% da comissão padrão.
SNIPER_GROUP_ID = 22

# Modos de visualização do simulador (espelham as três áreas das planilhas
# por loja/coordenador: PROJEÇÃO, REALIZADO e SIMULADOR).
VIEW_PROJECAO = 'projecao'
VIEW_REALIZADO = 'realizado'
VIEW_SIMULADOR = 'simulador'

VIEW_CHOICES = [
    (VIEW_PROJECAO, 'Projeção'),
    (VIEW_REALIZADO, 'Realizado'),
    (VIEW_SIMULADOR, 'Simulador'),
]

# Pilares utilizados no formulário do simulador (mesmas chaves de hunter)
SIMULATOR_INPUT_PILLARS = [
    ('movel', 'Móvel'),
    ('fixa', 'Fixa'),
    ('smartphones', 'Smartphones'),
    ('eletronicos_a', 'Eletrônicos - A'),
    ('eletronicos_b', 'Eletrônicos - B'),
    ('essenciais_a', 'Essenciais - A'),
    ('essenciais_b', 'Essenciais - B'),
    ('seguros', 'Seguros'),
    ('sva', 'SVA'),
]

# Versão exibida na UI: Eletrônicos A/B e Essenciais A/B aparecem unificados.
# As chaves continuam usando "_a" para que o backend trate o valor digitado
# como o total do pilar (B fica zero).
SIMULATOR_INPUT_PILLARS_DISPLAY = [
    ('movel', 'Móvel'),
    ('fixa', 'Fixa'),
    ('smartphones', 'Smartphones'),
    ('eletronicos_a', 'Eletrônicos'),
    ('essenciais_a', 'Essenciais'),
    ('seguros', 'Seguros'),
    ('sva', 'SVA'),
]

# Pilares usados nos seletores de Hunter (também unificam A/B).
HUNTER_PILLARS = [
    ('movel', 'Móvel'),
    ('fixa', 'Fixa'),
    ('smartphones', 'Smartphones'),
    ('eletronicos', 'Eletrônicos'),
    ('essenciais', 'Essenciais'),
    ('seguros', 'Seguros'),
    ('sva', 'SVA'),
]

WORKBOOK_FILES = {
    ROLE_CONSULTOR: 'CONSULTOR.xlsx',
    ROLE_GERENTE: 'GERENTE.xlsx',
    ROLE_COORDENADOR: 'COORDENADOR.xlsx',
}


@dataclass(frozen=True)
class FactorRangeSpec:
    key: str
    label: str
    cell_range: str
    columns: Tuple[str, ...] = ()


# Rótulos de coluna padrão por tipo de tabela (mais intuitivo na edição).
_COLS_3_PCT = ('Ating. Mín (%)', 'Ating. Máx (%)', 'Taxa (%)')
_COLS_4_PCT = ('Ating. Mín (%)', 'Ating. Máx (%)', 'Taxa A (%)', 'Taxa B (%)')
_COLS_PDV_3 = ('Ating. PDV Mín (%)', 'Ating. PDV Máx (%)', 'Taxa (%)')
_COLS_PDV_4 = ('Ating. PDV Mín (%)', 'Ating. PDV Máx (%)', 'Taxa A (%)', 'Taxa B (%)')
_COLS_SVA_PREMIUM = ('Qtd Mínima', 'Receita Mín (R$)', 'Taxa (%)')
_COLS_SVA_FIXED = ('Taxa Fixa (%)',)


FACTOR_RANGE_SPECS = {
    ROLE_CONSULTOR: [
        FactorRangeSpec('essenciais_commission', 'Essenciais - Comissão', 'B6:E8', _COLS_4_PCT),
        FactorRangeSpec('essenciais_premium_individual', 'Essenciais - Premiação Individual', 'B9:E10', _COLS_4_PCT),
        FactorRangeSpec('essenciais_premium_pdv', 'Essenciais - Premiação PDV', 'B11:E11', _COLS_PDV_4),
        FactorRangeSpec('movel_commission', 'Móvel - Comissão', 'I4:K6', _COLS_3_PCT),
        FactorRangeSpec('movel_premium_individual', 'Móvel - Premiação Individual', 'I7:K8', _COLS_3_PCT),
        FactorRangeSpec('movel_premium_pdv', 'Móvel - Premiação PDV', 'I9:K9', _COLS_PDV_3),
        FactorRangeSpec('fixa_commission', 'Fixa - Comissão', 'I12:K14', _COLS_3_PCT),
        FactorRangeSpec('fixa_premium_individual', 'Fixa - Premiação Individual', 'I15:K16', _COLS_3_PCT),
        FactorRangeSpec('fixa_premium_pdv', 'Fixa - Premiação PDV', 'I17:K17', _COLS_PDV_3),
        FactorRangeSpec('eletronicos_commission', 'Eletrônicos - Comissão', 'B16:E18', _COLS_4_PCT),
        FactorRangeSpec('eletronicos_premium_individual', 'Eletrônicos - Premiação Individual', 'B19:E20', _COLS_4_PCT),
        FactorRangeSpec('eletronicos_premium_pdv', 'Eletrônicos - Premiação PDV', 'B21:E21', _COLS_PDV_4),
        FactorRangeSpec('smartphones_commission', 'Smartphones - Comissão', 'B25:D26', _COLS_3_PCT),
        FactorRangeSpec('smartphones_premium_individual', 'Smartphones - Premiação Individual', 'B27:D27', _COLS_3_PCT),
        FactorRangeSpec('smartphones_premium_pdv', 'Smartphones - Premiação PDV', 'B28:D28', _COLS_PDV_3),
        FactorRangeSpec('seguros_commission', 'Seguros - Comissão', 'O4:Q6', _COLS_3_PCT),
        FactorRangeSpec('seguros_premium_individual', 'Seguros - Premiação Individual', 'O7:Q8', _COLS_3_PCT),
        FactorRangeSpec('seguros_premium_pdv', 'Seguros - Premiação PDV', 'O9:Q9', _COLS_PDV_3),
        FactorRangeSpec('sva_commission', 'SVA - Comissão (Vendeu Ganhou)', 'Q12:Q12', _COLS_SVA_FIXED),
        FactorRangeSpec('sva_premium_individual', 'SVA - Premiação Individual', 'O15:Q16', _COLS_SVA_PREMIUM),
        FactorRangeSpec('sva_premium_pdv', 'SVA - Premiação PDV', 'O17:Q17', _COLS_SVA_PREMIUM),
    ],
    ROLE_GERENTE: [
        FactorRangeSpec('essenciais_commission', 'Essenciais - Comissão', 'B34:E36', _COLS_4_PCT),
        FactorRangeSpec('essenciais_premium_individual', 'Essenciais - Premiação Individual', 'B37:E38', _COLS_4_PCT),
        FactorRangeSpec('essenciais_premium_pdv', 'Essenciais - Premiação PDV', 'B39:E39', _COLS_PDV_4),
        FactorRangeSpec('movel_commission', 'Móvel - Comissão', 'I32:K34', _COLS_3_PCT),
        FactorRangeSpec('movel_premium_individual', 'Móvel - Premiação Individual', 'I35:K36', _COLS_3_PCT),
        FactorRangeSpec('movel_premium_pdv', 'Móvel - Premiação PDV', 'I37:K37', _COLS_PDV_3),
        FactorRangeSpec('fixa_commission', 'Fixa - Comissão', 'I40:K42', _COLS_3_PCT),
        FactorRangeSpec('fixa_premium_individual', 'Fixa - Premiação Individual', 'I43:K44', _COLS_3_PCT),
        FactorRangeSpec('fixa_premium_pdv', 'Fixa - Premiação PDV', 'I45:K45', _COLS_PDV_3),
        FactorRangeSpec('eletronicos_commission', 'Eletrônicos - Comissão', 'B44:E46', _COLS_4_PCT),
        FactorRangeSpec('eletronicos_premium_individual', 'Eletrônicos - Premiação Individual', 'B47:E48', _COLS_4_PCT),
        FactorRangeSpec('eletronicos_premium_pdv', 'Eletrônicos - Premiação PDV', 'B49:E49', _COLS_PDV_4),
        FactorRangeSpec('smartphones_commission', 'Smartphones - Comissão', 'B53:D54', _COLS_3_PCT),
        FactorRangeSpec('smartphones_premium_individual', 'Smartphones - Premiação Individual', 'B55:D55', _COLS_3_PCT),
        FactorRangeSpec('smartphones_premium_pdv', 'Smartphones - Premiação PDV', 'B56:D56', _COLS_PDV_3),
        FactorRangeSpec('seguros_commission', 'Seguros - Comissão', 'O32:Q34', _COLS_3_PCT),
        FactorRangeSpec('seguros_premium_individual', 'Seguros - Premiação Individual', 'O35:Q36', _COLS_3_PCT),
        FactorRangeSpec('seguros_premium_pdv', 'Seguros - Premiação PDV', 'O37:Q37', _COLS_PDV_3),
        FactorRangeSpec('sva_commission', 'SVA - Comissão (Vendeu Ganhou)', 'Q40:Q40', _COLS_SVA_FIXED),
        FactorRangeSpec('sva_premium_individual', 'SVA - Premiação Individual', 'O43:Q44', _COLS_SVA_PREMIUM),
        FactorRangeSpec('sva_premium_pdv', 'SVA - Premiação PDV', 'O45:Q45', _COLS_SVA_PREMIUM),
    ],
    ROLE_COORDENADOR: [
        FactorRangeSpec('essenciais_commission', 'Essenciais - Comissão', 'B63:E67', _COLS_4_PCT),
        FactorRangeSpec('movel_commission', 'Móvel - Comissão', 'I61:K65', _COLS_3_PCT),
        FactorRangeSpec('fixa_commission', 'Fixa - Comissão', 'I68:K72', _COLS_3_PCT),
        FactorRangeSpec('eletronicos_commission', 'Eletrônicos - Comissão', 'B72:E76', _COLS_4_PCT),
        FactorRangeSpec('smartphones_commission', 'Smartphones - Comissão', 'B80:D83', _COLS_3_PCT),
        FactorRangeSpec('seguros_commission', 'Seguros - Comissão', 'O61:Q65', _COLS_3_PCT),
        FactorRangeSpec('sva_commission', 'SVA - Comissão (Vendeu Ganhou)', 'Q68:Q68', _COLS_SVA_FIXED),
    ],
}


DEFAULT_META_BY_ROLE = {
    ROLE_CONSULTOR: {
        'bonus_6_7_rate': 0.1,
        'hunter2_rate': 0.05,
        'hunter3_rate': 0.15,
    },
    ROLE_GERENTE: {
        'bonus_6_7_rate': 0.1,
        'hunter2_rate': 0.05,
        'hunter3_rate': 0.15,
    },
    ROLE_COORDENADOR: {
        'bonus_6_7_rate': 0.1,
        'hunter2_rate': 0.05,
        'hunter3_rate': 0.05,
        'sniper_rate': 0.75,
    },
}


PILLAR_ORDER = [
    'movel',
    'fixa',
    'smartphones',
    'eletronicos_a',
    'eletronicos_b',
    'essenciais_a',
    'essenciais_b',
    'seguros',
    'sva',
]


PILLAR_LABELS = {
    'movel': 'Móvel',
    'fixa': 'Fixa',
    'smartphones': 'Smartphones',
    'eletronicos_a': 'Eletrônicos - A',
    'eletronicos_b': 'Eletrônicos - B',
    'essenciais_a': 'Essenciais - A',
    'essenciais_b': 'Essenciais - B',
    'seguros': 'Seguros',
    'sva': 'SVA',
}


PILLAR_CONFIG_CONSULTOR = {
    'movel': {
        'meta_col': 'META_MOVEL',
        'proj_col': 'PROJ_MOVEL',
        'commission_key': 'movel_commission',
        'premium_key': 'movel_premium_individual',
        'pdv_premium_key': 'movel_premium_pdv',
        'rate_col': 3,
        'pdv_rate_col': 3,
    },
    'fixa': {
        'meta_col': 'META_FIXA',
        'proj_col': 'PROJ_FIXA',
        'commission_key': 'fixa_commission',
        'premium_key': 'fixa_premium_individual',
        'pdv_premium_key': 'fixa_premium_pdv',
        'rate_col': 3,
        'pdv_rate_col': 3,
    },
    'smartphones': {
        'meta_col': 'META_SMARTPHONE',
        'proj_col': 'PROJ_APARELHO',
        'commission_key': 'smartphones_commission',
        'premium_key': 'smartphones_premium_individual',
        'pdv_premium_key': 'smartphones_premium_pdv',
        'rate_col': 3,
        'pdv_rate_col': 3,
    },
    'eletronicos_a': {
        'meta_col': 'META_ACESSORIO',
        'proj_col': 'PROJ_ELETRO_A',
        'proj_col_b': 'PROJ_ELETRO_B',
        'commission_key': 'eletronicos_commission',
        'premium_key': 'eletronicos_premium_individual',
        'pdv_premium_key': 'eletronicos_premium_pdv',
        'rate_col': 3,
        'pdv_rate_col': 3,
    },
    'eletronicos_b': {
        'meta_col': 'META_ACESSORIO',
        'proj_col': 'PROJ_ELETRO_B',
        'commission_key': 'eletronicos_commission',
        'premium_key': 'eletronicos_premium_individual',
        'pdv_premium_key': 'eletronicos_premium_pdv',
        'rate_col': 4,
        'pdv_rate_col': 4,
    },
    'essenciais_a': {
        'meta_col': 'META_ESSENCIAIS',
        'proj_col': 'PROJ_ESSEN_A',
        'proj_col_b': 'PROJ_ESSEN_B',
        'commission_key': 'essenciais_commission',
        'premium_key': 'essenciais_premium_individual',
        'pdv_premium_key': 'essenciais_premium_pdv',
        'rate_col': 3,
        'pdv_rate_col': 3,
    },
    'essenciais_b': {
        'meta_col': 'META_ESSENCIAIS',
        'proj_col': 'PROJ_ESSEN_B',
        'commission_key': 'essenciais_commission',
        'premium_key': 'essenciais_premium_individual',
        'pdv_premium_key': 'essenciais_premium_pdv',
        'rate_col': 4,
        'pdv_rate_col': 4,
    },
    'seguros': {
        'meta_col': 'META_SEGUROS',
        'proj_col': 'PROJ_SEGURO',
        'commission_key': 'seguros_commission',
        'premium_key': 'seguros_premium_individual',
        'pdv_premium_key': 'seguros_premium_pdv',
        'rate_col': 3,
        'pdv_rate_col': 3,
    },
    'sva': {
        'meta_col': 'META_SVA',
        'proj_col': 'PROJ_SVA',
        'commission_key': 'sva_commission',
        'premium_key': 'sva_premium_individual',
        'pdv_premium_key': 'sva_premium_pdv',
        'rate_col': 3,
        'pdv_rate_col': 3,
    },
}

PILLAR_CONFIG_GERENTE = {
    **PILLAR_CONFIG_CONSULTOR,
}

PILLAR_CONFIG_COORDENADOR = {
    'movel': {
        'meta_col': 'META_MOVEL',
        'proj_col': 'PROJ_MOVEL',
        'commission_key': 'movel_commission',
        'rate_col': 3,
    },
    'fixa': {
        'meta_col': 'META_FIXA',
        'proj_col': 'PROJ_FIXA',
        'commission_key': 'fixa_commission',
        'rate_col': 3,
    },
    'smartphones': {
        'meta_col': 'META_SMARTPHONE',
        'proj_col': 'PROJ_APARELHO',
        'commission_key': 'smartphones_commission',
        'rate_col': 3,
    },
    'eletronicos_a': {
        'meta_col': 'META_ACESSORIO',
        'proj_col': 'PROJ_ELETRO_A',
        'proj_col_b': 'PROJ_ELETRO_B',
        'commission_key': 'eletronicos_commission',
        'rate_col': 3,
    },
    'eletronicos_b': {
        'meta_col': 'META_ACESSORIO',
        'proj_col': 'PROJ_ELETRO_B',
        'commission_key': 'eletronicos_commission',
        'rate_col': 4,
    },
    'essenciais_a': {
        'meta_col': 'META_ESSENCIAIS',
        'proj_col': 'PROJ_ESSEN_A',
        'proj_col_b': 'PROJ_ESSEN_B',
        'commission_key': 'essenciais_commission',
        'rate_col': 3,
    },
    'essenciais_b': {
        'meta_col': 'META_ESSENCIAIS',
        'proj_col': 'PROJ_ESSEN_B',
        'commission_key': 'essenciais_commission',
        'rate_col': 4,
    },
    'seguros': {
        'meta_col': 'META_SEGUROS',
        'proj_col': 'PROJ_SEGURO',
        'commission_key': 'seguros_commission',
        'rate_col': 3,
    },
    'sva': {
        'meta_col': 'META_SVA',
        'proj_col': 'PROJ_SVA',
        'commission_key': 'sva_commission',
        'rate_col': 1,
    },
}


def normalize_text(value: Any) -> str:
    raw = str(value or '').strip().upper()
    raw = unicodedata.normalize('NFKD', raw)
    return ''.join(ch for ch in raw if not unicodedata.combining(ch))


def to_float(value: Any) -> float:
    try:
        if value is None:
            return 0.0
        if isinstance(value, str):
            s = value.strip()
            if not s:
                return 0.0
            # remove prefixo monetário e espaços
            s = s.replace('R$', '').replace('r$', '').replace(' ', '')
            # aceita formato brasileiro "1.234,56" e formato simples "1234.56"
            if ',' in s:
                s = s.replace('.', '').replace(',', '.')
            return float(s)
        return float(value)
    except (TypeError, ValueError):
        return 0.0


def get_workbook_path(role: str) -> str:
    filename = WORKBOOK_FILES.get(role)
    if not filename:
        raise ValueError(f"Arquivo de planilha não definido para role {role}.")
    return str(settings.BASE_DIR / filename)


def read_range(ws, cell_range: str) -> List[List[Optional[float]]]:
    values: List[List[Optional[float]]] = []
    for row in ws[cell_range]:
        row_values = []
        for cell in row:
            if cell.value is None:
                row_values.append(None)
            else:
                try:
                    row_values.append(float(cell.value))
                except (TypeError, ValueError):
                    row_values.append(None)
        values.append(row_values)
    return values


def load_default_factor_data(role: str) -> Dict[str, Any]:
    path = get_workbook_path(role)
    workbook = load_workbook(path, data_only=True)
    sheet = workbook['FATORES']
    ranges = {}
    for spec in FACTOR_RANGE_SPECS[role]:
        ranges[spec.key] = read_range(sheet, spec.cell_range)

    meta = DEFAULT_META_BY_ROLE.get(role, {}).copy()
    simulator_sheet = workbook.sheetnames[4] if len(workbook.sheetnames) > 4 else workbook.sheetnames[-1]
    simulator_ws = workbook[simulator_sheet]

    if role in (ROLE_CONSULTOR, ROLE_GERENTE):
        bonus_value = simulator_ws['S15'].value
        meta['bonus_6_7_rate'] = to_float(bonus_value) or meta.get('bonus_6_7_rate', 0.1)
    elif role == ROLE_COORDENADOR:
        bonus_value = simulator_ws['J15'].value
        meta['bonus_6_7_rate'] = to_float(bonus_value) or meta.get('bonus_6_7_rate', 0.1)

    return {
        'ranges': ranges,
        'meta': meta,
    }


def get_factor_set(role: str, updated_by: Optional[User] = None) -> SimulatorFactorSet:
    defaults = load_default_factor_data(role)
    factor_set, created = SimulatorFactorSet.objects.get_or_create(
        role=role,
        defaults={
            'data': defaults,
            'updated_by': updated_by,
        },
    )

    if not created:
        data = factor_set.data or {}
        ranges = data.get('ranges', {})
        meta = data.get('meta', {})
        refreshed = False
        for spec in FACTOR_RANGE_SPECS[role]:
            if spec.key not in ranges:
                ranges[spec.key] = defaults['ranges'].get(spec.key, [])
                refreshed = True
        for key, value in defaults.get('meta', {}).items():
            if key not in meta:
                meta[key] = value
                refreshed = True
        if refreshed:
            factor_set.data = {'ranges': ranges, 'meta': meta}
            if updated_by:
                factor_set.updated_by = updated_by
            factor_set.save()

    return factor_set


def load_dataframe(role: str, sheet_name: str) -> pd.DataFrame:
    cache_key = f"simulator_df_{role}_{sheet_name}"
    df = cache.get(cache_key)
    if df is None:
        df = pd.read_excel(get_workbook_path(role), sheet_name=sheet_name, engine='openpyxl')
        cache.set(cache_key, df, 300)
    return df


def find_row_by_name(df: pd.DataFrame, col: str, name: str) -> Optional[pd.Series]:
    if col not in df.columns:
        return None
    target = normalize_text(name)
    for _, row in df.iterrows():
        value = normalize_text(row.get(col, ''))
        if value == target:
            return row
    for _, row in df.iterrows():
        value = normalize_text(row.get(col, ''))
        if target and value and (target in value or value in target):
            return row
    return None


def xlookup(df: pd.DataFrame, key_col: str, key: str, target_col: str) -> float:
    row = find_row_by_name(df, key_col, key)
    if row is None:
        return 0.0
    return to_float(row.get(target_col))


def sumifs(df: pd.DataFrame, sum_col: str, filter_col: str, filter_value: str) -> float:
    if sum_col not in df.columns or filter_col not in df.columns:
        return 0.0
    target = normalize_text(filter_value)
    total = 0.0
    for _, row in df.iterrows():
        value = normalize_text(row.get(filter_col, ''))
        if value == target:
            total += to_float(row.get(sum_col))
    return total


def get_pdvs_of_coord(df: pd.DataFrame, coord_name: str) -> List[str]:
    """Lista PDVs únicos (normalizados) da planilha que pertencem a uma COORDENAÇÃO."""
    if not coord_name or 'COORDENAÇÃO' not in df.columns or 'PDV' not in df.columns:
        return []
    target = normalize_text(coord_name)
    seen: set = set()
    result: List[str] = []
    for _, row in df.iterrows():
        if normalize_text(row.get('COORDENAÇÃO', '')) != target:
            continue
        pdv_val = str(row.get('PDV') or '').strip()
        if not pdv_val:
            continue
        key = normalize_text(pdv_val)
        if key in seen:
            continue
        seen.add(key)
        result.append(pdv_val)
    return result


def vlookup(value: float, table: List[List[Optional[float]]], col_index: int) -> float:
    if not table:
        return 0.0
    best_row = None
    for row in table:
        if not row:
            continue
        threshold = row[0] if row[0] is not None else None
        if threshold is None:
            continue
        if value >= threshold:
            best_row = row
    if best_row is None:
        return 0.0
    idx = col_index - 1
    if idx < 0 or idx >= len(best_row):
        return 0.0
    return to_float(best_row[idx])


def pdv_threshold_rate(table: List[List[Optional[float]]], col_index: int) -> Tuple[float, float]:
    if not table or not table[0]:
        return 0.0, 0.0
    threshold = to_float(table[0][0])
    idx = col_index - 1
    rate = 0.0
    if 0 <= idx < len(table[0]):
        rate = to_float(table[0][idx])
    return threshold, rate


def get_user_role(user: User) -> str:
    if user.is_superuser or getattr(user, 'hierarchy', None) == 'SUPERADMIN':
        return ROLE_SUPERADMIN

    # Grupo COORDENADORES tem prioridade (membros podem ter qualquer hierarquia).
    coord_group = CommunicationGroup.objects.filter(name__icontains='COORDENADORES').first()
    if coord_group and coord_group.members.filter(id=user.id).exists():
        return ROLE_COORDENADOR

    if getattr(user, 'hierarchy', None) == 'SUPERVISOR':
        primary_sector = getattr(user, 'primary_sector', None)
        sector_name = (primary_sector.name if primary_sector else '').lower()
        if 'comercial' in sector_name:
            return ROLE_COORDENADOR

    if getattr(user, 'hierarchy', None) == 'PADRAO':
        gerente_group = CommunicationGroup.objects.filter(name__icontains='GERENTES').first()
        if gerente_group and gerente_group.members.filter(id=user.id).exists():
            return ROLE_GERENTE
        return ROLE_CONSULTOR

    return ROLE_CONSULTOR


def get_coordinator_sectors(user: User) -> List[Sector]:
    access = CoordinatorStoreAccess.objects.filter(coordinator=user).first()
    if access and access.sectors.exists():
        return list(access.sectors.all())
    return list(Sector.objects.filter(name__icontains='loja'))


def _group_member_ids(name_filter: str, exact: bool = False) -> List[int]:
    """Retorna ids dos membros de um CommunicationGroup pelo nome."""
    qs = CommunicationGroup.objects.all()
    if exact:
        qs = qs.filter(name__iexact=name_filter)
    else:
        qs = qs.filter(name__icontains=name_filter)
    group = qs.first()
    if not group:
        return []
    return list(group.members.values_list('id', flat=True))


def get_simulator_excluded_user_ids() -> set:
    """IDs de usuários que devem ser ocultados na seleção do /simulator:
    - PADRÃO que pertencem aos grupos ADMINS ou RECEPCIONISTAS
    """
    admin_ids = set(_group_member_ids('ADMINS', exact=True))
    recep_ids = set(_group_member_ids('RECEPCIONISTAS'))

    excluded: set = set()
    padrao_in_admin_or_recep = User.objects.filter(
        hierarchy='PADRAO', id__in=(admin_ids | recep_ids)
    ).values_list('id', flat=True)
    excluded.update(padrao_in_admin_or_recep)
    return excluded


def get_all_consultors() -> List[User]:
    gerente_group = CommunicationGroup.objects.filter(name__icontains='GERENTES').first()
    gerente_ids = []
    if gerente_group:
        gerente_ids = list(gerente_group.members.values_list('id', flat=True))
    coord_group = CommunicationGroup.objects.filter(name__icontains='COORDENADORES').first()
    coord_ids = []
    if coord_group:
        coord_ids = list(coord_group.members.values_list('id', flat=True))
    excluded = get_simulator_excluded_user_ids()
    return list(
        User.objects.filter(hierarchy='PADRAO', is_active=True)
        .exclude(id__in=gerente_ids)
        .exclude(id__in=coord_ids)
        .exclude(id__in=excluded)
        .order_by('first_name', 'last_name')
    )


def get_all_gerentes() -> List[User]:
    gerente_group = CommunicationGroup.objects.filter(name__icontains='GERENTES').first()
    if not gerente_group:
        return []
    coord_group = CommunicationGroup.objects.filter(name__icontains='COORDENADORES').first()
    coord_ids = []
    if coord_group:
        coord_ids = list(coord_group.members.values_list('id', flat=True))
    excluded = get_simulator_excluded_user_ids()
    return list(
        gerente_group.members.filter(is_active=True)
        .exclude(id__in=coord_ids)
        .exclude(id__in=excluded)
        .order_by('first_name', 'last_name')
    )


def get_all_coordinators() -> List[User]:
    """Coordenadores = membros do grupo COORDENADORES (CommunicationGroup)."""
    coord_group = CommunicationGroup.objects.filter(name__icontains='COORDENADORES').first()
    if not coord_group:
        return []
    excluded = get_simulator_excluded_user_ids()
    return list(
        coord_group.members.filter(is_active=True)
        .exclude(id__in=excluded)
        .order_by('first_name', 'last_name')
    )


def get_hunter_levels_from_request(request) -> Dict[str, int]:
    levels = {}
    for key in ['movel', 'fixa', 'smartphones', 'eletronicos', 'essenciais', 'seguros', 'sva']:
        value = request.GET.get(f'hunter_{key}', request.POST.get(f'hunter_{key}', '0'))
        try:
            levels[key] = int(value)
        except (TypeError, ValueError):
            levels[key] = 0
    return levels


def all_pillars_ok(att_map: Dict[str, float], coordinator_name: str, threshold: float) -> bool:
    is_ariel = normalize_text(coordinator_name) == 'ARIEL'
    if is_ariel:
        required = ['movel', 'smartphones', 'eletronicos', 'essenciais', 'seguros', 'sva']
    else:
        required = ['movel', 'fixa', 'smartphones', 'eletronicos', 'essenciais', 'seguros', 'sva']
    return all(att_map.get(key, 0.0) >= threshold for key in required)


def bonus_6_7_ok(att_map: Dict[str, float], coordinator_name: str) -> bool:
    is_ariel = normalize_text(coordinator_name) == 'ARIEL'
    if is_ariel:
        required = ['movel', 'smartphones', 'eletronicos', 'essenciais', 'seguros', 'sva']
    else:
        required = ['movel', 'fixa', 'smartphones', 'eletronicos', 'essenciais', 'seguros', 'sva']
    if not required:
        return False
    return min(att_map.get(key, 0.0) for key in required) >= 1.0


def build_pillar_values(meta: float, proj: float) -> Tuple[float, float, float]:
    attainment = proj / meta if meta else 0.0
    return meta, proj, attainment


def build_group_values(meta: float, proj_a: float, proj_b: float) -> Tuple[float, float, float, float]:
    total_proj = proj_a + proj_b
    attainment = total_proj / meta if meta else 0.0
    return meta, proj_a, proj_b, attainment


def _easter_sunday(year: int) -> date:
    """Calcula o domingo de Páscoa (algoritmo de Meeus/Jones/Butcher)."""
    a = year % 19
    b = year // 100
    c = year % 100
    d = b // 4
    e = b % 4
    f = (b + 8) // 25
    g = (b - f + 1) // 3
    h = (19 * a + b - d - g + 15) % 30
    i = c // 4
    k = c % 4
    l = (32 + 2 * e + 2 * i - h - k) % 7
    m = (a + 11 * h + 22 * l) // 451
    month = (h + l - 7 * m + 114) // 31
    day = ((h + l - 7 * m + 114) % 31) + 1
    return date(year, month, day)


def _br_national_holidays(year: int) -> set:
    """Feriados nacionais brasileiros (fixos + móveis baseados na Páscoa)."""
    easter = _easter_sunday(year)
    return {
        date(year, 1, 1),   # Confraternização Universal
        easter - timedelta(days=48),  # Carnaval (segunda)
        easter - timedelta(days=47),  # Carnaval (terça)
        easter - timedelta(days=2),   # Sexta-feira Santa
        date(year, 4, 21),  # Tiradentes
        date(year, 5, 1),   # Dia do Trabalho
        easter + timedelta(days=60),  # Corpus Christi
        date(year, 9, 7),   # Independência
        date(year, 10, 12), # N. S. Aparecida
        date(year, 11, 2),  # Finados
        date(year, 11, 15), # Proclamação da República
        date(year, 12, 25), # Natal
    }


def get_business_days_info(reference_date: Optional[date] = None) -> Tuple[int, int]:
    """Retorna (dias_uteis_passados_ate_ontem, dias_uteis_totais_no_mes).

    Considera seg-sex e exclui feriados nacionais brasileiros (fixos + móveis).
    Usa sempre D-1 (ontem) como corte — o dia atual não conta porque os
    dados de venda só ficam consolidados no fechamento do dia anterior.
    """
    if reference_date is None:
        reference_date = timezone.localdate()
    cutoff = reference_date - timedelta(days=1)
    year, month = reference_date.year, reference_date.month
    _, days_in_month = calendar.monthrange(year, month)
    br_holidays = _br_national_holidays(year)

    total_du = 0
    passed_du = 0
    for day in range(1, days_in_month + 1):
        d = date(year, month, day)
        if d.weekday() >= 5:  # sáb/dom
            continue
        if d in br_holidays:
            continue
        total_du += 1
        if d <= cutoff:
            passed_du += 1
    return passed_du, total_du


def project_from_realized(realized_value: float, du_passed: int, du_total: int) -> float:
    """Projeta valor mensal: realizado / DU_passados * DU_totais.

    Se ainda não houve dia útil no mês, retorna o próprio realizado.
    """
    if du_passed <= 0:
        return float(realized_value or 0.0)
    return float(realized_value or 0.0) / du_passed * du_total


def _get_sim_input(simulator_inputs: Optional[Dict[str, Any]], pillar: str, field: str, default: float = 0.0) -> float:
    """Lê um valor enviado pelo formulário 'Simulador' para um pilar."""
    if not simulator_inputs:
        return default
    raw = simulator_inputs.get(f"{pillar}__{field}")
    if raw is None or raw == '':
        return default
    return to_float(raw)


def _get_sim_input_optional(simulator_inputs: Optional[Dict[str, Any]], pillar: str, field: str) -> Optional[float]:
    """Igual ao _get_sim_input, mas devolve None se o usuário não preencheu."""
    if not simulator_inputs:
        return None
    raw = simulator_inputs.get(f"{pillar}__{field}")
    if raw is None or raw == '':
        return None
    return to_float(raw)


# ---------------------------------------------------------------------------
# Metas vindas do módulo Power BI (/power-bi/metas/)
# ---------------------------------------------------------------------------

POWER_BI_PILAR_MAP = {
    'MOVEL': 'movel',
    'MÓVEL': 'movel',
    'FIXA': 'fixa',
    'SMARTPHONE': 'smartphones',
    'SMARTPHONES': 'smartphones',
    'ELETRONICOS': 'eletronicos',
    'ELETRÔNICOS': 'eletronicos',
    'ESSENCIAIS': 'essenciais',
    'SEGURO': 'seguros',
    'SEGUROS': 'seguros',
    'SVA': 'sva',
}


def get_store_name_from_user(user) -> str:
    """Extrai o nome da loja a partir do setor do usuário.

    Aceita setores como "Loja Anchieta", "LOJA - JARDIM CAMBURI", etc.
    Retorna a string normalizada (maiúsculas, sem acentos extras) usada nas
    planilhas Power BI / Simulador.
    """
    sector = getattr(user, 'sector', None) or getattr(user, 'primary_sector', None)
    if not sector:
        return ''
    raw = (sector.name or '').strip()
    if not raw:
        return ''
    lowered = raw.lower()
    if 'loja' not in lowered:
        return raw.upper()
    cleaned = raw
    for prefix in ('Loja', 'LOJA', 'loja'):
        if cleaned.startswith(prefix):
            cleaned = cleaned[len(prefix):]
            break
    cleaned = cleaned.lstrip(' -–—:').strip()
    return cleaned.upper()


def _latest_goal_upload():
    from power_bi.models import GoalUpload
    return GoalUpload.objects.order_by('-year', '-month').first()


def get_metas_from_power_bi(user_name: str = '', store_name: str = '') -> Dict[str, float]:
    """Devolve metas (chave: pilar do simulador) para o consultor ou para o PDV.

    - Se ``user_name`` for fornecido, busca em METAS_CN_REAL.
    - Se ``store_name`` for fornecido, busca em META_PDV_REAL.
    - Pilares ausentes ficam em 0.0.
    """
    from power_bi.models import GoalEntry

    upload = _latest_goal_upload()
    if not upload:
        return {}

    qs = GoalEntry.objects.filter(upload=upload)
    if user_name:
        qs = qs.filter(sheet_type='METAS_CN_REAL')
        target = normalize_text(user_name)
        rows = [g for g in qs if normalize_text(g.user_name) == target]
    elif store_name:
        qs = qs.filter(sheet_type='META_PDV_REAL')
        target = normalize_text(store_name)
        rows = [g for g in qs if normalize_text(g.store_name) == target]
    else:
        return {}

    result: Dict[str, float] = {}
    for entry in rows:
        pilar_key = POWER_BI_PILAR_MAP.get((entry.pilar or '').upper().strip())
        if not pilar_key:
            continue
        try:
            value = float(entry.goal_value or 0)
        except (TypeError, ValueError):
            value = 0.0
        result[pilar_key] = result.get(pilar_key, 0.0) + value
    return result


def get_pdv_metas_for_coordinator(coord_name: str) -> Dict[str, float]:
    """Soma metas de PDV (META_PDV_REAL) para todos os PDVs do coordenador."""
    from power_bi.models import GoalEntry

    upload = _latest_goal_upload()
    if not upload:
        return {}

    target = normalize_text(coord_name)
    result: Dict[str, float] = {}
    qs = GoalEntry.objects.filter(upload=upload, sheet_type='META_PDV_REAL')
    for entry in qs:
        coord_in_row = normalize_text((entry.row_data or {}).get('COORDENAÇÃO') or '')
        if coord_in_row != target:
            continue
        pilar_key = POWER_BI_PILAR_MAP.get((entry.pilar or '').upper().strip())
        if not pilar_key:
            continue
        try:
            value = float(entry.goal_value or 0)
        except (TypeError, ValueError):
            value = 0.0
        result[pilar_key] = result.get(pilar_key, 0.0) + value
    return result


# ---------------------------------------------------------------------------
# Pós-processamento para unir Eletrônicos A/B e Essenciais A/B em uma única
# linha de exibição (mantendo os cálculos internos com taxas distintas).
# ---------------------------------------------------------------------------

_MERGE_GROUPS = {
    'eletronicos': ('eletronicos_a', 'eletronicos_b', 'Eletrônicos'),
    'essenciais': ('essenciais_a', 'essenciais_b', 'Essenciais'),
}


def _merge_grouped_rows(rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """Une linhas A+B (Eletrônicos / Essenciais) em uma única linha somada."""
    by_key = {row.get('key'): row for row in rows}
    new_rows: List[Dict[str, Any]] = []
    consumed = set()

    for row in rows:
        key = row.get('key')
        if key in consumed:
            continue
        merge_pair = None
        for base, (key_a, key_b, label) in _MERGE_GROUPS.items():
            if key == key_a:
                merge_pair = (base, key_a, key_b, label)
                break
            if key == key_b:
                merge_pair = (base, key_a, key_b, label)
                break
        if not merge_pair:
            new_rows.append(row)
            continue

        base, key_a, key_b, label = merge_pair
        a = by_key.get(key_a)
        b = by_key.get(key_b)
        consumed.add(key_a)
        consumed.add(key_b)
        if a is None and b is None:
            continue

        def _sum(field, default=0.0):
            va = (a or {}).get(field) if a else None
            vb = (b or {}).get(field) if b else None
            va = va if isinstance(va, (int, float)) else default
            vb = vb if isinstance(vb, (int, float)) else default
            return va + vb

        meta = (a or {}).get('meta') if a else (b or {}).get('meta')
        if meta is None:
            meta = (b or {}).get('meta', 0.0)
        proj = _sum('proj')
        attainment = (proj / meta) if meta else 0.0
        commission_value = _sum('commission_value')
        premium_value = _sum('premium_value')
        total_individual = _sum('total_individual')
        pdv_premium_value = _sum('pdv_premium_value')
        total_with_pdv = _sum('total_with_pdv')
        hunter2_value = _sum('hunter2_value')
        hunter3_value = _sum('hunter3_value')
        # Taxas efetivas combinadas (R$ / base)
        commission_rate = (commission_value / proj) if proj else 0.0
        premium_rate = (premium_value / proj) if proj else 0.0

        merged = {
            'key': base,
            'label': label,
            'meta': meta or 0.0,
            'proj': proj,
            'attainment': attainment,
            'commission_rate': commission_rate,
            'premium_rate': premium_rate,
            'commission_value': commission_value,
            'premium_value': premium_value,
            'total_individual': total_individual,
            'pdv_meta': (a or {}).get('pdv_meta') if a else (b or {}).get('pdv_meta'),
            'pdv_proj': (a or {}).get('pdv_proj') if a else (b or {}).get('pdv_proj'),
            'pdv_attainment': (a or {}).get('pdv_attainment') if a else (b or {}).get('pdv_attainment'),
            'pdv_premium_rate': (a or {}).get('pdv_premium_rate') if a else (b or {}).get('pdv_premium_rate'),
            'pdv_premium_value': pdv_premium_value,
            'total_with_pdv': total_with_pdv,
            'coord_meta': (a or {}).get('coord_meta') if a else (b or {}).get('coord_meta'),
            'coord_proj': (a or {}).get('coord_proj') if a else (b or {}).get('coord_proj'),
            'coord_attainment': (a or {}).get('coord_attainment') if a else (b or {}).get('coord_attainment'),
            'hunter2_value': hunter2_value,
            'hunter3_value': hunter3_value,
        }
        new_rows.append(merged)

    return new_rows



def compute_consultor_simulation(
    user: User,
    factor_data: Dict[str, Any],
    hunter_levels: Optional[Dict[str, int]] = None,
    view_mode: str = VIEW_PROJECAO,
    simulator_inputs: Optional[Dict[str, Any]] = None,
) -> Dict[str, Any]:
    realized = load_dataframe(ROLE_CONSULTOR, 'REALIZADO')
    projection = load_dataframe(ROLE_CONSULTOR, 'PROJEÇÃO')

    user_name = user.get_full_name() or user.first_name or user.email
    real_row = find_row_by_name(realized, 'CONSULTOR', user_name)
    proj_row = find_row_by_name(projection, 'CONSULTOR', user_name)

    # Se o usuário não existir na planilha, ainda permitimos o cálculo: as
    # informações ficam em zero (modo Realizado/Projeção) ou vêm dos inputs do
    # próprio usuário (modo Simulador). O motor de comissionamento aplica os
    # mesmos fatores carregados do XLSX.
    user_in_sheet = real_row is not None or proj_row is not None
    if real_row is None:
        real_row = pd.Series(dtype=object)
    if proj_row is None:
        proj_row = pd.Series(dtype=object)

    pdv = str(real_row.get('PDV') or proj_row.get('PDV') or '').strip()
    if not pdv and getattr(user, 'sector', None):
        pdv = user.sector.name or ''
    coord_name = str(proj_row.get('COORDENAÇÃO') or '').strip()

    # Metas vêm sempre da planilha REALIZADO (META_*), mas no modo Simulador
    # podem ser sobrescritas pelo usuário.
    base_meta = {
        'movel': to_float(real_row.get('META_MOVEL')),
        'fixa': to_float(real_row.get('META_FIXA')),
        'smartphones': to_float(real_row.get('META_SMARTPHONE')),
        'eletronicos': to_float(real_row.get('META_ACESSORIO')),
        'essenciais': to_float(real_row.get('META_ESSENCIAIS')),
        'seguros': to_float(real_row.get('META_SEGUROS')),
        'sva': to_float(real_row.get('META_SVA')),
    }

    # Sobrescreve com as metas oficiais cadastradas em /power-bi/metas/.
    pb_metas = get_metas_from_power_bi(user_name=user_name)
    if not pb_metas:
        store_name = get_store_name_from_user(user) or pdv
        if store_name:
            pb_metas = get_metas_from_power_bi(store_name=store_name)
    for k, v in (pb_metas or {}).items():
        if v:
            base_meta[k] = v

    # Valores do indivíduo conforme o modo selecionado.
    if view_mode == VIEW_REALIZADO:
        # Realizado individual: agora vem do MySQL (vendas_produtos + vendas_servicos).
        mysql_real = get_realized_sales_from_mysql(vendor=user_name)
        ind_values = {
            'movel': mysql_real.get('movel', 0.0),
            'fixa': mysql_real.get('fixa', 0.0),
            'smartphones': mysql_real.get('smartphones', 0.0),
            # MySQL não separa A/B: todo o valor vai em _a, _b = 0 (a soma A+B é o que importa).
            'eletronicos_a': mysql_real.get('eletronicos', 0.0),
            'eletronicos_b': 0.0,
            'essenciais_a': mysql_real.get('essenciais', 0.0),
            'essenciais_b': 0.0,
            'seguros': mysql_real.get('seguros', 0.0),
            'sva': mysql_real.get('sva', 0.0),
        }
        fixa_quantity = mysql_real.get('fixa_qty', 0.0)
        fixa_revenue = ind_values['fixa']
    elif view_mode == VIEW_SIMULADOR:
        # Simulador: usa valores informados pelo usuário.
        # Aceita chaves unificadas para Eletrônicos/Essenciais (campo "_a" recebe o total).
        ind_values = {p: _get_sim_input(simulator_inputs, p, 'real') for p, _ in SIMULATOR_INPUT_PILLARS}
        # Permite sobrescrever metas individuais
        for key in base_meta:
            override = _get_sim_input_optional(simulator_inputs, key, 'meta')
            if override is not None:
                base_meta[key] = override
        fixa_quantity = _get_sim_input(simulator_inputs, 'fixa', 'qty')
        # Receita estimada de Fixa = qtd simulada × ticket médio do realizado até o momento.
        _real_consultor = get_realized_sales_from_mysql(vendor=user_name)
        _real_fixa_qty = _real_consultor.get('fixa_qty', 0.0) or 0.0
        _real_fixa_rev = _real_consultor.get('fixa', 0.0) or 0.0
        _ticket_medio_fixa = (_real_fixa_rev / _real_fixa_qty) if _real_fixa_qty > 0 else 0.0
        fixa_revenue = fixa_quantity * _ticket_medio_fixa
        ind_values['fixa'] = fixa_revenue
    else:  # VIEW_PROJECAO
        # Projeção dinâmica: pega o realizado do MySQL e projeta pelo DU.
        # Fórmula: projeção = realizado / DU_passados_até_hoje * DU_totais_do_mês
        mysql_real = get_realized_sales_from_mysql(vendor=user_name)
        du_passed, du_total = get_business_days_info()
        ind_values = {
            'movel': project_from_realized(mysql_real.get('movel', 0.0), du_passed, du_total),
            'fixa': project_from_realized(mysql_real.get('fixa', 0.0), du_passed, du_total),
            'smartphones': project_from_realized(mysql_real.get('smartphones', 0.0), du_passed, du_total),
            # MySQL não separa A/B: todo o valor projetado vai em _a, _b = 0.
            'eletronicos_a': project_from_realized(mysql_real.get('eletronicos', 0.0), du_passed, du_total),
            'eletronicos_b': 0.0,
            'essenciais_a': project_from_realized(mysql_real.get('essenciais', 0.0), du_passed, du_total),
            'essenciais_b': 0.0,
            'seguros': project_from_realized(mysql_real.get('seguros', 0.0), du_passed, du_total),
            'sva': project_from_realized(mysql_real.get('sva', 0.0), du_passed, du_total),
        }
        fixa_quantity = project_from_realized(mysql_real.get('fixa_qty', 0.0), du_passed, du_total)
        fixa_revenue = ind_values['fixa']

    meta_movel, proj_movel, att_movel = build_pillar_values(base_meta['movel'], ind_values['movel'])
    meta_fixa, proj_fixa, att_fixa = build_pillar_values(base_meta['fixa'], ind_values['fixa'])
    # Fixa: o atingimento é calculado pela quantidade vendida ÷ meta (que também está em quantidade).
    att_fixa = (fixa_quantity / meta_fixa) if meta_fixa else 0.0
    meta_smart, proj_smart, att_smart = build_pillar_values(base_meta['smartphones'], ind_values['smartphones'])
    meta_eletro, proj_eletro_a, proj_eletro_b, att_eletro = build_group_values(
        base_meta['eletronicos'], ind_values['eletronicos_a'], ind_values['eletronicos_b'],
    )
    meta_ess, proj_ess_a, proj_ess_b, att_ess = build_group_values(
        base_meta['essenciais'], ind_values['essenciais_a'], ind_values['essenciais_b'],
    )
    meta_seg, proj_seg, att_seg = build_pillar_values(base_meta['seguros'], ind_values['seguros'])
    meta_sva, proj_sva, att_sva = build_pillar_values(base_meta['sva'], ind_values['sva'])

    att_map = {
        'movel': att_movel,
        'fixa': att_fixa,
        'smartphones': att_smart,
        'eletronicos': att_eletro,
        'essenciais': att_ess,
        'seguros': att_seg,
        'sva': att_sva,
    }

    # PDV: meta sempre = soma das metas dos consultores do PDV
    pdv_meta = {
        'movel': sumifs(realized, 'META_MOVEL', 'PDV', pdv),
        'fixa': sumifs(realized, 'META_FIXA', 'PDV', pdv),
        'smartphones': sumifs(realized, 'META_SMARTPHONE', 'PDV', pdv),
        'eletronicos': sumifs(realized, 'META_ACESSORIO', 'PDV', pdv),
        'essenciais': sumifs(realized, 'META_ESSENCIAIS', 'PDV', pdv),
        'seguros': sumifs(realized, 'META_SEGUROS', 'PDV', pdv),
        'sva': sumifs(realized, 'META_SVA', 'PDV', pdv),
    }

    # Sobrescreve metas do PDV com valores do Power BI (META_PDV_REAL)
    store_for_pdv = get_store_name_from_user(user) or pdv
    if store_for_pdv:
        pdv_pb = get_metas_from_power_bi(store_name=store_for_pdv)
        for k, v in (pdv_pb or {}).items():
            if v:
                pdv_meta[k] = v

    # PDV: realizado conforme o modo
    if view_mode == VIEW_REALIZADO:
        # Realizado do PDV vem do MySQL (loja do consultor).
        pdv_lookup = get_store_name_from_user(user) or pdv
        mysql_pdv = get_realized_sales_from_mysql(pdv=pdv_lookup)
        pdv_proj = {
            'movel': mysql_pdv.get('movel', 0.0),
            'fixa': mysql_pdv.get('fixa', 0.0),
            'smartphones': mysql_pdv.get('smartphones', 0.0),
            'eletronicos': mysql_pdv.get('eletronicos', 0.0),
            'essenciais': mysql_pdv.get('essenciais', 0.0),
            'seguros': mysql_pdv.get('seguros', 0.0),
            'sva': mysql_pdv.get('sva', 0.0),
        }
    elif view_mode == VIEW_SIMULADOR:
        pdv_proj = {
            'movel': _get_sim_input(simulator_inputs, 'movel', 'realpdv'),
            # Fixa: PDV é em QUANTIDADE de vendas (campo 'qtypdv'), não receita.
            'fixa': _get_sim_input(simulator_inputs, 'fixa', 'qtypdv'),
            'smartphones': _get_sim_input(simulator_inputs, 'smartphones', 'realpdv'),
            'eletronicos': _get_sim_input(simulator_inputs, 'eletronicos_a', 'realpdv')
                + _get_sim_input(simulator_inputs, 'eletronicos_b', 'realpdv'),
            'essenciais': _get_sim_input(simulator_inputs, 'essenciais_a', 'realpdv')
                + _get_sim_input(simulator_inputs, 'essenciais_b', 'realpdv'),
            'seguros': _get_sim_input(simulator_inputs, 'seguros', 'realpdv'),
            'sva': _get_sim_input(simulator_inputs, 'sva', 'realpdv'),
        }
    else:
        # VIEW_PROJECAO: projeção dinâmica do PDV via MySQL D-1 (ontem) projetado por DU.
        pdv_lookup = get_store_name_from_user(user) or pdv
        mysql_pdv = get_realized_sales_from_mysql(pdv=pdv_lookup)
        du_passed, du_total = get_business_days_info()
        pdv_proj = {
            'movel': project_from_realized(mysql_pdv.get('movel', 0.0), du_passed, du_total),
            'fixa': project_from_realized(mysql_pdv.get('fixa', 0.0), du_passed, du_total),
            'smartphones': project_from_realized(mysql_pdv.get('smartphones', 0.0), du_passed, du_total),
            'eletronicos': project_from_realized(mysql_pdv.get('eletronicos', 0.0), du_passed, du_total),
            'essenciais': project_from_realized(mysql_pdv.get('essenciais', 0.0), du_passed, du_total),
            'seguros': project_from_realized(mysql_pdv.get('seguros', 0.0), du_passed, du_total),
            'sva': project_from_realized(mysql_pdv.get('sva', 0.0), du_passed, du_total),
        }
    pdv_att = {
        key: (pdv_proj[key] / pdv_meta[key] if pdv_meta[key] else 0.0)
        for key in pdv_meta
    }

    # COORDENAÇÃO: meta = soma das metas dos consultores da coordenação.
    coord_meta = {
        'movel': sumifs(realized, 'META_MOVEL', 'COORDENAÇÃO', coord_name),
        'fixa': sumifs(realized, 'META_FIXA', 'COORDENAÇÃO', coord_name),
        'smartphones': sumifs(realized, 'META_SMARTPHONE', 'COORDENAÇÃO', coord_name),
        'eletronicos': sumifs(realized, 'META_ACESSORIO', 'COORDENAÇÃO', coord_name),
        'essenciais': sumifs(realized, 'META_ESSENCIAIS', 'COORDENAÇÃO', coord_name),
        'seguros': sumifs(realized, 'META_SEGUROS', 'COORDENAÇÃO', coord_name),
        'sva': sumifs(realized, 'META_SVA', 'COORDENAÇÃO', coord_name),
    }
    if view_mode == VIEW_REALIZADO:
        coord_pdvs = get_pdvs_of_coord(realized, coord_name) or get_pdvs_of_coord(projection, coord_name)
        mysql_coord = get_realized_sales_from_mysql(pdvs=coord_pdvs) if coord_pdvs else get_realized_sales_from_mysql(coord_name=coord_name)
        coord_proj = {k: mysql_coord.get(k, 0.0) for k in ['movel','fixa','smartphones','eletronicos','essenciais','seguros','sva']}
    else:
        # VIEW_PROJECAO / VIEW_SIMULADOR: projeção dinâmica da Coordenação via MySQL D-1 projetado por DU.
        coord_pdvs = get_pdvs_of_coord(realized, coord_name) or get_pdvs_of_coord(projection, coord_name)
        mysql_coord = get_realized_sales_from_mysql(pdvs=coord_pdvs) if coord_pdvs else get_realized_sales_from_mysql(coord_name=coord_name)
        du_passed_c, du_total_c = get_business_days_info()
        coord_proj = {
            'movel': project_from_realized(mysql_coord.get('movel', 0.0), du_passed_c, du_total_c),
            'fixa': project_from_realized(mysql_coord.get('fixa', 0.0), du_passed_c, du_total_c),
            'smartphones': project_from_realized(mysql_coord.get('smartphones', 0.0), du_passed_c, du_total_c),
            'eletronicos': project_from_realized(mysql_coord.get('eletronicos', 0.0), du_passed_c, du_total_c),
            'essenciais': project_from_realized(mysql_coord.get('essenciais', 0.0), du_passed_c, du_total_c),
            'seguros': project_from_realized(mysql_coord.get('seguros', 0.0), du_passed_c, du_total_c),
            'sva': project_from_realized(mysql_coord.get('sva', 0.0), du_passed_c, du_total_c),
        }
    coord_att = {
        key: (coord_proj[key] / coord_meta[key] if coord_meta[key] else 0.0)
        for key in coord_meta
    }

    meta_config = factor_data.get('meta', {})
    ranges = factor_data.get('ranges', {})
    hunter_levels = hunter_levels or {}

    all_ok = all_pillars_ok(att_map, coord_name, 0.695)

    rows = []
    total_p = 0.0
    total_h2 = 0.0
    total_h3 = 0.0

    for key in PILLAR_ORDER:
        config = PILLAR_CONFIG_CONSULTOR.get(key)
        if not config:
            continue

        label = PILLAR_LABELS.get(key, key)
        rate_col = config.get('rate_col', 3)
        pdv_rate_col = config.get('pdv_rate_col', rate_col)
        commission_key = config.get('commission_key')
        premium_key = config.get('premium_key')
        pdv_premium_key = config.get('pdv_premium_key')

        if key == 'movel':
            meta = meta_movel
            proj = proj_movel
            att = att_movel
            pdv_att_ref = pdv_att['movel']
        elif key == 'fixa':
            meta = meta_fixa
            proj = proj_fixa
            att = att_fixa
            pdv_att_ref = pdv_att['smartphones']
        elif key == 'smartphones':
            meta = meta_smart
            proj = proj_smart
            att = att_smart
            pdv_att_ref = pdv_att['smartphones']
        elif key == 'eletronicos_a':
            meta = meta_eletro
            proj = proj_eletro_a
            att = att_eletro
            pdv_att_ref = pdv_att['eletronicos']
        elif key == 'eletronicos_b':
            meta = None
            proj = proj_eletro_b
            att = att_eletro
            pdv_att_ref = pdv_att['eletronicos']
        elif key == 'essenciais_a':
            meta = meta_ess
            proj = proj_ess_a
            att = att_ess
            pdv_att_ref = pdv_att['essenciais']
        elif key == 'essenciais_b':
            meta = None
            proj = proj_ess_b
            att = att_ess
            pdv_att_ref = pdv_att['essenciais']
        elif key == 'seguros':
            meta = meta_seg
            proj = proj_seg
            att = att_seg
            pdv_att_ref = pdv_att['seguros']
        else:
            meta = meta_sva
            proj = proj_sva
            att = att_sva
            pdv_att_ref = pdv_att['sva']

        commission_rate = 0.0
        if commission_key == 'sva_commission':
            commission_range = ranges.get(commission_key, [[0.0]])
            commission_rate = to_float(commission_range[0][0]) if att >= 0 else 0.0
        else:
            commission_rate = vlookup(att, ranges.get(commission_key, []), rate_col)

        premium_rate = 0.0
        if premium_key:
            premium_rate = vlookup(att, ranges.get(premium_key, []), rate_col)

        commission_value = proj * commission_rate
        premium_value = proj * premium_rate if all_ok else 0.0
        total_individual = commission_value + premium_value

        pdv_group_key = 'sva' if key == 'sva' else key.replace('_a', '').replace('_b', '')
        pdv_meta_value = pdv_meta.get(pdv_group_key, 0.0)
        pdv_proj_value = pdv_proj.get(pdv_group_key, 0.0)
        pdv_att_value = pdv_proj_value / pdv_meta_value if pdv_meta_value else 0.0
        coord_meta_value = coord_meta.get(pdv_group_key, 0.0)
        coord_proj_value = coord_proj.get(pdv_group_key, 0.0)
        coord_att_value = coord_proj_value / coord_meta_value if coord_meta_value else 0.0
        if key.endswith('_b'):
            pdv_meta_value = None
            pdv_proj_value = None
            pdv_att_value = None
            coord_meta_value = None
            coord_proj_value = None
            coord_att_value = None
        pdv_threshold, pdv_rate = pdv_threshold_rate(ranges.get(pdv_premium_key, []), pdv_rate_col)
        pdv_rate_value = pdv_rate if pdv_att_ref >= pdv_threshold else 0.0
        pdv_premium_value = proj * pdv_rate_value if (att >= 1.0 and all_ok) else 0.0

        total_with_pdv = total_individual + pdv_premium_value

        hunter_key = key.replace('_a', '').replace('_b', '')
        hunter_level = hunter_levels.get(hunter_key, 0)
        hunter2_rate = meta_config.get('hunter2_rate', 0.0)
        hunter3_rate = meta_config.get('hunter3_rate', 0.0)
        hunter2_value = total_with_pdv * hunter2_rate if hunter_level == 2 else 0.0
        hunter3_value = total_with_pdv * hunter3_rate if hunter_level == 3 else 0.0

        total_p += total_with_pdv
        total_h2 += hunter2_value
        total_h3 += hunter3_value

        rows.append({
            'key': key,
            'label': label,
            'meta': meta,
            'proj': proj,
            'attainment': att,
            'commission_rate': commission_rate,
            'premium_rate': premium_rate,
            'commission_value': commission_value,
            'premium_value': premium_value,
            'total_individual': total_individual,
            'pdv_meta': pdv_meta_value,
            'pdv_proj': pdv_proj_value,
            'pdv_attainment': pdv_att_value,
            'pdv_premium_rate': pdv_rate_value,
            'pdv_premium_value': pdv_premium_value,
            'total_with_pdv': total_with_pdv,
            'coord_meta': coord_meta_value,
            'coord_proj': coord_proj_value,
            'coord_attainment': coord_att_value,
            'hunter2_value': hunter2_value,
            'hunter3_value': hunter3_value,
            'quantity': fixa_quantity if key == 'fixa' else None,
            'revenue': fixa_revenue if key == 'fixa' else None,
        })

    bonus_rate = meta_config.get('bonus_6_7_rate', 0.0)
    bonus_value = (total_p + total_h2 + total_h3) * bonus_rate if bonus_6_7_ok(att_map, coord_name) else 0.0

    rows = _merge_grouped_rows(rows)
    return {
        'user_name': user_name,
        'first_name': user.first_name or user_name.split(' ')[0],
        'pdv': pdv,
        'coordinator': coord_name,
        'view_mode': view_mode,
        'rows': rows,
        'totals': {
            'total_with_pdv': total_p,
            'hunter2': total_h2,
            'hunter3': total_h3,
            'bonus_6_7': bonus_value,
            'ganho_total': total_p + total_h2 + total_h3 + bonus_value,
        },
    }


def compute_gerente_simulation(
    user: User,
    factor_data: Dict[str, Any],
    hunter_levels: Optional[Dict[str, int]] = None,
    view_mode: str = VIEW_PROJECAO,
    simulator_inputs: Optional[Dict[str, Any]] = None,
) -> Dict[str, Any]:
    realized = load_dataframe(ROLE_GERENTE, 'REALIZADO')
    projection = load_dataframe(ROLE_GERENTE, 'PROJEÇÃO')

    # Resolve PDV usando o nome de loja normalizado (ex.: "Loja Anchieta" → "ANCHIETA")
    # para que bata com o valor das planilhas.
    store_name = get_store_name_from_user(user)
    pdv = store_name
    if not pdv and getattr(user, 'sector', None):
        pdv = (user.sector.name or '').strip()
    # Sem PDV vinculado: o gerente ainda pode usar o Simulador (valores
    # informados manualmente). As metas/projeções da planilha ficarão zeradas.

    coord_name = ''
    if pdv:
        target_pdv = normalize_text(pdv)
        for _, row in realized.iterrows():
            if normalize_text(row.get('PDV', '')) == target_pdv:
                coord_name = str(row.get('COORDENAÇÃO') or '').strip()
                if coord_name:
                    break
        # Fallback: tenta na planilha de PROJEÇÃO se não achou em REALIZADO.
        if not coord_name:
            for _, row in projection.iterrows():
                if normalize_text(row.get('PDV', '')) == target_pdv:
                    coord_name = str(row.get('COORDENAÇÃO') or '').strip()
                    if coord_name:
                        break

    meta_map = {
        'movel': sumifs(realized, 'META_MOVEL', 'PDV', pdv),
        'fixa': sumifs(realized, 'META_FIXA', 'PDV', pdv),
        'smartphones': sumifs(realized, 'META_SMARTPHONE', 'PDV', pdv),
        'eletronicos': sumifs(realized, 'META_ACESSORIO', 'PDV', pdv),
        'essenciais': sumifs(realized, 'META_ESSENCIAIS', 'PDV', pdv),
        'seguros': sumifs(realized, 'META_SEGUROS', 'PDV', pdv),
        'sva': sumifs(realized, 'META_SVA', 'PDV', pdv),
    }
    # Sobrescreve com metas oficiais (META_PDV_REAL)
    store_for_pdv = store_name or pdv
    if store_for_pdv:
        pdv_pb = get_metas_from_power_bi(store_name=store_for_pdv)
        for k, v in (pdv_pb or {}).items():
            if v:
                meta_map[k] = v

    if view_mode == VIEW_REALIZADO:
        # Gerente: realizado da loja inteira via MySQL.
        mysql_pdv = get_realized_sales_from_mysql(pdv=pdv)
        proj_map = {
            'movel': mysql_pdv.get('movel', 0.0),
            'fixa': mysql_pdv.get('fixa', 0.0),
            'smartphones': mysql_pdv.get('smartphones', 0.0),
            'eletronicos': mysql_pdv.get('eletronicos', 0.0),
            'essenciais': mysql_pdv.get('essenciais', 0.0),
            'seguros': mysql_pdv.get('seguros', 0.0),
            'sva': mysql_pdv.get('sva', 0.0),
        }
        eletro_a_pdv = mysql_pdv.get('eletronicos', 0.0)
        eletro_b_pdv = 0.0
        ess_a_pdv = mysql_pdv.get('essenciais', 0.0)
        ess_b_pdv = 0.0
        fixa_quantity = mysql_pdv.get('fixa_qty', 0.0)
        fixa_revenue = proj_map['fixa']
    elif view_mode == VIEW_SIMULADOR:
        # Para Gerente em modo Simulador: usuário define real (PDV) por pilar.
        # Aceita tanto a chave unificada ('eletronicos'/'essenciais') quanto _a/_b.
        eletro_a_in = _get_sim_input(simulator_inputs, 'eletronicos_a', 'real')
        eletro_b_in = _get_sim_input(simulator_inputs, 'eletronicos_b', 'real')
        ess_a_in = _get_sim_input(simulator_inputs, 'essenciais_a', 'real')
        ess_b_in = _get_sim_input(simulator_inputs, 'essenciais_b', 'real')
        proj_map = {
            'movel': _get_sim_input(simulator_inputs, 'movel', 'real'),
            # Fixa: input do gerente para PDV/coordenação é em QUANTIDADE ('qty').
            'fixa': _get_sim_input(simulator_inputs, 'fixa', 'qty'),
            'smartphones': _get_sim_input(simulator_inputs, 'smartphones', 'real'),
            'eletronicos': eletro_a_in + eletro_b_in,
            'essenciais': ess_a_in + ess_b_in,
            'seguros': _get_sim_input(simulator_inputs, 'seguros', 'real'),
            'sva': _get_sim_input(simulator_inputs, 'sva', 'real'),
        }
        eletro_a_pdv = eletro_a_in
        eletro_b_pdv = eletro_b_in
        ess_a_pdv = ess_a_in
        ess_b_pdv = ess_b_in
        fixa_quantity = _get_sim_input(simulator_inputs, 'fixa', 'qty')
        # Receita estimada de Fixa = qtd simulada × ticket médio do realizado da loja.
        _real_pdv = get_realized_sales_from_mysql(pdv=pdv)
        _real_fixa_qty = _real_pdv.get('fixa_qty', 0.0) or 0.0
        _real_fixa_rev = _real_pdv.get('fixa', 0.0) or 0.0
        _ticket_medio_fixa = (_real_fixa_rev / _real_fixa_qty) if _real_fixa_qty > 0 else 0.0
        fixa_revenue = fixa_quantity * _ticket_medio_fixa
        proj_map['fixa'] = fixa_revenue
        # Permite sobrescrever metas
        for key in list(meta_map.keys()):
            override = _get_sim_input_optional(simulator_inputs, key, 'meta')
            if override is not None:
                meta_map[key] = override
    else:
        # Projeção dinâmica do PDV: realizado MySQL × DU.
        mysql_pdv = get_realized_sales_from_mysql(pdv=pdv)
        du_passed, du_total = get_business_days_info()
        proj_map = {
            'movel': project_from_realized(mysql_pdv.get('movel', 0.0), du_passed, du_total),
            'fixa': project_from_realized(mysql_pdv.get('fixa', 0.0), du_passed, du_total),
            'smartphones': project_from_realized(mysql_pdv.get('smartphones', 0.0), du_passed, du_total),
            'eletronicos': project_from_realized(mysql_pdv.get('eletronicos', 0.0), du_passed, du_total),
            'essenciais': project_from_realized(mysql_pdv.get('essenciais', 0.0), du_passed, du_total),
            'seguros': project_from_realized(mysql_pdv.get('seguros', 0.0), du_passed, du_total),
            'sva': project_from_realized(mysql_pdv.get('sva', 0.0), du_passed, du_total),
        }
        eletro_a_pdv = proj_map['eletronicos']
        eletro_b_pdv = 0.0
        ess_a_pdv = proj_map['essenciais']
        ess_b_pdv = 0.0
        fixa_quantity = project_from_realized(mysql_pdv.get('fixa_qty', 0.0), du_passed, du_total)
        fixa_revenue = proj_map['fixa']

    coord_meta = {
        'movel': sumifs(realized, 'META_MOVEL', 'COORDENAÇÃO', coord_name),
        'fixa': sumifs(realized, 'META_FIXA', 'COORDENAÇÃO', coord_name),
        'smartphones': sumifs(realized, 'META_SMARTPHONE', 'COORDENAÇÃO', coord_name),
        'eletronicos': sumifs(realized, 'META_ACESSORIO', 'COORDENAÇÃO', coord_name),
        'essenciais': sumifs(realized, 'META_ESSENCIAIS', 'COORDENAÇÃO', coord_name),
        'seguros': sumifs(realized, 'META_SEGUROS', 'COORDENAÇÃO', coord_name),
        'sva': sumifs(realized, 'META_SVA', 'COORDENAÇÃO', coord_name),
    }
    if view_mode == VIEW_REALIZADO:
        # Realizado do coordenador via MySQL (soma de todas as lojas que ele coordena).
        coord_pdvs = get_pdvs_of_coord(realized, coord_name) or get_pdvs_of_coord(projection, coord_name)
        mysql_coord = get_realized_sales_from_mysql(pdvs=coord_pdvs) if coord_pdvs else get_realized_sales_from_mysql(coord_name=coord_name)
        coord_proj = {
            'movel': mysql_coord.get('movel', 0.0),
            'fixa': mysql_coord.get('fixa', 0.0),
            'smartphones': mysql_coord.get('smartphones', 0.0),
            'eletronicos': mysql_coord.get('eletronicos', 0.0),
            'essenciais': mysql_coord.get('essenciais', 0.0),
            'seguros': mysql_coord.get('seguros', 0.0),
            'sva': mysql_coord.get('sva', 0.0),
        }
    else:
        # VIEW_PROJECAO / VIEW_SIMULADOR: coord_proj via MySQL D-1 projetado por DU.
        coord_pdvs = get_pdvs_of_coord(realized, coord_name) or get_pdvs_of_coord(projection, coord_name)
        mysql_coord = get_realized_sales_from_mysql(pdvs=coord_pdvs) if coord_pdvs else get_realized_sales_from_mysql(coord_name=coord_name)
        du_passed_c, du_total_c = get_business_days_info()
        coord_proj = {
            'movel': project_from_realized(mysql_coord.get('movel', 0.0), du_passed_c, du_total_c),
            'fixa': project_from_realized(mysql_coord.get('fixa', 0.0), du_passed_c, du_total_c),
            'smartphones': project_from_realized(mysql_coord.get('smartphones', 0.0), du_passed_c, du_total_c),
            'eletronicos': project_from_realized(mysql_coord.get('eletronicos', 0.0), du_passed_c, du_total_c),
            'essenciais': project_from_realized(mysql_coord.get('essenciais', 0.0), du_passed_c, du_total_c),
            'seguros': project_from_realized(mysql_coord.get('seguros', 0.0), du_passed_c, du_total_c),
            'sva': project_from_realized(mysql_coord.get('sva', 0.0), du_passed_c, du_total_c),
        }

    att_map = {
        key: (proj_map[key] / meta_map[key] if meta_map[key] else 0.0)
        for key in meta_map
    }
    # Fixa: atingimento por quantidade (qty_vendida ÷ meta_qty).
    att_map['fixa'] = (fixa_quantity / meta_map['fixa']) if meta_map['fixa'] else 0.0

    meta_config = factor_data.get('meta', {})
    ranges = factor_data.get('ranges', {})
    hunter_levels = hunter_levels or {}

    all_ok = all_pillars_ok(att_map, coord_name, 0.695)

    rows = []
    total_p = 0.0
    total_h2 = 0.0
    total_h3 = 0.0

    for key in PILLAR_ORDER:
        config = PILLAR_CONFIG_GERENTE.get(key)
        if not config:
            continue

        label = PILLAR_LABELS.get(key, key)
        rate_col = config.get('rate_col', 3)
        pdv_rate_col = config.get('pdv_rate_col', rate_col)
        commission_key = config.get('commission_key')
        premium_key = config.get('premium_key')
        pdv_premium_key = config.get('pdv_premium_key')

        if key == 'movel':
            meta = meta_map['movel']
            proj = proj_map['movel']
            att = att_map['movel']
            pdv_att_ref = coord_proj['movel'] / coord_meta['movel'] if coord_meta['movel'] else 0.0
        elif key == 'fixa':
            meta = meta_map['fixa']
            proj = proj_map['fixa']
            att = att_map['fixa']
            pdv_att_ref = coord_proj['fixa'] / coord_meta['fixa'] if coord_meta['fixa'] else 0.0
        elif key == 'smartphones':
            meta = meta_map['smartphones']
            proj = proj_map['smartphones']
            att = att_map['smartphones']
            pdv_att_ref = coord_proj['smartphones'] / coord_meta['smartphones'] if coord_meta['smartphones'] else 0.0
        elif key == 'eletronicos_a':
            meta = meta_map['eletronicos']
            proj = eletro_a_pdv
            att = att_map['eletronicos']
            pdv_att_ref = coord_proj['eletronicos'] / coord_meta['eletronicos'] if coord_meta['eletronicos'] else 0.0
        elif key == 'eletronicos_b':
            meta = None
            proj = eletro_b_pdv
            att = att_map['eletronicos']
            pdv_att_ref = coord_proj['eletronicos'] / coord_meta['eletronicos'] if coord_meta['eletronicos'] else 0.0
        elif key == 'essenciais_a':
            meta = meta_map['essenciais']
            proj = ess_a_pdv
            att = att_map['essenciais']
            pdv_att_ref = coord_proj['essenciais'] / coord_meta['essenciais'] if coord_meta['essenciais'] else 0.0
        elif key == 'essenciais_b':
            meta = None
            proj = ess_b_pdv
            att = att_map['essenciais']
            pdv_att_ref = coord_proj['essenciais'] / coord_meta['essenciais'] if coord_meta['essenciais'] else 0.0
        elif key == 'seguros':
            meta = meta_map['seguros']
            proj = proj_map['seguros']
            att = att_map['seguros']
            pdv_att_ref = coord_proj['seguros'] / coord_meta['seguros'] if coord_meta['seguros'] else 0.0
        else:
            meta = meta_map['sva']
            proj = proj_map['sva']
            att = att_map['sva']
            pdv_att_ref = coord_proj['sva'] / coord_meta['sva'] if coord_meta['sva'] else 0.0

        commission_rate = 0.0
        if commission_key == 'sva_commission':
            commission_range = ranges.get(commission_key, [[0.0]])
            commission_rate = to_float(commission_range[0][0]) if att >= 0 else 0.0
        else:
            commission_rate = vlookup(att, ranges.get(commission_key, []), rate_col)

        premium_rate = 0.0
        if premium_key:
            premium_rate = vlookup(att, ranges.get(premium_key, []), rate_col)

        commission_value = proj * commission_rate
        premium_value = proj * premium_rate if all_ok else 0.0
        total_individual = commission_value + premium_value

        pdv_group_key = 'sva' if key == 'sva' else key.replace('_a', '').replace('_b', '')
        pdv_meta_value = coord_meta.get(pdv_group_key, 0.0)
        pdv_proj_value = coord_proj.get(pdv_group_key, 0.0)
        pdv_att_value = pdv_proj_value / pdv_meta_value if pdv_meta_value else 0.0
        if key.endswith('_b'):
            pdv_meta_value = None
            pdv_proj_value = None
            pdv_att_value = None
        pdv_threshold, pdv_rate = pdv_threshold_rate(ranges.get(pdv_premium_key, []), pdv_rate_col)
        pdv_rate_value = pdv_rate if pdv_att_ref >= pdv_threshold else 0.0
        pdv_premium_value = proj * pdv_rate_value if (att >= 1.0 and all_ok) else 0.0

        total_with_pdv = total_individual + pdv_premium_value

        hunter_key = key.replace('_a', '').replace('_b', '')
        hunter_level = hunter_levels.get(hunter_key, 0)
        hunter2_rate = meta_config.get('hunter2_rate', 0.0)
        hunter3_rate = meta_config.get('hunter3_rate', 0.0)
        hunter2_value = total_with_pdv * hunter2_rate if hunter_level == 2 else 0.0
        hunter3_value = total_with_pdv * hunter3_rate if hunter_level == 3 else 0.0

        total_p += total_with_pdv
        total_h2 += hunter2_value
        total_h3 += hunter3_value

        rows.append({
            'key': key,
            'label': label,
            'meta': meta,
            'proj': proj,
            'attainment': att,
            'commission_rate': commission_rate,
            'premium_rate': premium_rate,
            'commission_value': commission_value,
            'premium_value': premium_value,
            'total_individual': total_individual,
            'pdv_meta': pdv_meta_value,
            'pdv_proj': pdv_proj_value,
            'pdv_attainment': pdv_att_value,
            'pdv_premium_rate': pdv_rate_value,
            'pdv_premium_value': pdv_premium_value,
            'total_with_pdv': total_with_pdv,
            'coord_meta': pdv_meta_value,
            'coord_proj': pdv_proj_value,
            'coord_attainment': pdv_att_value,
            'hunter2_value': hunter2_value,
            'hunter3_value': hunter3_value,
            'quantity': fixa_quantity if key == 'fixa' else None,
            'revenue': fixa_revenue if key == 'fixa' else None,
        })

    bonus_rate = meta_config.get('bonus_6_7_rate', 0.0)
    bonus_value = (total_p + total_h2 + total_h3) * bonus_rate if bonus_6_7_ok(att_map, coord_name) else 0.0

    rows = _merge_grouped_rows(rows)
    return {
        'user_name': user.get_full_name() or user.first_name or user.email,
        'first_name': user.first_name or (user.get_full_name() or user.email).split(' ')[0],
        'pdv': pdv,
        'coordinator': coord_name,
        'view_mode': view_mode,
        'rows': rows,
        'totals': {
            'total_with_pdv': total_p,
            'hunter2': total_h2,
            'hunter3': total_h3,
            'bonus_6_7': bonus_value,
            'ganho_total': total_p + total_h2 + total_h3 + bonus_value,
        },
    }


def compute_coordenador_simulation(
    user: User,
    factor_data: Dict[str, Any],
    hunter_levels: Optional[Dict[str, int]] = None,
    view_mode: str = VIEW_PROJECAO,
    simulator_inputs: Optional[Dict[str, Any]] = None,
) -> Dict[str, Any]:
    realized = load_dataframe(ROLE_COORDENADOR, 'REALIZADO')
    projection = load_dataframe(ROLE_COORDENADOR, 'PROJEÇÃO')

    coord_name = user.first_name or user.get_full_name() or user.email

    meta_map = {
        'movel': sumifs(realized, 'META_MOVEL', 'COORDENAÇÃO', coord_name),
        'fixa': sumifs(realized, 'META_FIXA', 'COORDENAÇÃO', coord_name),
        'smartphones': sumifs(realized, 'META_SMARTPHONE', 'COORDENAÇÃO', coord_name),
        'eletronicos': sumifs(realized, 'META_ACESSORIO', 'COORDENAÇÃO', coord_name),
        'essenciais': sumifs(realized, 'META_ESSENCIAIS', 'COORDENAÇÃO', coord_name),
        'seguros': sumifs(realized, 'META_SEGUROS', 'COORDENAÇÃO', coord_name),
        'sva': sumifs(realized, 'META_SVA', 'COORDENAÇÃO', coord_name),
    }
    # Sobrescreve com soma das metas de PDV cadastradas em /power-bi/metas/
    coord_pb = get_pdv_metas_for_coordinator(coord_name)
    for k, v in (coord_pb or {}).items():
        if v:
            meta_map[k] = v
    proj_map = {
        'movel': sumifs(projection, 'PROJ_MOVEL', 'COORDENAÇÃO', coord_name),
        'fixa': sumifs(projection, 'PROJ_FIXA', 'COORDENAÇÃO', coord_name),
        'smartphones': sumifs(projection, 'PROJ_APARELHO', 'COORDENAÇÃO', coord_name),
        'eletronicos': sumifs(projection, 'PROJ_ELETRO_A', 'COORDENAÇÃO', coord_name) + sumifs(projection, 'PROJ_ELETRO_B', 'COORDENAÇÃO', coord_name),
        'essenciais': sumifs(projection, 'PROJ_ESSEN_A', 'COORDENAÇÃO', coord_name) + sumifs(projection, 'PROJ_ESSEN_B', 'COORDENAÇÃO', coord_name),
        'seguros': sumifs(projection, 'PROJ_SEGURO', 'COORDENAÇÃO', coord_name),
        'sva': sumifs(projection, 'PROJ_SVA', 'COORDENAÇÃO', coord_name),
    }
    eletro_a = sumifs(projection, 'PROJ_ELETRO_A', 'COORDENAÇÃO', coord_name)
    eletro_b = sumifs(projection, 'PROJ_ELETRO_B', 'COORDENAÇÃO', coord_name)
    ess_a = sumifs(projection, 'PROJ_ESSEN_A', 'COORDENAÇÃO', coord_name)
    ess_b = sumifs(projection, 'PROJ_ESSEN_B', 'COORDENAÇÃO', coord_name)
    fixa_quantity = sumifs(projection, 'PROJ_BL', 'COORDENAÇÃO', coord_name) if 'PROJ_BL' in projection.columns else 0.0
    fixa_revenue = proj_map['fixa']

    if view_mode == VIEW_REALIZADO:
        # Coordenador: realizado de todas as lojas coordenadas via MySQL.
        coord_pdvs = get_pdvs_of_coord(realized, coord_name) or get_pdvs_of_coord(projection, coord_name)
        mysql_coord = get_realized_sales_from_mysql(pdvs=coord_pdvs) if coord_pdvs else get_realized_sales_from_mysql(coord_name=coord_name)
        proj_map = {
            'movel': mysql_coord.get('movel', 0.0),
            'fixa': mysql_coord.get('fixa', 0.0),
            'smartphones': mysql_coord.get('smartphones', 0.0),
            'eletronicos': mysql_coord.get('eletronicos', 0.0),
            'essenciais': mysql_coord.get('essenciais', 0.0),
            'seguros': mysql_coord.get('seguros', 0.0),
            'sva': mysql_coord.get('sva', 0.0),
        }
        eletro_a = mysql_coord.get('eletronicos', 0.0)
        eletro_b = 0.0
        ess_a = mysql_coord.get('essenciais', 0.0)
        ess_b = 0.0
        fixa_quantity = mysql_coord.get('fixa_qty', 0.0)
        fixa_revenue = proj_map['fixa']
    elif view_mode == VIEW_SIMULADOR:
        eletro_a_in = _get_sim_input(simulator_inputs, 'eletronicos_a', 'real')
        eletro_b_in = _get_sim_input(simulator_inputs, 'eletronicos_b', 'real')
        ess_a_in = _get_sim_input(simulator_inputs, 'essenciais_a', 'real')
        ess_b_in = _get_sim_input(simulator_inputs, 'essenciais_b', 'real')
        proj_map = {
            'movel': _get_sim_input(simulator_inputs, 'movel', 'real'),
            # Fixa: coordenador informa QUANTIDADE de vendas ('qty'), não receita.
            'fixa': _get_sim_input(simulator_inputs, 'fixa', 'qty'),
            'smartphones': _get_sim_input(simulator_inputs, 'smartphones', 'real'),
            'eletronicos': eletro_a_in + eletro_b_in,
            'essenciais': ess_a_in + ess_b_in,
            'seguros': _get_sim_input(simulator_inputs, 'seguros', 'real'),
            'sva': _get_sim_input(simulator_inputs, 'sva', 'real'),
        }
        eletro_a = eletro_a_in
        eletro_b = eletro_b_in
        ess_a = ess_a_in
        ess_b = ess_b_in
        fixa_quantity = _get_sim_input(simulator_inputs, 'fixa', 'qty')
        # Receita estimada de Fixa = qtd simulada × ticket médio do realizado da coordenação.
        coord_pdvs = get_pdvs_of_coord(realized, coord_name) or get_pdvs_of_coord(projection, coord_name)
        _real_coord = get_realized_sales_from_mysql(pdvs=coord_pdvs) if coord_pdvs else get_realized_sales_from_mysql(coord_name=coord_name)
        _real_fixa_qty = _real_coord.get('fixa_qty', 0.0) or 0.0
        _real_fixa_rev = _real_coord.get('fixa', 0.0) or 0.0
        _ticket_medio_fixa = (_real_fixa_rev / _real_fixa_qty) if _real_fixa_qty > 0 else 0.0
        fixa_revenue = fixa_quantity * _ticket_medio_fixa
        proj_map['fixa'] = fixa_revenue
        for key in list(meta_map.keys()):
            override = _get_sim_input_optional(simulator_inputs, key, 'meta')
            if override is not None:
                meta_map[key] = override
    else:
        # VIEW_PROJECAO: projeção dinâmica = realizado MySQL / DU_passados * DU_totais.
        coord_pdvs = get_pdvs_of_coord(realized, coord_name) or get_pdvs_of_coord(projection, coord_name)
        mysql_coord = get_realized_sales_from_mysql(pdvs=coord_pdvs) if coord_pdvs else get_realized_sales_from_mysql(coord_name=coord_name)
        du_passed, du_total = get_business_days_info()
        proj_map = {
            'movel': project_from_realized(mysql_coord.get('movel', 0.0), du_passed, du_total),
            'fixa': project_from_realized(mysql_coord.get('fixa', 0.0), du_passed, du_total),
            'smartphones': project_from_realized(mysql_coord.get('smartphones', 0.0), du_passed, du_total),
            'eletronicos': project_from_realized(mysql_coord.get('eletronicos', 0.0), du_passed, du_total),
            'essenciais': project_from_realized(mysql_coord.get('essenciais', 0.0), du_passed, du_total),
            'seguros': project_from_realized(mysql_coord.get('seguros', 0.0), du_passed, du_total),
            'sva': project_from_realized(mysql_coord.get('sva', 0.0), du_passed, du_total),
        }
        eletro_a = proj_map['eletronicos']
        eletro_b = 0.0
        ess_a = proj_map['essenciais']
        ess_b = 0.0
        fixa_quantity = project_from_realized(mysql_coord.get('fixa_qty', 0.0), du_passed, du_total)
        fixa_revenue = proj_map['fixa']

    att_map = {
        key: (proj_map[key] / meta_map[key] if meta_map[key] else 0.0)
        for key in meta_map
    }
    # Fixa: atingimento por quantidade (qty_vendida ÷ meta_qty).
    att_map['fixa'] = (fixa_quantity / meta_map['fixa']) if meta_map['fixa'] else 0.0

    meta_config = factor_data.get('meta', {})
    ranges = factor_data.get('ranges', {})
    hunter_levels = hunter_levels or {}

    rows = []
    total_commission = 0.0
    total_h2 = 0.0
    total_h3 = 0.0

    for key in PILLAR_ORDER:
        config = PILLAR_CONFIG_COORDENADOR.get(key)
        if not config:
            continue

        label = PILLAR_LABELS.get(key, key)
        rate_col = config.get('rate_col', 3)
        commission_key = config.get('commission_key')

        if key == 'movel':
            meta = meta_map['movel']
            proj = proj_map['movel']
            att = att_map['movel']
        elif key == 'fixa':
            meta = meta_map['fixa']
            proj = proj_map['fixa']
            att = att_map['fixa']
        elif key == 'smartphones':
            meta = meta_map['smartphones']
            proj = proj_map['smartphones']
            att = att_map['smartphones']
        elif key == 'eletronicos_a':
            meta = meta_map['eletronicos']
            proj = eletro_a
            att = att_map['eletronicos']
        elif key == 'eletronicos_b':
            meta = None
            proj = eletro_b
            att = att_map['eletronicos']
        elif key == 'essenciais_a':
            meta = meta_map['essenciais']
            proj = ess_a
            att = att_map['essenciais']
        elif key == 'essenciais_b':
            meta = None
            proj = ess_b
            att = att_map['essenciais']
        elif key == 'seguros':
            meta = meta_map['seguros']
            proj = proj_map['seguros']
            att = att_map['seguros']
        else:
            meta = meta_map['sva']
            proj = proj_map['sva']
            att = att_map['sva']

        if commission_key == 'sva_commission':
            commission_range = ranges.get(commission_key, [[0.0]])
            commission_rate = to_float(commission_range[0][0]) if att >= 0 else 0.0
        else:
            commission_rate = vlookup(att, ranges.get(commission_key, []), rate_col)

        commission_value = proj * commission_rate

        hunter_key = key.replace('_a', '').replace('_b', '')
        hunter_level = hunter_levels.get(hunter_key, 0)
        hunter2_rate = meta_config.get('hunter2_rate', 0.0)
        hunter3_rate = meta_config.get('hunter3_rate', 0.0)
        hunter2_value = commission_value * hunter2_rate if hunter_level == 2 else 0.0
        hunter3_value = commission_value * hunter3_rate if hunter_level == 3 else 0.0

        total_commission += commission_value
        total_h2 += hunter2_value
        total_h3 += hunter3_value

        rows.append({
            'key': key,
            'label': label,
            'meta': meta,
            'proj': proj,
            'attainment': att,
            'commission_rate': commission_rate,
            'commission_value': commission_value,
            'hunter2_value': hunter2_value,
            'hunter3_value': hunter3_value,
            'quantity': fixa_quantity if key == 'fixa' else None,
            'revenue': fixa_revenue if key == 'fixa' else None,
        })

    bonus_rate = meta_config.get('bonus_6_7_rate', 0.0)
    bonus_value = total_commission * bonus_rate if bonus_6_7_ok(att_map, coord_name) else 0.0
    total_coordinator = total_commission + total_h2 + total_h3 + bonus_value
    sniper_rate = meta_config.get('sniper_rate', 0.75)

    # Usuários que pertencem ao grupo SNIPER (id=22 em /users/manage/groups/)
    # recebem 75% da comissão de coordenador.
    is_sniper = False
    try:
        is_sniper = user.communication_groups.filter(id=SNIPER_GROUP_ID).exists()
    except Exception:
        is_sniper = False

    if is_sniper:
        for row in rows:
            row['commission_value'] = row.get('commission_value', 0.0) * sniper_rate
            row['hunter2_value'] = row.get('hunter2_value', 0.0) * sniper_rate
            row['hunter3_value'] = row.get('hunter3_value', 0.0) * sniper_rate
        total_commission *= sniper_rate
        total_h2 *= sniper_rate
        total_h3 *= sniper_rate
        bonus_value *= sniper_rate
        total_coordinator *= sniper_rate

    rows = _merge_grouped_rows(rows)
    return {
        'user_name': user.get_full_name() or user.first_name or user.email,
        'first_name': user.first_name or (user.get_full_name() or user.email).split(' ')[0],
        'coordinator': coord_name,
        'view_mode': view_mode,
        'rows': rows,
        'is_sniper': is_sniper,
        'totals': {
            'total_commission': total_commission,
            'hunter2': total_h2,
            'hunter3': total_h3,
            'bonus_6_7': bonus_value,
            'ganho_total': total_coordinator,
            'ganho_total_sniper': total_coordinator if is_sniper else total_coordinator * sniper_rate,
        },
    }


def update_factor_sets_from_post(post_data: Dict[str, Any], updated_by: User) -> None:
    """Atualiza os fatores a partir do formulário interativo (campos numéricos).

    Se o usuário marcar reset__<role>, o conjunto será recarregado a partir do
    arquivo XLSX correspondente.
    """
    with transaction.atomic():
        for role in [ROLE_CONSULTOR, ROLE_GERENTE, ROLE_COORDENADOR]:
            factor_set = get_factor_set(role, updated_by=updated_by)

            if post_data.get(f"reset__{role}"):
                fresh = load_default_factor_data(role)
                factor_set.data = fresh
                factor_set.updated_by = updated_by
                factor_set.save()
                continue

            data = factor_set.data or {}
            ranges = data.get('ranges', {})
            meta = data.get('meta', {})

            for spec in FACTOR_RANGE_SPECS[role]:
                key = spec.key
                rows = [list(r) for r in ranges.get(key, [])]
                for r_index, row in enumerate(rows):
                    for c_index, _ in enumerate(row):
                        field = f"range__{role}__{key}__{r_index}__{c_index}"
                        if field in post_data:
                            raw = post_data.get(field)
                            if raw == '' or raw is None:
                                rows[r_index][c_index] = None
                            else:
                                rows[r_index][c_index] = to_float(raw)
                ranges[key] = rows

            for meta_key in DEFAULT_META_BY_ROLE.get(role, {}):
                field = f"meta__{role}__{meta_key}"
                if field in post_data:
                    raw = post_data.get(field)
                    if raw == '' or raw is None:
                        continue
                    meta[meta_key] = to_float(raw)

            data['ranges'] = ranges
            data['meta'] = meta
            factor_set.data = data
            factor_set.updated_by = updated_by
            factor_set.save()
