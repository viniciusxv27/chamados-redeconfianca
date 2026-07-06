from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_POST
from django.core.paginator import Paginator
from django.db import transaction
from django.utils import timezone
from decimal import Decimal
from datetime import timedelta

from .models import Prize, PrizeCategory, Redemption, CSTransaction, PrizeDiscount
from users.models import User


@login_required
def marketplace_view(request):
    """Marketplace de prêmios atualizado"""
    categories = PrizeCategory.objects.filter(active=True)
    
    # Filtros
    category_filter = request.GET.get('category')
    search = request.GET.get('search', '')
    ordering = request.GET.get('ordering', 'relevance')
    
    prizes = Prize.objects.filter(is_active=True)
    
    if category_filter:
        prizes = prizes.filter(category_id=category_filter)
    
    if search:
        prizes = prizes.filter(name__icontains=search)
    
    # Ordenação
    ordering_map = {
        'value_asc': ('value_cs', 'name'),
        'value_desc': ('-value_cs', 'name'),
        'recent': ('-created_at',),
        'relevance': ('-has_discount', '-priority', 'name'),
    }
    order_fields = ordering_map.get(ordering, ordering_map['relevance'])
    prizes = prizes.order_by(*order_fields)
    
    # Paginação
    paginator = Paginator(prizes, 12)
    page_number = request.GET.get('page')
    prizes_page = paginator.get_page(page_number)
    
    context = {
        'prizes': prizes_page,
        'categories': categories,
        'user_balance': request.user.balance_cs,
        'current_category': category_filter,
        'search': search,
        'current_ordering': ordering,
    }
    return render(request, 'prizes/marketplace.html', context)


@login_required
@require_POST
def redeem_prize(request, prize_id):
    """Resgatar um prêmio"""
    prize = get_object_or_404(Prize, id=prize_id, is_active=True)
    
    if not prize.available:
        return JsonResponse({'success': False, 'error': 'Prêmio não disponível'})
    
    # Usar valor final (com desconto se aplicável)
    final_value = prize.final_value
    
    if request.user.balance_cs < final_value:
        return JsonResponse({'success': False, 'error': 'Saldo insuficiente'})
    
    try:
        with transaction.atomic():
            # Criar resgate
            redemption = Redemption.objects.create(
                user=request.user,
                prize=prize,
                status='PENDENTE',
                original_value=prize.value_cs,
                discount_value=prize.discount_amount,
                final_value=final_value
            )
            
            # Debitar saldo (usar valor final)
            request.user.balance_cs -= final_value
            request.user.save()
            
            # Registrar transação
            CSTransaction.objects.create(
                user=request.user,
                amount=-final_value,
                transaction_type='REDEMPTION',
                description=f'Resgate: {prize.name}',
                related_redemption=redemption,
                status='APPROVED',  # Resgates são aprovados automaticamente
                created_by=request.user
            )
            
            messages.success(request, f'Prêmio "{prize.name}" resgatado com sucesso! Aguardando aprovação.')
            return JsonResponse({'success': True})
            
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
def manage_prizes(request):
    """Gerenciar prêmios (admin)"""
    if not request.user.can_manage_users():
        messages.error(request, 'Acesso negado.')
        return redirect('marketplace')
    
    prizes = Prize.objects.all().order_by('-created_at')
    categories = PrizeCategory.objects.all()
    
    # Buscar resgates recentes (últimos 15 dias)
    fifteen_days_ago = timezone.now() - timedelta(days=15)
    recent_redemptions = Redemption.objects.filter(
        redeemed_at__gte=fifteen_days_ago
    ).select_related('user', 'prize').order_by('-redeemed_at')[:10]
    
    context = {
        'prizes': prizes,
        'categories': categories,
        'recent_redemptions': recent_redemptions,
    }
    return render(request, 'prizes/manage.html', context)


@login_required
def create_prize(request):
    """Criar novo prêmio"""
    if not request.user.can_manage_users():
        messages.error(request, 'Acesso negado.')
        return redirect('marketplace')
    
    if request.method == 'POST':
        try:
            # Calcular desconto se aplicável
            has_discount = request.POST.get('has_discount') == 'on'
            discount_percentage = None
            discounted_value = None
            
            if has_discount and request.POST.get('discount_percentage'):
                discount_percentage = Decimal(request.POST['discount_percentage'])
                value_cs = Decimal(request.POST['value_cs'])
                discounted_value = value_cs - (value_cs * discount_percentage / 100)
            
            prize = Prize.objects.create(
                name=request.POST['name'],
                description=request.POST['description'],
                category_id=request.POST.get('category') or None,
                value_cs=Decimal(request.POST['value_cs']),
                has_discount=has_discount,
                discount_percentage=discount_percentage,
                discounted_value=discounted_value,
                priority=request.POST.get('priority', 'NORMAL'),
                stock=int(request.POST.get('stock', 0)),
                unlimited_stock=request.POST.get('unlimited_stock') == 'on',
                terms=request.POST.get('terms', ''),
                valid_until=request.POST.get('valid_until') or None,
                created_by=request.user
            )
            
            if request.FILES.get('image'):
                prize.image = request.FILES['image']
                prize.save()
            
            messages.success(request, f'Prêmio "{prize.name}" criado com sucesso!')
            return redirect('manage_prizes')
            
        except Exception as e:
            messages.error(request, f'Erro ao criar prêmio: {str(e)}')
    
    categories = PrizeCategory.objects.filter(active=True)
    return render(request, 'prizes/create.html', {'categories': categories})


@login_required
def edit_prize(request, prize_id):
    """Editar prêmio existente"""
    if not request.user.can_manage_users():
        messages.error(request, 'Acesso negado.')
        return redirect('marketplace')
    
    prize = get_object_or_404(Prize, id=prize_id)
    
    if request.method == 'POST':
        try:
            # Calcular desconto se aplicável
            has_discount = request.POST.get('has_discount') == 'on'
            discount_percentage = None
            discounted_value = None
            
            if has_discount and request.POST.get('discount_percentage'):
                discount_percentage = Decimal(request.POST['discount_percentage'])
                value_cs = Decimal(request.POST['value_cs'])
                discounted_value = value_cs - (value_cs * discount_percentage / 100)
            
            # Atualizar dados do prêmio
            prize.name = request.POST['name']
            prize.description = request.POST['description']
            prize.category_id = request.POST.get('category') or None
            prize.value_cs = Decimal(request.POST['value_cs'])
            prize.has_discount = has_discount
            prize.discount_percentage = discount_percentage
            prize.discounted_value = discounted_value
            prize.priority = request.POST.get('priority', 'NORMAL')
            prize.stock = int(request.POST.get('stock', 0))
            prize.unlimited_stock = request.POST.get('unlimited_stock') == 'on'
            prize.terms = request.POST.get('terms', '')
            prize.valid_until = request.POST.get('valid_until') or None
            prize.is_active = request.POST.get('is_active') == 'on'
            
            # Gerenciar imagem
            if request.FILES.get('image'):
                prize.image = request.FILES['image']
            elif request.POST.get('remove_image') == 'on':
                prize.image = None
            
            prize.save()
            
            messages.success(request, f'Prêmio "{prize.name}" atualizado com sucesso!')
            return redirect('manage_prizes')
            
        except Exception as e:
            messages.error(request, f'Erro ao atualizar prêmio: {str(e)}')
    
    categories = PrizeCategory.objects.filter(active=True)
    redemptions = Redemption.objects.filter(prize=prize).select_related('user').order_by('-redeemed_at')[:10]
    
    context = {
        'prize': prize,
        'categories': categories,
        'redemptions': redemptions,
    }
    return render(request, 'prizes/edit.html', context)


@login_required
def manage_categories(request):
    """Gerenciar categorias de prêmios"""
    if not request.user.can_manage_users():
        messages.error(request, 'Acesso negado.')
        return redirect('marketplace')
    
    if request.method == 'POST':
        try:
            PrizeCategory.objects.create(
                name=request.POST['name'],
                description=request.POST.get('description', ''),
                icon=request.POST.get('icon', 'fas fa-gift'),
                color=request.POST.get('color', 'blue'),
            )
            messages.success(request, 'Categoria criada com sucesso!')
        except Exception as e:
            messages.error(request, f'Erro ao criar categoria: {str(e)}')
    
    categories = PrizeCategory.objects.all()
    return render(request, 'prizes/categories.html', {'categories': categories})


@login_required
def redemption_history(request):
    """Histórico de resgates do usuário"""
    redemptions = Redemption.objects.filter(user=request.user).order_by('-redeemed_at')
    
    paginator = Paginator(redemptions, 10)
    page_number = request.GET.get('page')
    redemptions_page = paginator.get_page(page_number)
    
    return render(request, 'prizes/redemption_history.html', {'redemptions': redemptions_page})


@login_required
def manage_redemptions(request):
    """Gerenciar resgates (admin)"""
    if not request.user.can_manage_users():
        messages.error(request, 'Acesso negado.')
        return redirect('marketplace')
    
    from django.db.models import Q
    
    # Filtros
    search = request.GET.get('search', '')
    status_filter = request.GET.get('status', '')
    
    redemptions = Redemption.objects.select_related('user', 'user__sector', 'prize', 'prize__category').prefetch_related('user__sectors').all()

    if search:
        redemptions = redemptions.filter(
            Q(user__first_name__icontains=search) |
            Q(user__last_name__icontains=search) |
            Q(user__email__icontains=search) |
            Q(prize__name__icontains=search)
        )
    
    if status_filter:
        redemptions = redemptions.filter(status=status_filter)
    
    redemptions = redemptions.order_by('-redeemed_at')
    
    # Estatísticas
    stats = {
        'pending': Redemption.objects.filter(status='PENDENTE').count(),
        'approved': Redemption.objects.filter(status='APROVADO').count(),
        'delivered': Redemption.objects.filter(status='ENTREGUE').count(),
        'canceled': Redemption.objects.filter(status='CANCELADO').count(),
    }
    
    # Paginação
    paginator = Paginator(redemptions, 20)
    page_number = request.GET.get('page')
    redemptions_page = paginator.get_page(page_number)
    
    context = {
        'redemptions': redemptions_page,
        'current_status': status_filter,
        'status_choices': Redemption.STATUS_CHOICES,
        'stats': stats,
    }
    return render(request, 'prizes/manage_redemptions.html', context)


@login_required
def export_redemptions_excel(request):
    """Exportar resgates para Excel (respeita os filtros aplicados)"""
    if not request.user.can_manage_users():
        messages.error(request, 'Acesso negado.')
        return redirect('marketplace')

    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    from django.db.models import Q

    # Mesmos filtros da listagem
    search = request.GET.get('search', '')
    status_filter = request.GET.get('status', '')

    redemptions = Redemption.objects.select_related('user', 'user__sector', 'prize', 'prize__category', 'approved_by').prefetch_related('user__sectors').all()

    if search:
        redemptions = redemptions.filter(
            Q(user__first_name__icontains=search) |
            Q(user__last_name__icontains=search) |
            Q(user__email__icontains=search) |
            Q(prize__name__icontains=search)
        )

    if status_filter:
        redemptions = redemptions.filter(status=status_filter)

    redemptions = redemptions.order_by('-redeemed_at')

    # Criar workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Resgates"

    headers = [
        'ID', 'Usuário', 'E-mail', 'Setor', 'Telefone', 'Prêmio', 'Categoria', 'Custo (C$)',
        'Status', 'Data do Resgate', 'Data da Aprovação', 'Data da Entrega',
        'Aprovado/Atualizado por', 'Observações'
    ]

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    status_map = dict(Redemption.STATUS_CHOICES)

    for row, redemption in enumerate(redemptions, 2):
        primary_sector = redemption.user.primary_sector
        ws.cell(row=row, column=1, value=redemption.id)
        ws.cell(row=row, column=2, value=redemption.user.full_name)
        ws.cell(row=row, column=3, value=redemption.user.email or "")
        ws.cell(row=row, column=4, value=primary_sector.name if primary_sector else "")
        ws.cell(row=row, column=5, value=redemption.user.phone or "")
        ws.cell(row=row, column=6, value=redemption.prize.name)
        ws.cell(row=row, column=7, value=redemption.prize.category.name if redemption.prize.category else "")
        ws.cell(row=row, column=8, value=float(redemption.prize.value_cs))
        ws.cell(row=row, column=9, value=status_map.get(redemption.status, redemption.status))
        ws.cell(row=row, column=10, value=redemption.redeemed_at.strftime("%d/%m/%Y %H:%M") if redemption.redeemed_at else "")
        ws.cell(row=row, column=11, value=redemption.approved_at.strftime("%d/%m/%Y %H:%M") if redemption.approved_at else "")
        ws.cell(row=row, column=12, value=redemption.delivered_at.strftime("%d/%m/%Y %H:%M") if redemption.delivered_at else "")
        ws.cell(row=row, column=13, value=redemption.approved_by.full_name if redemption.approved_by else "")
        ws.cell(row=row, column=14, value=redemption.notes or redemption.delivery_notes or "")

    # Ajustar largura das colunas
    for col in range(1, len(headers) + 1):
        column_letter = get_column_letter(col)
        max_length = 0
        for cell in ws[column_letter]:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except Exception:
                pass
        ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="resgates_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    wb.save(response)
    return response


@login_required
def update_redemption_status(request, redemption_id):
    """Atualizar status do resgate"""
    if not request.user.can_manage_users():
        return JsonResponse({'success': False, 'error': 'Acesso negado'})
    
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Método não permitido'})
    
    redemption = get_object_or_404(Redemption, id=redemption_id)
    
    # Aceitar dados JSON ou form data
    if request.content_type == 'application/json':
        import json
        data = json.loads(request.body)
        new_status = data.get('status')
        delivery_notes = data.get('delivery_notes', '')
    else:
        new_status = request.POST.get('status')
        delivery_notes = request.POST.get('delivery_notes', '')
    
    if new_status not in dict(Redemption.STATUS_CHOICES):
        return JsonResponse({'success': False, 'error': 'Status inválido'})
    
    try:
        old_status = redemption.status
        redemption.status = new_status
        redemption.approved_by = request.user
        
        if new_status == 'APROVADO':
            redemption.approved_at = timezone.now()
            if delivery_notes:
                redemption.delivery_notes = delivery_notes
        elif new_status == 'ENTREGUE':
            redemption.delivered_at = timezone.now()
            if delivery_notes:
                redemption.delivery_notes = delivery_notes
        elif new_status == 'CANCELADO':
            if delivery_notes:
                redemption.notes = delivery_notes
        
        redemption.save()
        
        # Gerenciar estoque baseado no status
        if new_status == 'CANCELADO' and old_status != 'CANCELADO':
            # Se cancelado, devolver o C$ para o usuário
            redemption.user.balance_cs += redemption.prize.value_cs
            redemption.user.save()
            
            # Registrar transação de devolução
            CSTransaction.objects.create(
                user=redemption.user,
                amount=redemption.prize.value_cs,
                transaction_type='CREDIT',
                description=f'Devolução por cancelamento: {redemption.prize.name}',
                related_redemption=redemption,
                status='APPROVED',  # Devoluções são aprovadas automaticamente
                created_by=request.user
            )
        
        # Log da ação
        from core.middleware import log_action
        log_action(
            request.user,
            'REDEMPTION_STATUS_UPDATE',
            f'Status do resgate #{redemption.id} alterado de {old_status} para {new_status}',
            request
        )
        
        return JsonResponse({'success': True})
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
@require_POST
def notify_pickup(request, redemption_id):
    """Notifica o usuário que o prêmio está na SEDE e define a data limite de retirada."""
    if not request.user.can_manage_users():
        return JsonResponse({'success': False, 'error': 'Acesso negado'})

    redemption = get_object_or_404(Redemption, id=redemption_id)

    if redemption.status != 'APROVADO':
        return JsonResponse({'success': False, 'error': 'Apenas resgates aprovados podem ser notificados para retirada.'})

    # Aceitar dados JSON ou form data
    if request.content_type == 'application/json':
        import json
        data = json.loads(request.body)
        deadline_raw = (data.get('pickup_deadline') or '').strip()
    else:
        deadline_raw = (request.POST.get('pickup_deadline') or '').strip()

    if not deadline_raw:
        return JsonResponse({'success': False, 'error': 'Informe a data limite para retirada.'})

    from datetime import datetime
    try:
        pickup_deadline = datetime.strptime(deadline_raw, '%Y-%m-%d').date()
    except ValueError:
        return JsonResponse({'success': False, 'error': 'Data inválida.'})

    if pickup_deadline < timezone.now().date():
        return JsonResponse({'success': False, 'error': 'A data limite não pode ser no passado.'})

    try:
        redemption.pickup_deadline = pickup_deadline
        redemption.pickup_notified_at = timezone.now()
        redemption.save(update_fields=['pickup_deadline', 'pickup_notified_at'])

        # Notificação para o usuário
        from core.models import Notification
        deadline_str = pickup_deadline.strftime('%d/%m/%Y')
        Notification.objects.create(
            user=redemption.user,
            title='Prêmio disponível para retirada',
            message=(
                f'Seu resgate de "{redemption.prize.name}" está na SEDE. '
                f'Retire até {deadline_str}.'
            ),
            notification_type='SYSTEM',
            related_object_id=redemption.id,
            related_url='/prizes/my-redemptions/',
        )

        # Log da ação
        from core.middleware import log_action
        log_action(
            request.user,
            'REDEMPTION_PICKUP_NOTIFY',
            f'Usuário notificado sobre retirada do resgate #{redemption.id} até {deadline_str}',
            request
        )

        return JsonResponse({'success': True})

    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
@require_POST
def cancel_redemption(request, redemption_id):
    """Cancelar resgate - Admin ou próprio usuário"""
    redemption = get_object_or_404(Redemption, id=redemption_id)
    
    # Verificar permissão: admin ou próprio usuário
    if not (request.user.can_manage_users() or redemption.user == request.user):
        return JsonResponse({'success': False, 'error': 'Acesso negado'})
    
    # Só pode cancelar se estiver pendente ou aprovado
    if redemption.status not in ['PENDENTE', 'APROVADO']:
        return JsonResponse({'success': False, 'error': 'Este resgate não pode ser cancelado'})
    
    try:
        with transaction.atomic():
            old_status = redemption.status
            redemption.status = 'CANCELADO'
            redemption.approved_by = request.user
            
            # Adicionar nota sobre quem cancelou
            if request.user == redemption.user:
                redemption.notes = f'Cancelado pelo próprio usuário em {timezone.now().strftime("%d/%m/%Y %H:%M")}'
            else:
                redemption.notes = f'Cancelado por {request.user.get_full_name()} em {timezone.now().strftime("%d/%m/%Y %H:%M")}'
            
            redemption.save()
            
            # Devolver o C$ para o usuário
            redemption.user.balance_cs += redemption.prize.value_cs
            redemption.user.save()
            
            # Registrar transação de devolução
            CSTransaction.objects.create(
                user=redemption.user,
                amount=redemption.prize.value_cs,
                transaction_type='REFUND',
                description=f'Devolução por cancelamento: {redemption.prize.name}',
                related_redemption=redemption,
                status='APPROVED',
                created_by=request.user
            )
            
            # Log da ação
            from core.middleware import log_action
            log_action(
                request.user,
                'REDEMPTION_CANCELLED',
                f'Resgate #{redemption.id} ({redemption.prize.name}) cancelado',
                request
            )
            
            return JsonResponse({
                'success': True, 
                'message': f'Resgate cancelado com sucesso! C${redemption.prize.value_cs} foi devolvido.'
            })
            
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
def manage_discounts(request):
    """Gerenciar descontos (admin)"""
    if not request.user.can_manage_users():
        messages.error(request, 'Acesso negado.')
        return redirect('marketplace')
    
    status_filter = request.GET.get('status', 'active')
    discounts = PrizeDiscount.objects.all()
    
    if status_filter == 'active':
        discounts = discounts.filter(is_active=True)
    elif status_filter == 'inactive':
        discounts = discounts.filter(is_active=False)
    elif status_filter == 'expired':
        discounts = discounts.filter(valid_until__lt=timezone.now().date())
    
    discounts = discounts.order_by('-created_at')
    
    # Estatísticas
    total_discounts = PrizeDiscount.objects.count()
    active_discounts = PrizeDiscount.objects.filter(is_active=True).count()
    
    paginator = Paginator(discounts, 20)
    page_number = request.GET.get('page')
    discounts_page = paginator.get_page(page_number)
    
    context = {
        'discounts': discounts_page,
        'current_status': status_filter,
        'total_discounts': total_discounts,
        'active_discounts': active_discounts,
    }
    return render(request, 'prizes/discount_list.html', context)


@login_required
def create_discount(request):
    """Criar novo desconto"""
    if not request.user.can_manage_users():
        messages.error(request, 'Acesso negado.')
        return redirect('marketplace')
    
    if request.method == 'POST':
        try:
            discount = PrizeDiscount.objects.create(
                name=request.POST['name'],
                code=request.POST['code'].upper().strip(),
                description=request.POST.get('description', ''),
                discount_type=request.POST['discount_type'],
                discount_value=Decimal(request.POST['discount_value']),
                min_purchase_value=Decimal(request.POST.get('min_purchase_value', 0)),
                max_discount_value=Decimal(request.POST['max_discount_value']) if request.POST.get('max_discount_value') else None,
                valid_from=request.POST['valid_from'],
                valid_until=request.POST['valid_until'],
                max_uses=int(request.POST['max_uses']) if request.POST.get('max_uses') else None,
                created_by=request.user
            )
            
            # Adicionar categorias aplicáveis
            if request.POST.getlist('categories'):
                discount.applies_to_categories.set(request.POST.getlist('categories'))
            
            messages.success(request, f'Desconto "{discount.name}" criado com sucesso!')
            return redirect('manage_discounts')
            
        except Exception as e:
            messages.error(request, f'Erro ao criar desconto: {str(e)}')
    
    categories = PrizeCategory.objects.filter(active=True)
    context = {
        'categories': categories,
        'is_edit': False,
    }
    return render(request, 'prizes/discount_form.html', context)


@login_required
def edit_discount(request, discount_id):
    """Editar desconto"""
    if not request.user.can_manage_users():
        messages.error(request, 'Acesso negado.')
        return redirect('marketplace')
    
    discount = get_object_or_404(PrizeDiscount, id=discount_id)
    
    if request.method == 'POST':
        try:
            discount.name = request.POST['name']
            discount.code = request.POST['code'].upper().strip()
            discount.description = request.POST.get('description', '')
            discount.discount_type = request.POST['discount_type']
            discount.discount_value = Decimal(request.POST['discount_value'])
            discount.min_purchase_value = Decimal(request.POST.get('min_purchase_value', 0))
            discount.max_discount_value = Decimal(request.POST['max_discount_value']) if request.POST.get('max_discount_value') else None
            discount.valid_from = request.POST['valid_from']
            discount.valid_until = request.POST['valid_until']
            discount.max_uses = int(request.POST['max_uses']) if request.POST.get('max_uses') else None
            discount.is_active = request.POST.get('is_active') == 'on'
            discount.save()
            
            # Atualizar categorias aplicáveis
            if request.POST.getlist('categories'):
                discount.applies_to_categories.set(request.POST.getlist('categories'))
            else:
                discount.applies_to_categories.clear()
            
            messages.success(request, f'Desconto "{discount.name}" atualizado com sucesso!')
            return redirect('manage_discounts')
            
        except Exception as e:
            messages.error(request, f'Erro ao atualizar desconto: {str(e)}')
    
    categories = PrizeCategory.objects.filter(active=True)
    context = {
        'discount': discount,
        'categories': categories,
        'is_edit': True,
    }
    return render(request, 'prizes/discount_form.html', context)


@login_required
@require_POST
def delete_discount(request, discount_id):
    """Excluir desconto"""
    if not request.user.can_manage_users():
        return JsonResponse({'success': False, 'error': 'Acesso negado'})
    
    discount = get_object_or_404(PrizeDiscount, id=discount_id)
    
    # Verificar se há resgates usando este desconto
    redemptions_count = Redemption.objects.filter(discount=discount).count()
    
    if redemptions_count > 0:
        return JsonResponse({
            'success': False, 
            'error': f'Este desconto não pode ser excluído pois está sendo usado em {redemptions_count} resgate(s). Você pode desativá-lo.'
        })
    
    try:
        discount_name = discount.name
        discount.delete()
        
        # Log da ação
        from core.middleware import log_action
        log_action(
            request.user,
            'DISCOUNT_DELETED',
            f'Desconto "{discount_name}" excluído',
            request
        )
        
        return JsonResponse({'success': True, 'message': f'Desconto "{discount_name}" excluído com sucesso!'})
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
def validate_discount_code(request):
    """Validar código de desconto (AJAX)"""
    code = request.GET.get('code', '').upper().strip()
    prize_id = request.GET.get('prize_id')
    
    if not code:
        return JsonResponse({'valid': False, 'error': 'Código não informado'})
    
    try:
        discount = PrizeDiscount.objects.get(code=code)
        prize = get_object_or_404(Prize, id=prize_id)
        
        if not discount.is_valid:
            return JsonResponse({'valid': False, 'error': 'Desconto inválido ou expirado'})
        
        if not discount.can_apply_to_prize(prize):
            return JsonResponse({'valid': False, 'error': 'Este desconto não se aplica a este prêmio'})
        
        discount_amount = discount.calculate_discount(prize.value_cs)
        final_value = prize.value_cs - discount_amount
        
        return JsonResponse({
            'valid': True,
            'discount_name': discount.name,
            'discount_type': discount.get_discount_type_display(),
            'discount_value': str(discount.discount_value),
            'discount_amount': str(discount_amount),
            'original_value': str(prize.value_cs),
            'final_value': str(final_value),
        })
        
    except PrizeDiscount.DoesNotExist:
        return JsonResponse({'valid': False, 'error': 'Código de desconto não encontrado'})
    except Exception as e:
        return JsonResponse({'valid': False, 'error': str(e)})
