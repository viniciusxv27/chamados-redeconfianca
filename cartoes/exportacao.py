"""Exportação em Excel do extrato e do relatório de conciliação."""
from datetime import date
from decimal import Decimal
from io import BytesIO

from django.http import HttpResponse

ZERO = Decimal('0.00')

CABECALHO = {'bold': True, 'fundo': 'FF4C1D95'}
DINHEIRO = 'R$ #,##0.00'
DATA_BR = 'DD/MM/YYYY'


def _estilizar_cabecalho(planilha, colunas):
    from openpyxl.styles import Alignment, Font, PatternFill
    fundo = PatternFill('solid', fgColor=CABECALHO['fundo'])
    for i, (titulo, largura) in enumerate(colunas, start=1):
        celula = planilha.cell(row=1, column=i, value=titulo)
        celula.font = Font(bold=True, color='FFFFFFFF')
        celula.fill = fundo
        celula.alignment = Alignment(horizontal='center', vertical='center')
        planilha.column_dimensions[celula.column_letter].width = largura
    planilha.freeze_panes = 'A2'


def _resposta(livro, nome):
    buffer = BytesIO()
    livro.save(buffer)
    buffer.seek(0)
    resposta = HttpResponse(
        buffer.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resposta['Content-Disposition'] = f'attachment; filename="{nome}"'
    return resposta


def extrato_excel(cartao, gastos, inicio=None, fim=None):
    """Extrato do cartão no período, uma linha por gasto."""
    from openpyxl import Workbook

    livro = Workbook()
    aba = livro.active
    aba.title = 'Extrato'
    _estilizar_cabecalho(aba, [
        ('Data', 12), ('Estabelecimento', 38), ('Categoria', 22),
        ('Valor (R$)', 14), ('Origem', 10), ('Lançado por', 26),
        ('Chamado', 12), ('Descrição', 46),
    ])

    linha = 2
    for gasto in gastos:
        aba.cell(row=linha, column=1, value=gasto.data_gasto).number_format = DATA_BR
        aba.cell(row=linha, column=2, value=gasto.estabelecimento or '')
        aba.cell(row=linha, column=3, value=gasto.categoria_gasto or '')
        aba.cell(row=linha, column=4, value=float(gasto.valor)).number_format = DINHEIRO
        aba.cell(row=linha, column=5, value=gasto.get_origem_display())
        aba.cell(row=linha, column=6,
                 value=getattr(gasto.criado_por, 'full_name', '') or '')
        aba.cell(row=linha, column=7, value=gasto.ticket_id or '')
        aba.cell(row=linha, column=8, value=gasto.descricao or '')
        linha += 1

    if linha > 2:
        from openpyxl.styles import Font
        aba.cell(row=linha, column=3, value='TOTAL').font = Font(bold=True)
        total = aba.cell(row=linha, column=4, value=f'=SUM(D2:D{linha - 1})')
        total.number_format = DINHEIRO
        total.font = Font(bold=True)

    periodo = ''
    if inicio or fim:
        periodo = f"_{(inicio or date.min):%Y%m%d}-{(fim or date.today()):%Y%m%d}"
    return _resposta(livro, f'extrato_cartao_{cartao.last4}{periodo}.xlsx')


def conciliacao_excel(cartao, relatorio, referencia=None):
    """Relatório de conciliação em abas: divergências, não localizados, conferidos."""
    from openpyxl import Workbook
    from openpyxl.styles import Font

    livro = Workbook()

    # ── Resumo ──────────────────────────────────────────────────────────────
    resumo = livro.active
    resumo.title = 'Resumo'
    resumo.column_dimensions['A'].width = 42
    resumo.column_dimensions['B'].width = 18
    linhas = [
        ('Cartão', f'{cartao.titulo} (final {cartao.last4})'),
        ('Fatura de referência', referencia.strftime('%m/%Y') if referencia else '—'),
        ('', ''),
        ('Total da fatura', float(relatorio['total_fatura'])),
        ('Total lançado no portal', float(relatorio['total_extrato'])),
        ('Diferença', float(relatorio['diferenca_total'])),
        ('', ''),
        ('Lançamentos conferidos', len(relatorio['conferidos'])),
        ('Divergências de valor', len(relatorio['divergentes'])),
        ('Na fatura e não lançados', len(relatorio['so_na_fatura'])),
        ('Lançados e ausentes da fatura', len(relatorio['so_no_extrato'])),
    ]
    for i, (rotulo, valor) in enumerate(linhas, start=1):
        resumo.cell(row=i, column=1, value=rotulo).font = Font(bold=bool(rotulo))
        celula = resumo.cell(row=i, column=2, value=valor)
        if isinstance(valor, float):
            celula.number_format = DINHEIRO

    # ── Divergências de valor ───────────────────────────────────────────────
    aba = livro.create_sheet('Divergências')
    _estilizar_cabecalho(aba, [
        ('Data na fatura', 14), ('Estabelecimento (fatura)', 34), ('Valor na fatura', 16),
        ('Data no portal', 14), ('Estabelecimento (portal)', 34), ('Valor no portal', 16),
        ('Diferença', 14),
    ])
    for i, item in enumerate(relatorio['divergentes'], start=2):
        f, g = item['fatura'], item['gasto']
        aba.cell(row=i, column=1, value=f['data']).number_format = DATA_BR
        aba.cell(row=i, column=2, value=f['estabelecimento'])
        aba.cell(row=i, column=3, value=float(f['valor'])).number_format = DINHEIRO
        aba.cell(row=i, column=4, value=g.data_gasto).number_format = DATA_BR
        aba.cell(row=i, column=5, value=g.estabelecimento or '')
        aba.cell(row=i, column=6, value=float(g.valor)).number_format = DINHEIRO
        aba.cell(row=i, column=7, value=float(item['diferenca'])).number_format = DINHEIRO

    # ── Na fatura, sem lançamento ───────────────────────────────────────────
    aba = livro.create_sheet('Não lançados')
    _estilizar_cabecalho(aba, [
        ('Data', 14), ('Estabelecimento', 40), ('Parcela', 10), ('Valor (R$)', 16)])
    for i, item in enumerate(relatorio['so_na_fatura'], start=2):
        aba.cell(row=i, column=1, value=item['data']).number_format = DATA_BR
        aba.cell(row=i, column=2, value=item['estabelecimento'])
        aba.cell(row=i, column=3, value=item['parcela'] or '')
        aba.cell(row=i, column=4, value=float(item['valor'])).number_format = DINHEIRO

    # ── Lançado no portal, sem cobrança ─────────────────────────────────────
    aba = livro.create_sheet('Sem cobrança')
    _estilizar_cabecalho(aba, [
        ('Data', 14), ('Estabelecimento', 40), ('Valor (R$)', 16), ('Lançado por', 26)])
    for i, gasto in enumerate(relatorio['so_no_extrato'], start=2):
        aba.cell(row=i, column=1, value=gasto.data_gasto).number_format = DATA_BR
        aba.cell(row=i, column=2, value=gasto.estabelecimento or '')
        aba.cell(row=i, column=3, value=float(gasto.valor)).number_format = DINHEIRO
        aba.cell(row=i, column=4,
                 value=getattr(gasto.criado_por, 'full_name', '') or '')

    # ── Conferidos ──────────────────────────────────────────────────────────
    aba = livro.create_sheet('Conferidos')
    _estilizar_cabecalho(aba, [
        ('Data na fatura', 14), ('Estabelecimento', 40), ('Valor (R$)', 16),
        ('Data no portal', 14), ('Lançado por', 26)])
    for i, item in enumerate(relatorio['conferidos'], start=2):
        f, g = item['fatura'], item['gasto']
        aba.cell(row=i, column=1, value=f['data']).number_format = DATA_BR
        aba.cell(row=i, column=2, value=f['estabelecimento'])
        aba.cell(row=i, column=3, value=float(f['valor'])).number_format = DINHEIRO
        aba.cell(row=i, column=4, value=g.data_gasto).number_format = DATA_BR
        aba.cell(row=i, column=5,
                 value=getattr(g.criado_por, 'full_name', '') or '')

    marca = f"_{referencia:%Y%m}" if referencia else ''
    return _resposta(livro, f'conciliacao_cartao_{cartao.last4}{marca}.xlsx')
