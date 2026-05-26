"""Serviços do módulo Fibras: sync do MySQL e notificações."""
from __future__ import annotations

import re
import unicodedata
from datetime import date, timedelta
from decimal import Decimal
from typing import Iterable, Optional

from django.contrib.auth import get_user_model
from django.utils import timezone

from simulator.sql_realizado import _mysql_config

from .models import Fibra, FibraStatusHistory, PlanilhaOrdemInconsistente


User = get_user_model()


def _normalize(value) -> str:
    if value is None:
        return ''
    text = str(value).strip()
    text = unicodedata.normalize('NFKD', text)
    text = ''.join(ch for ch in text if not unicodedata.combining(ch))
    return ' '.join(text.upper().split())


def _digits_only(value) -> str:
    """Retorna apenas os dígitos de um identificador (remove letras, '-', espaços)."""
    if value is None:
        return ''
    return re.sub(r'\D+', '', str(value))


def fetch_fibras_from_mysql(
    *, year: Optional[int] = None, month: Optional[int] = None,
) -> list[dict]:
    """Lê as vendas de Fibra do MySQL do mês corrente.

    Fibra = ``vendas_servicos`` com ``pilar = 'FIXA'``.
    """
    try:
        import pymysql  # type: ignore
    except ImportError:
        return []

    now = timezone.now()
    year = year or now.year
    month = month or now.month

    try:
        conn = pymysql.connect(**_mysql_config())
    except Exception:
        return []

    rows: list[dict] = []
    try:
        with conn.cursor() as cur:
            cur.execute(
                """
                SELECT
                    `ID_da_venda`,
                    `Nº_protocolo`,
                    `CPF_do_cliente`,
                    `Nome_do_cliente`,
                    `endereço_do_cliente`,
                    `Nº_acesso`,
                    `Plano_novo`,
                    COALESCE(`receita_calculada`, `Receita`, 0) AS valor,
                    `PDV`,
                    `Nome_do_vendedor`,
                    `data_da_venda`,
                    `pilar`,
                    `Serviço_Técnico`,
                    `Venda_ativa`
                FROM vendas_servicos
                WHERE YEAR(data_da_venda) = %s
                  AND MONTH(data_da_venda) = %s
                  AND UPPER(`pilar`) = 'FIXA'
                """,
                (year, month),
            )
            cols = [
                'numero_da_venda', 'numero_protocolo', 'cpf', 'cliente', 'endereco', 'numero_acesso',
                'plano', 'valor', 'pdv', 'vendedor', 'data_da_venda',
                'pilar', 'servico_tecnico', 'venda_ativa',
            ]
            for row in cur.fetchall():
                rows.append(dict(zip(cols, row)))
    finally:
        try:
            conn.close()
        except Exception:
            pass

    return rows


def fetch_fibras_by_protocols(protocols: Iterable[str]) -> list[dict]:
    """Busca no MySQL todas as vendas cujos ``Nº_protocolo`` estejam em ``protocols``.

    Não filtra por mês/pilar — usa só o protocolo para garantir o cruzamento
    direto com o que veio da planilha.
    """
    protocols = [p for p in {str(p).strip() for p in protocols} if p]
    if not protocols:
        return []

    try:
        import pymysql  # type: ignore
    except ImportError:
        return []

    try:
        conn = pymysql.connect(**_mysql_config())
    except Exception:
        return []

    rows: list[dict] = []
    cols = [
        'numero_da_venda', 'numero_protocolo', 'cpf', 'cliente', 'endereco', 'numero_acesso',
        'plano', 'valor', 'pdv', 'vendedor', 'data_da_venda',
        'pilar', 'servico_tecnico', 'venda_ativa',
    ]
    try:
        with conn.cursor() as cur:
            CHUNK = 500
            for i in range(0, len(protocols), CHUNK):
                chunk = protocols[i:i + CHUNK]
                placeholders = ','.join(['%s'] * len(chunk))
                cur.execute(
                    f"""
                    SELECT
                        `ID_da_venda`,
                        `Nº_protocolo`,
                        `CPF_do_cliente`,
                        `Nome_do_cliente`,
                        `endereço_do_cliente`,
                        `Nº_acesso`,
                        `Plano_novo`,
                        COALESCE(`receita_calculada`, `Receita`, 0) AS valor,
                        `PDV`,
                        `Nome_do_vendedor`,
                        `data_da_venda`,
                        `pilar`,
                        `Serviço_Técnico`,
                        `Venda_ativa`
                    FROM vendas_servicos
                    WHERE `Nº_protocolo` IN ({placeholders})
                    """,
                    chunk,
                )
                for row in cur.fetchall():
                    rows.append(dict(zip(cols, row)))
    finally:
        try:
            conn.close()
        except Exception:
            pass

    return rows


def _upsert_fibra_from_mysql_row(r: dict) -> tuple[Optional[Fibra], bool]:
    """Cria ou atualiza uma Fibra a partir de um dict do MySQL.

    Retorna ``(fibra, created)``. Se o número da venda for vazio devolve
    ``(None, False)``.
    """
    numero = (str(r.get('numero_da_venda') or '')).strip()
    if not numero:
        return None, False

    venda_ativa = (str(r.get('venda_ativa') or '')).strip()
    cancelada = venda_ativa == '0'

    common = {
        'numero_protocolo': (str(r.get('numero_protocolo') or '')).strip(),
        'cpf': (r.get('cpf') or '').strip() if r.get('cpf') else '',
        'cliente': (r.get('cliente') or '').strip() if r.get('cliente') else '',
        'endereco': (r.get('endereco') or '').strip() if r.get('endereco') else '',
        'numero_acesso': (str(r.get('numero_acesso') or '')).strip(),
        'plano': (r.get('plano') or '').strip() if r.get('plano') else '',
        'valor': Decimal(str(r.get('valor') or 0)),
        'pdv': (r.get('pdv') or '').strip() if r.get('pdv') else '',
        'vendedor': (r.get('vendedor') or '').strip() if r.get('vendedor') else '',
        'data_da_venda': r.get('data_da_venda'),
        'pilar': (r.get('pilar') or '').strip() if r.get('pilar') else '',
        'servico_tecnico': (r.get('servico_tecnico') or '').strip() if r.get('servico_tecnico') else '',
    }

    existing = Fibra.objects.filter(numero_da_venda=numero).first()
    if existing:
        for field, value in common.items():
            setattr(existing, field, value)
        if cancelada and existing.status != Fibra.STATUS_CANCELADO:
            existing.status = Fibra.STATUS_CANCELADO
        existing.save()
        return existing, False

    initial_status = Fibra.STATUS_CANCELADO if cancelada else Fibra.STATUS_AGENDADO
    fibra = Fibra.objects.create(
        numero_da_venda=numero,
        status=initial_status,
        **common,
    )
    return fibra, True


def sync_fibras(*, year: Optional[int] = None, month: Optional[int] = None) -> dict:
    """Atualiza/insere registros de Fibra a partir do MySQL.

    Para vendas já existentes localmente NÃO sobrescreve ``status`` nem
    ``retorno_myrella`` (Myrella já pode ter mexido). Exceção: se a venda
    foi cancelada no MySQL (``Venda_ativa='0'``), o status local vira
    ``STATUS_CANCELADO`` automaticamente.

    Retorna estatísticas (created, updated).
    """
    raw = fetch_fibras_from_mysql(year=year, month=month)
    created = 0
    updated = 0
    for r in raw:
        _, was_created = _upsert_fibra_from_mysql_row(r)
        if was_created:
            created += 1
        else:
            updated += 1

    return {'created': created, 'updated': updated, 'total_in_source': len(raw)}


# ---------------------------------------------------------------------------
# Filtragem por usuário (consultor / coordenador / gerente)
# ---------------------------------------------------------------------------

def is_gerente(user) -> bool:
    """Gerente = hierarchy PADRAO e está no grupo de comunicação 'GERENTES'."""
    if getattr(user, 'hierarchy', '') != 'PADRAO':
        return False
    try:
        return user.communication_groups.filter(name__iexact='GERENTES').exists()
    except Exception:
        return False


def _user_loja_names(user) -> list[str]:
    """Nome(s) de loja do gerente, a partir de ``user.sector`` e ``user.sectors``."""
    names: list[str] = []
    sector = getattr(user, 'sector', None)
    if sector and getattr(sector, 'name', ''):
        names.append(sector.name.strip())
    try:
        for s in user.sectors.all():
            n = (getattr(s, 'name', '') or '').strip()
            if n and n not in names:
                names.append(n)
    except Exception:
        pass
    return names


def fibras_for_user(user) -> 'models.QuerySet[Fibra]':
    """Devolve um queryset das fibras visíveis para o usuário.

    - Padrão consultor: apenas as fibras em que ele é o vendedor.
    - Gerente (PADRAO + grupo GERENTES): todas as fibras do(s) seu(s) setor(es)
      (match por PDV contendo o nome da loja).
    - Coordenador/Admin/SuperAdmin: todas (filtros aplicados na view).
    """
    from django.db.models import Q

    qs = Fibra.objects.all()
    hierarchy = getattr(user, 'hierarchy', 'PADRAO')
    if hierarchy != 'PADRAO':
        return qs

    if is_gerente(user):
        lojas = _user_loja_names(user)
        if not lojas:
            return qs.none()
        cond = Q()
        for nome in lojas:
            cond |= Q(pdv__icontains=nome)
        return qs.filter(cond)

    nome = _normalize(user.get_full_name() or user.username)
    ids = [
        f.id for f in qs.only('id', 'vendedor')
        if _normalize(f.vendedor) == nome
    ]
    return qs.filter(id__in=ids)


# ---------------------------------------------------------------------------
# Importação da planilha diária
# ---------------------------------------------------------------------------

# Mapeia palavras-chave do texto bruto de status na planilha para o status
# interno do Fibra. A correspondência é por substring após normalização,
# avaliada na ordem declarada (a primeira que casar vence).
_PLANILHA_STATUS_MAP = (
    # Códigos numéricos da planilha "STATUS INSTALACAO"
    ('1.',          Fibra.STATUS_INSTALADO),   # 1.1 INSTALADA
    ('3.5',         Fibra.STATUS_AGENDADO),    # 3.5 AGENDADA
    ('3.8',         Fibra.STATUS_CANCELADO),   # 3.8 CANCELADA
    ('3.6',         Fibra.STATUS_PROBLEMA),    # 3.6 NENHUM REGISTRO ENCONTRADO NO WFM
    ('3.4',         Fibra.STATUS_PROBLEMA),    # 3.4 PENDENTE TECNICA
    ('3.1',         Fibra.STATUS_PENDENTE),    # 3.1 PENDENTE ENRIQUECIMENTO
    ('3.2',         Fibra.STATUS_PENDENTE),    # 3.2 PENDENTE AGENDAMENTO
    ('3.3',         Fibra.STATUS_PENDENTE),    # 3.3 PENDENTE RETENÇÃO
    ('2.',          Fibra.STATUS_CANCELADO),   # 2.x CANCELAMENTO
    # Fallback por palavras-chave
    ('INSTALAD',    Fibra.STATUS_INSTALADO),
    ('CONCLU',      Fibra.STATUS_INSTALADO),
    ('ATIVAD',      Fibra.STATUS_INSTALADO),
    ('AGENDAD',     Fibra.STATUS_AGENDADO),
    ('CANCELAD',    Fibra.STATUS_CANCELADO),
    ('FRAUDE',      Fibra.STATUS_CANCELADO),
    ('PENDENTE',    Fibra.STATUS_PENDENTE),
    ('AGUARDAN',    Fibra.STATUS_PENDENTE),
    ('NENHUM',      Fibra.STATUS_PROBLEMA),
    ('PROBLEM',     Fibra.STATUS_PROBLEMA),
    ('IMPRODUT',    Fibra.STATUS_PROBLEMA),
)


def _map_planilha_status(raw: str) -> Optional[str]:
    n = _normalize(raw)
    if not n:
        return None
    for needle, status in _PLANILHA_STATUS_MAP:
        if needle in n:
            return status
    return None


def import_planilha_xlsx(file_obj, *, by_user=None) -> dict:
    """Importa a planilha diária e atualiza as fibras casando por protocolo.

    Lê o arquivo .xlsx com ``openpyxl``, localiza:
      - Coluna do protocolo (header ``ORDEM`` — equivalente ao ``Nº_protocolo``
        no MySQL — ou, como fallback, header contendo ``PROTOCOLO``).
      - Coluna de status (header contém ``STATUS`` ou ``SITUA``;
        prioriza ``STATUS INSTALACAO``).

    Para cada linha:
      - Tenta achar a ``Fibra`` local por ``numero_protocolo``.
      - Se não existir, busca no MySQL (``vendas_servicos.Nº_protocolo``) e
        faz upsert local, então casa novamente.
      - Atualiza ``ordem_planilha`` (posição da linha), ``status`` (via
        :func:`_map_planilha_status`) e ``status_planilha_raw``.

    Retorna ``{'matched', 'updated_status', 'created_from_mysql', 'not_found',
    'not_found_count', 'rows'}``.
    """
    try:
        import openpyxl  # type: ignore
    except ImportError as e:
        raise RuntimeError('openpyxl não instalado no servidor.') from e

    wb = openpyxl.load_workbook(file_obj, data_only=True, read_only=True)
    ws = wb.active

    # Encontra a linha de cabeçalho.
    header_row_idx = None
    headers: list[str] = []
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=15, values_only=True), start=1):
        norm = [_normalize(c) for c in row]
        if any(c == 'ORDEM' or 'PROTOCOLO' in c for c in norm):
            header_row_idx = i
            headers = norm
            break
    if header_row_idx is None:
        raise RuntimeError("Coluna 'ORDEM' (protocolo) não encontrada na planilha.")

    # Prioriza 'ORDEM' exato; cai para qualquer coluna com 'PROTOCOLO'.
    proto_idx = next((j for j, c in enumerate(headers) if c == 'ORDEM'), None)
    if proto_idx is None:
        proto_idx = next((j for j, c in enumerate(headers) if 'PROTOCOLO' in c), None)
    if proto_idx is None:
        raise RuntimeError("Coluna 'ORDEM' (protocolo) não encontrada na planilha.")

    # Prioriza 'STATUS INSTALACAO'; cai para qualquer coluna com STATUS/SITUA.
    status_idx = next(
        (j for j, c in enumerate(headers) if 'STATUS' in c and 'INSTALAC' in c),
        None,
    )
    if status_idx is None:
        status_idx = next(
            (j for j, c in enumerate(headers) if 'STATUS' in c or 'SITUA' in c),
            None,
        )

    # Coluna 'SLA AGENDA' (ou contendo SLA + AGEND); fallback só SLA.
    sla_idx = next(
        (j for j, c in enumerate(headers) if 'SLA' in c and 'AGEND' in c),
        None,
    )
    if sla_idx is None:
        sla_idx = next((j for j, c in enumerate(headers) if 'SLA' in c), None)

    # Coluna 'MOTIVO' (usada em vendas com problema).
    motivo_idx = next((j for j, c in enumerate(headers) if 'MOTIVO' in c), None)

    # 1ª passada: coleta protocolos e status_raw mantendo a ordem da planilha.
    parsed: list[tuple[str, str, str, str]] = []  # (proto, status_raw, sla, motivo)
    for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
        if not row or all(c is None or str(c).strip() == '' for c in row):
            continue
        proto = (str(row[proto_idx]).strip() if proto_idx < len(row) and row[proto_idx] is not None else '')
        status_raw = ''
        if status_idx is not None and status_idx < len(row) and row[status_idx] is not None:
            status_raw = str(row[status_idx]).strip()
        sla_val = ''
        if sla_idx is not None and sla_idx < len(row) and row[sla_idx] is not None:
            sla_val = str(row[sla_idx]).strip()
        motivo_val = ''
        if motivo_idx is not None and motivo_idx < len(row) and row[motivo_idx] is not None:
            motivo_val = str(row[motivo_idx]).strip()
        parsed.append((proto, status_raw, sla_val, motivo_val))

    total_rows = len(parsed)
    protocols = [p for p, _, _, _ in parsed if p]

    # Cross com o banco local; o que faltar buscamos no MySQL e fazemos upsert.
    existing_map: dict[str, Fibra] = {
        f.numero_protocolo: f
        for f in Fibra.objects.filter(numero_protocolo__in=protocols)
    }
    missing = [p for p in protocols if p not in existing_map]
    created_from_mysql = 0
    if missing:
        for r in fetch_fibras_by_protocols(missing):
            fibra, was_created = _upsert_fibra_from_mysql_row(r)
            if fibra and was_created:
                created_from_mysql += 1
        # Recarrega após upsert.
        existing_map = {
            f.numero_protocolo: f
            for f in Fibra.objects.filter(numero_protocolo__in=protocols)
        }

    # --- Fallback: para protocolos da planilha que ainda não bateram,
    # compara apenas pelos dígitos (remove letras e "-" do protocolo do banco).
    # Isso evita marcar como "Não localizada" vendas que diferem apenas por
    # prefixo/sufixo alfanumérico no Nº de Protocolo armazenado.
    pending_planilha = [p for p in protocols if p not in existing_map]
    if pending_planilha:
        digits_to_planilha: dict[str, str] = {}
        for p in pending_planilha:
            d = _digits_only(p)
            if d:
                digits_to_planilha.setdefault(d, p)
        if digits_to_planilha:
            # Varre Fibras com protocolo preenchido procurando match por dígitos.
            for f in (
                Fibra.objects.exclude(numero_protocolo='')
                .only('id', 'numero_protocolo')
                .iterator()
            ):
                d = _digits_only(f.numero_protocolo)
                if not d:
                    continue
                planilha_proto = digits_to_planilha.get(d)
                if planilha_proto and planilha_proto not in existing_map:
                    existing_map[planilha_proto] = f

    matched = 0
    updated_status = 0
    not_found: list[str] = []
    now = timezone.now()

    # Mantém um mapa proto -> status_raw para registrar inconsistências.
    unmatched_status: dict[str, str] = {}

    for ordem, (proto, status_raw, sla_val, motivo_val) in enumerate(parsed, start=1):
        if not proto:
            continue
        fibra = existing_map.get(proto)
        if not fibra:
            not_found.append(proto)
            unmatched_status[proto] = status_raw
            continue
        matched += 1
        fibra.ordem_planilha = ordem
        fibra.last_planilha_at = now
        fibra.status_planilha_raw = status_raw[:120]
        fibra.sla_agenda = sla_val[:120]
        fibra.motivo_planilha = motivo_val[:255]
        update_fields = [
            'ordem_planilha', 'last_planilha_at', 'status_planilha_raw',
            'sla_agenda', 'motivo_planilha',
        ]

        mapped = _map_planilha_status(status_raw)
        if mapped and mapped != fibra.status:
            old = fibra.status
            fibra.status = mapped
            update_fields.append('status')
            updated_status += 1
            try:
                FibraStatusHistory.objects.create(
                    fibra=fibra,
                    status_anterior=old,
                    status_novo=mapped,
                    retorno=f'Import planilha: {status_raw[:80]}',
                    alterado_por=by_user,
                )
            except Exception:
                pass
            _notify_vendor(fibra, old, mapped)

        fibra.save(update_fields=update_fields)

    # --- Persistência de inconsistências (ORDEMs da planilha sem match local).
    inc_created = 0
    inc_updated = 0
    for proto, status_raw in unmatched_status.items():
        obj, was_created = PlanilhaOrdemInconsistente.objects.get_or_create(
            ordem=proto,
            defaults={'status_raw': status_raw[:120]},
        )
        if was_created:
            inc_created += 1
        else:
            obj.status_raw = status_raw[:120]
            obj.occurrences = (obj.occurrences or 0) + 1
            obj.save(update_fields=['status_raw', 'occurrences', 'last_seen_at'])
            inc_updated += 1

    # --- Sweep: qualquer inconsistência cuja ordem agora bata com algum
    # Fibra.numero_protocolo é considerada resolvida e removida.
    inc_resolved = _sweep_inconsistencias()

    # --- Limpeza: fibras FIXA que estavam em planilhas anteriores mas não
    # vieram nesta importação podem sair, EXCETO se foram "mexidas" (retorno
    # da ilha preenchido, histórico de status com alteração manual, mensagens
    # de chat) ou estão "em tratativa" (incidente aberto/em andamento).
    pruned = _prune_stale_fibras(import_ts=now)

    return {
        'rows': total_rows,
        'matched': matched,
        'updated_status': updated_status,
        'created_from_mysql': created_from_mysql,
        'not_found': not_found[:20],
        'not_found_count': len(not_found),
        'inconsistencias_novas': inc_created,
        'inconsistencias_atualizadas': inc_updated,
        'inconsistencias_resolvidas': inc_resolved,
        'removidas': pruned,
    }


def _sweep_inconsistencias() -> int:
    """Remove ordens inconsistentes que agora batem com algum protocolo no DB.

    Faz dois passes: (1) match direto pelo ``numero_protocolo`` e
    (2) match por dígitos apenas (remove letras e ``-``) para cobrir casos
    em que o protocolo no banco tem prefixo/sufixo alfanumérico.
    """
    pendentes = list(
        PlanilhaOrdemInconsistente.objects.values_list('ordem', flat=True)
    )
    if not pendentes:
        return 0

    achados: set[str] = set()
    direct = set(
        Fibra.objects.filter(numero_protocolo__in=pendentes)
        .exclude(numero_protocolo='')
        .values_list('numero_protocolo', flat=True)
    )
    achados.update(direct)

    # Passe 2: comparação por dígitos apenas.
    pendentes_restantes = [p for p in pendentes if p not in direct]
    if pendentes_restantes:
        digits_to_ordem: dict[str, str] = {}
        for p in pendentes_restantes:
            d = _digits_only(p)
            if d:
                digits_to_ordem.setdefault(d, p)
        if digits_to_ordem:
            for proto in (
                Fibra.objects.exclude(numero_protocolo='')
                .values_list('numero_protocolo', flat=True)
                .iterator()
            ):
                d = _digits_only(proto)
                if d and d in digits_to_ordem:
                    achados.add(digits_to_ordem[d])

    if not achados:
        return 0
    deleted, _ = PlanilhaOrdemInconsistente.objects.filter(ordem__in=achados).delete()
    return int(deleted)


def _prune_stale_fibras(*, import_ts) -> int:
    """Remove fibras FIXA que sumiram da planilha atual.

    Critérios para apagar:
      - pilar contém 'FIXA' (ou está vazio — fibras criadas só via planilha);
      - já tinha sido vista em uma planilha antes (``last_planilha_at`` não nulo)
        e essa última visita é anterior à importação atual (``import_ts``).

    Critérios de PROTEÇÃO (NÃO apaga):
      - ``retorno_myrella`` preenchido (a ilha já interagiu);
      - existe ``FibraIncidente`` em status ``aberto`` ou ``em_tratativa``;
      - existe ``FibraStatusHistory`` com ``alterado_por`` não nulo
        (alteração manual de status feita por alguém);
      - existe ``FibraChatMessage`` (chat reverso teve mensagem).
    """
    from django.db.models import Q, Exists, OuterRef

    from .models import (
        FibraChatMessage, FibraIncidente, FibraStatusHistory,
    )

    base = Fibra.objects.filter(
        last_planilha_at__isnull=False,
        last_planilha_at__lt=import_ts,
    ).filter(Q(pilar__icontains='FIXA') | Q(pilar=''))

    if not base.exists():
        return 0

    tratativa = FibraIncidente.objects.filter(
        fibra_id=OuterRef('pk'),
        status__in=[FibraIncidente.STATUS_ABERTO, FibraIncidente.STATUS_EM_TRATATIVA],
    )
    historico_manual = FibraStatusHistory.objects.filter(
        fibra_id=OuterRef('pk'),
        alterado_por__isnull=False,
    )
    chat_msg = FibraChatMessage.objects.filter(chat__fibra_id=OuterRef('pk'))

    candidatas = (
        base.annotate(
            _tratativa=Exists(tratativa),
            _hist=Exists(historico_manual),
            _chat=Exists(chat_msg),
        )
        .exclude(_tratativa=True)
        .exclude(_hist=True)
        .exclude(_chat=True)
        .exclude(retorno_myrella__gt='')
    )

    deleted, _detail = candidatas.delete()
    return int(deleted)


# ---------------------------------------------------------------------------
# Mudança de status + notificação ao vendedor
# ---------------------------------------------------------------------------

def change_status(fibra: Fibra, new_status: str, *, changed_by, retorno: str = '') -> None:
    if new_status not in dict(Fibra.STATUS_CHOICES):
        raise ValueError(f"Status inválido: {new_status}")
    old = fibra.status
    fibra.status = new_status
    if retorno:
        fibra.retorno_myrella = retorno
    fibra.save(update_fields=['status', 'retorno_myrella', 'last_synced_at'])

    FibraStatusHistory.objects.create(
        fibra=fibra,
        status_anterior=old,
        status_novo=new_status,
        retorno=retorno,
        alterado_por=changed_by,
    )

    _notify_vendor(fibra, old, new_status)


def _notify_vendor(fibra: Fibra, old_status: str, new_status: str) -> None:
    """Notifica o vendedor (se cadastrado no portal) sobre a mudança de status."""
    if old_status == new_status:
        return
    vendor_user = _find_vendor_user(fibra.vendedor)
    if not vendor_user:
        return

    labels = dict(Fibra.STATUS_CHOICES)
    msg_map = {
        Fibra.STATUS_INSTALADO: 'foi instalada',
        Fibra.STATUS_AGENDADO: 'foi reagendada',
        Fibra.STATUS_CANCELADO: 'foi cancelada',
        Fibra.STATUS_PENDENTE: 'voltou para pendente',
        Fibra.STATUS_PROBLEMA: 'foi sinalizada com problema',
    }
    acao = msg_map.get(new_status, f'mudou para {labels.get(new_status, new_status)}')
    title = f"Fibra {labels.get(new_status, new_status)}"
    message = (
        f"Sua fibra para o cliente {fibra.cliente or '—'}, "
        f"venda {fibra.numero_da_venda}, valor R$ {fibra.valor} {acao}."
    )

    try:
        from core.notifications import NotificationMixin  # type: ignore
        NotificationMixin.create_notification(
            user=vendor_user,
            title=title,
            message=message,
            notification_type='fibra_status',
            related_object_id=fibra.id,
            related_url=f"/fibras/{fibra.id}/",
        )
    except Exception:
        # Fallback: não bloqueia a operação se notificações estiverem indisponíveis.
        pass


def _find_vendor_user(vendedor_nome: str):
    if not vendedor_nome:
        return None
    target = _normalize(vendedor_nome)
    for u in User.objects.filter(is_active=True).only('id', 'first_name', 'last_name', 'username'):
        full = _normalize(f"{u.first_name} {u.last_name}".strip() or u.username)
        if full == target:
            return u
    return None
