from __future__ import annotations

from decimal import Decimal

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db.models import Count, Sum
from django.http import HttpResponse, HttpResponseRedirect, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.utils import timezone
from django.views.decorators.http import require_POST

from .models import Fibra, FibraIncidente, FibraChat, FibraChatMessage
from .services import change_status, fibras_for_user, import_planilha_xlsx, is_gerente, sync_fibras


def _is_ilha(user) -> bool:
    """Quem pode mexer em status / chat reverso / fechar incidentes."""
    return user.is_superuser or getattr(user, 'hierarchy', '') in (
        'ADMIN', 'SUPERADMIN', 'SUPERVISOR', 'ADMINISTRATIVO',
    )


def _can_import_planilha(user) -> bool:
    return _is_ilha(user) or is_gerente(user)


@login_required
def kanban(request):
    from django.db.models import Q as _Q
    from django.db.models.functions import Length
    qs = fibras_for_user(request.user)
    # filtros simples
    pdv = (request.GET.get('pdv') or '').strip()
    vendedor = (request.GET.get('vendedor') or '').strip()
    venda = (request.GET.get('venda') or '').strip()
    ordem = (request.GET.get('ordem') or '').strip()
    # B2B: documento (CPF/CNPJ) com mais de 11 caracteres = empresa.
    b2b = (request.GET.get('b2b') or '').strip() in ('1', 'true', 'on', 'sim')
    if b2b:
        qs = qs.annotate(_doc_len=Length('cpf')).filter(_doc_len__gt=11)
    if pdv:
        qs = qs.filter(pdv__icontains=pdv)
    if vendedor:
        qs = qs.filter(vendedor__icontains=vendedor)
    if venda:
        qs = qs.filter(numero_da_venda__icontains=venda)
    if ordem:
        # Aceita o nº de ORDEM da planilha (== numero_protocolo) ou a posição
        # numérica (ordem_planilha). Tenta os dois para ser flexível.
        cond = _Q(numero_protocolo__icontains=ordem)
        if ordem.isdigit():
            cond |= _Q(ordem_planilha=int(ordem))
        qs = qs.filter(cond)

    # Coluna "Não localizado": vendas de fibra cujo protocolo não aparece na
    # planilha atual (last_planilha_at antes do último import OU nulo). Só
    # considera vendas FIXA não canceladas/instaladas para não poluir a tela.
    # IMPORTANTE: calculada ANTES das colunas por status para que possamos
    # excluir esses IDs das outras colunas e evitar duplicação no kanban.
    from django.db.models import Max, Q
    from .models import PlanilhaOrdemInconsistente
    latest_import = Fibra.objects.aggregate(m=Max('last_planilha_at'))['m']
    nao_qs = qs.filter(
        Q(pilar__iexact='fixa') | Q(pilar__icontains='fixa') | Q(pilar=''),
    ).exclude(status__in=[Fibra.STATUS_CANCELADO, Fibra.STATUS_INSTALADO])
    if latest_import:
        nao_qs = nao_qs.filter(
            Q(last_planilha_at__isnull=True) | Q(last_planilha_at__lt=latest_import)
        )
    else:
        nao_qs = nao_qs.filter(last_planilha_at__isnull=True)
    nao_qs = nao_qs.order_by('-data_da_venda', '-id')
    nao_items = list(nao_qs)
    nao_ids = {f.id for f in nao_items}

    colunas = []
    for key, label in Fibra.STATUS_CHOICES:
        # Ordena pela posição na planilha (quando disponível) e depois pela data.
        # Exclui fibras que estão na coluna "Não localizado" para não duplicar.
        col_qs = (
            qs.filter(status=key)
              .exclude(id__in=nao_ids)
              .order_by('ordem_planilha', '-data_da_venda')
        )
        colunas.append({
            'key': key,
            'label': label,
            'items': list(col_qs),
            'count': col_qs.count(),
            'total': col_qs.aggregate(t=Sum('valor'))['t'] or Decimal('0'),
        })

    colunas.append({
        'key': 'nao_localizado',
        'label': 'Não localizado',
        'items': nao_items,
        'count': len(nao_items),
        'total': nao_qs.aggregate(t=Sum('valor'))['t'] or Decimal('0'),
    })

    inconsistencias = list(
        PlanilhaOrdemInconsistente.objects.order_by('-last_seen_at')[:200]
    )

    return render(request, 'fibras/kanban.html', {
        'colunas': colunas,
        'is_ilha': _is_ilha(request.user),
        'can_import': _can_import_planilha(request.user),
        'filtro_pdv': pdv,
        'filtro_vendedor': vendedor,
        'filtro_venda': venda,
        'filtro_ordem': ordem,
        'filtro_b2b': b2b,
        'status_choices': Fibra.STATUS_CHOICES,
        'inconsistencias': inconsistencias,
        'inconsistencias_count': len(inconsistencias),
    })


@login_required
def lista(request):
    """Visão em lista (tabela) alternativa ao kanban."""
    from django.db.models import Q as _Q
    qs = fibras_for_user(request.user)
    pdv = (request.GET.get('pdv') or '').strip()
    vendedor = (request.GET.get('vendedor') or '').strip()
    status = (request.GET.get('status') or '').strip()
    venda = (request.GET.get('venda') or '').strip()
    ordem = (request.GET.get('ordem') or '').strip()
    if pdv:
        qs = qs.filter(pdv__icontains=pdv)
    if vendedor:
        qs = qs.filter(vendedor__icontains=vendedor)
    if status:
        qs = qs.filter(status=status)
    if venda:
        qs = qs.filter(numero_da_venda__icontains=venda)
    if ordem:
        cond = _Q(numero_protocolo__icontains=ordem)
        if ordem.isdigit():
            cond |= _Q(ordem_planilha=int(ordem))
        qs = qs.filter(cond)

    return render(request, 'fibras/lista.html', {
        'fibras': qs.order_by('ordem_planilha', '-data_da_venda', '-id'),
        'is_ilha': _is_ilha(request.user),
        'can_import': _can_import_planilha(request.user),
        'filtro_pdv': pdv,
        'filtro_vendedor': vendedor,
        'filtro_status': status,
        'filtro_venda': venda,
        'filtro_ordem': ordem,
        'status_choices': Fibra.STATUS_CHOICES,
    })


@login_required
def detail(request, pk: int):
    fibra = get_object_or_404(Fibra, pk=pk)
    incidentes = fibra.incidentes.select_related('aberto_por').all()
    historico = fibra.status_history.select_related('alterado_por')[:50]
    return render(request, 'fibras/detail.html', {
        'fibra': fibra,
        'incidentes': incidentes,
        'historico': historico,
        'is_ilha': _is_ilha(request.user),
        'status_choices': Fibra.STATUS_CHOICES,
    })


@login_required
@require_POST
def change_status_view(request, pk: int):
    if not _is_ilha(request.user):
        messages.error(request, 'Apenas a Ilha pode alterar o status da fibra.')
        return redirect('fibras:detail', pk=pk)
    fibra = get_object_or_404(Fibra, pk=pk)
    new_status = (request.POST.get('status') or '').strip()
    retorno = (request.POST.get('retorno') or '').strip()
    try:
        change_status(fibra, new_status, changed_by=request.user, retorno=retorno)
        messages.success(request, 'Status atualizado.')
    except ValueError as e:
        messages.error(request, str(e))
    return redirect('fibras:detail', pk=pk)


@login_required
@require_POST
def change_status_ajax(request, pk: int):
    """Endpoint usado pelo drag&drop do Kanban — retorna JSON.

    Mantido para compatibilidade, porém o status passa a ser ditado pela
    planilha diária: por isso só a Ilha (Supervisor/Admin) pode mudar manualmente.
    """
    if not _is_ilha(request.user):
        return JsonResponse({'ok': False, 'error': 'permission'}, status=403)
    fibra = get_object_or_404(Fibra, pk=pk)
    new_status = (request.POST.get('status') or '').strip()
    try:
        change_status(fibra, new_status, changed_by=request.user, retorno='')
    except ValueError as e:
        return JsonResponse({'ok': False, 'error': str(e)}, status=400)
    return JsonResponse({
        'ok': True,
        'pk': fibra.pk,
        'status': fibra.status,
        'status_label': fibra.get_status_display(),
    })


@login_required
@require_POST
def importar_planilha(request):
    """Upload manual da planilha diária (.xlsx). Atualiza o Kanban/Lista."""
    if not _can_import_planilha(request.user):
        messages.error(request, 'Sem permissão para importar planilha.')
        return redirect('fibras:kanban')
    file_obj = request.FILES.get('planilha')
    if not file_obj:
        messages.error(request, 'Selecione o arquivo .xlsx da planilha.')
        return redirect('fibras:kanban')
    if not file_obj.name.lower().endswith('.xlsx'):
        messages.error(request, 'Envie um arquivo .xlsx.')
        return redirect('fibras:kanban')
    try:
        stats = import_planilha_xlsx(file_obj, by_user=request.user)
    except Exception as e:
        messages.error(request, f'Falha ao importar planilha: {e}')
        return redirect('fibras:kanban')
    messages.success(
        request,
        f"Planilha importada: {stats['matched']} fibras casadas, "
        f"{stats['updated_status']} com status atualizado, "
        f"{stats['not_found_count']} protocolos não encontrados "
        f"(de {stats['rows']} linhas). "
        f"Inconsistências: +{stats.get('inconsistencias_novas', 0)} novas, "
        f"{stats.get('inconsistencias_resolvidas', 0)} resolvidas. "
        f"Removidas (sumiram da planilha e sem tratativa): {stats.get('removidas', 0)}."
    )
    return redirect('fibras:kanban')


@login_required
@require_POST
def abrir_incidente(request, pk: int):
    fibra = get_object_or_404(Fibra, pk=pk)
    observacao = (request.POST.get('observacao') or '').strip()
    if not observacao:
        messages.error(request, 'Descreva a observação do incidente.')
        return redirect('fibras:detail', pk=pk)

    incidente = FibraIncidente.objects.create(
        fibra=fibra, aberto_por=request.user, observacao=observacao,
    )

    # Cria/vincula um SupportChat na categoria "SUPORTE FIXA" para que a
    # tratativa apareça também em /projects/support/admin/template/ e o chat
    # seja único entre vendedor/gerente e a atendente do suporte.
    try:
        from projects.models_chat import SupportCategory, SupportChat, SupportChatMessage

        sector = getattr(request.user, 'sector', None)
        category = SupportCategory.objects.filter(name__iexact='SUPORTE FIXA').first()
        if not category and sector:
            category = SupportCategory.objects.create(
                name='SUPORTE FIXA',
                sector=sector,
                description='Tratativas de incidentes de Fibra (módulo /fibras/).',
            )
        if category:
            chat = SupportChat.objects.create(
                user=request.user,
                category=category,
                sector=category.sector or sector,
                title=f"Fibra {fibra.numero_da_venda} — {fibra.cliente or 'cliente'}",
                priority='ALTA',
                status='ABERTO',
            )
            SupportChatMessage.objects.create(
                chat=chat, user=request.user, message=observacao,
            )
            incidente.support_chat = chat
            incidente.save(update_fields=['support_chat'])
    except Exception:
        # Não bloqueia o fluxo se o módulo de suporte estiver indisponível.
        pass

    messages.success(request, 'Tratativa aberta para a Ilha (visível também em Suporte → SUPORTE FIXA).')
    return redirect('fibras:detail', pk=pk)


@login_required
def chat_view(request, pk: int):
    """Chat da tratativa. Se houver SupportChat vinculado, usa-o como fonte única."""
    fibra = get_object_or_404(Fibra, pk=pk)
    incidente = fibra.incidentes.filter(support_chat__isnull=False).order_by('-aberto_em').first()

    if incidente and incidente.support_chat_id:
        support_chat = incidente.support_chat
        mensagens_support = support_chat.messages.select_related('user').all()
        # Marca como visualizado para o autor (some o badge de "resposta nova").
        if incidente.aberto_por_id == request.user.id:
            incidente.last_opener_view_at = timezone.now()
            incidente.save(update_fields=['last_opener_view_at'])
        return render(request, 'fibras/chat.html', {
            'fibra': fibra,
            'support_chat': support_chat,
            'mensagens_support': mensagens_support,
            'incidente': incidente,
            'is_ilha': _is_ilha(request.user),
        })

    # Fallback: chat legado (FibraChat).
    chat, _ = FibraChat.objects.get_or_create(fibra=fibra)
    mensagens = chat.mensagens.select_related('autor').all()
    chat.mensagens.filter(lida_em__isnull=True).exclude(autor=request.user).update(
        lida_em=timezone.now()
    )
    return render(request, 'fibras/chat.html', {
        'fibra': fibra, 'chat': chat, 'mensagens': mensagens,
        'is_ilha': _is_ilha(request.user),
    })


@login_required
@require_POST
def chat_post(request, pk: int):
    fibra = get_object_or_404(Fibra, pk=pk)
    texto = (request.POST.get('texto') or '').strip()
    if not texto:
        return redirect('fibras:chat', pk=pk)

    # Se houver incidente com SupportChat, escreve lá (chat unificado).
    incidente = fibra.incidentes.filter(support_chat__isnull=False).order_by('-aberto_em').first()
    if incidente and incidente.support_chat_id:
        try:
            from projects.models_chat import SupportChatMessage
            SupportChatMessage.objects.create(
                chat=incidente.support_chat, user=request.user, message=texto,
            )
            return redirect('fibras:chat', pk=pk)
        except Exception:
            pass

    chat, _ = FibraChat.objects.get_or_create(fibra=fibra)
    FibraChatMessage.objects.create(chat=chat, autor=request.user, texto=texto)
    return redirect('fibras:chat', pk=pk)


@login_required
@require_POST
def sync_now(request):
    if not _is_ilha(request.user):
        messages.error(request, 'Sem permissão.')
        return redirect('fibras:kanban')
    stats = sync_fibras()
    messages.success(
        request,
        f"Sincronização concluída: {stats['created']} novas, {stats['updated']} atualizadas "
        f"(de {stats['total_in_source']} no MySQL).",
    )
    return redirect('fibras:kanban')


@login_required
def relatorio(request):
    qs = fibras_for_user(request.user)
    total = qs.count() or 1  # evita divisão por zero
    receita_total = qs.aggregate(t=Sum('valor'))['t'] or Decimal('0')

    # Bloco "Não localizada": vendas FIXA (não canceladas/instaladas) cujo
    # protocolo não bateu com nenhuma ORDEM da planilha mais recente.
    # Calculado ANTES dos blocos por status para podermos excluir esses IDs
    # das outras categorias e evitar dupla contagem.
    from django.db.models import Max, Q
    latest_import = Fibra.objects.aggregate(m=Max('last_planilha_at'))['m']
    nao_qs = qs.filter(
        Q(pilar__iexact='fixa') | Q(pilar__icontains='fixa') | Q(pilar=''),
    ).exclude(status__in=[Fibra.STATUS_CANCELADO, Fibra.STATUS_INSTALADO])
    if latest_import:
        nao_qs = nao_qs.filter(
            Q(last_planilha_at__isnull=True) | Q(last_planilha_at__lt=latest_import)
        )
    else:
        nao_qs = nao_qs.filter(last_planilha_at__isnull=True)
    qtd_nao_loc = nao_qs.count()
    receita_nao_loc = nao_qs.aggregate(t=Sum('valor'))['t'] or Decimal('0')
    nao_ids = set(nao_qs.values_list('id', flat=True))

    # Detalhe "Não localizado": uma linha por número da venda (numero_da_venda
    # é único no modelo), com PDV e Vendedor para o card do relatório.
    nao_loc_detalhes = list(
        nao_qs.order_by('-data_da_venda', '-id')
        .values('numero_da_venda', 'pdv', 'vendedor')
    )

    blocos = []
    for key, label in Fibra.STATUS_CHOICES:
        sub = qs.filter(status=key).exclude(id__in=nao_ids)
        cnt = sub.count()
        receita = sub.aggregate(t=Sum('valor'))['t'] or Decimal('0')
        blocos.append({
            'key': key,
            'label': label,
            'qtd': cnt,
            'receita': receita,
            'percent': round((cnt / total) * 100, 1),
        })

    blocos.append({
        'key': 'nao_localizado',
        'label': 'Não localizada',
        'qtd': qtd_nao_loc,
        'receita': receita_nao_loc,
        'percent': round((qtd_nao_loc / total) * 100, 1),
    })

    # principal motivo de cancelamento (texto do retorno_myrella mais frequente)
    motivos = (
        qs.filter(status=Fibra.STATUS_CANCELADO)
        .exclude(retorno_myrella='')
        .values('retorno_myrella')
        .annotate(n=Count('id'))
        .order_by('-n')[:5]
    )

    # KPIs principais
    qtd_cancel = qs.filter(status=Fibra.STATUS_CANCELADO).count()
    qtd_install = qs.filter(status=Fibra.STATUS_INSTALADO).count()
    receita_cancel = qs.filter(status=Fibra.STATUS_CANCELADO).aggregate(
        t=Sum('valor'))['t'] or Decimal('0')
    receita_install = qs.filter(status=Fibra.STATUS_INSTALADO).aggregate(
        t=Sum('valor'))['t'] or Decimal('0')
    real_total = qs.count() or 1

    # Top 5 vendedores por receita
    top_vendedores = list(
        qs.exclude(vendedor='')
        .values('vendedor')
        .annotate(qtd=Count('id'), receita=Sum('valor'))
        .order_by('-receita')[:5]
    )

    # Dataset para os gráficos (Chart.js) — serializado em JSON no template.
    chart_labels = [b['label'] for b in blocos]
    chart_qtd = [b['qtd'] for b in blocos]
    chart_receita = [float(b['receita']) for b in blocos]
    chart_colors = ['#3B82F6', '#F59E0B', '#EF4444', '#10B981', '#6B7280', '#A855F7']

    return render(request, 'fibras/relatorio.html', {
        'blocos': blocos,
        'total': qs.count(),
        'receita_total': receita_total,
        'motivos': motivos,
        'qtd_cancel': qtd_cancel,
        'qtd_install': qtd_install,
        'receita_cancel': receita_cancel,
        'receita_install': receita_install,
        'percent_cancel': round((qtd_cancel / real_total) * 100, 1),
        'percent_install': round((qtd_install / real_total) * 100, 1),
        'qtd_nao_loc': qtd_nao_loc,
        'receita_nao_loc': receita_nao_loc,
        'percent_nao_loc': round((qtd_nao_loc / real_total) * 100, 1),
        'nao_loc_detalhes': nao_loc_detalhes,
        'top_vendedores': top_vendedores,
        'chart_labels': chart_labels,
        'chart_qtd': chart_qtd,
        'chart_receita': chart_receita,
        'chart_colors': chart_colors,
    })


@login_required
def export_excel(request):
    """Exporta as fibras do mês corrente para um arquivo .xlsx.

    Aceita ``?status=<status>`` para exportar apenas um status específico
    (ex.: ``cancelado`` para exportar todas as fibras canceladas).
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    qs = fibras_for_user(request.user).order_by('-data_da_venda', '-id')

    status = (request.GET.get('status') or '').strip()
    status_valido = dict(Fibra.STATUS_CHOICES)
    nome_arquivo = 'fibras'
    if status in status_valido:
        qs = qs.filter(status=status)
        nome_arquivo = f'fibras_{status}'

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Fibras'

    headers = [
        'Nº Venda', 'CPF', 'Cliente', 'Endereço', 'Nº Acesso', 'Plano',
        'Valor', 'PDV', 'Vendedor', 'Data da Venda', 'Pilar',
        'Serviço Técnico', 'Status', 'Retorno Myrella',
        'Primeira sincronização', 'Última sincronização',
    ]
    ws.append(headers)

    header_fill = PatternFill('solid', fgColor='FF6B35')
    header_font = Font(bold=True, color='FFFFFF', name='Calibri', size=11)
    for col_idx, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    for f in qs:
        ws.append([
            f.numero_da_venda, f.cpf, f.cliente, f.endereco, f.numero_acesso,
            f.plano, float(f.valor or 0), f.pdv, f.vendedor,
            f.data_da_venda.strftime('%d/%m/%Y') if f.data_da_venda else '',
            f.pilar, f.servico_tecnico, f.get_status_display(),
            f.retorno_myrella,
            timezone.localtime(f.first_seen_at).strftime('%d/%m/%Y %H:%M') if f.first_seen_at else '',
            timezone.localtime(f.last_synced_at).strftime('%d/%m/%Y %H:%M') if f.last_synced_at else '',
        ])

    widths = [16, 16, 28, 36, 16, 28, 12, 22, 24, 14, 14, 18, 16, 40, 20, 20]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = 'A2'

    filename = f'{nome_arquivo}_{timezone.now().strftime("%Y%m%d_%H%M")}.xlsx'
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


@login_required
def export_canceladas_csv(request):
    """Exporta as fibras *canceladas* no formato da planilha de Gestão
    Documental (a mesma usada em /validad1), com o status atualizado e a
    Receita zerada (cancelamento não gera receita)."""
    import csv
    from validad1.views import GESTAO_DOCUMENTAL_HEADER

    qs = (
        fibras_for_user(request.user)
        .filter(status=Fibra.STATUS_CANCELADO)
        .order_by('-data_da_venda', '-id')
    )

    response = HttpResponse(content_type='text/csv; charset=utf-8')
    filename = f'fibras_canceladas_{timezone.now().strftime("%Y%m%d_%H%M")}.csv'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    response.write('\ufeff')  # BOM para abrir corretamente no Excel

    writer = csv.writer(response, delimiter=';')
    writer.writerow(GESTAO_DOCUMENTAL_HEADER)

    for f in qs:
        valor_plano = f'{float(f.valor or 0):.2f}'
        situacao = f.get_status_display()                       # "Cancelado"
        motivo_cancelamento = (f.motivo_planilha or '').replace('\n', ' ').strip()
        observacao = (f.retorno_myrella or '').replace('\n', ' ').strip()
        data_venda = f.data_da_venda.strftime('%d/%m/%Y') if f.data_da_venda else ''
        row = [
            f.pdv,                  # Filial
            data_venda,             # Data Venda
            '',                     # Data Ativação
            '',                     # Status BKO
            f.servico_tecnico,      # Serviço
            f.numero_da_venda,      # ID Venda
            '',                     # Modelo
            '',                     # IMEI
            f.plano,                # Plano
            valor_plano,            # Valor plano
            '',                     # Plano Antigo
            '',                     # Valor plano antigo
            '',                     # Vencimento Fatura
            f.vendedor,             # Vendedor
            f.cliente,              # Cliente
            f.cpf,                  # CPF/CNPJ
            f.numero_acesso,        # Nº Acesso
            '',                     # Nº Portado
            '',                     # Servico Instalado
            '',                     # Data Agendada de Instalação
            '',                     # Armário
            '',                     # Data Instalação
            situacao,               # Situação do serviço (status atualizado)
            '0.00',                 # Receita (zerada — cancelamento)
            '',                     # Status GED
            '',                     # Sim Card
            '',                     # COD
            '',                     # Fidelização
            '',                     # Zerar Remuneração
            '',                     # Gerar Price
            '',                     # Nº Solicitação NEXT / RPON
            f.numero_protocolo,     # Nº Protocolo GED
            '',                     # Nº Protocolo GED Portabilidade
            '',                     # Motivo Reprovação
            '',                     # Data Digitalização
            '',                     # Usuário Inserção
            '',                     # Usuário Alteração
            observacao,             # Observação
            motivo_cancelamento,    # Motivo Cancelamento
            '',                     # Comissão de Serviço do Vendedor
            '',                     # Comissão de Aparelho do Vendedor
            '',                     # Data Última Ativação
            '',                     # Data de alteração
            '',                     # Usuário Cancelamento
            '',                     # Data Cancelamento
            '',                     # Líder de Equipe
            '',                     # Status Primeira Fatura
            '',                     # Data Primeira Fatura
            '',                     # Status Segunda Fatura
            '',                     # Data Segunda Fatura
            '',                     # Status Terceira Fatura
            '',                     # Data Terceira Fatura
            '0.00',                 # Valor Receita (zerada — cancelamento)
            '',                     # Valor Considerado TFP
            '',                     # Valor Penalizado TFP
        ]
        writer.writerow(row)

    return response