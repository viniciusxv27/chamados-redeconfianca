import os
from io import BytesIO

from django.conf import settings
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.utils import timezone
from django.db.models import Q
from django.core.files.base import ContentFile

from .models import FolhaPonto, FolhaPontoManagerPermission
from .pdf_parser import (
    extract_all_folhas, extract_pages_pdf, normalize_name, clean_cpf,
)
from .periodicity import MENSAL, SEMANAL, annotate_periodicity, is_semanal
from users.models import User

HIERARCHY_RANK = {
    'PADRAO': 0,
    'ADMINISTRATIVO': 1,
    'SUPERVISOR': 2,
    'ADMIN': 3,
    'SUPERADMIN': 4,
}


def is_superadmin(user):
    return HIERARCHY_RANK.get(user.hierarchy, 0) >= HIERARCHY_RANK['SUPERADMIN']


def can_manage_folhaponto(user):
    """Superadmins sempre podem; os demais precisam estar liberados em
    FolhaPontoManagerPermission (gerenciada em /folha-ponto/admin/acessos/)."""
    if not user.is_authenticated:
        return False
    if is_superadmin(user):
        return True
    return FolhaPontoManagerPermission.objects.filter(user=user).exists()


def _read_pdf_bytes(field_file):
    """Lê os bytes de um FileField, com fallback para storage legado no S3."""
    if not field_file or not field_file.name:
        return None
    try:
        field_file.open('rb')
        data = field_file.read()
        field_file.close()
        if data:
            return data
    except Exception:
        pass

    storage_location = getattr(field_file.storage, 'location', '')
    if not getattr(settings, 'USE_S3', False) or storage_location != 'media':
        return None
    if field_file.name.startswith('media/'):
        return None
    try:
        from storages.backends.s3boto3 import S3Boto3Storage
        legacy_storage = S3Boto3Storage()
        if legacy_storage.exists(field_file.name):
            with legacy_storage.open(field_file.name, 'rb') as legacy_file:
                data = legacy_file.read()
            if data:
                try:
                    field_file.save(os.path.basename(field_file.name), ContentFile(data), save=True)
                except Exception:
                    pass
                return data
    except Exception:
        pass
    return None


def _client_ip(request):
    xff = request.META.get('HTTP_X_FORWARDED_FOR')
    if xff:
        return xff.split(',')[0].strip()
    return request.META.get('REMOTE_ADDR', '')


def _find_user_by_name(employee_name, users_cache):
    """Encontra um usuário do portal a partir do nome extraído do PDF."""
    if not employee_name:
        return None

    emp_normalized = normalize_name(employee_name)
    emp_words = emp_normalized.split()
    if not emp_words:
        return None

    for user in users_cache:
        user_full = normalize_name(f"{user.first_name} {user.last_name}")
        if user_full == emp_normalized:
            return user

    for user in users_cache:
        user_full = normalize_name(f"{user.first_name} {user.last_name}")
        if user_full and (user_full in emp_normalized or emp_normalized in user_full):
            return user

    for user in users_cache:
        first = normalize_name(user.first_name)
        last = normalize_name(user.last_name)
        if first and last and len(emp_words) >= 2:
            if first == emp_words[0] and last == emp_words[-1]:
                return user

    for user in users_cache:
        first = normalize_name(user.first_name)
        last = normalize_name(user.last_name)
        if first and last:
            last_words = last.split()
            if first in emp_words and all(lw in emp_words for lw in last_words):
                return user

    for user in users_cache:
        user_full = normalize_name(f"{user.first_name} {user.last_name}")
        user_words = user_full.split()
        if len(user_words) >= 2 and all(uw in emp_words for uw in user_words):
            return user

    return None


def _match_user(cpf, employee_name, users_cache):
    """Tenta casar por CPF (mais confiável) e depois por nome."""
    if cpf:
        cpf_clean = clean_cpf(cpf)
        if cpf_clean:
            user = User.objects.filter(cpf__icontains=cpf_clean).first()
            if user:
                return user
    return _find_user_by_name(employee_name, users_cache)


# ─── Área Pessoal ────────────────────────────────────────────────────────────

@login_required
def my_folhas(request):
    folhas = FolhaPonto.objects.filter(user=request.user)

    year_filter = request.GET.get('year')
    if year_filter:
        folhas = folhas.filter(year=year_filter)

    years = (
        FolhaPonto.objects.filter(user=request.user)
        .values_list('year', flat=True).distinct().order_by('-year')
    )

    folhas = annotate_periodicity(folhas)

    return render(request, 'folhaponto/my_folhas.html', {
        'folhas': folhas,
        'years': years,
        'selected_year': year_filter,
        'pending_signature': sum(1 for f in folhas if f.can_sign),
    })


@login_required
def folha_detail(request, pk):
    folha = get_object_or_404(FolhaPonto, pk=pk)
    if folha.user != request.user and not can_manage_folhaponto(request.user):
        messages.error(request, 'Sem permissão para visualizar esta folha de ponto.')
        return redirect('folhaponto:my_folhas')

    annotate_periodicity([folha])

    return render(request, 'folhaponto/folha_detail.html', {'folha': folha})


@login_required
def folha_pdf(request, pk):
    """Serve o PDF recortado (página do colaborador)."""
    folha = get_object_or_404(FolhaPonto, pk=pk)
    if folha.user != request.user and not can_manage_folhaponto(request.user):
        messages.error(request, 'Sem permissão.')
        return redirect('folhaponto:my_folhas')

    if not folha.pdf_file or not folha.pdf_file.name:
        messages.error(request, 'PDF não disponível para esta folha de ponto.')
        return redirect('folhaponto:folha_detail', pk=folha.pk)

    pdf_content = _read_pdf_bytes(folha.pdf_file)
    if not pdf_content:
        try:
            return redirect(folha.pdf_file.url)
        except Exception:
            messages.error(request, 'Não foi possível acessar o arquivo PDF.')
            return redirect('folhaponto:folha_detail', pk=folha.pk)

    response = HttpResponse(pdf_content, content_type='application/pdf')
    response['Content-Disposition'] = (
        f'inline; filename="folha_ponto_{folha.user.pk}_{folha.year}_{folha.month:02d}.pdf"'
    )
    return response


@login_required
def folha_signed_pdf(request, pk):
    """Serve o PDF do documento + folha de Certificado de Assinatura Digital."""
    folha = get_object_or_404(FolhaPonto, pk=pk)
    if folha.user != request.user and not can_manage_folhaponto(request.user):
        messages.error(request, 'Sem permissão.')
        return redirect('folhaponto:my_folhas')

    if not folha.is_signed:
        messages.error(request, 'Esta folha de ponto ainda não foi assinada.')
        return redirect('folhaponto:folha_detail', pk=folha.pk)

    pdf_content = _read_pdf_bytes(folha.pdf_file) if folha.pdf_file else None

    from core.signature_cert import build_signed_pdf

    extra = []
    if folha.total_trabalhadas:
        extra.append(f'Horas trabalhadas: {folha.total_trabalhadas}  |  '
                     f'Saldo do período: {folha.total_saldo or "—"}')
    if folha.dias_faltosos:
        extra.append(f'Dias faltosos: {folha.dias_faltosos}')

    signed_str = timezone.localtime(folha.signed_at).strftime('%d/%m/%Y às %H:%M') \
        if folha.signed_at else ''

    signed_bytes = build_signed_pdf(
        pdf_content,
        doc_title=f'Folha de Ponto – {folha.period_display}',
        person_name=folha.user.full_name,
        cpf=folha.user.cpf or folha.cpf,
        signed_at_str=signed_str,
        ip=folha.signature_ip,
        record_id=folha.pk,
        signature_data_url=folha.signature_image,
        hash_value=folha.signature_hash,
        extra_lines=extra,
    )

    if not signed_bytes:
        messages.error(request, 'Não foi possível gerar o documento assinado.')
        return redirect('folhaponto:folha_detail', pk=folha.pk)

    response = HttpResponse(signed_bytes, content_type='application/pdf')
    response['Content-Disposition'] = (
        f'attachment; filename="folha_ponto_assinada_{folha.user.pk}_{folha.year}_{folha.month:02d}.pdf"'
    )
    return response


@login_required
def api_sign_folha(request, pk):
    if request.method != 'POST':
        return JsonResponse({'error': 'Método não permitido'}, status=405)

    folha = get_object_or_404(FolhaPonto, pk=pk)
    if folha.user != request.user:
        return JsonResponse({'error': 'Sem permissão para assinar esta folha de ponto.'}, status=403)
    if folha.is_signed:
        return JsonResponse({'error': 'Esta folha de ponto já foi assinada.'}, status=409)
    # Folha semanal é prévia do período em aberto (reimportada toda semana):
    # só o fechamento mensal é assinado. Bloqueio no servidor, porque esconder
    # o botão no template não impede um POST direto.
    if is_semanal(folha):
        return JsonResponse({
            'error': 'Esta folha é semanal (período em aberto) e não deve ser assinada. '
                     'A assinatura fica disponível quando o período for fechado.',
        }, status=409)

    import json
    import hashlib

    try:
        body = json.loads(request.body)
    except (json.JSONDecodeError, ValueError):
        return JsonResponse({'error': 'Dados inválidos.'}, status=400)

    signature_data = body.get('signature', '')
    if not signature_data:
        return JsonResponse({'error': 'Assinatura não fornecida.'}, status=400)
    if not signature_data.startswith('data:image/png;base64,'):
        return JsonResponse({'error': 'Formato de assinatura inválido.'}, status=400)

    hash_content = (
        f"{folha.pk}|{folha.user.pk}|{folha.month}|{folha.year}|"
        f"{folha.total_trabalhadas}|{folha.total_saldo}|{signature_data[:100]}"
    )
    signature_hash = hashlib.sha256(hash_content.encode('utf-8')).hexdigest()

    folha.signed_at = timezone.now()
    folha.signature_image = signature_data
    folha.signature_ip = _client_ip(request)
    folha.signature_user_agent = request.META.get('HTTP_USER_AGENT', '')[:500]
    folha.signature_hash = signature_hash
    folha.save()

    return JsonResponse({
        'success': True,
        'message': 'Folha de ponto assinada com sucesso.',
        'signed_at': timezone.localtime(folha.signed_at).strftime('%d/%m/%Y às %H:%M'),
        'hash': signature_hash[:16] + '...',
    })


# ─── Área Administrativa (SUPERADMIN) ────────────────────────────────────────

@login_required
def admin_folhas(request):
    if not can_manage_folhaponto(request.user):
        messages.error(request, 'Acesso restrito.')
        return redirect('folhaponto:my_folhas')

    folhas = FolhaPonto.objects.select_related('user', 'uploaded_by').all()

    search = request.GET.get('q', '').strip()
    year_filter = request.GET.get('year')
    month_filter = request.GET.get('month')

    if search:
        folhas = folhas.filter(
            Q(user__first_name__icontains=search) |
            Q(user__last_name__icontains=search) |
            Q(employee_name__icontains=search) |
            Q(cpf__icontains=search)
        )
    if year_filter:
        folhas = folhas.filter(year=year_filter)
    if month_filter:
        folhas = folhas.filter(month=month_filter)

    years = FolhaPonto.objects.values_list('year', flat=True).distinct().order_by('-year')

    # Periodicidade é derivada da competência mais recente de cada colaborador,
    # então precisa ser calculada antes de aplicar o filtro de periodicidade.
    folhas = annotate_periodicity(folhas)

    periodicity_filter = request.GET.get('periodicity')
    if periodicity_filter in (SEMANAL, MENSAL):
        folhas = [f for f in folhas if f.periodicity == periodicity_filter]

    return render(request, 'folhaponto/admin_folhas.html', {
        'folhas': folhas,
        'years': years,
        'search': search,
        'selected_year': year_filter,
        'selected_month': month_filter,
        'selected_periodicity': periodicity_filter,
        'month_choices': FolhaPonto.MONTH_CHOICES,
        'semanal_count': sum(1 for f in folhas if f.is_semanal),
        'mensal_count': sum(1 for f in folhas if not f.is_semanal),
    })


@login_required
def admin_import(request):
    if not can_manage_folhaponto(request.user):
        messages.error(request, 'Acesso restrito.')
        return redirect('folhaponto:my_folhas')

    users = User.objects.filter(is_active=True).order_by('first_name', 'last_name')
    current_year = timezone.now().year

    return render(request, 'folhaponto/admin_import.html', {
        'users': users,
        'month_choices': FolhaPonto.MONTH_CHOICES,
        'current_year': current_year,
        'year_range': range(current_year - 2, current_year + 2),
    })


@login_required
def admin_access(request):
    """Gerencia quem pode administrar a Folha de Ponto além dos superadmins."""
    if not is_superadmin(request.user):
        messages.error(request, 'Apenas superadministradores gerenciam os acessos da Folha de Ponto.')
        return redirect('folhaponto:my_folhas')

    if request.method == 'POST':
        action = request.POST.get('action')
        target = User.objects.filter(pk=request.POST.get('user_id')).first()
        if not target:
            messages.error(request, 'Usuário não encontrado.')
            return redirect('folhaponto:admin_access')

        if action == 'grant':
            if is_superadmin(target):
                messages.info(request, f'{target.full_name} já é superadmin e tem acesso.')
            else:
                FolhaPontoManagerPermission.objects.get_or_create(
                    user=target, defaults={'granted_by': request.user},
                )
                messages.success(request, f'Acesso liberado para {target.full_name}.')
        elif action == 'revoke':
            FolhaPontoManagerPermission.objects.filter(user=target).delete()
            messages.success(request, f'Acesso removido de {target.full_name}.')
        return redirect('folhaponto:admin_access')

    permissions = (
        FolhaPontoManagerPermission.objects
        .select_related('user', 'user__sector', 'granted_by')
        .order_by('user__first_name', 'user__last_name')
    )
    available_users = (
        User.objects.filter(is_active=True)
        .exclude(pk__in=permissions.values_list('user_id', flat=True))
        .exclude(hierarchy='SUPERADMIN')
        .order_by('first_name', 'last_name')
    )

    return render(request, 'folhaponto/admin_access.html', {
        'permissions': permissions,
        'available_users': available_users,
    })


def _model_fields(record):
    """Filtra apenas as chaves que são campos do modelo (descarta '_key', '_pages', 'error', month/year)."""
    ignore = {'_key', '_pages', 'error', 'month', 'year'}
    return {k: v for k, v in record.items() if k not in ignore and not k.startswith('_')}


def _wants_overwrite(request):
    return str(request.POST.get('overwrite', '')).lower() in ('1', 'true', 'on', 'yes')


def _upsert_folha(target_user, month, year, pdf_content, page_number, uploaded_by, fields, overwrite):
    """Cria a folha do mês; se já existir e `overwrite`, atualiza no lugar.

    Retorna (folha, status), com status em {'created', 'updated', 'exists'}.
    Ao sobrescrever, a assinatura anterior é invalidada — o documento mudou — e
    o PDF antigo é removido do storage.
    """
    existing = FolhaPonto.objects.filter(user=target_user, month=month, year=year).first()

    if existing and not overwrite:
        return existing, 'exists'

    if existing:
        old_file_name = existing.pdf_file.name if existing.pdf_file else None
        existing.pdf_file = pdf_content
        existing.pdf_page_number = page_number
        existing.uploaded_by = uploaded_by
        for field, value in fields.items():
            setattr(existing, field, value)
        existing.signed_at = None
        existing.signature_image = ''
        existing.signature_ip = None
        existing.signature_user_agent = ''
        existing.signature_hash = ''
        existing.save()
        if old_file_name and old_file_name != existing.pdf_file.name:
            try:
                existing.pdf_file.storage.delete(old_file_name)
            except Exception:
                pass
        return existing, 'updated'

    folha = FolhaPonto.objects.create(
        user=target_user, month=month, year=year,
        pdf_file=pdf_content, pdf_page_number=page_number,
        uploaded_by=uploaded_by, **fields,
    )
    return folha, 'created'


@login_required
def api_import_folha(request):
    """Importação individual: funcionário escolhido manualmente."""
    if request.method != 'POST':
        return JsonResponse({'error': 'Método não permitido'}, status=405)
    if not can_manage_folhaponto(request.user):
        return JsonResponse({'error': 'Acesso restrito'}, status=403)

    user_id = request.POST.get('user_id')
    month = request.POST.get('month')
    year = request.POST.get('year')
    pdf_file = request.FILES.get('pdf_file')

    if not all([user_id, month, year, pdf_file]):
        return JsonResponse({'error': 'Preencha todos os campos obrigatórios.'}, status=400)

    try:
        target_user = User.objects.get(pk=user_id)
    except User.DoesNotExist:
        return JsonResponse({'error': 'Usuário não encontrado.'}, status=404)

    month = int(month)
    year = int(year)
    overwrite = _wants_overwrite(request)

    if not overwrite and FolhaPonto.objects.filter(user=target_user, month=month, year=year).exists():
        return JsonResponse({
            'error': f'Já existe folha de ponto para {target_user.full_name} em '
                     f'{dict(FolhaPonto.MONTH_CHOICES)[month]}/{year}. '
                     f'Marque "Sobrescrever" para substituí-la.'
        }, status=409)

    pdf_bytes = pdf_file.read()
    all_folhas = extract_all_folhas(BytesIO(pdf_bytes))

    record = {}
    pages = None
    if all_folhas and not ('error' in all_folhas[0] and not all_folhas[0].get('employee_name')):
        record = all_folhas[0]
        pages = record.get('_pages')

    # Importação individual: ignora o nome do PDF (funcionário já escolhido)
    fields = _model_fields(record)
    fields.pop('employee_name', None)

    # Recorta apenas as páginas do colaborador (se for PDF completo)
    pdf_content_to_save = pdf_bytes
    page_number = pages[0] if pages else 0
    if pages:
        cut = extract_pages_pdf(pdf_bytes, pages)
        if cut:
            pdf_content_to_save = cut
            page_number = pages[0]

    safe_name = f"folha_ponto_{target_user.pk}_{year}_{month:02d}.pdf"
    content_file = ContentFile(pdf_content_to_save, name=safe_name)

    folha, status = _upsert_folha(
        target_user, month, year, content_file, page_number,
        request.user, fields, overwrite=overwrite,
    )

    verb = 'atualizada' if status == 'updated' else 'importada'
    return JsonResponse({
        'success': True,
        'status': status,
        'message': f'Folha de ponto {verb} para {target_user.full_name} – {folha.period_display}',
        'folha_id': folha.pk,
    })


@login_required
def api_bulk_import(request):
    """Importa vários PDFs (cada um pode conter 1 ou vários colaboradores)."""
    if request.method != 'POST':
        return JsonResponse({'error': 'Método não permitido'}, status=405)
    if not can_manage_folhaponto(request.user):
        return JsonResponse({'error': 'Acesso restrito'}, status=403)

    month = request.POST.get('month')
    year = request.POST.get('year')
    files = request.FILES.getlist('pdf_files')

    if not month or not year or not files:
        return JsonResponse({'error': 'Informe mês, ano e selecione os arquivos.'}, status=400)

    month = int(month)
    year = int(year)
    overwrite = _wants_overwrite(request)
    results = []
    unmatched_names = []
    users_cache = list(User.objects.filter(is_active=True))

    for pdf_file in files:
        pdf_bytes = pdf_file.read()
        all_folhas = extract_all_folhas(BytesIO(pdf_bytes))

        if len(all_folhas) == 1 and 'error' in all_folhas[0] and not all_folhas[0].get('employee_name'):
            results.append({'file': pdf_file.name, 'status': 'error',
                            'message': f'Erro ao processar: {all_folhas[0]["error"]}'})
            continue
        if not all_folhas:
            results.append({'file': pdf_file.name, 'status': 'skip',
                            'message': 'Nenhuma folha de ponto encontrada no PDF.'})
            continue

        for record in all_folhas:
            if 'error' in record and not record.get('employee_name'):
                continue
            employee_name = record.get('employee_name', '')
            cpf = record.get('cpf', '')
            pages = record.get('_pages') or []

            target_user = _match_user(cpf, employee_name, users_cache)
            if not target_user:
                results.append({'file': pdf_file.name, 'status': 'skip',
                                'message': f'Funcionário não encontrado: {employee_name or "?"}'})
                unmatched_names.append({'nome_pdf': employee_name or 'Nome não identificado', 'cpf': cpf})
                continue

            fields = _model_fields(record)
            safe_name = f"folha_ponto_{target_user.pk}_{year}_{month:02d}.pdf"

            if len(all_folhas) == 1 and len(pages) <= 1:
                pdf_content = ContentFile(pdf_bytes, name=safe_name)
                saved_page = pages[0] if pages else 0
            else:
                cut = extract_pages_pdf(pdf_bytes, pages) if pages else None
                pdf_content = ContentFile(cut or pdf_bytes, name=safe_name)
                saved_page = pages[0] if pages else 0

            _, status = _upsert_folha(
                target_user, month, year, pdf_content, saved_page,
                request.user, fields, overwrite=overwrite,
            )

            if status == 'exists':
                results.append({'file': pdf_file.name, 'status': 'skip',
                                'message': f'Já existe para {target_user.full_name} '
                                           f'(marque "Sobrescrever" para substituir)'})
            elif status == 'updated':
                results.append({'file': pdf_file.name, 'status': 'success',
                                'message': f'Atualizado para {target_user.full_name}'})
            else:
                results.append({'file': pdf_file.name, 'status': 'success',
                                'message': f'Importado para {target_user.full_name}'})

    success_count = sum(1 for r in results if r['status'] == 'success')
    return JsonResponse({
        'results': results,
        'total': len(results),
        'success_count': success_count,
        'unmatched': unmatched_names,
    })


@login_required
def api_process_full_pdf(request):
    """Processa o PDF completo do mês: recorta por colaborador e cria/atualiza."""
    if request.method != 'POST':
        return JsonResponse({'error': 'Método não permitido'}, status=405)
    if not can_manage_folhaponto(request.user):
        return JsonResponse({'error': 'Acesso restrito'}, status=403)

    month = request.POST.get('month')
    year = request.POST.get('year')
    pdf_file = request.FILES.get('pdf_file')

    if not month or not year or not pdf_file:
        return JsonResponse({'error': 'Informe mês, ano e o PDF completo.'}, status=400)

    try:
        month = int(month)
        year = int(year)
    except (TypeError, ValueError):
        return JsonResponse({'error': 'Mês/ano inválidos.'}, status=400)

    if not (pdf_file.name or '').lower().endswith('.pdf'):
        return JsonResponse({'error': 'Arquivo inválido. Envie um PDF (.pdf).'}, status=400)

    pdf_bytes = pdf_file.read()
    all_folhas = extract_all_folhas(BytesIO(pdf_bytes))

    if len(all_folhas) == 1 and 'error' in all_folhas[0] and not all_folhas[0].get('employee_name'):
        return JsonResponse({'error': all_folhas[0]['error']}, status=400)
    if not all_folhas:
        return JsonResponse({'error': 'Nenhuma folha de ponto encontrada no PDF enviado.'}, status=400)

    users_cache = list(User.objects.filter(is_active=True))
    results = []
    unmatched_names = []
    created_count = 0
    updated_count = 0

    for record in all_folhas:
        if 'error' in record and not record.get('employee_name'):
            continue
        employee_name = record.get('employee_name', '')
        cpf = record.get('cpf', '')
        pages = record.get('_pages') or []

        target_user = _match_user(cpf, employee_name, users_cache)
        if not target_user:
            results.append({'status': 'skip', 'message': f'Funcionário não encontrado: {employee_name or "?"}'})
            unmatched_names.append({'nome_pdf': employee_name or 'Nome não identificado', 'cpf': cpf})
            continue

        if not pages:
            results.append({'status': 'error', 'message': f'Página não identificada para {target_user.full_name}.'})
            continue

        cut = extract_pages_pdf(pdf_bytes, pages)
        if not cut:
            results.append({'status': 'error', 'message': f'Não foi possível recortar a folha de {target_user.full_name}.'})
            continue

        safe_name = f"folha_ponto_{target_user.pk}_{year}_{month:02d}.pdf"
        pdf_content = ContentFile(cut, name=safe_name)
        fields = _model_fields(record)

        # Este fluxo é sempre idempotente: reprocessar o PDF do mês atualiza o
        # que já existe, permitindo reimportar a competência quantas vezes for.
        _, status = _upsert_folha(
            target_user, month, year, pdf_content, pages[0],
            request.user, fields, overwrite=True,
        )
        if status == 'updated':
            updated_count += 1
            results.append({'status': 'updated', 'message': f'Atualizado: {target_user.full_name}'})
        else:
            created_count += 1
            results.append({'status': 'created', 'message': f'Criado: {target_user.full_name}'})

    return JsonResponse({
        'results': results,
        'total': len(results),
        'success_count': created_count + updated_count,
        'created_count': created_count,
        'updated_count': updated_count,
        'unmatched': unmatched_names,
        'message': f'Processamento concluído: {created_count} criado(s), {updated_count} atualizado(s).',
    })


@login_required
def admin_delete_folha(request, pk):
    if request.method != 'POST':
        return JsonResponse({'error': 'Método não permitido'}, status=405)
    if not can_manage_folhaponto(request.user):
        return JsonResponse({'error': 'Acesso restrito'}, status=403)

    folha = get_object_or_404(FolhaPonto, pk=pk)
    info = str(folha)
    try:
        folha.pdf_file.delete(save=False)
    except Exception:
        pass
    folha.delete()
    messages.success(request, f'Folha de ponto excluída: {info}')
    return redirect('folhaponto:admin_folhas')


@login_required
def admin_reupload_folha_pdf(request, pk):
    if request.method != 'POST':
        return JsonResponse({'error': 'Método não permitido'}, status=405)
    if not can_manage_folhaponto(request.user):
        return JsonResponse({'error': 'Acesso restrito'}, status=403)

    folha = get_object_or_404(FolhaPonto, pk=pk)
    new_pdf = request.FILES.get('pdf_file')
    if not new_pdf:
        messages.error(request, 'Selecione um arquivo PDF para reenviar.')
        return redirect('folhaponto:admin_folhas')
    if not (new_pdf.name or '').lower().endswith('.pdf'):
        messages.error(request, 'Arquivo inválido. Envie um PDF (.pdf).')
        return redirect('folhaponto:admin_folhas')

    new_bytes = new_pdf.read()
    safe_name = f"folha_ponto_{folha.user_id}_{folha.year}_{folha.month:02d}.pdf"
    old_file_name = folha.pdf_file.name if folha.pdf_file else None

    folha.pdf_file = ContentFile(new_bytes, name=safe_name)
    folha.pdf_page_number = 0
    folha.uploaded_by = request.user
    folha.signed_at = None
    folha.signature_image = ''
    folha.signature_ip = None
    folha.signature_user_agent = ''
    folha.signature_hash = ''
    folha.save()

    if old_file_name and old_file_name != folha.pdf_file.name:
        try:
            folha.pdf_file.storage.delete(old_file_name)
        except Exception:
            pass

    messages.success(request, f'PDF reenviado para {folha.user.full_name} ({folha.period_display}).')
    return redirect('folhaponto:admin_folhas')


@login_required
def api_bulk_delete(request):
    if request.method != 'POST':
        return JsonResponse({'error': 'Método não permitido'}, status=405)
    if not can_manage_folhaponto(request.user):
        return JsonResponse({'error': 'Acesso restrito'}, status=403)

    month = request.POST.get('month')
    year = request.POST.get('year')
    if not month or not year:
        return JsonResponse({'error': 'Mês e ano são obrigatórios.'}, status=400)

    try:
        month = int(month)
        year = int(year)
    except (ValueError, TypeError):
        return JsonResponse({'error': 'Mês e ano inválidos.'}, status=400)

    folhas = FolhaPonto.objects.filter(month=month, year=year)
    count = folhas.count()
    if count == 0:
        return JsonResponse({'error': f'Nenhuma folha encontrada para {month:02d}/{year}.'}, status=404)

    for f in folhas:
        try:
            f.pdf_file.delete(save=False)
        except Exception:
            pass
    folhas.delete()

    month_name = dict(FolhaPonto.MONTH_CHOICES).get(month, str(month))
    messages.success(request, f'{count} folhas de ponto de {month_name}/{year} excluídas.')
    return JsonResponse({'success': True, 'deleted': count, 'message': f'{count} folhas excluídas.'})


@login_required
def export_signature_report(request):
    """Relatório CSV de assinaturas de folha de ponto de um mês/ano."""
    if not can_manage_folhaponto(request.user):
        messages.error(request, 'Acesso restrito.')
        return redirect('folhaponto:admin_folhas')

    import csv

    month = request.GET.get('month')
    year = request.GET.get('year')
    if not month or not year:
        messages.error(request, 'Selecione mês e ano para exportar.')
        return redirect('folhaponto:admin_folhas')

    try:
        month = int(month)
        year = int(year)
    except (ValueError, TypeError):
        messages.error(request, 'Mês e ano inválidos.')
        return redirect('folhaponto:admin_folhas')

    month_name = dict(FolhaPonto.MONTH_CHOICES).get(month, str(month))
    folhas = FolhaPonto.objects.filter(month=month, year=year).select_related('user', 'user__sector')
    all_users = User.objects.filter(is_active=True).select_related('sector').order_by('first_name', 'last_name')
    users_with = set(folhas.values_list('user_id', flat=True))

    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = f'attachment; filename="assinaturas_folha_ponto_{year}_{month:02d}.csv"'
    response.write('﻿')
    writer = csv.writer(response, delimiter=';')

    # Folha semanal (período em aberto do colaborador) não é assinável, então
    # não pode entrar na conta de "não assinadas" — senão o relatório acusaria
    # pendência de quem não tem o que assinar.
    folhas = annotate_periodicity(folhas.order_by('user__first_name', 'user__last_name'))

    total = len(folhas)
    signed_count = sum(1 for f in folhas if f.is_signed)
    semanal_count = sum(1 for f in folhas if f.is_semanal and not f.is_signed)
    pending_count = sum(1 for f in folhas if f.can_sign)

    writer.writerow([f'Relatório de Assinaturas – Folha de Ponto {month_name}/{year}'])
    writer.writerow([])
    writer.writerow(['Total de folhas', total])
    writer.writerow(['Assinadas', signed_count])
    writer.writerow(['Pendentes de assinatura', pending_count])
    writer.writerow(['Semanais (período em aberto, não assinável)', semanal_count])
    writer.writerow(['Usuários sem folha', all_users.count() - len(users_with)])
    writer.writerow([])
    writer.writerow(['Funcionário', 'CPF', 'Setor', 'Periodicidade', 'Status',
                     'Data Assinatura', 'IP', 'Hash'])

    for f in folhas:
        sector_name = f.user.sector.name if f.user.sector else ''
        if f.is_signed:
            status = 'Assinado'
            signed_date = timezone.localtime(f.signed_at).strftime('%d/%m/%Y %H:%M') if f.signed_at else ''
        elif f.is_semanal:
            status = 'Não se aplica (semanal)'
            signed_date = ''
        else:
            status = 'Não assinado'
            signed_date = ''
        writer.writerow([
            f.user.full_name, f.user.cpf or f.cpf, sector_name, f.periodicity_label,
            status, signed_date,
            f.signature_ip or '', f.signature_hash[:16] if f.signature_hash else '',
        ])

    users_without = all_users.exclude(id__in=users_with)
    if users_without.exists():
        writer.writerow([])
        writer.writerow(['--- Usuários ativos SEM folha de ponto ---'])
        writer.writerow(['Funcionário', 'CPF', 'Setor'])
        for u in users_without:
            writer.writerow([u.full_name, u.cpf or '', u.sector.name if u.sector else ''])

    return response


@login_required
def api_download_unmatched_excel(request):
    if request.method != 'POST':
        return JsonResponse({'error': 'Método não permitido'}, status=405)
    if not can_manage_folhaponto(request.user):
        return JsonResponse({'error': 'Acesso restrito'}, status=403)

    import json
    try:
        body = json.loads(request.body)
        unmatched = body.get('unmatched', [])
    except (json.JSONDecodeError, AttributeError):
        return JsonResponse({'error': 'Dados inválidos'}, status=400)

    if not unmatched:
        return JsonResponse({'error': 'Nenhum nome para exportar'}, status=400)

    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Não Encontrados'

    ws.append(['Funcionários da Folha de Ponto sem cadastro no portal'])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)
    ws['A1'].font = Font(bold=True, size=14)
    ws.append([])

    headers = ['Nome no PDF', 'CPF']
    ws.append(headers)
    for col_idx, _ in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_idx)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='4F46E5', end_color='4F46E5', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')

    for item in unmatched:
        ws.append([item.get('nome_pdf', ''), item.get('cpf', '')])

    ws.column_dimensions['A'].width = 45
    ws.column_dimensions['B'].width = 20

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    response = HttpResponse(
        buf.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename="nao_encontrados_folha_ponto.xlsx"'
    return response
