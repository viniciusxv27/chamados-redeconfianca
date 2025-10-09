from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse, HttpResponseForbidden
from django.views.decorators.http import require_POST, require_http_methods
from django.utils import timezone
from django.db import models
from tickets.models import Ticket, Category
from .models import Notification, AdminChecklistSectorTask, AdminChecklistTemplate, DailyAdminChecklist, AdminChecklistTask, AdminChecklistAssignment
from .middleware import log_action
from users.models import User
import json


@login_required
def marketplace(request):
    """Marketplace de recompensas C$"""
    context = {
        'user_balance': request.user.balance_cs,
        'featured_items': [],  # Placeholder para itens em destaque
        'categories': [],      # Placeholder para categorias
    }
    return render(request, 'marketplace/index.html', context)


@login_required  
def training_module(request):
    """Módulo de treinamentos"""
    context = {
        'courses': [],         # Placeholder para cursos
        'user_progress': {},   # Placeholder para progresso do usuário
    }
    return render(request, 'training/index.html', context)


@login_required
def dashboard(request):
    """Dashboard analítico"""
    from django.db.models import Count, Q
    from django.db import models
    from datetime import datetime, timedelta
    from users.models import User
    from communications.models import Communication
    
    user = request.user
    
    # Filtrar tickets baseado na hierarquia do usuário
    # TODOS os usuários sempre veem seus próprios chamados (created_by=user)
    base_filter = models.Q(created_by=user)
    
    if user.can_view_all_tickets():
        # Admin vê todos os tickets (incluindo fechados)
        user_tickets = Ticket.objects.all()
    elif user.can_view_sector_tickets():
        # Supervisores veem: seus próprios tickets + tickets dos setores + tickets atribuídos
        user_sectors = list(user.sectors.all())
        if user.sector:
            user_sectors.append(user.sector)
        
        user_tickets = Ticket.objects.filter(
            base_filter |  # Sempre inclui próprios tickets
            models.Q(sector__in=user_sectors) |
            models.Q(assigned_to=user) |
            models.Q(additional_assignments__user=user, additional_assignments__is_active=True)
        ).distinct()
    else:
        # Usuários comuns veem: seus próprios tickets + tickets atribuídos
        # Excluindo tickets fechados
        user_tickets = Ticket.objects.filter(
            base_filter |  # Sempre inclui próprios tickets
            models.Q(assigned_to=user) |
            models.Q(additional_assignments__user=user, additional_assignments__is_active=True)
        ).exclude(status='FECHADO').distinct()
    
    # Estatísticas de tickets (baseadas nos tickets filtrados)
    total_tickets = user_tickets.count()
    open_tickets = user_tickets.filter(status='ABERTO').count()
    closed_tickets = user_tickets.filter(status='FECHADO').count()
    pending_tickets = user_tickets.filter(status='EM_ANDAMENTO').count()
    
    # Estatísticas de usuários (mantém global)
    total_users = User.objects.count()
    active_users = User.objects.filter(is_active=True).count()
    
    # Estatísticas de comunicações (mantém global)
    total_communications = Communication.objects.count()
    recent_communications = Communication.objects.filter(
        created_at__gte=datetime.now() - timedelta(days=7)
    ).count()
    
    # Comunicados fixados na dashboard
    pinned_communications = Communication.objects.filter(
        is_pinned=True,
        send_to_all=True
    ).order_by('-created_at')[:3]
    
    # Tickets por categoria (baseados nos tickets filtrados)
    tickets_by_category = user_tickets.values('category__name').annotate(
        count=Count('id')
    ).order_by('-count')[:5]
    
    # Tickets recentes (apenas os últimos 3 dos setores do usuário)
    recent_tickets = user_tickets.order_by('-created_at')[:3]
    
    # Comunicações recentes
    recent_comms = Communication.objects.order_by('-created_at')[:5]
    
    context = {
        'stats': {
            'total_tickets': total_tickets,
            'open_tickets': open_tickets,
            'closed_tickets': closed_tickets,
            'pending_tickets': pending_tickets,
            'total_users': total_users,
            'active_users': active_users,
            'total_communications': total_communications,
            'recent_communications': recent_communications,
        },
        'pinned_communications': pinned_communications,
        'tickets_by_category': tickets_by_category,
        'recent_tickets': recent_tickets,
        'recent_communications': recent_comms,
        'charts_data': {
            'categories': [item['category__name'] for item in tickets_by_category],
            'values': [item['count'] for item in tickets_by_category],
        }
    }
    return render(request, 'dashboard.html', context)


@login_required
def anonymous_report(request):
    """Canal de denúncias anônimas"""
    if request.method == 'POST':
        title = request.POST.get('title')
        description = request.POST.get('description')
        urgency = request.POST.get('urgency', 'NORMAL')
        
        try:
            # Criar ticket anônimo
            anonymous_category, created = Category.objects.get_or_create(
                name='Denúncia Anônima',
                defaults={'description': 'Denúncias enviadas anonimamente'}
            )
            
            ticket = Ticket.objects.create(
                title=f"[ANÔNIMO] {title}",
                description=description,
                category=anonymous_category,
                created_by=request.user,  # Sistema registra quem criou, mas é tratado como anônimo
                priority='ALTA' if urgency == 'URGENT' else 'NORMAL',
                status='ABERTO',
                is_anonymous=True  # Flag para identificar denúncias anônimas
            )
            
            messages.success(request, 'Denúncia enviada com sucesso! Será tratada com total confidencialidade.')
            return redirect('anonymous_report')
            
        except Exception as e:
            messages.error(request, f'Erro ao enviar denúncia: {str(e)}')
    
    context = {
        'urgency_choices': [
            ('NORMAL', 'Normal'),
            ('URGENT', 'Urgente'),
        ]
    }
    return render(request, 'core/anonymous_report.html', context)


@login_required
def manage_reports(request):
    """Visualizar e gerenciar denúncias - Apenas Superadmins"""
    if request.user.hierarchy != 'SUPERADMIN':
        messages.error(request, 'Você não tem permissão para acessar esta área.')
        return redirect('dashboard')
    
    from tickets.models import Ticket
    from django.db.models import Q
    from django.core.paginator import Paginator
    
    # Buscar tickets anônimos (denúncias)
    reports = Ticket.objects.filter(is_anonymous=True).select_related(
        'category', 'created_by', 'assigned_to', 'sector'
    ).order_by('-created_at')
    
    # Filtros
    status_filter = request.GET.get('status')
    priority_filter = request.GET.get('priority')
    search = request.GET.get('search')
    
    if status_filter and status_filter != 'all':
        reports = reports.filter(status=status_filter)
    
    if priority_filter and priority_filter != 'all':
        reports = reports.filter(priority=priority_filter)
    
    if search:
        reports = reports.filter(
            Q(title__icontains=search) | 
            Q(description__icontains=search)
        )
    
    # Paginação
    paginator = Paginator(reports, 10)
    page_number = request.GET.get('page')
    reports = paginator.get_page(page_number)
    
    # Estatísticas
    stats = {
        'total': Ticket.objects.filter(is_anonymous=True).count(),
        'pending': Ticket.objects.filter(is_anonymous=True, status='ABERTO').count(),
        'in_progress': Ticket.objects.filter(is_anonymous=True, status='EM_ANDAMENTO').count(),
        'resolved': Ticket.objects.filter(is_anonymous=True, status='RESOLVIDO').count(),
        'urgent': Ticket.objects.filter(is_anonymous=True, priority='ALTA').count(),
    }
    
    context = {
        'reports': reports,
        'stats': stats,
        'status_choices': Ticket.STATUS_CHOICES,
        'priority_choices': Ticket.PRIORITY_CHOICES,
        'current_status': status_filter,
        'current_priority': priority_filter,
        'search': search,
    }
    return render(request, 'core/manage_reports.html', context)


@login_required  
def report_detail(request, report_id):
    """Ver detalhes de uma denúncia - Apenas Superadmins"""
    if request.user.hierarchy != 'SUPERADMIN':
        messages.error(request, 'Você não tem permissão para acessar esta área.')
        return redirect('dashboard')
    
    from tickets.models import Ticket
    from django.shortcuts import get_object_or_404
    
    report = get_object_or_404(Ticket, id=report_id, is_anonymous=True)
    
    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'update_status':
            new_status = request.POST.get('status')
            old_status = report.status
            report.status = new_status
            report.save()
            
            log_action(
                request.user,
                'REPORT_STATUS_UPDATE', 
                f'Status da denúncia #{report.id} alterado de {old_status} para {new_status}',
                request
            )
            
            messages.success(request, f'Status da denúncia atualizado para {report.get_status_display()}')
            
        elif action == 'assign':
            from users.models import User
            assigned_to_id = request.POST.get('assigned_to')
            if assigned_to_id:
                assigned_user = User.objects.get(id=assigned_to_id)
                report.assigned_to = assigned_user
                report.status = 'EM_ANDAMENTO'
                report.save()
                
                log_action(
                    request.user,
                    'REPORT_ASSIGNMENT',
                    f'Denúncia #{report.id} atribuída para {assigned_user.full_name}',
                    request
                )
                
                messages.success(request, f'Denúncia atribuída para {assigned_user.full_name}')
        
        elif action == 'add_note':
            from tickets.models import TicketComment
            note = request.POST.get('note')
            if note.strip():
                TicketComment.objects.create(
                    ticket=report,
                    author=request.user,
                    comment=f"[NOTA ADMINISTRATIVA] {note}",
                    is_internal=True
                )
                
                log_action(
                    request.user,
                    'REPORT_NOTE_ADDED',
                    f'Nota administrativa adicionada à denúncia #{report.id}',
                    request
                )
                
                messages.success(request, 'Nota administrativa adicionada com sucesso')
        
        return redirect('report_detail', report_id=report.id)
    
    # Buscar usuários que podem ser responsáveis
    from users.models import User
    assignable_users = User.objects.filter(
        hierarchy__in=['SUPERADMIN', 'ADMINISTRATIVO', 'SUPERVISOR']
    ).order_by('first_name')
    
    # Buscar comentários/notas
    comments = report.comments.all().order_by('created_at')
    
    context = {
        'report': report,
        'assignable_users': assignable_users,
        'comments': comments,
        'status_choices': Ticket.STATUS_CHOICES,
    }
    return render(request, 'core/report_detail.html', context)


# ========================
# NOTIFICATIONS API VIEWS
# ========================

@login_required
def notifications_api_view(request):
    """API para listar notificações do usuário"""
    notifications = request.user.notifications.all()[:10]  # Últimas 10
    
    notifications_data = []
    for notification in notifications:
        notifications_data.append({
            'id': notification.id,
            'title': notification.title,
            'message': notification.message,
            'notification_type': notification.notification_type,
            'is_read': notification.is_read,
            'created_at': notification.created_at.isoformat(),
            'related_url': notification.related_url,
        })
    
    unread_count = request.user.notifications.filter(is_read=False).count()
    
    return JsonResponse({
        'notifications': notifications_data,
        'unread_count': unread_count
    })


@login_required
def notifications_count_api_view(request):
    """API para contar notificações não lidas"""
    count = request.user.notifications.filter(is_read=False).count()
    return JsonResponse({'count': count})


@login_required
@require_POST
def notification_mark_read_api_view(request, notification_id):
    """API para marcar notificação como lida"""
    try:
        notification = request.user.notifications.get(id=notification_id)
        notification.mark_as_read()
        return JsonResponse({'success': True})
    except Notification.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Notificação não encontrada'})


@login_required
@require_POST
def notifications_mark_all_read_api_view(request):
    """API para marcar todas as notificações como lidas"""
    request.user.notifications.filter(is_read=False).update(is_read=True)
    return JsonResponse({'success': True})


# ===== CHECKLIST ADMINISTRATIVO VIEWS =====
from django.utils import timezone
from datetime import date, timedelta
from django.core.paginator import Paginator
from django.db.models import Count, Q, Avg
from django.http import HttpResponse
import csv
from io import BytesIO
import openpyxl
from .models import (
    AdminChecklistTemplate, 
    DailyAdminChecklist, 
    AdminChecklistTask, 
    AdminChecklistAssignment
)


@login_required
def admin_checklist_dashboard(request):
    """Dashboard principal do checklist administrativo"""
    user = request.user
    today = date.today()
    
    # Verificar permissões
    if not user.hierarchy in ['SUPERADMIN', 'ADMINISTRATIVO', 'SUPERVISOR']:
        messages.error(request, 'Acesso negado ao checklist administrativo.')
        return redirect('dashboard')
    
    # Obter ou criar checklist do dia
    daily_checklist, created = DailyAdminChecklist.objects.get_or_create(
        date=today,
        defaults={
            'created_by': user
        }
    )
    
    # Se foi criado agora, gerar as tarefas do dia
    if created:
        generate_daily_tasks(daily_checklist)
    
    # Estatísticas gerais
    stats = daily_checklist.get_completion_stats()
    sector_stats = daily_checklist.get_sector_stats()
    
    # Tarefas do usuário (se não for superadmin)
    user_tasks = []
    if user.hierarchy != 'SUPERADMIN':
        user_sectors = list(user.sectors.all())
        if user.sector:
            user_sectors.append(user.sector)
        
        user_tasks = daily_checklist.tasks.filter(
            Q(template__sector__in=user_sectors) |
            Q(assigned_to=user)
        ).select_related('template', 'assigned_to', 'reviewed_by')
    
    # Dados para gráficos (últimos 7 dias)
    chart_data = get_checklist_chart_data()
    
    # Checklists recentes (últimos 7 dias)
    recent_checklists = DailyAdminChecklist.objects.filter(
        date__gte=today - timedelta(days=7)
    ).order_by('-date')[:5]
    
    # Adicionar estatísticas aos checklists recentes
    for checklist in recent_checklists:
        stats_data = checklist.get_completion_stats()
        checklist.total_tasks = stats_data['total']
        checklist.completed_tasks = stats_data['completed'] + stats_data['approved']
        checklist.completion_percentage = stats_data['completion_percentage']
    
    context = {
        'daily_checklist': daily_checklist,
        'stats': stats,
        'sector_stats': sector_stats,
        'user_tasks': user_tasks,
        'chart_data': chart_data,
        'recent_checklists': recent_checklists,
        'user': user,
        'today': today,
    }
    
    return render(request, 'core/admin_checklist_dashboard.html', context)


@login_required
def admin_checklist_tasks(request):
    """Lista de tarefas do checklist administrativo"""
    user = request.user
    today = date.today()
    
    # Filtros
    date_filter = request.GET.get('date', today.strftime('%Y-%m-%d'))
    sector_filter = request.GET.get('sector', '')
    status_filter = request.GET.get('status', '')
    
    try:
        filter_date = timezone.datetime.strptime(date_filter, '%Y-%m-%d').date()
    except ValueError:
        filter_date = today
    
    # Obter checklist da data
    try:
        daily_checklist = DailyAdminChecklist.objects.get(date=filter_date)
    except DailyAdminChecklist.DoesNotExist:
        daily_checklist = None
        tasks = AdminChecklistTask.objects.none()
    
    if daily_checklist:
        tasks = daily_checklist.tasks.select_related(
            'template', 'template__sector', 'assigned_to', 'reviewed_by'
        )
        
        # Aplicar filtros
        if user.hierarchy != 'SUPERADMIN':
            user_sectors = list(user.sectors.all())
            if user.sector:
                user_sectors.append(user.sector)
            tasks = tasks.filter(template__sector__in=user_sectors)
        
        if sector_filter:
            tasks = tasks.filter(template__sector_id=sector_filter)
        
        if status_filter:
            tasks = tasks.filter(status=status_filter)
    
    # Paginação
    paginator = Paginator(tasks, 20)
    page = request.GET.get('page')
    tasks_page = paginator.get_page(page)
    
    # Setores para filtro
    if user.hierarchy == 'SUPERADMIN':
        from users.models import Sector
        sectors = Sector.objects.filter(admin_checklist_templates__isnull=False).distinct()
    else:
        user_sectors = list(user.sectors.all())
        if user.sector:
            user_sectors.append(user.sector)
        sectors = user_sectors
    
    context = {
        'tasks': tasks_page,
        'daily_checklist': daily_checklist,
        'sectors': sectors,
        'date_filter': date_filter,
        'sector_filter': sector_filter,
        'status_filter': status_filter,
        'user': user,
    }
    
    return render(request, 'core/admin_checklist_tasks.html', context)


@login_required
@require_POST
def admin_checklist_task_action(request, task_id):
    """Ações nas tarefas do checklist (completar, aprovar, rejeitar)"""
    user = request.user
    
    try:
        task = AdminChecklistTask.objects.select_related(
            'template', 'template__sector'
        ).get(id=task_id)
    except AdminChecklistTask.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Tarefa não encontrada'})
    
    action = request.POST.get('action')
    notes = request.POST.get('notes', '')
    
    if action == 'complete':
        if task.mark_completed(user, notes):
            return JsonResponse({
                'success': True, 
                'message': 'Tarefa marcada como concluída',
                'status': task.get_status_display()
            })
        else:
            return JsonResponse({'success': False, 'error': 'Sem permissão para executar esta tarefa'})
    
    elif action == 'approve':
        if task.approve(user, notes):
            return JsonResponse({
                'success': True, 
                'message': 'Tarefa aprovada',
                'status': task.get_status_display()
            })
        else:
            return JsonResponse({'success': False, 'error': 'Sem permissão para aprovar esta tarefa'})
    
    elif action == 'reject':
        if task.reject(user, notes):
            return JsonResponse({
                'success': True, 
                'message': 'Tarefa rejeitada. Retornará para pendente.',
                'status': task.get_status_display()
            })
        else:
            return JsonResponse({'success': False, 'error': 'Sem permissão para rejeitar esta tarefa'})
    
    return JsonResponse({'success': False, 'error': 'Ação inválida'})


@login_required
def admin_checklist_templates(request):
    """Gerenciamento de templates de checklist administrativo"""
    user = request.user
    
    # Apenas superadmins podem gerenciar templates
    if user.hierarchy != 'SUPERADMIN':
        messages.error(request, 'Acesso negado. Apenas superadmins podem gerenciar templates.')
        return redirect('admin_checklist:dashboard')
    
    templates = AdminChecklistTemplate.objects.select_related('sector', 'created_by').order_by('sector__name', 'title')
    
    context = {
        'templates': templates,
        'user': user,
    }
    
    return render(request, 'core/admin_checklist_templates.html', context)


@login_required
def admin_checklist_assignments(request):
    """Gerenciamento de atribuições de tarefas"""
    user = request.user
    
    # Apenas superadmins podem gerenciar atribuições
    if user.hierarchy != 'SUPERADMIN':
        messages.error(request, 'Acesso negado.')
        return redirect('admin_checklist:dashboard')
    
    # Filtros
    status_filter = request.GET.get('status', 'all')
    sector_filter = request.GET.get('sector')
    
    # Base query
    assignments = AdminChecklistAssignment.objects.select_related(
        'task_template', 'task_template__sector', 'user', 'assigned_by'
    )
    
    # Aplicar filtros
    if status_filter == 'active':
        assignments = assignments.filter(is_active=True)
    elif status_filter == 'inactive':
        assignments = assignments.filter(is_active=False)
    
    if sector_filter:
        from users.models import Sector
        try:
            sector = Sector.objects.get(id=sector_filter)
            assignments = assignments.filter(task_template__sector=sector)
        except Sector.DoesNotExist:
            pass
    
    assignments = assignments.order_by('-assigned_at')
    
    # Estatísticas
    total_assignments = assignments.count()
    stats = {
        'active': assignments.filter(is_active=True).count(),
        'inactive': assignments.filter(is_active=False).count(),
        'total': total_assignments,
    }
    
    # Paginação
    from django.core.paginator import Paginator
    paginator = Paginator(assignments, 25)
    page_number = request.GET.get('page')
    assignments_page = paginator.get_page(page_number)
    
    # Setores para filtro
    from users.models import Sector
    sectors = Sector.objects.all().order_by('name')
    
    context = {
        'assignments': assignments_page,
        'sectors': sectors,
        'status_filter': status_filter,
        'sector_filter': sector_filter,
        'stats': stats,
        'user': user,
    }
    
    return render(request, 'core/admin_checklist_assignments.html', context)


@login_required
def admin_checklist_create_assignment(request):
    """Criar nova atribuição de template para usuário"""
    user = request.user
    
    # Apenas superadmins podem criar atribuições
    if user.hierarchy != 'SUPERADMIN':
        messages.error(request, 'Acesso negado.')
        return redirect('admin_checklist:dashboard')
    
    if request.method == 'POST':
        template_id = request.POST.get('template_id')
        user_id = request.POST.get('user_id')
        notes = request.POST.get('notes', '')
        
        try:
            template = AdminChecklistTemplate.objects.get(id=template_id)
            assigned_user = User.objects.get(id=user_id)
            
            # Verificar se já existe uma atribuição ativa
            existing = AdminChecklistAssignment.objects.filter(
                task_template=template,
                user=assigned_user,
                is_active=True
            ).first()
            
            if existing:
                messages.warning(request, f'Já existe uma atribuição ativa deste template para {assigned_user.get_full_name()}.')
            else:
                # Criar nova atribuição
                assignment = AdminChecklistAssignment.objects.create(
                    task_template=template,
                    user=assigned_user,
                    assigned_by=user,
                    notes=notes,
                    is_active=True
                )
                
                messages.success(request, f'Atribuição criada com sucesso para {assigned_user.get_full_name()}.')
                
        except (AdminChecklistTemplate.DoesNotExist, User.DoesNotExist):
            messages.error(request, 'Template ou usuário não encontrado.')
        except Exception as e:
            messages.error(request, f'Erro ao criar atribuição: {str(e)}')
    
    # Buscar templates e usuários disponíveis
    templates = AdminChecklistTemplate.objects.select_related('sector', 'created_by').order_by('sector__name', 'title')
    users = User.objects.filter(is_active=True).order_by('first_name', 'last_name', 'email')
    
    context = {
        'templates': templates,
        'users': users,
        'user': user,
    }
    
    return render(request, 'core/admin_checklist_create_assignment.html', context)


@login_required
@require_http_methods(["POST"])
def admin_checklist_toggle_assignment(request, assignment_id):
    """Ativar/desativar uma atribuição"""
    user = request.user
    
    # Apenas superadmins podem modificar atribuições
    if user.hierarchy != 'SUPERADMIN':
        return JsonResponse({'success': False, 'message': 'Acesso negado.'}, status=403)
    
    try:
        assignment = AdminChecklistAssignment.objects.get(id=assignment_id)
        assignment.is_active = not assignment.is_active
        assignment.save()
        
        action = 'ativada' if assignment.is_active else 'desativada'
        return JsonResponse({
            'success': True, 
            'message': f'Atribuição {action} com sucesso.',
            'is_active': assignment.is_active
        })
        
    except AdminChecklistAssignment.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Atribuição não encontrada.'}, status=404)
    except Exception as e:
        return JsonResponse({'success': False, 'message': f'Erro: {str(e)}'}, status=500)


@login_required
@require_http_methods(["POST"])
def admin_checklist_delete_assignment(request, assignment_id):
    """Deletar uma atribuição"""
    user = request.user
    
    # Apenas superadmins podem deletar atribuições
    if user.hierarchy != 'SUPERADMIN':
        return JsonResponse({'success': False, 'message': 'Acesso negado.'}, status=403)
    
    try:
        assignment = AdminChecklistAssignment.objects.get(id=assignment_id)
        user_name = assignment.user.get_full_name()
        template_title = assignment.task_template.title
        
        assignment.delete()
        
        return JsonResponse({
            'success': True, 
            'message': f'Atribuição de "{template_title}" para {user_name} removida com sucesso.'
        })
        
    except AdminChecklistAssignment.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Atribuição não encontrada.'}, status=404)
    except Exception as e:
        return JsonResponse({'success': False, 'message': f'Erro: {str(e)}'}, status=500)


@login_required
def admin_checklist_reports(request):
    """Relatórios e exportação do checklist administrativo"""
    user = request.user
    
    # Filtros para relatório
    date_from = request.GET.get('date_from', (date.today() - timedelta(days=30)).strftime('%Y-%m-%d'))
    date_to = request.GET.get('date_to', date.today().strftime('%Y-%m-%d'))
    sector_filter = request.GET.get('sector', '')
    export_format = request.GET.get('export', '')
    
    try:
        start_date = timezone.datetime.strptime(date_from, '%Y-%m-%d').date()
        end_date = timezone.datetime.strptime(date_to, '%Y-%m-%d').date()
    except ValueError:
        start_date = date.today() - timedelta(days=30)
        end_date = date.today()
    
    # Query base
    tasks_query = AdminChecklistTask.objects.select_related(
        'checklist', 'template', 'template__sector', 'assigned_to', 'reviewed_by'
    ).filter(checklist__date__range=[start_date, end_date])
    
    # Filtros de permissão
    if user.hierarchy != 'SUPERADMIN':
        user_sectors = list(user.sectors.all())
        if user.sector:
            user_sectors.append(user.sector)
        tasks_query = tasks_query.filter(template__sector__in=user_sectors)
    
    if sector_filter:
        tasks_query = tasks_query.filter(template__sector_id=sector_filter)
    
    # Exportar se solicitado
    if export_format in ['csv', 'xlsx']:
        return export_checklist_data(tasks_query, export_format, start_date, end_date)
    
    # Estatísticas para o período
    period_stats = calculate_period_stats(tasks_query, start_date, end_date)
    
    # Setores para filtro
    if user.hierarchy == 'SUPERADMIN':
        from users.models import Sector
        sectors = Sector.objects.filter(admin_checklist_templates__isnull=False).distinct()
    else:
        user_sectors = list(user.sectors.all())
        if user.sector:
            user_sectors.append(user.sector)
        sectors = user_sectors
    
    context = {
        'period_stats': period_stats,
        'sectors': sectors,
        'date_from': date_from,
        'date_to': date_to,
        'sector_filter': sector_filter,
        'user': user,
    }
    
    return render(request, 'core/admin_checklist_reports.html', context)


# ===== FUNÇÕES AUXILIARES =====
def generate_daily_tasks(daily_checklist):
    """Gera as tarefas do dia baseadas nos templates ativos"""
    templates = AdminChecklistTemplate.objects.filter(is_active=True).select_related('sector')
    
    for template in templates:
        # Verificar se já existe uma tarefa para este template
        if not daily_checklist.tasks.filter(template=template).exists():
            # Verificar se há atribuição específica
            assignment = AdminChecklistAssignment.objects.filter(
                task_template=template, 
                is_active=True
            ).first()
            
            AdminChecklistTask.objects.create(
                checklist=daily_checklist,
                template=template,
                assigned_to=assignment.user if assignment else None
            )


def get_checklist_chart_data():
    """Dados para gráficos dos últimos 7 dias"""
    end_date = date.today()
    start_date = end_date - timedelta(days=6)
    
    chart_data = {
        'labels': [],
        'completion_data': [],
        'approval_data': [],
    }
    
    for i in range(7):
        current_date = start_date + timedelta(days=i)
        chart_data['labels'].append(current_date.strftime('%d/%m'))
        
        try:
            checklist = DailyAdminChecklist.objects.get(date=current_date)
            stats = checklist.get_completion_stats()
            chart_data['completion_data'].append(stats['completion_percentage'])
            chart_data['approval_data'].append(stats['approval_percentage'])
        except DailyAdminChecklist.DoesNotExist:
            chart_data['completion_data'].append(0)
            chart_data['approval_data'].append(0)
    
    return chart_data


def calculate_period_stats(tasks_query, start_date, end_date):
    """Calcula estatísticas para um período"""
    total_tasks = tasks_query.count()
    
    if total_tasks == 0:
        return {
            'total_tasks': 0,
            'completion_rate': 0,
            'approval_rate': 0,
            'avg_completion_time': 0,
            'sector_performance': {},
        }
    
    completed_tasks = tasks_query.filter(status__in=['COMPLETED', 'APPROVED']).count()
    approved_tasks = tasks_query.filter(status='APPROVED').count()
    
    # Tempo médio de conclusão (em horas)
    completed_with_times = tasks_query.filter(
        started_at__isnull=False, 
        completed_at__isnull=False
    )
    
    avg_completion_time = 0
    if completed_with_times.exists():
        total_time = sum([
            (task.completed_at - task.started_at).total_seconds() / 3600
            for task in completed_with_times
        ])
        avg_completion_time = round(total_time / completed_with_times.count(), 2)
    
    # Performance por setor
    sector_performance = {}
    sectors = tasks_query.values_list('template__sector__name', flat=True).distinct()
    
    for sector_name in sectors:
        if sector_name:
            sector_tasks = tasks_query.filter(template__sector__name=sector_name)
            sector_total = sector_tasks.count()
            sector_completed = sector_tasks.filter(status__in=['COMPLETED', 'APPROVED']).count()
            sector_approved = sector_tasks.filter(status='APPROVED').count()
            
            sector_performance[sector_name] = {
                'total': sector_total,
                'completed': sector_completed,
                'approved': sector_approved,
                'completion_rate': round((sector_completed / sector_total) * 100) if sector_total > 0 else 0,
                'approval_rate': round((sector_approved / sector_total) * 100) if sector_total > 0 else 0,
            }
    
    return {
        'total_tasks': total_tasks,
        'completed_tasks': completed_tasks,
        'approved_tasks': approved_tasks,
        'completion_rate': round((completed_tasks / total_tasks) * 100),
        'approval_rate': round((approved_tasks / total_tasks) * 100),
        'avg_completion_time': avg_completion_time,
        'sector_performance': sector_performance,
    }


def export_checklist_data(tasks_query, format_type, start_date, end_date):
    """Exporta dados do checklist em CSV ou Excel"""
    
    if format_type == 'csv':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="checklist_admin_{start_date}_{end_date}.csv"'
        
        writer = csv.writer(response)
        writer.writerow([
            'Data', 'Setor', 'Tarefa', 'Status', 'Atribuído Para', 
            'Iniciado Em', 'Concluído Em', 'Revisado Por', 'Notas da Execução'
        ])
        
        for task in tasks_query:
            writer.writerow([
                task.checklist.date,
                task.template.sector.name,
                task.template.title,
                task.get_status_display(),
                task.assigned_to.get_full_name() if task.assigned_to else '',
                task.started_at.strftime('%d/%m/%Y %H:%M') if task.started_at else '',
                task.completed_at.strftime('%d/%m/%Y %H:%M') if task.completed_at else '',
                task.reviewed_by.get_full_name() if task.reviewed_by else '',
                task.execution_notes,
            ])
        
        return response
    
    elif format_type == 'xlsx':
        output = BytesIO()
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = 'Checklist Administrativo'
        
        # Headers
        headers = [
            'Data', 'Setor', 'Tarefa', 'Status', 'Atribuído Para', 
            'Iniciado Em', 'Concluído Em', 'Revisado Por', 'Notas da Execução'
        ]
        for col, header in enumerate(headers, 1):
            worksheet.cell(row=1, column=col, value=header)
        
        # Data
        for row, task in enumerate(tasks_query, 2):
            worksheet.cell(row=row, column=1, value=task.checklist.date)
            worksheet.cell(row=row, column=2, value=task.template.sector.name)
            worksheet.cell(row=row, column=3, value=task.template.title)
            worksheet.cell(row=row, column=4, value=task.get_status_display())
            worksheet.cell(row=row, column=5, value=task.assigned_to.get_full_name() if task.assigned_to else '')
            worksheet.cell(row=row, column=6, value=task.started_at.strftime('%d/%m/%Y %H:%M') if task.started_at else '')
            worksheet.cell(row=row, column=7, value=task.completed_at.strftime('%d/%m/%Y %H:%M') if task.completed_at else '')
            worksheet.cell(row=row, column=8, value=task.reviewed_by.get_full_name() if task.reviewed_by else '')
            worksheet.cell(row=row, column=9, value=task.execution_notes)
        
        workbook.save(output)
        output.seek(0)
        
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="checklist_admin_{start_date}_{end_date}.xlsx"'
        
        return response


@login_required
def admin_checklist_detail_view(request, checklist_id):
    """Visualização detalhada de um checklist específico"""
    checklist = get_object_or_404(DailyAdminChecklist, id=checklist_id)
    
    # Verificar permissões
    user_sectors = []
    if hasattr(request.user, 'sector') and request.user.sector:
        user_sectors.append(request.user.sector.id)
    
    # Filtrar tarefas baseado nas permissões
    if request.user.hierarchy == 'SUPERADMIN':
        tasks = checklist.tasks.all()
    else:
        tasks = checklist.tasks.filter(template__sector_id__in=user_sectors)
    
    tasks = tasks.select_related('template', 'template__sector', 'assigned_to', 'reviewed_by')
    
    # Estatísticas do checklist
    stats = {
        'total': tasks.count(),
        'pending': tasks.filter(status='PENDING').count(),
        'in_progress': tasks.filter(status='IN_PROGRESS').count(),
        'completed': tasks.filter(status='COMPLETED').count(),
        'approved': tasks.filter(status='APPROVED').count(),
        'rejected': tasks.filter(status='REJECTED').count(),
    }
    
    # Cálculo da porcentagem de conclusão
    if stats['total'] > 0:
        completed_tasks = stats['completed'] + stats['approved']
        stats['completion_percentage'] = round((completed_tasks / stats['total']) * 100, 1)
    else:
        stats['completion_percentage'] = 0
    
    # Agrupar tarefas por setor
    tasks_by_sector = {}
    for task in tasks:
        sector_name = task.template.sector.name
        if sector_name not in tasks_by_sector:
            tasks_by_sector[sector_name] = []
        tasks_by_sector[sector_name].append(task)
    
    # Usuários do setor para atribuição (apenas para superadmin)
    sector_users = []
    if request.user.hierarchy == 'SUPERADMIN':
        from users.models import User
        sector_users = User.objects.filter(
            sector__in=[task.template.sector for task in tasks],
            is_active=True
        ).select_related('sector').distinct()
    
    context = {
        'checklist': checklist,
        'tasks': tasks,
        'stats': stats,
        'tasks_by_sector': tasks_by_sector,
        'sector_users': sector_users,
        'can_assign': request.user.hierarchy == 'SUPERADMIN',
        'can_review': request.user.hierarchy in ['SUPERADMIN', 'ADMINISTRATIVO'],
        'user_sector_id': request.user.sector.id if hasattr(request.user, 'sector') and request.user.sector else None
    }
    
    return render(request, 'core/admin_checklist_detail.html', context)


@login_required
@require_http_methods(["POST"])
def admin_checklist_assign_task(request):
    """Atribuir tarefa a um usuário específico"""
    if request.user.hierarchy != 'SUPERADMIN':
        return JsonResponse({'success': False, 'error': 'Sem permissão para atribuir tarefas'})
    
    try:
        task_id = request.POST.get('task_id')
        user_id = request.POST.get('user_id')
        
        if not task_id or not user_id:
            return JsonResponse({'success': False, 'error': 'Parâmetros obrigatórios não fornecidos'})
        
        task = AdminChecklistTask.objects.get(id=task_id)
        user = User.objects.get(id=user_id)
        
        # Verificar se o usuário pertence ao setor da tarefa
        if user.sector != task.template.sector:
            return JsonResponse({
                'success': False, 
                'error': f'Usuário não pertence ao setor {task.template.sector.name}'
            })
        
        # Atribuir a tarefa
        task.assigned_to = user
        task.save()
        
        # Criar registro de atribuição
        AdminChecklistAssignment.objects.create(
            task=task,
            user=user,
            assigned_by=request.user
        )
        
        return JsonResponse({
            'success': True,
            'message': f'Tarefa atribuída para {user.get_full_name()}',
            'assigned_user': user.get_full_name(),
            'assigned_user_id': user.id
        })
        
    except AdminChecklistTask.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Tarefa não encontrada'})
    except User.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Usuário não encontrado'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
@require_http_methods(["POST"])
def admin_checklist_bulk_assign(request):
    """Atribuir múltiplas tarefas em lote"""
    if request.user.hierarchy != 'SUPERADMIN':
        return JsonResponse({'success': False, 'error': 'Sem permissão para atribuir tarefas'})
    
    try:
        task_ids = request.POST.getlist('task_ids[]')
        user_id = request.POST.get('user_id')
        
        if not task_ids or not user_id:
            return JsonResponse({'success': False, 'error': 'Parâmetros obrigatórios não fornecidos'})
        
        user = User.objects.get(id=user_id)
        tasks = AdminChecklistTask.objects.filter(id__in=task_ids)
        
        assigned_count = 0
        errors = []
        
        for task in tasks:
            # Verificar se o usuário pertence ao setor da tarefa
            if user.sector != task.template.sector:
                errors.append(f'Tarefa "{task.template.title}" - usuário não pertence ao setor {task.template.sector.name}')
                continue
            
            # Atribuir a tarefa
            task.assigned_to = user
            task.save()
            
            # Criar registro de atribuição
            AdminChecklistAssignment.objects.create(
                task=task,
                user=user,
                assigned_by=request.user
            )
            
            assigned_count += 1
        
        response_data = {
            'success': True,
            'message': f'{assigned_count} tarefa(s) atribuída(s) para {user.get_full_name()}',
            'assigned_count': assigned_count,
            'assigned_user': user.get_full_name()
        }
        
        if errors:
            response_data['warnings'] = errors
        
        return JsonResponse(response_data)
        
    except User.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Usuário não encontrado'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
@require_http_methods(["GET", "POST"])
def admin_checklist_add_activity(request):
    """Adicionar nova atividade (template) ao sistema"""
    if request.user.hierarchy != 'SUPERADMIN':
        return JsonResponse({'success': False, 'error': 'Sem permissão para adicionar atividades'})
    
    if request.method == 'GET':
        # Retornar formulário para adicionar atividade
        from users.models import Sector
        sectors = Sector.objects.all().order_by('name')
        
        context = {
            'sectors': sectors,
            'priority_choices': AdminChecklistTemplate.PRIORITY_CHOICES if hasattr(AdminChecklistTemplate, 'PRIORITY_CHOICES') else [
                ('LOW', 'Baixa'),
                ('MEDIUM', 'Média'),
                ('HIGH', 'Alta'),
                ('URGENT', 'Urgente'),
            ]
        }
        
        return render(request, 'core/admin_checklist_add_activity.html', context)
    
    elif request.method == 'POST':
        try:
            # Validar dados
            title = request.POST.get('title', '').strip()
            description = request.POST.get('description', '').strip()
            sector_id = request.POST.get('sector_id')
            priority = request.POST.get('priority', 'MEDIUM')
            estimated_time = request.POST.get('estimated_time_minutes')
            instructions = request.POST.get('instructions', '').strip()
            is_active = request.POST.get('is_active') == 'on'
            add_to_today = request.POST.get('add_to_today') == 'on'
            
            if not all([title, description, sector_id, estimated_time]):
                return JsonResponse({'success': False, 'error': 'Todos os campos obrigatórios devem ser preenchidos'})
            
            # Criar template
            from users.models import Sector
            sector = Sector.objects.get(id=sector_id)
            
            template = AdminChecklistTemplate.objects.create(
                title=title,
                description=description,
                sector=sector,
                priority=priority,
                estimated_time_minutes=int(estimated_time),
                instructions=instructions,
                is_active=is_active,
                created_by=request.user
            )
            
            # Se solicitado, adicionar ao checklist de hoje
            tasks_added = 0
            if add_to_today:
                from datetime import date
                today_checklist = DailyAdminChecklist.objects.filter(date=date.today()).first()
                
                if today_checklist:
                    # Verificar se já existe tarefa deste template no checklist de hoje
                    existing_task = AdminChecklistTask.objects.filter(
                        checklist=today_checklist,
                        template=template
                    ).first()
                    
                    if not existing_task:
                        AdminChecklistTask.objects.create(
                            checklist=today_checklist,
                            template=template,
                            status='PENDING'
                        )
                        tasks_added = 1
            
            response_data = {
                'success': True,
                'message': f'Atividade "{template.title}" criada com sucesso!',
                'template_id': template.id
            }
            
            if tasks_added:
                response_data['message'] += f' E adicionada ao checklist de hoje.'
            
            return JsonResponse(response_data)
            
        except Sector.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Setor não encontrado'})
        except ValueError:
            return JsonResponse({'success': False, 'error': 'Tempo estimado deve ser um número válido'})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})


@login_required
def admin_checklist_user_tasks(request):
    """Visualizar tarefas do usuário logado"""
    user = request.user
    
    # Filtros
    status_filter = request.GET.get('status', '')
    date_filter = request.GET.get('date', '')
    
    # Query base - tarefas do usuário
    tasks_query = AdminChecklistTask.objects.filter(assigned_to=user)
    
    # Aplicar filtros
    if status_filter:
        tasks_query = tasks_query.filter(status=status_filter)
    
    if date_filter:
        from datetime import datetime
        try:
            filter_date = datetime.strptime(date_filter, '%Y-%m-%d').date()
            tasks_query = tasks_query.filter(checklist__date=filter_date)
        except ValueError:
            pass
    
    tasks_query = tasks_query.select_related(
        'checklist', 'template', 'template__sector', 'reviewed_by'
    ).order_by('-checklist__date', 'template__priority', 'template__title')
    
    # Paginação
    from django.core.paginator import Paginator
    paginator = Paginator(tasks_query, 20)
    page_number = request.GET.get('page')
    tasks = paginator.get_page(page_number)
    
    # Estatísticas do usuário
    user_stats = {
        'total': AdminChecklistTask.objects.filter(assigned_to=user).count(),
        'pending': AdminChecklistTask.objects.filter(assigned_to=user, status='PENDING').count(),
        'in_progress': AdminChecklistTask.objects.filter(assigned_to=user, status='IN_PROGRESS').count(),
        'completed': AdminChecklistTask.objects.filter(assigned_to=user, status='COMPLETED').count(),
        'approved': AdminChecklistTask.objects.filter(assigned_to=user, status='APPROVED').count(),
        'rejected': AdminChecklistTask.objects.filter(assigned_to=user, status='REJECTED').count(),
    }
    
    context = {
        'tasks': tasks,
        'user_stats': user_stats,
        'status_filter': status_filter,
        'date_filter': date_filter,
        'status_choices': AdminChecklistTask.STATUS_CHOICES,
        'can_execute': user.hierarchy in ['SUPERADMIN', 'ADMINISTRATIVO', 'SUPERVISOR'],
    }
    
    return render(request, 'core/admin_checklist_user_tasks.html', context)


# ===== VIEWS PARA TAREFAS DE SETOR =====

@login_required
def admin_checklist_sector_tasks(request):
    """Lista tarefas do setor para o checklist administrativo"""
    user = request.user
    
    # Verificar permissões
    if user.hierarchy == 'PADRAO':
        return HttpResponseForbidden("Acesso negado")
    
    # Obter setores do usuário
    user_sectors = list(user.sectors.all())
    if user.sector:
        user_sectors.append(user.sector)
    
    # Filtrar tarefas do setor
    sector_tasks = AdminChecklistSectorTask.objects.filter(
        sector__in=user_sectors
    ).order_by('-created_at')
    
    # Filtros
    status_filter = request.GET.get('status', 'all')
    if status_filter == 'pending':
        sector_tasks = sector_tasks.filter(is_approved=False)
    elif status_filter == 'approved':
        sector_tasks = sector_tasks.filter(is_approved=True)
    
    date_filter = request.GET.get('date')
    if date_filter:
        sector_tasks = sector_tasks.filter(date_requested=date_filter)
    
    context = {
        'sector_tasks': sector_tasks,
        'user_sectors': user_sectors,
        'status_filter': status_filter,
        'date_filter': date_filter,
        'can_create': user.hierarchy in ['SUPERADMIN', 'ADMINISTRATIVO', 'SUPERVISOR'],
    }
    
    return render(request, 'core/admin_checklist_sector_tasks.html', context)


@login_required
def admin_checklist_create_sector_task(request):
    """Criar nova tarefa de setor para o checklist administrativo"""
    user = request.user
    
    # Verificar permissões
    if user.hierarchy not in ['SUPERADMIN', 'ADMINISTRATIVO', 'SUPERVISOR']:
        return HttpResponseForbidden("Acesso negado")
    
    if request.method == 'POST':
        from users.models import Sector
        
        title = request.POST.get('title', '').strip()
        description = request.POST.get('description', '').strip()
        instructions = request.POST.get('instructions', '').strip()
        sector_id = request.POST.get('sector')
        priority = request.POST.get('priority', 'MEDIUM')
        estimated_time = request.POST.get('estimated_time', 30)
        date_requested = request.POST.get('date_requested')
        
        # Validações
        errors = []
        
        if not title:
            errors.append("Título é obrigatório")
        if not description:
            errors.append("Descrição é obrigatória")
        if not sector_id:
            errors.append("Setor é obrigatório")
        if not date_requested:
            errors.append("Data solicitada é obrigatória")
        
        # Verificar se o setor existe e se o usuário tem acesso
        try:
            sector = Sector.objects.get(id=sector_id)
            user_sectors = list(user.sectors.all())
            if user.sector:
                user_sectors.append(user.sector)
            
            if user.hierarchy != 'SUPERADMIN' and sector not in user_sectors:
                errors.append("Você não tem permissão para criar tarefas para este setor")
        except Sector.DoesNotExist:
            errors.append("Setor não encontrado")
            sector = None
        
        if not errors:
            # Criar a tarefa
            AdminChecklistSectorTask.objects.create(
                title=title,
                description=description,
                instructions=instructions,
                sector=sector,
                priority=priority,
                estimated_time_minutes=int(estimated_time),
                date_requested=date_requested,
                created_by=user,
            )
            
            messages.success(request, f"Tarefa '{title}' criada com sucesso! Aguardando aprovação do superadmin.")
            return redirect('admin_checklist:sector_tasks')
        else:
            for error in errors:
                messages.error(request, error)
    
    # Obter setores disponíveis
    from users.models import Sector
    
    if user.hierarchy == 'SUPERADMIN':
        available_sectors = Sector.objects.all()
    else:
        available_sectors = list(user.sectors.all())
        if user.sector:
            available_sectors.append(user.sector)
    
    context = {
        'available_sectors': available_sectors,
        'priorities': AdminChecklistSectorTask._meta.get_field('priority').choices,
    }
    
    return render(request, 'core/admin_checklist_create_sector_task.html', context)


@login_required
def admin_checklist_approve_sector_tasks(request):
    """Aprovar tarefas de setor (apenas superadmin)"""
    user = request.user
    
    # Apenas superadmin pode aprovar
    if user.hierarchy != 'SUPERADMIN':
        return HttpResponseForbidden("Apenas superadmin pode aprovar tarefas")
    
    if request.method == 'POST':
        action = request.POST.get('action')
        task_ids = request.POST.getlist('task_ids')
        
        if not task_ids:
            messages.error(request, "Nenhuma tarefa selecionada")
            return redirect('admin_checklist:approve_sector_tasks')
        
        tasks = AdminChecklistSectorTask.objects.filter(
            id__in=task_ids,
            is_approved=False
        )
        
        if action == 'approve':
            # Aprovar tarefas selecionadas
            for task in tasks:
                task.is_approved = True
                task.approved_by = user
                task.approved_at = timezone.now()
                task.save()
                
                # Adicionar à checklist do dia solicitado (se existir)
                try:
                    daily_checklist = DailyAdminChecklist.objects.get(date=task.date_requested)
                    AdminChecklistTask.objects.create(
                        checklist=daily_checklist,
                        sector_task=task,
                        title=task.title,
                        description=task.description,
                        instructions=task.instructions,
                        status='PENDING'
                    )
                except DailyAdminChecklist.DoesNotExist:
                    # Se não existe checklist para o dia, criar
                    daily_checklist = DailyAdminChecklist.objects.create(
                        date=task.date_requested,
                        created_by=user
                    )
                    AdminChecklistTask.objects.create(
                        checklist=daily_checklist,
                        sector_task=task,
                        title=task.title,
                        description=task.description,
                        instructions=task.instructions,
                        status='PENDING'
                    )
            
            messages.success(request, f"{tasks.count()} tarefa(s) aprovada(s) com sucesso!")
            
        elif action == 'reject':
            # Rejeitar (excluir) tarefas selecionadas
            count = tasks.count()
            tasks.delete()
            messages.success(request, f"{count} tarefa(s) rejeitada(s)")
        
        return redirect('admin_checklist:approve_sector_tasks')
    
    # Listar tarefas pendentes de aprovação
    pending_tasks = AdminChecklistSectorTask.objects.filter(
        is_approved=False
    ).order_by('date_requested', '-created_at')
    
    context = {
        'pending_tasks': pending_tasks,
    }
    
    return render(request, 'core/admin_checklist_approve_sector_tasks.html', context)
