from decimal import Decimal, InvalidOperation
import os
import re
import unicodedata
from collections import defaultdict
from urllib.parse import unquote, urlparse

from django.db.models import Count
from django.db.models.functions import TruncDate
from django.db import transaction
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone

import openpyxl
from openpyxl.styles import Font, PatternFill
from django.utils import timezone

import openpyxl

from .forms import GoalUploadForm, PowerBIReportForm
from .models import GoalEntry, GoalUpload, PowerBIAccessLog, PowerBIReport


DEFAULT_MYSQL_URI = 'mysql://redeconfiancaadm:redeconfianca2025@painel.dev.redeconfianca.com.br:3306/rede_confianca_data'


def _is_superadmin(user):
    return user.is_superuser or user.hierarchy == 'SUPERADMIN'


def _is_standard_user(user):
    return user.hierarchy in ['PADRAO', 'PADRÃO'] and not user.is_superuser


def _is_gerentes_group_user(user):
    """Verifica se o usuario esta no grupo GERENTES via CommunicationGroup."""
    try:
        from communications.models import CommunicationGroup

        gerente_group = CommunicationGroup.objects.filter(name__icontains='GERENTES').first()
        if gerente_group:
            print(f"[is_user_gerente] Verificando se {user.get_full_name()} esta no grupo GERENTES")
            return user in gerente_group.members.all()
    except Exception:
        pass
    return False


def _get_user_store_candidates(user):
    candidates = set()
    if getattr(user, 'pdv', None):
        normalized_pdv = _normalize_text(user.pdv)
        if normalized_pdv:
            candidates.add(normalized_pdv)
    if getattr(user, 'sector', None) and getattr(user.sector, 'name', None):
        normalized_sector = _normalize_text(user.sector.name)
        if normalized_sector:
            candidates.add(normalized_sector)
    for sector in user.sectors.all():
        normalized_sector = _normalize_text(getattr(sector, 'name', ''))
        if normalized_sector:
            candidates.add(normalized_sector)
    return candidates


def _get_user_primary_store(user):
    if getattr(user, 'pdv', None):
        normalized_pdv = _normalize_text(user.pdv)
        if normalized_pdv and not _is_network_store(normalized_pdv):
            return normalized_pdv

    if getattr(user, 'sector', None) and getattr(user.sector, 'name', None):
        normalized_sector = _normalize_text(user.sector.name)
        if normalized_sector and not _is_network_store(normalized_sector):
            return normalized_sector

    first_sector = user.sectors.order_by('id').first()
    if first_sector and getattr(first_sector, 'name', None):
        normalized_sector = _normalize_text(first_sector.name)
        if normalized_sector and not _is_network_store(normalized_sector):
            return normalized_sector

    return ''


def _is_network_store(store_name):
    normalized = _normalize_text(store_name)
    if not normalized:
        return False
    return (
        normalized == 'REDE'
        or normalized.startswith('REDE ')
        or normalized.endswith(' REDE')
        or ' TOTAL REDE' in normalized
        or normalized == 'TOTAL REDE'
    )


def _normalize_store_key(value):
    normalized = _normalize_text(value)
    for prefix in ('LOJA ', 'PDV ', 'FILIAL '):
        if normalized.startswith(prefix):
            normalized = normalized[len(prefix):].strip()
    return normalized


def _stores_match(store_a, store_b):
    a = _normalize_store_key(store_a)
    b = _normalize_store_key(store_b)
    if not a or not b:
        return False
    if a == b:
        return True
    if a in b or b in a:
        return True
    a_words = set(a.split())
    b_words = set(b.split())
    common = a_words & b_words
    return len(common) >= 2 or (len(common) == 1 and min(len(a_words), len(b_words)) == 1)


def _normalize_text(value):
    if value is None:
        return ''
    text = str(value).strip()
    text = unicodedata.normalize('NFKD', text)
    text = ''.join(ch for ch in text if not unicodedata.combining(ch))
    return ' '.join(text.upper().split())


def _normalize_sheet_name(value):
    return _normalize_text(value).replace('_', '').replace(' ', '')


def _is_fixa_pilar(pilar):
    return _normalize_text(pilar) == 'FIXA'


def _store_key(value):
    """Chave de comparacao entre o setor do gerente e o PDV da planilha.

    Alem de caixa/acentos (via _normalize_text), remove a palavra LOJA e os
    separadores que sobram, para que "Loja - Jardim Camburi" e o PDV
    "JARDIM CAMBURI" resultem na mesma chave.
    """
    text = re.sub(r'\bLOJA\b', ' ', _normalize_text(value))
    return ' '.join(text.strip(' -–—:').split())


def _gerente_names_by_store():
    """Mapeia cada loja para os nomes dos gerentes do grupo GERENTES.

    O vinculo gerente -> PDV vem do setor principal do gerente. Retorna
    tambem se o grupo existe, para diferenciar "grupo ausente" de
    "grupo sem membros com setor".
    """
    from communications.models import CommunicationGroup

    group = CommunicationGroup.objects.filter(name__iexact='GERENTES').first()
    if group is None:
        return {}, False

    names_by_store = defaultdict(list)
    members = group.members.filter(is_active=True).select_related('sector').prefetch_related('sectors')
    for member in members:
        sector = member.primary_sector
        if sector is None:
            continue
        store = _store_key(sector.name)
        name = (member.full_name or '').strip()
        if not store or not name:
            continue
        names_by_store[store].append(name)

    return names_by_store, True


def _get_goals_mysql_config():
    mysql_uri = os.getenv('MYSQL_URI', DEFAULT_MYSQL_URI)
    parsed = urlparse(mysql_uri)

    if parsed.scheme.lower() != 'mysql':
        raise ValueError('MYSQL_URI invalida: o schema deve ser mysql://')

    if not parsed.hostname or not parsed.path or parsed.path == '/':
        raise ValueError('MYSQL_URI invalida: host e database sao obrigatorios.')

    return {
        'host': parsed.hostname,
        'port': parsed.port or 3306,
        'user': unquote(parsed.username or ''),
        'password': unquote(parsed.password or ''),
        'database': parsed.path.lstrip('/'),
        'charset': 'utf8mb4',
    }


def _parse_decimal(value):
    if value is None:
        return None
    if isinstance(value, Decimal):
        return value
    if isinstance(value, (int, float)):
        try:
            return Decimal(str(value)).quantize(Decimal('0.01'))
        except (InvalidOperation, ValueError):
            return None

    text = str(value).strip()
    if not text:
        return None

    text = text.replace('R$', '').replace(' ', '')
    if ',' in text and '.' in text:
        text = text.replace('.', '').replace(',', '.')
    else:
        text = text.replace(',', '.')

    try:
        return Decimal(text).quantize(Decimal('0.01'))
    except (InvalidOperation, ValueError):
        return None


PILAR_COLUMNS = [
    ('MOVEL', 'MOVEL'),
    ('FIXA', 'FIXA'),
    ('SVA', 'SVA'),
    ('SEGURO', 'SEGURO'),
    ('SMARTPHONE', 'SMARTPHONE'),
    ('ELETRONICOS', 'ELETRONICOS'),
    ('ESSENCIAIS', 'ESSENCIAIS'),
]


def _find_index_by_aliases(headers, aliases):
    normalized = [_normalize_text(h) for h in headers]
    for idx, header in enumerate(normalized):
        if any(alias in header for alias in aliases):
            return idx
    return None


def _extract_entries_cn_real(headers, rows):
    consultor_idx = _find_index_by_aliases(headers, ['CONSULTOR', 'CN', 'NOME'])
    pdv_idx = _find_index_by_aliases(headers, ['PDV', 'LOJA', 'FILIAL'])
    pilar_indexes = {}
    normalized_headers = [_normalize_text(h) for h in headers]

    for key, label in PILAR_COLUMNS:
        idx = None
        for header_idx, header in enumerate(normalized_headers):
            if key in header:
                idx = header_idx
                break
        pilar_indexes[label] = idx

    entries = []
    for row_number, row in enumerate(rows, start=2):
        if all(cell in (None, '') for cell in row):
            continue

        user_name = ''
        store_name = ''
        if consultor_idx is not None and consultor_idx < len(row):
            user_name = '' if row[consultor_idx] is None else str(row[consultor_idx]).strip()
        if pdv_idx is not None and pdv_idx < len(row):
            store_name = '' if row[pdv_idx] is None else str(row[pdv_idx]).strip()

        row_data = {}
        for col_idx, header in enumerate(headers):
            if not header:
                continue
            row_data[header] = row[col_idx] if col_idx < len(row) else None

        for pilar_name, idx in pilar_indexes.items():
            if idx is None or idx >= len(row):
                continue
            goal_value = _parse_decimal(row[idx])
            if goal_value is None:
                continue

            entries.append({
                'sheet_type': GoalEntry.SHEET_CN_REAL,
                'user_name': user_name,
                'store_name': store_name,
                'pilar': pilar_name,
                'goal_value': goal_value,
                'row_number': row_number,
                'row_data': row_data,
            })

    return entries


def _extract_entries_pdv_real(headers, rows):
    pdv_idx = _find_index_by_aliases(headers, ['PDV', 'LOJA', 'FILIAL'])
    pilar_indexes = {}
    normalized_headers = [_normalize_text(h) for h in headers]

    for key, label in PILAR_COLUMNS:
        idx = None
        for header_idx, header in enumerate(normalized_headers):
            if key in header:
                idx = header_idx
                break
        pilar_indexes[label] = idx

    entries = []
    for row_number, row in enumerate(rows, start=2):
        if all(cell in (None, '') for cell in row):
            continue

        store_name = ''
        if pdv_idx is not None and pdv_idx < len(row):
            store_name = '' if row[pdv_idx] is None else str(row[pdv_idx]).strip()

        if _is_network_store(store_name):
            continue

        row_data = {}
        for col_idx, header in enumerate(headers):
            if not header:
                continue
            row_data[header] = row[col_idx] if col_idx < len(row) else None

        for pilar_name, idx in pilar_indexes.items():
            if idx is None or idx >= len(row):
                continue
            goal_value = _parse_decimal(row[idx])
            if goal_value is None:
                continue

            entries.append({
                'sheet_type': GoalEntry.SHEET_PDV_REAL,
                'user_name': '',
                'store_name': store_name,
                'pilar': pilar_name,
                'goal_value': goal_value,
                'row_number': row_number,
                'row_data': row_data,
            })

    return entries


def _extract_pcn_by_consultor(headers, rows):
    """Mapeia o nome do consultor para o valor da coluna '% CN' na sheet METAS CN REAL."""
    consultor_idx = _find_index_by_aliases(headers, ['CONSULTOR', 'NOME'])

    normalized_headers = [_normalize_text(h) for h in headers]
    pcn_idx = None
    for header_idx, header in enumerate(normalized_headers):
        if header == '% CN' or header == '%CN':
            pcn_idx = header_idx
            break
    if pcn_idx is None:
        pcn_idx = _find_index_by_aliases(headers, ['% CN', '%CN'])

    pcn_by_consultor = {}
    if consultor_idx is None or pcn_idx is None:
        return pcn_by_consultor

    for row in rows:
        if all(cell in (None, '') for cell in row):
            continue
        if consultor_idx >= len(row):
            continue
        consultor = '' if row[consultor_idx] is None else str(row[consultor_idx]).strip()
        if not consultor:
            continue
        pcn_value = ''
        if pcn_idx < len(row) and row[pcn_idx] is not None:
            pcn_value = str(row[pcn_idx]).strip()
        # primeira ocorrencia prevalece
        pcn_by_consultor.setdefault(consultor, pcn_value)

    return pcn_by_consultor


def _pcn_from_row_data(row_data):
    """Recupera o valor da coluna '% CN' a partir do row_data salvo na GoalEntry."""
    for key, value in (row_data or {}).items():
        if _normalize_text(key).replace(' ', '') == '%CN':
            return '' if value is None else str(value).strip()
    return ''


def _update_users_pcn(pcn_by_consultor):
    """Atualiza o campo PCN dos usuarios casando pelo nome do consultor.

    Retorna a quantidade de usuarios efetivamente atualizados.
    """
    from users.models import User

    if not pcn_by_consultor:
        return 0

    normalized_pcn = {}
    for consultor, pcn_value in pcn_by_consultor.items():
        key = _normalize_text(consultor)
        if key:
            normalized_pcn.setdefault(key, pcn_value)

    if not normalized_pcn:
        return 0

    users_to_update = []
    for user in User.objects.all():
        full_name_key = _normalize_text(f'{user.first_name} {user.last_name}')
        if not full_name_key:
            continue
        if full_name_key in normalized_pcn:
            new_pcn = normalized_pcn[full_name_key] or ''
            if user.pcn != new_pcn:
                user.pcn = new_pcn
                users_to_update.append(user)

    if users_to_update:
        User.objects.bulk_update(users_to_update, ['pcn'])

    return len(users_to_update)


def _find_sector_for_store(sectors, store_name):
    """Encontra o Setor correspondente a uma loja (PDV) das metas.

    Prioriza correspondencia exata (apos normalizar e remover prefixo LOJA/PDV/FILIAL)
    e, em ultimo caso, usa a correspondencia aproximada de _stores_match.
    """
    target = _normalize_store_key(store_name)
    if not target:
        return None
    for sector in sectors:
        if _normalize_store_key(sector.name) == target:
            return sector
    for sector in sectors:
        if _stores_match(sector.name, store_name):
            return sector
    return None


def _update_users_sectors(store_by_consultor):
    """Atualiza o setor principal e os setores dos usuarios.

    Casa o usuario pelo nome do consultor (sheet METAS CN REAL) e o setor pela
    loja (coluna PDV) informada na mesma linha. Mantem os demais setores ja
    vinculados ao usuario, apenas adicionando o setor da loja e definindo-o como
    setor principal.

    Retorna a quantidade de usuarios efetivamente atualizados.
    """
    from users.models import Sector, User

    if not store_by_consultor:
        return 0

    normalized_store = {}
    for consultor, store in store_by_consultor.items():
        key = _normalize_text(consultor)
        if key and store:
            normalized_store.setdefault(key, store)

    if not normalized_store:
        return 0

    sectors = list(Sector.objects.all())
    if not sectors:
        return 0

    updated = 0
    for user in User.objects.all():
        full_name_key = _normalize_text(f'{user.first_name} {user.last_name}')
        if not full_name_key or full_name_key not in normalized_store:
            continue

        sector = _find_sector_for_store(sectors, normalized_store[full_name_key])
        if sector is None:
            continue

        changed = False
        if user.sector_id != sector.id:
            user.sector = sector
            user.save(update_fields=['sector'])
            changed = True
        if not user.sectors.filter(id=sector.id).exists():
            user.sectors.add(sector)
            changed = True
        if changed:
            updated += 1

    return updated


def _build_store_by_consultor(entries):
    """Monta o mapa nome do consultor -> loja a partir das linhas da sheet CN REAL."""
    store_by_consultor = {}
    for entry in entries:
        if entry.get('sheet_type') != GoalEntry.SHEET_CN_REAL:
            continue
        name = (entry.get('user_name') or '').strip()
        store = (entry.get('store_name') or '').strip()
        if name and store:
            store_by_consultor.setdefault(name, store)
    return store_by_consultor


def _load_goal_entries_from_workbook(uploaded_file):
    workbook = openpyxl.load_workbook(uploaded_file, data_only=True)

    required = {
        'METASCNREAL': GoalEntry.SHEET_CN_REAL,
        'METAPDVREAL': GoalEntry.SHEET_PDV_REAL,
    }

    normalized_to_real = {
        _normalize_sheet_name(sheet_name): sheet_name for sheet_name in workbook.sheetnames
    }

    missing = [name for name in required.keys() if name not in normalized_to_real]
    if missing:
        raise ValueError('As sheets obrigatorias METAS  CN REAL e META PDV REAL nao foram encontradas.')

    all_entries = []
    pcn_by_consultor = {}

    cn_sheet = workbook[normalized_to_real['METASCNREAL']]
    cn_rows = list(cn_sheet.iter_rows(values_only=True))
    if cn_rows:
        cn_headers = [str(cell).strip() if cell is not None else '' for cell in cn_rows[0]]
        all_entries.extend(_extract_entries_cn_real(cn_headers, cn_rows[1:]))
        pcn_by_consultor = _extract_pcn_by_consultor(cn_headers, cn_rows[1:])

    pdv_sheet = workbook[normalized_to_real['METAPDVREAL']]
    pdv_rows = list(pdv_sheet.iter_rows(values_only=True))
    if pdv_rows:
        pdv_headers = [str(cell).strip() if cell is not None else '' for cell in pdv_rows[0]]
        all_entries.extend(_extract_entries_pdv_real(pdv_headers, pdv_rows[1:]))

    return all_entries, pcn_by_consultor


def _visible_reports_for(user):
    reports = (
        PowerBIReport.objects.filter(is_active=True)
        .prefetch_related('allowed_groups', 'allowed_sectors', 'allowed_users')
        .order_by('sort_order', 'name')
    )
    if _is_superadmin(user):
        return reports
    return [report for report in reports if report.is_visible_to(user)]


def _build_manage_access_dashboard_context():
    access_logs = PowerBIAccessLog.objects.select_related('user', 'report')

    by_user = (
        access_logs
        .values('user_id', 'user__first_name', 'user__last_name', 'user__username')
        .annotate(total_accesses=Count('id'), active_days=Count(TruncDate('accessed_at'), distinct=True))
        .order_by('-total_accesses', 'user__first_name', 'user__last_name', 'user__username')[:20]
    )

    by_day = (
        access_logs
        .annotate(day=TruncDate('accessed_at'))
        .values('day')
        .annotate(total_accesses=Count('id'), users_count=Count('user_id', distinct=True))
        .order_by('-day')[:30]
    )

    by_report = (
        access_logs
        .values('report_id', 'report__name')
        .annotate(total_accesses=Count('id'), users_count=Count('user_id', distinct=True))
        .order_by('-total_accesses', 'report__name')[:20]
    )

    latest_accesses = access_logs.order_by('-accessed_at')[:20]

    totals = access_logs.aggregate(
        total_accesses=Count('id'),
        unique_users=Count('user_id', distinct=True),
        reports_accessed=Count('report_id', distinct=True),
    )

    return {
        'access_total': totals['total_accesses'] or 0,
        'access_unique_users': totals['unique_users'] or 0,
        'access_reports_used': totals['reports_accessed'] or 0,
        'access_by_user': by_user,
        'access_by_day': by_day,
        'access_by_report': by_report,
        'access_latest': latest_accesses,
    }


@login_required
def power_bi_list_view(request):
    reports = _visible_reports_for(request.user)
    return render(
        request,
        'power_bi/list.html',
        {
            'reports': reports,
        }
    )


@login_required
def power_bi_viewer(request, report_id):
    report = get_object_or_404(
        PowerBIReport.objects.prefetch_related('allowed_groups', 'allowed_sectors', 'allowed_users'),
        id=report_id,
        is_active=True,
    )

    if not report.is_visible_to(request.user):
        messages.error(request, 'Voce nao tem permissao para visualizar este BI.')
        return redirect('power_bi:list')

    PowerBIAccessLog.objects.create(
        report=report,
        user=request.user,
    )

    return render(
        request,
        'power_bi/viewer.html',
        {
            'report': report,
        }
    )


@login_required
def manage_power_bi_view(request):
    if not _is_superadmin(request.user):
        messages.error(request, 'Apenas SUPERADMIN pode gerenciar os links de Power BI.')
        return redirect('dashboard')

    reports = PowerBIReport.objects.all().prefetch_related('allowed_groups', 'allowed_sectors', 'allowed_users')
    form = PowerBIReportForm()

    context = {
        'reports': reports,
        'form': form,
        'editing_report': None,
    }
    context.update(_build_manage_access_dashboard_context())

    return render(
        request,
        'power_bi/manage.html',
        context
    )


@login_required
def create_power_bi_view(request):
    if not _is_superadmin(request.user):
        messages.error(request, 'Apenas SUPERADMIN pode gerenciar os links de Power BI.')
        return redirect('dashboard')

    if request.method != 'POST':
        return redirect('power_bi:manage')

    form = PowerBIReportForm(request.POST, request.FILES)
    if form.is_valid():
        form.save()
        messages.success(request, 'BI criado com sucesso.')
        return redirect('power_bi:manage')

    reports = PowerBIReport.objects.all().prefetch_related('allowed_groups', 'allowed_sectors', 'allowed_users')
    context = {
        'reports': reports,
        'form': form,
        'editing_report': None,
    }
    context.update(_build_manage_access_dashboard_context())

    return render(
        request,
        'power_bi/manage.html',
        context
    )


@login_required
def edit_power_bi_view(request, report_id):
    if not _is_superadmin(request.user):
        messages.error(request, 'Apenas SUPERADMIN pode gerenciar os links de Power BI.')
        return redirect('dashboard')

    report = get_object_or_404(PowerBIReport, id=report_id)

    if request.method == 'POST':
        form = PowerBIReportForm(request.POST, request.FILES, instance=report)
        if form.is_valid():
            form.save()
            messages.success(request, 'BI atualizado com sucesso.')
            return redirect('power_bi:manage')
    else:
        form = PowerBIReportForm(instance=report)

    reports = PowerBIReport.objects.all().prefetch_related('allowed_groups', 'allowed_sectors', 'allowed_users')
    context = {
        'reports': reports,
        'form': form,
        'editing_report': report,
    }
    context.update(_build_manage_access_dashboard_context())

    return render(
        request,
        'power_bi/manage.html',
        context
    )


@login_required
def delete_power_bi_view(request, report_id):
    if not _is_superadmin(request.user):
        messages.error(request, 'Apenas SUPERADMIN pode gerenciar os links de Power BI.')
        return redirect('dashboard')

    if request.method != 'POST':
        return redirect('power_bi:manage')

    report = get_object_or_404(PowerBIReport, id=report_id)
    report.delete()
    messages.success(request, 'BI removido com sucesso.')
    return redirect('power_bi:manage')


@login_required
def export_power_bi_excel_view(request):
    if not _is_superadmin(request.user):
        messages.error(request, 'Apenas SUPERADMIN pode exportar os dados de Power BI.')
        return redirect('dashboard')

    reports = PowerBIReport.objects.all().prefetch_related('allowed_groups', 'allowed_sectors', 'allowed_users')
    access_logs = (
        PowerBIAccessLog.objects
        .select_related('report', 'user', 'user__sector')
        .prefetch_related('user__sectors')
        .order_by('-accessed_at')
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Power BI'

    headers = [
        'Nome',
        'Descricao',
        'Icone',
        'Cor Predominante',
        'Link Embed',
        'Ativo',
        'Ordem',
        'Hierarquias',
        'Grupos',
        'Setores',
        'Usuarios Especificos',
        'Criado Em',
        'Atualizado Em',
    ]

    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='1F2937', end_color='1F2937', fill_type='solid')

    for col_index, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_index, value=header)
        cell.font = header_font
        cell.fill = header_fill

    for row_index, report in enumerate(reports, 2):
        groups = ', '.join(report.allowed_groups.values_list('name', flat=True))
        sectors = ', '.join(report.allowed_sectors.values_list('name', flat=True))
        users = ', '.join(
            [
                user.full_name if user.full_name else user.email
                for user in report.allowed_users.all()
            ]
        )
        hierarchies = ', '.join(report.allowed_hierarchies or [])

        ws.cell(row=row_index, column=1, value=report.name)
        ws.cell(row=row_index, column=2, value=report.description)
        ws.cell(row=row_index, column=3, value=report.icon_class)
        ws.cell(row=row_index, column=4, value=report.predominant_color)
        ws.cell(row=row_index, column=5, value=report.embed_url)
        ws.cell(row=row_index, column=6, value='Sim' if report.is_active else 'Nao')
        ws.cell(row=row_index, column=7, value=report.sort_order)
        ws.cell(row=row_index, column=8, value=hierarchies)
        ws.cell(row=row_index, column=9, value=groups)
        ws.cell(row=row_index, column=10, value=sectors)
        ws.cell(row=row_index, column=11, value=users)
        ws.cell(row=row_index, column=12, value=report.created_at.strftime('%d/%m/%Y %H:%M') if report.created_at else '')
        ws.cell(row=row_index, column=13, value=report.updated_at.strftime('%d/%m/%Y %H:%M') if report.updated_at else '')

    for column_cells in ws.columns:
        max_length = 0
        column = column_cells[0].column_letter
        for cell in column_cells:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column].width = min(max_length + 2, 60)

    ws_access = wb.create_sheet(title='acessos')
    access_headers = [
        'Usuario',
        'Setor do Usuario',
        'Em qual BI entrou',
        'Horario',
        'Dia',
    ]

    for col_index, header in enumerate(access_headers, 1):
        cell = ws_access.cell(row=1, column=col_index, value=header)
        cell.font = header_font
        cell.fill = header_fill

    for row_index, log in enumerate(access_logs, 2):
        full_name = log.user.get_full_name().strip()
        user_name = full_name if full_name else (log.user.full_name or log.user.username)

        user_sector = ''
        if getattr(log.user, 'sector', None) and getattr(log.user.sector, 'name', None):
            user_sector = log.user.sector.name
        elif log.user.sectors.exists():
            user_sector = ', '.join(log.user.sectors.values_list('name', flat=True))

        local_accessed_at = timezone.localtime(log.accessed_at)

        ws_access.cell(row=row_index, column=1, value=user_name)
        ws_access.cell(row=row_index, column=2, value=user_sector)
        ws_access.cell(row=row_index, column=3, value=log.report.name if log.report else '')
        ws_access.cell(row=row_index, column=4, value=local_accessed_at.strftime('%H:%M'))
        ws_access.cell(row=row_index, column=5, value=local_accessed_at.strftime('%d/%m/%Y'))

    for column_cells in ws_access.columns:
        max_length = 0
        column = column_cells[0].column_letter
        for cell in column_cells:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws_access.column_dimensions[column].width = min(max_length + 2, 60)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = (
        f'attachment; filename="power_bi_dashboard_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    )

    wb.save(response)
    return response


@login_required
def goals_list_view(request):
    is_standard_user = _is_standard_user(request.user)
    is_gerente_user = is_standard_user and _is_gerentes_group_user(request.user)
    is_cn_user = is_standard_user and not is_gerente_user

    selected_year = request.GET.get('year')
    selected_month = request.GET.get('month')

    uploads = GoalUpload.objects.order_by('-year', '-month', '-updated_at')
    current_upload = None

    # Todos os perfis podem escolher a competencia (mes/ano).
    if selected_year and selected_month:
        current_upload = uploads.filter(year=selected_year, month=selected_month).first()
    if current_upload is None:
        current_upload = uploads.first()

    entries = GoalEntry.objects.none()
    selected_store = '' if (is_standard_user or is_gerente_user) else request.GET.get('store', '').strip()
    selected_pilar = '' if (is_standard_user or is_gerente_user) else request.GET.get('pilar', '').strip()
    selected_cn = request.GET.get('cn', '').strip() if is_gerente_user else ''
    selected_seller = request.GET.get('seller', '').strip() if not is_cn_user else ''

    if current_upload:
        entries = GoalEntry.objects.filter(upload=current_upload).order_by('sheet_type', 'store_name', 'pilar', 'user_name')

        if is_cn_user:
            entries = entries.filter(sheet_type=GoalEntry.SHEET_CN_REAL)
            user_tokens = {
                _normalize_text(request.user.full_name),
                _normalize_text(request.user.get_full_name()),
                _normalize_text(request.user.first_name),
                _normalize_text(request.user.last_name),
                _normalize_text(request.user.username),
            }
            user_tokens = {token for token in user_tokens if token}

            filtered_ids = []
            for entry in entries:
                entry_name = _normalize_text(entry.user_name)
                if not entry_name:
                    continue
                if any(token == entry_name for token in user_tokens):
                    filtered_ids.append(entry.id)
            entries = entries.filter(id__in=filtered_ids)
        elif is_gerente_user:
            manager_store = _get_user_primary_store(request.user)
            manager_entry_ids = []
            for entry in entries:
                entry_store = _normalize_text(entry.store_name)
                if manager_store and _stores_match(entry_store, manager_store):
                    manager_entry_ids.append(entry.id)
            entries = entries.filter(id__in=manager_entry_ids)
        else:
            if selected_store:
                store_target = _normalize_text(selected_store)
                entry_ids = [
                    entry.id for entry in entries
                    if _normalize_text(entry.store_name) == store_target
                ]
                entries = entries.filter(id__in=entry_ids)

            if selected_pilar:
                pilar_target = _normalize_text(selected_pilar)
                entry_ids = [
                    entry.id for entry in entries
                    if _normalize_text(entry.pilar) == pilar_target
                ]
                entries = entries.filter(id__in=entry_ids)

    fixa_as_percentage = bool(getattr(current_upload, 'fixa_as_percentage', False)) if current_upload else False

    cn_entries = entries.filter(sheet_type=GoalEntry.SHEET_CN_REAL)

    if is_cn_user and current_upload:
        pdv_store_entries = GoalEntry.objects.filter(
            upload=current_upload,
            sheet_type=GoalEntry.SHEET_PDV_REAL,
        )
        user_store = _get_user_primary_store(request.user)
        pdv_store_ids = []
        for entry in pdv_store_entries:
            entry_store = _normalize_text(entry.store_name)
            if user_store and _stores_match(entry_store, user_store):
                pdv_store_ids.append(entry.id)
        pdv_entries = pdv_store_entries.filter(id__in=pdv_store_ids)
    elif is_gerente_user or not is_standard_user:
        pdv_entries = entries.filter(sheet_type=GoalEntry.SHEET_PDV_REAL)
    else:
        pdv_entries = GoalEntry.objects.none()

    available_cns = []
    available_sellers = sorted({entry.user_name for entry in cn_entries if entry.user_name})

    if is_gerente_user:
        available_cns = available_sellers
        if selected_cn:
            selected_cn_normalized = _normalize_text(selected_cn)
            cn_ids = [
                entry.id for entry in cn_entries
                if _normalize_text(entry.user_name) == selected_cn_normalized
            ]
            cn_entries = cn_entries.filter(id__in=cn_ids)
    elif selected_seller:
        selected_seller_normalized = _normalize_text(selected_seller)
        cn_ids = [
            entry.id for entry in cn_entries
            if _normalize_text(entry.user_name) == selected_seller_normalized
        ]
        cn_entries = cn_entries.filter(id__in=cn_ids)

    all_stores = []
    all_pilares = []
    if current_upload and not (is_standard_user or is_gerente_user):
        all_entries = GoalEntry.objects.filter(upload=current_upload)
        all_stores = sorted({entry.store_name for entry in all_entries if entry.store_name and not _is_network_store(entry.store_name)})
        all_pilares = sorted({entry.pilar for entry in all_entries if entry.pilar})

    cn_total_value = sum(
        (item.goal_value or Decimal('0.00'))
        for item in cn_entries
        if not (fixa_as_percentage and _is_fixa_pilar(item.pilar))
    )
    pdv_total_value = sum(
        (item.goal_value or Decimal('0.00'))
        for item in pdv_entries
        if not (fixa_as_percentage and _is_fixa_pilar(item.pilar))
    )
    # Quando vendedor estiver filtrado, o total nao deve agregar a rede inteira.
    if selected_cn or selected_seller:
        total_value = cn_total_value
    else:
        # PDV representa o total da loja; quando existir, evita duplicidade com soma de CN.
        total_value = pdv_total_value if pdv_total_value > Decimal('0.00') else cn_total_value

    pilar_cn_map = defaultdict(lambda: Decimal('0.00'))
    pilar_pdv_map = defaultdict(lambda: Decimal('0.00'))
    for item in cn_entries:
        if item.goal_value is not None:
            pilar_cn_map[item.pilar or 'SEM PILAR'] += item.goal_value
    for item in pdv_entries:
        if item.goal_value is not None:
            pilar_pdv_map[item.pilar or 'SEM PILAR'] += item.goal_value

    cn_pilar_rows = []
    cn_max_total = max(pilar_cn_map.values()) if pilar_cn_map else Decimal('0.00')
    cn_max_total = cn_max_total if cn_max_total > Decimal('0.00') else Decimal('1.00')
    for pilar in sorted(pilar_cn_map.keys()):
        total = pilar_cn_map[pilar]
        is_quantity = fixa_as_percentage and _is_fixa_pilar(pilar)
        cn_pilar_rows.append({
            'pilar': pilar,
            'total': total,
            'is_quantity': is_quantity,
            'bar_percent': float((total / cn_max_total) * Decimal('100')),
        })

    manager_cn_table_rows = []
    manager_cn_table_columns = [
        'SVA',
        'SEGURO',
        'MOVEL',
        'FIXA',
        'SMARTPHONE',
        'ELETRONICOS',
        'ESSENCIAIS',
    ]
    if is_gerente_user:
        by_seller = defaultdict(lambda: defaultdict(lambda: Decimal('0.00')))
        for item in cn_entries:
            seller = item.user_name or 'SEM CONSULTOR'
            pilar = _normalize_text(item.pilar)
            if item.goal_value is None:
                continue
            by_seller[seller][pilar] += item.goal_value

        for seller in sorted(by_seller.keys()):
            row_total = Decimal('0.00')
            row = {
                'seller': seller,
                'ordered_cells': [],
            }
            for column in manager_cn_table_columns:
                value = by_seller[seller].get(column, Decimal('0.00'))
                is_quantity = fixa_as_percentage and _is_fixa_pilar(column)
                row['ordered_cells'].append({
                    'value': value,
                    'is_quantity': is_quantity,
                })
                if not is_quantity:
                    row_total += value
            row['row_total'] = row_total
            manager_cn_table_rows.append(row)

    pdv_pilar_rows = []
    pdv_max_total = max(pilar_pdv_map.values()) if pilar_pdv_map else Decimal('0.00')
    pdv_max_total = pdv_max_total if pdv_max_total > Decimal('0.00') else Decimal('1.00')
    for pilar in sorted(pilar_pdv_map.keys()):
        total = pilar_pdv_map[pilar]
        is_quantity = fixa_as_percentage and _is_fixa_pilar(pilar)
        pdv_pilar_rows.append({
            'pilar': pilar,
            'total': total,
            'is_quantity': is_quantity,
            'bar_percent': float((total / pdv_max_total) * Decimal('100')),
        })

    consultant_totals = defaultdict(lambda: Decimal('0.00'))
    for item in cn_entries:
        if item.goal_value is None:
            continue
        if fixa_as_percentage and _is_fixa_pilar(item.pilar):
            continue
        consultant_key = item.user_name or 'SEM CONSULTOR'
        consultant_totals[consultant_key] += item.goal_value
    top_consultants = sorted(
        [{'name': key, 'total': value} for key, value in consultant_totals.items()],
        key=lambda x: x['total'],
        reverse=True
    )[:10]

    store_totals = defaultdict(lambda: Decimal('0.00'))
    if is_gerente_user or not is_standard_user:
        for item in pdv_entries:
            if item.goal_value is None:
                continue
            if fixa_as_percentage and _is_fixa_pilar(item.pilar):
                continue
            store_key = item.store_name or 'SEM LOJA'
            store_totals[store_key] += item.goal_value
    top_stores = sorted(
        [{'store': key, 'total': value} for key, value in store_totals.items()],
        key=lambda x: x['total'],
        reverse=True
    )[:10]

    context = {
        'uploads': uploads[:24],
        'current_upload': current_upload,
        'selected_store': selected_store,
        'selected_pilar': selected_pilar,
        'all_stores': all_stores,
        'all_pilares': all_pilares,
        'is_standard_user': is_standard_user,
        'is_gerente_user': is_gerente_user,
        'is_cn_user': is_cn_user,
        'selected_cn': selected_cn,
        'selected_seller': selected_seller,
        'available_cns': available_cns,
        'available_sellers': available_sellers,
        'total_value': total_value,
        'entries_count': entries.count() if hasattr(entries, 'count') else len(entries),
        'cn_total': cn_total_value,
        'pdv_total': pdv_total_value,
        'cn_entries_count': cn_entries.count(),
        'pdv_entries_count': pdv_entries.count(),
        'cn_pilar_rows': cn_pilar_rows,
        'pdv_pilar_rows': pdv_pilar_rows,
        'top_consultants': top_consultants,
        'top_stores': top_stores,
        'manager_cn_table_columns': manager_cn_table_columns,
        'manager_cn_table_rows': manager_cn_table_rows,
        'fixa_as_percentage': fixa_as_percentage,
        'now': timezone.now(),
    }
    return render(request, 'power_bi/goals_list.html', context)


@login_required
def manage_goals_view(request):
    if not _is_superadmin(request.user):
        messages.error(request, 'Apenas SUPERADMIN pode gerenciar metas.')
        return redirect('dashboard')

    form = GoalUploadForm()
    uploads = GoalUpload.objects.order_by('-year', '-month', '-updated_at')

    return render(
        request,
        'power_bi/manage_goals.html',
        {
            'form': form,
            'uploads': uploads,
        }
    )


@login_required
def upload_goals_view(request):
    if not _is_superadmin(request.user):
        messages.error(request, 'Apenas SUPERADMIN pode gerenciar metas.')
        return redirect('dashboard')

    if request.method != 'POST':
        return redirect('power_bi:manage_goals')

    form = GoalUploadForm(request.POST, request.FILES)
    if not form.is_valid():
        uploads = GoalUpload.objects.order_by('-year', '-month', '-updated_at')
        return render(
            request,
            'power_bi/manage_goals.html',
            {
                'form': form,
                'uploads': uploads,
            }
        )

    uploaded_file = form.cleaned_data['file']
    year = form.cleaned_data['year']
    month = form.cleaned_data['month']
    fixa_as_percentage = form.cleaned_data.get('fixa_as_percentage', False)

    try:
        uploaded_file.seek(0)
        parsed_entries, pcn_by_consultor = _load_goal_entries_from_workbook(uploaded_file)
    except Exception as exc:
        messages.error(request, f'Erro ao processar planilha de metas: {exc}')
        uploads = GoalUpload.objects.order_by('-year', '-month', '-updated_at')
        return render(
            request,
            'power_bi/manage_goals.html',
            {
                'form': form,
                'uploads': uploads,
            }
        )

    with transaction.atomic():
        upload, _ = GoalUpload.objects.get_or_create(year=year, month=month)
        upload.source_file_name = uploaded_file.name
        upload.fixa_as_percentage = fixa_as_percentage
        upload.uploaded_by = request.user
        upload.save()

        upload.entries.all().delete()
        GoalEntry.objects.bulk_create([
            GoalEntry(upload=upload, **entry) for entry in parsed_entries
        ])

        updated_pcn = _update_users_pcn(pcn_by_consultor)
        updated_sectors = _update_users_sectors(_build_store_by_consultor(parsed_entries))

    messages.success(
        request,
        f'Metas de {month:02d}/{year} importadas com sucesso ({len(parsed_entries)} linhas). '
        f'PCN atualizado em {updated_pcn} usuario(s). '
        f'Setor atualizado em {updated_sectors} usuario(s).'
    )
    return redirect('power_bi:manage_goals')


@login_required
def delete_goals_upload_view(request, upload_id):
    if not _is_superadmin(request.user):
        messages.error(request, 'Apenas SUPERADMIN pode gerenciar metas.')
        return redirect('dashboard')

    if request.method != 'POST':
        return redirect('power_bi:manage_goals')

    upload = get_object_or_404(GoalUpload, id=upload_id)
    period = f'{upload.month:02d}/{upload.year}'
    upload.delete()
    messages.success(request, f'Competencia {period} excluida com sucesso.')
    return redirect('power_bi:manage_goals')


@login_required
def sync_goals_upload_to_mysql_view(request, upload_id):
    if not _is_superadmin(request.user):
        messages.error(request, 'Apenas SUPERADMIN pode sincronizar metas.')
        return redirect('dashboard')

    if request.method != 'POST':
        return redirect('power_bi:manage_goals')

    upload = get_object_or_404(GoalUpload, id=upload_id)
    entries = GoalEntry.objects.filter(
        upload=upload,
        sheet_type=GoalEntry.SHEET_PDV_REAL,
    ).order_by('id')

    if not entries.exists():
        messages.warning(request, 'Nao ha metas para sincronizar nesta competencia.')
        return redirect('power_bi:manage_goals')

    cn_entries = GoalEntry.objects.filter(
        upload=upload,
        sheet_type=GoalEntry.SHEET_CN_REAL,
    ).only('store_name', 'user_name')

    hc_by_store = defaultdict(set)
    for cn_entry in cn_entries:
        store_normalized = _normalize_text(cn_entry.store_name)
        seller_normalized = _normalize_text(cn_entry.user_name)
        if not store_normalized or not seller_normalized:
            continue
        hc_by_store[store_normalized].add(seller_normalized)

    grouped_rows = {}
    for entry in entries:
        if entry.goal_value is None:
            continue

        pdv_raw = (entry.store_name or '').strip()
        pilar = (entry.pilar or '').strip()
        if not pdv_raw:
            continue

        unidade = 'Unidade' if (upload.fixa_as_percentage and _is_fixa_pilar(pilar)) else 'Valor'
        key = (_normalize_text(pdv_raw), _normalize_text(pilar), unidade)

        grouped_rows[key] = {
            'valor': entry.goal_value,
            'pdv': pdv_raw,
            'pilar': pilar,
            'unidade': unidade,
            'hc': len(hc_by_store.get(_normalize_text(pdv_raw), set())),
        }

    rows = [
        (
            row['valor'],
            row['pdv'],
            upload.month,
            upload.year,
            row['unidade'],
            row['pilar'],
            row['hc'],
        )
        for row in grouped_rows.values()
    ]

    if not rows:
        messages.warning(request, 'Nenhum item valido encontrado para sincronizar.')
        return redirect('power_bi:manage_goals')

    # Metas por consultor (sheet METAS CN REAL): uma linha por CN/pilar, com a % CN.
    cn_meta_entries = GoalEntry.objects.filter(
        upload=upload,
        sheet_type=GoalEntry.SHEET_CN_REAL,
    ).order_by('id')

    cn_grouped_rows = {}
    for entry in cn_meta_entries:
        if entry.goal_value is None:
            continue

        cn_name = (entry.user_name or '').strip()
        pdv_raw = (entry.store_name or '').strip()
        pilar = (entry.pilar or '').strip()
        if not cn_name:
            continue

        unidade = 'Unidade' if (upload.fixa_as_percentage and _is_fixa_pilar(pilar)) else 'Valor'
        pcn = _pcn_from_row_data(entry.row_data)
        key = (_normalize_text(cn_name), _normalize_text(pdv_raw), _normalize_text(pilar), unidade)

        cn_grouped_rows[key] = {
            'valor': entry.goal_value,
            'pdv': pdv_raw,
            'cn': cn_name,
            'pilar': pilar,
            'unidade': unidade,
            'pcn': pcn,
        }

    cn_rows = [
        (
            row['valor'],
            row['pdv'],
            row['cn'],
            upload.month,
            upload.year,
            row['unidade'],
            row['pilar'],
            row['pcn'],
        )
        for row in cn_grouped_rows.values()
    ]

    # Metas por gerente: as mesmas linhas de loja, replicadas para cada gerente
    # do grupo GERENTES cujo setor principal aponte para aquele PDV.
    gerentes_by_store, gerente_group_found = _gerente_names_by_store()

    gerente_rows = []
    pdvs_sem_gerente = set()
    for row in grouped_rows.values():
        gerente_names = gerentes_by_store.get(_store_key(row['pdv']), [])
        if not gerente_names:
            pdvs_sem_gerente.add(row['pdv'])
            continue
        for gerente_name in gerente_names:
            gerente_rows.append((
                row['valor'],
                row['pdv'],
                gerente_name,
                upload.month,
                upload.year,
                row['unidade'],
                row['pilar'],
                row['hc'],
            ))

    try:
        import pymysql
    except Exception:
        messages.error(request, 'Dependencia PyMySQL nao encontrada no ambiente para sincronizacao.')
        return redirect('power_bi:manage_goals')

    try:
        mysql_config = _get_goals_mysql_config()
        connection = pymysql.connect(**mysql_config)
        try:
            with connection.cursor() as cursor:
                # DDL primeiro: no MySQL, CREATE TABLE faz commit implicito e
                # encerraria a transacao no meio dos DELETE/INSERT abaixo.
                cursor.execute(
                    'CREATE TABLE IF NOT EXISTS metas_cn ('
                    '  id INT AUTO_INCREMENT PRIMARY KEY,'
                    '  valor DECIMAL(14,2),'
                    '  pdv VARCHAR(255),'
                    '  cn VARCHAR(255),'
                    '  mes_ref INT,'
                    '  ano_ref INT,'
                    '  unidade VARCHAR(20),'
                    '  pilar VARCHAR(100),'
                    '  pcn VARCHAR(20)'
                    ') DEFAULT CHARSET=utf8mb4'
                )
                cursor.execute(
                    'CREATE TABLE IF NOT EXISTS metas_gerente ('
                    '  id INT AUTO_INCREMENT PRIMARY KEY,'
                    '  valor DECIMAL(14,2),'
                    '  pdv VARCHAR(255),'
                    '  nome_gerente VARCHAR(255),'
                    '  mes_ref INT,'
                    '  ano_ref INT,'
                    '  unidade VARCHAR(20),'
                    '  pilar VARCHAR(100),'
                    '  hc INT'
                    ') DEFAULT CHARSET=utf8mb4'
                )

                cursor.execute(
                    'DELETE FROM metas WHERE mes_ref = %s AND ano_ref = %s',
                    (upload.month, upload.year),
                )
                cursor.executemany(
                    'INSERT INTO metas (valor, pdv, mes_ref, ano_ref, unidade, pilar, hc) VALUES (%s, %s, %s, %s, %s, %s, %s)',
                    rows,
                )

                cursor.execute(
                    'DELETE FROM metas_cn WHERE mes_ref = %s AND ano_ref = %s',
                    (upload.month, upload.year),
                )
                if cn_rows:
                    cursor.executemany(
                        'INSERT INTO metas_cn (valor, pdv, cn, mes_ref, ano_ref, unidade, pilar, pcn) '
                        'VALUES (%s, %s, %s, %s, %s, %s, %s, %s)',
                        cn_rows,
                    )

                cursor.execute(
                    'DELETE FROM metas_gerente WHERE mes_ref = %s AND ano_ref = %s',
                    (upload.month, upload.year),
                )
                if gerente_rows:
                    cursor.executemany(
                        'INSERT INTO metas_gerente (valor, pdv, nome_gerente, mes_ref, ano_ref, unidade, pilar, hc) '
                        'VALUES (%s, %s, %s, %s, %s, %s, %s, %s)',
                        gerente_rows,
                    )
            connection.commit()
        finally:
            connection.close()
    except Exception as exc:
        messages.error(request, f'Erro ao sincronizar metas no MySQL: {exc}')
        return redirect('power_bi:manage_goals')

    messages.success(
        request,
        f'Metas de {upload.month:02d}/{upload.year} sincronizadas com sucesso no banco MySQL '
        f'({len(rows)} linhas de loja, {len(cn_rows)} linhas de CN, {len(gerente_rows)} linhas de gerente).',
    )

    if not gerente_group_found:
        messages.warning(
            request,
            'Grupo GERENTES nao encontrado em /users/manage/groups/: a tabela metas_gerente ficou sem linhas.',
        )
    elif pdvs_sem_gerente:
        amostra = ', '.join(sorted(pdvs_sem_gerente)[:8])
        restantes = len(pdvs_sem_gerente) - 8
        if restantes > 0:
            amostra += f' e mais {restantes}'
        messages.warning(
            request,
            f'{len(pdvs_sem_gerente)} PDV(s) sem gerente vinculado ficaram fora de metas_gerente: {amostra}.',
        )

    return redirect('power_bi:manage_goals')
