from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse
from django.views.decorators.http import require_POST
from django.utils import timezone
from django.db.models import Q
from django.contrib import messages
from django.db import models

from .models import Activity
from .models_chat import TaskChat, TaskChatMessage, SupportChat, SupportChatMessage, SupportAgent, SupportCategory, SupportChatRating, SupportChatFile, SupportChatFile
from users.models import User


@login_required
def get_task_chat(request, activity_id):
    """Buscar mensagens do chat de uma tarefa"""
    activity = get_object_or_404(Activity, id=activity_id)
    
    # Verificar se o usuário pode acessar esta tarefa
    if not (activity.created_by == request.user or 
            activity.responsible_user == request.user or 
            request.user.hierarchy in ['ADMIN', 'SUPERADMIN']):
        return JsonResponse({'success': False, 'error': 'Acesso negado'})
    
    # Criar chat se não existir
    chat, created = TaskChat.objects.get_or_create(activity=activity)
    
    # Buscar mensagens
    messages = TaskChatMessage.objects.filter(chat=chat).select_related('user')
    
    messages_data = []
    for msg in messages:
        messages_data.append({
            'id': msg.id,
            'user': {
                'id': msg.user.id,
                'name': msg.user.get_full_name(),
                'avatar': msg.user.first_name[0].upper() if msg.user.first_name else 'U'
            },
            'message': msg.message,
            'created_at': msg.created_at.strftime('%d/%m/%Y %H:%M'),
            'is_own': msg.user == request.user
        })
    
    return JsonResponse({
        'success': True,
        'messages': messages_data,
        'activity': {
            'id': activity.id,
            'name': activity.name,
            'created_by': activity.created_by.get_full_name(),
            'responsible': activity.responsible_user.get_full_name() if activity.responsible_user else None
        }
    })


@login_required
@require_POST
def send_task_message(request, activity_id):
    """Enviar mensagem no chat de uma tarefa"""
    activity = get_object_or_404(Activity, id=activity_id)
    
    # Verificar se o usuário pode acessar esta tarefa
    if not (activity.created_by == request.user or 
            activity.responsible_user == request.user or 
            request.user.hierarchy in ['ADMIN', 'SUPERADMIN']):
        return JsonResponse({'success': False, 'error': 'Acesso negado'})
    
    message_text = request.POST.get('message', '').strip()
    if not message_text:
        return JsonResponse({'success': False, 'error': 'Mensagem não pode estar vazia'})
    
    # Criar chat se não existir
    chat, created = TaskChat.objects.get_or_create(activity=activity)
    
    # Criar mensagem
    message = TaskChatMessage.objects.create(
        chat=chat,
        user=request.user,
        message=message_text
    )
    
    return JsonResponse({
        'success': True,
        'message': {
            'id': message.id,
            'user': {
                'id': request.user.id,
                'name': request.user.get_full_name(),
                'avatar': request.user.first_name[0].upper() if request.user.first_name else 'U'
            },
            'message': message.message,
            'created_at': message.created_at.strftime('%d/%m/%Y %H:%M'),
            'is_own': True
        }
    })


@login_required
def support_chat_list(request):
    """Listar chats de suporte do usuário"""
    user_chats = SupportChat.objects.filter(user=request.user).order_by('-updated_at')
    
    # Se for agente de suporte ou supervisor+, mostrar todos os chats
    is_support_agent = hasattr(request.user, 'support_agent') and request.user.support_agent.is_active
    is_supervisor_or_higher = request.user.hierarchy in ['SUPERVISOR', 'ADMIN', 'SUPERADMIN', 'ADMINISTRATIVO'] or request.user.is_superuser
    
    if is_support_agent or is_supervisor_or_higher:
        all_chats = SupportChat.objects.all().order_by('-updated_at')
    else:
        all_chats = user_chats
    
    chats_data = []
    for chat in all_chats[:20]:  # Limitar a 20 chats
        last_message = chat.messages.last()
        queue_position = chat.get_queue_position() if hasattr(chat, 'get_queue_position') else None
        chats_data.append({
            'id': chat.id,
            'title': chat.title,
            'protocol': getattr(chat, 'protocol', None),
            'status': chat.get_status_display(),
            'status_code': chat.status,
            'priority': chat.get_priority_display(),
            'user': chat.user.get_full_name(),
            'assigned_to': chat.assigned_to.get_full_name() if chat.assigned_to else None,
            'last_message': last_message.message[:100] if last_message else 'Sem mensagens',
            'updated_at': chat.updated_at.strftime('%d/%m/%Y %H:%M'),
            'is_own': chat.user == request.user,
            'queue_position': queue_position
        })
    
    return JsonResponse({
        'success': True,
        'chats': chats_data,
        'is_support_agent': is_support_agent
    })


@login_required
def get_support_chat(request, chat_id):
    """Buscar mensagens de um chat de suporte"""
    chat = get_object_or_404(SupportChat, id=chat_id)
    
    # Verificar se o usuário pode acessar este chat
    is_support_agent = hasattr(request.user, 'support_agent') and request.user.support_agent.is_active
    is_supervisor_or_higher = request.user.hierarchy in ['SUPERVISOR', 'ADMIN', 'SUPERADMIN', 'ADMINISTRATIVO'] or request.user.is_superuser
    
    if not (chat.user == request.user or is_support_agent or is_supervisor_or_higher):
        return JsonResponse({'success': False, 'error': 'Acesso negado'})
    
    # Buscar mensagens
    messages = SupportChatMessage.objects.filter(chat=chat).select_related('user').prefetch_related('files')
    
    # Filtrar mensagens internas se não for agente de suporte ou supervisor+
    if not (is_support_agent or is_supervisor_or_higher):
        messages = messages.filter(is_internal=False)
    
    messages_data = []
    for msg in messages:
        msg_data = {
            'id': msg.id,
            'user': {
                'id': msg.user.id,
                'name': msg.user.get_full_name(),
                'avatar': msg.user.first_name[0].upper() if msg.user.first_name else 'U',
                'is_support': hasattr(msg.user, 'support_agent')
            },
            'message': msg.message,
            'is_internal': msg.is_internal,
            'created_at': msg.created_at.strftime('%d/%m/%Y %H:%M'),
            'is_own': msg.user == request.user,
            'files': []
        }
        
        # Adicionar arquivos
        for file in msg.files.all():
            msg_data['files'].append({
                'id': file.id,
                'type': file.file_type,
                'name': file.original_name,
                'url': file.file.url,
                'size': file.file_size
            })
        
        messages_data.append(msg_data)
    
    # Calcular posição na fila
    queue_position = chat.get_queue_position() if hasattr(chat, 'get_queue_position') else None
    
    return JsonResponse({
        'success': True,
        'messages': messages_data,
        'chat': {
            'id': chat.id,
            'title': chat.title,
            'protocol': getattr(chat, 'protocol', None),
            'status': chat.status,
            'get_status_display': chat.get_status_display(),
            'priority': chat.priority.lower() if chat.priority else 'media',
            'get_priority_display': chat.get_priority_display(),
            'queue_position': queue_position,
            'user': {
                'id': chat.user.id,
                'get_full_name': chat.user.get_full_name()
            },
            'sector': {
                'id': chat.sector.id,
                'name': chat.sector.name
            } if chat.sector else None,
            'category': {
                'id': chat.category.id,
                'name': chat.category.name
            } if chat.category else None,
            'assigned_to': {
                'id': chat.assigned_to.id,
                'get_full_name': chat.assigned_to.get_full_name()
            } if chat.assigned_to else None,
            'is_own': chat.user == request.user
        },
        'is_support_agent': is_support_agent
    })


@login_required
@require_POST
def create_support_chat(request):
    """Criar novo chat de suporte"""
    title = request.POST.get('title', '').strip()
    message = request.POST.get('message', '').strip()
    sector_id = request.POST.get('sector_id')
    category_id = request.POST.get('category_id')
    
    if not title or not message:
        return JsonResponse({'success': False, 'error': 'Título e mensagem são obrigatórios'})
    
    # Importar modelos necessários
    from users.models import Sector
    
    # Criar chat
    chat_data = {
        'user': request.user,
        'title': title,
        'status': 'AGUARDANDO',  # Começa na fila
    }
    
    if sector_id:
        try:
            sector = Sector.objects.get(id=sector_id)
            chat_data['sector'] = sector
        except Sector.DoesNotExist:
            pass
    
    if category_id:
        try:
            category = SupportCategory.objects.get(id=category_id)
            chat_data['category'] = category
        except SupportCategory.DoesNotExist:
            pass
    
    chat = SupportChat.objects.create(**chat_data)
    
    # Criar primeira mensagem
    SupportChatMessage.objects.create(
        chat=chat,
        user=request.user,
        message=message
    )
    
    # Calcular posição na fila
    queue_position = chat.get_queue_position() if hasattr(chat, 'get_queue_position') else None
    
    return JsonResponse({
        'success': True,
        'chat_id': chat.id,
        'protocol': getattr(chat, 'protocol', None),
        'queue_position': queue_position,
        'message': f'Chat de suporte criado! Protocolo: {getattr(chat, "protocol", "N/A")}'
    })


@login_required
@require_POST
def send_support_message(request, chat_id):
    """Enviar mensagem no chat de suporte"""
    import json
    
    chat = get_object_or_404(SupportChat, id=chat_id)
    
    # Verificar se o usuário pode acessar este chat
    is_support_agent = hasattr(request.user, 'support_agent') and request.user.support_agent.is_active
    is_supervisor_or_higher = request.user.hierarchy in ['SUPERVISOR', 'ADMIN', 'SUPERADMIN', 'ADMINISTRATIVO'] or request.user.is_superuser
    
    if not (chat.user == request.user or is_support_agent or is_supervisor_or_higher):
        return JsonResponse({'success': False, 'error': 'Acesso negado'})
    
    # Aceitar JSON ou form data
    if request.content_type == 'application/json':
        try:
            data = json.loads(request.body)
            message_text = data.get('message', '').strip()
            is_internal = data.get('is_internal') == True and (is_support_agent or is_supervisor_or_higher)
        except json.JSONDecodeError:
            return JsonResponse({'success': False, 'error': 'JSON inválido'})
    else:
        message_text = request.POST.get('message', '').strip()
        is_internal = request.POST.get('is_internal') == 'true' and (is_support_agent or is_supervisor_or_higher)
    
    if not message_text:
        return JsonResponse({'success': False, 'error': 'Mensagem não pode estar vazia'})
    
    # Criar mensagem
    message = SupportChatMessage.objects.create(
        chat=chat,
        user=request.user,
        message=message_text,
        is_internal=is_internal
    )
    
    # Atualizar status do chat se necessário
    if chat.status == 'ABERTO' and (is_support_agent or is_supervisor_or_higher):
        chat.status = 'EM_ANDAMENTO'
        chat.assigned_to = request.user
        chat.save()
    
    return JsonResponse({
        'success': True,
        'message': {
            'id': message.id,
            'user': {
                'id': request.user.id,
                'name': request.user.get_full_name(),
                'avatar': request.user.first_name[0].upper() if request.user.first_name else 'U',
                'is_support': hasattr(request.user, 'support_agent')
            },
            'message': message.message,
            'is_internal': message.is_internal,
            'created_at': message.created_at.strftime('%d/%m/%Y %H:%M'),
            'is_own': True,
            'files': []
        }
    })


@login_required
def get_sectors(request):
    """Buscar TODOS os setores para o formulário de suporte - qualquer usuário pode abrir chat para qualquer setor"""
    from users.models import Sector
    
    # Retornar todos os setores - usuários devem poder abrir chat de suporte para qualquer setor
    sectors = Sector.objects.all().order_by('name')
    
    sectors_data = []
    for s in sectors:
        sectors_data.append({'id': s.id, 'name': s.name})
    
    # Ordenar por nome
    sectors_data.sort(key=lambda x: x['name'])
    
    return JsonResponse({'success': True, 'sectors': sectors_data})


@login_required
def get_user_sectors(request):
    """Buscar apenas os setores do usuário logado - para gerenciamento de agentes/categorias"""
    from users.models import Sector
    
    # Retornar apenas setores do supervisor
    if request.user.is_superuser:
        sectors = Sector.objects.all().order_by('name')
    else:
        user_sectors_list = list(request.user.sectors.all())
        if request.user.sector:
            user_sectors_list.append(request.user.sector)
        sectors = Sector.objects.filter(id__in=[s.id for s in user_sectors_list]).order_by('name')
    
    sectors_data = [{'id': s.id, 'name': s.name} for s in sectors]
    
    return JsonResponse({'success': True, 'sectors': sectors_data})


@login_required
def get_sector_categories(request, sector_id):
    """Buscar categorias de um setor específico"""
    from .models_chat import SupportCategory
    from users.models import Sector
    
    # Verificar se é o setor Ilha de Qualidade
    try:
        sector = Sector.objects.get(id=sector_id)
    except Sector.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Setor não encontrado'})
    
    categories = SupportCategory.objects.filter(
        sector_id=sector_id, 
        is_active=True
    ).order_by('name')
    
    categories_data = [{'id': c.id, 'name': c.name, 'description': c.description} for c in categories]
    
    return JsonResponse({'success': True, 'categories': categories_data})


@login_required
@require_POST
def upload_chat_file(request, chat_id):
    """Upload de arquivos no chat de suporte"""
    import json
    from .models_chat import SupportChatFile
    
    chat = get_object_or_404(SupportChat, id=chat_id)
    
    # Verificar se pode enviar arquivos neste chat
    if chat.user != request.user and not hasattr(request.user, 'support_agent'):
        return JsonResponse({'success': False, 'error': 'Acesso negado'})
    
    if 'file' not in request.FILES:
        return JsonResponse({'success': False, 'error': 'Nenhum arquivo enviado'})
    
    uploaded_file = request.FILES['file']
    
    # Determinar tipo do arquivo
    file_type = 'DOCUMENT'
    if uploaded_file.content_type.startswith('image/'):
        file_type = 'IMAGE'
    elif uploaded_file.content_type.startswith('video/'):
        file_type = 'VIDEO'
    elif uploaded_file.content_type.startswith('audio/'):
        file_type = 'AUDIO'
    
    # Criar mensagem primeiro
    message_text = request.POST.get('message', f'Arquivo enviado: {uploaded_file.name}')
    message = SupportChatMessage.objects.create(
        chat=chat,
        user=request.user,
        message=message_text
    )
    
    # Criar arquivo
    chat_file = SupportChatFile.objects.create(
        chat=chat,
        message=message,
        file=uploaded_file,
        file_type=file_type,
        original_name=uploaded_file.name,
        file_size=uploaded_file.size
    )
    
    return JsonResponse({
        'success': True,
        'file': {
            'id': chat_file.id,
            'type': file_type,
            'name': chat_file.original_name,
            'size': chat_file.file_size,
            'url': chat_file.file.url
        }
    })


@login_required
@require_POST
def rate_support_chat(request, chat_id):
    """Avaliar atendimento de suporte"""
    import json
    from .models_chat import SupportChatRating
    
    chat = get_object_or_404(SupportChat, id=chat_id)
    
    # Apenas o dono do chat pode avaliar
    if chat.user != request.user:
        return JsonResponse({'success': False, 'error': 'Acesso negado'})
    
    data = json.loads(request.body)
    rating_value = data.get('rating')
    feedback = data.get('feedback', '')
    
    if not rating_value or rating_value not in [1, 2, 3, 4, 5]:
        return JsonResponse({'success': False, 'error': 'Avaliação inválida'})
    
    # Criar ou atualizar avaliação
    rating, created = SupportChatRating.objects.get_or_create(
        chat=chat,
        defaults={'rating': rating_value, 'feedback': feedback}
    )
    
    if not created:
        rating.rating = rating_value
        rating.feedback = feedback
        rating.save()
    
    # Se a avaliação for ruim (1 ou 2), reabrir o chat
    if rating_value <= 2:
        chat.status = 'ABERTO'
        chat.assigned_to = None
        chat.save()
        
        # Criar mensagem automática
        SupportChatMessage.objects.create(
            chat=chat,
            user=request.user,
            message=f"Chat reaberto devido à avaliação negativa. Feedback: {feedback}" if feedback else "Chat reaberto devido à avaliação negativa."
        )
        
        return JsonResponse({
            'success': True, 
            'message': 'Avaliação registrada. O chat foi reaberto para nova análise.',
            'reopened': True
        })
    else:
        chat.status = 'FECHADO'
        chat.closed_at = timezone.now()
        chat.save()
        
        return JsonResponse({
            'success': True, 
            'message': 'Obrigado pela sua avaliação! O chat foi finalizado.',
            'closed': True
        })


@login_required
def support_admin_dashboard(request):
    """Dashboard administrativo do suporte"""
    from django.db.models import Count, Q, Avg
    from .models_chat import SupportChat, SupportChatRating
    from users.models import Sector
    
    # Verificar se é agente de suporte
    is_support_agent = SupportAgent.objects.filter(
        user=request.user,
        is_active=True
    ).exists()
    
    # Verificar permissões de acesso: SUPERVISOR ou maior, ou agente de suporte
    if not (request.user.hierarchy in ['SUPERVISOR', 'ADMIN', 'SUPERADMIN', 'ADMINISTRATIVO'] or is_support_agent or request.user.is_superuser):
        return JsonResponse({'success': False, 'error': 'Acesso negado'})
    
    # Obter setores do usuário
    if request.user.is_superuser:
        user_sectors = Sector.objects.all()
        base_filter = Q()  # Superuser vê tudo
    else:
        user_sectors_list = list(request.user.sectors.all())
        if request.user.sector:
            user_sectors_list.append(request.user.sector)
        user_sectors = Sector.objects.filter(id__in=[s.id for s in user_sectors_list])
        base_filter = Q(sector__in=user_sectors)
    
    # Estatísticas filtradas por setor
    total_chats = SupportChat.objects.filter(base_filter).count()
    open_chats = SupportChat.objects.filter(base_filter, status__in=['AGUARDANDO', 'ABERTO']).count()
    in_progress_chats = SupportChat.objects.filter(base_filter, status='EM_ANDAMENTO').count()
    resolved_chats = SupportChat.objects.filter(base_filter, status='RESOLVIDO').count()
    
    # Chats por prioridade (filtrados)
    priority_stats = SupportChat.objects.filter(base_filter).values('priority').annotate(count=Count('id'))
    
    # Avaliação média (filtrada)
    avg_rating = SupportChatRating.objects.filter(chat__in=SupportChat.objects.filter(base_filter)).aggregate(avg=Avg('rating'))['avg'] or 0
    
    # Chats recentes (filtrados por setor)
    recent_chats = SupportChat.objects.filter(base_filter).select_related('user', 'assigned_to', 'sector', 'category').order_by('-created_at')[:20]
    
    # Agentes de suporte (filtrados por setor)
    if request.user.is_superuser:
        agents_qs = SupportAgent.objects.filter(is_active=True)
    else:
        agents_qs = SupportAgent.objects.filter(is_active=True, sectors__in=user_sectors).distinct()
    
    agents_qs = agents_qs.select_related('user').prefetch_related('sectors')
    
    # Serializar agentes para JSON
    agents_data = []
    for agent in agents_qs:
        agents_data.append({
            'id': agent.id,
            'user': {
                'id': agent.user.id,
                'name': agent.user.get_full_name(),
                'email': agent.user.email
            },
            'can_assign_tickets': agent.can_assign_tickets,
            'sectors': [{'id': s.id, 'name': s.name} for s in agent.sectors.all()]
        })
    
    # Serializar chats para JSON (se for AJAX request)
    recent_chats_data = []
    for chat in recent_chats:
        recent_chats_data.append({
            'id': chat.id,
            'title': chat.title,
            'status': chat.status,
            'get_status_display': chat.get_status_display(),
            'priority': chat.priority.lower() if chat.priority else 'media',
            'get_priority_display': chat.get_priority_display(),
            'user': {
                'id': chat.user.id,
                'get_full_name': chat.user.get_full_name()
            },
            'sector': {
                'id': chat.sector.id,
                'name': chat.sector.name
            } if chat.sector else None,
            'category': {
                'id': chat.category.id,
                'name': chat.category.name
            } if chat.category else None,
            'assigned_to': {
                'id': chat.assigned_to.id,
                'get_full_name': chat.assigned_to.get_full_name()
            } if chat.assigned_to else None,
            'created_at': chat.created_at.isoformat()
        })
    
    context = {
        'stats': {
            'total': total_chats,
            'open': open_chats,
            'in_progress': in_progress_chats,
            'resolved': resolved_chats,
            'avg_rating': round(avg_rating, 1)
        },
        'priority_stats': list(priority_stats),
        'recent_chats': recent_chats_data,
        'agents': agents_data,
        'is_support_agent': is_support_agent,
        'user_sectors': [{'id': s.id, 'name': s.name} for s in user_sectors]
    }
    
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.GET.get('format') == 'json':
        return JsonResponse(context, safe=False)
    
    return render(request, 'support/admin_dashboard.html', context)


@login_required
def manage_support_categories(request):
    """Gerenciar categorias de suporte"""
    # Verificar permissões de admin: SUPERVISOR ou maior
    if not (request.user.hierarchy in ['SUPERVISOR', 'ADMIN', 'SUPERADMIN', 'ADMINISTRATIVO'] or request.user.is_superuser):
        if request.method == 'POST':
            return JsonResponse({'success': False, 'error': 'Acesso negado'})
        else:
            messages.error(request, 'Você não tem permissão para acessar esta área.')
            return redirect('core:home')
    
    from .models_chat import SupportCategory
    from users.models import Sector
    
    # Obter setores do usuário
    if request.user.is_superuser:
        user_sectors = Sector.objects.all()
    else:
        user_sectors = list(request.user.sectors.all())
        if request.user.sector:
            user_sectors.append(request.user.sector)
        user_sectors = Sector.objects.filter(id__in=[s.id for s in user_sectors])
    
    if request.method == 'POST':
        import json
        data = json.loads(request.body)
        
        if data.get('action') == 'create':
            sector = get_object_or_404(Sector, id=data['sector_id'])
            
            # Verificar se o usuário pode criar categoria neste setor
            if not request.user.is_superuser and sector not in user_sectors:
                return JsonResponse({'success': False, 'error': 'Você não pode criar categorias neste setor'})
            
            category = SupportCategory.objects.create(
                name=data['name'],
                sector=sector,
                description=data.get('description', '')
            )
            return JsonResponse({'success': True, 'category_id': category.id})
        
        elif data.get('action') == 'update':
            category = get_object_or_404(SupportCategory, id=data['category_id'])
            
            # Verificar se o usuário pode editar categoria deste setor
            if not request.user.is_superuser and category.sector not in user_sectors:
                return JsonResponse({'success': False, 'error': 'Você não pode editar categorias deste setor'})
            
            category.name = data['name']
            category.description = data.get('description', '')
            category.save()
            return JsonResponse({'success': True})
        
        elif data.get('action') == 'delete':
            category = get_object_or_404(SupportCategory, id=data['category_id'])
            
            # Verificar se o usuário pode deletar categoria deste setor
            if not request.user.is_superuser and category.sector not in user_sectors:
                return JsonResponse({'success': False, 'error': 'Você não pode deletar categorias deste setor'})
            
            category.delete()
            return JsonResponse({'success': True})
    
    # Filtrar categorias pelos setores do usuário
    categories = SupportCategory.objects.filter(sector__in=user_sectors).select_related('sector').order_by('sector__name', 'name')
    
    return render(request, 'support/manage_categories.html', {
        'categories': categories,
        'sectors': user_sectors.order_by('name')
    })


@login_required
def manage_support_agents(request):
    """Gerenciar agentes de suporte"""
    # Verificar permissões de admin: SUPERVISOR ou maior
    if not (request.user.hierarchy in ['SUPERVISOR', 'ADMIN', 'SUPERADMIN', 'ADMINISTRATIVO'] or request.user.is_superuser):
        if request.method == 'POST':
            return JsonResponse({'success': False, 'error': 'Acesso negado'})
        else:
            messages.error(request, 'Você não tem permissão para acessar esta área.')
            return redirect('core:home')
    
    from .models_chat import SupportAgent
    from users.models import Sector
    
    # Obter setores do usuário
    if request.user.is_superuser:
        user_sectors = Sector.objects.all()
    else:
        user_sectors = list(request.user.sectors.all())
        if request.user.sector:
            user_sectors.append(request.user.sector)
        user_sectors = Sector.objects.filter(id__in=[s.id for s in user_sectors])
    
    if request.method == 'POST':
        import json
        data = json.loads(request.body)
        
        if data.get('action') == 'create':
            user = get_object_or_404(User, id=data['user_id'])
            sector_ids = data.get('sector_ids', [])
            
            # Verificar se o supervisor pode atribuir para esses setores
            if not request.user.is_superuser:
                user_sector_ids = [s.id for s in user_sectors]
                if not all(sid in user_sector_ids for sid in sector_ids):
                    return JsonResponse({'success': False, 'error': 'Você não pode atribuir agentes para setores que não gerencia'})
            
            agent, created = SupportAgent.objects.get_or_create(
                user=user,
                defaults={'can_assign_tickets': data.get('can_assign_tickets', False)}
            )
            
            # Adicionar setores
            if sector_ids:
                agent.sectors.set(sector_ids)
            
            return JsonResponse({'success': True, 'created': created})
        
        elif data.get('action') == 'update':
            agent = get_object_or_404(SupportAgent, id=data['agent_id'])
            sector_ids = data.get('sector_ids', [])
            
            # Verificar se o supervisor pode editar agente desses setores
            if not request.user.is_superuser:
                user_sector_ids = [s.id for s in user_sectors]
                if not all(sid in user_sector_ids for sid in sector_ids):
                    return JsonResponse({'success': False, 'error': 'Você não pode editar agentes de setores que não gerencia'})
            
            agent.can_assign_tickets = data.get('can_assign_tickets', False)
            agent.is_active = data.get('is_active', True)
            agent.save()
            
            # Atualizar setores
            if sector_ids:
                agent.sectors.set(sector_ids)
            
            return JsonResponse({'success': True})
        
        elif data.get('action') == 'delete':
            agent = get_object_or_404(SupportAgent, id=data['agent_id'])
            
            # Verificar se o supervisor pode deletar agente
            if not request.user.is_superuser:
                agent_sector_ids = [s.id for s in agent.sectors.all()]
                user_sector_ids = [s.id for s in user_sectors]
                if not any(sid in user_sector_ids for sid in agent_sector_ids):
                    return JsonResponse({'success': False, 'error': 'Você não pode deletar agentes de setores que não gerencia'})
            
            agent.delete()
            return JsonResponse({'success': True})
    
    # Filtrar agentes que atendem os setores do supervisor
    if request.user.is_superuser:
        agents = SupportAgent.objects.all()
    else:
        agents = SupportAgent.objects.filter(sectors__in=user_sectors).distinct()
    
    agents = agents.select_related('user').prefetch_related('sectors').order_by('user__first_name')
    
    # Usuários disponíveis dos setores do supervisor
    available_users = User.objects.filter(
        is_active=True,
        hierarchy__in=['ADMIN', 'SUPERVISOR', 'FUNCIONARIO']
    ).exclude(support_agent__isnull=False).order_by('first_name')
    
    # Filtrar usuários pelos setores se não for superuser
    if not request.user.is_superuser:
        available_users = available_users.filter(
            Q(sectors__in=user_sectors) | Q(sector__in=user_sectors)
        ).distinct()
    
    return render(request, 'support/manage_agents.html', {
        'agents': agents,
        'available_users': available_users,
        'user_sectors': user_sectors.order_by('name')
    })


def support_admin_template(request):
    """Template do dashboard administrativo"""
    # Permitir acesso para SUPERVISOR ou hierarquia maior, além de agentes de suporte
    is_supervisor_or_higher = request.user.hierarchy in ['SUPERVISOR', 'ADMIN', 'SUPERADMIN', 'ADMINISTRATIVO'] or request.user.is_superuser
    
    # Verificar se é agente de suporte
    is_support_agent = SupportAgent.objects.filter(
        user=request.user, 
        is_active=True
    ).exists()
    
    # Permitir acesso se for supervisor+ OU agente de suporte
    if not (is_supervisor_or_higher or is_support_agent):
        messages.error(request, 'Você não tem permissão para acessar esta área.')
        return redirect('core:home')
    
    # Obter setores do usuário
    from users.models import Sector
    if request.user.is_superuser:
        user_sectors = Sector.objects.all()
    else:
        user_sectors_list = list(request.user.sectors.all())
        if request.user.sector:
            user_sectors_list.append(request.user.sector)
        user_sectors = Sector.objects.filter(id__in=[s.id for s in user_sectors_list])
    
    # Filtrar chats por setores do usuário
    if request.user.is_superuser:
        chats_filter = Q()
    else:
        chats_filter = Q(sector__in=user_sectors)
    
    # Estatísticas básicas para o template (filtradas por setor)
    stats = {
        'total': SupportChat.objects.filter(chats_filter).count(),
        'open': SupportChat.objects.filter(chats_filter, status__in=['AGUARDANDO', 'ABERTO']).count(),
        'in_progress': SupportChat.objects.filter(chats_filter, status='EM_ANDAMENTO').count(),
        'resolved': SupportChat.objects.filter(chats_filter, status='RESOLVIDO').count(),
        'avg_rating': round(SupportChatRating.objects.filter(
            chat__in=SupportChat.objects.filter(chats_filter)
        ).aggregate(avg_rating=models.Avg('rating'))['avg_rating'] or 0, 1)
    }
    
    # Lista de setores do usuário para o template
    user_sectors_data = [
        {'id': sector.id, 'name': sector.name} 
        for sector in user_sectors
    ]
    
    return render(request, 'support/admin_dashboard.html', {
        'stats': stats,
        'is_support_agent': is_support_agent,
        'user_sectors': user_sectors_data
    })


def assign_chat_to_agent(request, chat_id):
    """Atribui um chat a um agente"""
    if not request.user.is_staff:
        return JsonResponse({'error': 'Acesso negado'}, status=403)
    
    if request.method == 'POST':
        try:
            chat = SupportChat.objects.get(id=chat_id)
            chat.assigned_to = request.user
            chat.status = 'EM_ANDAMENTO'
            chat.save()
            
            return JsonResponse({'success': True})
        except SupportChat.DoesNotExist:
            return JsonResponse({'error': 'Chat não encontrado'}, status=404)
    
    return JsonResponse({'error': 'Método inválido'}, status=405)


def support_metrics(request):
    """Métricas de suporte para supervisores"""
    # Permitir acesso para SUPERVISOR ou hierarquia maior
    if not (request.user.is_superuser or request.user.hierarchy in ['SUPERADMIN', 'ADMIN', 'SUPERVISOR', 'ADMINISTRATIVO']):
        messages.error(request, 'Você não tem permissão para acessar esta área.')
        return redirect('core:home')
    
    from datetime import datetime, timedelta
    from django.db.models import Count, Avg, Q
    import json
    
    # Período selecionado (padrão: 30 dias)
    period_days = int(request.GET.get('period', 30))
    start_date = datetime.now() - timedelta(days=period_days)
    
    # Estatísticas principais
    total_tickets = SupportChat.objects.filter(created_at__gte=start_date).count()
    resolved_tickets = SupportChat.objects.filter(
        created_at__gte=start_date, 
        status='FECHADO'
    ).count()
    
    resolution_rate = round((resolved_tickets / total_tickets * 100) if total_tickets > 0 else 0, 1)
    
    # Tempo médio de resposta (simulado - seria calculado com base nas mensagens)
    avg_response_time = round(4.5, 1)  # Em horas
    
    # Avaliação média
    avg_rating = round(SupportChatRating.objects.filter(
        chat__created_at__gte=start_date
    ).aggregate(avg_rating=Avg('rating'))['avg_rating'] or 0, 1)
    
    # Dados para gráficos (simulados para demonstração)
    daily_labels = []
    daily_tickets = []
    daily_resolved = []
    
    for i in range(7):  # Últimos 7 dias
        date = datetime.now() - timedelta(days=i)
        daily_labels.insert(0, date.strftime('%d/%m'))
        daily_tickets.insert(0, SupportChat.objects.filter(
            created_at__date=date.date()
        ).count())
        daily_resolved.insert(0, SupportChat.objects.filter(
            created_at__date=date.date(),
            status='FECHADO'
        ).count())
    
    # Distribuição por status
    status_data = [
        SupportChat.objects.filter(status__in=['AGUARDANDO', 'ABERTO']).count(),
        SupportChat.objects.filter(status='EM_ANDAMENTO').count(),
        SupportChat.objects.filter(status='RESOLVIDO').count(),
        SupportChat.objects.filter(status='FECHADO').count(),
    ]
    
    # Top categorias
    top_categories = SupportCategory.objects.annotate(
        count=Count('supportchat')
    ).order_by('-count')[:5]
    
    categories_data = []
    total_cats = sum(cat.count for cat in top_categories)
    for cat in top_categories:
        percentage = round((cat.count / total_cats * 100) if total_cats > 0 else 0, 1)
        categories_data.append({
            'name': cat.name,
            'count': cat.count,
            'percentage': percentage
        })
    
    # Performance dos agentes (simulado)
    from django.contrib.auth import get_user_model
    User = get_user_model()
    
    agents = User.objects.filter(is_staff=True)[:5]
    agent_performance = []
    for agent in agents:
        tickets_resolved = SupportChat.objects.filter(
            assigned_to=agent,
            status='FECHADO',
            created_at__gte=start_date
        ).count()
        
        avg_agent_rating = SupportChatRating.objects.filter(
            chat__assigned_to=agent,
            chat__created_at__gte=start_date
        ).aggregate(avg_rating=Avg('rating'))['avg_rating'] or 0
        
        agent_performance.append({
            'name': agent.get_full_name(),
            'tickets_resolved': tickets_resolved,
            'avg_rating': round(avg_agent_rating, 1),
            'avg_time': round(3.2 + (agent.id % 3), 1)  # Simulado
        })
    
    # Tickets recentes
    recent_tickets = SupportChat.objects.select_related(
        'user', 'assigned_to'
    ).order_by('-created_at')[:10]
    
    metrics = {
        'total_tickets': total_tickets,
        'tickets_growth': 15,  # Simulado
        'avg_response_time': avg_response_time,
        'time_improvement': 8,  # Simulado
        'resolution_rate': resolution_rate,
        'resolution_improvement': 12,  # Simulado
        'avg_rating': avg_rating,
        'daily_labels': json.dumps(daily_labels),
        'daily_tickets': json.dumps(daily_tickets),
        'daily_resolved': json.dumps(daily_resolved),
        'status_labels': json.dumps(['Abertos', 'Em Andamento', 'Resolvidos', 'Fechados']),
        'status_data': json.dumps(status_data),
        'top_categories': categories_data,
        'agent_performance': agent_performance,
        'recent_tickets': recent_tickets
    }
    
    if request.headers.get('Accept') == 'application/json':
        return JsonResponse(metrics)
    
    return render(request, 'support/metrics_dashboard.html', {
        'metrics': metrics
    })


def export_metrics_report(request):
    """Exporta relatório de métricas em Excel"""
    from datetime import datetime, timedelta
    from django.db.models import Count, Avg, Q
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from django.http import HttpResponse
    
    # Permitir acesso para SUPERVISOR ou maior
    if not (request.user.is_superuser or request.user.hierarchy in ['SUPERADMIN', 'ADMIN', 'SUPERVISOR', 'ADMINISTRATIVO']):
        return JsonResponse({'error': 'Acesso negado'}, status=403)
    
    # Período selecionado (padrão: 30 dias)
    period_days = int(request.GET.get('period', 30))
    start_date = datetime.now() - timedelta(days=period_days)
    
    # Criar workbook
    wb = Workbook()
    
    # ========== ABA 1: Resumo Geral ==========
    ws_resumo = wb.active
    ws_resumo.title = 'Resumo Geral'
    
    # Estilos
    header_font = Font(bold=True, color='FFFFFF', size=12)
    header_fill = PatternFill(start_color='FF6B35', end_color='FF6B35', fill_type='solid')
    title_font = Font(bold=True, size=14)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Título
    ws_resumo['A1'] = f'Relatório de Métricas de Suporte - Últimos {period_days} dias'
    ws_resumo['A1'].font = title_font
    ws_resumo.merge_cells('A1:E1')
    ws_resumo['A2'] = f'Gerado em: {datetime.now().strftime("%d/%m/%Y %H:%M")}'
    
    # Estatísticas principais
    total_tickets = SupportChat.objects.filter(created_at__gte=start_date).count()
    resolved_tickets = SupportChat.objects.filter(created_at__gte=start_date, status='FECHADO').count()
    resolution_rate = round((resolved_tickets / total_tickets * 100) if total_tickets > 0 else 0, 1)
    avg_rating = round(SupportChatRating.objects.filter(
        chat__created_at__gte=start_date
    ).aggregate(avg_rating=Avg('rating'))['avg_rating'] or 0, 1)
    
    # Headers das métricas
    ws_resumo['A4'] = 'Métrica'
    ws_resumo['B4'] = 'Valor'
    for col in ['A', 'B']:
        ws_resumo[f'{col}4'].font = header_font
        ws_resumo[f'{col}4'].fill = header_fill
        ws_resumo[f'{col}4'].border = border
    
    # Dados das métricas
    metricas = [
        ('Total de Atendimentos', total_tickets),
        ('Atendimentos Resolvidos', resolved_tickets),
        ('Taxa de Resolução', f'{resolution_rate}%'),
        ('Avaliação Média', f'{avg_rating}/5'),
    ]
    
    for i, (nome, valor) in enumerate(metricas, start=5):
        ws_resumo[f'A{i}'] = nome
        ws_resumo[f'B{i}'] = valor
        ws_resumo[f'A{i}'].border = border
        ws_resumo[f'B{i}'].border = border
    
    # Ajustar larguras
    ws_resumo.column_dimensions['A'].width = 30
    ws_resumo.column_dimensions['B'].width = 20
    
    # ========== ABA 2: Por Categoria ==========
    ws_cat = wb.create_sheet('Por Categoria')
    
    # Headers
    headers_cat = ['Categoria', 'Setor', 'Total', 'Abertos', 'Em Andamento', 'Resolvidos', 'Fechados']
    for col, header in enumerate(headers_cat, start=1):
        cell = ws_cat.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
    
    # Dados por categoria
    categories = SupportCategory.objects.annotate(
        total=Count('support_chats', filter=Q(support_chats__created_at__gte=start_date)),
        abertos=Count('support_chats', filter=Q(support_chats__created_at__gte=start_date, support_chats__status__in=['AGUARDANDO', 'ABERTO'])),
        em_andamento=Count('support_chats', filter=Q(support_chats__created_at__gte=start_date, support_chats__status='EM_ANDAMENTO')),
        resolvidos=Count('support_chats', filter=Q(support_chats__created_at__gte=start_date, support_chats__status='RESOLVIDO')),
        fechados=Count('support_chats', filter=Q(support_chats__created_at__gte=start_date, support_chats__status='FECHADO'))
    ).select_related('sector').order_by('-total')
    
    for row, cat in enumerate(categories, start=2):
        ws_cat.cell(row=row, column=1, value=cat.name).border = border
        ws_cat.cell(row=row, column=2, value=cat.sector.name if cat.sector else 'N/A').border = border
        ws_cat.cell(row=row, column=3, value=cat.total).border = border
        ws_cat.cell(row=row, column=4, value=cat.abertos).border = border
        ws_cat.cell(row=row, column=5, value=cat.em_andamento).border = border
        ws_cat.cell(row=row, column=6, value=cat.resolvidos).border = border
        ws_cat.cell(row=row, column=7, value=cat.fechados).border = border
    
    # Ajustar larguras
    for col in range(1, 8):
        ws_cat.column_dimensions[get_column_letter(col)].width = 18
    
    # ========== ABA 3: Por Agente ==========
    ws_agente = wb.create_sheet('Por Agente')
    
    # Headers
    headers_agente = ['Agente', 'Total Atendimentos', 'Resolvidos', 'Taxa de Resolução', 'Avaliação Média']
    for col, header in enumerate(headers_agente, start=1):
        cell = ws_agente.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
    
    # Dados por agente
    from django.contrib.auth import get_user_model
    User = get_user_model()
    
    agents = SupportAgent.objects.filter(is_active=True).select_related('user')
    row_num = 2
    for agent in agents:
        total = SupportChat.objects.filter(assigned_to=agent.user, created_at__gte=start_date).count()
        resolved = SupportChat.objects.filter(assigned_to=agent.user, created_at__gte=start_date, status='FECHADO').count()
        rate = round((resolved / total * 100) if total > 0 else 0, 1)
        avg = SupportChatRating.objects.filter(
            chat__assigned_to=agent.user,
            chat__created_at__gte=start_date
        ).aggregate(avg_rating=Avg('rating'))['avg_rating'] or 0
        
        if total > 0:
            ws_agente.cell(row=row_num, column=1, value=agent.user.get_full_name()).border = border
            ws_agente.cell(row=row_num, column=2, value=total).border = border
            ws_agente.cell(row=row_num, column=3, value=resolved).border = border
            ws_agente.cell(row=row_num, column=4, value=f'{rate}%').border = border
            ws_agente.cell(row=row_num, column=5, value=round(avg, 1)).border = border
            row_num += 1
    
    # Ajustar larguras
    for col in range(1, 6):
        ws_agente.column_dimensions[get_column_letter(col)].width = 22
    
    # ========== ABA 4: Todos os Chats ==========
    ws_chats = wb.create_sheet('Todos os Chats')
    
    # Headers
    headers_chats = ['ID', 'Protocolo', 'Título', 'Cliente', 'Setor', 'Categoria', 'Status', 'Agente', 'Criado em', 'Fechado em']
    for col, header in enumerate(headers_chats, start=1):
        cell = ws_chats.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
    
    # Todos os chats do período
    chats = SupportChat.objects.filter(
        created_at__gte=start_date
    ).select_related('user', 'sector', 'category', 'assigned_to').order_by('-created_at')
    
    for row, chat in enumerate(chats, start=2):
        ws_chats.cell(row=row, column=1, value=chat.id).border = border
        ws_chats.cell(row=row, column=2, value=getattr(chat, 'protocol', '') or '').border = border
        ws_chats.cell(row=row, column=3, value=chat.title).border = border
        ws_chats.cell(row=row, column=4, value=chat.user.get_full_name()).border = border
        ws_chats.cell(row=row, column=5, value=chat.sector.name if chat.sector else 'N/A').border = border
        ws_chats.cell(row=row, column=6, value=chat.category.name if chat.category else 'N/A').border = border
        ws_chats.cell(row=row, column=7, value=chat.get_status_display()).border = border
        ws_chats.cell(row=row, column=8, value=chat.assigned_to.get_full_name() if chat.assigned_to else 'Não atribuído').border = border
        ws_chats.cell(row=row, column=9, value=chat.created_at.strftime('%d/%m/%Y %H:%M')).border = border
        ws_chats.cell(row=row, column=10, value=chat.closed_at.strftime('%d/%m/%Y %H:%M') if chat.closed_at else '').border = border
    
    # Ajustar larguras
    widths = [8, 18, 40, 25, 20, 25, 15, 25, 18, 18]
    for col, width in enumerate(widths, start=1):
        ws_chats.column_dimensions[get_column_letter(col)].width = width
    
    # Preparar resposta
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f'metricas_suporte_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response


@login_required
def get_support_categories_api(request):
    """API para listar categorias de suporte do setor do usuário"""
    import json
    from users.models import Sector
    
    if request.method == 'GET':
        # Obter setores do usuário
        if request.user.is_superuser:
            user_sectors = Sector.objects.all()
        else:
            user_sectors_list = list(request.user.sectors.all())
            if request.user.sector:
                user_sectors_list.append(request.user.sector)
            user_sectors = Sector.objects.filter(id__in=[s.id for s in user_sectors_list])
        
        # Mostrar categorias dos setores do usuário
        categories = SupportCategory.objects.filter(
            sector__in=user_sectors,
            is_active=True
        ).select_related('sector')
        
        categories_data = [{
            'id': cat.id,
            'name': cat.name,
            'description': cat.description,
            'sector': {
                'id': cat.sector.id,
                'name': cat.sector.name
            } if cat.sector else None,
            'is_active': cat.is_active
        } for cat in categories]
        
        return JsonResponse({'success': True, 'categories': categories_data})
    
    elif request.method == 'POST':
        # Permitir SUPERVISOR ou hierarquia maior gerenciar categorias
        if request.user.hierarchy not in ['SUPERVISOR', 'ADMIN', 'SUPERADMIN', 'ADMINISTRATIVO'] and not request.user.is_superuser:
            return JsonResponse({'success': False, 'error': 'Acesso negado'}, status=403)
        
        try:
            data = json.loads(request.body)
            action = data.get('action')
            
            # Obter setores do usuário para validações
            from users.models import Sector
            if request.user.is_superuser:
                user_sectors = Sector.objects.all()
            else:
                user_sectors_list = list(request.user.sectors.all())
                if request.user.sector:
                    user_sectors_list.append(request.user.sector)
                user_sectors = Sector.objects.filter(id__in=[s.id for s in user_sectors_list])
            
            if action == 'create':
                sector = get_object_or_404(Sector, id=data['sector_id'])
                
                # Validar que o setor pertence ao supervisor (exceto SUPERUSER)
                if not request.user.is_superuser and sector not in user_sectors:
                    return JsonResponse({'success': False, 'error': 'Você não pode criar categorias para este setor'}, status=403)
                
                category = SupportCategory.objects.create(
                    name=data['name'],
                    sector=sector,
                    description=data.get('description', ''),
                    is_active=True
                )
                
                return JsonResponse({
                    'success': True,
                    'category': {
                        'id': category.id,
                        'name': category.name,
                        'description': category.description,
                        'sector': {'id': sector.id, 'name': sector.name}
                    }
                })
            
            elif action == 'update':
                category = get_object_or_404(SupportCategory, id=data['category_id'])
                
                # Validar que o supervisor gerencia o setor da categoria
                if not request.user.is_superuser and category.sector not in user_sectors:
                    return JsonResponse({'success': False, 'error': 'Você não pode editar categorias de setores que não gerencia'}, status=403)
                
                category.name = data.get('name', category.name)
                category.description = data.get('description', category.description)
                category.is_active = data.get('is_active', category.is_active)
                category.save()
                
                return JsonResponse({'success': True})
            
            elif action == 'delete':
                category = get_object_or_404(SupportCategory, id=data['category_id'])
                
                # Validar que o supervisor gerencia o setor da categoria
                if not request.user.is_superuser and category.sector not in user_sectors:
                    return JsonResponse({'success': False, 'error': 'Você não pode deletar categorias de setores que não gerencia'}, status=403)
                
                category.delete()
                return JsonResponse({'success': True})
            
            return JsonResponse({'success': False, 'error': 'Ação inválida'})
            
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)}, status=400)
    
    return JsonResponse({'success': False, 'error': 'Método não permitido'}, status=405)


@login_required
def get_support_agents_api(request):
    """API para listar e gerenciar agentes de suporte"""
    import json
    from users.models import Sector
    
    if request.method == 'GET':
        # Obter setores do usuário
        if request.user.is_superuser:
            user_sectors = Sector.objects.all()
        else:
            user_sectors = list(request.user.sectors.all())
            if request.user.sector:
                user_sectors.append(request.user.sector)
        
        # Listar agentes dos setores do usuário
        if request.user.is_superuser:
            agents = SupportAgent.objects.filter(is_active=True)
        else:
            agents = SupportAgent.objects.filter(
                is_active=True,
                sectors__in=user_sectors
            ).distinct()
        
        agents = agents.select_related('user', 'user__sector').prefetch_related('sectors')
        
        agents_data = [{
            'id': agent.id,
            'user': {
                'id': agent.user.id,
                'name': agent.user.get_full_name(),
                'email': agent.user.email,
                'sector': agent.user.sector.name if agent.user.sector else 'N/A'
            },
            'sectors': [{'id': s.id, 'name': s.name} for s in agent.sectors.all()],
            'can_assign_tickets': agent.can_assign_tickets,
            'is_active': agent.is_active
        } for agent in agents]
        
        return JsonResponse({'success': True, 'agents': agents_data})
    
    elif request.method == 'POST':
        # Permitir SUPERVISOR ou hierarquia maior gerenciar agentes
        if request.user.hierarchy not in ['SUPERVISOR', 'ADMIN', 'SUPERADMIN', 'ADMINISTRATIVO'] and not request.user.is_superuser:
            return JsonResponse({'success': False, 'error': 'Acesso negado'}, status=403)
        
        try:
            data = json.loads(request.body)
            action = data.get('action')
            
            # Obter setores do usuário para validações
            from users.models import Sector
            if request.user.is_superuser:
                user_sectors = Sector.objects.all()
            else:
                user_sectors = list(request.user.sectors.all())
                if request.user.sector:
                    user_sectors.append(request.user.sector)
            
            if action == 'create':
                user = get_object_or_404(User, id=data['user_id'])
                
                # Verificar se já é agente
                if hasattr(user, 'support_agent'):
                    return JsonResponse({
                        'success': False,
                        'error': 'Usuário já é um agente de suporte'
                    }, status=400)
                
                # Validar que o usuário pertence aos setores do supervisor (exceto SUPERUSER)
                if not request.user.is_superuser:
                    user_in_same_sector = user.sectors.filter(id__in=[s.id for s in user_sectors]).exists()
                    if not user_in_same_sector:
                        return JsonResponse({
                            'success': False,
                            'error': 'Você só pode adicionar agentes dos seus setores'
                        }, status=403)
                
                agent = SupportAgent.objects.create(
                    user=user,
                    can_assign_tickets=data.get('can_assign_tickets', False),
                    is_active=True
                )
                
                return JsonResponse({
                    'success': True,
                    'agent': {
                        'id': agent.id,
                        'user': {
                            'id': user.id,
                            'name': user.get_full_name(),
                            'email': user.email,
                            'sector': user.sector.name if user.sector else 'N/A'
                        },
                        'can_assign_tickets': agent.can_assign_tickets,
                        'is_active': agent.is_active
                    }
                })
            
            elif action == 'update':
                agent = get_object_or_404(SupportAgent, id=data['agent_id'])
                
                # Validar se o supervisor gerencia os setores do agente
                if not request.user.is_superuser:
                    agent_sector_ids = [s.id for s in agent.sectors.all()]
                    user_sector_ids = [s.id for s in user_sectors]
                    if not any(sid in user_sector_ids for sid in agent_sector_ids):
                        return JsonResponse({'success': False, 'error': 'Você não pode editar agentes de setores que não gerencia'}, status=403)
                
                agent.can_assign_tickets = data.get('can_assign_tickets', agent.can_assign_tickets)
                agent.is_active = data.get('is_active', agent.is_active)
                agent.save()
                
                return JsonResponse({'success': True})
            
            elif action == 'delete':
                agent = get_object_or_404(SupportAgent, id=data['agent_id'])
                
                # Validar se o supervisor gerencia os setores do agente
                if not request.user.is_superuser:
                    agent_sector_ids = [s.id for s in agent.sectors.all()]
                    user_sector_ids = [s.id for s in user_sectors]
                    if not any(sid in user_sector_ids for sid in agent_sector_ids):
                        return JsonResponse({'success': False, 'error': 'Você não pode deletar agentes de setores que não gerencia'}, status=403)
                
                agent.delete()
                return JsonResponse({'success': True})
            
            return JsonResponse({'success': False, 'error': 'Ação inválida'})
            
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)}, status=400)
    
    return JsonResponse({'success': False, 'error': 'Método não permitido'}, status=405)


@login_required
def get_available_users_api(request):
    """API para listar usuários que podem se tornar agentes - filtrado por setores do supervisor"""
    # Permitir SUPERVISOR ou hierarquia maior
    if request.user.hierarchy not in ['SUPERVISOR', 'ADMIN', 'SUPERADMIN', 'ADMINISTRATIVO'] and not request.user.is_superuser:
        return JsonResponse({"success": False, "error": "Acesso negado"}, status=403)
    
    # Obter setores do usuário
    from users.models import Sector
    if request.user.is_superuser:
        user_sectors = Sector.objects.all()
    else:
        user_sectors = list(request.user.sectors.all())
        if request.user.sector:
            user_sectors.append(request.user.sector)
    
    # Filtrar por setor específico se fornecido
    sector_id = request.GET.get('sector_id')
    if sector_id:
        try:
            sector = Sector.objects.get(id=sector_id)
            # Verificar se o usuário tem permissão para este setor
            if not request.user.is_superuser and sector not in user_sectors:
                return JsonResponse({"success": False, "error": "Você não pode acessar este setor"}, status=403)
            
            # Buscar usuários deste setor específico
            available_users = User.objects.filter(
                Q(is_active=True) & (Q(sectors=sector) | Q(sector=sector))
            ).distinct().select_related("sector").order_by("first_name")
        except Sector.DoesNotExist:
            return JsonResponse({"success": False, "error": "Setor não encontrado"}, status=404)
    else:
        # Sem setor específico - mostrar todos dos setores do supervisor
        if request.user.is_superuser:
            available_users = User.objects.filter(
                is_active=True
            ).select_related("sector").order_by("first_name")
        else:
            available_users = User.objects.filter(
                Q(is_active=True) & (Q(sectors__in=user_sectors) | Q(sector__in=user_sectors))
            ).distinct().select_related("sector").order_by("first_name")
    
    users_data = [{
        "id": user.id,
        "name": user.get_full_name(),
        "email": user.email,
        "sector": user.sector.name if user.sector else "N/A",
        "sectors": [{"id": s.id, "name": s.name} for s in user.sectors.all()]
    } for user in available_users]
    
    return JsonResponse({"success": True, "users": users_data})


@login_required
@require_POST
def assign_chat_to_agent(request, chat_id):
    """Atribui um chat a um agente (Assumir Atendimento)"""
    # Verificar se é agente de suporte
    is_support_agent = hasattr(request.user, 'support_agent') and request.user.support_agent.is_active
    
    if not (is_support_agent or request.user.hierarchy in ['ADMIN', 'SUPERADMIN']):
        return JsonResponse({'success': False, 'error': 'Apenas agentes de suporte podem assumir atendimentos'}, status=403)
    
    try:
        chat = get_object_or_404(SupportChat, id=chat_id)
        
        # Validar que o chat pertence a um setor do agente (exceto SUPERADMIN)
        if request.user.hierarchy != 'SUPERADMIN':
            user_sectors = request.user.sectors.all()
            if chat.sector not in user_sectors:
                return JsonResponse({
                    'success': False, 
                    'error': 'Você só pode assumir atendimentos do seu setor'
                }, status=403)
        
        # Verificar se já está atribuído
        if chat.assigned_to and chat.assigned_to != request.user:
            return JsonResponse({
                'success': False,
                'error': f'Este chat já está sendo atendido por {chat.assigned_to.get_full_name()}'
            }, status=400)
        
        # Atribuir chat ao agente
        chat.assigned_to = request.user
        
        # Atualizar status se ainda estiver ABERTO
        if chat.status == 'ABERTO':
            chat.status = 'EM_ANDAMENTO'
        
        chat.save()
        
        # Criar mensagem automática
        SupportChatMessage.objects.create(
            chat=chat,
            user=request.user,
            message=f'Atendimento assumido por {request.user.get_full_name()}',
            is_internal=True
        )
        
        return JsonResponse({
            'success': True,
            'message': 'Atendimento assumido com sucesso',
            'assigned_to': request.user.get_full_name(),
            'status': chat.get_status_display()
        })
        
    except Exception as e:
        import traceback
        print(f"Erro ao assumir atendimento: {traceback.format_exc()}")
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@login_required
@require_POST
def resolve_support_chat(request, chat_id):
    """Marca um chat como resolvido"""
    try:
        chat = get_object_or_404(SupportChat, id=chat_id)
        
        # Apenas o atendente responsável pode marcar como resolvido
        if chat.assigned_to != request.user and request.user.hierarchy != 'SUPERADMIN':
            return JsonResponse({
                'success': False,
                'error': 'Apenas o atendente responsável pode marcar como resolvido'
            }, status=403)
        
        chat.status = 'RESOLVIDO'
        chat.save()
        
        # Criar mensagem automática
        SupportChatMessage.objects.create(
            chat=chat,
            user=request.user,
            message='Ticket marcado como resolvido. Aguardando confirmação do cliente.',
            is_internal=True
        )
        
        return JsonResponse({
            'success': True,
            'message': 'Chat marcado como resolvido',
            'status': chat.get_status_display()
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@login_required
@require_POST
def close_support_chat(request, chat_id):
    """Fecha um chat de suporte definitivamente"""
    try:
        chat = get_object_or_404(SupportChat, id=chat_id)
        
        # Apenas o atendente responsável ou admin pode fechar
        if chat.assigned_to != request.user and request.user.hierarchy != 'SUPERADMIN':
            return JsonResponse({
                'success': False,
                'error': 'Apenas o atendente responsável pode fechar o ticket'
            }, status=403)
        
        chat.close_chat()  # Usa o método do model que define status e closed_at
        
        # Criar mensagem automática
        SupportChatMessage.objects.create(
            chat=chat,
            user=request.user,
            message='Ticket fechado.',
            is_internal=True
        )
        
        return JsonResponse({
            'success': True,
            'message': 'Chat fechado com sucesso',
            'status': chat.get_status_display()
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@login_required
@require_POST
def delete_support_chat(request, chat_id):
    """Exclui um chat de suporte (apenas para SUPERVISOR ou maior)"""
    try:
        chat = get_object_or_404(SupportChat, id=chat_id)
        
        # Verificar permissão: SUPERVISOR ou maior
        is_supervisor_or_higher = request.user.hierarchy in ['SUPERVISOR', 'ADMIN', 'SUPERADMIN', 'ADMINISTRATIVO'] or request.user.is_superuser
        
        if not is_supervisor_or_higher:
            return JsonResponse({
                'success': False,
                'error': 'Apenas supervisores ou hierarquia maior podem excluir chats'
            }, status=403)
        
        # Salvar informações para log
        chat_title = chat.title
        chat_user = chat.user.get_full_name()
        
        # Deletar o chat (isso também deletará mensagens, arquivos, etc via CASCADE)
        chat.delete()
        
        return JsonResponse({
            'success': True,
            'message': f'Chat "{chat_title}" de {chat_user} foi excluído com sucesso'
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@login_required
def support_metrics(request):
    """
    Retorna métricas e estatísticas do suporte
    """
    # Verificar permissão: SUPERVISOR ou maior, ou agente de suporte
    is_supervisor_or_higher = request.user.hierarchy in ['SUPERVISOR', 'ADMIN', 'SUPERADMIN', 'ADMINISTRATIVO'] or request.user.is_superuser
    is_support_agent = SupportAgent.objects.filter(user=request.user, is_active=True).exists()
    
    if not (is_supervisor_or_higher or is_support_agent):
        return JsonResponse({'success': False, 'error': 'Sem permissão para acessar métricas'}, status=403)
    
    try:
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')
        
        if not start_date or not end_date:
            return JsonResponse({'success': False, 'error': 'Datas obrigatórias'}, status=400)
        
        # Converter strings para datetime
        from datetime import datetime, timedelta
        start = datetime.strptime(start_date, '%Y-%m-%d')
        end = datetime.strptime(end_date, '%Y-%m-%d')
        end = end + timedelta(days=1)  # Incluir o dia final completo
        
        # Filtrar por setores do usuário
        if request.user.hierarchy == 'SUPERADMIN':
            base_filter = Q()
        else:
            user_sectors = request.user.sectors.all()
            base_filter = Q(sector__in=user_sectors)
        
        # Tickets no período
        tickets = SupportChat.objects.filter(
            base_filter,
            created_at__gte=start,
            created_at__lt=end
        )
        
        total_tickets = tickets.count()
        resolved_tickets = tickets.filter(status='RESOLVIDO').count()
        
        # Tempo médio de resolução (usando closed_at ao invés de resolved_at)
        resolved_with_time = tickets.filter(status='RESOLVIDO', closed_at__isnull=False)
        if resolved_with_time.exists():
            total_time = sum([
                (ticket.closed_at - ticket.created_at).total_seconds() / 3600 
                for ticket in resolved_with_time
            ])
            avg_time = total_time / resolved_with_time.count()
            avg_time_str = f"{int(avg_time)}h"
        else:
            avg_time_str = "0h"
        
        # Taxa de satisfação (baseado em avaliações)
        from django.db.models import Avg
        rated_tickets = tickets.filter(rating__isnull=False)
        if rated_tickets.exists():
            avg_rating = rated_tickets.aggregate(Avg('rating'))['rating__avg']
            satisfaction_rate = f"{int((avg_rating / 5) * 100)}%"
        else:
            satisfaction_rate = "N/A"
        
        # Métricas por agente
        agents_stats = []
        agents = SupportAgent.objects.filter(is_active=True)
        
        if request.user.hierarchy != 'SUPERADMIN':
            agents = agents.filter(user__sectors__in=user_sectors).distinct()
        
        for agent in agents:
            agent_tickets = tickets.filter(assigned_to=agent.user)
            total = agent_tickets.count()
            resolved = agent_tickets.filter(status='RESOLVIDO').count()
            rate = int((resolved / total * 100)) if total > 0 else 0
            
            if total > 0:  # Só incluir agentes com tickets
                agents_stats.append({
                    'name': agent.user.get_full_name(),
                    'total': total,
                    'resolved': resolved,
                    'rate': rate
                })
        
        # Ordenar por total de tickets
        agents_stats.sort(key=lambda x: x['total'], reverse=True)
        
        # Métricas por categoria
        categories_stats = []
        categories = SupportCategory.objects.all()
        
        if request.user.hierarchy != 'SUPERADMIN':
            categories = categories.filter(sector__in=user_sectors)
        
        for category in categories:
            cat_tickets = tickets.filter(category=category)
            total = cat_tickets.count()
            
            if total > 0:  # Só incluir categorias com tickets
                categories_stats.append({
                    'name': category.name,
                    'total': total,
                    'open': cat_tickets.filter(status__in=['AGUARDANDO', 'ABERTO']).count(),
                    'resolved': cat_tickets.filter(status='RESOLVIDO').count()
                })
        
        # Ordenar por total
        categories_stats.sort(key=lambda x: x['total'], reverse=True)
        
        return JsonResponse({
            'success': True,
            'metrics': {
                'total_tickets': total_tickets,
                'resolved_tickets': resolved_tickets,
                'avg_resolution_time': avg_time_str,
                'satisfaction_rate': satisfaction_rate,
                'by_agent': agents_stats,
                'by_category': categories_stats
            }
        })
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


# ============ NOVAS VIEWS PARA TEMPO REAL E MELHORIAS ============

@login_required
def poll_dashboard_updates(request):
    """Polling para atualizações em tempo real do dashboard"""
    from django.db.models import Q
    from users.models import Sector
    import logging
    
    logger = logging.getLogger(__name__)
    
    # Verificar permissões
    is_support_agent = SupportAgent.objects.filter(
        user=request.user,
        is_active=True
    ).exists()
    is_supervisor_or_higher = request.user.hierarchy in ['SUPERVISOR', 'ADMIN', 'SUPERADMIN', 'ADMINISTRATIVO'] or request.user.is_superuser
    
    if not (is_supervisor_or_higher or is_support_agent):
        return JsonResponse({'success': False, 'error': 'Acesso negado'})
    
    # Obter timestamp da última verificação
    last_check = request.GET.get('last_check')
    if last_check:
        try:
            from datetime import datetime
            last_check_time = datetime.fromisoformat(last_check.replace('Z', '+00:00'))
            logger.info(f'[Polling] Last check time: {last_check_time}')
        except:
            last_check_time = None
            logger.warning('[Polling] Failed to parse last_check time')
    else:
        last_check_time = None
        logger.info('[Polling] No last_check provided')
    
    # Obter setores do usuário
    if request.user.is_superuser:
        user_sectors = Sector.objects.all()
        base_filter = Q()
        logger.info('[Polling] User is superuser - see all tickets')
    else:
        user_sectors_list = list(request.user.sectors.all())
        if request.user.sector:
            user_sectors_list.append(request.user.sector)
        user_sectors = Sector.objects.filter(id__in=[s.id for s in user_sectors_list])
        base_filter = Q(sector__in=user_sectors)
        logger.info(f'[Polling] User sectors: {[s.name for s in user_sectors]}')
    
    # Verificar se há novos tickets desde a última verificação
    has_updates = False
    new_tickets = []
    
    if last_check_time:
        new_chats = SupportChat.objects.filter(
            base_filter,
            created_at__gt=last_check_time
        ).select_related('user', 'assigned_to', 'sector', 'category').order_by('-created_at')[:5]
        
        logger.info(f'[Polling] Found {new_chats.count()} new chats since {last_check_time}')
        
        if new_chats.exists():
            has_updates = True
            for chat in new_chats:
                logger.info(f'[Polling] New chat: {chat.id} - {chat.title} - Status: {chat.status}')
                new_tickets.append({
                    'id': chat.id,
                    'title': chat.title,
                    'protocol': getattr(chat, 'protocol', None),
                    'status': chat.status,
                    'get_status_display': chat.get_status_display(),
                    'priority': chat.priority.lower() if chat.priority else 'media',
                    'get_priority_display': chat.get_priority_display(),
                    'user': {
                        'id': chat.user.id,
                        'get_full_name': chat.user.get_full_name()
                    },
                    'sector': {
                        'id': chat.sector.id,
                        'name': chat.sector.name
                    } if chat.sector else None,
                    'category': {
                        'id': chat.category.id,
                        'name': chat.category.name
                    } if chat.category else None,
                    'assigned_to': {
                        'id': chat.assigned_to.id,
                        'get_full_name': chat.assigned_to.get_full_name()
                    } if chat.assigned_to else None,
                    'created_at': chat.created_at.isoformat()
                })
    
    # Estatísticas atualizadas
    stats = {
        'total': SupportChat.objects.filter(base_filter).count(),
        'open': SupportChat.objects.filter(base_filter, status__in=['AGUARDANDO', 'ABERTO']).count(),
        'in_progress': SupportChat.objects.filter(base_filter, status='EM_ANDAMENTO').count(),
        'resolved': SupportChat.objects.filter(base_filter, status='RESOLVIDO').count(),
    }
    
    logger.info(f'[Polling] Stats: {stats}')
    logger.info(f'[Polling] Has updates: {has_updates}, New tickets count: {len(new_tickets)}')
    
    return JsonResponse({
        'success': True,
        'has_updates': has_updates,
        'new_tickets': new_tickets,
        'stats': stats,
        'timestamp': timezone.now().isoformat()
    })


@login_required
def poll_chat_updates(request, chat_id):
    """Polling para atualizações em tempo real do chat"""
    import json
    
    chat = get_object_or_404(SupportChat, id=chat_id)
    
    # Verificar se o usuário pode acessar este chat
    is_support_agent = hasattr(request.user, 'support_agent') and request.user.support_agent.is_active
    is_supervisor_or_higher = request.user.hierarchy in ['SUPERVISOR', 'ADMIN', 'SUPERADMIN', 'ADMINISTRATIVO'] or request.user.is_superuser
    
    if not (chat.user == request.user or is_support_agent or is_supervisor_or_higher):
        return JsonResponse({'success': False, 'error': 'Acesso negado'})
    
    # Obter último ID de mensagem visto (para retornar apenas novas)
    last_message_id = request.GET.get('last_message_id', 0)
    try:
        last_message_id = int(last_message_id)
    except:
        last_message_id = 0
    
    # Buscar novas mensagens
    messages = SupportChatMessage.objects.filter(
        chat=chat,
        id__gt=last_message_id
    ).select_related('user').prefetch_related('files')
    
    # Filtrar mensagens internas se não for agente de suporte ou supervisor+
    if not (is_support_agent or is_supervisor_or_higher):
        messages = messages.filter(is_internal=False)
    
    messages_data = []
    for msg in messages:
        msg_data = {
            'id': msg.id,
            'user': {
                'id': msg.user.id,
                'name': msg.user.get_full_name(),
                'avatar': msg.user.first_name[0].upper() if msg.user.first_name else 'U',
                'is_support': hasattr(msg.user, 'support_agent')
            },
            'message': msg.message,
            'is_internal': msg.is_internal,
            'created_at': msg.created_at.strftime('%d/%m/%Y %H:%M'),
            'is_own': msg.user == request.user,
            'files': []
        }
        
        # Adicionar arquivos da mensagem
        for file in msg.files.all():
            msg_data['files'].append({
                'id': file.id,
                'type': file.file_type,
                'name': file.original_name,
                'url': file.file.url,
                'size': file.file_size
            })
        
        messages_data.append(msg_data)
    
    # Calcular posição na fila
    queue_position = chat.get_queue_position() if hasattr(chat, 'get_queue_position') else None
    
    return JsonResponse({
        'success': True,
        'messages': messages_data,
        'chat_status': chat.status,
        'chat_status_display': chat.get_status_display(),
        'assigned_to': chat.assigned_to.get_full_name() if chat.assigned_to else None,
        'queue_position': queue_position,
        'protocol': getattr(chat, 'protocol', None)
    })


@login_required
def get_queue_status(request):
    """Retorna status da fila de chats por setor"""
    from users.models import Sector
    
    sector_id = request.GET.get('sector_id')
    
    if sector_id:
        sectors = Sector.objects.filter(id=sector_id)
    else:
        sectors = Sector.objects.all()
    
    queue_data = []
    for sector in sectors:
        waiting_count = SupportChat.objects.filter(
            sector=sector,
            status__in=['AGUARDANDO', 'ABERTO']
        ).count()
        
        in_progress_count = SupportChat.objects.filter(
            sector=sector,
            status='EM_ANDAMENTO'
        ).count()
        
        queue_data.append({
            'sector_id': sector.id,
            'sector_name': sector.name,
            'waiting': waiting_count,
            'in_progress': in_progress_count
        })
    
    return JsonResponse({
        'success': True,
        'queue': queue_data
    })


@login_required
def get_user_queue_position(request, chat_id):
    """Retorna a posição do usuário na fila para um chat específico"""
    chat = get_object_or_404(SupportChat, id=chat_id)
    
    # Verificar se o usuário pode acessar este chat
    if chat.user != request.user and not hasattr(request.user, 'support_agent'):
        return JsonResponse({'success': False, 'error': 'Acesso negado'})
    
    queue_position = chat.get_queue_position() if hasattr(chat, 'get_queue_position') else None
    
    # Contar total na fila do setor
    total_in_queue = 0
    if chat.sector:
        total_in_queue = SupportChat.objects.filter(
            sector=chat.sector,
            status__in=['AGUARDANDO', 'ABERTO']
        ).count()
    
    return JsonResponse({
        'success': True,
        'position': queue_position,
        'total_in_queue': total_in_queue,
        'status': chat.status,
        'protocol': getattr(chat, 'protocol', None)
    })


@login_required
@require_POST  
def upload_chat_file_with_message(request, chat_id):
    """Upload de arquivos no chat de suporte com mensagem"""
    chat = get_object_or_404(SupportChat, id=chat_id)
    
    # Verificar permissão
    is_support_agent = hasattr(request.user, 'support_agent') and request.user.support_agent.is_active
    is_supervisor_or_higher = request.user.hierarchy in ['SUPERVISOR', 'ADMIN', 'SUPERADMIN', 'ADMINISTRATIVO'] or request.user.is_superuser
    
    if not (chat.user == request.user or is_support_agent or is_supervisor_or_higher):
        return JsonResponse({'success': False, 'error': 'Acesso negado'})
    
    if 'file' not in request.FILES:
        return JsonResponse({'success': False, 'error': 'Nenhum arquivo enviado'})
    
    uploaded_file = request.FILES['file']
    
    # Validar tamanho (max 25MB)
    if uploaded_file.size > 25 * 1024 * 1024:
        return JsonResponse({'success': False, 'error': 'Arquivo muito grande (máximo 25MB)'})
    
    # Determinar tipo do arquivo
    file_type = 'DOCUMENT'
    content_type = uploaded_file.content_type.lower()
    if content_type.startswith('image/'):
        file_type = 'IMAGE'
    elif content_type.startswith('video/'):
        file_type = 'VIDEO'
    elif content_type.startswith('audio/'):
        file_type = 'AUDIO'
    
    # Criar mensagem primeiro
    message_text = request.POST.get('message', '').strip()
    if not message_text:
        message_text = f'📎 {uploaded_file.name}'
    
    is_internal = request.POST.get('is_internal') == 'true' and (is_support_agent or is_supervisor_or_higher)
    
    message = SupportChatMessage.objects.create(
        chat=chat,
        user=request.user,
        message=message_text,
        is_internal=is_internal
    )
    
    # Criar arquivo
    chat_file = SupportChatFile.objects.create(
        chat=chat,
        message=message,
        file=uploaded_file,
        file_type=file_type,
        original_name=uploaded_file.name,
        file_size=uploaded_file.size
    )
    
    # Atualizar status do chat se necessário
    if chat.status in ['AGUARDANDO', 'ABERTO'] and (is_support_agent or is_supervisor_or_higher):
        chat.status = 'EM_ANDAMENTO'
        chat.assigned_to = request.user
        chat.save()
    
    return JsonResponse({
        'success': True,
        'message': {
            'id': message.id,
            'user': {
                'id': request.user.id,
                'name': request.user.get_full_name(),
                'avatar': request.user.first_name[0].upper() if request.user.first_name else 'U',
                'is_support': is_support_agent
            },
            'message': message.message,
            'is_internal': message.is_internal,
            'created_at': message.created_at.strftime('%d/%m/%Y %H:%M'),
            'is_own': True,
            'files': [{
                'id': chat_file.id,
                'type': file_type,
                'name': chat_file.original_name,
                'url': chat_file.file.url,
                'size': chat_file.file_size
            }]
        }
    })


@login_required
@require_POST
def update_chat_status(request, chat_id):
    """Atualiza o status de um chat via drag and drop"""
    # Permitir acesso para supervisores ou agentes de suporte
    is_supervisor_or_higher = request.user.hierarchy in ['SUPERVISOR', 'ADMIN', 'SUPERADMIN', 'ADMINISTRATIVO'] or request.user.is_superuser
    is_support_agent = SupportAgent.objects.filter(
        user=request.user, 
        is_active=True
    ).exists()
    
    if not (is_supervisor_or_higher or is_support_agent):
        return JsonResponse({'success': False, 'error': 'Acesso negado'}, status=403)
    
    try:
        chat = SupportChat.objects.get(id=chat_id)
    except SupportChat.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Chat não encontrado'}, status=404)
    
    # Obter novo status do request
    new_status = request.POST.get('status', '').upper()
    
    # Validar status
    valid_statuses = ['AGUARDANDO', 'ABERTO', 'EM_ANDAMENTO', 'RESOLVIDO', 'FECHADO']
    if new_status not in valid_statuses:
        return JsonResponse({'success': False, 'error': 'Status inválido'}, status=400)
    
    # Atualizar status
    chat.status = new_status
    
    # Se movendo para EM_ANDAMENTO e não tem responsável, atribuir ao usuário atual
    if new_status == 'EM_ANDAMENTO' and not chat.assigned_to:
        chat.assigned_to = request.user
    
    chat.save()
    
    return JsonResponse({
        'success': True,
        'message': 'Status atualizado com sucesso',
        'chat': {
            'id': chat.id,
            'status': chat.status,
            'assigned_to': chat.assigned_to.get_full_name() if chat.assigned_to else None
        }
    })
