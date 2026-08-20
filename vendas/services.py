"""Importação da tabela de preços (planilha oficial, multi-aba) para ItemPreco.

Estratégia genérica e resiliente (as abas têm layouts diferentes e linhas de
título): para cada aba de dados, detecta a linha de cabeçalho pela maior
quantidade de palavras-chave, mapeia as colunas conhecidas (nome, valor, plano,
sistema, grupamento, cód SAP/sistema) e guarda TODO o resto em ``extra`` (JSON).
Reimportar substitui por (categoria, nome, cod_sap) via update_or_create.
"""
import unicodedata
from decimal import Decimal, InvalidOperation

from django.utils import timezone

from .models import ItemPreco

# Abas de dados a importar (as demais — índice/alterações/auxiliar — são ignoradas).
SHEETS_ALVO = [
    'PLANOS', 'PRODUTOS', 'SMARTPHONES', 'SMARTPHONES_Demo', 'WATCHES_PL',
    'ELETRÔNICOS_LP_Conectados', 'ELETRÔNICOS_LP_Não Conectados',
    'ELETRÔNICOS_PL_Conectados', 'ELETRÔNICOS_PL_Não Conectados',
    'VITRINE', 'DEVICES_ESPECIAIS', 'PRODUTOS B2B',
]

_CAMPO_KEYWORDS = {
    'nome': ['NOME DO PLANO', 'NOME', 'PRODUTO', 'MODELO COMERCIAL', 'MODELO', 'DESCRICAO', 'APARELHO'],
    'valor': ['VALOR', 'PRECO', 'PRECO CLIENTE'],
    'plano': ['PLANO'],
    'sistema': ['SISTEMA'],
    'grupamento': ['GRUPAMENTO'],
    'cod_sap': ['COD SAP', 'CODIGO SAP'],
    'cod_sistema': ['COD SISTEMA', 'COD SIST'],
}


def _norm(s):
    if s is None:
        return ''
    txt = str(s).strip()
    txt = unicodedata.normalize('NFKD', txt)
    txt = ''.join(c for c in txt if not unicodedata.combining(c))
    return ' '.join(txt.upper().split())


def _to_decimal(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        try:
            return Decimal(str(v))
        except (InvalidOperation, ValueError):
            return None
    txt = str(v).replace('R$', '').replace(' ', '').strip()
    if not txt:
        return None
    if ',' in txt and '.' in txt:
        txt = txt.replace('.', '').replace(',', '.')
    elif ',' in txt:
        txt = txt.replace(',', '.')
    try:
        return Decimal(txt)
    except (InvalidOperation, ValueError):
        return None


def _match_field(header_norm):
    """Dado um cabeçalho normalizado, devolve o nome do campo mapeado, ou None."""
    for campo, kws in _CAMPO_KEYWORDS.items():
        for kw in kws:
            if kw in header_norm:
                return campo
    return None


def _detect_header_row(rows, max_scan=15):
    """Índice da linha com mais colunas reconhecíveis (>=2 campos distintos)."""
    best_idx, best_score = None, 1
    for i, row in enumerate(rows[:max_scan]):
        campos = set()
        for cell in row:
            campo = _match_field(_norm(cell))
            if campo:
                campos.add(campo)
        if len(campos) > best_score:
            best_score, best_idx = len(campos), i
    return best_idx


def importar_tabela_precos(file_obj, sheets=None):
    """Importa as abas alvo do arquivo para ItemPreco. Devolve um resumo por aba."""
    import openpyxl

    wb = openpyxl.load_workbook(file_obj, data_only=True, read_only=True)
    alvo = sheets or SHEETS_ALVO
    agora = timezone.now()
    resumo = {'importados': 0, 'por_categoria': {}, 'abas_ignoradas': [], 'erros': []}

    for sheet_name in wb.sheetnames:
        if sheet_name not in alvo:
            continue
        try:
            ws = wb[sheet_name]
            rows = [list(r) for r in ws.iter_rows(values_only=True)]
            if not rows:
                continue
            header_idx = _detect_header_row(rows)
            if header_idx is None:
                resumo['abas_ignoradas'].append(sheet_name)
                continue

            headers = rows[header_idx]
            # Mapa: índice de coluna -> campo conhecido; e cabeçalho legível.
            col_campo = {}
            col_titulo = {}
            for ci, h in enumerate(headers):
                titulo = str(h).strip() if h is not None else ''
                if not titulo:
                    continue
                col_titulo[ci] = titulo
                campo = _match_field(_norm(h))
                if campo and campo not in col_campo.values():
                    col_campo[ci] = campo

            categoria = sheet_name[:60]
            count = 0
            for row in rows[header_idx + 1:]:
                dados = {campo: None for campo in _CAMPO_KEYWORDS}
                extra = {}
                for ci, titulo in col_titulo.items():
                    val = row[ci] if ci < len(row) else None
                    if ci in col_campo:
                        dados[col_campo[ci]] = val
                    elif val is not None and str(val).strip():
                        extra[titulo] = str(val).strip()

                nome = (str(dados['nome']).strip() if dados['nome'] is not None else '')
                if not nome or nome == '-':
                    continue

                ItemPreco.objects.update_or_create(
                    categoria=categoria,
                    nome=nome[:200],
                    cod_sap=(str(dados['cod_sap']).strip()[:40] if dados['cod_sap'] else ''),
                    defaults={
                        'plano': (str(dados['plano']).strip()[:200] if dados['plano'] else ''),
                        'sistema': (str(dados['sistema']).strip()[:60] if dados['sistema'] else ''),
                        'grupamento': (str(dados['grupamento']).strip()[:120] if dados['grupamento'] else ''),
                        'cod_sistema': (str(dados['cod_sistema']).strip()[:40] if dados['cod_sistema'] else ''),
                        'valor': _to_decimal(dados['valor']),
                        'extra': extra,
                        'ativo': True,
                        'importado_em': agora,
                    },
                )
                count += 1

            resumo['por_categoria'][categoria] = count
            resumo['importados'] += count
        except Exception as exc:  # noqa: BLE001 - uma aba problemática não derruba o import
            resumo['erros'].append(f'{sheet_name}: {exc}')

    return resumo
