import csv
import json
from datetime import timedelta
from decimal import Decimal, InvalidOperation
from io import BytesIO
from urllib.parse import urlencode

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.paginator import EmptyPage, PageNotAnInteger, Paginator
from django.db import transaction
from django.db.models import Count, DecimalField, F, Q, Sum
from django.db.models.functions import TruncDate
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from django.utils.dateparse import parse_date, parse_datetime
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from users.models import Sector, User

from .models import ItemPreco, Venda, VendaProduto, VendaServico
from .permissions import (can_access_vendas, is_superadmin, pode_gerenciar_precos,
                          vendas_do_usuario)
from .services import importar_tabela_precos


def _deny(request):
    messages.error(request, 'Acesso restrito.')
    return redirect('home')


def _parse_decimal(raw):
    if raw in (None, ''):
        return None
    txt = str(raw).replace('R$', '').replace(' ', '').strip()
    if ',' in txt and '.' in txt:
        txt = txt.replace('.', '').replace(',', '.')
    elif ',' in txt:
        txt = txt.replace(',', '.')
    try:
        return Decimal(txt)
    except (InvalidOperation, ValueError):
        return None


def _filtrar_vendas(request):
    """Aplica os filtros de GET e devolve (queryset, filtros_dict, filter_query_string)."""
    # O recorte vem antes de qualquer filtro: vendedor só enxerga o que é dele,
    # e nenhum parâmetro de URL contorna isso.
    qs = vendas_do_usuario(
        request.user, Venda.objects.select_related('loja', 'vendedor').all())

    search = request.GET.get('search', '').strip()
    date_from = request.GET.get('date_from', '').strip()
    date_to = request.GET.get('date_to', '').strip()
    loja_id = request.GET.get('loja', '').strip()
    vendedor_id = request.GET.get('vendedor', '').strip()
    tipo_venda = request.GET.get('tipo_venda', '').strip()
    comprovante = request.GET.get('comprovante', '').strip()

    if search:
        qs = qs.filter(
            Q(cliente_nome__icontains=search) | Q(cliente_cpf__icontains=search)
            | Q(pdv_nome__icontains=search) | Q(id__icontains=search)
        )
    if date_from and parse_date(date_from):
        qs = qs.filter(data_venda__date__gte=parse_date(date_from))
    if date_to and parse_date(date_to):
        qs = qs.filter(data_venda__date__lte=parse_date(date_to))
    if loja_id.isdigit():
        qs = qs.filter(loja_id=int(loja_id))
    if vendedor_id.isdigit():
        qs = qs.filter(vendedor_id=int(vendedor_id))
    if tipo_venda:
        qs = qs.filter(tipo_venda=tipo_venda)
    if comprovante:
        qs = qs.filter(comprovante_fiscal=comprovante)

    filtros = {
        'search': search, 'date_from': date_from, 'date_to': date_to,
        'loja': loja_id, 'vendedor': vendedor_id, 'tipo_venda': tipo_venda,
        'comprovante': comprovante,
    }
    filter_query_string = urlencode({k: v for k, v in filtros.items() if v})
    return qs, filtros, filter_query_string


@login_required
def dashboard(request):
    if not can_access_vendas(request.user):
        return _deny(request)

    qs, filtros, filter_query_string = _filtrar_vendas(request)

    export_format = request.GET.get('export', '')
    if export_format in ('csv', 'xlsx'):
        return venda_export(request)

    qs = qs.order_by('-data_venda', '-id')

    try:
        per_page = int(request.GET.get('per_page', '25'))
        if per_page not in (25, 50, 100, 200):
            per_page = 25
    except (ValueError, TypeError):
        per_page = 25

    paginator = Paginator(qs, per_page)
    page = request.GET.get('page')
    try:
        vendas_page = paginator.page(page)
    except PageNotAnInteger:
        vendas_page = paginator.page(1)
    except EmptyPage:
        vendas_page = paginator.page(paginator.num_pages)

    context = {
        'vendas': vendas_page,
        'paginator': paginator,
        'filtros': filtros,
        'filter_query_string': filter_query_string,
        'per_page': per_page,
        'lojas': Sector.objects.all().order_by('name'),
        'vendedores': User.objects.filter(is_active=True).order_by('first_name', 'last_name'),
        'comprovante_choices': Venda.COMPROVANTE_CHOICES,
        'is_superadmin': is_superadmin(request.user),
        'pode_gerenciar_precos': pode_gerenciar_precos(request.user),
        'aba': 'vendas',
    }
    context.update(_indicadores(qs, request.user))
    return render(request, 'vendas/dashboard.html', context)


def _indicadores(qs, user):
    """Os números do topo do dashboard, já no recorte de quem está olhando.

    Tudo sai do mesmo queryset filtrado da listagem — o painel e a tabela
    contam a mesma coisa, senão o total do topo brigaria com a soma da tabela.
    """
    ids = list(qs.values_list('id', flat=True))

    produtos = VendaProduto.objects.filter(venda_id__in=ids)
    servicos = VendaServico.objects.filter(venda_id__in=ids)

    receita_produtos = produtos.aggregate(
        t=Sum(F('valor_venda') * F('qtde'), output_field=DecimalField()))['t'] or Decimal('0')
    receita_servicos = servicos.aggregate(t=Sum('valor_plano'))['t'] or Decimal('0')
    total = receita_produtos + receita_servicos

    quantidade = len(ids)
    pecas = produtos.aggregate(n=Sum('qtde'))['n'] or 0

    hoje = timezone.localdate()
    inicio_mes = hoje.replace(day=1)

    # Série dos últimos 30 dias, para o gráfico de barras.
    serie = {}
    for linha in (qs.filter(data_venda__date__gte=hoje - timedelta(days=29))
                  .annotate(dia=TruncDate('data_venda'))
                  .values('dia').annotate(n=Count('id')).order_by('dia')):
        serie[linha['dia']] = linha['n']
    dias_serie = []
    for i in range(29, -1, -1):
        d = hoje - timedelta(days=i)
        dias_serie.append({'dia': d, 'n': serie.get(d, 0)})
    pico = max((x['n'] for x in dias_serie), default=0) or 1
    for x in dias_serie:
        x['altura'] = round(x['n'] / pico * 100)

    def _ranking(campo, rotulo, limite=6):
        linhas = list(qs.exclude(**{f'{campo}__isnull': True})
                      .values(campo, rotulo).annotate(n=Count('id')).order_by('-n')[:limite])
        maior = max((l['n'] for l in linhas), default=0) or 1
        for l in linhas:
            l['fatia'] = round(l['n'] / maior * 100)
            l['nome'] = l.get(rotulo) or '—'
        return linhas

    indicadores = {
        'kpi_total': total,
        'kpi_quantidade': quantidade,
        'kpi_ticket': (total / quantidade) if quantidade else Decimal('0'),
        'kpi_pecas': pecas,
        'kpi_produtos': receita_produtos,
        'kpi_servicos': receita_servicos,
        'kpi_no_mes': qs.filter(data_venda__date__gte=inicio_mes).count(),
        'kpi_hoje': qs.filter(data_venda__date=hoje).count(),
        'serie_dias': dias_serie,
        'top_produtos': list(produtos.values('nome_produto')
                             .annotate(n=Sum('qtde'),
                                       total=Sum(F('valor_venda') * F('qtde'),
                                                 output_field=DecimalField()))
                             .order_by('-total')[:6]),
        'top_servicos': list(servicos.values('servico')
                             .annotate(n=Count('id'), total=Sum('valor_plano'))
                             .order_by('-total')[:6]),
    }
    # O ranking de gente e de loja só faz sentido para quem vê mais de uma.
    if is_superadmin(user):
        indicadores['por_loja'] = _ranking('loja', 'loja__name')
        indicadores['por_vendedor'] = _ranking('vendedor', 'vendedor__first_name')
    return indicadores


@login_required
def venda_detail(request, pk):
    if not can_access_vendas(request.user):
        return _deny(request)
    # O recorte vale aqui também: sem isso, trocar o número na URL abriria a
    # venda de qualquer um.
    venda = get_object_or_404(
        vendas_do_usuario(request.user, Venda.objects.select_related('loja', 'vendedor')
                          .prefetch_related('produtos', 'servicos')), pk=pk,
    )
    return render(request, 'vendas/venda_detail.html', {
        'venda': venda,
        'is_superadmin': is_superadmin(request.user),
        'aba': 'vendas',
    })


@login_required
def venda_create(request):
    if not can_access_vendas(request.user):
        return _deny(request)

    if request.method == 'POST':
        data_raw = request.POST.get('data_venda', '').strip()
        data_venda = parse_datetime(data_raw) if data_raw else None
        if data_venda is None and data_raw:
            d = parse_date(data_raw)
            if d:
                data_venda = timezone.datetime(d.year, d.month, d.day)
        if data_venda is None:
            data_venda = timezone.now()
        if timezone.is_naive(data_venda):
            data_venda = timezone.make_aware(data_venda)

        loja_id = request.POST.get('loja')
        vendedor_id = request.POST.get('vendedor')

        try:
            produtos = json.loads(request.POST.get('produtos_json') or '[]')
            servicos = json.loads(request.POST.get('servicos_json') or '[]')
            if not isinstance(produtos, list):
                produtos = []
            if not isinstance(servicos, list):
                servicos = []
        except (json.JSONDecodeError, TypeError):
            produtos, servicos = [], []

        if not produtos and not servicos:
            messages.error(request, 'Adicione ao menos um produto ou serviço à venda.')
            return render(request, 'vendas/venda_form.html', _venda_form_context(request))

        with transaction.atomic():
            venda = Venda.objects.create(
                loja_id=int(loja_id) if (loja_id or '').isdigit() else None,
                pdv_nome=request.POST.get('pdv_nome', '').strip(),
                uf=request.POST.get('uf', '').strip()[:2],
                vendedor_id=int(vendedor_id) if (vendedor_id or '').isdigit() else None,
                estoque_avancado=(request.POST.get('estoque_avancado') == 'sim'),
                cliente_nome=request.POST.get('cliente_nome', '').strip(),
                cliente_cpf=request.POST.get('cliente_cpf', '').strip(),
                tipo_venda=request.POST.get('tipo_venda', '').strip(),
                comprovante_fiscal=request.POST.get('comprovante_fiscal', 'NFCE'),
                data_venda=data_venda,
                observacao=request.POST.get('observacao', '').strip(),
                created_by=request.user,
            )
            for p in produtos:
                nome = (p.get('nome_produto') or '').strip()
                if not nome:
                    continue
                VendaProduto.objects.create(
                    venda=venda, nome_produto=nome[:200],
                    tipo_produto=(p.get('tipo_produto') or '')[:120],
                    categoria=(p.get('categoria') or '')[:120],
                    subcategoria=(p.get('subcategoria') or '')[:120],
                    marca=(p.get('marca') or '')[:80],
                    modelo=(p.get('modelo') or '')[:200],
                    sku=(p.get('sku') or '')[:60],
                    serial=(p.get('serial') or '')[:120],
                    cor=(p.get('cor') or '')[:60],
                    qtde=int(p.get('qtde') or 1),
                    valor_venda=_parse_decimal(p.get('valor_venda')) or Decimal('0'),
                    plano=(p.get('plano') or '')[:200],
                    pilar=(p.get('pilar') or '')[:40],
                    preco_id=p.get('preco_id') or None,
                )
            for s in servicos:
                nome = (s.get('servico') or '').strip()
                if not nome:
                    continue
                VendaServico.objects.create(
                    venda=venda, servico=nome[:200],
                    servico_tecnico=(s.get('servico_tecnico') or '')[:200],
                    tipo_plano=(s.get('tipo_plano') or '')[:120],
                    plano_novo=(s.get('plano_novo') or '')[:200],
                    grupamento=(s.get('grupamento') or '')[:120],
                    numero_acesso=(s.get('numero_acesso') or '')[:40],
                    valor_plano=_parse_decimal(s.get('valor_plano')) or Decimal('0'),
                    receita=_parse_decimal(s.get('receita')),
                    status_servico=(s.get('status_servico') or '')[:60],
                    pilar=(s.get('pilar') or '')[:40],
                    preco_id=s.get('preco_id') or None,
                )

        messages.success(request, f'Venda #{venda.id} lançada com sucesso.')
        return redirect('vendas:venda_detail', pk=venda.id)

    return render(request, 'vendas/venda_form.html', _venda_form_context(request))


def _venda_form_context(request):
    return {
        'aba': 'nova',
        'is_superadmin': is_superadmin(request.user),
        'lojas': Sector.objects.all().order_by('name'),
        'vendedores': User.objects.filter(is_active=True).order_by('first_name', 'last_name'),
        'comprovante_choices': Venda.COMPROVANTE_CHOICES,
        'now': timezone.localtime().strftime('%Y-%m-%dT%H:%M'),
        'precos_categorias': list(
            ItemPreco.objects.filter(ativo=True).values_list('categoria', flat=True).distinct().order_by('categoria')
        ),
    }


# ---------------------------------------------------------------------------
# Exportação (reproduz os relatórios anexados: produto e serviço analítico)
# ---------------------------------------------------------------------------

_PRODUTO_HEADERS = ['Filial', 'UF', 'Produto', 'Tipo Produto', 'Categoria', 'Subcategoria',
                    'Nº Venda', 'Marca', 'Modelo', 'SKU', 'Serial', 'Cor', 'Nome Cliente',
                    'CPF/CNPJ', 'Vendedor', 'Plano', 'Tabela de preço', 'Data da venda',
                    'Qtde', 'Valor de Venda', 'Pilar']

_SERVICO_HEADERS = ['Filial', 'UF', 'Serviço', 'Serviço Técnico', 'Nº Venda', 'Data',
                    'Vendedor', 'Nome Cliente', 'CPF/CNPJ', 'Plano', 'Tipo do Plano',
                    'Grupamento', 'Nº de Acesso', 'Valor do Plano', 'Receita',
                    'Status do Serviço', 'Pilar']


def _produto_rows(vendas_qs):
    for vp in VendaProduto.objects.filter(venda__in=vendas_qs).select_related('venda', 'venda__loja', 'venda__vendedor'):
        v = vp.venda
        yield [
            v.loja.name if v.loja else v.pdv_nome, v.uf, vp.nome_produto, vp.tipo_produto,
            vp.categoria, vp.subcategoria, v.id, vp.marca, vp.modelo, vp.sku, vp.serial, vp.cor,
            v.cliente_nome, v.cliente_cpf,
            (v.vendedor.get_full_name() if v.vendedor else ''), vp.plano, vp.tabela_preco,
            v.data_venda.strftime('%d/%m/%Y %H:%M') if v.data_venda else '',
            vp.qtde, vp.valor_venda, vp.pilar,
        ]


def _servico_rows(vendas_qs):
    for vs in VendaServico.objects.filter(venda__in=vendas_qs).select_related('venda', 'venda__loja', 'venda__vendedor'):
        v = vs.venda
        yield [
            v.loja.name if v.loja else v.pdv_nome, v.uf, vs.servico, vs.servico_tecnico, v.id,
            v.data_venda.strftime('%d/%m/%Y %H:%M') if v.data_venda else '',
            (v.vendedor.get_full_name() if v.vendedor else ''), v.cliente_nome, v.cliente_cpf,
            vs.plano_novo, vs.tipo_plano, vs.grupamento, vs.numero_acesso,
            vs.valor_plano, vs.receita, vs.status_servico, vs.pilar,
        ]


@login_required
def venda_export(request):
    if not can_access_vendas(request.user):
        return _deny(request)

    qs, _filtros, _qsstr = _filtrar_vendas(request)
    tipo = request.GET.get('tipo', 'produto')
    fmt = request.GET.get('export', 'csv')
    headers = _PRODUTO_HEADERS if tipo == 'produto' else _SERVICO_HEADERS
    rows = _produto_rows(qs) if tipo == 'produto' else _servico_rows(qs)
    stamp = timezone.now().strftime('%Y%m%d_%H%M%S')
    base = f'vendas_{tipo}_{stamp}'

    if fmt == 'xlsx':
        wb = Workbook()
        ws = wb.active
        ws.title = tipo.capitalize()
        header_fill = PatternFill(start_color='6D28D9', end_color='6D28D9', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        for col, h in enumerate(headers, start=1):
            c = ws.cell(row=1, column=col, value=h)
            c.fill = header_fill
            c.font = header_font
            c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        r = 2
        for row in rows:
            for col, val in enumerate(row, start=1):
                ws.cell(row=r, column=col, value=(float(val) if isinstance(val, Decimal) else val))
            r += 1
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 20
        ws.freeze_panes = 'A2'
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        resp = HttpResponse(buf.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        resp['Content-Disposition'] = f'attachment; filename="{base}.xlsx"'
        return resp

    # CSV padrão brasileiro (delimitador ';', BOM para o Excel).
    resp = HttpResponse(content_type='text/csv; charset=utf-8')
    resp['Content-Disposition'] = f'attachment; filename="{base}.csv"'
    resp.write('﻿')
    writer = csv.writer(resp, delimiter=';')
    writer.writerow(headers)
    for row in rows:
        writer.writerow(['' if v is None else v for v in row])
    return resp


# ---------------------------------------------------------------------------
# Tabela de preços
# ---------------------------------------------------------------------------

@login_required
def precos(request):
    if not can_access_vendas(request.user):
        return _deny(request)

    itens = ItemPreco.objects.all()
    categoria = request.GET.get('categoria', '').strip()
    search = request.GET.get('search', '').strip()
    if categoria:
        itens = itens.filter(categoria=categoria)
    if search:
        itens = itens.filter(Q(nome__icontains=search) | Q(plano__icontains=search) | Q(cod_sap__icontains=search))
    itens = itens.order_by('categoria', 'nome')

    paginator = Paginator(itens, 50)
    try:
        page = paginator.page(request.GET.get('page'))
    except PageNotAnInteger:
        page = paginator.page(1)
    except EmptyPage:
        page = paginator.page(paginator.num_pages)

    context = {
        'aba': 'precos',
        'pode_gerenciar_precos': pode_gerenciar_precos(request.user),
        'itens': page,
        'paginator': paginator,
        'categoria': categoria,
        'search': search,
        'categorias': list(ItemPreco.objects.values_list('categoria', flat=True).distinct().order_by('categoria')),
        'total': ItemPreco.objects.count(),
    }
    return render(request, 'vendas/precos_list.html', context)


@login_required
def precos_import(request):
    if not pode_gerenciar_precos(request.user):
        return _deny(request)

    if request.method == 'POST':
        f = request.FILES.get('arquivo')
        if not f or not (f.name or '').lower().endswith(('.xlsx', '.xlsm')):
            messages.error(request, 'Envie a planilha (.xlsx).')
            return redirect('vendas:precos_import')
        try:
            resumo = importar_tabela_precos(f)
            partes = ', '.join(f'{k}: {v}' for k, v in resumo['por_categoria'].items())
            messages.success(request, f"{resumo['importados']} itens importados. {partes}")
            if resumo['erros']:
                messages.warning(request, 'Avisos: ' + ' | '.join(resumo['erros'][:5]))
        except Exception as exc:  # noqa: BLE001
            messages.error(request, f'Falha ao importar: {exc}')
        return redirect('vendas:precos')

    return render(request, 'vendas/precos_import.html', {'aba': 'precos'})


@login_required
def precos_create(request):
    if not pode_gerenciar_precos(request.user):
        return _deny(request)

    if request.method == 'POST':
        nome = request.POST.get('nome', '').strip()
        categoria = request.POST.get('categoria', '').strip()
        if not nome or not categoria:
            messages.error(request, 'Informe categoria e nome.')
        else:
            ItemPreco.objects.create(
                categoria=categoria[:60], nome=nome[:200],
                plano=request.POST.get('plano', '').strip()[:200],
                sistema=request.POST.get('sistema', '').strip()[:60],
                grupamento=request.POST.get('grupamento', '').strip()[:120],
                cod_sap=request.POST.get('cod_sap', '').strip()[:40],
                cod_sistema=request.POST.get('cod_sistema', '').strip()[:40],
                valor=_parse_decimal(request.POST.get('valor')),
            )
            messages.success(request, 'Item de preço cadastrado.')
            return redirect('vendas:precos')

    return render(request, 'vendas/precos_form.html', {
        'aba': 'precos',
        'categorias': list(ItemPreco.objects.values_list('categoria', flat=True).distinct().order_by('categoria')),
    })


@login_required
def precos_buscar(request):
    """Autocomplete da tabela de preços (JSON) para o formulário de venda."""
    if not can_access_vendas(request.user):
        return HttpResponse(status=403)
    from django.http import JsonResponse

    q = request.GET.get('q', '').strip()
    if len(q) < 2:
        return JsonResponse({'results': []})
    qs = (
        ItemPreco.objects.filter(ativo=True)
        .filter(Q(nome__icontains=q) | Q(plano__icontains=q) | Q(cod_sap__icontains=q))
        .order_by('categoria', 'nome')[:20]
    )
    results = [{
        'id': it.id, 'categoria': it.categoria, 'nome': it.nome, 'plano': it.plano,
        'sistema': it.sistema, 'grupamento': it.grupamento, 'cod_sap': it.cod_sap,
        'valor': (str(it.valor) if it.valor is not None else ''),
        'extra': it.extra if isinstance(it.extra, dict) else {},
    } for it in qs]
    return JsonResponse({'results': results})
