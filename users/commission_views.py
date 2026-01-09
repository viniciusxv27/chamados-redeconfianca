"""
Views para exibição de dados de comissionamento
Busca dados da planilha Excel do OneDrive/SharePoint
Design inspirado na identidade visual VIVO

Visões:
- CN: Vê seu comissionamento, valores por pilar, lista de vendas
- Gerente: Vê todos os CNs da loja, valores de cada CN, atingimentos por pilar
- Coordenador: Vê todos os gerentes e CNs, atingimento das lojas
"""
import requests
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.conf import settings
from django.core.cache import cache
from django.db.models import Q
import json
import pandas as pd
from io import BytesIO
from users.models import User, Sector, SystemConfig
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ============================================================================
# CONFIGURAÇÕES DAS PLANILHAS (valores padrão - podem ser alterados no banco)
# ============================================================================

# Valores padrão (usados se não houver configuração no banco)
DEFAULT_EXCEL_COMISSAO_URL = "https://1drv.ms/x/c/871ee1819c7e2faa/IQDp2ONpM_88ToSopGzjlPVdASX6rIFB3_ENXnJcGN3e7Go?e=Q3fDRz"
DEFAULT_EXCEL_VENDAS_URL = "https://1drv.ms/x/c/871ee1819c7e2faa/IQAVeQ-dgEiBTYG0UlK7URSLAQ5r634qBo9-GicO2D8ZfmY"
DEFAULT_EXCEL_BASE_PAGAMENTO_URL = "https://1drv.ms/x/c/871ee1819c7e2faa/IQBHZkNccF89Tb0x1dXfoLhiAT8Q5C_fzHlIyUnc2L2FJVs?e=vAO4OX"
DEFAULT_EXCEL_BASE_EXCLUSAO_URL = "https://1drv.ms/x/c/871ee1819c7e2faa/IQBryBteOg4sS4cBwU1tIgKoATfi6qmYB8eRrIaTpyP8Qhc?e=pye3Sj"


def get_excel_urls():
    """
    Obtém as URLs das planilhas Excel do banco de dados.
    Retorna valores padrão se não houver configuração.
    """
    try:
        config = SystemConfig.get_config()
        return {
            'comissao': config.excel_comissao_url or DEFAULT_EXCEL_COMISSAO_URL,
            'vendas': config.excel_vendas_url or DEFAULT_EXCEL_VENDAS_URL,
            'base_pagamento': config.excel_base_pagamento_url or DEFAULT_EXCEL_BASE_PAGAMENTO_URL,
            'base_exclusao': config.excel_base_exclusao_url or DEFAULT_EXCEL_BASE_EXCLUSAO_URL,
        }
    except Exception:
        # Se der erro ao acessar banco, usar valores padrão
        return {
            'comissao': DEFAULT_EXCEL_COMISSAO_URL,
            'vendas': DEFAULT_EXCEL_VENDAS_URL,
            'base_pagamento': DEFAULT_EXCEL_BASE_PAGAMENTO_URL,
            'base_exclusao': DEFAULT_EXCEL_BASE_EXCLUSAO_URL,
        }


# Manter variáveis globais para compatibilidade (serão substituídas gradualmente)
EXCEL_COMISSAO_URL = DEFAULT_EXCEL_COMISSAO_URL
EXCEL_VENDAS_URL = DEFAULT_EXCEL_VENDAS_URL
EXCEL_BASE_PAGAMENTO_URL = DEFAULT_EXCEL_BASE_PAGAMENTO_URL
EXCEL_BASE_EXCLUSAO_URL = DEFAULT_EXCEL_BASE_EXCLUSAO_URL

# Nome das sheets na planilha de comissionamento
SHEET_GERENTE = "REMUNERAÇÃO GERENTE"
SHEET_CN = "REMUNERAÇÃO CN"
SHEET_COORDENADOR = "REMUNERAÇÃO COO e SNP"

# Nome da sheet na planilha de vendas (será detectado automaticamente)
SHEET_VENDAS = "VENDAS"

# Grupos que identificam cada tipo de usuário
GERENTE_GROUP_NAME = "GERENTES (CHECKLIST)"
COORDENADOR_GROUP_NAME = "COORDENADORES"


# ============================================================================
# FUNÇÕES HELPER PARA VERIFICAR TIPO DE USUÁRIO
# ============================================================================

def is_user_gerente(user):
    """Verifica se o usuário está no grupo GERENTES (CHECKLIST)"""
    from communications.models import CommunicationGroup
    try:
        gerente_group = CommunicationGroup.objects.filter(name__icontains="GERENTES").first()
        if gerente_group:
            return user in gerente_group.members.all()
    except:
        pass
    return False


def is_user_coordenador(user):
    """Verifica se o usuário está no grupo COORDENADORES"""
    from communications.models import CommunicationGroup
    try:
        coord_group = CommunicationGroup.objects.filter(name__icontains="COORDENADORES").first()
        if coord_group:
            return user in coord_group.members.all()
    except:
        pass
    return False


def get_user_role(user):
    """
    Retorna o papel do usuário no sistema de comissionamento:
    - 'coordenador': Pode ver todos os gerentes e CNs
    - 'gerente': Pode ver CNs da sua loja
    - 'cn': Vê apenas seu próprio comissionamento
    """
    if is_user_coordenador(user):
        return 'coordenador'
    elif is_user_gerente(user):
        return 'gerente'
    else:
        return 'cn'


def get_sector_users(user, include_gerentes=False):
    """Retorna lista de usuários do mesmo setor"""
    if not user.sector:
        return User.objects.none()
    
    queryset = User.objects.filter(
        sector=user.sector,
        hierarchy='PADRAO',
        is_active=True
    ).exclude(id=user.id).order_by('first_name', 'last_name')
    
    return queryset


def get_all_cns_for_gerente(user):
    """Retorna todos os CNs que o gerente pode ver (mesmo setor/loja)"""
    if not user.sector:
        return User.objects.none()
    
    return User.objects.filter(
        sector=user.sector,
        hierarchy='PADRAO',
        is_active=True
    ).exclude(id=user.id).order_by('first_name', 'last_name')


def get_coordenador_scope(user):
    """
    Retorna os usuários que o coordenador pode ver.
    Coordenador vê todos os gerentes e CNs.
    """
    from communications.models import CommunicationGroup
    
    # Pegar todos os gerentes
    gerentes = []
    gerente_group = CommunicationGroup.objects.filter(name__icontains="GERENTES").first()
    if gerente_group:
        gerentes = list(gerente_group.members.filter(is_active=True))
    
    # Pegar todos os CNs (usuários PADRAO que não são gerentes)
    gerente_ids = [g.id for g in gerentes]
    cns = User.objects.filter(
        hierarchy='PADRAO',
        is_active=True
    ).exclude(id__in=gerente_ids + [user.id])
    
    return {
        'gerentes': gerentes,
        'cns': cns,
    }


def get_lojas_do_coordenador(user):
    """Retorna as lojas/setores que o coordenador supervisiona"""
    from communications.models import CommunicationGroup
    
    # Pegar todos os gerentes para identificar as lojas
    gerente_group = CommunicationGroup.objects.filter(name__icontains="GERENTES").first()
    if not gerente_group:
        return []
    
    gerentes = gerente_group.members.filter(is_active=True)
    
    # Agrupar por setor (loja)
    lojas = {}
    for gerente in gerentes:
        if gerente.sector:
            if gerente.sector.id not in lojas:
                lojas[gerente.sector.id] = {
                    'setor': gerente.sector,
                    'gerente': gerente,
                    'cns': []
                }
    
    # Adicionar CNs a cada loja
    for loja_id, loja_data in lojas.items():
        loja_data['cns'] = User.objects.filter(
            sector_id=loja_id,
            hierarchy='PADRAO',
            is_active=True
        ).exclude(id=loja_data['gerente'].id)
    
    return list(lojas.values())


# ============================================================================
# FUNÇÕES DE DOWNLOAD E PROCESSAMENTO DE PLANILHAS
# ============================================================================

def get_excel_download_url(share_url):
    """
    Converte URL de compartilhamento do OneDrive para URL de download direto
    """
    import base64
    import re
    
    if 'download=1' in share_url:
        return share_url
    
    if 'resid=' in share_url:
        match = re.search(r'resid=([^&]+)', share_url)
        if match:
            resid = match.group(1)
            return f"https://onedrive.live.com/download?resid={resid}"
    
    encoded_url = base64.urlsafe_b64encode(share_url.encode()).decode()
    encoded_url = encoded_url.rstrip('=')
    sharing_token = f"u!{encoded_url}"
    
    return f"https://api.onedrive.com/v1.0/shares/{sharing_token}/root/content"


def download_excel_file(excel_url, cache_key_prefix="excel"):
    """
    Baixa arquivo Excel do OneDrive e retorna como BytesIO
    """
    cache_key = f"{cache_key_prefix}_file_content"
    cached_content = cache.get(cache_key)
    
    if cached_content:
        return BytesIO(cached_content), None
    
    download_urls = []
    download_urls.append(get_excel_download_url(excel_url))
    
    if '?' in excel_url:
        download_urls.append(excel_url + '&download=1')
    else:
        download_urls.append(excel_url + '?download=1')
    
    if 'IQ' in excel_url:
        import re
        match = re.search(r'(IQ[A-Za-z0-9_-]+)', excel_url)
        if match:
            file_id = match.group(1)
            download_urls.append(f"https://onedrive.live.com/download.aspx?resid={file_id}")
    
    response = None
    last_error = None
    
    for url in download_urls:
        try:
            headers = {
                'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
            }
            response = requests.get(url, timeout=30, headers=headers, allow_redirects=True)
            
            if response.status_code == 200:
                content_type = response.headers.get('Content-Type', '')
                if 'excel' in content_type or 'spreadsheet' in content_type or 'octet-stream' in content_type or len(response.content) > 1000:
                    break
                else:
                    last_error = f'Conteúdo não é Excel: {content_type}'
                    response = None
            else:
                last_error = f'HTTP {response.status_code}'
                response = None
        except Exception as e:
            last_error = str(e)
            continue
    
    if response is None or response.status_code != 200:
        return None, f'Erro ao baixar planilha. Último erro: {last_error}'
    
    # Cache por 5 minutos
    cache.set(cache_key, response.content, 300)
    
    return BytesIO(response.content), None


def fetch_excel_data(sheet_name, user_name, excel_url=None):
    """
    Busca dados da planilha Excel do OneDrive
    Retorna os dados do usuário específico
    """
    if excel_url is None:
        urls = get_excel_urls()
        excel_url = urls['comissao']
        
    cache_key = f"commission_data_{sheet_name}_{user_name.replace(' ', '_')}"
    cached_data = cache.get(cache_key)
    
    if cached_data:
        return cached_data
    
    try:
        excel_file, error = download_excel_file(excel_url, f"comissao_{sheet_name}")
        if error:
            return {'error': error}
        
        try:
            df = pd.read_excel(excel_file, sheet_name=sheet_name, engine='openpyxl')
        except Exception as e:
            return {'error': f'Erro ao ler planilha: {str(e)}'}
        
        nome_col = None
        for col in df.columns:
            if 'NOME' in str(col).upper():
                nome_col = col
                break
        
        if nome_col is None:
            nome_col = df.columns[0]
        
        def normalize_name(name):
            if pd.isna(name):
                return ""
            return str(name).strip().upper()
        
        user_name_normalized = normalize_name(user_name)
        
        user_row = None
        for idx, row in df.iterrows():
            row_name = normalize_name(row.get(nome_col, ''))
            if row_name == user_name_normalized:
                user_row = row
                break
            elif user_name_normalized in row_name or row_name in user_name_normalized:
                user_row = row
                break
        
        if user_row is None:
            return {'error': 'Usuário não encontrado na planilha', 'user_name': user_name}
        
        data = {}
        for col in df.columns:
            value = user_row.get(col)
            if pd.isna(value):
                data[str(col)] = None
            elif isinstance(value, (int, float)):
                data[str(col)] = float(value) if isinstance(value, float) else int(value)
            else:
                data[str(col)] = str(value)
        
        result = {
            'success': True,
            'data': data,
            'sheet': sheet_name
        }
        
        cache.set(cache_key, result, 300)
        
        return result
        
    except requests.exceptions.Timeout:
        return {'error': 'Tempo limite excedido ao acessar a planilha'}
    except Exception as e:
        return {'error': f'Erro ao processar dados: {str(e)}'}


def fetch_all_users_from_sheet(sheet_name, excel_url=None):
    """
    Busca todos os usuários de uma sheet da planilha
    Retorna lista com dados de todos os usuários
    """
    if excel_url is None:
        urls = get_excel_urls()
        excel_url = urls['comissao']
        
    cache_key = f"commission_all_users_{sheet_name}"
    cached_data = cache.get(cache_key)
    
    if cached_data:
        return cached_data
    
    try:
        excel_file, error = download_excel_file(excel_url, f"comissao_{sheet_name}")
        if error:
            return {'error': error}
        
        try:
            df = pd.read_excel(excel_file, sheet_name=sheet_name, engine='openpyxl')
        except Exception as e:
            return {'error': f'Erro ao ler planilha: {str(e)}'}
        
        nome_col = None
        for col in df.columns:
            if 'NOME' in str(col).upper():
                nome_col = col
                break
        
        if nome_col is None:
            nome_col = df.columns[0]
        
        users_data = []
        for idx, row in df.iterrows():
            nome = row.get(nome_col)
            if pd.isna(nome) or not str(nome).strip():
                continue
            
            data = {}
            for col in df.columns:
                value = row.get(col)
                if pd.isna(value):
                    data[str(col)] = None
                elif isinstance(value, (int, float)):
                    data[str(col)] = float(value) if isinstance(value, float) else int(value)
                else:
                    data[str(col)] = str(value)
            
            users_data.append({
                'nome': str(nome).strip(),
                'data': data
            })
        
        result = {
            'success': True,
            'users': users_data,
            'sheet': sheet_name
        }
        
        cache.set(cache_key, result, 300)
        
        return result
        
    except Exception as e:
        return {'error': f'Erro ao processar dados: {str(e)}'}


def fetch_metas_por_pilar(user_name, is_gerente=False, user_sector=None):
    """
    Busca Pago e Exclusão por pilar da planilha BASE_PAGAMENTO.
    
    Planilha: BASE_PAGAMENTO
    Sheet: Planilha1
    Colunas: Filial, UF, Nº da Venda, Data, Vendedor, Nome Cliente, CPF/CNPJ, 
             Plano/Produto, NUMERO ACESSO, Receita, COORDENACAO, PILAR
    
    Filtra por Filial (para gerente) ou Vendedor (para CN)
    Agrupa por PILAR e soma Receita
    """
    # Cache key diferente para gerente vs CN
    cache_suffix = f"gerente_{user_sector}" if is_gerente else f"cn_{user_name}"
    cache_key = f"metas_pilar_v2_{cache_suffix.replace(' ', '_')}"
    
    # Verificar cache
    cached_data = cache.get(cache_key)
    if cached_data:
        print(f"[fetch_metas_por_pilar] Usando cache para {cache_suffix}")
        return cached_data
    
    # Inicializar metas
    metas = {
        'movel': {'total': 0, 'pago': 0, 'exclusao': 0},
        'fixa': {'total': 0, 'pago': 0, 'exclusao': 0},
        'smartphone': {'total': 0, 'pago': 0, 'exclusao': 0},
        'eletronicos': {'total': 0, 'pago': 0, 'exclusao': 0},
        'essenciais': {'total': 0, 'pago': 0, 'exclusao': 0},
        'seguro': {'total': 0, 'pago': 0, 'exclusao': 0},
        'sva': {'total': 0, 'pago': 0, 'exclusao': 0},
    }
    
    # Mapeamento de texto do pilar para chave
    pilar_mapping = {
        'MÓVEL': 'movel',
        'FIXA': 'fixa',
        'SMART': 'smartphone',
        'SMARTPHONE': 'smartphone',
        'ELETRO': 'eletronicos',
        'ELETRÔNICO': 'eletronicos',
        'ELETRONICOS': 'eletronicos',
        'ELETRÔNICOS': 'eletronicos',
        'ESSEN': 'essenciais',
        'ESSENCIAL': 'essenciais',
        'ESSENCIAIS': 'essenciais',
        'SEG': 'seguro',
        'SEGURO': 'seguro',
        'SEGUROS': 'seguro',
        'SVA': 'sva',
    }
    
    def normalize_text(text):
        if pd.isna(text):
            return ""
        return str(text).strip().upper()
    
    def normalize_sector(sector_name):
        """Normaliza nome do setor para comparação."""
        if not sector_name:
            return ""
        normalized = str(sector_name).strip().upper()
        # Remover prefixos comuns
        prefixes = ['LOJA ', 'LOJA_', 'PDV ', 'PDV_']
        for prefix in prefixes:
            if normalized.startswith(prefix):
                normalized = normalized[len(prefix):]
        return normalized
    
    def sectors_match(sector1, sector2):
        """Verifica se dois setores são equivalentes"""
        s1 = normalize_sector(sector1)
        s2 = normalize_sector(sector2)
        if not s1 or not s2:
            return False
        # Comparação exata
        if s1 == s2:
            return True
        # Um contém o outro
        if s1 in s2 or s2 in s1:
            return True
        # Comparar palavras principais
        words1 = set(s1.split())
        words2 = set(s2.split())
        common = words1 & words2
        if len(common) >= 1 and len(common) >= min(len(words1), len(words2)) * 0.5:
            return True
        return False
    
    user_name_normalized = normalize_text(user_name)
    user_sector_normalized = normalize_sector(user_sector) if user_sector else ""
    
    print(f"[fetch_metas_por_pilar] user_name={user_name}, is_gerente={is_gerente}, user_sector={user_sector}")
    print(f"[fetch_metas_por_pilar] user_sector_normalized={user_sector_normalized}")
    
    # Obter URLs dinâmicas
    urls = get_excel_urls()
    
    try:
        # Baixar planilha BASE_PAGAMENTO (apenas para dados de pagamento)
        excel_file, error = download_excel_file(urls['base_pagamento'], "base_pagamento")
        if error:
            print(f"[fetch_metas_por_pilar] Erro ao baixar planilha BASE_PAGAMENTO: {error}")
            return metas
        
        # Ler a sheet "Planilha1" da BASE_PAGAMENTO
        try:
            excel_file.seek(0)
            df = pd.read_excel(excel_file, sheet_name='Planilha1', engine='openpyxl')
            print(f"[fetch_metas_por_pilar] Colunas da BASE_PAGAMENTO: {list(df.columns)}")
            print(f"[fetch_metas_por_pilar] Total de linhas: {len(df)}")
        except Exception as e:
            print(f"[fetch_metas_por_pilar] Erro ao ler planilha: {e}")
            return metas
        
        # Encontrar colunas necessárias
        filial_col = None
        receita_col = None
        pilar_col = None
        vendedor_col = None
        
        for col in df.columns:
            col_upper = str(col).strip().upper()
            if col_upper == 'FILIAL':
                filial_col = col
            elif col_upper == 'RECEITA':
                receita_col = col
            elif col_upper == 'PILAR':
                pilar_col = col
            elif col_upper == 'VENDEDOR':
                vendedor_col = col
        
        print(f"[fetch_metas_por_pilar] filial_col={filial_col}, receita_col={receita_col}, pilar_col={pilar_col}, vendedor_col={vendedor_col}")
        
        if receita_col is None or pilar_col is None:
            print(f"[fetch_metas_por_pilar] ERRO: Colunas obrigatórias não encontradas")
            return metas
        
        # Debug: mostrar valores únicos de filial e pilar
        if filial_col:
            filiais_unicas = df[filial_col].dropna().unique()[:15]
            print(f"[fetch_metas_por_pilar] Filiais na planilha (amostra): {list(filiais_unicas)}")
        
        pilares_unicos = df[pilar_col].dropna().unique()
        print(f"[fetch_metas_por_pilar] Pilares na planilha: {list(pilares_unicos)}")
        
        # Debug: Filtrar apenas por filial EXATA "SERRA SEDE" e contar
        df_serra = df[df[filial_col].astype(str).str.strip().str.upper() == 'SERRA SEDE']
        print(f"[DEBUG] Linhas com Filial EXATA 'SERRA SEDE': {len(df_serra)}")
        
        # Contar e somar por pilar apenas para SERRA SEDE exato
        for pilar_nome in df_serra[pilar_col].dropna().unique():
            df_pilar = df_serra[df_serra[pilar_col] == pilar_nome]
            soma = df_pilar[receita_col].sum()
            count = len(df_pilar)
            print(f"[DEBUG SERRA SEDE] Pilar={pilar_nome}, Qtd={count}, Soma Receita={soma}")
        
        # Processar cada linha - USAR COMPARAÇÃO EXATA para filial
        for idx, row in df.iterrows():
            # Verificar filtro (Filial para gerente, Vendedor para CN)
            match = False
            
            if is_gerente and filial_col:
                row_filial = str(row.get(filial_col, '')).strip().upper()
                # Comparação EXATA - normalizar ambos para comparar
                match = (row_filial == user_sector_normalized)
            elif not is_gerente and vendedor_col:
                row_vendedor = normalize_text(row.get(vendedor_col, ''))
                match = (row_vendedor == user_name_normalized or 
                        user_name_normalized in row_vendedor or 
                        row_vendedor in user_name_normalized)
            
            if match:
                # Identificar o pilar desta linha
                pilar_val = normalize_text(row.get(pilar_col, ''))
                
                # Mapear para chave do pilar
                pilar_key = None
                for key, value in pilar_mapping.items():
                    if key in pilar_val or pilar_val == key:
                        pilar_key = value
                        break
                
                if pilar_key:
                    # Somar receita ao pago
                    val = row.get(receita_col)
                    if pd.notna(val):
                        try:
                            valor_float = float(val)
                            metas[pilar_key]['pago'] += valor_float
                        except (ValueError, TypeError):
                            pass
        
        # =====================================================
        # PROCESSAR PLANILHA DE EXCLUSÃO
        # =====================================================
        try:
            excel_file_exclusao, error_exclusao = download_excel_file(urls['base_exclusao'], "base_exclusao")
            if error_exclusao:
                print(f"[fetch_metas_por_pilar] Erro ao baixar planilha BASE_EXCLUSAO: {error_exclusao}")
                # Continuar sem dados de exclusão
                df_exclusao = None
            else:
                excel_file_exclusao.seek(0)
                df_exclusao = pd.read_excel(excel_file_exclusao, sheet_name='Planilha1', engine='openpyxl')
                print(f"[fetch_metas_por_pilar] Colunas da BASE_EXCLUSAO: {list(df_exclusao.columns)}")
                print(f"[fetch_metas_por_pilar] Total de linhas EXCLUSÃO: {len(df_exclusao)}")
            
            if df_exclusao is not None:
                # Encontrar colunas na planilha de exclusão
                filial_col_exc = None
                receita_col_exc = None
                pilar_col_exc = None
                
                for col in df_exclusao.columns:
                    col_upper = str(col).strip().upper()
                    if col_upper == 'FILIAL':
                        filial_col_exc = col
                    elif col_upper == 'RECEITA':
                        receita_col_exc = col
                    elif col_upper == 'PILAR':
                        pilar_col_exc = col
                
                print(f"[fetch_metas_por_pilar] EXCLUSÃO - filial_col={filial_col_exc}, receita_col={receita_col_exc}, pilar_col={pilar_col_exc}")
            else:
                filial_col_exc = None
                receita_col_exc = None
                pilar_col_exc = None
            
            if filial_col_exc and receita_col_exc and pilar_col_exc and df_exclusao is not None:
                # Encontrar coluna vendedor para filtro de CN
                vendedor_col_exc = None
                for col in df_exclusao.columns:
                    col_upper = str(col).strip().upper()
                    if col_upper == 'VENDEDOR':
                        vendedor_col_exc = col
                        break
                
                # Debug: mostrar informações baseado no tipo de usuário
                if is_gerente:
                    df_exc_filtrado = df_exclusao[df_exclusao[filial_col_exc].astype(str).str.strip().str.upper() == user_sector_normalized]
                    print(f"[DEBUG EXCLUSÃO GERENTE] Filtrando por Filial='{user_sector_normalized}': {len(df_exc_filtrado)} linhas")
                else:
                    if vendedor_col_exc:
                        # Para CN, criar uma lista de possíveis matches do nome
                        df_exc_filtrado = df_exclusao[
                            df_exclusao[vendedor_col_exc].astype(str).str.strip().str.upper().str.contains(user_name_normalized.replace(' ', ''), na=False, case=False)
                        ]
                        print(f"[DEBUG EXCLUSÃO CN] Filtrando por Vendedor contendo '{user_name_normalized}': {len(df_exc_filtrado)} linhas")
                    else:
                        print(f"[DEBUG EXCLUSÃO CN] ERRO: Coluna 'Vendedor' não encontrada!")
                        df_exc_filtrado = pd.DataFrame()
                
                # Mostrar soma por pilar no filtro
                if len(df_exc_filtrado) > 0:
                    for pilar_nome in df_exc_filtrado[pilar_col_exc].dropna().unique():
                        df_pilar = df_exc_filtrado[df_exc_filtrado[pilar_col_exc] == pilar_nome]
                        soma = df_pilar[receita_col_exc].sum()
                        count = len(df_pilar)
                        tipo = "GERENTE" if is_gerente else "CN"
                        print(f"[DEBUG EXCLUSÃO {tipo}] Pilar={pilar_nome}, Qtd={count}, Soma Receita={soma}")
                
                # Processar cada linha da sheet de exclusão
                for idx, row in df_exclusao.iterrows():
                    # Verificar filtro (Filial para gerente, Vendedor para CN)
                    match = False
                    
                    if is_gerente and filial_col_exc:
                        row_filial = str(row.get(filial_col_exc, '')).strip().upper()
                        # Comparação EXATA
                        match = (row_filial == user_sector_normalized)
                    elif not is_gerente and vendedor_col_exc:
                        # Para CN, buscar pelo nome do vendedor
                        row_vendedor = str(row.get(vendedor_col_exc, '')).strip().upper()
                        row_vendedor_normalized = row_vendedor.replace(' ', '')
                        # Match se o nome do usuário está contido no vendedor ou vice-versa
                        match = (user_name_normalized.replace(' ', '') in row_vendedor_normalized or 
                                row_vendedor_normalized in user_name_normalized.replace(' ', ''))
                    
                    if match:
                        pilar_val = normalize_text(row.get(pilar_col_exc, ''))
                        
                        # Mapear para chave do pilar
                        pilar_key = None
                        for key, value in pilar_mapping.items():
                            if key in pilar_val or pilar_val == key:
                                pilar_key = value
                                break
                        
                        if pilar_key:
                            val = row.get(receita_col_exc)
                            if pd.notna(val):
                                try:
                                    valor_float = float(val)
                                    metas[pilar_key]['exclusao'] += valor_float
                                except (ValueError, TypeError):
                                    pass
            else:
                print(f"[fetch_metas_por_pilar] AVISO: Colunas obrigatórias não encontradas na planilha EXCLUSÃO ou planilha não disponível")
                
        except Exception as e:
            print(f"[fetch_metas_por_pilar] Erro ao processar planilha EXCLUSÃO: {e}")
        
        # Calcular totais
        total_geral_pago = 0
        total_geral_exclusao = 0
        for pilar_key in metas:
            pago = metas[pilar_key]['pago']
            exclusao = metas[pilar_key]['exclusao']
            total_geral_pago += pago
            total_geral_exclusao += exclusao
            # Total = Pago + Exclusão para calcular porcentagem
            metas[pilar_key]['total'] = (pago + exclusao) if (pago + exclusao) > 0 else 1
        
        print(f"[fetch_metas_por_pilar] RESULTADO FINAL: {metas}")
        print(f"[fetch_metas_por_pilar] Total geral pago: {total_geral_pago}, Total geral exclusão: {total_geral_exclusao}")
        
        # Salvar no cache por 5 minutos
        cache.set(cache_key, metas, 300)
        
        return metas
        
    except Exception as e:
        print(f"[fetch_metas_por_pilar] Erro geral: {e}")
        import traceback
        traceback.print_exc()
        return metas


def fetch_vendas_data(user_name):
    """
    Busca dados de vendas do usuário da planilha de vendas e metas
    """
    cache_key = f"vendas_data_{user_name.replace(' ', '_')}"
    cached_data = cache.get(cache_key)
    
    if cached_data:
        return cached_data
    
    try:
        excel_file, error = download_excel_file(EXCEL_VENDAS_URL, "vendas")
        if error:
            return {'error': error, 'vendas': [], 'metas_pilar': {}}
        
        try:
            # Tentar ler a primeira sheet se SHEET_VENDAS não existir
            try:
                df = pd.read_excel(excel_file, sheet_name=SHEET_VENDAS, engine='openpyxl')
            except:
                excel_file.seek(0)
                df = pd.read_excel(excel_file, sheet_name=0, engine='openpyxl')
        except Exception as e:
            return {'error': f'Erro ao ler planilha de vendas: {str(e)}', 'vendas': [], 'metas_pilar': {}}
        
        # Encontrar coluna de nome
        nome_col = None
        for col in df.columns:
            col_upper = str(col).upper()
            if 'NOME' in col_upper or 'VENDEDOR' in col_upper or 'CN' in col_upper:
                nome_col = col
                break
        
        if nome_col is None:
            nome_col = df.columns[0]
        
        def normalize_name(name):
            if pd.isna(name):
                return ""
            return str(name).strip().upper()
        
        user_name_normalized = normalize_name(user_name)
        
        # Buscar todas as linhas do usuário (pode ter múltiplas vendas)
        user_rows = []
        for idx, row in df.iterrows():
            row_name = normalize_name(row.get(nome_col, ''))
            if row_name == user_name_normalized or user_name_normalized in row_name or row_name in user_name_normalized:
                row_data = {}
                for col in df.columns:
                    value = row.get(col)
                    if pd.isna(value):
                        row_data[str(col)] = None
                    elif isinstance(value, (int, float)):
                        row_data[str(col)] = float(value) if isinstance(value, float) else int(value)
                    else:
                        row_data[str(col)] = str(value)
                user_rows.append(row_data)
        
        # Buscar metas por pilar das sheets específicas (Total, Pago, Exclusão)
        metas_pilar = fetch_metas_por_pilar(user_name)
        
        result = {
            'success': True,
            'vendas': user_rows,
            'total_vendas': len(user_rows),
            'metas_pilar': metas_pilar
        }
        
        cache.set(cache_key, result, 300)
        
        return result
        
    except Exception as e:
        return {'error': f'Erro ao processar vendas: {str(e)}', 'vendas': [], 'metas_pilar': {}}


def fetch_all_vendas():
    """
    Busca todas as vendas da planilha
    """
    cache_key = "vendas_all_data"
    cached_data = cache.get(cache_key)
    
    if cached_data:
        return cached_data
    
    try:
        excel_file, error = download_excel_file(EXCEL_VENDAS_URL, "vendas_all")
        if error:
            return {'error': error, 'vendas': [], 'colunas': []}
        
        try:
            try:
                df = pd.read_excel(excel_file, sheet_name=SHEET_VENDAS, engine='openpyxl')
            except:
                excel_file.seek(0)
                df = pd.read_excel(excel_file, sheet_name=0, engine='openpyxl')
        except Exception as e:
            return {'error': f'Erro ao ler planilha de vendas: {str(e)}', 'vendas': [], 'colunas': []}
        
        all_vendas = []
        for idx, row in df.iterrows():
            row_data = {}
            for col in df.columns:
                value = row.get(col)
                if pd.isna(value):
                    row_data[str(col)] = None
                elif isinstance(value, (int, float)):
                    row_data[str(col)] = float(value) if isinstance(value, float) else int(value)
                else:
                    row_data[str(col)] = str(value)
            all_vendas.append(row_data)
        
        result = {
            'success': True,
            'vendas': all_vendas,
            'colunas': list(df.columns),
            'total': len(all_vendas)
        }
        
        cache.set(cache_key, result, 300)
        
        return result
        
    except Exception as e:
        return {'error': f'Erro ao processar vendas: {str(e)}', 'vendas': [], 'colunas': []}


# ============================================================================
# FUNÇÕES DE PROCESSAMENTO DE DADOS
# ============================================================================

def safe_float(value, default=0):
    """Converte valor para float de forma segura"""
    if value is None:
        return default
    try:
        return float(value)
    except:
        return default


def normalize_col_name(col_name):
    """
    Normaliza nome de coluna removendo acentos e convertendo para uppercase.
    Usado para comparação flexível de nomes de colunas.
    """
    import unicodedata
    if col_name is None:
        return ""
    s = str(col_name).upper().strip()
    # Remove acentos
    s = unicodedata.normalize('NFD', s)
    s = ''.join(c for c in s if unicodedata.category(c) != 'Mn')
    return s


def find_column_value(data, search_patterns):
    """
    Busca um valor nas colunas do data usando padrões flexíveis.
    Retorna o primeiro valor encontrado (não zero/none).
    
    search_patterns: lista de padrões para buscar (ex: ['%ATING_MOVEL_1', 'ATING_MOVEL_1'])
    """
    # Primeiro, tentar busca exata
    for pattern in search_patterns:
        val = data.get(pattern)
        if val is not None and safe_float(val) != 0:
            return safe_float(val)
    
    # Depois, busca normalizada (sem acentos, case-insensitive)
    for pattern in search_patterns:
        pattern_norm = normalize_col_name(pattern)
        for col in data.keys():
            col_norm = normalize_col_name(col)
            if col_norm == pattern_norm:
                val = data.get(col)
                if val is not None and safe_float(val) != 0:
                    return safe_float(val)
    
    # Por último, busca parcial
    for pattern in search_patterns:
        pattern_norm = normalize_col_name(pattern)
        for col in data.keys():
            col_norm = normalize_col_name(col)
            if pattern_norm in col_norm or col_norm in pattern_norm:
                val = data.get(col)
                if val is not None and safe_float(val) != 0:
                    return safe_float(val)
    
    return 0


def convert_percentage(value):
    """
    Converte valor decimal para percentual se necessário.
    Valores <= 3 são considerados decimais (ex: 0.92 = 92%, 1.5 = 150%)
    """
    if value is None or value == 0:
        return 0
    if value <= 3:
        return value * 100
    return value


def get_month_names():
    """
    Retorna os nomes dos meses M1, M2, M3 baseado no mês atual.
    Também retorna o mês de referência (2 meses atrás).
    Para os atingimentos por pilar:
    - pct_1 (último): 2 meses atrás
    - pct_2 (penúltimo): 3 meses atrás
    - pct_3 (antepenúltimo): 4 meses atrás
    """
    from datetime import datetime
    from dateutil.relativedelta import relativedelta
    
    meses_pt = {
        1: 'Janeiro', 2: 'Fevereiro', 3: 'Março', 4: 'Abril',
        5: 'Maio', 6: 'Junho', 7: 'Julho', 8: 'Agosto',
        9: 'Setembro', 10: 'Outubro', 11: 'Novembro', 12: 'Dezembro'
    }
    
    hoje = datetime.now()
    
    # Mês de referência: 2 meses atrás
    data_referencia = hoje - relativedelta(months=2)
    mes_referencia = meses_pt[data_referencia.month]
    ano_referencia = data_referencia.year
    
    # Meses para os atingimentos (pct_1, pct_2, pct_3)
    data_pct_1 = hoje - relativedelta(months=2)  # Último: 2 meses atrás
    data_pct_2 = hoje - relativedelta(months=3)  # Penúltimo: 3 meses atrás
    data_pct_3 = hoje - relativedelta(months=4)  # Antepenúltimo: 4 meses atrás
    
    # Meses antigos para gráficos (mantendo compatibilidade)
    def get_month_back(months_back):
        m = hoje.month - months_back
        y = hoje.year
        while m <= 0:
            m += 12
            y -= 1
        return m
    
    return {
        'm1': meses_pt[get_month_back(2)],
        'm2': meses_pt[get_month_back(3)],
        'm3': meses_pt[get_month_back(4)],
        'mes_referencia': mes_referencia,
        'ano_referencia': ano_referencia,
        # Novos: meses para atingimentos
        'mes_pct_1': meses_pt[data_pct_1.month],
        'mes_pct_2': meses_pt[data_pct_2.month],
        'mes_pct_3': meses_pt[data_pct_3.month],
    }


def fetch_iq_data(user_name):
    """
    Busca dados de IQ MÓVEL e IQ FIXA da planilha ÍNDICE DE QUALIDADE.
    Sheet: ÍNDICE DE QUALIDADE
    Usa a primeira linha como cabeçalho.
    Coluna de busca: COLABORADOR
    Retorna: {'iq_movel': float, 'iq_fixa': float} em percentual (ex: 0.05 = 5%)
    """
    cache_key = f"iq_data_{user_name.replace(' ', '_')}"
    cached_data = cache.get(cache_key)
    
    if cached_data:
        return cached_data
    
    iq_result = {'iq_movel': 0, 'iq_fixa': 0}
    
    def normalize_text(text):
        if pd.isna(text):
            return ""
        return str(text).strip().upper()
    
    user_name_normalized = normalize_text(user_name)
    
    # Obter URLs dinâmicas
    urls = get_excel_urls()
    
    try:
        # Baixar planilha de comissão
        excel_file, error = download_excel_file(urls['comissao'], "comissao_iq")
        if error:
            print(f"[fetch_iq_data] Erro ao baixar planilha: {error}")
            return iq_result
        
        # Ler sheet ÍNDICE DE QUALIDADE com primeira linha como cabeçalho
        try:
            excel_file.seek(0)
            df = pd.read_excel(excel_file, sheet_name='ÍNDICE DE QUALIDADE', header=0, engine='openpyxl')
            print(f"[fetch_iq_data] Colunas: {list(df.columns)}")
            print(f"[fetch_iq_data] Total linhas: {len(df)}")
            if len(df) > 0:
                print(f"[fetch_iq_data] Primeira linha: {df.iloc[0].to_dict()}")
        except Exception as e:
            print(f"[fetch_iq_data] Erro ao ler sheet ÍNDICE DE QUALIDADE: {e}")
            return iq_result
        
        # Encontrar colunas necessárias - IQ Móvel (coluna 9) e IQ Fixa (coluna 13)
        colaborador_col = None
        iq_movel_col = None
        iq_fixa_col = None
        
        for col in df.columns:
            col_str = str(col).strip()
            col_upper = col_str.upper()
            if 'COLABORADOR' in col_upper or col_upper == 'NOME':
                colaborador_col = col
            # Buscar colunas IQ de forma flexível (pode ter espaços, acentos diferentes, etc)
            # Usar IQ Móvel (coluna 9) e IQ Fixa (coluna 13) - são as colunas corretas
            # Não usar as colunas 2 e 3 que são da loja
            elif 'IQ' in col_upper and ('MOVEL' in col_upper or 'MÓVEL' in col_upper):
                # Verificar se não é da loja (evitar colunas 2 e 3)
                if not iq_movel_col or ('LOJA' not in col_upper and 'PDV' not in col_upper):
                    iq_movel_col = col
            elif 'IQ' in col_upper and 'FIXA' in col_upper:
                # Verificar se não é da loja (evitar colunas 2 e 3)
                if not iq_fixa_col or ('LOJA' not in col_upper and 'PDV' not in col_upper):
                    iq_fixa_col = col
        
        print(f"[fetch_iq_data] colaborador_col={colaborador_col}, iq_movel_col={iq_movel_col}, iq_fixa_col={iq_fixa_col}")
        
        if colaborador_col is None:
            print(f"[fetch_iq_data] Coluna COLABORADOR não encontrada")
            return iq_result
        
        # Buscar usuário
        for idx, row in df.iterrows():
            row_name = normalize_text(row.get(colaborador_col, ''))
            
            # Match exato ou parcial
            match = False
            if row_name == user_name_normalized:
                match = True
            elif user_name_normalized in row_name or row_name in user_name_normalized:
                match = True
            
            if match:
                if iq_movel_col:
                    val = row.get(iq_movel_col)
                    if pd.notna(val):
                        try:
                            iq_result['iq_movel'] = float(val)
                        except:
                            pass
                
                if iq_fixa_col:
                    val = row.get(iq_fixa_col)
                    if pd.notna(val):
                        try:
                            iq_result['iq_fixa'] = float(val)
                        except:
                            pass
                
                print(f"[fetch_iq_data] Encontrado: {row_name}, IQ Móvel={iq_result['iq_movel']}, IQ Fixa={iq_result['iq_fixa']}")
                break
        
        # Cache por 5 minutos
        cache.set(cache_key, iq_result, 300)
        
        return iq_result
        
    except Exception as e:
        print(f"[fetch_iq_data] Erro geral: {e}")
        return iq_result


def fetch_vendedores_por_filial(user_sector):
    """
    Busca todos os vendedores e suas comissões por pilar da BASE_PAGAMENTO para uma filial.
    Busca também os percentuais de atingimento da planilha REMUNERAÇÃO CN.
    Retorna no formato esperado pelo template (com lista de pilares).
    """
    if not user_sector:
        return []
    
    # Normalizar nome do setor
    user_sector_normalized = str(user_sector).strip().upper()
    prefixes = ['LOJA ', 'LOJA_', 'PDV ', 'PDV_']
    for prefix in prefixes:
        if user_sector_normalized.startswith(prefix):
            user_sector_normalized = user_sector_normalized[len(prefix):]
    
    # Mapeamento de pilares - ordem importa, mais específico primeiro
    pilar_mapping = {
        'MÓVEL': 'movel', 'MOVEL': 'movel', 'MOV': 'movel',
        'FIXA': 'fixa', 'FIX': 'fixa',
        'SMARTPHONE': 'smartphone', 'SMART': 'smartphone', 'SMARTP': 'smartphone',
        'ELETRONICOS': 'eletronicos', 'ELETRÔNICOS': 'eletronicos', 'ELETRO': 'eletronicos', 'ELETRÔNICO': 'eletronicos',
        'ESSENCIAIS': 'essenciais', 'ESSEN': 'essenciais', 'ESSENCIAL': 'essenciais',
        'SEGURO': 'seguro', 'SEGUROS': 'seguro', 'SEG': 'seguro',
        'SVA': 'sva',
    }
    
    # Ordem dos pilares para o template
    pilares_ordem = ['movel', 'fixa', 'smartphone', 'eletronicos', 'essenciais', 'seguro', 'sva']
    pilares_nomes = {
        'movel': 'Móvel', 'fixa': 'Fixa', 'smartphone': 'Smartphone',
        'eletronicos': 'Eletrônicos', 'essenciais': 'Essenciais', 'seguro': 'Seguro', 'sva': 'SVA'
    }
    
    # Mapeamento de colunas %ATING_ da planilha REMUNERAÇÃO CN
    ating_cols_map = {
        'movel': ['%ATING_MOVEL_1', '%ATING_MÓVEL', 'ATING_MOVEL', '%MOVEL'],
        'fixa': ['%ATING_FIXA_1', 'ATING_FIXA', '%FIXA'],
        'smartphone': ['%ATING_SMART_1', 'ATING_SMART', '%SMART', '%ATING_SMARTPHONE'],
        'eletronicos': ['%ATING_ELETRO_1', 'ATING_ELETRO', '%ELETRO', '%ATING_ELETRONICOS'],
        'essenciais': ['%ATING_ESSEN_1', 'ATING_ESSEN', '%ESSEN', '%ATING_ESSENCIAIS'],
        'seguro': ['%ATING_SEG_1', 'ATING_SEG', '%SEG', '%ATING_SEGURO'],
        'sva': ['%ATING_SVA_1', 'ATING_SVA', '%SVA'],
    }
    
    # Obter URLs dinâmicas
    urls = get_excel_urls()
    
    try:
        # 1. Buscar dados da planilha REMUNERAÇÃO CN (percentuais de atingimento)
        excel_file_rem, error = download_excel_file(urls['comissao'], "remuneracao_cn")
        ating_por_vendedor = {}  # {vendedor_nome_upper: {pilar: pct}}
        remuneracao_por_vendedor = {}  # {vendedor_nome_upper: remuneracao_final}
        
        if not error:
            try:
                excel_file_rem.seek(0)
                df_rem = pd.read_excel(excel_file_rem, sheet_name='REMUNERAÇÃO CN', engine='openpyxl')
                
                # Encontrar coluna de nome
                nome_col_rem = None
                for col in df_rem.columns:
                    if 'NOME' in str(col).upper():
                        nome_col_rem = col
                        break
                
                # Encontrar coluna PDV para filtrar por filial
                pdv_col_rem = None
                for col in df_rem.columns:
                    if 'PDV' in str(col).upper() or 'FILIAL' in str(col).upper():
                        pdv_col_rem = col
                        break
                
                # Encontrar coluna de remuneração final
                rem_final_col = None
                for col in df_rem.columns:
                    col_upper = str(col).upper()
                    if 'REMUNERAÇÃO' in col_upper or 'REMUNERACAO' in col_upper:
                        if 'FINAL' in col_upper or 'TOTAL' in col_upper:
                            rem_final_col = col
                            break
                if not rem_final_col:
                    for col in df_rem.columns:
                        col_upper = str(col).upper()
                        if 'REMUNERAÇÃO' in col_upper or 'REMUNERACAO' in col_upper:
                            rem_final_col = col
                            break
                
                # Mapear colunas de atingimento - PRIORIZAR sempre colunas com _1
                ating_cols = {}
                for pilar_key, possible_cols in ating_cols_map.items():
                    # Primeira tentativa: buscar apenas colunas com _1
                    for col in df_rem.columns:
                        col_upper = str(col).upper().replace(' ', '_')
                        # Priorizar colunas que terminam com _1
                        if '_1' in col_upper:
                            for possible in possible_cols:
                                if '_1' in possible.upper() and (possible.upper() in col_upper or col_upper == possible.upper()):
                                    ating_cols[pilar_key] = col
                                    break
                        if pilar_key in ating_cols:
                            break
                    
                    # Segunda tentativa: se não encontrou com _1, buscar outras variações
                    if pilar_key not in ating_cols:
                        for col in df_rem.columns:
                            col_upper = str(col).upper().replace(' ', '_')
                            for possible in possible_cols:
                                if possible.upper() in col_upper or col_upper == possible.upper():
                                    ating_cols[pilar_key] = col
                                    break
                            if pilar_key in ating_cols:
                                break
                
                # Filtrar por filial se possível
                if pdv_col_rem:
                    df_rem['pdv_norm'] = df_rem[pdv_col_rem].astype(str).str.strip().str.upper()
                    # Filtrar por filial
                    mask = df_rem['pdv_norm'].str.contains(user_sector_normalized, case=False, na=False)
                    df_rem_filial = df_rem[mask]
                    if df_rem_filial.empty:
                        df_rem_filial = df_rem  # Se não encontrar, usar todos
                else:
                    df_rem_filial = df_rem
                
                # Extrair dados de cada vendedor
                for idx, row in df_rem_filial.iterrows():
                    if nome_col_rem:
                        nome = str(row[nome_col_rem]).strip().upper() if pd.notna(row[nome_col_rem]) else ''
                        if not nome:
                            continue
                        
                        # Percentuais de atingimento
                        ating_por_vendedor[nome] = {}
                        for pilar_key, col in ating_cols.items():
                            val = row[col]
                            if pd.notna(val):
                                try:
                                    pct = float(val)
                                    # Se valor < 3, provavelmente está em decimal (0.95 = 95%)
                                    if pct < 3:
                                        pct = pct * 100
                                    ating_por_vendedor[nome][pilar_key] = pct
                                except (ValueError, TypeError):
                                    pass
                        
                        # Remuneração final
                        if rem_final_col and pd.notna(row[rem_final_col]):
                            try:
                                remuneracao_por_vendedor[nome] = float(row[rem_final_col])
                            except (ValueError, TypeError):
                                pass
                
            except Exception as e:
                pass  # Silently handle errors reading REMUNERAÇÃO CN
        
        # 2. Buscar dados de comissão da BASE_PAGAMENTO
        excel_file, error = download_excel_file(urls['base_pagamento'], "base_pagamento")
        if error:
            return []
        
        excel_file.seek(0)
        df = pd.read_excel(excel_file, sheet_name='Planilha1', engine='openpyxl')
        
        # Colunas
        colunas_lower = {col.lower(): col for col in df.columns}
        filial_col = colunas_lower.get('filial') or df.columns[df.columns.str.lower().str.contains('filial')].tolist()[0] if any(df.columns.str.lower().str.contains('filial')) else None
        vendedor_col = colunas_lower.get('vendedor') or df.columns[df.columns.str.lower().str.contains('vendedor')].tolist()[0] if any(df.columns.str.lower().str.contains('vendedor')) else None
        receita_col = colunas_lower.get('receita') or df.columns[df.columns.str.lower().str.contains('receita')].tolist()[0] if any(df.columns.str.lower().str.contains('receita')) else None
        pilar_col = colunas_lower.get('pilar') or df.columns[df.columns.str.lower().str.contains('pilar')].tolist()[0] if any(df.columns.str.lower().str.contains('pilar')) else None
        
        if not all([filial_col, vendedor_col, receita_col, pilar_col]):
            return []
        
        # Filtrar por filial exata
        df['filial_norm'] = df[filial_col].astype(str).str.strip().str.upper()
        df_filial = df[df['filial_norm'] == user_sector_normalized]
        
        if df_filial.empty:
            return []
        
        # Agrupar por vendedor e pilar
        vendedores = {}
        for idx, row in df_filial.iterrows():
            vendedor_nome = str(row[vendedor_col]).strip().upper() if pd.notna(row[vendedor_col]) else ''
            if not vendedor_nome:
                continue
                
            if vendedor_nome not in vendedores:
                vendedores[vendedor_nome] = {
                    'nome': vendedor_nome.title(),
                    'nome_upper': vendedor_nome,
                    'pilares_valores': {p: 0 for p in pilares_ordem},
                    'comissao_total': 0,
                    'remuneracao_final': 0,
                }
            
            # Identificar pilar
            pilar_val = str(row[pilar_col]).strip().upper() if pd.notna(row[pilar_col]) else ''
            pilar_key = None
            
            # Tentarmatch mais preciso primeiro
            if pilar_val in pilar_mapping:
                pilar_key = pilar_mapping[pilar_val]
            else:
                # Tentativa mais flexível - buscar parte do texto
                for key, value in pilar_mapping.items():
                    if key in pilar_val:
                        pilar_key = value
                        break
            
            # Debug quando não encontrar pilar
            if not pilar_key:
                print(f"[DEBUG] Pilar não mapeado: '{pilar_val}' para vendedor {vendedor_nome}")
            
            # Somar receita
            if pilar_key:
                receita = row[receita_col]
                if pd.notna(receita):
                    try:
                        valor = float(receita)
                        vendedores[vendedor_nome]['pilares_valores'][pilar_key] += valor
                        vendedores[vendedor_nome]['comissao_total'] += valor
                    except (ValueError, TypeError):
                        pass
        
        # Converter para formato esperado pelo template
        result = []
        for vendedor_nome, dados in vendedores.items():
            nome_upper = dados['nome_upper']
            
            # Buscar percentual de atingimento - tentar match exato primeiro, depois fuzzy
            vendedor_ating = None
            if nome_upper in ating_por_vendedor:
                vendedor_ating = ating_por_vendedor[nome_upper]
            else:
                # Tentar match parcial/fuzzy
                for ating_nome in ating_por_vendedor.keys():
                    # Remover espaços extras e comparar
                    nome_clean = ' '.join(nome_upper.split())
                    ating_clean = ' '.join(ating_nome.split())
                    if nome_clean == ating_clean:
                        vendedor_ating = ating_por_vendedor[ating_nome]
                        break
                    # Comparar primeiras palavras (nome e sobrenome)
                    nome_parts = nome_clean.split()
                    ating_parts = ating_clean.split()
                    if len(nome_parts) >= 2 and len(ating_parts) >= 2:
                        if nome_parts[0] == ating_parts[0] and nome_parts[-1] == ating_parts[-1]:
                            vendedor_ating = ating_por_vendedor[ating_nome]
                            break
            
            # Criar lista de pilares com percentual de ATINGIMENTO
            pilares = []
            for p in pilares_ordem:
                valor = dados['pilares_valores'][p]
                # Buscar percentual de atingimento da planilha REMUNERAÇÃO CN
                pct = 0
                if vendedor_ating and p in vendedor_ating:
                    pct = vendedor_ating[p]
                
                pilares.append({
                    'nome': pilares_nomes[p],
                    'pct': pct,
                    'comissao': valor,
                })
            
            # Remuneração final da planilha REMUNERAÇÃO CN ou usar comissao_total
            remuneracao = remuneracao_por_vendedor.get(nome_upper, dados['comissao_total'])
            
            # Buscar user_id no Django para o link "Ver detalhes"
            user_id = None
            nome_parts = nome_upper.split()
            if len(nome_parts) >= 2:
                user_obj = User.objects.filter(
                    first_name__iexact=nome_parts[0],
                    last_name__icontains=nome_parts[-1],
                    is_active=True
                ).first()
                if user_obj:
                    user_id = user_obj.id
            if not user_id and len(nome_parts) >= 1:
                user_obj = User.objects.filter(
                    first_name__iexact=nome_parts[0],
                    is_active=True
                ).first()
                if user_obj:
                    user_id = user_obj.id
            
            result.append({
                'nome': dados['nome'],
                'user_id': user_id,
                'pilares': pilares,
                'comissao_total': dados['comissao_total'],
                'remuneracao_final': remuneracao,
                # Campos individuais para o gráfico de representatividade
                'movel_comissao': dados['pilares_valores']['movel'],
                'fixa_comissao': dados['pilares_valores']['fixa'],
                'smartphone_comissao': dados['pilares_valores']['smartphone'],
                'eletronicos_comissao': dados['pilares_valores']['eletronicos'],
                'essenciais_comissao': dados['pilares_valores']['essenciais'],
                'seguro_comissao': dados['pilares_valores']['seguro'],
                'sva_comissao': dados['pilares_valores']['sva'],
            })
        
        return result
        
    except Exception as e:
        return []


def process_commission_data(data, is_gerente=False, metas_pilar=None, iq_data=None):
    """
    Processa os dados brutos da planilha e organiza em seções.
    metas_pilar: dict com {pilar_key: {'total': x, 'pago': y, 'exclusao': z}}
    iq_data: dict com {'iq_movel': float, 'iq_fixa': float} em percentual decimal
    """
    if metas_pilar is None:
        metas_pilar = {}
    if iq_data is None:
        iq_data = {'iq_movel': 0, 'iq_fixa': 0}
    
    # DEBUG: Mostrar todas as colunas disponíveis
    print(f"[DEBUG process_commission_data] Colunas disponíveis: {list(data.keys())[:30]}")  # Primeiras 30 colunas
    
    # DEBUG: Buscar colunas que contenham 'COM_' e 'SMARTPHONE' ou 'ELETRO'
    colunas_com = [col for col in data.keys() if 'COM_' in str(col).upper()]
    print(f"[DEBUG process_commission_data] Colunas COM_: {colunas_com}")
    
    processed = {
        'info': {
            'nome': data.get('NOME', ''),
            'cargo': data.get('CARGO', ''),
            'pdv': data.get('PDV', ''),
            'coordenador': data.get('COORDENADOR (A)', ''),
        },
        'pilares': [],
        'comissoes': {},
        'bonus': {},
        'alto_desempenho': {},
        'hunter': {},
        'aceleradores': {},
        'descontos': {},
        'totais': {},
        'charts': {},
        'iq': iq_data,  # Adicionar dados de IQ
    }
    
    # Mapeamento de chave para metas_pilar
    pilar_to_meta_key = {
        'MOVEL': 'movel',
        'FIXA': 'fixa',
        'SMART': 'smartphone',
        'ELETRO': 'eletronicos',
        'ESSEN': 'essenciais',
        'SEG': 'seguro',
        'SVA': 'sva',
    }
    
    # Pilares com atingimentos
    pilares_config = [
        {'nome': 'Móvel', 'key': 'MOVEL', 'icon': '📱', 'color': '#660099'},
        {'nome': 'Fixa', 'key': 'FIXA', 'icon': '🏠', 'color': '#9B30FF'},
        {'nome': 'Smartphone', 'key': 'SMART', 'icon': '📲', 'color': '#BA55D3'},
        {'nome': 'Eletrônicos', 'key': 'ELETRO', 'icon': '💻', 'color': '#9370DB'},
        {'nome': 'Essenciais', 'key': 'ESSEN', 'icon': '⭐', 'color': '#8A2BE2'},
        {'nome': 'Seguro', 'key': 'SEG', 'icon': '🛡️', 'color': '#7B68EE'},
        {'nome': 'SVA', 'key': 'SVA', 'icon': '📦', 'color': '#6A5ACD'},
    ]
    
    for pilar in pilares_config:
        key = pilar['key']
        cart_key = f'ATING_CART_{key}' if is_gerente else f'ATING_{key}_PDV'
        
        # DEBUG: Mostrar colunas disponíveis relacionadas a ATING
        ating_cols_debug = [col for col in data.keys() if 'ATING' in str(col).upper() or '%' in str(col)]
        if not hasattr(process_commission_data, '_debug_ating_shown'):
            print(f"[DEBUG ATING COLS] Colunas com ATING ou %: {ating_cols_debug}")
            process_commission_data._debug_ating_shown = True
        
        # Padrões para buscar pct_3 (antepenúltimo mês)
        patterns_3 = [
            f'%ATING_{key}_3', f'%ATING_{key} 3', f'ATING_{key}_3',
            f'%ATING {key}_3', f'% ATING_{key}_3', f'%ATING_{key}_M3',
            f'ATING_{key}_M3', f'%{key}_3', f'%{key} 3'
        ]
        # Padrões para buscar pct_2 (penúltimo mês)
        patterns_2 = [
            f'%ATING_{key}_2', f'%ATING_{key} 2', f'ATING_{key}_2',
            f'%ATING {key}_2', f'% ATING_{key}_2', f'%ATING_{key}_M2',
            f'ATING_{key}_M2', f'%{key}_2', f'%{key} 2'
        ]
        # Padrões para buscar pct_1 (último mês / atual)
        patterns_1 = [
            f'%ATING_{key}_1', f'%ATING_{key} 1', f'ATING_{key}_1',
            f'%ATING {key}_1', f'% ATING_{key}_1', f'%ATING_{key}',
            f'%ATING_{key}_M1', f'ATING_{key}_M1', f'%{key}_1',
            f'%{key} 1', f'%{key}'
        ]
        
        # Usar função flexível para buscar valores
        pct_3_raw = find_column_value(data, patterns_3)
        pct_2_raw = find_column_value(data, patterns_2)
        pct_1_raw = find_column_value(data, patterns_1)
        
        # DEBUG: Mostrar valores encontrados
        print(f"[DEBUG Ating] {key}: pct_1={pct_1_raw}, pct_2={pct_2_raw}, pct_3={pct_3_raw} (is_gerente={is_gerente})")
        if pct_1_raw == 0 and pct_2_raw == 0 and pct_3_raw == 0:
            print(f"[DEBUG Ating ZERO] Todas as tentativas de busca para {key} retornaram 0")
        
        # Buscar metas (Total, Pago, Exclusão) deste pilar
        meta_key = pilar_to_meta_key.get(key, key.lower())
        pilar_metas = metas_pilar.get(meta_key, {'total': 0, 'pago': 0, 'exclusao': 0})
        
        pilar_data = {
            'nome': pilar['nome'],
            'icon': pilar['icon'],
            'color': pilar['color'],
            'pct_3': convert_percentage(pct_3_raw),
            'pct_2': convert_percentage(pct_2_raw),
            'pct_1': convert_percentage(pct_1_raw),
            'carteira': safe_float(data.get(cart_key) or data.get(f'ATING_CART_{key}')),
            'habilitado': data.get(f'H_{pilar["nome"].upper()}') or data.get(f'H_{key}'),
            # Dados de metas: Total, Pago, Exclusão
            'total': pilar_metas.get('total', 0),
            'pago': pilar_metas.get('pago', 0),
            'exclusao': pilar_metas.get('exclusao', 0),
        }
        
        ating_values = [pilar_data['pct_3'], pilar_data['pct_2'], pilar_data['pct_1']]
        valid_values = [v for v in ating_values if v > 0]
        pilar_data['media'] = sum(valid_values) / len(valid_values) if valid_values else 0
        
        processed['pilares'].append(pilar_data)
    
    # Comissões por pilar - tentar múltiplas variações de nomes
    comissoes_valores = {
        'movel': safe_float(data.get('COM_MÓVEL') or data.get('COM_MOVEL')),
        'fixa': safe_float(data.get('COM_FIXA')),
        'smartphone': safe_float(data.get('COM_SMARTPHONE') or data.get('COM_SMART')),
        'eletronicos_a': safe_float(data.get('COM_ELETRONICOS - A') or data.get('COM_ELETRÔNICOS - A') or data.get('COM_ELETRO - A') or data.get('COM_ELETRONICOS-A') or data.get('COM_ELETRONICOS A')),
        'eletronicos_b': safe_float(data.get('COM_ELETRONICOS - B') or data.get('COM_ELETRÔNICOS - B') or data.get('COM_ELETRO - B') or data.get('COM_ELETRONICOS-B') or data.get('COM_ELETRONICOS B')),
        'essenciais_a': safe_float(data.get('COM_ESSENCIAIS - A') or data.get('COM_ESSEN - A') or data.get('COM_ESSENCIAIS-A') or data.get('COM_ESSENCIAIS A')),
        'essenciais_b': safe_float(data.get('COM_ESSENCIAIS - B') or data.get('COM_ESSEN - B') or data.get('COM_ESSENCIAIS-B') or data.get('COM_ESSENCIAIS B')),
        'seguro': safe_float(data.get('COM_SEGURO') or data.get('COM_SEG')),
        'sva': safe_float(data.get('COM_SVA')),
    }
    
    # DEBUG: Mostrar valores de comissões encontrados
    print(f"[DEBUG Comissões] smartphone={comissoes_valores['smartphone']}, eletronicos_a={comissoes_valores['eletronicos_a']}, eletronicos_b={comissoes_valores['eletronicos_b']}")
    
    total_comissoes = safe_float(data.get('Total Comissão'))
    if total_comissoes == 0:
        total_comissoes = sum(comissoes_valores.values())
    comissoes_valores['total'] = total_comissoes
    processed['comissoes'] = comissoes_valores
    
    # Bônus Carteira - nomes exatos das colunas da planilha REMUNERAÇÃO GERENTE/CN
    bonus_valores = {
        'movel': safe_float(data.get('BONUS_CARTEIRA_MÓVEL')),
        'fixa': safe_float(data.get('BONU_CARTEIRA_FIXA')),  # Nota: BONU sem S (nome real na planilha)
        'smartphone': safe_float(data.get('BONUTS_CARTEIRA_SMARTPHONE')),
        'eletronicos_a': safe_float(data.get('BONUS_CARTEIRA_ELETRONICOS - A')),
        'eletronicos_b': safe_float(data.get('BONUS_CARTEIRA_ELETRONICOS - B')),
        'essenciais_a': safe_float(data.get('BONUS_CARTEIRA_ESSENCIAIS - A')),
        'essenciais_b': safe_float(data.get('BONUS_CARTEIRA_ESSENCIAIS - B')),
        'seguro': safe_float(data.get('BONUS_CARTEIRA_SEGURO')),
        'sva': safe_float(data.get('BONUS_CARTEIRA_SVA')),
    }
    # Juntar eletrônicos e essenciais em valores únicos para exibição
    bonus_valores['eletronicos'] = bonus_valores['eletronicos_a'] + bonus_valores['eletronicos_b']
    bonus_valores['essenciais'] = bonus_valores['essenciais_a'] + bonus_valores['essenciais_b']
    
    # Total Bônus Carteira = TOTAL BONUS LOJA (soma dos valores individuais de bônus carteira)
    total_bonus = safe_float(data.get('TOTAL BONUS LOJA'))
    if total_bonus == 0:
        total_bonus = sum([bonus_valores['movel'], bonus_valores['fixa'], bonus_valores['smartphone'], 
                          bonus_valores['eletronicos'], bonus_valores['essenciais'], 
                          bonus_valores['seguro'], bonus_valores['sva']])
    bonus_valores['total'] = total_bonus
    processed['bonus'] = bonus_valores
    
    # Alto Desempenho
    alto_desempenho_valores = {
        'movel': safe_float(data.get('ALTO_DESEM_MÓVEL')),
        'fixa': safe_float(data.get('ALTO_DESEM_FIXA')),
        'smartphone': safe_float(data.get('ALTO_DESEM_SMARTPHONE')),
        'eletronicos_a': safe_float(data.get('ALTO_DESEM_ELETRONICOS - A')),
        'eletronicos_b': safe_float(data.get('ALTO_DESEM_ELETRONICOS - B')),
        'essenciais_a': safe_float(data.get('ALTO_DESEM_ESSENCIAIS - A')),
        'essenciais_b': safe_float(data.get('ALTO_DESEM_ESSENCIAIS - B')),
        'seguro': safe_float(data.get('ALTO_DESEM_SEGURO')),
        'sva': safe_float(data.get('ALTO_DESEM_SVA')),
    }
    total_alto_desempenho = safe_float(data.get('TOTAL PREMIAÇÃO ALTO DESEMPENHO'))
    if total_alto_desempenho == 0:
        total_alto_desempenho = sum(alto_desempenho_valores.values())
    alto_desempenho_valores['total'] = total_alto_desempenho
    processed['alto_desempenho'] = alto_desempenho_valores
    
    # Hunter - valores monetários das colunas HUNTER_
    processed['hunter'] = {
        'movel': safe_float(data.get('HUNTER_MOVEL') or data.get('HUNTER_MÓVEL')),
        'fixa': safe_float(data.get('HUNTER_FIXA')),
        'smartphone': safe_float(data.get('HUNTER_SMARTPHONE')),
        'eletronicos': safe_float(data.get('HUNTER_ELETRONICOS') or data.get('HUNTER_ELETRÔNICOS')),
        'essenciais': safe_float(data.get('HUNTER_ESSENCIAIS')),
        'seguro': safe_float(data.get('HUNTER_SEGUROS') or data.get('HUNTER_SEGURO')),
        'sva': safe_float(data.get('HUNTER_SVA')),
    }
    # Calcular total do Hunter
    processed['hunter']['total'] = sum([
        processed['hunter']['movel'],
        processed['hunter']['fixa'],
        processed['hunter']['smartphone'],
        processed['hunter']['eletronicos'],
        processed['hunter']['essenciais'],
        processed['hunter']['seguro'],
        processed['hunter']['sva'],
    ])
    
    # Habilitação Hunter (H_MÓVEL, H_FIXA, etc.) - indica se hunter está ativo
    # Estes são valores da planilha que indicam habilitação
    bonus_hunter_valores = {
        'movel': data.get('H_MÓVEL') or data.get('H_MOVEL'),
        'fixa': data.get('H_FIXA'),
        'smartphone': data.get('H_SMARTPHONE'),
        'eletronicos': data.get('H_ELETRONICOS') or data.get('H_ELETRÔNICOS'),
        'essenciais': data.get('H_ESSENCIAIS'),
        'seguro': data.get('H_SEGURO'),
        'sva': data.get('H_SVA'),
    }
    # Se H_ tem valor numérico, somar; senão é apenas flag
    total_h = 0
    for key, val in bonus_hunter_valores.items():
        try:
            if val and isinstance(val, (int, float)):
                total_h += float(val)
        except:
            pass
    bonus_hunter_valores['total'] = total_h
    processed['bonus_hunter'] = bonus_hunter_valores
    
    # Aceleradores e IQ
    processed['aceleradores'] = {
        'seis_pilares': safe_float(data.get('ACELERADOR_6 PILARES')),
        'total_aceleradores': safe_float(data.get('REMUNERAÇÃO TOTAL + ACELERADORES')),
        'iq_acelerador_movel': safe_float(data.get('IQ_ACELADOR MÓVEL')),
        'iq_acelerador_fixa': safe_float(data.get('IQ_ACELERADOR FIXA')),
        'iq_deflator_movel': safe_float(data.get('IQ_DEFLATOR MÓVEL')),
        'iq_deflator_fixa': safe_float(data.get('IQ_DEFLATOR FIXA')),
        'remuneracao_iq': safe_float(data.get('REMUNERÇÃO COM IQ')),
    }
    
    # Descontos
    processed['descontos'] = {
        'advertencia': safe_float(data.get('DESCONTO_ADVERTÊNCIA')),
        'excedente': safe_float(data.get('DESC_PAGAMENTO EXCEDENTE DO MÊS ANTERIOR')),
        'price': safe_float(data.get('DESCONTO_PRICE')),
        'total': safe_float(data.get('DESCONTO_TOTAL')),
    }
    
    # Totais
    if is_gerente:
        # Buscar REMUNERAÇÃO FINAL TOTAL de forma flexível
        remun_final_total = 0
        for key_name in data.keys():
            key_upper = str(key_name).upper()
            if 'REMUNERA' in key_upper and 'FINAL' in key_upper and 'TOTAL' in key_upper:
                remun_final_total = safe_float(data.get(key_name))
                if remun_final_total > 0:
                    print(f"[DEBUG] Remuneração Final Total encontrada em: '{key_name}' = {remun_final_total}")
                    break
        
        # Se não encontrou, tentar as variações fixas
        if remun_final_total == 0:
            remun_final_total = (
                safe_float(data.get('REMUNERAÇÃO FINAL TOTAL')) or
                safe_float(data.get('REMUNERACAO FINAL TOTAL')) or
                safe_float(data.get('REMUNERAÇÃO_FINAL_TOTAL')) or
                safe_float(data.get('REMUNERACAO_FINAL_TOTAL'))
            )
        
        processed['totais'] = {
            'comissao_cargo': safe_float(data.get('COMISSÃO_GERENTE') or data.get('COMISSÃO GERENTE')),
            'premiacao_cargo': safe_float(data.get('PREMIAÇÃO_GERENTE') or data.get('PREMIAÇÃO GERENTE')),
            'iq_cargo': safe_float(data.get('IQ_GERENTE') or data.get('IQ GERENTE')),
            'remuneracao_cargo': (
                safe_float(data.get('REMUNERAÇÃO FINAL GERENTE')) or
                safe_float(data.get('REMUNERAÇÃO_FINAL_GERENTE')) or
                safe_float(data.get('REMUNERACAO FINAL GERENTE'))
            ),
            'total_comissao_premiacao': (
                safe_float(data.get('TOTAL COMISSÃO + PREMIAÇÃO')) or
                safe_float(data.get('TOTAL COMISSAO + PREMIACAO'))
            ),
            'remuneracao_final': remun_final_total,
            'tfp_movel': safe_float(data.get('TFP MÓVEL') or data.get('TFP MOVEL')),
            'tfp_fixa': safe_float(data.get('TFP FIXA')),
        }
        processed['totais']['cargo_label'] = 'Gerente'
    else:
        # DEBUG: Mostrar todas as colunas disponíveis para CN
        print(f"[DEBUG CN] Colunas disponíveis: {list(data.keys())}")
        
        processed['totais'] = {
            'comissao_cargo': safe_float(data.get('COMISSÃO_CN') or data.get('COMISSÃO CN')),
            'premiacao_cargo': safe_float(data.get('PREMIAÇÃO_CN') or data.get('PREMIAÇÃO CN')),
            'iq_cargo': safe_float(data.get('CN_PLENO') or data.get('CN PLENO')),
            'remuneracao_cargo': (
                safe_float(data.get('REMUNERAÇÃO_FINAL')) or
                safe_float(data.get('REMUNERAÇÃO FINAL')) or
                safe_float(data.get('REMUNERACAO_FINAL')) or
                safe_float(data.get('REMUNERACAO FINAL'))
            ),
            'total_comissao_premiacao': (
                safe_float(data.get('TOTAL COMISSÃO + PREMIAÇÃO')) or
                safe_float(data.get('TOTAL COMISSAO + PREMIACAO'))
            ),
        }
        
        # Buscar REMUNERAÇÃO FINAL de forma flexível para CN
        remun_final_cn = 0
        for key_name in data.keys():
            key_upper = str(key_name).upper()
            if 'REMUNERA' in key_upper and 'FINAL' in key_upper:
                remun_final_cn = safe_float(data.get(key_name))
                if remun_final_cn > 0:
                    print(f"[DEBUG CN] Remuneração Final encontrada em: '{key_name}' = {remun_final_cn}")
                    break
        
        # Se não encontrou, tentar as variações fixas
        if remun_final_cn == 0:
            remun_final_cn = (
                safe_float(data.get('REMUNERAÇÃO_FINAL_TOTAL')) or
                safe_float(data.get('REMUNERAÇÃO FINAL TOTAL')) or
                safe_float(data.get('REMUNERACAO_FINAL_TOTAL')) or
                safe_float(data.get('REMUNERACAO FINAL TOTAL')) or
                safe_float(data.get('REMUNERAÇÃO FINAL')) or
                safe_float(data.get('REMUNERACAO FINAL'))
            )
        
        processed['totais']['remuneracao_final'] = remun_final_cn
        processed['totais'].update({
            'tfp_movel': safe_float(data.get('TFP MÓVEL') or data.get('TFP MOVEL')),
            'tfp_fixa': safe_float(data.get('TFP FIXA')),
            'cn_pleno': safe_float(data.get('CN PLENO') or data.get('CN_PLENO')),
            'cn_lider': safe_float(data.get('CN LÍDER') or data.get('CN LIDER') or data.get('CN_LIDER')),
            'campanha': safe_float(data.get('CAMPANHA')),
        })
        processed['totais']['cargo_label'] = 'CN'
        
        print(f"[DEBUG CN] Remuneração Final Total: {processed['totais']['remuneracao_final']}")
    
    # Calcular soma (comissões + bônus + alto desempenho + hunter) para cada pilar
    pilar_keys = ['movel', 'fixa', 'smartphone', 'eletronicos', 'essenciais', 'seguro', 'sva']
    for i, pilar_key in enumerate(pilar_keys):
        if i < len(processed['pilares']):
            # Comissão do pilar
            if pilar_key == 'eletronicos':
                com_valor = processed['comissoes'].get('eletronicos_a', 0) + processed['comissoes'].get('eletronicos_b', 0)
                ad_valor = processed['alto_desempenho'].get('eletronicos_a', 0) + processed['alto_desempenho'].get('eletronicos_b', 0)
            elif pilar_key == 'essenciais':
                com_valor = processed['comissoes'].get('essenciais_a', 0) + processed['comissoes'].get('essenciais_b', 0)
                ad_valor = processed['alto_desempenho'].get('essenciais_a', 0) + processed['alto_desempenho'].get('essenciais_b', 0)
            else:
                com_valor = processed['comissoes'].get(pilar_key, 0)
                ad_valor = processed['alto_desempenho'].get(pilar_key, 0)
            
            bonus_valor = processed['bonus'].get(pilar_key, 0)
            # Hunter agora usa 'seguro' consistentemente
            hunter_valor = processed['hunter'].get(pilar_key, 0)
            
            # Guardar a soma de comissões+bônus+alto+hunter em campo separado
            soma = com_valor + bonus_valor + ad_valor + hunter_valor
            processed['pilares'][i]['soma'] = soma
            
            # Calcular soma com IQ (IQ negativo = deflator/desconto, IQ positivo = acréscimo)
            # soma_com_iq = soma + (soma * iq_pct)
            if pilar_key == 'movel':
                iq_pct = iq_data.get('iq_movel', 0)
                ajuste_iq = soma * iq_pct
                processed['pilares'][i]['soma_com_iq'] = soma + ajuste_iq
                processed['pilares'][i]['iq_pct'] = iq_pct * 100  # Percentual para exibição (ex: -5 para -5%)
                processed['pilares'][i]['iq_multiplicador'] = 100 + (iq_pct * 100)  # Multiplicador (ex: 95 para -5%)
                processed['pilares'][i]['iq_valor'] = iq_data.get('iq_movel', 0)  # Valor bruto do IQ
                processed['pilares'][i]['ajuste_iq'] = ajuste_iq  # Valor do ajuste aplicado
            elif pilar_key == 'fixa':
                iq_pct = iq_data.get('iq_fixa', 0)
                ajuste_iq = soma * iq_pct
                processed['pilares'][i]['soma_com_iq'] = soma + ajuste_iq
                processed['pilares'][i]['iq_pct'] = iq_pct * 100  # Percentual para exibição (ex: -5 para -5%)
                processed['pilares'][i]['iq_multiplicador'] = 100 + (iq_pct * 100)  # Multiplicador (ex: 95 para -5%)
                processed['pilares'][i]['iq_valor'] = iq_data.get('iq_fixa', 0)  # Valor bruto do IQ
                processed['pilares'][i]['ajuste_iq'] = ajuste_iq  # Valor do ajuste aplicado
            else:
                # Outros pilares não têm IQ, soma_com_iq = soma
                processed['pilares'][i]['soma_com_iq'] = soma
                processed['pilares'][i]['iq_pct'] = None
                processed['pilares'][i]['iq_multiplicador'] = None
                processed['pilares'][i]['iq_valor'] = None
                processed['pilares'][i]['ajuste_iq'] = 0
    
    # Calcular totais de Pago e Exclusão (soma de todos os pilares)
    total_pago = sum(p['pago'] for p in processed['pilares'])
    total_exclusao = sum(p['exclusao'] for p in processed['pilares'])
    total_soma = sum(p.get('soma', 0) for p in processed['pilares'])
    total_soma_com_iq = sum(p.get('soma_com_iq', 0) for p in processed['pilares'])
    processed['totais']['total_pago'] = total_pago
    processed['totais']['total_exclusao'] = total_exclusao
    processed['totais']['total_pilares'] = total_soma
    processed['totais']['total_pilares_com_iq'] = total_soma_com_iq
    
    processed['habilitados'] = {
        'pilares_67': data.get('6/7PILARES'),
    }
    
    # Dados para gráficos
    processed['charts'] = {
        'atingimentos': {
            'labels': [p['nome'] for p in processed['pilares']],
            'data': [p['media'] for p in processed['pilares']],
            'colors': [p['color'] for p in processed['pilares']],
        },
        'comissoes': {
            'labels': ['Móvel', 'Fixa', 'Smart', 'Eletro', 'Essen', 'Seguro', 'SVA'],
            'data': [
                processed['comissoes']['movel'],
                processed['comissoes']['fixa'],
                processed['comissoes']['smartphone'],
                processed['comissoes']['eletronicos_a'] + processed['comissoes']['eletronicos_b'],
                processed['comissoes']['essenciais_a'] + processed['comissoes']['essenciais_b'],
                processed['comissoes']['seguro'],
                processed['comissoes']['sva'],
            ],
        },
        'bonus': {
            'labels': ['Móvel', 'Fixa', 'Smart', 'Eletro', 'Essen', 'Seguro', 'SVA'],
            'data': [
                processed['bonus']['movel'],
                processed['bonus']['fixa'],
                processed['bonus']['smartphone'],
                processed['bonus']['eletronicos'],
                processed['bonus']['essenciais'],
                processed['bonus']['seguro'],
                processed['bonus']['sva'],
            ],
        },
        'alto_desempenho': {
            'labels': ['Móvel', 'Fixa', 'Smart', 'Eletro', 'Essen', 'Seguro', 'SVA'],
            'data': [
                processed['alto_desempenho']['movel'],
                processed['alto_desempenho']['fixa'],
                processed['alto_desempenho']['smartphone'],
                processed['alto_desempenho']['eletronicos_a'] + processed['alto_desempenho']['eletronicos_b'],
                processed['alto_desempenho']['essenciais_a'] + processed['alto_desempenho']['essenciais_b'],
                processed['alto_desempenho']['seguro'],
                processed['alto_desempenho']['sva'],
            ],
        },
        'hunter': {
            'labels': ['Móvel', 'Fixa', 'Smart', 'Eletro', 'Essen', 'Seguro', 'SVA'],
            'data': [
                processed['hunter']['movel'],
                processed['hunter']['fixa'],
                processed['hunter']['smartphone'],
                processed['hunter']['eletronicos'],
                processed['hunter']['essenciais'],
                processed['hunter']['seguro'],
                processed['hunter']['sva'],
            ],
        },
        'composicao': {
            'labels': ['Comissão', 'Bônus', 'Alto Desempenho', 'Hunter', 'Aceleradores'],
            'data': [
                processed['comissoes']['total'],
                processed['bonus']['total'],
                processed['alto_desempenho']['total'],
                sum(processed['hunter'].values()),
                processed['aceleradores']['seis_pilares'],
            ],
        },
    }
    
    processed['raw_data'] = data
    
    return processed


def process_simple_commission_data(data, is_gerente=False):
    """
    Versão simplificada para listar múltiplos usuários
    """
    pilares_config = [
        {'nome': 'Móvel', 'key': 'MOVEL', 'com_key': 'COM_MÓVEL', 'field': 'movel_comissao'},
        {'nome': 'Fixa', 'key': 'FIXA', 'com_key': 'COM_FIXA', 'field': 'fixa_comissao'},
        {'nome': 'Smartphone', 'key': 'SMART', 'com_key': 'COM_SMARTPHONE', 'field': 'smartphone_comissao'},
        {'nome': 'Eletrônicos', 'key': 'ELETRO', 'com_key': 'COM_ELETRONICOS', 'field': 'eletronicos_comissao'},
        {'nome': 'Essenciais', 'key': 'ESSEN', 'com_key': 'COM_ESSENCIAIS', 'field': 'essenciais_comissao'},
        {'nome': 'Seguro', 'key': 'SEG', 'com_key': 'COM_SEGURO', 'field': 'seguro_comissao'},
        {'nome': 'SVA', 'key': 'SVA', 'com_key': 'COM_SVA', 'field': 'sva_comissao'},
    ]
    
    pilares = []
    result = {
        'nome': data.get('NOME', ''),
        'pdv': data.get('PDV', ''),
        'remuneracao_final': safe_float(data.get('REMUNERAÇÃO_FINAL_TOTAL') or data.get('REMUNERAÇÃO FINAL TOTAL')),
        'comissao_total': safe_float(data.get('Total Comissão')),
    }
    
    for pilar in pilares_config:
        key = pilar['key']
        
        # Buscar percentual de atingimento com fallbacks para variações
        if pilar['nome'] == 'Eletrônicos':
            pct_1_raw = safe_float(
                data.get('%ATING_ELETRO_1') or
                data.get('%ATING_ELETRONICOS_1') or
                data.get('%ATING_ELETRÔNICOS_1') or
                data.get('ATING_ELETRO_1') or
                data.get('%ATING ELETRO_1') or
                data.get('%ATING_ELETRO 1')
            )
        elif pilar['nome'] == 'Smartphone':
            pct_1_raw = safe_float(
                data.get('%ATING_SMART_1') or
                data.get('%ATING_SMARTPHONE_1') or
                data.get('ATING_SMART_1')
            )
        elif pilar['nome'] == 'Móvel':
            pct_1_raw = safe_float(
                data.get('%ATING_MOVEL_1') or
                data.get('%ATING_MÓVEL_1') or
                data.get('ATING_MOVEL_1')
            )
        elif pilar['nome'] == 'Essenciais':
            pct_1_raw = safe_float(
                data.get('%ATING_ESSEN_1') or
                data.get('%ATING_ESSENCIAIS_1') or
                data.get('ATING_ESSEN_1')
            )
        elif pilar['nome'] == 'Seguro':
            pct_1_raw = safe_float(
                data.get('%ATING_SEG_1') or
                data.get('%ATING_SEGURO_1') or
                data.get('ATING_SEG_1')
            )
        else:
            # Fallback padrão
            pct_1_raw = safe_float(data.get(f'%ATING_{key}_1'))
        
        # Buscar comissão do pilar com fallbacks para variações de nomes
        if pilar['nome'] == 'Eletrônicos':
            # Tentar múltiplas variações para Eletrônicos
            comissao = safe_float(
                data.get('COM_ELETRONICOS - A') or 
                data.get('COM_ELETRÔNICOS - A') or 
                data.get('COM_ELETRO - A') or 
                data.get('COM_ELETRONICOS-A') or 
                data.get('COM_ELETRONICOS A') or
                data.get('COM_ELETRONICOS - B') or 
                data.get('COM_ELETRÔNICOS - B') or 
                data.get('COM_ELETRO - B') or 
                data.get('COM_ELETRONICOS-B') or 
                data.get('COM_ELETRONICOS B') or
                data.get('COM_ELETRONICOS') or
                data.get('COM_ELETRÔNICOS') or
                data.get('COM_ELETRO')
            )
            # Se encontrou A ou B, somar ambos
            eletro_a = safe_float(
                data.get('COM_ELETRONICOS - A') or 
                data.get('COM_ELETRÔNICOS - A') or 
                data.get('COM_ELETRO - A') or 
                data.get('COM_ELETRONICOS-A') or 
                data.get('COM_ELETRONICOS A')
            )
            eletro_b = safe_float(
                data.get('COM_ELETRONICOS - B') or 
                data.get('COM_ELETRÔNICOS - B') or 
                data.get('COM_ELETRO - B') or 
                data.get('COM_ELETRONICOS-B') or 
                data.get('COM_ELETRONICOS B')
            )
            comissao = eletro_a + eletro_b if (eletro_a > 0 or eletro_b > 0) else comissao
        elif pilar['nome'] == 'Smartphone':
            comissao = safe_float(data.get('COM_SMARTPHONE') or data.get('COM_SMART'))
        elif pilar['nome'] == 'Móvel':
            comissao = safe_float(data.get('COM_MÓVEL') or data.get('COM_MOVEL'))
        elif pilar['nome'] == 'Essenciais':
            # Tentar múltiplas variações para Essenciais
            essen_a = safe_float(
                data.get('COM_ESSENCIAIS - A') or 
                data.get('COM_ESSEN - A') or 
                data.get('COM_ESSENCIAIS-A') or 
                data.get('COM_ESSENCIAIS A')
            )
            essen_b = safe_float(
                data.get('COM_ESSENCIAIS - B') or 
                data.get('COM_ESSEN - B') or 
                data.get('COM_ESSENCIAIS-B') or 
                data.get('COM_ESSENCIAIS B')
            )
            comissao = essen_a + essen_b
        elif pilar['nome'] == 'Seguro':
            comissao = safe_float(data.get('COM_SEGURO') or data.get('COM_SEG'))
        else:
            # Fallback padrão para outros pilares
            comissao = safe_float(data.get(pilar['com_key'])) or safe_float(data.get(f"COM_{pilar['nome'].upper()}"))
        
        pilares.append({
            'nome': pilar['nome'],
            'pct': convert_percentage(pct_1_raw),
            'comissao': comissao,
        })
        
        # Adicionar campo individual de comissão para o gráfico
        result[pilar['field']] = comissao
    
    result['pilares'] = pilares
    return result


def process_coordenador_commission_data(data, metas_pilar=None):
    """
    Processa os dados de comissionamento específicos para Coordenador.
    A planilha de COO/SNP tem colunas diferentes da de Gerente/CN.
    """
    processed = {
        'info': {
            'nome': data.get('NOME', ''),
            'cargo': 'Coordenador',
        },
        'pilares': [],
        'comissoes': {},
        'bonus': {},
        'alto_desempenho': {},
        'hunter': {},
        'aceleradores': {},
        'descontos': {},
        'totais': {},
        'charts': {}
    }
    
    # Inicializar metas_pilar se não fornecido
    if metas_pilar is None:
        metas_pilar = {}
    
    # Mapeamento de chave para metas_pilar
    pilar_to_meta_key = {
        'MOVEL': 'movel',
        'FIXA': 'fixa',
        'SMART': 'smartphone',
        'ELETRO': 'eletronicos',
        'ESSEN': 'essenciais',
        'SEG': 'seguro',
        'SVA': 'sva',
    }
    
    # Pilares com atingimentos
    pilares_config = [
        {'nome': 'Móvel', 'key': 'MOVEL', 'icon': '📱', 'color': '#660099'},
        {'nome': 'Fixa', 'key': 'FIXA', 'icon': '🏠', 'color': '#9B30FF'},
        {'nome': 'Smartphone', 'key': 'SMART', 'icon': '📲', 'color': '#BA55D3'},
        {'nome': 'Eletrônicos', 'key': 'ELETRO', 'icon': '💻', 'color': '#9370DB'},
        {'nome': 'Essenciais', 'key': 'ESSEN', 'icon': '⭐', 'color': '#8A2BE2'},
        {'nome': 'Seguro', 'key': 'SEG', 'icon': '🛡️', 'color': '#7B68EE'},
        {'nome': 'SVA', 'key': 'SVA', 'icon': '📦', 'color': '#6A5ACD'},
    ]
    
    for pilar in pilares_config:
        key = pilar['key']
        
        pct_3_raw = safe_float(data.get(f'%ATING_{key}_3'))
        pct_2_raw = safe_float(data.get(f'%ATING_{key}_2'))
        pct_1_raw = safe_float(data.get(f'%ATING_{key}_1'))
        
        # DEBUG: Mostrar valores encontrados
        print(f"[DEBUG Ating COORD] {key}: pct_1={pct_1_raw}, pct_2={pct_2_raw}, pct_3={pct_3_raw}")
        
        # Buscar metas (Total, Pago, Exclusão) deste pilar
        meta_key = pilar_to_meta_key.get(key, key.lower())
        pilar_metas = metas_pilar.get(meta_key, {'total': 0, 'pago': 0, 'exclusao': 0})
        
        pilar_data = {
            'nome': pilar['nome'],
            'icon': pilar['icon'],
            'color': pilar['color'],
            'pct_3': convert_percentage(pct_3_raw),
            'pct_2': convert_percentage(pct_2_raw),
            'pct_1': convert_percentage(pct_1_raw),
            'habilitado': data.get(f'H_{pilar["nome"].upper()}') or data.get(f'H_{key}'),
            # Dados de metas: Total, Pago, Exclusão
            'total': pilar_metas.get('total', 0),
            'pago': pilar_metas.get('pago', 0),
            'exclusao': pilar_metas.get('exclusao', 0),
        }
        
        ating_values = [pilar_data['pct_3'], pilar_data['pct_2'], pilar_data['pct_1']]
        valid_values = [v for v in ating_values if v > 0]
        pilar_data['media'] = sum(valid_values) / len(valid_values) if valid_values else 0
        
        processed['pilares'].append(pilar_data)
    
    # Comissões por pilar
    comissoes_valores = {
        'movel': safe_float(data.get('COM_MÓVEL')),
        'fixa': safe_float(data.get('COM_FIXA')),
        'smartphone': safe_float(data.get('COM_SMARTPHONE')),
        'eletronicos_a': safe_float(data.get('COM_ELETRONICOS - A')),
        'eletronicos_b': safe_float(data.get('COM_ELETRONICOS - B')),
        'essenciais_a': safe_float(data.get('COM_ESSENCIAIS - A')),
        'essenciais_b': safe_float(data.get('COM_ESSENCIAIS - B')),
        'seguro': safe_float(data.get('COM_SEGURO')),
        'sva': safe_float(data.get('COM_SVA')),
    }
    # Total de comissões = REMUNERÇÃO PILARES
    total_comissoes = safe_float(data.get('REMUNERÇÃO PILARES'))
    if total_comissoes == 0:
        total_comissoes = sum(comissoes_valores.values())
    comissoes_valores['total'] = total_comissoes
    processed['comissoes'] = comissoes_valores
    
    # Coordenador não tem Bônus Carteira separado
    processed['bonus'] = {
        'movel': 0, 'fixa': 0, 'smartphone': 0,
        'eletronicos_a': 0, 'eletronicos_b': 0, 'eletronicos': 0,
        'essenciais_a': 0, 'essenciais_b': 0, 'essenciais': 0,
        'seguro': 0, 'sva': 0, 'total': 0
    }
    
    # Coordenador não tem Alto Desempenho separado
    processed['alto_desempenho'] = {
        'movel': 0, 'fixa': 0, 'smartphone': 0,
        'eletronicos_a': 0, 'eletronicos_b': 0,
        'essenciais_a': 0, 'essenciais_b': 0,
        'seguro': 0, 'sva': 0, 'total': 0
    }
    
    # Hunter
    processed['hunter'] = {
        'movel': safe_float(data.get('HUNTER_MOVEL')),
        'fixa': safe_float(data.get('HUNTER_FIXA')),
        'smartphone': safe_float(data.get('HUNTER_SMARTPHONE')),
        'eletronicos': safe_float(data.get('HUNTER_ELETRONICOS')),
        'essenciais': safe_float(data.get('HUNTER_ESSENCIAIS')),
        'seguro': safe_float(data.get('HUNTER_SEGUROS')),
        'sva': safe_float(data.get('HUNTER_SVA')),
    }
    # Calcular total do Hunter
    processed['hunter']['total'] = sum([
        processed['hunter']['movel'],
        processed['hunter']['fixa'],
        processed['hunter']['smartphone'],
        processed['hunter']['eletronicos'],
        processed['hunter']['essenciais'],
        processed['hunter']['seguro'],
        processed['hunter']['sva'],
    ])
    
    # Bônus Hunter - Coordenador não tem
    processed['bonus_hunter'] = {
        'movel': 0, 'fixa': 0, 'smartphone': 0,
        'eletronicos': 0, 'essenciais': 0, 'seguro': 0, 'sva': 0, 'total': 0
    }
    
    # Aceleradores e IQ
    processed['aceleradores'] = {
        'seis_pilares': safe_float(data.get('6/7_PILARES')),
        'campanha': safe_float(data.get('CAMPANHA')),
        'total_aceleradores': safe_float(data.get('REMUNERÇÃO BRUTA')),
        'iq_acelerador_movel': safe_float(data.get('IQ_ACELADOR MÓVEL')),
        'iq_acelerador_fixa': safe_float(data.get('IQ_ACELERADOR FIXA')),
        'iq_deflator_movel': safe_float(data.get('IQ_DEFLATOR MÓVEL')),
        'iq_deflator_fixa': safe_float(data.get('IQ_DEFLATOR FIXA')),
        'remuneracao_iq': safe_float(data.get('REMUNERÇÃO COM IQ')),
    }
    
    # Descontos
    processed['descontos'] = {
        'advertencia': safe_float(data.get('DESCONTO_ADVERTÊNCIA')),
        'excedente': safe_float(data.get('DESC_PAGAMENTO EXCEDENTE DO MÊS ANTERIOR')),
        'price': safe_float(data.get('DESCONTO_PRICE')),
        'total': safe_float(data.get('DESCONTO_TOTAL')),
    }
    
    # Totais específicos de Coordenador
    processed['totais'] = {
        'comissao_cargo': safe_float(data.get('COMISSÃO_COO/SNP')),
        'premiacao_cargo': safe_float(data.get('PREMIAÇÃO_COO/SNP')),
        'remuneracao_pilares': safe_float(data.get('REMUNERÇÃO PILARES')),
        'remuneracao_bruta': safe_float(data.get('REMUNERÇÃO BRUTA')),
        'remuneracao_iq': safe_float(data.get('REMUNERÇÃO COM IQ')),
        'remuneracao_final': safe_float(data.get(' REMUNERAÇÃO FINAL COO/SNP')),  # Nota: tem espaço no início
        'cargo_label': 'Coordenador',
    }
    
    processed['habilitados'] = {
        'pilares_67': data.get('M|F|T|A'),
    }
    
    # Dados para gráficos
    processed['charts'] = {
        'atingimentos': {
            'labels': [p['nome'] for p in processed['pilares']],
            'data': [p['media'] for p in processed['pilares']],
            'colors': [p['color'] for p in processed['pilares']],
        },
        'comissoes': {
            'labels': ['Móvel', 'Fixa', 'Smart', 'Eletro', 'Essen', 'Seguro', 'SVA'],
            'data': [
                processed['comissoes']['movel'],
                processed['comissoes']['fixa'],
                processed['comissoes']['smartphone'],
                processed['comissoes']['eletronicos_a'] + processed['comissoes']['eletronicos_b'],
                processed['comissoes']['essenciais_a'] + processed['comissoes']['essenciais_b'],
                processed['comissoes']['seguro'],
                processed['comissoes']['sva'],
            ],
        },
        'composicao': {
            'labels': ['Comissão', 'Hunter', 'Campanha'],
            'data': [
                processed['comissoes']['total'],
                sum(processed['hunter'].values()),
                processed['aceleradores']['campanha'],
            ],
        },
    }
    
    processed['raw_data'] = data
    
    return processed


def fetch_lojas_por_coordenador(coordenador_nome):
    """
    Busca as lojas (PDVs) que pertencem a um coordenador específico.
    Lê a sheet REMUNERAÇÃO GERENTE e filtra pela coluna COORDENADOR (A).
    """
    cache_key = f"lojas_coordenador_{coordenador_nome.replace(' ', '_')}"
    cached_data = cache.get(cache_key)
    
    if cached_data:
        return cached_data
    
    # Obter URLs dinâmicas
    urls = get_excel_urls()
    
    try:
        excel_file, error = download_excel_file(urls['comissao'], "lojas_coord")
        if error:
            return []
        
        df = pd.read_excel(excel_file, sheet_name=SHEET_GERENTE, engine='openpyxl')
        
        coord_col = None
        for col in df.columns:
            if 'COORDENADOR' in str(col).upper():
                coord_col = col
                break
        
        if coord_col is None:
            return []
        
        def normalize_name(name):
            if pd.isna(name):
                return ""
            return str(name).strip().upper()
        
        coord_normalized = normalize_name(coordenador_nome)
        
        lojas = []
        for idx, row in df.iterrows():
            row_coord = normalize_name(row.get(coord_col, ''))
            # Comparar primeiro nome
            if row_coord and (row_coord == coord_normalized or 
                             coord_normalized in row_coord or 
                             row_coord in coord_normalized):
                pdv = row.get('PDV', '')
                gerente = row.get('NOME', '')
                if pd.notna(pdv) and str(pdv).strip():
                    print(str(gerente).strip())
                    print(pd.notna(gerente))
                    lojas.append({
                        'pdv': str(pdv).strip(),
                        'gerente': str(gerente).strip() if pd.notna(gerente) else '',
                    })
        
        cache.set(cache_key, lojas, 300)
        return lojas
        
    except Exception as e:
        return []


# ============================================================================
# VIEWS
# ============================================================================

@login_required
def commission_view(request):
    """
    View principal para exibir dados de comissionamento
    Redireciona para a visão apropriada baseado no tipo de usuário
    """
    user = request.user
    
    # Bloqueio de acesso: apenas setores 24 e 29
    SETORES_PERMITIDOS = [24, 29]
    user_sector_id = user.sector.id if hasattr(user, 'sector') and user.sector else None
    
    if user_sector_id not in SETORES_PERMITIDOS:
        from django.http import HttpResponseForbidden
        return HttpResponseForbidden("Acesso negado. Você não tem permissão para acessar esta página.")
    
    role = get_user_role(user)
    
    # Coordenador tem visão especial
    if role == 'coordenador':
        return commission_coordenador_view(request)
    
    # Gerente pode ver equipe ou seu próprio
    if role == 'gerente':
        return commission_gerente_view(request)
    
    # CN padrão
    return commission_cn_view(request)


@login_required
def commission_cn_view(request):
    """
    Visão do CN:
    - Vê seu comissionamento
    - Valores por pilar
    - Lista de vendas
    """
    user = request.user
    
    # Determina qual sheet usar
    is_gerente = is_user_gerente(user)
    sheet_name = SHEET_GERENTE if is_gerente else SHEET_CN
    
    # Nome para busca na planilha
    user_full_name = user.get_full_name() or user.first_name
    
    # Setor do usuário
    user_sector = user.sector.name if hasattr(user, 'sector') and user.sector else None
    
    # Buscar dados de comissionamento
    result = fetch_excel_data(sheet_name, user_full_name)
    
    # Buscar dados de vendas
    vendas_result = fetch_vendas_data(user_full_name)
    
    # Obter nomes dos meses
    meses = get_month_names()
    
    context = {
        'user': user,
        'target_user': user,
        'viewing_other': False,
        'is_gerente': is_gerente,
        'target_is_gerente': is_gerente,
        'sector_users': [],
        'result': result,
        'meses': meses,
        'vendas': vendas_result.get('vendas', []) if vendas_result.get('success') else [],
        'role': 'cn',
    }
    
    if result.get('success'):
        # Obter metas por pilar - CN busca por nome, Gerente busca por filial
        metas_pilar = fetch_metas_por_pilar(user_full_name, is_gerente, user_sector)
        # Buscar dados de IQ
        iq_data = fetch_iq_data(user_full_name)
        print(f"[commission_dashboard] IQ Data: {iq_data}")
        print(f"[commission_dashboard] Metas Pilar: {metas_pilar}")
        processed = process_commission_data(result['data'], is_gerente, metas_pilar, iq_data)
        print(f"[commission_dashboard] Remuneração Final: {processed['totais'].get('remuneracao_final')}")
        print(f"[commission_dashboard] Pilares[0] pct_1: {processed['pilares'][0].get('pct_1') if processed['pilares'] else 'N/A'}")
        context['data'] = processed
        context['charts_json'] = json.dumps(processed['charts'])
    
    return render(request, 'users/commission.html', context)


@login_required
def commission_gerente_view(request):
    """
    Visão do Gerente:
    - Consegue ver todos os CNs da loja
    - Consegue ver valores que cada CN está gerando
    - Atingimentos por pilar
    """
    user = request.user
    
    # Verificar se está visualizando um CN específico
    viewing_user_id = request.GET.get('user')
    viewing_user = None
    
    # Buscar CNs do setor
    sector_cns = get_all_cns_for_gerente(user)
    
    if viewing_user_id:
        try:
            viewing_user = User.objects.get(id=viewing_user_id, is_active=True)
            # Verificar se o CN é do mesmo setor
            if viewing_user.sector != user.sector:
                messages.error(request, 'Este usuário não pertence à sua loja.')
                return redirect('commission')
        except User.DoesNotExist:
            messages.error(request, 'Usuário não encontrado.')
            return redirect('commission')
    
    # Define qual usuário buscar dados
    if viewing_user:
        target_user = viewing_user
        target_is_gerente = is_user_gerente(viewing_user)
    else:
        target_user = user
        target_is_gerente = True
    
    sheet_name = SHEET_GERENTE if target_is_gerente else SHEET_CN
    user_full_name = target_user.get_full_name() or target_user.first_name
    
    # Buscar dados de comissionamento
    result = fetch_excel_data(sheet_name, user_full_name)
    
    # Buscar resumo de todos os vendedores da loja diretamente da BASE_PAGAMENTO
    target_sector = target_user.sector.name if hasattr(target_user, 'sector') and target_user.sector else None
    cns_resumo = fetch_vendedores_por_filial(target_sector)
    
    meses = get_month_names()
    
    context = {
        'user': user,
        'target_user': target_user,
        'viewing_other': viewing_user is not None,
        'is_gerente': True,
        'target_is_gerente': target_is_gerente,
        'sector_users': sector_cns,
        'cns_resumo': cns_resumo,
        'result': result,
        'meses': meses,
        'role': 'gerente',
    }
    
    if result.get('success'):
        # Buscar metas por pilar - Gerente busca por filial, CN busca por nome
        target_name = target_user.get_full_name() or target_user.first_name
        target_sector = target_user.sector.name if hasattr(target_user, 'sector') and target_user.sector else None
        print(f"[commission_gerente_view] target_name={target_name}, target_is_gerente={target_is_gerente}, target_sector={target_sector}")
        metas_pilar = fetch_metas_por_pilar(target_name, target_is_gerente, target_sector)
        print(f"[commission_gerente_view] metas_pilar retornado: {metas_pilar}")
        # Buscar dados de IQ
        iq_data = fetch_iq_data(target_name)
        processed = process_commission_data(result['data'], target_is_gerente, metas_pilar, iq_data)
        context['data'] = processed
        context['charts_json'] = json.dumps(processed['charts'])
    
    return render(request, 'users/commission_gerente.html', context)


@login_required
def commission_coordenador_view(request):
    """
    Visão do Coordenador:
    - Consegue ver todos os gerentes e CNs
    - Consegue ver atingimento das lojas
    """
    user = request.user
    
    # Verificar se está visualizando um usuário específico
    viewing_user_id = request.GET.get('user')
    viewing_user = None
    
    if viewing_user_id:
        try:
            viewing_user = User.objects.get(id=viewing_user_id, is_active=True)
        except User.DoesNotExist:
            messages.error(request, 'Usuário não encontrado.')
            return redirect('commission')
    
    # Primeiro nome do coordenador para filtrar
    user_first_name = user.first_name.strip().upper() if user.first_name else ''
    
    # Buscar lojas deste coordenador específico da planilha
    lojas_coordenador = fetch_lojas_por_coordenador(user_first_name)
    pdvs_coordenador = [l['pdv'].upper() for l in lojas_coordenador]
    
    # Buscar dados de todos os gerentes e filtrar pelos que pertencem a este coordenador
    gerentes_data = []
    all_gerentes_result = fetch_all_users_from_sheet(SHEET_GERENTE)
    
    if all_gerentes_result.get('success'):
        for gerente_data in all_gerentes_result['users']:
            pdv = str(gerente_data['data'].get('PDV', '')).strip().upper()
            coord = str(gerente_data['data'].get('COORDENADOR (A)', '')).strip().upper()
            
            # Filtrar apenas gerentes deste coordenador
            if user_first_name in coord or pdv in pdvs_coordenador:
                processed = process_simple_commission_data(gerente_data['data'], is_gerente=True)
                processed['pdv'] = gerente_data['data'].get('PDV', '')
                processed['coordenador'] = gerente_data['data'].get('COORDENADOR (A)', '')
                gerentes_data.append(processed)
    
    # Criar lojas diretamente a partir de gerentes_data (mesma fonte que Desempenho dos Gerentes)
    lojas = []
    for gerente in gerentes_data:
        pdv = gerente.get('pdv', '')
        nome = gerente.get('nome', '')
        # Buscar o gerente no Django pelo nome
        gerente_user = None
        cns = []
        sector = None
        
        if nome:
            # Tentar encontrar o usuário pelo nome
            nome_parts = nome.split()
            if len(nome_parts) >= 2:
                gerente_user = User.objects.filter(
                    first_name__iexact=nome_parts[0],
                    last_name__icontains=nome_parts[-1],
                    is_active=True
                ).first()
            if not gerente_user:
                gerente_user = User.objects.filter(
                    first_name__iexact=nome_parts[0] if nome_parts else nome,
                    is_active=True
                ).first()
            
            if gerente_user and gerente_user.sector:
                sector = gerente_user.sector
                cns = User.objects.filter(
                    sector=sector,
                    is_active=True
                ).exclude(
                    groups__name__in=['Gerente', 'Gerentes', 'Coordenador', 'Coordenadores']
                )
        
        lojas.append({
            'pdv': pdv,
            'nome_gerente': nome,
            'setor': sector,
            'gerente': gerente_user,
            'cns': cns,
            'gerente_data': gerente,  # Dados da planilha
        })
    
    # Se visualizando usuário específico
    if viewing_user:
        target_is_gerente = is_user_gerente(viewing_user)
        sheet_name = SHEET_GERENTE if target_is_gerente else SHEET_CN
        user_full_name = viewing_user.get_full_name() or viewing_user.first_name
        result = fetch_excel_data(sheet_name, user_full_name)
    else:
        # Coordenador visualizando seus próprios dados
        # Na planilha de coordenador, busca pelo PRIMEIRO NOME apenas
        sheet_name = SHEET_COORDENADOR
        result = fetch_excel_data(sheet_name, user_first_name)
    
    meses = get_month_names()
    
    # Calcular resumo de comissões por loja para o gráfico de representatividade
    lojas_resumo = []
    for loja in lojas:
        gerente_data = loja.get('gerente_data', {})
        if gerente_data:
            pilares = gerente_data.get('pilares', [])
            # Cada pilar tem 'nome', 'pct', 'comissao'
            lojas_resumo.append({
                'pdv': loja['pdv'],
                'movel': pilares[0].get('comissao', 0) if len(pilares) > 0 else 0,
                'fixa': pilares[1].get('comissao', 0) if len(pilares) > 1 else 0,
                'smartphone': pilares[2].get('comissao', 0) if len(pilares) > 2 else 0,
                'eletronicos': pilares[3].get('comissao', 0) if len(pilares) > 3 else 0,
                'essenciais': pilares[4].get('comissao', 0) if len(pilares) > 4 else 0,
                'seguro': pilares[5].get('comissao', 0) if len(pilares) > 5 else 0,
                'sva': pilares[6].get('comissao', 0) if len(pilares) > 6 else 0,
                'total': gerente_data.get('remuneracao_final', 0),
            })
    
    context = {
        'user': user,
        'target_user': viewing_user if viewing_user else user,
        'viewing_other': viewing_user is not None,
        'lojas': lojas,
        'lojas_coordenador': lojas_coordenador,
        'lojas_resumo': lojas_resumo,
        'lojas_resumo_json': json.dumps(lojas_resumo),
        'gerentes_data': gerentes_data,
        'result': result,
        'meses': meses,
        'role': 'coordenador',
    }
    
    if viewing_user and result.get('success'):
        target_is_gerente = is_user_gerente(viewing_user)
        # Buscar metas por pilar - Gerente busca por filial, CN busca por nome
        target_name = viewing_user.get_full_name() or viewing_user.first_name
        target_sector = viewing_user.sector.name if hasattr(viewing_user, 'sector') and viewing_user.sector else None
        metas_pilar = fetch_metas_por_pilar(target_name, target_is_gerente, target_sector)
        # Buscar dados de IQ
        iq_data = fetch_iq_data(target_name)
        processed = process_commission_data(result['data'], target_is_gerente, metas_pilar, iq_data)
        context['data'] = processed
        context['charts_json'] = json.dumps(processed['charts'])
        context['target_is_gerente'] = target_is_gerente
    elif not viewing_user and result.get('success'):
        # Coordenador visualizando seus próprios dados - usar função específica
        # Buscar metas por pilar para o coordenador
        coord_name = user.get_full_name() or user.first_name
        coord_sector = user.sector.name if hasattr(user, 'sector') and user.sector else None
        metas_pilar = fetch_metas_por_pilar(coord_name, is_gerente=False, user_sector=coord_sector)
        processed = process_coordenador_commission_data(result['data'], metas_pilar)
        context['data'] = processed
        context['charts_json'] = json.dumps(processed['charts'])
        context['target_is_gerente'] = False
        context['is_coordenador'] = True
    
    return render(request, 'users/commission_coordenador.html', context)


@login_required
def commission_api(request):
    """
    API para buscar dados de comissionamento via AJAX
    """
    user = request.user
    user_id = request.GET.get('user_id')
    role = get_user_role(user)
    
    # Verificar permissões
    if user_id:
        if role == 'cn':
            return JsonResponse({'error': 'Não autorizado'}, status=403)
        
        try:
            target_user = User.objects.get(id=user_id, is_active=True)
            
            # Gerente só pode ver CNs do seu setor
            if role == 'gerente' and target_user.sector != user.sector:
                return JsonResponse({'error': 'Não autorizado'}, status=403)
                
        except User.DoesNotExist:
            return JsonResponse({'error': 'Usuário não encontrado'}, status=404)
    else:
        target_user = user
    
    target_is_gerente = is_user_gerente(target_user)
    sheet_name = SHEET_GERENTE if target_is_gerente else SHEET_CN
    user_full_name = target_user.get_full_name() or target_user.first_name
    target_sector = target_user.sector.name if hasattr(target_user, 'sector') and target_user.sector else None
    
    result = fetch_excel_data(sheet_name, user_full_name)
    
    if result.get('success'):
        # Buscar metas por pilar - Gerente busca por filial, CN busca por nome
        metas_pilar = fetch_metas_por_pilar(user_full_name, target_is_gerente, target_sector)
        # Buscar dados de IQ
        iq_data = fetch_iq_data(user_full_name)
        processed = process_commission_data(result['data'], target_is_gerente, metas_pilar, iq_data)
        return JsonResponse({
            'success': True,
            'data': processed,
            'user_name': user_full_name
        })
    
    return JsonResponse(result)


@login_required
def commission_refresh(request):
    """
    Força atualização dos dados de comissionamento (limpa cache)
    """
    # Limpa todos os caches de comissionamento
    from django.core.cache import cache
    
    # Como não podemos listar as chaves no cache padrão, 
    # limpamos as chaves conhecidas
    user = request.user
    user_full_name = (user.get_full_name() or user.first_name).replace(' ', '_')
    
    for sheet in [SHEET_GERENTE, SHEET_CN]:
        cache.delete(f"commission_data_{sheet}_{user_full_name}")
        cache.delete(f"commission_all_users_{sheet}")
        cache.delete(f"comissao_{sheet}_file_content")
    
    cache.delete("vendas_all_data")
    cache.delete(f"vendas_data_{user_full_name}")
    cache.delete("vendas_file_content")
    cache.delete("vendas_all_file_content")
    
    # Limpa também caches de pilares e IQ
    cache.delete("metas_por_pilar")
    cache.delete("iq_data")
    cache.delete("vendedores_por_filial")
    cache.delete("lojas_por_coordenador")
    
    messages.success(request, 'Cache de dados de comissionamento limpo com sucesso!')
    
    # Redireciona de volta para a página de origem, se especificada
    referer = request.META.get('HTTP_REFERER', '')
    if 'system-config' in referer:
        return redirect('system_config')
    return redirect('commission')


@login_required
def export_commission_excel(request):
    """
    Exporta dados de comissionamento para Excel
    """
    user = request.user
    role = get_user_role(user)
    
    wb = Workbook()
    ws = wb.active
    ws.title = 'Comissionamento'
    
    # Estilos
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='660099', end_color='660099', fill_type='solid')
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    if role == 'coordenador':
        # Exportar dados de todos os gerentes
        headers = ['Nome', 'PDV', 'Móvel %', 'Fixa %', 'Smart %', 'Eletro %', 'Essen %', 'Seguro %', 'SVA %', 'Remuneração Final']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
        
        all_result = fetch_all_users_from_sheet(SHEET_GERENTE)
        if all_result.get('success'):
            for row, user_data in enumerate(all_result['users'], start=2):
                processed = process_simple_commission_data(user_data['data'], is_gerente=True)
                ws.cell(row=row, column=1, value=processed['nome']).border = border
                ws.cell(row=row, column=2, value=processed['pdv']).border = border
                for col, pilar in enumerate(processed['pilares'], start=3):
                    ws.cell(row=row, column=col, value=f"{pilar['pct']:.1f}%").border = border
                ws.cell(row=row, column=10, value=f"R$ {processed['remuneracao_final']:.2f}").border = border
    
    elif role == 'gerente':
        # Exportar dados dos CNs do setor
        headers = ['Nome', 'Móvel %', 'Fixa %', 'Smart %', 'Eletro %', 'Essen %', 'Seguro %', 'SVA %', 'Comissão Total', 'Remuneração Final']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
        
        sector_cns = get_all_cns_for_gerente(user)
        all_result = fetch_all_users_from_sheet(SHEET_CN)
        
        row = 2
        if all_result.get('success'):
            for cn in sector_cns:
                cn_full_name = (cn.get_full_name() or cn.first_name).upper()
                for user_data in all_result['users']:
                    if cn_full_name in user_data['nome'].upper() or user_data['nome'].upper() in cn_full_name:
                        processed = process_simple_commission_data(user_data['data'])
                        ws.cell(row=row, column=1, value=processed['nome']).border = border
                        for col, pilar in enumerate(processed['pilares'], start=2):
                            ws.cell(row=row, column=col, value=f"{pilar['pct']:.1f}%").border = border
                        ws.cell(row=row, column=9, value=f"R$ {processed['comissao_total']:.2f}").border = border
                        ws.cell(row=row, column=10, value=f"R$ {processed['remuneracao_final']:.2f}").border = border
                        row += 1
                        break
    
    else:
        # CN - Exportar dados próprios
        user_full_name = user.get_full_name() or user.first_name
        user_sector = user.sector.name if hasattr(user, 'sector') and user.sector else None
        result = fetch_excel_data(SHEET_CN, user_full_name)
        
        if result.get('success'):
            # Buscar metas por pilar - CN busca por nome
            metas_pilar = fetch_metas_por_pilar(user_full_name, False, user_sector)
            # Buscar dados de IQ
            iq_data = fetch_iq_data(user_full_name)
            processed = process_commission_data(result['data'], False, metas_pilar, iq_data)
            
            ws['A1'] = 'Comissionamento - ' + user_full_name
            ws['A1'].font = Font(bold=True, size=14)
            
            ws['A3'] = 'Pilar'
            ws['B3'] = 'Atingimento'
            ws['C3'] = 'Comissão'
            for col in ['A', 'B', 'C']:
                ws[f'{col}3'].font = header_font
                ws[f'{col}3'].fill = header_fill
            
            row = 4
            for pilar in processed['pilares']:
                ws.cell(row=row, column=1, value=pilar['nome'])
                ws.cell(row=row, column=2, value=f"{pilar['media']:.1f}%")
                row += 1
    
    # Ajustar larguras
    for col in range(1, 12):
        ws.column_dimensions[get_column_letter(col)].width = 15
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    from datetime import datetime
    filename = f'comissionamento_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response


@login_required
def api_vendas_por_pilar(request):
    """
    API para buscar vendas de um pilar específico via AJAX.
    Retorna lista de vendas filtradas por pilar (Pago e Exclusão).
    
    Usa EXCEL_BASE_PAGAMENTO_URL, sheets: Planilha1 (Pago) e EXCLUSAO
    
    Filtros por papel:
    - CN: Filtra por coluna VENDEDOR = nome do usuário
    - Gerente: Filtra por coluna PDV = nome da loja do usuário
    - Coordenador: Filtra por coluna COORDENACAO = primeiro nome do usuário
    """
    user = request.user
    pilar = request.GET.get('pilar', '').upper()
    
    if not pilar:
        return JsonResponse({'error': 'Pilar não especificado'}, status=400)
    
    # Identificar papel do usuário
    role = get_user_role(user)
    is_coordenador = role == 'coordenador'
    is_gerente = role == 'gerente'
    
    # Informações do usuário para busca
    user_full_name = user.get_full_name() or user.first_name or ""
    user_first_name = user.first_name.strip().upper() if user.first_name else ""
    user_last_name = user.last_name.strip().upper() if user.last_name else ""
    user_pdv = user.sector.name if hasattr(user, 'sector') and user.sector else ""
    
    print(f"[api_vendas_por_pilar] Usuário: {user_full_name}, Pilar: {pilar}, Role: {role}")
    print(f"[api_vendas_por_pilar] first_name: {user_first_name}, last_name: {user_last_name}, PDV: {user_pdv}")
    
    # Mapeamento de pilar - inclui variações com e sem acento
    pilar_mapping = {
        'MOVEL': ['MÓVEL', 'MOVEL'],
        'MÓVEL': ['MÓVEL', 'MOVEL'],
        'FIXA': ['FIXA'],
        'SMARTPHONE': ['SMART', 'SMARTPHONE'],
        'ELETRONICOS': ['ELETRO', 'ELETRÔNICO', 'ELETRONICOS', 'ELETRÔNICOS'],
        'ESSENCIAIS': ['ESSEN', 'ESSENCIAL', 'ESSENCIAIS'],
        'SEGURO': ['SEG', 'SEGURO', 'SEGUROS'],
        'SVA': ['SVA'],
    }
    
    # Normalizar o pilar recebido
    pilar_filtros = pilar_mapping.get(pilar, [pilar])
    
    def normalize_text(text):
        if pd.isna(text):
            return ""
        return str(text).strip().upper()
    
    def remove_accents(text):
        """Remove acentos de uma string para comparação"""
        import unicodedata
        if not text:
            return ""
        # Normaliza para forma decomposta e remove marcas diacríticas
        nfkd = unicodedata.normalize('NFKD', text)
        return ''.join(c for c in nfkd if not unicodedata.combining(c))
    
    def pilar_match_check(row_pilar, filtros):
        """Verifica se o pilar da linha corresponde aos filtros"""
        row_clean = remove_accents(row_pilar)
        for filtro in filtros:
            filtro_clean = remove_accents(filtro)
            # Comparar versões sem acento
            if filtro_clean in row_clean or row_clean == filtro_clean:
                return True
            # Comparar versões originais (com acento)
            if filtro in row_pilar or row_pilar == filtro:
                return True
        return False
    
    user_name_normalized = normalize_text(user_full_name)
    
    vendas_pago = []
    vendas_exclusao = []
    colunas = []
    
    # Função auxiliar para verificar match de nome (CN - coluna VENDEDOR)
    def nome_match(row_vendedor):
        """Verifica se o vendedor da linha corresponde ao usuário CN"""
        if not row_vendedor:
            return False
        vendedor_upper = normalize_text(row_vendedor)
        
        # Match exato
        if vendedor_upper == user_name_normalized:
            return True
        
        # Nome completo contém ou está contido
        if user_name_normalized in vendedor_upper or vendedor_upper in user_name_normalized:
            return True
        
        # Primeiro nome
        if user_first_name and user_first_name in vendedor_upper:
            return True
        
        # Primeiro + último nome
        if user_first_name and user_last_name:
            if user_first_name in vendedor_upper and user_last_name in vendedor_upper:
                return True
        
        return False
    
    # Função auxiliar para verificar match de coordenador (coluna COORDENACAO)
    def coordenador_match(row_coordenacao):
        """Verifica se a coordenação da linha corresponde ao coordenador"""
        if not row_coordenacao:
            return False
        coord_upper = normalize_text(row_coordenacao)
        
        # Coordenador: filtra pelo primeiro nome
        if user_first_name and user_first_name in coord_upper:
            return True
        
        # Ou nome completo
        if user_name_normalized in coord_upper or coord_upper in user_name_normalized:
            return True
        
        return False
    
    # Função auxiliar para verificar match de PDV (Gerente - coluna PDV)
    def pdv_match(row_pdv):
        """Verifica se o PDV da linha corresponde ao PDV do gerente"""
        if not row_pdv or not user_pdv:
            return False
        pdv_upper = normalize_text(row_pdv)
        pdv_user_upper = normalize_text(user_pdv)
        
        # Match exato
        if pdv_upper == pdv_user_upper:
            return True
        
        # Contém
        if pdv_user_upper in pdv_upper or pdv_upper in pdv_user_upper:
            return True
        
        return False
    
    # Obter URLs dinâmicas
    urls = get_excel_urls()
    
    try:
        # Baixar planilha BASE_PAGAMENTO
        excel_file, error = download_excel_file(urls['base_pagamento'], "base_pagamento_vendas")
        if error:
            print(f"[api_vendas_por_pilar] Erro download: {error}")
            return JsonResponse({'error': f'Erro ao baixar planilha: {error}'}, status=500)
        
        # ================== SHEET PAGO (Planilha1) ==================
        try:
            excel_file.seek(0)
            df_pago = pd.read_excel(excel_file, sheet_name='Planilha1', engine='openpyxl')
            colunas = list(df_pago.columns)
            print(f"[api_vendas_por_pilar] Colunas Planilha1: {colunas}")
            print(f"[api_vendas_por_pilar] Total linhas Planilha1: {len(df_pago)}")
        except Exception as e:
            return JsonResponse({'error': f'Erro ao ler planilha: {str(e)}'}, status=500)
        
        # Encontrar colunas necessárias
        pilar_col = None
        vendedor_col = None
        coordenacao_col = None
        pdv_col = None
        
        for col in df_pago.columns:
            col_upper = str(col).strip().upper()
            if col_upper == 'PILAR':
                pilar_col = col
            elif col_upper == 'VENDEDOR':
                vendedor_col = col
            elif col_upper in ['COORDENACAO', 'COORDENAÇÃO', 'COORDENADOR']:
                coordenacao_col = col
            elif col_upper == 'FILIAL':
                pdv_col = col
        
        print(f"[api_vendas_por_pilar] pilar_col={pilar_col}, vendedor_col={vendedor_col}, coordenacao_col={coordenacao_col}, pdv_col={pdv_col}")
        
        if pilar_col is None:
            return JsonResponse({'error': 'Coluna PILAR não encontrada'}, status=500)
        
        # Verificar coluna necessária baseada no papel
        if is_coordenador:
            if coordenacao_col is None:
                return JsonResponse({'error': 'Coluna COORDENACAO não encontrada'}, status=500)
            print(f"[api_vendas_por_pilar] Coordenador filtrando por COORDENACAO")
        elif is_gerente:
            if pdv_col is None:
                return JsonResponse({'error': 'Coluna PDV não encontrada'}, status=500)
            print(f"[api_vendas_por_pilar] Gerente filtrando por PDV: {user_pdv}")
        else:
            if vendedor_col is None:
                return JsonResponse({'error': 'Coluna VENDEDOR não encontrada'}, status=500)
            # Debug: mostrar alguns vendedores únicos
            vendedores_unicos = df_pago[vendedor_col].dropna().unique()[:10]
            print(f"[api_vendas_por_pilar] CN filtrando por VENDEDOR. Amostra: {list(vendedores_unicos)}")
        
        # Filtrar vendas PAGO
        for idx, row in df_pago.iterrows():
            # Verificar pilar usando função com suporte a acentos
            row_pilar = normalize_text(row.get(pilar_col, ''))
            if not pilar_match_check(row_pilar, pilar_filtros):
                continue
            
            # Filtrar baseado no papel do usuário
            match = False
            if is_coordenador:
                # Coordenador: filtra por COORDENACAO
                row_coord = row.get(coordenacao_col, '')
                match = coordenador_match(row_coord)
            elif is_gerente:
                # Gerente: filtra por PDV
                row_pdv = row.get(pdv_col, '')
                match = pdv_match(row_pdv)
            else:
                # CN: filtra por VENDEDOR
                row_vendedor = row.get(vendedor_col, '')
                match = nome_match(row_vendedor)
            
            if match:
                row_data = {}
                for col in df_pago.columns:
                    value = row.get(col)
                    if pd.isna(value):
                        row_data[str(col)] = None
                    elif isinstance(value, (int, float)):
                        row_data[str(col)] = float(value) if isinstance(value, float) else int(value)
                    else:
                        row_data[str(col)] = str(value)
                vendas_pago.append(row_data)
        
        print(f"[api_vendas_por_pilar] Vendas PAGO encontradas: {len(vendas_pago)}")
        
        # ================== SHEET EXCLUSÃO ==================
        try:
            excel_file_exclusao, error_exclusao = download_excel_file(urls['base_exclusao'], "base_exclusao")
            if error_exclusao:
                print(f"[api_vendas_por_pilar] Erro ao baixar EXCLUSAO: {error_exclusao}")
                df_exclusao = None
            else:
                excel_file_exclusao.seek(0)
                df_exclusao = pd.read_excel(excel_file_exclusao, sheet_name='Planilha1', engine='openpyxl')
                print(f"[api_vendas_por_pilar] Total linhas EXCLUSAO: {len(df_exclusao)}")
            
            if df_exclusao is not None:
                pilar_col_exc = None
                vendedor_col_exc = None
                coordenacao_col_exc = None
                pdv_col_exc = None
                
                for col in df_exclusao.columns:
                    col_upper = str(col).strip().upper()
                    if col_upper == 'PILAR':
                        pilar_col_exc = col
                    elif col_upper == 'VENDEDOR':
                        vendedor_col_exc = col
                    elif col_upper in ['COORDENACAO', 'COORDENAÇÃO', 'COORDENADOR']:
                        coordenacao_col_exc = col
                    elif col_upper == 'FILIAL':
                        pdv_col_exc = col
                
                print(f"[api_vendas_por_pilar] EXCLUSAO pilar_col={pilar_col_exc}, vendedor_col={vendedor_col_exc}, coordenacao_col={coordenacao_col_exc}, pdv_col={pdv_col_exc}")
                
                # Verificar se temos as colunas necessárias
                can_process = pilar_col_exc is not None
                if is_coordenador:
                    can_process = can_process and coordenacao_col_exc is not None
                elif is_gerente:
                    can_process = can_process and pdv_col_exc is not None
                else:
                    can_process = can_process and vendedor_col_exc is not None
                
                if can_process:
                    for idx, row in df_exclusao.iterrows():
                        row_pilar = normalize_text(row.get(pilar_col_exc, ''))
                        # Usar função com suporte a acentos
                        if not pilar_match_check(row_pilar, pilar_filtros):
                            continue
                        
                        # Filtrar baseado no papel do usuário
                        match = False
                        if is_coordenador:
                            # Coordenador: filtra por COORDENACAO
                            row_coord = row.get(coordenacao_col_exc, '')
                            match = coordenador_match(row_coord)
                        elif is_gerente:
                            # Gerente: filtra por PDV
                            row_pdv = row.get(pdv_col_exc, '')
                            match = pdv_match(row_pdv)
                        else:
                            # CN: filtra por VENDEDOR
                            row_vendedor = row.get(vendedor_col_exc, '')
                            match = nome_match(row_vendedor)
                        
                        if match:
                            row_data = {}
                            for col in df_exclusao.columns:
                                value = row.get(col)
                                if pd.isna(value):
                                    row_data[str(col)] = None
                                elif isinstance(value, (int, float)):
                                    row_data[str(col)] = float(value) if isinstance(value, float) else int(value)
                                else:
                                    row_data[str(col)] = str(value)
                            row_data['_tipo'] = 'exclusao'
                            vendas_exclusao.append(row_data)
                
                print(f"[api_vendas_por_pilar] Vendas EXCLUSAO encontradas: {len(vendas_exclusao)}")
            else:
                print(f"[api_vendas_por_pilar] df_exclusao é None, nenhuma venda de exclusão processada")
        except Exception as e:
            print(f"[api_vendas_por_pilar] Erro ao ler EXCLUSAO: {e}")
        
        # Calcular totais de receita
        receita_col_name = None
        for col in colunas:
            if 'RECEITA' in str(col).upper():
                receita_col_name = col
                break
        
        valor_total_pago = 0
        valor_total_exclusao = 0
        
        if receita_col_name:
            for venda in vendas_pago:
                val = venda.get(receita_col_name) or venda.get('Receita') or venda.get('RECEITA') or 0
                try:
                    valor_total_pago += float(val) if val else 0
                except:
                    pass
            
            for venda in vendas_exclusao:
                val = venda.get(receita_col_name) or venda.get('Receita') or venda.get('RECEITA') or 0
                try:
                    valor_total_exclusao += float(val) if val else 0
                except:
                    pass
        
        return JsonResponse({
            'success': True,
            'pilar': pilar,
            'vendas_pago': vendas_pago,
            'vendas_exclusao': vendas_exclusao,
            'total_pago': len(vendas_pago),
            'total_exclusao': len(vendas_exclusao),
            'valor_total_pago': valor_total_pago,
            'valor_total_exclusao': valor_total_exclusao,
            'colunas': colunas,
            'debug': {
                'user_name': user_full_name,
                'user_first_name': user_first_name,
                'user_last_name': user_last_name,
                'role': role,
                'is_coordenador': is_coordenador,
            }
        })
        
    except Exception as e:
        return JsonResponse({'error': f'Erro ao processar: {str(e)}'}, status=500)


@login_required
def export_vendas_pilar_excel(request):
    """
    Exporta vendas de um pilar específico para Excel.
    """
    import json
    from datetime import datetime
    
    pilar = request.GET.get('pilar', 'Pilar')
    vendas_json = request.GET.get('vendas', '[]')
    
    try:
        vendas = json.loads(vendas_json)
    except:
        vendas = []
    
    wb = Workbook()
    ws = wb.active
    ws.title = f'Vendas {pilar}'
    
    # Estilos
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='660099', end_color='660099', fill_type='solid')
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    if vendas:
        # Obter colunas da primeira venda
        colunas = [col for col in vendas[0].keys() if not col.startswith('_')]
        
        # Header
        for col_idx, col_name in enumerate(colunas, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
        
        # Dados
        for row_idx, venda in enumerate(vendas, 2):
            for col_idx, col_name in enumerate(colunas, 1):
                value = venda.get(col_name, '')
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = border
        
        # Ajustar larguras
        for col_idx in range(1, len(colunas) + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 18
    else:
        ws['A1'] = 'Sem vendas para exibir'
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f'vendas_{pilar}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response
