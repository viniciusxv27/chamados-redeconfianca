from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_POST, require_http_methods
from django.utils import timezone
from django.db.models import Sum, Max
from django.db import transaction
from decimal import Decimal
from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from .models import User, Sector, UserSession, normalize_cpf, RequiredDocument, UserDocument
from .serializers import UserSerializer, SectorSerializer
from core.middleware import log_action
import json


def _parse_bool_excel(value, default=False):
    if value is None or value == '':
        return default
    return str(value).strip().lower() in ['sim', 'true', '1', 'yes', 'y']


def _parse_date_excel(value):
    if not value:
        return None
    from datetime import datetime
    if isinstance(value, str):
        value = value.strip()
        if not value:
            return None
        for fmt in ('%d/%m/%Y', '%Y-%m-%d'):
            try:
                return datetime.strptime(value, fmt).date()
            except Exception:
                continue
        return None
    return value


def login_view(request):
    if request.user.is_authenticated:
        return redirect('home')
    
    if request.method == 'POST':
        email = request.POST.get('email')
        password = request.POST.get('password')
        
        user = authenticate(request, username=email, password=password)
        if user is not None:
            login(request, user)
            log_action(user, 'USER_LOGIN', f'Login realizado: {email}', request)
            return redirect('home')
        else:
            messages.error(request, 'Email ou senha inválidos.')
    
    return render(request, 'users/login.html')

def forgot_password_view(request):
    if request.user.is_authenticated:
        return redirect('home')
    
    if request.method == 'POST':
        email = request.POST.get('email')
        try:
            user = User.objects.get(email=email)
            # Aqui você implementaria o envio do email de recuperação de senha
            messages.success(request, 'Instruções para recuperação de senha foram enviadas para seu email.')
        except User.DoesNotExist:
            messages.error(request, 'Email não encontrado.')
    
    return render(request, 'users/forgot_password.html')

@login_required
def logout_view(request):
    user = request.user
    log_action(user, 'USER_LOGOUT', f'Logout realizado: {user.email}', request)
    logout(request)
    return redirect('login')


@login_required
def dashboard_view(request):
    from tickets.models import Ticket
    from communications.models import Communication
    from django.db.models import Q, Count
    from django.db import models
    from django.utils import timezone
    
    user = request.user
    
    # Filtro base: TODOS os usuários sempre veem seus próprios chamados
    base_filter = models.Q(created_by=user)
    
    # Estatísticas de tickets
    if user.can_view_all_tickets():
        # Admin vê todos os tickets (incluindo fechados)
        user_tickets = Ticket.objects.all()
    elif user.can_view_sector_tickets():
        # Supervisores veem: seus próprios tickets + tickets dos setores + tickets atribuídos
        user_sectors = list(user.sectors.all())
        if user.sector:
            user_sectors.append(user.sector)
        
        user_tickets = Ticket.objects.filter(
            base_filter |  # Sempre inclui próprios tickets
            models.Q(sector__in=user_sectors) |
            models.Q(assigned_to=user)
        ).distinct()
    else:
        # Usuários comuns veem: seus próprios tickets + tickets atribuídos
        # Excluindo tickets fechados
        user_tickets = Ticket.objects.filter(
            base_filter |  # Sempre inclui próprios tickets
            models.Q(assigned_to=user) |
            models.Q(additional_assignments__user=user, additional_assignments__is_active=True)
        ).exclude(status='FECHADO').distinct()
    
    # Chamados recentes - mesma lógica
    if user.can_view_all_tickets():
        # Admin vê todos os tickets
        sector_recent_tickets = Ticket.objects.all()
    elif user.can_view_sector_tickets():
        # Supervisores veem: seus próprios tickets + tickets dos setores + tickets atribuídos
        user_sectors = list(user.sectors.all())
        if user.sector:
            user_sectors.append(user.sector)
        
        sector_recent_tickets = Ticket.objects.filter(
            base_filter |  # Sempre inclui próprios tickets
            models.Q(sector__in=user_sectors) |
            models.Q(assigned_to=user)
        ).distinct()
    else:
        # Usuários comuns veem: seus próprios tickets + tickets atribuídos
        # Excluindo tickets fechados
        sector_recent_tickets = Ticket.objects.filter(
            base_filter |  # Sempre inclui próprios tickets
            models.Q(assigned_to=user) |
            models.Q(additional_assignments__user=user, additional_assignments__is_active=True)
        ).exclude(status='FECHADO').distinct()
    ticket_stats = {
        'total': user_tickets.count(),
        'abertos': user_tickets.filter(status='ABERTO').count(),
        'em_andamento': user_tickets.filter(status='EM_ANDAMENTO').count(),
        'aguardando_aprovacao': user_tickets.filter(status='AGUARDANDO_APROVACAO').count(),
        'resolvidos': user_tickets.filter(status='RESOLVIDO').count(),
        'fechados': user_tickets.filter(status='FECHADO').count(),
        'overdue': user_tickets.filter(due_date__lt=timezone.now(), status__in=['ABERTO', 'EM_ANDAMENTO']).count(),
    }
    
    # Comunicados não lidos (apenas os ativos)
    now = timezone.now()
    
    unread_communications = Communication.objects.filter(
        Q(recipients=user) | Q(send_to_all=True)
    ).filter(
        Q(active_from__isnull=True) | Q(active_from__lte=now)
    ).filter(
        Q(active_until__isnull=True) | Q(active_until__gte=now)
    ).exclude(
        communicationread__user=user
    ).distinct()[:5]
    
    # Tickets recentes (sempre do setor, exceto superadmin) - limitado a 3
    recent_tickets = sector_recent_tickets.order_by('-created_at')[:3]
    
    # Tickets em atraso que o usuário pode ver
    overdue_tickets = user_tickets.filter(
        due_date__lt=timezone.now(),
        status__in=['ABERTO', 'EM_ANDAMENTO']
    ).order_by('due_date')[:5]
    
    # Stats para o template (compatível com o formato antigo)
    stats = {
        'total_tickets': ticket_stats['total'],
        'open_tickets': ticket_stats['abertos'],
        'pending_tickets': ticket_stats['em_andamento'],
        'closed_tickets': ticket_stats['fechados'],
        'overdue_tickets': ticket_stats['overdue'],
        'active_users': User.objects.filter(is_active=True).count(),
    }
    
    context = {
        'user': user,
        'user_balance': user.balance_cs,
        'ticket_stats': ticket_stats,
        'stats': stats,
        'unread_communications': unread_communications,
        'recent_tickets': recent_tickets,
        'overdue_tickets': overdue_tickets,
    }
    return render(request, 'dashboard.html', context)


@login_required
def admin_panel_view(request):
    """Painel administrativo para superadmin/administrativo"""
    if not request.user.can_manage_users():
        messages.error(request, 'Você não tem permissão para acessar esta área.')
        return redirect('dashboard')
    
    from prizes.models import CSTransaction
    from tickets.models import Ticket
    from django.db.models import Sum
    
    # Calcular C$ em circulação (soma de todos os saldos dos usuários)
    total_cs_circulation = User.objects.aggregate(
        total=Sum('balance_cs')
    )['total'] or 0
    
    # Contar chamados abertos
    open_tickets_count = Ticket.objects.filter(
        status__in=['ABERTO', 'EM_ANDAMENTO']
    ).count()
    
    # Contar transações C$ pendentes
    pending_transactions_count = CSTransaction.objects.filter(
        status='PENDING',
        transaction_type='CREDIT'
    ).count()
    
    context = {
        'total_users': User.objects.count(),
        'total_sectors': Sector.objects.count(),
        'recent_users': User.objects.order_by('-created_at')[:5],
        'pending_transactions_count': pending_transactions_count,
        'total_cs_circulation': total_cs_circulation,
        'open_tickets_count': open_tickets_count,
        'user': request.user,
    }
    return render(request, 'admin/panel.html', context)


@login_required
def manage_users_view(request):
    """Gerenciar usuários"""
    if not request.user.can_manage_users():
        messages.error(request, 'Você não tem permissão para acessar esta área.')
        return redirect('dashboard')
    
    users = User.objects.all().select_related('sector')

    total_users_count = users.count()
    active_users_count = users.filter(is_active=True).count()
    superadmin_count = users.filter(hierarchy='SUPERADMIN').count()
    
    context = {
        'users': users,
        'sectors': Sector.objects.all(),
        'user': request.user,
        'total_users_count': total_users_count,
        'active_users_count': active_users_count,
        'superadmin_count': superadmin_count,
    }
    return render(request, 'admin/users.html', context)


@login_required
def manage_sessions_view(request):
    """Lista todas as sessões ativas dos usuários (IP, localização, dispositivo)."""
    if not request.user.can_access_admin_panel():
        messages.error(request, 'Você não tem permissão para acessar esta área.')
        return redirect('dashboard')

    from django.contrib.sessions.models import Session

    now = timezone.now()
    # Chaves de sessões do Django que ainda são válidas (não expiradas).
    valid_keys = set(
        Session.objects.filter(expire_date__gt=now).values_list('session_key', flat=True)
    )

    # Remove registros órfãos (sessões já expiradas/derrubadas) para manter a lista fiel.
    UserSession.objects.exclude(session_key__in=valid_keys).delete()

    sessions = list(
        UserSession.objects.select_related('user', 'user__sector')
        .filter(session_key__in=valid_keys)
        .order_by('user__first_name', 'user__last_name', '-last_activity')
    )

    current_key = request.session.session_key
    for s in sessions:
        s.is_current = (s.session_key == current_key)

    context = {
        'user': request.user,
        'sessions': sessions,
        'total_sessions': len(sessions),
        'online_users_count': len({s.user_id for s in sessions}),
        'mobile_sessions_count': sum(1 for s in sessions if s.is_mobile),
    }
    return render(request, 'admin/sessions.html', context)


@login_required
@require_POST
def terminate_session_view(request):
    """Derruba uma sessão específica pelo session_key."""
    if not request.user.can_access_admin_panel():
        messages.error(request, 'Você não tem permissão para realizar esta ação.')
        return redirect('dashboard')

    from django.contrib.sessions.models import Session

    session_key = request.POST.get('session_key', '').strip()
    if not session_key:
        messages.error(request, 'Sessão inválida.')
        return redirect('manage_sessions')

    target = UserSession.objects.filter(session_key=session_key).select_related('user').first()
    target_label = target.user.get_full_name() or target.user.email if target else 'usuário'

    Session.objects.filter(session_key=session_key).delete()
    UserSession.objects.filter(session_key=session_key).delete()

    log_action(
        request.user,
        'ADMIN_ACTION',
        f'Sessão derrubada ({session_key[:8]}...) do usuário {target_label}',
        request,
    )

    if session_key == request.session.session_key:
        messages.success(request, 'Sua própria sessão foi encerrada.')
        return redirect('login')

    messages.success(request, f'Sessão de {target_label} derrubada com sucesso.')
    return redirect('manage_sessions')


@login_required
def export_users_excel(request):
    """Exportar dados de usuários em Excel"""
    if not request.user.can_manage_users():
        messages.error(request, 'Você não tem permissão para realizar esta ação.')
        return redirect('dashboard')
    
    # Criar workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Usuários"
    
    # Definir cabeçalhos
    headers = [
        'Username', 'Email', 'Nome Completo', 'Telefone', 
        'Setor ID', 'Setor Nome', 'Hierarquia', 'Saldo C$', 'Ativo', 
        'Data Criação', 'Último Login',
        'CPF', 'PIS', 'Cargo', 'Data Nascimento', 'Data Admissão', 'Se Há Janela de Experiencia', 'Data de Demissão',
        'Login/Código', 'PDV', 'Bairro', 'Cidade',
        'Perfil DISC', 'Tamanho Blusa', 'Tamanho Calça'
    ]
    
    # Estilizar cabeçalhos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Buscar dados dos usuários
    users = User.objects.all().select_related('sector').order_by('first_name', 'last_name')
    
    # Preencher dados
    for row, user in enumerate(users, 2):
        ws.cell(row=row, column=1, value=user.username)
        ws.cell(row=row, column=2, value=user.email)
        ws.cell(row=row, column=3, value=f"{user.first_name} {user.last_name}".strip())
        ws.cell(row=row, column=4, value=user.phone or "")
        ws.cell(row=row, column=5, value=user.sector.id if user.sector else "")
        ws.cell(row=row, column=6, value=user.sector.name if user.sector else "")
        ws.cell(row=row, column=7, value=user.hierarchy)
        ws.cell(row=row, column=8, value=float(user.balance_cs))
        ws.cell(row=row, column=9, value="Sim" if user.is_active else "Não")
        ws.cell(row=row, column=10, value=user.date_joined.strftime("%Y-%m-%d %H:%M:%S"))
        ws.cell(row=row, column=11, value=user.last_login.strftime("%Y-%m-%d %H:%M:%S") if user.last_login else "")
        # Novos campos de RH
        ws.cell(row=row, column=12, value=normalize_cpf(user.cpf))
        ws.cell(row=row, column=13, value=user.pis or "")
        ws.cell(row=row, column=14, value=user.job_title or "")
        ws.cell(row=row, column=15, value=user.birth_date.strftime("%d/%m/%Y") if user.birth_date else "")
        ws.cell(row=row, column=16, value=user.admission_date.strftime("%d/%m/%Y") if user.admission_date else "")
        ws.cell(row=row, column=17, value="Sim" if user.has_experience_window else "Não")
        ws.cell(row=row, column=18, value=user.demission_date.strftime("%d/%m/%Y") if user.demission_date else "")
        ws.cell(row=row, column=19, value=user.login_code or "")
        ws.cell(row=row, column=20, value=user.pdv or "")
        ws.cell(row=row, column=21, value=user.neighborhood or "")
        ws.cell(row=row, column=22, value=user.city or "")
        ws.cell(row=row, column=23, value=user.disc_profile or "")
        ws.cell(row=row, column=24, value=user.uniform_size_shirt or "")
        ws.cell(row=row, column=25, value=user.uniform_size_pants or "")
    
    # Ajustar largura das colunas
    for col in range(1, len(headers) + 1):
        column_letter = get_column_letter(col)
        max_length = 0
        for row in ws[column_letter]:
            try:
                if len(str(row.value)) > max_length:
                    max_length = len(str(row.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Preparar response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="usuarios_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    
    # Salvar workbook na response
    wb.save(response)
    
    # Log da ação
    log_action(
        request.user,
        'USER_EXPORT',
        f'Exportação de dados de usuários realizada',
        request
    )
    
    return response


@login_required
def import_users_excel(request):
    """Importar usuários de Excel"""
    if not request.user.can_manage_users():
        messages.error(request, 'Você não tem permissão para realizar esta ação.')
        return redirect('manage_users')
    
    if request.method == 'POST':
        if 'excel_file' not in request.FILES:
            messages.error(request, 'Nenhum arquivo foi enviado.')
            return redirect('manage_users')
        
        excel_file = request.FILES['excel_file']
        
        try:
            # Ler o arquivo Excel
            wb = openpyxl.load_workbook(excel_file)
            ws = wb.active
            
            created_count = 0
            updated_count = 0
            error_count = 0
            errors = []
            
            # Processar cada linha (pular cabeçalho)
            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
                try:
                    # Suporta planilha antiga (23 colunas) e nova (25 colunas)
                    row_data = list(row)
                    has_window_columns = len(row_data) >= 25
                    if len(row_data) >= 25:
                        row_data = row_data[:25]
                        (username, email, full_name, phone, sector_id, sector_name, hierarchy,
                         balance_cs, is_active, date_created, last_login,
                         cpf, pis, job_title, birth_date, admission_date, has_experience_window,
                         demission_date, login_code, pdv, neighborhood, city,
                         disc_profile, uniform_size_shirt, uniform_size_pants) = row_data
                    else:
                        row_data = row_data + [None] * (23 - len(row_data))
                        (username, email, full_name, phone, sector_id, sector_name, hierarchy,
                         balance_cs, is_active, date_created, last_login,
                         cpf, pis, job_title, birth_date, admission_date,
                         login_code, pdv, neighborhood, city,
                         disc_profile, uniform_size_shirt, uniform_size_pants) = row_data[:23]
                        has_experience_window = None
                        demission_date = None
                    
                    # Separar nome completo: primeira palavra = first_name, restante = last_name
                    name_parts = (full_name or '').strip().split()
                    first_name = name_parts[0] if name_parts else ''
                    last_name = ' '.join(name_parts[1:]) if len(name_parts) > 1 else ''

                    # Segurança: o importador não pode atribuir hierarquia acima da sua.
                    # Ignora o valor da planilha quando não permitido (novo vira PADRAO;
                    # existente mantém a hierarquia atual).
                    if hierarchy and not request.user.can_assign_hierarchy(hierarchy):
                        hierarchy = None

                    if not email:  # Email é obrigatório
                        continue
                    
                    # Buscar setor
                    sector = None
                    if sector_id:
                        try:
                            sector = Sector.objects.get(id=int(sector_id))
                        except Sector.DoesNotExist:
                            pass
                    
                    # Processar datas e campos normalizados
                    birth_date_parsed = _parse_date_excel(birth_date)
                    admission_date_parsed = _parse_date_excel(admission_date)
                    demission_date_parsed = _parse_date_excel(demission_date)
                    has_experience_window_parsed = _parse_bool_excel(has_experience_window, default=False) if has_window_columns else None
                    normalized_cpf = normalize_cpf(cpf)
                    
                    # Verificar se usuário já existe
                    user, created = User.objects.get_or_create(
                        email=email,
                        defaults={
                            'username': username or email,
                            'first_name': first_name or '',
                            'last_name': last_name or '',
                            'phone': phone or '',
                            'sector': sector,
                            'hierarchy': hierarchy or 'PADRAO',
                            'balance_cs': Decimal(str(balance_cs)) if balance_cs else Decimal('0'),
                            'is_active': _parse_bool_excel(is_active, default=True),
                            'cpf': normalized_cpf,
                            'pis': pis or '',
                            'job_title': job_title or '',
                            'birth_date': birth_date_parsed,
                            'admission_date': admission_date_parsed,
                            'has_experience_window': has_experience_window_parsed if has_experience_window_parsed is not None else False,
                            'demission_date': demission_date_parsed,
                            'login_code': login_code or '',
                            'pdv': pdv or '',
                            'neighborhood': neighborhood or '',
                            'city': city or '',
                            'disc_profile': disc_profile or '',
                            'uniform_size_shirt': uniform_size_shirt or '',
                            'uniform_size_pants': uniform_size_pants or '',
                        }
                    )
                    
                    if created:
                        # Definir senha padrão para novos usuários
                        user.set_password('123456')  # Senha padrão
                        user.save()
                        created_count += 1
                    else:
                        # Atualizar usuário existente
                        user.username = username or user.username
                        user.first_name = first_name or user.first_name
                        user.last_name = last_name or user.last_name
                        user.phone = phone or user.phone
                        if sector:
                            user.sector = sector
                        if hierarchy:
                            user.hierarchy = hierarchy
                        if balance_cs is not None:
                            user.balance_cs = Decimal(str(balance_cs))
                        if is_active is not None:
                            user.is_active = _parse_bool_excel(is_active, default=user.is_active)
                        # Novos campos
                        if cpf not in [None, '']:
                            user.cpf = normalized_cpf
                        if pis:
                            user.pis = pis
                        if job_title:
                            user.job_title = job_title
                        if birth_date_parsed:
                            user.birth_date = birth_date_parsed
                        if admission_date_parsed:
                            user.admission_date = admission_date_parsed
                        if has_window_columns:
                            user.has_experience_window = has_experience_window_parsed
                            user.demission_date = demission_date_parsed
                        if login_code:
                            user.login_code = login_code
                        if pdv:
                            user.pdv = pdv
                        if neighborhood:
                            user.neighborhood = neighborhood
                        if city:
                            user.city = city
                        if disc_profile:
                            user.disc_profile = disc_profile
                        if uniform_size_shirt:
                            user.uniform_size_shirt = uniform_size_shirt
                        if uniform_size_pants:
                            user.uniform_size_pants = uniform_size_pants
                        user.save()
                        updated_count += 1
                        
                except Exception as e:
                    error_count += 1
                    errors.append(f'Linha {row_num}: {str(e)}')
                    continue
            
            # Mensagem de resultado
            if created_count > 0 or updated_count > 0:
                message = f'Importação concluída! {created_count} usuários criados, {updated_count} usuários atualizados.'
                if error_count > 0:
                    message += f' {error_count} erros encontrados.'
                messages.success(request, message)
            else:
                messages.warning(request, 'Nenhum usuário foi importado.')
            
            if errors:
                for error in errors[:5]:  # Mostrar apenas os primeiros 5 erros
                    messages.error(request, error)
            
            # Log da ação
            log_action(
                request.user,
                'USER_IMPORT',
                f'Importação de usuários: {created_count} criados, {updated_count} atualizados, {error_count} erros',
                request
            )
            
        except Exception as e:
            messages.error(request, f'Erro ao processar arquivo: {str(e)}')
    
    return redirect('manage_users')


@login_required
def import_colaboradores_csv(request):
    """Importar dados de colaboradores de CSV usando nome como chave"""
    if not request.user.can_manage_users():
        messages.error(request, 'Você não tem permissão para realizar esta ação.')
        return redirect('manage_users')
    
    if request.method == 'POST':
        if 'csv_file' not in request.FILES:
            messages.error(request, 'Nenhum arquivo foi enviado.')
            return redirect('manage_users')
        
        csv_file = request.FILES['csv_file']
        
        try:
            import csv
            import io
            from datetime import datetime
            
            # Ler o arquivo CSV
            decoded_file = csv_file.read().decode('utf-8-sig')
            reader = csv.DictReader(io.StringIO(decoded_file), delimiter=';')
            
            updated_count = 0
            not_found_count = 0
            error_count = 0
            errors = []
            not_found_names = []
            
            for row_num, row in enumerate(reader, 2):
                try:
                    # Pegar o nome do colaborador (chave)
                    colaborador_nome = row.get('Colaborador', '').strip()
                    
                    if not colaborador_nome:
                        continue
                    
                    # Normalizar nome (remover acentos e espaços extras)
                    import unicodedata
                    def normalize_name(name):
                        # Remover acentos
                        nfkd = unicodedata.normalize('NFKD', name)
                        without_accents = ''.join(c for c in nfkd if not unicodedata.combining(c))
                        # Converter para minúsculas e remover espaços extras
                        return ' '.join(without_accents.lower().split())
                    
                    # Separar nome: primeira palavra = first_name, restante = last_name
                    name_parts = colaborador_nome.split()
                    first_name = name_parts[0] if name_parts else ''
                    last_name = ' '.join(name_parts[1:]) if len(name_parts) > 1 else ''
                    full_name_normalized = normalize_name(colaborador_nome)
                    
                    # Tentar várias estratégias de busca
                    user = None
                    
                    # 1. Busca exata por first_name e last_name
                    user = User.objects.filter(
                        first_name__iexact=first_name,
                        last_name__iexact=last_name
                    ).first()
                    
                    # 2. Busca por nome completo concatenado
                    if not user:
                        for u in User.objects.all():
                            db_full_name = f"{u.first_name} {u.last_name}".strip()
                            if normalize_name(db_full_name) == full_name_normalized:
                                user = u
                                break
                    
                    # 3. Busca por primeiro nome + primeiro sobrenome (ignora nomes do meio)
                    if not user and len(name_parts) >= 2:
                        first = name_parts[0]
                        last = name_parts[-1]  # Último nome
                        for u in User.objects.filter(first_name__iexact=first):
                            if u.last_name and normalize_name(u.last_name).endswith(normalize_name(last)):
                                user = u
                                break
                            if u.last_name and normalize_name(last) in normalize_name(u.last_name):
                                user = u
                                break
                    
                    # 4. Busca contém (mais flexível)
                    if not user:
                        for u in User.objects.filter(first_name__iexact=first_name):
                            db_full = normalize_name(f"{u.first_name} {u.last_name}")
                            if full_name_normalized in db_full or db_full in full_name_normalized:
                                user = u
                                break
                    
                    if not user:
                        not_found_count += 1
                        not_found_names.append(colaborador_nome)
                        continue
                    
                    # Atualizar campos opcionais (apenas se tiver valor válido)
                    updated = False
                    
                    # Cargo
                    cargo = row.get('Cargo', '').strip()
                    if cargo and cargo != '#N/D':
                        user.job_title = cargo
                        updated = True
                    
                    # Data de Nascimento
                    birth_date_str = row.get('Data de Nascimento', '').strip()
                    if birth_date_str and birth_date_str != '#N/D':
                        try:
                            birth_date = datetime.strptime(birth_date_str, '%d/%m/%Y').date()
                            user.birth_date = birth_date
                            updated = True
                        except:
                            pass
                    
                    # CPF
                    cpf = row.get('CPF', '').strip()
                    if cpf and cpf != '#N/D':
                        user.cpf = normalize_cpf(cpf)
                        updated = True
                    
                    # PIS
                    pis = row.get('PIS', '').strip()
                    if pis and pis != '#N/D':
                        user.pis = pis
                        updated = True
                    
                    # Admissão
                    admission_str = row.get('Admissão', '').strip()
                    if admission_str and admission_str != '#N/D':
                        try:
                            admission_date = datetime.strptime(admission_str, '%d/%m/%Y').date()
                            user.admission_date = admission_date
                            updated = True
                        except:
                            pass
                    
                    # Login/Código (não altera username, salva em campo separado)
                    login = row.get('Login', '').strip()
                    if login and login != '#N/D':
                        user.login_code = login
                        updated = True
                    
                    # PDV
                    pdv = row.get('PDV', '').strip()
                    if pdv and pdv != '#N/D':
                        user.pdv = pdv
                        updated = True
                    
                    # Bairro
                    bairro = row.get('Bairro', '').strip()
                    if bairro and bairro != '#N/D':
                        user.neighborhood = bairro
                        updated = True
                    
                    # Cidade
                    cidade = row.get('Cidade', '').strip()
                    if cidade and cidade != '#N/D' and cidade != '0':
                        user.city = cidade
                        updated = True
                    
                    # Confianças - atualizar balance_cs se for um número válido
                    confiancas = row.get('Confianças', '').strip()
                    if confiancas and confiancas != '#N/D':
                        try:
                            # Tratar número com vírgula como decimal
                            confiancas = confiancas.replace(',', '.')
                            balance = Decimal(confiancas)
                            user.balance_cs = balance
                            updated = True
                        except:
                            pass
                    
                    if updated:
                        user.save()
                        updated_count += 1
                    
                except Exception as e:
                    error_count += 1
                    errors.append(f'Linha {row_num}: {str(e)}')
                    continue
            
            # Mensagem de resultado
            if updated_count > 0:
                message = f'Importação concluída! {updated_count} usuários atualizados.'
                if not_found_count > 0:
                    message += f' {not_found_count} colaboradores não encontrados.'
                if error_count > 0:
                    message += f' {error_count} erros encontrados.'
                messages.success(request, message)
            else:
                messages.warning(request, f'Nenhum usuário foi atualizado. {not_found_count} colaboradores não encontrados.')
            
            # Mostrar alguns nomes não encontrados
            if not_found_names:
                names_preview = ', '.join(not_found_names[:5])
                if len(not_found_names) > 5:
                    names_preview += f' e mais {len(not_found_names) - 5}...'
                messages.warning(request, f'Colaboradores não encontrados: {names_preview}')
            
            if errors:
                for error in errors[:5]:
                    messages.error(request, error)
            
            # Log da ação
            log_action(
                request.user,
                'USER_IMPORT_COLABORADORES',
                f'Importação de colaboradores: {updated_count} atualizados, {not_found_count} não encontrados, {error_count} erros',
                request
            )
            
        except Exception as e:
            messages.error(request, f'Erro ao processar arquivo: {str(e)}')
    
    return redirect('manage_users')


@login_required
def create_user_view(request):
    """Criar novo usuário"""
    if not request.user.can_manage_users():
        messages.error(request, 'Você não tem permissão para acessar esta área.')
        return redirect('dashboard')
    
    if request.method == 'POST':
        email = request.POST.get('email')
        username = request.POST.get('username')
        full_name = request.POST.get('full_name', '').strip()
        name_parts = full_name.split()
        first_name = name_parts[0] if name_parts else ''
        last_name = ' '.join(name_parts[1:]) if len(name_parts) > 1 else ''
        password = request.POST.get('password')
        sector_id = request.POST.get('sector')
        sectors_ids = request.POST.getlist('sectors')
        hierarchy = request.POST.get('hierarchy')
        phone = request.POST.get('phone')
        disc_profile = request.POST.get('disc_profile')
        uniform_size_shirt = request.POST.get('uniform_size_shirt')
        uniform_size_pants = request.POST.get('uniform_size_pants')
        is_active = request.POST.get('is_active') == 'on'

        # Situação do colaborador (ativo/inativo/afastado)
        valid_status = dict(User.STATUS_CHOICES)
        status = request.POST.get('status') or User.STATUS_ATIVO
        if status not in valid_status:
            status = User.STATUS_ATIVO
        inactivation_reason = request.POST.get('inactivation_reason', '').strip()
        leave_reason = request.POST.get('leave_reason', '').strip()
        leave_attachment = request.FILES.get('leave_attachment')
        vacation_start_date = request.POST.get('vacation_start_date', '').strip()
        vacation_end_date = request.POST.get('vacation_end_date', '').strip()

        # Campos de RH
        cpf = normalize_cpf(request.POST.get('cpf', ''))
        pis = request.POST.get('pis', '')
        job_title = request.POST.get('job_title', '')
        birth_date = request.POST.get('birth_date', '')
        admission_date = request.POST.get('admission_date', '')
        has_experience_window = request.POST.get('has_experience_window') == 'on'
        demission_date = request.POST.get('demission_date', '')
        login_code = request.POST.get('login_code', '')
        pdv = request.POST.get('pdv', '')
        neighborhood = request.POST.get('neighborhood', '')
        city = request.POST.get('city', '')
        
        context = {
            'sectors': Sector.objects.all(),
            'hierarchy_choices': request.user.assignable_hierarchy_choices(),
            'user': request.user,
        }

        # Segurança: não permitir criar usuário com hierarquia acima da própria
        if not request.user.can_assign_hierarchy(hierarchy):
            messages.error(request, 'Você não pode criar um usuário com hierarquia superior à sua.')
            return render(request, 'admin/create_user.html', context)

        if (status == User.STATUS_FERIAS and vacation_start_date and vacation_end_date
                and vacation_end_date < vacation_start_date):
            messages.error(request, 'O fim das férias não pode ser anterior ao início.')
            return render(request, 'admin/create_user.html', context)

        try:
            # Verificar se email já existe
            if User.objects.filter(email=email).exists():
                messages.error(request, 'Email já existe.')
                return render(request, 'admin/create_user.html', context)
            
            # Verificar se username já existe
            if User.objects.filter(username=username).exists():
                messages.error(request, 'Nome de usuário já existe.')
                return render(request, 'admin/create_user.html', context)
            
            sector = get_object_or_404(Sector, id=sector_id) if sector_id else None
            
            user = User.objects.create_user(
                username=username,
                email=email,
                password=password,
                first_name=first_name,
                last_name=last_name,
                sector=sector,
                hierarchy=hierarchy,
                phone=phone,
                disc_profile=disc_profile,
                uniform_size_shirt=uniform_size_shirt,
                uniform_size_pants=uniform_size_pants,
                is_active=is_active,
                status=status,
                inactivation_reason=inactivation_reason if status == User.STATUS_INATIVO else '',
                leave_reason=leave_reason if status == User.STATUS_AFASTADO else '',
                vacation_start_date=(vacation_start_date or None) if status == User.STATUS_FERIAS else None,
                vacation_end_date=(vacation_end_date or None) if status == User.STATUS_FERIAS else None,
                cpf=cpf,
                pis=pis,
                job_title=job_title,
                birth_date=birth_date if birth_date else None,
                admission_date=admission_date if admission_date else None,
                has_experience_window=has_experience_window,
                demission_date=demission_date if demission_date else None,
                login_code=login_code,
                pdv=pdv,
                neighborhood=neighborhood,
                city=city,
            )
            
            # Anexo de afastamento (somente quando a situação é "Afastado")
            if status == User.STATUS_AFASTADO and leave_attachment:
                user.leave_attachment = leave_attachment
                user.save()

            # Atualizar setores múltiplos
            sectors = Sector.objects.filter(id__in=sectors_ids) if sectors_ids else Sector.objects.none()
            user.sectors.set(sectors)

            # Se não tem setor principal definido, definir o primeiro da lista
            if not user.sector and sectors.exists():
                user.sector = sectors.first()
                user.save()

            log_action(
                request.user,
                'USER_CREATE',
                f'Usuário criado: {user.full_name} ({user.email})',
                request
            )
            
            messages.success(request, f'Usuário {user.full_name} criado com sucesso!')
            return redirect('manage_users')
            
        except Exception as e:
            messages.error(request, f'Erro ao criar usuário: {str(e)}')
    
    context = {
        'sectors': Sector.objects.all(),
        'hierarchy_choices': request.user.assignable_hierarchy_choices(),
        'user': request.user,
    }
    return render(request, 'admin/create_user.html', context)


@login_required
def edit_user_view(request, user_id):
    """Editar usuário existente"""
    if not request.user.can_manage_users():
        messages.error(request, 'Você não tem permissão para acessar esta área.')
        return redirect('dashboard')
    
    user_to_edit = get_object_or_404(User.objects.prefetch_related('sectors'), id=user_id)

    if request.method == 'POST':
        email = request.POST.get('email')
        username = request.POST.get('username')
        full_name = request.POST.get('full_name', '').strip()
        # Separar nome completo: primeira palavra = first_name, restante = last_name
        name_parts = full_name.split()
        first_name = name_parts[0] if name_parts else ''
        last_name = ' '.join(name_parts[1:]) if len(name_parts) > 1 else ''
        sector_id = request.POST.get('sector')
        sectors_ids = request.POST.getlist('sectors')  # Múltiplos setores
        hierarchy = request.POST.get('hierarchy')
        phone = request.POST.get('phone')
        disc_profile = request.POST.get('disc_profile')
        uniform_size_shirt = request.POST.get('uniform_size_shirt')
        uniform_size_pants = request.POST.get('uniform_size_pants')
        is_active = request.POST.get('is_active') == 'on'

        # Situação do colaborador (ativo/inativo/afastado)
        valid_status = dict(User.STATUS_CHOICES)
        status = request.POST.get('status') or User.STATUS_ATIVO
        if status not in valid_status:
            status = User.STATUS_ATIVO
        inactivation_reason = request.POST.get('inactivation_reason', '').strip()
        leave_reason = request.POST.get('leave_reason', '').strip()
        leave_return_date = request.POST.get('leave_return_date', '').strip()
        leave_attachment = request.FILES.get('leave_attachment')
        remove_leave_attachment = request.POST.get('remove_leave_attachment') == 'on'
        vacation_start_date = request.POST.get('vacation_start_date', '').strip()
        vacation_end_date = request.POST.get('vacation_end_date', '').strip()

        # Novos campos de RH
        cpf = normalize_cpf(request.POST.get('cpf', ''))
        pis = request.POST.get('pis', '')
        job_title = request.POST.get('job_title', '')
        birth_date = request.POST.get('birth_date', '')
        admission_date = request.POST.get('admission_date', '')
        has_experience_window = request.POST.get('has_experience_window') == 'on'
        demission_date = request.POST.get('demission_date', '')
        login_code = request.POST.get('login_code', '')
        pdv = request.POST.get('pdv', '')
        neighborhood = request.POST.get('neighborhood', '')
        city = request.POST.get('city', '')
        
        try:
            # Verificar se email já existe (exceto para o próprio usuário)
            if User.objects.filter(email=email).exclude(id=user_id).exists():
                messages.error(request, 'Email já existe.')
            else:
                # Segurança: não permitir gerenciar usuário de hierarquia superior
                # à própria, nem atribuir (a si mesmo ou a outro) uma hierarquia
                # acima da sua — evita autopromoção e escalonamento.
                if not request.user.can_manage_hierarchy_of(user_to_edit):
                    messages.error(request, 'Você não pode editar um usuário de hierarquia superior à sua.')
                    return redirect('edit_user', user_id=user_to_edit.id)

                if not request.user.can_assign_hierarchy(hierarchy):
                    messages.error(request, 'Você não pode definir uma hierarquia maior que a sua.')
                    return redirect('edit_user', user_id=user_to_edit.id)

                sector = get_object_or_404(Sector, id=sector_id) if sector_id else None
                
                user_to_edit.email = email
                user_to_edit.username = username
                user_to_edit.first_name = first_name
                user_to_edit.last_name = last_name
                user_to_edit.sector = sector
                user_to_edit.hierarchy = hierarchy
                user_to_edit.phone = phone
                user_to_edit.disc_profile = disc_profile
                user_to_edit.uniform_size_shirt = uniform_size_shirt
                user_to_edit.uniform_size_pants = uniform_size_pants
                user_to_edit.is_active = is_active

                # Salvar situação do colaborador, mantendo os motivos coerentes
                user_to_edit.status = status
                user_to_edit.inactivation_reason = inactivation_reason if status == User.STATUS_INATIVO else ''
                user_to_edit.leave_reason = leave_reason if status == User.STATUS_AFASTADO else ''
                if status == User.STATUS_AFASTADO:
                    user_to_edit.leave_return_date = leave_return_date or None
                    if leave_attachment:
                        user_to_edit.leave_attachment = leave_attachment
                    elif remove_leave_attachment:
                        user_to_edit.leave_attachment = None
                else:
                    # Ao sair do afastamento, descarta o retorno e o anexo vinculado
                    user_to_edit.leave_return_date = None
                    user_to_edit.leave_attachment = None

                if status == User.STATUS_FERIAS:
                    if vacation_start_date and vacation_end_date and vacation_end_date < vacation_start_date:
                        messages.error(request, 'O fim das férias não pode ser anterior ao início.')
                        return redirect('edit_user', user_id=user_to_edit.id)
                    user_to_edit.vacation_start_date = vacation_start_date or None
                    user_to_edit.vacation_end_date = vacation_end_date or None
                else:
                    # Ao sair das férias, limpa o período para não sobrar data solta
                    user_to_edit.vacation_start_date = None
                    user_to_edit.vacation_end_date = None

                # Salvar novos campos de RH
                user_to_edit.cpf = cpf
                user_to_edit.pis = pis
                user_to_edit.job_title = job_title
                user_to_edit.birth_date = birth_date if birth_date else None
                user_to_edit.admission_date = admission_date if admission_date else None
                user_to_edit.has_experience_window = has_experience_window
                user_to_edit.demission_date = demission_date if demission_date else None
                user_to_edit.login_code = login_code
                user_to_edit.pdv = pdv
                user_to_edit.neighborhood = neighborhood
                user_to_edit.city = city
                
                user_to_edit.save()
                
                # Atualizar setores múltiplos
                sectors = Sector.objects.filter(id__in=sectors_ids) if sectors_ids else Sector.objects.none()
                user_to_edit.sectors.set(sectors)
                
                # Se não tem setor principal definido, definir o primeiro da lista
                if not user_to_edit.sector and sectors.exists():
                    user_to_edit.sector = sectors.first()
                    user_to_edit.save()
                
                log_action(
                    request.user, 
                    'USER_UPDATE', 
                    f'Usuário editado: {user_to_edit.full_name} ({user_to_edit.email})',
                    request
                )
                
                messages.success(request, f'Usuário {user_to_edit.full_name} atualizado com sucesso!')
                return redirect('manage_users')
                
        except Exception as e:
            messages.error(request, f'Erro ao atualizar usuário: {str(e)}')
    
    context = {
        'user_to_edit': user_to_edit,
        'sectors': Sector.objects.all(),
        'hierarchy_choices': request.user.assignable_hierarchy_choices(),
        'user': request.user,
    }
    return render(request, 'admin/edit_user.html', context)


# ==========================================================================
# Pré-cadastro de colaboradores e documentos
# ==========================================================================

# Tipos de arquivo aceitos para documentos e limite de tamanho (público/admin)
DOCUMENT_ALLOWED_EXTENSIONS = ['.pdf', '.jpg', '.jpeg', '.png', '.webp', '.doc', '.docx']
DOCUMENT_MAX_SIZE = 20 * 1024 * 1024  # 20MB por documento


def _document_is_valid(uploaded_file):
    """Valida extensão e tamanho de um documento enviado. Retorna (ok, erro)."""
    import os
    if uploaded_file.size > DOCUMENT_MAX_SIZE:
        return False, f'O arquivo "{uploaded_file.name}" excede o limite de 20MB.'
    ext = os.path.splitext(uploaded_file.name)[1].lower()
    if ext not in DOCUMENT_ALLOWED_EXTENSIONS:
        return False, f'O arquivo "{uploaded_file.name}" tem um formato não permitido (use PDF, imagem ou Word).'
    return True, ''


def _store_user_document(target_user, uploaded_file, document_type=None, document_name='', uploaded_by=None):
    """Cria um UserDocument a partir de um arquivo enviado."""
    return UserDocument.objects.create(
        user=target_user,
        document_type=document_type,
        document_name=document_name or (document_type.name if document_type else ''),
        file=uploaded_file,
        original_filename=uploaded_file.name,
        file_size=uploaded_file.size,
        content_type=getattr(uploaded_file, 'content_type', '') or '',
        uploaded_by=uploaded_by,
    )


@login_required
def pre_register_user_view(request):
    """Cria um pré-cadastro (dados básicos) e gera um link para o colaborador concluir."""
    if not request.user.can_manage_users():
        messages.error(request, 'Você não tem permissão para acessar esta área.')
        return redirect('dashboard')

    # Mini relatório dos pré-cadastros + listas mostradas na página.
    pre_reg_pending = list(
        User.objects.filter(pre_registration_status=User.PRE_REG_PENDING)
        .select_related('sector')
        .order_by('-pre_registration_created_at')
    )
    # Concluídos aguardando análise: o gestor aprova ou reprova a partir daqui.
    pre_reg_completed = list(
        User.objects.filter(pre_registration_status=User.PRE_REG_COMPLETED)
        .select_related('sector')
        .prefetch_related('documents')
        .order_by('-pre_registration_completed_at')
    )
    pre_reg_rejected = list(
        User.objects.filter(pre_registration_status=User.PRE_REG_REJECTED)
        .select_related('sector', 'pre_registration_reviewed_by')
        .prefetch_related('pre_registration_extra_documents')
        .order_by('-pre_registration_reviewed_at')
    )
    pre_reg_stats = {
        'pending': len(pre_reg_pending),
        'completed': len(pre_reg_completed),
        'rejected': len(pre_reg_rejected),
        'approved': User.objects.filter(pre_registration_status=User.PRE_REG_APPROVED).count(),
    }
    pre_reg_stats['total'] = (
        pre_reg_stats['pending'] + pre_reg_stats['completed']
        + pre_reg_stats['rejected'] + pre_reg_stats['approved']
    )
    pre_reg_lists = {
        'pre_reg_pending': pre_reg_pending,
        'pre_reg_completed': pre_reg_completed,
        'pre_reg_rejected': pre_reg_rejected,
        'pre_reg_stats': pre_reg_stats,
        'required_documents': RequiredDocument.objects.filter(is_active=True),
    }

    if request.method == 'POST':
        full_name = request.POST.get('full_name', '').strip()
        email = request.POST.get('email', '').strip().lower()
        sector_id = request.POST.get('sector')
        hierarchy = request.POST.get('hierarchy') or 'PADRAO'
        job_title = request.POST.get('job_title', '').strip()
        phone = request.POST.get('phone', '').strip()
        admission_date = request.POST.get('admission_date', '').strip()

        context = {
            'sectors': Sector.objects.all(),
            'hierarchy_choices': request.user.assignable_hierarchy_choices(),
            'user': request.user,
            'form_data': request.POST,
            **pre_reg_lists,
        }

        if not full_name:
            messages.error(request, 'Informe o nome completo do colaborador.')
            return render(request, 'admin/pre_register_user.html', context)
        if not email:
            messages.error(request, 'Informe o e-mail do colaborador.')
            return render(request, 'admin/pre_register_user.html', context)
        if User.objects.filter(email=email).exists():
            messages.error(request, 'Já existe um usuário com este e-mail.')
            return render(request, 'admin/pre_register_user.html', context)

        # Segurança: não permitir hierarquia acima da do próprio usuário
        if User.hierarchy_rank(hierarchy) == -1:
            hierarchy = 'PADRAO'
        if not request.user.can_assign_hierarchy(hierarchy):
            messages.error(request, 'Você não pode definir uma hierarquia maior que a sua.')
            return render(request, 'admin/pre_register_user.html', context)

        name_parts = full_name.split()
        first_name = name_parts[0] if name_parts else ''
        last_name = ' '.join(name_parts[1:]) if len(name_parts) > 1 else ''

        from django.utils.crypto import get_random_string
        # Username único a partir do e-mail
        base_username = (email.split('@')[0] or 'colaborador')[:120]
        username = base_username
        suffix = 1
        while User.objects.filter(username=username).exists():
            suffix += 1
            username = f"{base_username}{suffix}"

        # Token único para o link de conclusão
        token = get_random_string(48)
        while User.objects.filter(pre_registration_token=token).exists():
            token = get_random_string(48)

        sector = get_object_or_404(Sector, id=sector_id) if sector_id else None

        try:
            new_user = User(
                username=username,
                email=email,
                first_name=first_name,
                last_name=last_name,
                sector=sector,
                hierarchy=hierarchy,
                job_title=job_title,
                phone=phone,
                admission_date=admission_date if admission_date else None,
                is_active=False,  # inativo até o colaborador concluir o pré-cadastro
                pre_registration_status=User.PRE_REG_PENDING,
                pre_registration_token=token,
                pre_registration_created_at=timezone.now(),
            )
            new_user.set_unusable_password()
            new_user.save()
            if sector:
                new_user.sectors.set([sector])

            log_action(
                request.user,
                'USER_CREATE',
                f'Pré-cadastro criado: {new_user.full_name} ({new_user.email})',
                request
            )
            messages.success(
                request,
                f'Pré-cadastro de {new_user.full_name} criado! Envie o link abaixo para o colaborador.'
            )
            return redirect('pre_register_link', user_id=new_user.id)
        except Exception as e:
            messages.error(request, f'Erro ao criar pré-cadastro: {str(e)}')
            return render(request, 'admin/pre_register_user.html', context)

    context = {
        'sectors': Sector.objects.all(),
        'hierarchy_choices': request.user.assignable_hierarchy_choices(),
        'user': request.user,
        **pre_reg_lists,
    }
    return render(request, 'admin/pre_register_user.html', context)


@login_required
def pre_register_link_view(request, user_id):
    """Exibe o link de pré-cadastro de um colaborador pendente (para copiar/enviar)."""
    if not request.user.can_manage_users():
        messages.error(request, 'Você não tem permissão para acessar esta área.')
        return redirect('dashboard')

    from django.urls import reverse
    target = get_object_or_404(User, id=user_id)
    if not target.pre_registration_token or target.pre_registration_status != User.PRE_REG_PENDING:
        messages.info(request, 'Este colaborador não possui um pré-cadastro pendente.')
        return redirect('view_user_profile', user_id=target.id)

    link = request.build_absolute_uri(
        reverse('complete_pre_registration', kwargs={'token': target.pre_registration_token})
    )
    context = {
        'user': request.user,
        'target': target,
        'link': link,
    }
    return render(request, 'admin/pre_register_link.html', context)


def _read_pre_registration_personal_data(request):
    """Lê e valida os dados pessoais do formulário de pré-cadastro/ajuste.

    Retorna (valores, erros). Os campos obrigatórios seguem a ficha cadastral:
    RG (número, emissão e órgão), CEP, filiação (pai e mãe) e cor/raça.
    """
    data = request.POST
    values = {
        'full_name': data.get('full_name', '').strip(),
        'cpf': normalize_cpf(data.get('cpf', '')),
        'pis': data.get('pis', '').strip(),
        'phone': data.get('phone', '').strip(),
        'birth_date': data.get('birth_date', '').strip(),
        'neighborhood': data.get('neighborhood', '').strip(),
        'city': data.get('city', '').strip(),
        'cep': data.get('cep', '').strip(),
        'uniform_size_shirt': data.get('uniform_size_shirt', '').strip(),
        'uniform_size_pants': data.get('uniform_size_pants', '').strip(),
        'rg': data.get('rg', '').strip(),
        'rg_issue_date': data.get('rg_issue_date', '').strip(),
        'rg_issuer': data.get('rg_issuer', '').strip(),
        'voter_title': data.get('voter_title', '').strip(),
        'voter_zone': data.get('voter_zone', '').strip(),
        'voter_section': data.get('voter_section', '').strip(),
        'father_name': data.get('father_name', '').strip(),
        'mother_name': data.get('mother_name', '').strip(),
        'skin_color': data.get('skin_color', '').strip(),
    }

    errors = []
    if not values['rg']:
        errors.append('Informe o número do RG.')
    if not values['rg_issue_date']:
        errors.append('Informe a data de emissão do RG.')
    if not values['rg_issuer']:
        errors.append('Informe o órgão emissor do RG.')
    if not values['cep']:
        errors.append('Informe o CEP.')
    if not values['father_name']:
        errors.append('Informe o nome completo do pai.')
    if not values['mother_name']:
        errors.append('Informe o nome completo da mãe.')
    if values['skin_color'] not in {c[0] for c in User.SKIN_COLOR_CHOICES}:
        errors.append('Selecione a cor/raça.')

    return values, errors


def _apply_pre_registration_personal_data(target, values):
    """Aplica no usuário os dados pessoais já lidos/validados do formulário."""
    full_name = values['full_name']
    if full_name:
        parts = full_name.split()
        target.first_name = parts[0] if parts else target.first_name
        target.last_name = ' '.join(parts[1:]) if len(parts) > 1 else ''
    target.cpf = values['cpf']
    target.pis = values['pis']
    if values['phone']:
        target.phone = values['phone']
    target.birth_date = values['birth_date'] or None
    target.neighborhood = values['neighborhood']
    target.city = values['city']
    target.cep = values['cep']
    target.uniform_size_shirt = values['uniform_size_shirt']
    target.uniform_size_pants = values['uniform_size_pants']
    target.rg = values['rg']
    target.rg_issue_date = values['rg_issue_date'] or None
    target.rg_issuer = values['rg_issuer']
    target.voter_title = values['voter_title']
    target.voter_zone = values['voter_zone']
    target.voter_section = values['voter_section']
    target.father_name = values['father_name']
    target.mother_name = values['mother_name']
    target.skin_color = values['skin_color']


def complete_pre_registration_view(request, token):
    """Página PÚBLICA onde o colaborador conclui o cadastro e anexa os documentos."""
    target = User.objects.filter(pre_registration_token=token).first()

    # Token inválido, expirado ou já utilizado
    if not target or target.pre_registration_status != User.PRE_REG_PENDING:
        return render(request, 'users/pre_register_invalid.html', status=404)

    required_documents = list(RequiredDocument.objects.filter(is_active=True))

    if request.method == 'POST':
        password = request.POST.get('password', '')
        password_confirm = request.POST.get('password_confirm', '')

        values, errors = _read_pre_registration_personal_data(request)
        if len(password) < 6:
            errors.append('A senha deve ter pelo menos 6 caracteres.')
        if password != password_confirm:
            errors.append('As senhas não conferem.')

        # Validar documentos (obrigatórios presentes; todos os enviados válidos)
        docs_to_save = []
        for doc in required_documents:
            uploaded = request.FILES.get(f'document_{doc.id}')
            if not uploaded:
                if doc.is_required:
                    errors.append(f'O documento "{doc.name}" é obrigatório.')
                continue
            ok, err = _document_is_valid(uploaded)
            if not ok:
                errors.append(err)
            else:
                docs_to_save.append((doc, uploaded))

        if errors:
            for e in errors:
                messages.error(request, e)
            context = {
                'target': target,
                'required_documents': required_documents,
                'skin_color_choices': User.SKIN_COLOR_CHOICES,
                'form_data': request.POST,
            }
            return render(request, 'users/pre_register_complete.html', context)

        # Atualizar dados do colaborador
        _apply_pre_registration_personal_data(target, values)

        # Auto-ativação: define a senha e ativa a conta
        target.set_password(password)
        target.is_active = True
        target.pre_registration_status = User.PRE_REG_COMPLETED
        target.pre_registration_completed_at = timezone.now()
        target.pre_registration_token = None  # invalida o link após a conclusão
        target.save()

        for doc, uploaded in docs_to_save:
            _store_user_document(target, uploaded, document_type=doc, document_name=doc.name, uploaded_by=target)

        log_action(
            target,
            'USER_UPDATE',
            f'Pré-cadastro concluído: {target.full_name} ({target.email})',
            request
        )
        return redirect('pre_registration_success')

    context = {
        'target': target,
        'required_documents': required_documents,
        'skin_color_choices': User.SKIN_COLOR_CHOICES,
        'form_data': None,
    }
    return render(request, 'users/pre_register_complete.html', context)


def pre_registration_success_view(request):
    """Página pública de confirmação após concluir o pré-cadastro."""
    return render(request, 'users/pre_register_success.html')


# Situações em que o gestor ainda pode aprovar/reprovar. APPROVED e REJECTED
# continuam na lista para permitir corrigir uma análise feita por engano — sem
# isso, um colaborador reprovado sem querer ficaria travado fora do portal.
PRE_REG_REVIEWABLE_STATUSES = (
    User.PRE_REG_COMPLETED,
    User.PRE_REG_APPROVED,
    User.PRE_REG_REJECTED,
)


def _safe_next_url(request, fallback='pre_register_user'):
    """Destino pós-ação vindo do formulário, recusando redirect para fora do site."""
    from django.utils.http import url_has_allowed_host_and_scheme

    next_url = request.POST.get('next')
    if next_url and url_has_allowed_host_and_scheme(
        next_url,
        allowed_hosts={request.get_host()},
        require_https=request.is_secure(),
    ):
        return next_url
    return fallback


@login_required
@require_POST
def approve_pre_registration_view(request, user_id):
    """Aprova um pré-cadastro já preenchido pelo colaborador."""
    if not request.user.can_review_pre_registrations():
        messages.error(request, 'Você não tem permissão para analisar pré-cadastros.')
        return redirect('dashboard')

    target = get_object_or_404(User, id=user_id)
    if target.pre_registration_status not in PRE_REG_REVIEWABLE_STATUSES:
        messages.error(request, 'Este pré-cadastro ainda não foi preenchido pelo colaborador.')
        return redirect('pre_register_user')

    target.pre_registration_status = User.PRE_REG_APPROVED
    target.pre_registration_rejection_reason = ''
    target.pre_registration_reviewed_at = timezone.now()
    target.pre_registration_reviewed_by = request.user
    target.save(update_fields=[
        'pre_registration_status',
        'pre_registration_rejection_reason',
        'pre_registration_reviewed_at',
        'pre_registration_reviewed_by',
        'updated_at',
    ])
    target.pre_registration_extra_documents.clear()

    log_action(
        request.user,
        'USER_UPDATE',
        f'Pré-cadastro aprovado: {target.full_name} ({target.email})',
        request
    )
    messages.success(request, f'Pré-cadastro de {target.full_name} aprovado.')
    return redirect(_safe_next_url(request))


@login_required
@require_POST
def reject_pre_registration_view(request, user_id):
    """Reprova um pré-cadastro, registrando o motivo e os documentos a reenviar."""
    if not request.user.can_review_pre_registrations():
        messages.error(request, 'Você não tem permissão para analisar pré-cadastros.')
        return redirect('dashboard')

    target = get_object_or_404(User, id=user_id)
    redirect_to = _safe_next_url(request)

    if target.pre_registration_status not in PRE_REG_REVIEWABLE_STATUSES:
        messages.error(request, 'Este pré-cadastro ainda não foi preenchido pelo colaborador.')
        return redirect('pre_register_user')

    reason = request.POST.get('rejection_reason', '').strip()
    if not reason:
        messages.error(request, 'Informe o motivo da reprovação para o colaborador saber o que ajustar.')
        return redirect(redirect_to)

    extra_document_ids = request.POST.getlist('extra_documents')
    extra_documents = list(
        RequiredDocument.objects.filter(id__in=extra_document_ids, is_active=True)
    ) if extra_document_ids else []

    target.pre_registration_status = User.PRE_REG_REJECTED
    target.pre_registration_rejection_reason = reason
    target.pre_registration_reviewed_at = timezone.now()
    target.pre_registration_reviewed_by = request.user
    target.save(update_fields=[
        'pre_registration_status',
        'pre_registration_rejection_reason',
        'pre_registration_reviewed_at',
        'pre_registration_reviewed_by',
        'updated_at',
    ])
    target.pre_registration_extra_documents.set(extra_documents)

    docs_info = f' | Documentos solicitados: {", ".join(d.name for d in extra_documents)}' if extra_documents else ''
    log_action(
        request.user,
        'USER_UPDATE',
        f'Pré-cadastro reprovado: {target.full_name} ({target.email}) | Motivo: {reason}{docs_info}',
        request
    )
    messages.success(
        request,
        f'Pré-cadastro de {target.full_name} reprovado. O colaborador verá o motivo ao acessar o portal.'
    )
    return redirect(redirect_to)


@login_required
def adjust_pre_registration_view(request):
    """Tela do colaborador cujo pré-cadastro foi reprovado: mostra o motivo e
    permite corrigir os dados/documentos e reenviar para análise."""
    target = request.user

    if not target.needs_pre_registration_adjustment():
        return redirect('home')

    requested_documents = list(target.pre_registration_extra_documents.all())
    requested_ids = {d.id for d in requested_documents}
    other_documents = [
        d for d in RequiredDocument.objects.filter(is_active=True)
        if d.id not in requested_ids
    ]

    if request.method == 'POST':
        values, errors = _read_pre_registration_personal_data(request)
        docs_to_save = []

        # Documentos pedidos pelo gestor são obrigatórios no reenvio.
        for doc in requested_documents:
            uploaded = request.FILES.get(f'document_{doc.id}')
            if not uploaded:
                errors.append(f'O documento "{doc.name}" foi solicitado e precisa ser anexado.')
                continue
            ok, err = _document_is_valid(uploaded)
            if not ok:
                errors.append(err)
            else:
                docs_to_save.append((doc, uploaded))

        # Os demais são opcionais (reenvio por iniciativa do colaborador).
        for doc in other_documents:
            uploaded = request.FILES.get(f'document_{doc.id}')
            if not uploaded:
                continue
            ok, err = _document_is_valid(uploaded)
            if not ok:
                errors.append(err)
            else:
                docs_to_save.append((doc, uploaded))

        if errors:
            for e in errors:
                messages.error(request, e)
            context = {
                'target': target,
                'requested_documents': requested_documents,
                'other_documents': other_documents,
                'skin_color_choices': User.SKIN_COLOR_CHOICES,
                'form_data': request.POST,
            }
            return render(request, 'users/pre_register_adjust.html', context)

        _apply_pre_registration_personal_data(target, values)

        # Volta para a fila de análise e libera o acesso ao portal.
        target.pre_registration_status = User.PRE_REG_COMPLETED
        target.pre_registration_rejection_reason = ''
        target.pre_registration_completed_at = timezone.now()
        target.pre_registration_reviewed_at = None
        target.pre_registration_reviewed_by = None
        target.save()
        target.pre_registration_extra_documents.clear()

        for doc, uploaded in docs_to_save:
            _store_user_document(target, uploaded, document_type=doc, document_name=doc.name, uploaded_by=target)

        log_action(
            target,
            'USER_UPDATE',
            f'Pré-cadastro reenviado após ajuste: {target.full_name} ({target.email})',
            request
        )
        messages.success(request, 'Cadastro reenviado para análise. Obrigado!')
        return redirect('home')

    context = {
        'target': target,
        'requested_documents': requested_documents,
        'other_documents': other_documents,
        'skin_color_choices': User.SKIN_COLOR_CHOICES,
        'form_data': None,
    }
    return render(request, 'users/pre_register_adjust.html', context)


@login_required
def view_user_profile_view(request, user_id):
    """Exibe o perfil completo do colaborador, com todos os dados e documentos."""
    if not request.user.can_manage_users():
        messages.error(request, 'Você não tem permissão para acessar esta área.')
        return redirect('dashboard')

    target = get_object_or_404(
        User.objects.prefetch_related('sectors', 'documents'),
        id=user_id
    )
    documents = target.documents.all().select_related('document_type', 'uploaded_by')

    # Documentos de assinatura eletrônica (app documentos): assinados e pendentes.
    from documentos.models import DocumentSignature
    signatures = (
        DocumentSignature.objects
        .filter(user=target)
        .select_related('document', 'document__category')
        .order_by('-created_at')
    )
    signed_documents = signatures.filter(signed_at__isnull=False)
    pending_documents = signatures.filter(signed_at__isnull=True)

    pre_registration_link = None
    if target.is_pre_registration_pending() and target.pre_registration_token:
        from django.urls import reverse
        pre_registration_link = request.build_absolute_uri(
            reverse('complete_pre_registration', kwargs={'token': target.pre_registration_token})
        )

    context = {
        'user': request.user,
        'target': target,
        'documents': documents,
        'signed_documents': signed_documents,
        'pending_documents': pending_documents,
        'required_documents': RequiredDocument.objects.filter(is_active=True),
        'pre_registration_link': pre_registration_link,
    }
    return render(request, 'admin/user_profile.html', context)


@login_required
def print_user_profile_view(request, user_id):
    """Ficha cadastral do colaborador em layout A4, pronta para impressão."""
    if not request.user.can_manage_users():
        messages.error(request, 'Você não tem permissão para acessar esta área.')
        return redirect('dashboard')

    target = get_object_or_404(
        User.objects.prefetch_related('sectors', 'documents'),
        id=user_id
    )

    context = {
        'user': request.user,
        'target': target,
        'documents': target.documents.all().select_related('document_type'),
        'printed_at': timezone.now(),
    }
    return render(request, 'admin/user_registration_form.html', context)


@login_required
@require_POST
def admin_upload_user_document_view(request, user_id):
    """Permite ao gestor anexar um documento ao colaborador pela tela de perfil."""
    if not request.user.can_manage_users():
        messages.error(request, 'Você não tem permissão para realizar esta ação.')
        return redirect('dashboard')

    target = get_object_or_404(User, id=user_id)
    uploaded = request.FILES.get('file')
    if not uploaded:
        messages.error(request, 'Selecione um arquivo para enviar.')
        return redirect('view_user_profile', user_id=target.id)

    ok, err = _document_is_valid(uploaded)
    if not ok:
        messages.error(request, err)
        return redirect('view_user_profile', user_id=target.id)

    doc_type_id = request.POST.get('document_type') or None
    doc_type = RequiredDocument.objects.filter(id=doc_type_id).first() if doc_type_id else None
    custom_name = request.POST.get('document_name', '').strip()

    _store_user_document(target, uploaded, document_type=doc_type, document_name=custom_name, uploaded_by=request.user)

    log_action(
        request.user,
        'USER_UPDATE',
        f'Documento adicionado ao colaborador {target.full_name} ({target.email})',
        request
    )
    messages.success(request, 'Documento adicionado com sucesso.')
    return redirect('view_user_profile', user_id=target.id)


@login_required
@require_POST
def delete_user_document_view(request, document_id):
    """Remove um documento de um colaborador."""
    if not request.user.can_manage_users():
        messages.error(request, 'Você não tem permissão para realizar esta ação.')
        return redirect('dashboard')

    doc = get_object_or_404(UserDocument, id=document_id)
    target_id = doc.user_id
    name = doc.display_name
    try:
        doc.file.delete(save=False)
    except Exception:
        pass
    doc.delete()

    log_action(request.user, 'USER_UPDATE', f'Documento removido de um colaborador: {name}', request)
    messages.success(request, 'Documento removido com sucesso.')
    return redirect('view_user_profile', user_id=target_id)


def _zip_entry_name(doc, used_names):
    """Nome de arquivo seguro e único (dentro do zip) para um documento."""
    import os
    import re

    ext = doc.file_extension or os.path.splitext(doc.file.name or '')[1].lower()
    base = doc.display_name or 'documento'
    # Remove o que não é seguro em nome de arquivo, preservando acentos.
    base = re.sub(r'[^\w\s.-]', '', base, flags=re.UNICODE).strip()
    base = re.sub(r'\s+', '_', base) or 'documento'
    if ext and base.lower().endswith(ext):
        base = base[:-len(ext)] or 'documento'

    name = f'{base}{ext}'
    counter = 1
    while name.lower() in used_names:
        counter += 1
        name = f'{base}_{counter}{ext}'
    used_names.add(name.lower())
    return name


@login_required
def download_user_documents_zip(request, user_id):
    """Baixa todos os documentos do colaborador compactados em um único .zip."""
    if not request.user.can_manage_users():
        messages.error(request, 'Você não tem permissão para acessar esta área.')
        return redirect('dashboard')

    target = get_object_or_404(User, id=user_id)
    documents = list(target.documents.all().select_related('document_type'))
    if not documents:
        messages.info(request, 'Este colaborador não possui documentos para baixar.')
        return redirect('view_user_profile', user_id=target.id)

    import io
    import zipfile
    from django.utils.text import slugify

    buffer = io.BytesIO()
    used_names = set()
    missing = []
    with zipfile.ZipFile(buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
        for doc in documents:
            try:
                doc.file.open('rb')
                try:
                    content = doc.file.read()
                finally:
                    doc.file.close()
            except Exception:
                # Arquivo ausente/ilegível no storage: não impede o download dos demais.
                missing.append(doc.display_name)
                continue
            zip_file.writestr(_zip_entry_name(doc, used_names), content)

    if len(missing) == len(documents):
        messages.error(request, 'Não foi possível ler os arquivos deste colaborador.')
        return redirect('view_user_profile', user_id=target.id)
    if missing:
        messages.warning(
            request,
            f'{len(missing)} documento(s) não puderam ser lidos e ficaram de fora do zip: {", ".join(missing)}.'
        )

    log_action(
        request.user,
        'USER_UPDATE',
        f'Download de todos os documentos de {target.full_name} ({target.email})',
        request
    )

    filename = f'documentos_{slugify(target.full_name) or target.id}.zip'
    response = HttpResponse(buffer.getvalue(), content_type='application/zip')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@login_required
def manage_required_documents_view(request):
    """Configuração (SUPERADMIN) dos documentos exigidos no pré-cadastro."""
    if not request.user.can_manage_required_documents():
        messages.error(request, 'Apenas o SUPERADMIN pode configurar os documentos exigidos.')
        return redirect('dashboard')

    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        description = request.POST.get('description', '').strip()
        is_required = request.POST.get('is_required') == 'on'
        is_active = request.POST.get('is_active') == 'on'
        order = request.POST.get('order', '0')
        if not name:
            messages.error(request, 'Informe o nome do documento.')
        else:
            RequiredDocument.objects.create(
                name=name,
                description=description,
                is_required=is_required,
                is_active=is_active,
                order=int(order) if str(order).isdigit() else 0,
            )
            messages.success(request, f'Documento "{name}" adicionado.')
        return redirect('manage_required_documents')

    context = {
        'user': request.user,
        'documents': RequiredDocument.objects.all(),
    }
    return render(request, 'admin/required_documents.html', context)


@login_required
@require_POST
def edit_required_document_view(request, doc_id):
    """Edita um documento exigido (SUPERADMIN)."""
    if not request.user.can_manage_required_documents():
        messages.error(request, 'Apenas o SUPERADMIN pode configurar os documentos exigidos.')
        return redirect('dashboard')

    doc = get_object_or_404(RequiredDocument, id=doc_id)
    name = request.POST.get('name', '').strip()
    if name:
        doc.name = name
    doc.description = request.POST.get('description', '').strip()
    doc.is_required = request.POST.get('is_required') == 'on'
    doc.is_active = request.POST.get('is_active') == 'on'
    order = request.POST.get('order', '')
    if str(order).isdigit():
        doc.order = int(order)
    doc.save()
    messages.success(request, f'Documento "{doc.name}" atualizado.')
    return redirect('manage_required_documents')


@login_required
@require_POST
def delete_required_document_view(request, doc_id):
    """Remove um documento exigido (SUPERADMIN)."""
    if not request.user.can_manage_required_documents():
        messages.error(request, 'Apenas o SUPERADMIN pode configurar os documentos exigidos.')
        return redirect('dashboard')

    doc = get_object_or_404(RequiredDocument, id=doc_id)
    name = doc.name
    doc.delete()
    messages.success(request, f'Documento "{name}" removido.')
    return redirect('manage_required_documents')


@login_required
def admin_change_user_password(request, user_id):
    """Alterar senha de usuário pelo admin"""
    if not request.user.can_manage_users():
        messages.error(request, 'Você não tem permissão para acessar esta área.')
        return redirect('dashboard')
    
    user_to_edit = get_object_or_404(User, id=user_id)
    
    if request.method == 'POST':
        new_password = request.POST.get('new_password')
        confirm_password = request.POST.get('confirm_password')
        
        if not new_password or len(new_password) < 6:
            messages.error(request, 'A senha deve ter no mínimo 6 caracteres.')
            return redirect('manage_users')
        
        if new_password != confirm_password:
            messages.error(request, 'As senhas não coincidem.')
            return redirect('manage_users')
        
        try:
            user_to_edit.set_password(new_password)
            user_to_edit.save()
            
            log_action(
                request.user, 
                'USER_PASSWORD_CHANGE', 
                f'Senha alterada pelo admin para: {user_to_edit.full_name} ({user_to_edit.email})',
                request
            )
            
            messages.success(request, f'Senha do usuário {user_to_edit.full_name} alterada com sucesso!')
        except Exception as e:
            messages.error(request, f'Erro ao alterar senha: {str(e)}')
    
    return redirect('manage_users')


@login_required
def manage_cs_view(request):
    """Gerenciar Confianças C$"""
    if not request.user.can_manage_cs():
        messages.error(request, 'Você não tem permissão para acessar esta área.')
        return redirect('dashboard')
    
    from django.core.paginator import Paginator
    from django.db.models import Q
    
    # Obter parâmetros de filtro
    search_query = request.GET.get('search', '')
    sector_filter = request.GET.get('sector', '')
    
    # Filtrar usuários
    users_queryset = User.objects.all().select_related('sector')
    
    if search_query:
        users_queryset = users_queryset.filter(
            Q(first_name__icontains=search_query) |
            Q(last_name__icontains=search_query) |
            Q(email__icontains=search_query)
        )
    
    if sector_filter:
        users_queryset = users_queryset.filter(sector_id=sector_filter)
    
    # Ordenar por nome
    users_queryset = users_queryset.order_by('first_name', 'last_name')
    
    # Paginação
    paginator = Paginator(users_queryset, 25)  # 25 usuários por página
    page_number = request.GET.get('page')
    users = paginator.get_page(page_number)
    
    # Calcular total em circulação corretamente
    total_circulation = User.objects.aggregate(
        total=Sum('balance_cs')
    )['total'] or Decimal('0')
    
    # Calcular média por usuário
    user_count = User.objects.count()
    average_per_user = total_circulation / user_count if user_count > 0 else Decimal('0')
    
    # Obter setores para filtro
    sectors = Sector.objects.all().order_by('name')
    
    context = {
        'users': users,
        'user': request.user,
        'total_circulation': total_circulation,
        'average_per_user': average_per_user,
        'sectors': sectors,
        'search_query': search_query,
        'sector_filter': sector_filter,
    }
    return render(request, 'admin/manage_cs.html', context)


@login_required
def export_cs_excel(request):
    """Exportar dados de Confianças em Excel"""
    if not request.user.can_manage_cs():
        messages.error(request, 'Você não tem permissão para realizar esta ação.')
        return redirect('dashboard')
    
    # Criar workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Confianças C$"
    
    # Definir cabeçalhos
    headers = ['Nome', 'Email', 'Setor', 'Saldo C$']
    
    # Estilizar cabeçalhos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Buscar dados dos usuários
    users = User.objects.all().select_related('sector').order_by('first_name', 'last_name')
    
    # Preencher dados
    for row, user in enumerate(users, 2):
        ws.cell(row=row, column=1, value=user.full_name)
        ws.cell(row=row, column=2, value=user.email)
        ws.cell(row=row, column=3, value=user.sector.name if user.sector else "Sem setor")
        ws.cell(row=row, column=4, value=float(user.balance_cs))
    
    # Ajustar largura das colunas
    for col in range(1, len(headers) + 1):
        column_letter = get_column_letter(col)
        max_length = 0
        for row in ws[column_letter]:
            try:
                if len(str(row.value)) > max_length:
                    max_length = len(str(row.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Adicionar linha de totais
    total_row = len(users) + 2
    ws.cell(row=total_row, column=1, value="TOTAL:")
    ws.cell(row=total_row, column=1).font = Font(bold=True)
    
    total_cs = sum(float(user.balance_cs) for user in users)
    ws.cell(row=total_row, column=4, value=total_cs)
    ws.cell(row=total_row, column=4).font = Font(bold=True)
    
    # Preparar response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="confiancas_cs_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    
    # Salvar workbook na response
    wb.save(response)
    
    # Log da ação
    log_action(
        request.user,
        'CS_EXPORT',
        f'Exportação de dados de Confianças C$ realizada',
        request
    )
    
    return response


@login_required 
def add_cs_view(request, user_id):
    """Adicionar C$ para usuário"""
    if not request.user.can_manage_cs():
        messages.error(request, 'Você não tem permissão para realizar esta ação.')
        return redirect('dashboard')
    
    target_user = get_object_or_404(User, id=user_id)
    
    if request.method == 'POST':
        amount = request.POST.get('amount')
        description = request.POST.get('description', '')
        expiration_date = request.POST.get('expiration_date', '')
        
        try:
            amount = float(amount)
            if amount <= 0:
                messages.error(request, 'Valor deve ser maior que zero.')
                return redirect('manage_cs')
            
            target_user.balance_cs += Decimal(str(amount))
            # Para créditos manuais, não adicionar diretamente ao saldo
            # A transação fica pendente até aprovação
            
            # Processar data de validade
            exp_date = None
            if expiration_date:
                from datetime import datetime
                try:
                    exp_date = datetime.strptime(expiration_date, '%Y-%m-%d').date()
                except ValueError:
                    pass
            
            # Registrar transação como pendente de aprovação
            from prizes.models import CSTransaction
            transaction = CSTransaction.objects.create(
                user=target_user,
                amount=amount,
                transaction_type='CREDIT',
                description=description or f'Adição manual de C$',
                status='PENDING',  # Transação fica pendente
                created_by=request.user,
                expiration_date=exp_date
            )
            
            log_action(
                request.user, 
                'CS_ADD_REQUEST', 
                f'Solicitado C$ {amount} para {target_user.full_name} (Aguardando aprovação)' + (f' - Validade: {exp_date}' if exp_date else ''),
                request
            )
            
            messages.success(request, f'Solicitação de C$ {amount} para {target_user.full_name} enviada para aprovação!')
            
        except ValueError:
            messages.error(request, 'Valor inválido.')
        except Exception as e:
            messages.error(request, f'Erro: {str(e)}')
    
    return redirect('manage_cs')


@login_required
def remove_cs_view(request, user_id):
    """Remover Confianças C$ (não precisa de aprovação)"""
    if not request.user.can_manage_cs():
        messages.error(request, 'Você não tem permissão para realizar esta ação.')
        return redirect('dashboard')
    
    if request.method == 'POST':
        try:
            target_user = User.objects.get(id=user_id)
            amount = Decimal(request.POST.get('amount', '0'))
            description = request.POST.get('description', '')
            
            if amount <= 0:
                messages.error(request, 'Valor deve ser maior que zero.')
                return redirect('manage_cs')
            
            if target_user.balance_cs < amount:
                messages.error(request, f'Saldo insuficiente. Saldo atual: C$ {target_user.balance_cs}')
                return redirect('manage_cs')
            
            # Remover diretamente do saldo (sem aprovação)
            target_user.balance_cs -= amount
            target_user.save()
            
            # Registrar transação como concluída
            from prizes.models import CSTransaction
            transaction = CSTransaction.objects.create(
                user=target_user,
                amount=-amount,  # Valor negativo para indicar remoção
                transaction_type='DEBIT',
                description=description or f'Remoção manual de C$',
                status='COMPLETED',
                created_by=request.user
            )
            
            log_action(
                request.user, 
                'CS_REMOVE', 
                f'Removido C$ {amount} de {target_user.full_name}',
                request
            )
            
            messages.success(request, f'C$ {amount} removido com sucesso de {target_user.full_name}!')
            
        except User.DoesNotExist:
            messages.error(request, 'Usuário não encontrado.')
        except ValueError:
            messages.error(request, 'Valor inválido.')
        except Exception as e:
            messages.error(request, f'Erro: {str(e)}')
    
    return redirect('manage_cs')


@login_required
def import_cs_excel_view(request):
    """Importar planilha de confianças (com aprovação)"""
    if not request.user.can_manage_cs():
        messages.error(request, 'Você não tem permissão para realizar esta ação.')
        return redirect('dashboard')
    
    if request.method == 'POST' and request.FILES.get('excel_file'):
        try:
            excel_file = request.FILES['excel_file']
            
            # Verificar se é um arquivo Excel
            if not excel_file.name.endswith(('.xlsx', '.xls')):
                messages.error(request, 'Por favor, envie um arquivo Excel (.xlsx ou .xls)')
                return redirect('manage_cs')
            
            # Ler o arquivo Excel
            wb = openpyxl.load_workbook(excel_file)
            ws = wb.active
            
            updates_requested = []
            errors = []
            
            # Pular o cabeçalho (linha 1)
            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    if not any(row):  # Pular linhas vazias
                        continue
                        
                    # Esperado: Nome, Email, Setor, Saldo C$
                    if len(row) < 4:
                        errors.append(f'Linha {row_num}: Dados insuficientes')
                        continue
                        
                    nome, email, setor, novo_saldo = row[:4]
                    
                    if not email or not novo_saldo:
                        errors.append(f'Linha {row_num}: Email ou saldo não informado')
                        continue
                    
                    # Encontrar usuário pelo email
                    try:
                        user = User.objects.get(email=email)
                        novo_saldo = Decimal(str(novo_saldo))
                        diferenca = novo_saldo - user.balance_cs
                        
                        if diferenca != 0:
                            updates_requested.append({
                                'user': user,
                                'saldo_atual': user.balance_cs,
                                'novo_saldo': novo_saldo,
                                'diferenca': diferenca,
                                'linha': row_num
                            })
                            
                    except User.DoesNotExist:
                        errors.append(f'Linha {row_num}: Usuário com email {email} não encontrado')
                    except (ValueError, TypeError):
                        errors.append(f'Linha {row_num}: Saldo inválido ({novo_saldo})')
                        
                except Exception as e:
                    errors.append(f'Linha {row_num}: Erro ao processar ({str(e)})')
            
            if not updates_requested and not errors:
                messages.info(request, 'Nenhuma atualização necessária. Todos os saldos já estão corretos.')
                return redirect('manage_cs')
            
            if errors:
                error_msg = 'Erros encontrados:\n' + '\n'.join(errors[:10])  # Mostrar apenas 10 primeiros erros
                if len(errors) > 10:
                    error_msg += f'\n... e mais {len(errors) - 10} erros'
                messages.error(request, error_msg)
            
            if updates_requested:
                # Criar transações pendentes para as atualizações
                from prizes.models import CSTransaction
                
                for update in updates_requested:
                    user = update['user']
                    diferenca = update['diferenca']
                    
                    transaction_type = 'CREDIT' if diferenca > 0 else 'DEBIT'
                    
                    CSTransaction.objects.create(
                        user=user,
                        amount=abs(diferenca),
                        transaction_type=transaction_type,
                        description=f'Importação de planilha - Linha {update["linha"]} - Saldo: C$ {update["saldo_atual"]} → C$ {update["novo_saldo"]}',
                        status='PENDING',
                        created_by=request.user
                    )
                
                log_action(
                    request.user, 
                    'CS_IMPORT_REQUEST', 
                    f'Solicitada importação de planilha com {len(updates_requested)} atualizações',
                    request
                )
                
                messages.success(request, 
                    f'Planilha processada! {len(updates_requested)} atualizações solicitadas e enviadas para aprovação.')
            
        except Exception as e:
            messages.error(request, f'Erro ao processar planilha: {str(e)}')
    
    return redirect('manage_cs')


@login_required
def manage_sectors_view(request):
    """Gerenciar setores"""
    if not request.user.can_manage_users():
        messages.error(request, 'Você não tem permissão para acessar esta área.')
        return redirect('dashboard')
    
    sectors = Sector.objects.all()
    context = {
        'sectors': sectors,
        'user': request.user,
    }
    return render(request, 'admin/sectors.html', context)


@login_required
def create_sector_view(request):
    """Criar novo setor"""
    if not request.user.can_manage_users():
        messages.error(request, 'Você não tem permissão para acessar esta área.')
        return redirect('dashboard')
    
    if request.method == 'POST':
        name = request.POST.get('name')
        description = request.POST.get('description', '')
        
        try:
            if Sector.objects.filter(name=name).exists():
                messages.error(request, 'Setor com este nome já existe.')
            else:
                sector = Sector.objects.create(
                    name=name,
                    description=description
                )
                
                log_action(
                    request.user, 
                    'SECTOR_CREATE', 
                    f'Setor criado: {sector.name}',
                    request
                )
                
                messages.success(request, f'Setor {sector.name} criado com sucesso!')
                return redirect('manage_sectors')
                
        except Exception as e:
            messages.error(request, f'Erro ao criar setor: {str(e)}')
    
    context = {
        'user': request.user,
    }
    return render(request, 'admin/create_sector.html', context)


@login_required
def edit_sector_view(request, sector_id):
    """Editar setor"""
    if not request.user.can_manage_users():
        messages.error(request, 'Você não tem permissão para acessar esta área.')
        return redirect('dashboard')
    
    sector = get_object_or_404(Sector, id=sector_id)
    
    if request.method == 'POST':
        name = request.POST.get('name')
        description = request.POST.get('description', '')
        
        try:
            if Sector.objects.filter(name=name).exclude(id=sector_id).exists():
                messages.error(request, 'Setor com este nome já existe.')
            else:
                old_name = sector.name
                sector.name = name
                sector.description = description
                sector.save()
                
                log_action(
                    request.user, 
                    'SECTOR_EDIT', 
                    f'Setor editado: {old_name} → {sector.name}',
                    request
                )
                
                messages.success(request, f'Setor {sector.name} atualizado com sucesso!')
                return redirect('manage_sectors')
                
        except Exception as e:
            messages.error(request, f'Erro ao editar setor: {str(e)}')
    
    context = {
        'sector': sector,
        'user': request.user,
    }
    return render(request, 'admin/edit_sector.html', context)


@login_required
def delete_sector_view(request, sector_id):
    """Deletar setor"""
    if not request.user.can_manage_users():
        messages.error(request, 'Você não tem permissão para acessar esta área.')
        return redirect('dashboard')
    
    sector = get_object_or_404(Sector, id=sector_id)
    
    if request.method == 'POST':
        # Verificar se há usuários usando este setor
        users_count = User.objects.filter(sector=sector).count()
        if users_count > 0:
            messages.error(request, f'Não é possível deletar o setor "{sector.name}" pois há {users_count} usuários vinculados a ele.')
            return redirect('manage_sectors')
        
        sector_name = sector.name
        sector.delete()
        
        log_action(
            request.user,
            'SECTOR_DELETE',
            f'Setor deletado: {sector_name}',
            request
        )
        
        messages.success(request, f'Setor "{sector_name}" deletado com sucesso!')
        return redirect('manage_sectors')
    
    # Contar usuários vinculados
    users_count = User.objects.filter(sector=sector).count()
    
    context = {
        'sector': sector,
        'users_count': users_count,
        'user': request.user,
    }
    return render(request, 'admin/delete_sector.html', context)


@login_required
def manage_categories_view(request):
    """Gerenciar categorias"""
    if not request.user.can_manage_users():
        messages.error(request, 'Você não tem permissão para acessar esta área.')
        return redirect('dashboard')
    
    from tickets.models import Category
    categories = Category.objects.all().select_related('sector')
    
    # Filtro por setor
    sector_filter = request.GET.get('sector', '')
    if sector_filter:
        try:
            sector_id = int(sector_filter)
            categories = categories.filter(sector_id=sector_id)
        except (ValueError, TypeError):
            pass
    
    # Busca por nome
    search = request.GET.get('search', '')
    if search:
        categories = categories.filter(name__icontains=search)
    
    context = {
        'categories': categories,
        'sectors': Sector.objects.all(),
        'sector_filter': sector_filter,
        'search': search,
        'user': request.user,
    }
    return render(request, 'admin/categories.html', context)


@login_required
def create_category_view(request):
    """Criar nova categoria"""
    if not request.user.can_manage_users():
        messages.error(request, 'Você não tem permissão para acessar esta área.')
        return redirect('dashboard')
    
    if request.method == 'POST':
        name = request.POST.get('name')
        sector_id = request.POST.get('sector')
        default_description = request.POST.get('default_description', '')
        webhook_url = request.POST.get('webhook_url', '')
        requires_approval = request.POST.get('requires_approval') == 'on'
        default_solution_time_hours = request.POST.get('default_solution_time_hours', 24)
        
        try:
            sector = get_object_or_404(Sector, id=sector_id)
            
            # Validar tempo de solução
            try:
                default_solution_time_hours = int(default_solution_time_hours)
                if default_solution_time_hours <= 0:
                    default_solution_time_hours = 24
            except (ValueError, TypeError):
                default_solution_time_hours = 24
            
            from tickets.models import Category
            if Category.objects.filter(name=name, sector=sector).exists():
                messages.error(request, 'Categoria com este nome já existe neste setor.')
            else:
                category = Category.objects.create(
                    name=name,
                    sector=sector,
                    default_description=default_description,
                    webhook_url=webhook_url,
                    requires_approval=requires_approval,
                    default_solution_time_hours=default_solution_time_hours
                )
                
                log_action(
                    request.user, 
                    'CATEGORY_CREATE', 
                    f'Categoria criada: {category.name} - {category.sector.name}',
                    request
                )
                
                messages.success(request, f'Categoria {category.name} criada com sucesso!')
                return redirect('manage_categories')
                
        except Exception as e:
            messages.error(request, f'Erro ao criar categoria: {str(e)}')
    
    context = {
        'sectors': Sector.objects.all(),
        'user': request.user,
    }
    return render(request, 'admin/create_category.html', context)


@login_required
def edit_category_view(request, category_id):
    """Editar categoria"""
    if not request.user.can_manage_users():
        messages.error(request, 'Você não tem permissão para acessar esta área.')
        return redirect('dashboard')
    
    from tickets.models import Category
    category = get_object_or_404(Category, id=category_id)
    
    if request.method == 'POST':
        name = request.POST.get('name')
        sector_id = request.POST.get('sector')
        default_description = request.POST.get('default_description', '')
        webhook_url = request.POST.get('webhook_url', '')
        requires_approval = request.POST.get('requires_approval') == 'on'
        is_active = request.POST.get('is_active') == 'on'
        default_solution_time_hours = request.POST.get('default_solution_time_hours', 24)
        
        try:
            sector = get_object_or_404(Sector, id=sector_id)
            
            # Validar tempo de solução
            try:
                default_solution_time_hours = int(default_solution_time_hours)
                if default_solution_time_hours <= 0:
                    default_solution_time_hours = 24
            except (ValueError, TypeError):
                default_solution_time_hours = 24
            
            # Verificar se já existe uma categoria com o mesmo nome no setor (excluindo a atual)
            if Category.objects.filter(name=name, sector=sector).exclude(id=category_id).exists():
                messages.error(request, 'Categoria com este nome já existe neste setor.')
            else:
                category.name = name
                category.sector = sector
                category.default_description = default_description
                category.webhook_url = webhook_url
                category.requires_approval = requires_approval
                category.is_active = is_active
                category.default_solution_time_hours = default_solution_time_hours
                category.save()
                
                log_action(
                    request.user, 
                    'CATEGORY_UPDATE', 
                    f'Categoria atualizada: {category.name} - {category.sector.name}',
                    request
                )
                
                messages.success(request, f'Categoria {category.name} atualizada com sucesso!')
                return redirect('manage_categories')
                
        except Exception as e:
            messages.error(request, f'Erro ao atualizar categoria: {str(e)}')
    
    context = {
        'category': category,
        'sectors': Sector.objects.all(),
        'user': request.user,
    }
    return render(request, 'admin/edit_category.html', context)


@login_required
def delete_category_view(request, category_id):
    """Deletar categoria"""
    if not request.user.can_manage_users():
        messages.error(request, 'Você não tem permissão para acessar esta área.')
        return redirect('dashboard')
    
    from tickets.models import Category
    category = get_object_or_404(Category, id=category_id)
    
    if request.method == 'POST':
        # Verificar se há chamados associados a esta categoria
        if category.ticket_set.exists():
            messages.error(request, 'Não é possível excluir esta categoria pois existem chamados associados a ela.')
            return redirect('manage_categories')
        
        category_name = category.name
        category.delete()
        
        log_action(
            request.user, 
            'CATEGORY_DELETE', 
            f'Categoria excluída: {category_name}',
            request
        )
        
        messages.success(request, f'Categoria {category_name} excluída com sucesso!')
        return redirect('manage_categories')
    
    context = {
        'category': category,
        'user': request.user,
    }
    return render(request, 'admin/delete_category.html', context)


class UserViewSet(viewsets.ModelViewSet):
    queryset = User.objects.all()
    serializer_class = UserSerializer
    permission_classes = [IsAuthenticated]
    
    def get_queryset(self):
        user = self.request.user
        if user.hierarchy == 'SUPERADMIN':
            return User.objects.all()
        elif user.hierarchy in ['ADMINISTRATIVO', 'SUPERVISOR']:
            return User.objects.filter(sector=user.sector)
        else:
            return User.objects.filter(id=user.id)
    
    @action(detail=True, methods=['post'])
    def update_balance(self, request, pk=None):
        user = self.get_object()
        target_user = request.user
        
        if not target_user.can_manage_cs():
            return Response(
                {'error': 'Você não tem permissão para alterar saldo C$'}, 
                status=status.HTTP_403_FORBIDDEN
            )
        
        amount = request.data.get('amount')
        operation = request.data.get('operation')  # 'add' or 'subtract'
        description = request.data.get('description', '')
        
        if not amount or not operation:
            return Response(
                {'error': 'Campos obrigatórios: amount, operation'}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            amount = float(amount)
            if operation == 'add':
                user.balance_cs += amount
            elif operation == 'subtract':
                user.balance_cs -= amount
            else:
                return Response(
                    {'error': 'Operação inválida'}, 
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Para débitos e ajustes, ainda aplicamos diretamente
            # Para créditos, deixamos pendente de aprovação
            transaction_status = 'APPROVED'  # Padrão para débitos e ajustes
            
            if operation == 'add':
                # Créditos ficam pendentes de aprovação - reverter a mudança no saldo
                user.balance_cs -= amount  # Desfaz a adição
                transaction_status = 'PENDING'
            
            user.save()
            
            # Registrar transação
            from prizes.models import CSTransaction
            transaction = CSTransaction.objects.create(
                user=user,
                amount=amount if operation == 'add' else amount,
                transaction_type='CREDIT' if operation == 'add' else 'DEBIT',
                description=description or f'Ajuste manual - {operation}',
                status=transaction_status,
                created_by=target_user
            )
            
            if operation == 'add':
                message = f'Solicitação de C$ {amount} para {user.full_name} enviada para aprovação'
                action = 'CS_ADD_REQUEST'
            else:
                message = f'C$ {amount} removido de {user.full_name}'
                action = 'CS_DEBIT'
            
            log_action(
                target_user, 
                action, 
                message,
                request
            )
            
            return Response({'message': 'Operação realizada com sucesso' if operation != 'add' else 'Solicitação enviada para aprovação'})
            
        except ValueError:
            return Response(
                {'error': 'Valor inválido'}, 
                status=status.HTTP_400_BAD_REQUEST
            )


class SectorViewSet(viewsets.ModelViewSet):
    queryset = Sector.objects.all()
    serializer_class = SectorSerializer
    permission_classes = [IsAuthenticated]


# ViewSets públicos (sem autenticação) para produção
class PublicUserViewSet(viewsets.ReadOnlyModelViewSet):
    """ViewSet público apenas para leitura de usuários"""
    queryset = User.objects.filter(is_active=True)
    serializer_class = UserSerializer
    permission_classes = []  # Sem autenticação
    
    def get_queryset(self):
        """Limitar campos expostos publicamente"""
        return User.objects.filter(is_active=True).values(
            'id', 'username', 'first_name', 'last_name', 'email', 'hierarchy'
        )


class PublicSectorViewSet(viewsets.ReadOnlyModelViewSet):
    """ViewSet público apenas para leitura de setores"""
    queryset = Sector.objects.all()
    serializer_class = SectorSerializer
    permission_classes = []  # Sem autenticação


@login_required
def manage_webhooks_view(request):
    """Gerenciar webhooks"""
    if not request.user.can_manage_users():
        messages.error(request, 'Você não tem permissão para acessar esta área.')
        return redirect('dashboard')
    
    from tickets.models import Webhook
    webhooks = Webhook.objects.all()
    
    context = {
        'webhooks': webhooks,
        'user': request.user,
    }
    return render(request, 'admin/webhooks.html', context)


def _normalize_commission_user_name(value):
    import unicodedata

    raw = str(value or '').strip().upper()
    raw = unicodedata.normalize('NFD', raw)
    without_accents = ''.join(ch for ch in raw if unicodedata.category(ch) != 'Mn')
    return ' '.join(without_accents.split())


def _snapshot_commission_users_for_reference(version_obj, updated_by):
    """Salva snapshot completo dos usuários por referência (mês/ano)."""
    from users.commission_views import fetch_all_users_from_sheet, SHEET_CN, SHEET_GERENTE, SHEET_COORDENADOR
    from users.models import CommissionUserReferenceHistory

    sheet_role_map = {
        SHEET_CN: 'cn',
        SHEET_GERENTE: 'gerente',
        SHEET_COORDENADOR: 'coordenador',
    }

    active_users = User.objects.filter(is_active=True)
    user_name_index = {}
    for db_user in active_users:
        normalized = _normalize_commission_user_name(db_user.get_full_name())
        if normalized and normalized not in user_name_index:
            user_name_index[normalized] = db_user.id

    history_rows = []
    for sheet_name, role in sheet_role_map.items():
        sheet_result = fetch_all_users_from_sheet(sheet_name, excel_url=version_obj.excel_comissao_url)
        if not sheet_result.get('success'):
            error_message = sheet_result.get('error') or f'Falha ao ler a sheet {sheet_name}'
            raise ValueError(error_message)

        for entry in sheet_result.get('users', []):
            user_name = str(entry.get('nome') or '').strip()
            if not user_name:
                continue
            normalized_name = _normalize_commission_user_name(user_name)
            history_rows.append(
                CommissionUserReferenceHistory(
                    year=version_obj.year,
                    month=version_obj.month,
                    user_name=user_name,
                    user_id=user_name_index.get(normalized_name),
                    role=role,
                    sheet_name=sheet_name,
                    source_version=version_obj,
                    row_data=entry.get('data') or {},
                    updated_by=updated_by,
                )
            )

    CommissionUserReferenceHistory.objects.filter(
        year=version_obj.year,
        month=version_obj.month,
    ).delete()

    if history_rows:
        CommissionUserReferenceHistory.objects.bulk_create(history_rows, batch_size=500)

    return len(history_rows)


@login_required
def system_config_view(request):
    """Gerenciar configurações do sistema (links das planilhas Excel)"""
    from users.models import SystemConfig, CommissionSpreadsheetVersion
    
    # Apenas superadmin pode acessar
    if request.user.hierarchy != 'SUPERADMIN':
        messages.error(request, 'Apenas Superadmin pode acessar as configurações do sistema.')
        return redirect('dashboard')
    
    config = SystemConfig.get_config()
    display_ref_year, display_ref_month = config.get_display_reference_month_year(base_date=timezone.now())

    active_version = CommissionSpreadsheetVersion.objects.filter(
        year=display_ref_year,
        month=display_ref_month,
    ).first()
    commission_versions = CommissionSpreadsheetVersion.objects.select_related('updated_by').all()

    available_references = list(
        CommissionSpreadsheetVersion.objects.values('year', 'month').order_by('-year', '-month')
    )
    available_year_choices = sorted({item['year'] for item in available_references}, reverse=True)
    available_month_choices = sorted({item['month'] for item in available_references})
    
    if request.method == 'POST':
        selected_month = request.POST.get('version_month', '').strip()
        selected_year = request.POST.get('version_year', '').strip()
        display_month = request.POST.get('display_reference_month', '').strip()
        display_year = request.POST.get('display_reference_year', '').strip()

        try:
            selected_month_int = int(selected_month)
            selected_year_int = int(selected_year)
        except (TypeError, ValueError):
            messages.error(request, 'Informe mês e ano válidos para a versão do comissionamento.')
            return redirect('system_config')

        if selected_month_int < 1 or selected_month_int > 12:
            messages.error(request, 'Mês inválido. Use um valor entre 1 e 12.')
            return redirect('system_config')

        if (display_month and not display_year) or (display_year and not display_month):
            messages.error(request, 'Informe mês e ano de exibição juntos.')
            return redirect('system_config')

        if display_month and display_year:
            try:
                display_month_int = int(display_month)
                display_year_int = int(display_year)
            except (TypeError, ValueError):
                messages.error(request, 'Mês e ano de exibição inválidos.')
                return redirect('system_config')
        else:
            display_month_int = selected_month_int
            display_year_int = selected_year_int

        if display_month_int < 1 or display_month_int > 12:
            messages.error(request, 'Mês de exibição inválido. Use um valor entre 1 e 12.')
            return redirect('system_config')

        display_matches_selected_version = (
            display_year_int == selected_year_int and display_month_int == selected_month_int
        )
        display_version_exists = CommissionSpreadsheetVersion.objects.filter(
            year=display_year_int,
            month=display_month_int,
        ).exists()
        if not (display_matches_selected_version or display_version_exists):
            messages.error(request, 'Selecione mês e ano de exibição já disponíveis no histórico de versões.')
            return redirect('system_config')

        commission_urls = {
            'excel_comissao_url': request.POST.get('excel_comissao_url', '').strip(),
            'excel_vendas_url': request.POST.get('excel_vendas_url', '').strip(),
            'excel_base_pagamento_url': request.POST.get('excel_base_pagamento_url', '').strip(),
            'excel_base_exclusao_url': request.POST.get('excel_base_exclusao_url', '').strip(),
        }

        if not all(commission_urls.values()):
            messages.error(request, 'Preencha todas as URLs de comissionamento para salvar a versão.')
            return redirect('system_config')

        try:
            with transaction.atomic():
                # Contestation phase selection (antes/pos)
                selected_phase = request.POST.get('contestacao_phase', 'pos').strip()
                if selected_phase not in ('antes', 'pos'):
                    selected_phase = 'pos'

                version_obj, created = CommissionSpreadsheetVersion.objects.update_or_create(
                    year=selected_year_int,
                    month=selected_month_int,
                    contestacao_phase=selected_phase,
                    defaults={
                        **commission_urls,
                        'updated_by': request.user,
                        # Toda alteração volta para rascunho: só vai ao ar quando
                        # liberada explicitamente em "Liberar Comissionamento".
                        'status': CommissionSpreadsheetVersion.STATUS_DRAFT,
                        'released_by': None,
                        'released_at': None,
                    }
                )

                # Mantém estes campos como fallback para instalações sem versão cadastrada.
                config.excel_comissao_url = commission_urls['excel_comissao_url']
                config.excel_vendas_url = commission_urls['excel_vendas_url']
                config.excel_base_pagamento_url = commission_urls['excel_base_pagamento_url']
                config.excel_base_exclusao_url = commission_urls['excel_base_exclusao_url']
                config.excel_contestacao_base_exclusao_url = request.POST.get('excel_contestacao_base_exclusao_url', '').strip()
                config.display_reference_month = display_month_int
                config.display_reference_year = display_year_int
                config.updated_by = request.user
                config.save()

                snapshot_count = _snapshot_commission_users_for_reference(version_obj, request.user)
        except ValueError as exc:
            messages.error(request, f'Erro ao salvar histórico completo dos usuários: {exc}')
            return redirect('system_config')
        
        # Limpar cache das planilhas para forçar reload
        from django.core.cache import cache
        cache.delete_many([
            'comissao_REMUNERAÇÃO CN_file_content',
            'comissao_REMUNERAÇÃO GERENTE_file_content', 
            'base_pagamento_file_content',
            'base_exclusao_file_content',
            'contestacao_base_exclusao_content',
            'vendas_file_content',
        ])
        
        operation = 'atualizada' if not created else 'criada'
        messages.success(
            request,
            (
                f'Versão {selected_month_int:02d}/{selected_year_int} {operation} como RASCUNHO. '
                f'Padrão de exibição definido para {display_month_int:02d}/{display_year_int}. '
                f'Histórico salvo para {snapshot_count} usuários. O cache foi limpo. '
                f'A versão só ficará visível para os usuários após clicar em "Liberar Comissionamento".'
            )
        )
        return redirect('system_config')

    if active_version:
        commission_form = {
            'excel_comissao_url': active_version.excel_comissao_url,
            'excel_vendas_url': active_version.excel_vendas_url,
            'excel_base_pagamento_url': active_version.excel_base_pagamento_url,
            'excel_base_exclusao_url': active_version.excel_base_exclusao_url,
            'version_month': active_version.month,
            'version_year': active_version.year,
            'version_phase': getattr(active_version, 'contestacao_phase', 'pos'),
            'display_reference_month': config.display_reference_month or active_version.month,
            'display_reference_year': config.display_reference_year or active_version.year,
        }
    else:
        commission_form = {
            'excel_comissao_url': config.excel_comissao_url,
            'excel_vendas_url': config.excel_vendas_url,
            'excel_base_pagamento_url': config.excel_base_pagamento_url,
            'excel_base_exclusao_url': config.excel_base_exclusao_url,
            'version_month': display_ref_month,
            'version_year': display_ref_year,
            'version_phase': 'pos',
            'display_reference_month': config.display_reference_month or display_ref_month,
            'display_reference_year': config.display_reference_year or display_ref_year,
        }

    meses_map = {
        1: 'Janeiro', 2: 'Fevereiro', 3: 'Março', 4: 'Abril',
        5: 'Maio', 6: 'Junho', 7: 'Julho', 8: 'Agosto',
        9: 'Setembro', 10: 'Outubro', 11: 'Novembro', 12: 'Dezembro',
    }

    display_month_choices = [
        (month_value, meses_map[month_value])
        for month_value in available_month_choices
        if month_value in meses_map
    ]
    if not display_month_choices:
        display_month_choices = [(display_ref_month, meses_map.get(display_ref_month, str(display_ref_month)))]

    if not available_year_choices:
        available_year_choices = [display_ref_year]
    
    from users.models import CommissionMonthlyTotal
    synced_totals = list(
        CommissionMonthlyTotal.objects.filter(
            year=display_ref_year, month=display_ref_month
        ).order_by('person_name')
    )
    synced_last = CommissionMonthlyTotal.objects.filter(
        year=display_ref_year, month=display_ref_month
    ).order_by('-synced_at').first()

    context = {
        'config': config,
        'commission_form': commission_form,
        'commission_versions': commission_versions,
        'active_version': active_version,
        'reference_month': display_ref_month,
        'reference_year': display_ref_year,
        'synced_totals': synced_totals,
        'synced_totals_count': len(synced_totals),
        'synced_last': synced_last,
        'month_choices': [
            (1, 'Janeiro'), (2, 'Fevereiro'), (3, 'Março'), (4, 'Abril'),
            (5, 'Maio'), (6, 'Junho'), (7, 'Julho'), (8, 'Agosto'),
            (9, 'Setembro'), (10, 'Outubro'), (11, 'Novembro'), (12, 'Dezembro'),
        ],
        'year_choices': list(range(display_ref_year - 3, display_ref_year + 3)),
        'display_month_choices': display_month_choices,
        'display_year_choices': available_year_choices,
        'user': request.user,
        'draft_versions': commission_versions.filter(status=CommissionSpreadsheetVersion.STATUS_DRAFT),
    }
    return render(request, 'admin/system_config.html', context)


@login_required
@require_POST
def release_commission_view(request):
    """Libera (publica) uma versão de comissionamento que está em rascunho.

    Apenas superadmins têm acesso. Enquanto a versão estiver em rascunho ela
    fica visível somente nesta tela; ao ser liberada passa a valer para todos
    os usuários."""
    from users.models import CommissionSpreadsheetVersion

    if request.user.hierarchy != 'SUPERADMIN':
        messages.error(request, 'Apenas Superadmin pode liberar o comissionamento.')
        return redirect('dashboard')

    version_id = request.POST.get('version_id')
    version = CommissionSpreadsheetVersion.objects.filter(pk=version_id).first()
    if not version:
        messages.error(request, 'Versão de comissionamento não encontrada.')
        return redirect('system_config')

    if version.status == CommissionSpreadsheetVersion.STATUS_RELEASED:
        messages.info(request, f'A versão {version.month:02d}/{version.year} já está liberada.')
        return redirect('system_config')

    version.status = CommissionSpreadsheetVersion.STATUS_RELEASED
    version.released_by = request.user
    version.released_at = timezone.now()
    version.save(update_fields=['status', 'released_by', 'released_at'])

    # Limpa o cache para que a versão liberada passe a valer imediatamente.
    from django.core.cache import cache
    cache.delete_many([
        'comissao_REMUNERAÇÃO CN_file_content',
        'comissao_REMUNERAÇÃO GERENTE_file_content',
        'base_pagamento_file_content',
        'base_exclusao_file_content',
        'contestacao_base_exclusao_content',
        'vendas_file_content',
    ])

    messages.success(
        request,
        f'Comissionamento {version.month:02d}/{version.year} liberado com sucesso. '
        f'Agora está visível para todos os usuários.'
    )
    return redirect('system_config')


def _extract_remuneracao_final(role, row_data):
    """Extrai o valor de remuneração final de uma linha (dict) da planilha."""
    row_data = row_data or {}

    def _to_decimal(value):
        if value is None or value == '':
            return None
        try:
            return Decimal(str(value).replace('R$', '').replace(' ', '').replace(',', '.'))
        except Exception:
            return None

    # Chaves conhecidas por papel
    candidate_keys = [
        'REMUNERACAO_FINAL_TOTAL', 'REMUNERACAO_FINAL', 'REMUNERAÇÃO_FINAL',
        ' REMUNERAÇÃO FINAL COO/SNP', 'REMUNERAÇÃO FINAL COO/SNP',
        'REMUNERAÇÃO FINAL', 'REMUNERACAO FINAL',
    ]
    for key in candidate_keys:
        if key in row_data:
            val = _to_decimal(row_data.get(key))
            if val is not None:
                return val

    # Fallback: qualquer coluna que contenha REMUNERA + FINAL
    for col, value in row_data.items():
        col_up = str(col).upper().replace('Ç', 'C').replace('Ã', 'A')
        if 'REMUNERA' in col_up and 'FINAL' in col_up:
            val = _to_decimal(value)
            if val is not None:
                return val

    return Decimal('0.00')


@login_required
@require_POST
def sync_commission_table_view(request):
    """Sincroniza a tabela de totais de comissionamento (nome, mês, ano, valor).

    Lê as planilhas de comissionamento do mês de referência e grava o valor total
    que cada pessoa receberá em CommissionMonthlyTotal. Para usuários do grupo
    'A PARTE', grava o valor calculado pelo modelo de comissionamento à parte.
    """
    from users.models import (
        SystemConfig, CommissionSpreadsheetVersion, CommissionMonthlyTotal,
        AParteCommissionConfig,
    )
    from users.commission_views import (
        fetch_all_users_from_sheet, get_excel_urls,
        SHEET_CN, SHEET_GERENTE, SHEET_COORDENADOR,
    )

    if request.user.hierarchy != 'SUPERADMIN':
        messages.error(request, 'Apenas Superadmin pode sincronizar a tabela.')
        return redirect('dashboard')

    config = SystemConfig.get_config()
    default_year, default_month = config.get_display_reference_month_year(base_date=timezone.now())
    try:
        ref_month = int(request.POST.get('month') or default_month)
        ref_year = int(request.POST.get('year') or default_year)
    except (TypeError, ValueError):
        ref_month, ref_year = default_month, default_year

    urls = get_excel_urls(ref_year, ref_month)
    version = CommissionSpreadsheetVersion.objects.filter(year=ref_year, month=ref_month).first()

    sheet_role_map = {
        SHEET_CN: 'cn',
        SHEET_GERENTE: 'gerente',
        SHEET_COORDENADOR: 'coordenador',
    }

    # Usuários "A parte": calculados pelo modelo próprio (não pela sheet).
    aparte_names = set()
    rows_by_key = {}  # (person_name_norm, role) -> CommissionMonthlyTotal

    try:
        from simulator.services import (
            get_all_aparte_users, get_network_metas_from_power_bi, compute_aparte_commission,
            get_aparte_factors,
        )
        from simulator.sql_realizado import get_network_realized_sales_from_mysql

        aparte_users = get_all_aparte_users()
        if aparte_users:
            metas_rede = get_network_metas_from_power_bi(ref_year, ref_month)
            realizado_rede = get_network_realized_sales_from_mysql(year=ref_year, month=ref_month)
            for au in aparte_users:
                cfg = AParteCommissionConfig.objects.filter(user=au).first()
                factors = get_aparte_factors(cfg)
                base = cfg.base_salary if cfg else 0
                calc = compute_aparte_commission(base, factors, metas_rede, realizado_rede)
                name = au.get_full_name() or au.first_name or au.email
                norm = _normalize_commission_user_name(name)
                aparte_names.add(norm)
                rows_by_key[(norm, 'aparte')] = CommissionMonthlyTotal(
                    year=ref_year, month=ref_month, person_name=name, role='aparte',
                    total_commission=Decimal(str(round(calc['total'], 2))),
                    source_version=version, synced_by=request.user,
                )
    except Exception as exc:
        messages.warning(request, f'Aviso ao calcular comissionamento "A parte": {exc}')

    erros = []
    for sheet_name, role in sheet_role_map.items():
        sheet_result = fetch_all_users_from_sheet(sheet_name, excel_url=urls['comissao'])
        if not sheet_result.get('success'):
            erros.append(sheet_result.get('error') or f'Falha ao ler {sheet_name}')
            continue
        for entry in sheet_result.get('users', []):
            name = str(entry.get('nome') or '').strip()
            if not name:
                continue
            norm = _normalize_commission_user_name(name)
            if norm in aparte_names:
                continue  # já contabilizado pelo modelo "A parte"
            total = _extract_remuneracao_final(role, entry.get('data') or {})
            rows_by_key[(norm, role)] = CommissionMonthlyTotal(
                year=ref_year, month=ref_month, person_name=name, role=role,
                total_commission=total, source_version=version, synced_by=request.user,
            )

    rows = list(rows_by_key.values())
    if not rows and erros:
        messages.error(request, 'Não foi possível sincronizar: ' + '; '.join(erros))
        return redirect('system_config')

    with transaction.atomic():
        CommissionMonthlyTotal.objects.filter(year=ref_year, month=ref_month).delete()
        if rows:
            CommissionMonthlyTotal.objects.bulk_create(rows, batch_size=500)

    msg = f'Tabela sincronizada: {len(rows)} pessoas em {ref_month:02d}/{ref_year}.'
    if erros:
        msg += ' Avisos: ' + '; '.join(erros)
    messages.success(request, msg)
    return redirect('system_config')


def _aparte_pillar_labels():
    return [
        ('movel', 'Móvel'), ('fixa', 'Fixa'), ('smartphones', 'Smartphones'),
        ('eletronicos', 'Eletrônicos'), ('essenciais', 'Essenciais'),
        ('seguros', 'Seguros'), ('sva', 'SVA'),
    ]


@login_required
def aparte_config_view(request):
    """Mini-sistema de comissionamento "A parte" (superadmin).

    Permite escolher um usuário do grupo 'A PARTE', definir o salário base e os
    fatores de cada pilar, e visualizar a comissão calculada por mês (usando a
    meta da rede de cada mês).
    """
    from users.models import AParteCommissionConfig, APARTE_DEFAULT_FACTORS
    from simulator.services import (
        get_all_aparte_users, get_network_metas_from_power_bi, compute_aparte_commission,
        get_aparte_factors, APARTE_PILLAR_ORDER,
    )

    if request.user.hierarchy != 'SUPERADMIN':
        messages.error(request, 'Apenas Superadmin pode acessar esta área.')
        return redirect('dashboard')

    aparte_users = get_all_aparte_users()
    selected_user = None
    selected_user_id = request.GET.get('user_id') or request.POST.get('user_id')
    if selected_user_id:
        selected_user = User.objects.filter(id=selected_user_id, is_active=True).first()

    config = None
    if selected_user:
        config, _ = AParteCommissionConfig.objects.get_or_create(user=selected_user)

    if request.method == 'POST' and selected_user and config:
        # Salário base
        try:
            base_raw = (request.POST.get('base_salary') or '0').replace('R$', '').replace(' ', '').replace('.', '').replace(',', '.')
            config.base_salary = Decimal(base_raw or '0')
        except Exception:
            config.base_salary = config.base_salary or Decimal('0.00')

        # Fatores: factor__<pilar>__<row>__<col>
        factors = get_aparte_factors(config)
        for pilar in APARTE_PILLAR_ORDER:
            table = [list(r) for r in factors.get(pilar, [])]
            for r_idx, frow in enumerate(table):
                for c_idx in range(len(frow)):
                    field = f"factor__{pilar}__{r_idx}__{c_idx}"
                    if field in request.POST:
                        raw = (request.POST.get(field) or '').replace(',', '.').strip()
                        try:
                            table[r_idx][c_idx] = float(raw) if raw != '' else 0.0
                        except ValueError:
                            pass
            factors[pilar] = table
        config.factors = factors
        config.is_active = request.POST.get('is_active') == 'on'
        config.updated_by = request.user
        config.save()
        messages.success(request, f'Configuração de {selected_user.get_full_name()} salva.')
        return redirect(f"{request.path}?user_id={selected_user.id}")

    # Montar dados para o template
    factor_tables = []
    monthly_preview = []
    if selected_user and config:
        factors = get_aparte_factors(config)
        for pilar, label in _aparte_pillar_labels():
            factor_tables.append({
                'key': pilar,
                'label': label,
                'rows': factors.get(pilar, []),
            })

        # Preview mensal: meses com metas no Power BI (+ mês atual).
        try:
            from power_bi.models import GoalUpload
            from simulator.sql_realizado import get_network_realized_sales_from_mysql
            uploads = list(GoalUpload.objects.order_by('-year', '-month')[:12])
            meses_pt = {
                1: 'Jan', 2: 'Fev', 3: 'Mar', 4: 'Abr', 5: 'Mai', 6: 'Jun',
                7: 'Jul', 8: 'Ago', 9: 'Set', 10: 'Out', 11: 'Nov', 12: 'Dez',
            }
            for up in uploads:
                metas_rede = get_network_metas_from_power_bi(up.year, up.month)
                realizado_rede = get_network_realized_sales_from_mysql(year=up.year, month=up.month)
                calc = compute_aparte_commission(config.base_salary, factors, metas_rede, realizado_rede)
                monthly_preview.append({
                    'label': f"{meses_pt.get(up.month, up.month)}/{up.year}",
                    'total': calc['total'],
                    'rows': calc['rows'],
                })
        except Exception:
            pass

    context = {
        'user': request.user,
        'aparte_users': aparte_users,
        'selected_user': selected_user,
        'config': config,
        'factor_tables': factor_tables,
        'monthly_preview': monthly_preview,
        'has_group': bool(aparte_users) or True,
    }
    return render(request, 'users/aparte_config.html', context)


@login_required
def create_webhook_view(request):
    """Criar novo webhook"""
    if not request.user.can_manage_users():
        messages.error(request, 'Você não tem permissão para acessar esta área.')
        return redirect('dashboard')
    
    if request.method == 'POST':
        from tickets.models import Webhook
        
        name = request.POST.get('name')
        url = request.POST.get('url')
        events = request.POST.getlist('events')
        is_active = request.POST.get('is_active') == 'on'
        
        try:
            webhook = Webhook.objects.create(
                name=name,
                url=url,
                events=events,
                is_active=is_active
            )
            
            log_action(request.user, 'WEBHOOK_CREATE', f'Webhook criado: {webhook.name}', request)
            messages.success(request, f'Webhook "{webhook.name}" criado com sucesso!')
            return redirect('manage_webhooks')
        except Exception as e:
            messages.error(request, f'Erro ao criar webhook: {str(e)}')
    
    context = {
        'user': request.user,
        'event_choices': [
            ('ticket_created', 'Chamado Criado'),
            ('ticket_updated', 'Chamado Atualizado'),
            ('ticket_closed', 'Chamado Fechado'),
            ('communication_sent', 'Comunicado Enviado'),
            ('user_created', 'Usuário Criado'),
        ]
    }
    return render(request, 'admin/create_webhook.html', context)


@login_required
def profile_view(request):
    """Visualizar perfil do usuário"""
    from tickets.models import Ticket
    
    # Contar chamados abertos do usuário
    user_tickets_count = Ticket.objects.filter(
        created_by=request.user,
        status__in=['OPEN', 'IN_PROGRESS', 'PENDING']
    ).count()
    
    context = {
        'user': request.user,
        'user_tickets_count': user_tickets_count,
    }
    return render(request, 'users/profile.html', context)


@login_required
def update_profile_view(request):
    """Atualizar perfil do usuário"""
    if request.method == 'POST':
        user = request.user
        user.first_name = request.POST.get('first_name', user.first_name)
        user.last_name = request.POST.get('last_name', user.last_name)
        user.email = request.POST.get('email', user.email)
        user.phone = request.POST.get('phone', user.phone)
        user.disc_profile = request.POST.get('disc_profile', user.disc_profile)
        user.uniform_size_shirt = request.POST.get('uniform_size_shirt', user.uniform_size_shirt)
        user.uniform_size_pants = request.POST.get('uniform_size_pants', user.uniform_size_pants)
        
        # Upload de foto de perfil
        if request.FILES.get('profile_picture'):
            user.profile_picture = request.FILES['profile_picture']
        
        user.save()
        
        messages.success(request, 'Perfil atualizado com sucesso!')
        return redirect('profile')

    return redirect('profile')


def _validate_profile_photo(photo):
    """Valida a imagem de perfil. Retorna a mensagem de erro, ou None se estiver ok."""
    if not photo:
        return 'Nenhuma imagem foi enviada.'

    content_type = (getattr(photo, 'content_type', '') or '').lower()
    if not content_type.startswith('image/'):
        return 'O arquivo enviado não é uma imagem válida.'

    # SVG é "image/*" mas pode conter script (XSS armazenado ao servir a mídia),
    # e não é um formato de foto de perfil válido — rejeita explicitamente.
    name = (getattr(photo, 'name', '') or '').lower()
    if content_type == 'image/svg+xml' or name.endswith('.svg'):
        return 'Imagens SVG não são permitidas. Envie JPG, PNG ou WEBP.'

    # Limite de 8 MB para evitar uploads muito grandes.
    if photo.size > 8 * 1024 * 1024:
        return 'A imagem deve ter no máximo 8 MB.'

    # Não confia só no content_type (enviado pelo cliente): confirma que o
    # arquivo é uma imagem raster real. Isso barra SVG/HTML disfarçados.
    try:
        from PIL import Image
        photo.seek(0)
        Image.open(photo).verify()
    except Exception:
        return 'Arquivo de imagem inválido ou corrompido.'
    finally:
        try:
            photo.seek(0)
        except Exception:
            pass

    return None


@login_required
@require_POST
def upload_profile_photo_ajax(request):
    """Upload rápido da foto de perfil via AJAX (usado pelo popup da Home)."""
    photo = request.FILES.get('profile_picture')
    error = _validate_profile_photo(photo)
    if error:
        return JsonResponse({'success': False, 'error': error}, status=400)

    user = request.user
    user.profile_picture = photo
    user.save()

    return JsonResponse({'success': True, 'photo_url': user.profile_picture.url})


@login_required
@require_POST
def admin_upload_user_photo(request, user_id):
    """Permite que um gestor troque a foto de perfil de outro usuário."""
    if not request.user.can_manage_users():
        return JsonResponse(
            {'success': False, 'error': 'Você não tem permissão para alterar a foto de outros usuários.'},
            status=403,
        )

    target = get_object_or_404(User, pk=user_id)

    photo = request.FILES.get('profile_picture')
    error = _validate_profile_photo(photo)
    if error:
        return JsonResponse({'success': False, 'error': error}, status=400)

    target.profile_picture = photo
    target.save(update_fields=['profile_picture'])

    return JsonResponse({
        'success': True,
        'photo_url': target.profile_picture.url,
        'user_id': target.pk,
    })


@login_required
def settings_view(request):
    """Visualizar configurações do usuário"""
    context = {
        'user': request.user,
    }
    return render(request, 'users/settings.html', context)


@login_required
def update_settings_view(request):
    """Atualizar configurações do usuário"""
    if request.method == 'POST':
        # Implementar atualização de configurações
        messages.success(request, 'Configurações atualizadas com sucesso!')
        return redirect('settings')
    
    return redirect('settings')


@login_required
def help_view(request):
    """Visualizar central de ajuda e tutoriais"""
    from core.models import Tutorial, TrainingCategory, TutorialProgress
    from django.core.paginator import Paginator
    
    # Filtros
    category_filter = request.GET.get('category', '')
    
    # Buscar categoria selecionada se existe
    current_category = None
    if category_filter:
        try:
            current_category = TrainingCategory.objects.get(id=category_filter, is_active=True)
        except TrainingCategory.DoesNotExist:
            pass
    
    # Buscar tutoriais ativos
    tutorials = Tutorial.objects.filter(is_active=True)
    
    if category_filter and current_category:
        tutorials = tutorials.filter(category=current_category)
    
    tutorials = tutorials.select_related('created_by', 'category').order_by('order', 'title')
    
    # Paginação
    paginator = Paginator(tutorials, 12)  # 12 tutoriais por página
    page = request.GET.get('page')
    tutorials = paginator.get_page(page)
    
    # Buscar progresso do usuário para todos os tutoriais (não só da página atual)
    all_tutorials = Tutorial.objects.filter(is_active=True)
    tutorial_progress = {}
    
    if all_tutorials:
        user_progress = TutorialProgress.objects.filter(
            tutorial__in=all_tutorials,
            user=request.user
        ).select_related('tutorial')
        
        for progress in user_progress:
            tutorial_progress[progress.tutorial.id] = progress
    
    # Calcular estatísticas do usuário
    total_tutorials = all_tutorials.count()
    completed_tutorials = sum(1 for p in tutorial_progress.values() if p.completed_at)
    viewed_tutorials = sum(1 for p in tutorial_progress.values() if p.viewed_at and not p.completed_at)
    completion_rate = int((completed_tutorials / total_tutorials * 100)) if total_tutorials > 0 else 0
    
    user_stats = {
        'total_tutorials': total_tutorials,
        'completed_tutorials': completed_tutorials,
        'viewed_tutorials': viewed_tutorials,
        'completion_rate': completion_rate,
    }
    
    # Buscar todas as categorias para filtros
    categories = TrainingCategory.objects.filter(
        is_active=True
    ).prefetch_related('tutorial_set').order_by('name')
    
    context = {
        'tutorials': tutorials,
        'categories': categories,
        'current_category': current_category,
        'selected_category': category_filter,
        'user_stats': user_stats,
        'tutorial_progress': tutorial_progress,
    }
    return render(request, 'help/tutorials.html', context)


@login_required
def tutorial_detail_view(request, tutorial_id):
    """Visualizar tutorial específico"""
    from core.models import Tutorial, TutorialProgress
    from core.middleware import log_action
    
    tutorial = get_object_or_404(Tutorial, id=tutorial_id, is_active=True)
    
    # Buscar ou criar progresso do usuário
    progress, created = TutorialProgress.objects.get_or_create(
        tutorial=tutorial,
        user=request.user
    )
    
    # Marcar como visualizado
    progress.mark_as_viewed()
    
    # Se for POST, marcar como concluído
    if request.method == 'POST' and request.POST.get('action') == 'complete':
        progress.mark_as_completed()
        
        log_action(
            request.user,
            'TUTORIAL_COMPLETE',
            f'Tutorial concluído: {tutorial.title}',
            request
        )
        
        messages.success(request, 'Tutorial marcado como concluído!')
        return redirect('tutorial_detail', tutorial_id=tutorial.id)
    
    # Estatísticas para quem criou o tutorial
    stats = None
    if request.user == tutorial.created_by or request.user.can_manage_users():
        stats = {
            'total_views': tutorial.get_viewers_count(),
            'total_completed': tutorial.get_completed_count(),
            'viewers': tutorial.get_all_viewers()[:10],  # Primeiros 10
            'completed_users': tutorial.get_all_completed()[:10],  # Primeiros 10
        }
    
    context = {
        'tutorial': tutorial,
        'progress': progress,
        'stats': stats,
    }
    return render(request, 'help/tutorial_detail.html', context)


def forgot_password_view(request):
    """Solicitar redefinição de senha — envia o link via WhatsApp (Z-API)."""
    if request.user.is_authenticated:
        return redirect('dashboard')

    if request.method == 'POST':
        email = (request.POST.get('email') or '').strip()

        if not email:
            messages.error(request, 'Por favor, insira seu email.')
            return render(request, 'users/forgot_password.html')

        # Mensagem genérica de sucesso para não revelar se o email existe.
        generic_success = (
            'Se o email existir em nosso sistema, enviaremos um WhatsApp com as '
            'instruções para o número cadastrado.'
        )

        try:
            user = User.objects.get(email__iexact=email, is_active=True)
        except User.DoesNotExist:
            messages.success(request, generic_success)
            return redirect('login')

        phone = (getattr(user, 'phone', '') or '').strip()
        if not phone:
            messages.error(
                request,
                'Não há um número de WhatsApp cadastrado para este email. '
                'Procure o administrador do sistema.',
            )
            return render(request, 'users/forgot_password.html')

        # Gerar token + link de redefinição.
        from django.contrib.auth.tokens import default_token_generator
        from django.utils.http import urlsafe_base64_encode
        from django.utils.encoding import force_bytes

        token = default_token_generator.make_token(user)
        uid = urlsafe_base64_encode(force_bytes(user.pk))
        reset_link = request.build_absolute_uri(f'/reset-password/{uid}/{token}/')

        wa_message = (
            f"Olá, {user.first_name or user.username}! 👋\n\n"
            "Recebemos uma solicitação para redefinir a sua senha no "
            "Sistema Rede Confiança.\n\n"
            f"Acesse o link abaixo para criar uma nova senha (válido por algumas horas):\n{reset_link}\n\n"
            "Se você não solicitou esta redefinição, basta ignorar esta mensagem."
        )

        try:
            from core.zapi import send_whatsapp_message
            ok, detail = send_whatsapp_message(phone, wa_message)
        except Exception:  # noqa: BLE001
            ok, detail = False, 'falha inesperada'

        if ok:
            messages.success(request, generic_success)
            return redirect('login')

        messages.error(
            request,
            'Não foi possível enviar a mensagem de WhatsApp agora. Tente novamente em instantes.',
        )
        return render(request, 'users/forgot_password.html')

    return render(request, 'users/forgot_password.html')


def reset_password_view(request, uidb64, token):
    """Redefinir senha com token"""
    if request.user.is_authenticated:
        return redirect('dashboard')
    
    from django.contrib.auth.tokens import default_token_generator
    from django.utils.http import urlsafe_base64_decode
    from django.utils.encoding import force_str
    
    try:
        uid = force_str(urlsafe_base64_decode(uidb64))
        user = User.objects.get(pk=uid)
    except (TypeError, ValueError, OverflowError, User.DoesNotExist):
        user = None
    
    if user is not None and default_token_generator.check_token(user, token):
        if request.method == 'POST':
            new_password = request.POST.get('new_password')
            confirm_password = request.POST.get('confirm_password')
            
            if not new_password or not confirm_password:
                messages.error(request, 'Todos os campos são obrigatórios.')
                return render(request, 'users/reset_password.html', {'validlink': True})
            
            if new_password != confirm_password:
                messages.error(request, 'As senhas não coincidem.')
                return render(request, 'users/reset_password.html', {'validlink': True})
            
            if len(new_password) < 8:
                messages.error(request, 'A senha deve ter pelo menos 8 caracteres.')
                return render(request, 'users/reset_password.html', {'validlink': True})
            
            try:
                user.set_password(new_password)
                user.save()
                
                log_action(
                    user,
                    'PASSWORD_RESET',
                    'Senha redefinida via email',
                    request
                )
                
                messages.success(request, 'Senha redefinida com sucesso! Faça login com sua nova senha.')
                return redirect('login')
                
            except Exception as e:
                messages.error(request, f'Erro ao redefinir senha: {str(e)}')
                return render(request, 'users/reset_password.html', {'validlink': True})
        
        return render(request, 'users/reset_password.html', {'validlink': True})
    else:
        return render(request, 'users/reset_password.html', {'validlink': False})


@login_required
def change_password_view(request):
    """Alterar senha do usuário"""
    if request.method == 'POST':
        current_password = request.POST.get('current_password')
        new_password = request.POST.get('new_password')
        confirm_password = request.POST.get('confirm_password')
        
        # Validações
        if not current_password or not new_password or not confirm_password:
            messages.error(request, 'Todos os campos são obrigatórios.')
            return render(request, 'users/change_password.html')
        
        # Verificar senha atual
        if not request.user.check_password(current_password):
            messages.error(request, 'Senha atual incorreta.')
            return render(request, 'users/change_password.html')
        
        # Verificar se as senhas novas coincidem
        if new_password != confirm_password:
            messages.error(request, 'As senhas não coincidem.')
            return render(request, 'users/change_password.html')
        
        # Validar força da senha
        if len(new_password) < 8:
            messages.error(request, 'A senha deve ter pelo menos 8 caracteres.')
            return render(request, 'users/change_password.html')
        
        # Verificar se a nova senha não é igual à atual
        if request.user.check_password(new_password):
            messages.error(request, 'A nova senha deve ser diferente da senha atual.')
            return render(request, 'users/change_password.html')
        
        try:
            # Alterar a senha
            request.user.set_password(new_password)
            request.user.save()
            
            # Log da ação
            log_action(
                request.user,
                'PASSWORD_CHANGE',
                'Senha alterada pelo usuário',
                request
            )
            
            # Fazer login novamente para manter a sessão
            from django.contrib.auth import update_session_auth_hash
            update_session_auth_hash(request, request.user)
            
            messages.success(request, 'Senha alterada com sucesso!')
            return redirect('profile')
            
        except Exception as e:
            messages.error(request, f'Erro ao alterar senha: {str(e)}')
            return render(request, 'users/change_password.html')
    
    return render(request, 'users/change_password.html')


@login_required
def manage_tutorials_view(request):
    """Gerenciar tutoriais (admin only)"""
    if not request.user.can_manage_users():
        messages.error(request, 'Você não tem permissão para acessar esta página.')
        return redirect('dashboard')
    
    from core.models import Tutorial, TrainingCategory
    
    # Filtros
    category_filter = request.GET.get('category', '')
    
    tutorials = Tutorial.objects.all()
    if category_filter:
        tutorials = tutorials.filter(category_id=category_filter)
    
    tutorials = tutorials.select_related('created_by', 'category').order_by('order', 'title')
    
    # Adicionar estatísticas para cada tutorial
    tutorials_with_stats = []
    for tutorial in tutorials:
        tutorials_with_stats.append({
            'tutorial': tutorial,
            'viewers_count': tutorial.get_viewers_count(),
            'completed_count': tutorial.get_completed_count(),
        })
    
    categories = TrainingCategory.objects.filter(is_active=True).order_by('name')
    
    context = {
        'tutorials_with_stats': tutorials_with_stats,
        'categories': categories,
        'selected_category': category_filter,
    }
    return render(request, 'admin/tutorials.html', context)


@login_required
def manage_training_categories_view(request):
    """Gerenciar categorias de treinamento"""
    if not request.user.can_manage_users():
        messages.error(request, 'Você não tem permissão para acessar esta página.')
        return redirect('dashboard')
    
    from core.models import TrainingCategory
    
    if request.method == 'POST':
        name = request.POST.get('name')
        description = request.POST.get('description', '')
        is_active = request.POST.get('is_active') == 'on'
        
        try:
            category = TrainingCategory.objects.create(
                name=name,
                description=description,
                is_active=is_active
            )
            
            log_action(
                request.user,
                'TRAINING_CATEGORY_CREATE',
                f'Categoria de treinamento criada: {category.name}',
                request
            )
            
            messages.success(request, f'Categoria "{category.name}" criada com sucesso!')
            return redirect('manage_training_categories')
            
        except Exception as e:
            messages.error(request, f'Erro ao criar categoria: {str(e)}')
            return redirect('manage_training_categories')
    
    categories = TrainingCategory.objects.all().order_by('name')
    
    context = {
        'categories': categories,
    }
    return render(request, 'admin/training_categories.html', context)


@login_required
def create_training_category_view(request):
    """Criar nova categoria de treinamento"""
    if not request.user.can_manage_users():
        messages.error(request, 'Você não tem permissão para realizar esta ação.')
        return redirect('dashboard')
    
    if request.method == 'POST':
        from core.models import TrainingCategory
        
        name = request.POST.get('name')
        description = request.POST.get('description', '')
        color = request.POST.get('color', '#3B82F6')
        icon = request.POST.get('icon', 'fas fa-graduation-cap')
        
        try:
            category = TrainingCategory.objects.create(
                name=name,
                description=description,
                color=color,
                icon=icon
            )
            
            log_action(
                request.user,
                'TRAINING_CATEGORY_CREATE',
                f'Categoria de treinamento criada: {category.name}',
                request
            )
            
            messages.success(request, f'Categoria "{category.name}" criada com sucesso!')
            return redirect('manage_training_categories')
            
        except Exception as e:
            messages.error(request, f'Erro ao criar categoria: {str(e)}')
    
    return render(request, 'admin/create_training_category.html')


@login_required
def edit_training_category_view(request, category_id):
    """Editar categoria de treinamento"""
    if not request.user.can_manage_users():
        if request.headers.get('Content-Type') == 'application/x-www-form-urlencoded':
            return JsonResponse({'success': False, 'error': 'Permissão negada'})
        messages.error(request, 'Você não tem permissão para realizar esta ação.')
        return redirect('dashboard')
    
    from core.models import TrainingCategory
    from django.http import JsonResponse
    
    category = get_object_or_404(TrainingCategory, id=category_id)
    
    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'toggle_status':
            is_active = request.POST.get('is_active') == 'true'
            category.is_active = is_active
            category.save()
            
            log_action(
                request.user,
                'TRAINING_CATEGORY_TOGGLE',
                f'Categoria {"ativada" if is_active else "desativada"}: {category.name}',
                request
            )
            
            return JsonResponse({
                'success': True, 
                'message': f'Categoria "{category.name}" {"ativada" if is_active else "desativada"} com sucesso!'
            })
            
        elif action == 'delete':
            category_name = category.name
            try:
                category.delete()
                
                log_action(
                    request.user,
                    'TRAINING_CATEGORY_DELETE',
                    f'Categoria de treinamento excluída: {category_name}',
                    request
                )
                
                return JsonResponse({
                    'success': True,
                    'message': f'Categoria "{category_name}" excluída com sucesso!'
                })
            except Exception as e:
                return JsonResponse({
                    'success': False,
                    'error': f'Erro ao excluir categoria: {str(e)}'
                })
        
        else:
            # Edição normal do formulário
            name = request.POST.get('name')
            description = request.POST.get('description', '')
            is_active = request.POST.get('is_active') == 'on'
            category_id = request.POST.get('category_id')
            
            try:
                if category_id:
                    # Editando categoria existente
                    category.name = name
                    category.description = description
                    category.is_active = is_active
                    category.save()
                    
                    log_action(
                        request.user,
                        'TRAINING_CATEGORY_UPDATE',
                        f'Categoria de treinamento atualizada: {category.name}',
                        request
                    )
                    
                    messages.success(request, f'Categoria "{category.name}" atualizada com sucesso!')
                else:
                    # Criando nova categoria
                    category = TrainingCategory.objects.create(
                        name=name,
                        description=description,
                        is_active=is_active
                    )
                    
                    log_action(
                        request.user,
                        'TRAINING_CATEGORY_CREATE',
                        f'Categoria de treinamento criada: {category.name}',
                        request
                    )
                    
                    messages.success(request, f'Categoria "{category.name}" criada com sucesso!')
                
                return redirect('manage_training_categories')
                
            except Exception as e:
                messages.error(request, f'Erro ao salvar categoria: {str(e)}')
    
    context = {
        'category': category,
    }
    return render(request, 'admin/edit_training_category.html', context)


@login_required
def create_tutorial_view(request):
    """Criar novo tutorial (admin only)"""
    if not request.user.can_manage_users():
        messages.error(request, 'Você não tem permissão para acessar esta página.')
        return redirect('dashboard')
    
    if request.method == 'POST':
        from core.models import Tutorial, TrainingCategory
        
        title = request.POST.get('title')
        description = request.POST.get('description')
        category_id = request.POST.get('category')
        pdf_file = request.FILES.get('pdf_file')
        order = request.POST.get('order', 0)
        
        try:
            category = None
            if category_id:
                category = get_object_or_404(TrainingCategory, id=category_id)
            
            tutorial = Tutorial.objects.create(
                title=title,
                description=description,
                category=category,
                pdf_file=pdf_file,
                order=order,
                created_by=request.user
            )
            
            log_action(
                request.user,
                'TUTORIAL_CREATE',
                f'Tutorial criado: {tutorial.title}',
                request
            )
            
            messages.success(request, 'Tutorial criado com sucesso!')
            return redirect('manage_tutorials')
            
        except Exception as e:
            messages.error(request, f'Erro ao criar tutorial: {str(e)}')
    
    from core.models import TrainingCategory
    categories = TrainingCategory.objects.filter(is_active=True).order_by('name')
    
    context = {
        'categories': categories,
    }
    return render(request, 'admin/create_tutorial.html', context)


@login_required
def manage_prizes_view(request):
    """Gerenciar prêmios (admin only)"""
    if not request.user.can_manage_users():
        messages.error(request, 'Você não tem permissão para acessar esta página.')
        return redirect('dashboard')
    
    from prizes.models import Prize, Redemption
    from django.db.models import Count, Sum, Q
    
    # Buscar prêmios com estatísticas de resgates
    prizes = Prize.objects.annotate(
        total_redemptions=Count('redemption'),
        pending_redemptions=Count('redemption', filter=Q(redemption__status='PENDENTE')),
        approved_redemptions=Count('redemption', filter=Q(redemption__status='APROVADO')),
        delivered_redemptions=Count('redemption', filter=Q(redemption__status='ENTREGUE')),
        total_cs_spent=Sum('redemption__prize__value_cs', filter=Q(redemption__status__in=['APROVADO', 'ENTREGUE']))
    ).order_by('-created_at')
    
    # Resgates recentes (últimos 15 dias, todos os status)
    from django.utils import timezone
    from datetime import timedelta
    fifteen_days_ago = timezone.now() - timedelta(days=15)
    recent_redemptions = Redemption.objects.filter(
        redeemed_at__gte=fifteen_days_ago
    ).select_related('user', 'prize').order_by('-redeemed_at')[:10]
    
    # Calcular total real de C$ em circulação somando os saldos de todos os usuários
    total_cs_circulation = User.objects.aggregate(
        total=Sum('balance_cs')
    )['total'] or Decimal('0')
    
    context = {
        'prizes': prizes,
        'recent_redemptions': recent_redemptions,
        'total_prizes': prizes.count(),
        'active_prizes': prizes.filter(is_active=True).count(),
        'pending_redemptions': Redemption.objects.filter(status='PENDENTE').count(),
        'total_cs_circulation': total_cs_circulation,
        'user': request.user,
    }
    return render(request, 'prizes/manage.html', context)


@login_required
def create_prize_view(request):
    """Criar novo prêmio (admin only)"""
    if not request.user.can_manage_users():
        messages.error(request, 'Você não tem permissão para acessar esta página.')
        return redirect('dashboard')
    
    if request.method == 'POST':
        from prizes.models import Prize
        
        name = request.POST.get('name')
        description = request.POST.get('description')
        value_cs = request.POST.get('value_cs')
        image = request.FILES.get('image')
        stock = request.POST.get('stock', 0)
        unlimited_stock = request.POST.get('unlimited_stock') == 'on'
        
        prize = Prize.objects.create(
            name=name,
            description=description,
            value_cs=value_cs,
            image=image,
            stock=stock if not unlimited_stock else 0,
            unlimited_stock=unlimited_stock
        )
        
        messages.success(request, 'Prêmio criado com sucesso!')
        return redirect('manage_prizes')
    
    return render(request, 'prizes/create.html')


@login_required
def edit_prize_view(request, prize_id):
    """Editar prêmio (admin only)"""
    if not request.user.can_manage_users():
        messages.error(request, 'Você não tem permissão para acessar esta página.')
        return redirect('dashboard')
    
    from prizes.models import Prize, PrizeCategory, Redemption
    prize = get_object_or_404(Prize, id=prize_id)
    
    if request.method == 'POST':
        prize.name = request.POST.get('name', prize.name)
        prize.description = request.POST.get('description', prize.description)
        prize.value_cs = request.POST.get('cost', prize.value_cs)  # Corrigindo para 'cost'
        
        # Processar categoria
        category_id = request.POST.get('category')
        if category_id:
            prize.category_id = category_id
        else:
            prize.category = None
        
        # Processar remoção de imagem
        if request.POST.get('remove_image'):
            if prize.image:
                prize.image.delete()
                prize.image = None
        
        # Processar upload de nova imagem
        if request.FILES.get('image'):
            if prize.image:
                prize.image.delete()
            prize.image = request.FILES.get('image')
        
        # Status ativo
        prize.is_active = request.POST.get('is_active') == 'on'
        
        # Estoque
        stock = request.POST.get('stock')
        if stock:
            prize.stock = int(stock)
        
        try:
            prize.save()
            
            log_action(
                request.user,
                'PRIZE_UPDATE',
                f'Prêmio atualizado: {prize.name}',
                request
            )
            
            messages.success(request, 'Prêmio atualizado com sucesso!')
            return redirect('manage_prizes')
            
        except Exception as e:
            messages.error(request, f'Erro ao atualizar prêmio: {str(e)}')
    
    # Buscar categorias e resgates
    categories = PrizeCategory.objects.all().order_by('name')
    redemptions = Redemption.objects.filter(prize=prize).order_by('-redeemed_at')
    
    context = {
        'prize': prize,
        'categories': categories,
        'redemptions': redemptions,
        'user': request.user,
    }
    return render(request, 'prizes/edit.html', context)


@login_required
def toggle_prize_status_view(request, prize_id):
    """Pausar/Despausar prêmio (admin only)"""
    if not request.user.can_manage_users():
        return JsonResponse({'success': False, 'error': 'Acesso negado'})
    
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Método não permitido'})
    
    from prizes.models import Prize
    prize = get_object_or_404(Prize, id=prize_id)
    
    try:
        prize.is_active = not prize.is_active
        prize.save()
        
        status_text = "ativado" if prize.is_active else "pausado"
        
        log_action(
            request.user,
            'PRIZE_STATUS_TOGGLE',
            f'Prêmio {status_text}: {prize.name}',
            request
        )
        
        return JsonResponse({
            'success': True, 
            'is_active': prize.is_active,
            'message': f'Prêmio {status_text} com sucesso!'
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
def prize_redemptions_view(request, prize_id):
    """Ver resgates de um prêmio específico (admin only)"""
    if not request.user.can_manage_users():
        messages.error(request, 'Você não tem permissão para acessar esta página.')
        return redirect('dashboard')
    
    from prizes.models import Prize, Redemption
    prize = get_object_or_404(Prize, id=prize_id)
    
    # Filtros
    status_filter = request.GET.get('status', '')
    
    redemptions = Redemption.objects.filter(prize=prize).select_related('user')
    
    if status_filter:
        redemptions = redemptions.filter(status=status_filter)
    
    redemptions = redemptions.order_by('-redeemed_at')
    
    # Estatísticas específicas do prêmio
    stats = {
        'total': redemptions.count(),
        'pending': redemptions.filter(status='PENDENTE').count(),
        'approved': redemptions.filter(status='APROVADO').count(),
        'delivered': redemptions.filter(status='ENTREGUE').count(),
        'canceled': redemptions.filter(status='CANCELADO').count(),
    }
    
    context = {
        'prize': prize,
        'redemptions': redemptions,
        'stats': stats,
        'current_status': status_filter,
        'user': request.user,
    }
    return render(request, 'prizes/prize_redemptions.html', context)


@login_required
def manage_redemptions_view(request):
    """Gerenciar resgates (admin only)"""
    if not request.user.can_manage_users():
        messages.error(request, 'Você não tem permissão para acessar esta página.')
        return redirect('dashboard')
    
    from prizes.models import Redemption
    from django.db.models import Q, Count
    
    # Filtros
    search = request.GET.get('search', '')
    status_filter = request.GET.get('status', '')
    
    redemptions = Redemption.objects.select_related('user', 'prize', 'prize__category').all()
    
    if search:
        redemptions = redemptions.filter(
            Q(user__first_name__icontains=search) |
            Q(user__last_name__icontains=search) |
            Q(user__email__icontains=search) |
            Q(prize__name__icontains=search)
        )
    
    if status_filter:
        redemptions = redemptions.filter(status=status_filter)
    
    redemptions = redemptions.order_by('-redeemed_at')
    
    # Estatísticas
    stats = {
        'pending': Redemption.objects.filter(status='PENDENTE').count(),
        'approved': Redemption.objects.filter(status='APROVADO').count(),
        'delivered': Redemption.objects.filter(status='ENTREGUE').count(),
        'canceled': Redemption.objects.filter(status='CANCELADO').count(),
    }
    
    context = {
        'redemptions': redemptions,
        'stats': stats,
        'user': request.user,
    }
    return render(request, 'prizes/manage_redemptions.html', context)


@login_required
def manage_groups_view(request):
    """Gerenciar grupos de comunicação"""
    if not request.user.can_manage_users():
        messages.error(request, 'Você não tem permissão para acessar esta página.')
        return redirect('dashboard')
    
    from communications.models import CommunicationGroup
    from django.db.models import Sum
    
    groups = CommunicationGroup.objects.all().order_by('name')
    
    # Calcular total de usuários em grupos (sem duplicatas)
    total_users_in_groups = User.objects.filter(
        communication_groups__isnull=False
    ).distinct().count()
    
    context = {
        'groups': groups,
        'total_users_in_groups': total_users_in_groups,
        'user': request.user,
    }
    return render(request, 'admin/groups.html', context)


@login_required
def create_group_view(request):
    """Criar novo grupo de comunicação"""
    if not request.user.can_manage_users():
        messages.error(request, 'Você não tem permissão para realizar esta ação.')
        return redirect('dashboard')
    
    if request.method == 'POST':
        name = request.POST.get('name')
        description = request.POST.get('description', '')
        user_ids = request.POST.getlist('users')
        
        try:
            from communications.models import CommunicationGroup
            group = CommunicationGroup.objects.create(
                name=name,
                description=description,
                created_by=request.user
            )
            
            if user_ids:
                users = User.objects.filter(id__in=user_ids)
                group.members.set(users)
            
            log_action(
                request.user,
                'GROUP_CREATE',
                f'Grupo criado: {group.name}',
                request
            )
            
            messages.success(request, f'Grupo "{group.name}" criado com sucesso!')
            return redirect('manage_groups')
            
        except Exception as e:
            messages.error(request, f'Erro ao criar grupo: {str(e)}')
    
    users = User.objects.filter(is_active=True).order_by('first_name', 'last_name')
    context = {
        'users': users,
        'user': request.user,
    }
    return render(request, 'admin/create_group.html', context)


@login_required
def edit_group_view(request, group_id):
    """Editar grupo de comunicação"""
    if not request.user.can_manage_users():
        messages.error(request, 'Você não tem permissão para realizar esta ação.')
        return redirect('dashboard')
    
    from communications.models import CommunicationGroup
    group = get_object_or_404(CommunicationGroup, id=group_id)
    
    if request.method == 'POST':
        name = request.POST.get('name')
        description = request.POST.get('description', '')
        user_ids = request.POST.getlist('users')
        
        try:
            group.name = name
            group.description = description
            group.save()
            
            if user_ids:
                users = User.objects.filter(id__in=user_ids)
                group.members.set(users)
            else:
                group.members.clear()
            
            log_action(
                request.user,
                'GROUP_UPDATE',
                f'Grupo atualizado: {group.name}',
                request
            )
            
            messages.success(request, f'Grupo "{group.name}" atualizado com sucesso!')
            return redirect('manage_groups')
            
        except Exception as e:
            messages.error(request, f'Erro ao atualizar grupo: {str(e)}')
    
    users = User.objects.filter(is_active=True).order_by('first_name', 'last_name')
    context = {
        'group': group,
        'users': users,
        'user': request.user,
    }
    return render(request, 'admin/edit_group.html', context)


@login_required
def delete_group_view(request, group_id):
    """Deletar grupo de comunicação"""
    if not request.user.can_manage_users():
        messages.error(request, 'Você não tem permissão para realizar esta ação.')
        return redirect('dashboard')
    
    from communications.models import CommunicationGroup
    group = get_object_or_404(CommunicationGroup, id=group_id)
    
    if request.method == 'POST':
        group_name = group.name
        group.delete()
        
        log_action(
            request.user,
            'GROUP_DELETE',
            f'Grupo deletado: {group_name}',
            request
        )
        
        messages.success(request, f'Grupo "{group_name}" deletado com sucesso!')
        return redirect('manage_groups')
    
    context = {
        'group': group,
        'user': request.user,
    }
    return render(request, 'admin/delete_group.html', context)


@login_required
def manage_prize_categories_view(request):
    """Gerenciar categorias de prêmios"""
    if not request.user.can_manage_users():
        messages.error(request, 'Você não tem permissão para acessar esta área.')
        return redirect('dashboard')
    
    from prizes.models import PrizeCategory
    categories = PrizeCategory.objects.all().order_by('name')
    
    context = {
        'categories': categories,
        'user': request.user,
    }
    return render(request, 'admin/manage_prize_categories.html', context)


@login_required
def create_prize_category_view(request):
    """Criar nova categoria de prêmio"""
    if not request.user.can_manage_users():
        messages.error(request, 'Você não tem permissão para realizar esta ação.')
        return redirect('dashboard')
    
    if request.method == 'POST':
        name = request.POST.get('name')
        description = request.POST.get('description', '')
        icon = request.POST.get('icon', 'fas fa-gift')
        color = request.POST.get('color', 'blue')
        active = request.POST.get('active') == 'on'
        
        try:
            from prizes.models import PrizeCategory
            category = PrizeCategory.objects.create(
                name=name,
                description=description,
                icon=icon,
                color=color,
                active=active
            )
            
            log_action(
                request.user,
                'PRIZE_CATEGORY_CREATE',
                f'Categoria de prêmio criada: {category.name}',
                request
            )
            
            messages.success(request, f'Categoria "{category.name}" criada com sucesso!')
            return redirect('manage_prize_categories')
            
        except Exception as e:
            messages.error(request, f'Erro ao criar categoria: {str(e)}')
    
    context = {
        'user': request.user,
    }
    return render(request, 'admin/create_prize_category.html', context)


@login_required
def edit_prize_category_view(request, category_id):
    """Editar categoria de prêmio"""
    if not request.user.can_manage_users():
        messages.error(request, 'Você não tem permissão para realizar esta ação.')
        return redirect('dashboard')
    
    from prizes.models import PrizeCategory
    category = get_object_or_404(PrizeCategory, id=category_id)
    
    if request.method == 'POST':
        category.name = request.POST.get('name')
        category.description = request.POST.get('description', '')
        category.icon = request.POST.get('icon', 'fas fa-gift')
        category.color = request.POST.get('color', 'blue')
        category.active = request.POST.get('active') == 'on'
        
        try:
            category.save()
            
            log_action(
                request.user,
                'PRIZE_CATEGORY_UPDATE',
                f'Categoria de prêmio atualizada: {category.name}',
                request
            )
            
            messages.success(request, f'Categoria "{category.name}" atualizada com sucesso!')
            return redirect('manage_prize_categories')
            
        except Exception as e:
            messages.error(request, f'Erro ao atualizar categoria: {str(e)}')
    
    context = {
        'category': category,
        'user': request.user,
    }
    return render(request, 'admin/edit_prize_category.html', context)


@login_required
def delete_prize_category_view(request, category_id):
    """Deletar categoria de prêmio"""
    if not request.user.can_manage_users():
        messages.error(request, 'Você não tem permissão para realizar esta ação.')
        return redirect('dashboard')
    
    from prizes.models import PrizeCategory
    category = get_object_or_404(PrizeCategory, id=category_id)
    
    if request.method == 'POST':
        category_name = category.name
        category.delete()
        
        log_action(
            request.user,
            'PRIZE_CATEGORY_DELETE',
            f'Categoria de prêmio deletada: {category_name}',
            request
        )
        
        messages.success(request, f'Categoria "{category_name}" deletada com sucesso!')
        return redirect('manage_prize_categories')
    
    context = {
        'category': category,
        'user': request.user,
    }
    return render(request, 'admin/delete_prize_category.html', context)


@login_required
def pending_cs_transactions_view(request):
    """Visualizar transações C$ pendentes de aprovação"""
    if not request.user.can_manage_users():
        messages.error(request, 'Acesso negado.')
        return redirect('dashboard')

    from prizes.models import CSTransaction
    from django.db.models import Sum, Count, Q

    pending_transactions = CSTransaction.objects.filter(
        status='PENDING',
        transaction_type='CREDIT'
    ).select_related('user', 'user__sector', 'created_by').order_by('-created_at')

    pending_stats = pending_transactions.aggregate(
        total=Sum('amount'),
        quantidade=Count('id'),
        solicitantes=Count('user', distinct=True),
    )

    # Movimentações resolvidas hoje (para o admin ter noção do ritmo de aprovação).
    today = timezone.localdate()
    resolved_today = CSTransaction.objects.filter(
        transaction_type='CREDIT',
        approved_at__date=today,
    ).aggregate(
        aprovadas=Count('id', filter=Q(status='APPROVED')),
        aprovadas_valor=Sum('amount', filter=Q(status='APPROVED')),
        rejeitadas=Count('id', filter=Q(status='REJECTED')),
    )

    context = {
        'pending_transactions': pending_transactions,
        'total_pending_amount': pending_stats['total'] or 0,
        'pending_count': pending_stats['quantidade'] or 0,
        'pending_users_count': pending_stats['solicitantes'] or 0,
        'approved_today_count': resolved_today['aprovadas'] or 0,
        'approved_today_amount': resolved_today['aprovadas_valor'] or 0,
        'rejected_today_count': resolved_today['rejeitadas'] or 0,
        'user': request.user,
    }
    return render(request, 'admin/pending_cs_transactions.html', context)


@login_required 
@require_POST
def approve_cs_transaction(request, transaction_id):
    """Aprovar uma transação C$ pendente"""
    if not request.user.can_manage_users():
        return JsonResponse({'error': 'Acesso negado. Apenas SUPERADMIN pode aprovar transações C$.'}, status=403)
    
    from prizes.models import CSTransaction
    from django.utils import timezone
    
    try:
        with transaction.atomic():
            cs_transaction = get_object_or_404(CSTransaction, id=transaction_id, status='PENDING')
            
            # Verificar se o usuário não está tentando aprovar sua própria solicitação
            if cs_transaction.created_by == request.user:
                return JsonResponse({
                    'error': 'Você não pode aprovar sua própria solicitação de C$. A aprovação deve ser feita por outro SUPERADMIN.'
                }, status=400)
            
            # Verificar se o aprovador é realmente SUPERADMIN
            if request.user.hierarchy != 'SUPERADMIN':
                return JsonResponse({
                    'error': 'Apenas usuários com hierarquia SUPERADMIN podem aprovar transações C$.'
                }, status=403)
            
            # Aprovar a transação
            cs_transaction.status = 'APPROVED'
            cs_transaction.approved_by = request.user
            cs_transaction.approved_at = timezone.now()
            cs_transaction.save()
            
            # Adicionar o valor ao saldo do usuário APENAS se for uma transação de crédito
            if cs_transaction.transaction_type == 'CREDIT':
                user = cs_transaction.user
                user.balance_cs += cs_transaction.amount
                user.save()
                
                log_action(
                    request.user,
                    'CS_APPROVE',
                    f'Transação C$ aprovada: +C$ {cs_transaction.amount} para {user.get_full_name()} por {request.user.get_full_name()}',
                    request
                )
                
                return JsonResponse({
                    'success': True,
                    'message': f'Transação aprovada com sucesso! C$ {cs_transaction.amount} foi adicionado ao saldo de {user.get_full_name()}.',
                    'new_balance': float(user.balance_cs)
                })
            else:
                log_action(
                    request.user,
                    'CS_APPROVE',
                    f'Transação C$ aprovada: {cs_transaction.transaction_type} de C$ {cs_transaction.amount} para {cs_transaction.user.get_full_name()} por {request.user.get_full_name()}',
                    request
                )
                
                return JsonResponse({
                    'success': True,
                    'message': f'Transação {cs_transaction.get_transaction_type_display().lower()} aprovada com sucesso!',
                })
            
    except CSTransaction.DoesNotExist:
        return JsonResponse({'error': 'Transação não encontrada ou já foi processada.'}, status=404)
    except Exception as e:
        return JsonResponse({'error': f'Erro interno do servidor: {str(e)}'}, status=500)


@login_required
@require_POST  
def reject_cs_transaction(request, transaction_id):
    """Rejeitar uma transação C$ pendente"""
    if not request.user.can_manage_users():
        return JsonResponse({'error': 'Acesso negado. Apenas SUPERADMIN pode rejeitar transações C$.'}, status=403)
    
    from prizes.models import CSTransaction
    from django.utils import timezone
    
    try:
        with transaction.atomic():
            cs_transaction = get_object_or_404(CSTransaction, id=transaction_id, status='PENDING')
            
            # Verificar se o usuário não está tentando rejeitar sua própria solicitação
            if cs_transaction.created_by == request.user:
                return JsonResponse({
                    'error': 'Você não pode rejeitar sua própria solicitação de C$. A rejeição deve ser feita por outro SUPERADMIN.'
                }, status=400)
            
            # Verificar se o aprovador é realmente SUPERADMIN
            if request.user.hierarchy != 'SUPERADMIN':
                return JsonResponse({
                    'error': 'Apenas usuários com hierarquia SUPERADMIN podem rejeitar transações C$.'
                }, status=403)
            
            # Obter motivo da rejeição (se fornecido)
            rejection_reason = request.POST.get('reason', '').strip()
            
            # Rejeitar a transação
            cs_transaction.status = 'REJECTED'
            cs_transaction.approved_by = request.user
            cs_transaction.approved_at = timezone.now()
            if rejection_reason:
                cs_transaction.rejection_reason = rejection_reason
            cs_transaction.save()
            
            log_action(
                request.user,
                'CS_REJECT',
                f'Transação C$ rejeitada: C$ {cs_transaction.amount} para {cs_transaction.user.get_full_name()} por {request.user.get_full_name()}' + 
                (f' - Motivo: {rejection_reason}' if rejection_reason else ''),
                request
            )
            
            return JsonResponse({
                'success': True,
                'message': f'Transação rejeitada com sucesso. C$ {cs_transaction.amount} para {cs_transaction.user.get_full_name()} foi negado.' +
                          (f' Motivo: {rejection_reason}' if rejection_reason else '')
            })
        
    except CSTransaction.DoesNotExist:
        return JsonResponse({'error': 'Transação não encontrada ou já foi processada.'}, status=404)
    except Exception as e:
        return JsonResponse({'error': f'Erro interno do servidor: {str(e)}'}, status=500)


# ===== CHECKLIST VIEWS =====
@login_required
def checklist_dashboard_view(request):
    """Dashboard de checklists do usuário"""
    from core.models import DailyChecklist, ChecklistItem
    from django.utils import timezone
    from datetime import date, timedelta
    
    today = date.today()
    user = request.user
    is_supervisor = user.hierarchy in ['SUPERVISOR', 'ADMIN', 'SUPERADMIN', 'ADMINISTRATIVO'] or user.is_staff
    
    # Usuários padrões também podem criar checklists para outros usuários padrões do mesmo setor
    can_create_checklist = is_supervisor or (user.hierarchy == 'PADRAO' and (user.sector or user.sectors.exists()))
    
    # Checklist de hoje do usuário
    today_checklist = DailyChecklist.objects.filter(
        user=user,
        date=today
    ).prefetch_related('items').first()
    
    # Checklists da semana do usuário (últimos 7 dias)
    week_ago = today - timedelta(days=7)
    week_checklists = DailyChecklist.objects.filter(
        user=user,
        date__gte=week_ago,
        date__lte=today
    ).prefetch_related('items').order_by('-date')
    
    # Para supervisores: checklists que criaram para usuários dos seus setores
    # Para SUPERADMIN: todos os checklists
    # Para usuários padrões: checklists que criaram para outros usuários padrões do mesmo setor
    sector_checklists = []
    if is_supervisor:
        if user.hierarchy == 'SUPERADMIN':
            # SUPERADMIN vê todos os checklists de todos os setores
            sector_checklists = DailyChecklist.objects.filter(
                date__gte=week_ago,
                date__lte=today
            ).select_related('user', 'template', 'user__sector').prefetch_related('items').order_by('-date')
        else:
            # Supervisores veem checklists de todos os seus setores
            from django.db.models import Q
            
            # Obter setores do usuário
            user_sectors = user.sectors.all()
            
            # Criar filtro para todos os setores do usuário
            sector_filter = Q()
            
            # Adicionar setores múltiplos
            if user_sectors.exists():
                sector_filter |= Q(sectors__in=user_sectors)
            
            # Adicionar setor principal se existir
            if user.sector:
                sector_filter |= Q(sector=user.sector)
            
            if sector_filter:
                # Buscar usuários dos setores
                sector_users = User.objects.filter(sector_filter).exclude(id=user.id).distinct()
                
                sector_checklists = DailyChecklist.objects.filter(
                    user__in=sector_users,
                    date__gte=week_ago,
                    date__lte=today
                ).select_related('user', 'template', 'user__sector').prefetch_related('items').order_by('-date')
    elif user.hierarchy == 'PADRAO' and can_create_checklist:
        # Usuários padrões veem checklists que criaram para outros usuários padrões do setor
        from django.db.models import Q
        
        # Obter setores do usuário
        user_sectors = user.sectors.all()
        
        # Criar filtro para todos os setores do usuário
        sector_filter = Q()
        
        # Adicionar setores múltiplos
        if user_sectors.exists():
            sector_filter |= Q(sectors__in=user_sectors)
        
        # Adicionar setor principal se existir
        if user.sector:
            sector_filter |= Q(sector=user.sector)
        
        if sector_filter:
            # Buscar checklists criados pelo usuário para outros usuários do mesmo setor
            sector_checklists = DailyChecklist.objects.filter(
                created_by=user,
                date__gte=week_ago,
                date__lte=today
            ).exclude(user=user).select_related('user', 'template', 'user__sector').prefetch_related('items').order_by('-date')
    
    # Estatísticas
    total_checklists = DailyChecklist.objects.filter(user=user).count()
    completed_checklists = DailyChecklist.objects.filter(
        user=user, 
        completed_at__isnull=False
    ).count()
    
    completion_rate = 0
    if total_checklists > 0:
        completion_rate = round((completed_checklists / total_checklists) * 100)
    
    # Templates criados pelo usuário
    from core.models import ChecklistTemplate
    user_templates = ChecklistTemplate.objects.filter(created_by=user, is_active=True).prefetch_related('items')
    
    # Estatísticas para supervisores e usuários padrões que criaram checklists
    sector_stats = {}
    if hasattr(sector_checklists, 'exists') and sector_checklists.exists():
        total_sector_checklists = sector_checklists.count()
        completed_sector_checklists = sector_checklists.filter(completed_at__isnull=False).count()
        sector_completion_rate = 0
        if total_sector_checklists > 0:
            sector_completion_rate = round((completed_sector_checklists / total_sector_checklists) * 100)
        
        sector_stats = {
            'total': total_sector_checklists,
            'completed': completed_sector_checklists,
            'completion_rate': sector_completion_rate
        }
    
    context = {
        'today_checklist': today_checklist,
        'week_checklists': week_checklists,
        'completion_rate': completion_rate,
        'total_checklists': total_checklists,
        'completed_checklists': completed_checklists,
        'user_templates': user_templates,
        'is_supervisor': is_supervisor,
        'can_create_checklist': can_create_checklist,
        'sector_checklists': sector_checklists,
        'sector_stats': sector_stats,
        'today': today,
    }
    return render(request, 'checklist/dashboard.html', context)


@login_required
def sector_checklists_view(request):
    """View para supervisores gerenciarem checklists do setor"""
    from core.models import DailyChecklist
    from django.utils import timezone
    from datetime import date, timedelta, datetime
    
    user = request.user
    is_supervisor = user.hierarchy in ['SUPERVISOR', 'ADMIN', 'SUPERADMIN', 'ADMINISTRATIVO'] or user.is_staff
    is_superadmin = user.hierarchy == 'SUPERADMIN'

    if not is_supervisor:
        messages.error(request, 'Você não tem permissão para acessar esta página.')
        return redirect('checklist_dashboard')

    # SUPERADMIN pode ver todos os setores, outros precisam estar vinculados a pelo menos um setor
    user_sectors = user.sectors.all()
    if not is_superadmin and not user_sectors.exists() and not user.sector:
        messages.error(request, 'Você precisa estar vinculado a um setor para gerenciar checklists.')
        return redirect('checklist_dashboard')

    # Filtros
    date_filter = request.GET.get('date', '')
    user_filter = request.GET.get('user', '')
    status_filter = request.GET.get('status', '')
    
    # Data padrão (últimos 30 dias)
    today = date.today()
    start_date = today - timedelta(days=30)
    
    if date_filter:
        try:
            filter_date = datetime.strptime(date_filter, '%Y-%m-%d').date()
            start_date = filter_date
        except ValueError:
            pass
    
    # Determinar usuários baseado na hierarquia
    if is_superadmin:
        # SUPERADMIN vê checklists de todos os setores
        sector_users = User.objects.exclude(id=user.id)
        checklists_query = DailyChecklist.objects.filter(
            date__gte=start_date
        ).select_related('user', 'template', 'user__sector').prefetch_related('items')
    else:
        # Supervisores veem checklists de todos os seus setores (tanto sectors quanto sector principal)
        from django.db.models import Q
        
        # Criar filtro para todos os setores do usuário
        sector_filter = Q()
        
        # Adicionar setores múltiplos
        if user_sectors.exists():
            sector_filter |= Q(sectors__in=user_sectors)
        
        # Adicionar setor principal se existir
        if user.sector:
            sector_filter |= Q(sector=user.sector)
        
        # Buscar usuários dos setores
        sector_users = User.objects.filter(sector_filter).exclude(id=user.id).distinct()
        
        checklists_query = DailyChecklist.objects.filter(
            user__in=sector_users,
            date__gte=start_date
        ).select_related('user', 'template', 'user__sector').prefetch_related('items')    # Aplicar filtros
    if user_filter:
        checklists_query = checklists_query.filter(user_id=user_filter)
    
    if status_filter == 'completed':
        checklists_query = checklists_query.filter(completed_at__isnull=False)
    elif status_filter == 'pending':
        checklists_query = checklists_query.filter(completed_at__isnull=True)
    
    checklists = checklists_query.order_by('-date', '-created_at')
    
    # Estatísticas
    total_checklists = checklists.count()
    completed_checklists = checklists.filter(completed_at__isnull=False).count()
    completion_rate = 0
    if total_checklists > 0:
        completion_rate = round((completed_checklists / total_checklists) * 100)
    
    # Determinar nome dos setores para o título
    if is_superadmin:
        sector_name = 'Todos os Setores'
    else:
        # Mostrar todos os setores do usuário
        sector_names = []
        if user_sectors.exists():
            sector_names.extend([sector.name for sector in user_sectors])
        if user.sector and user.sector.name not in sector_names:
            sector_names.append(user.sector.name)
        
        if sector_names:
            sector_name = ', '.join(sector_names)
        else:
            sector_name = 'Setor'

    context = {
        'checklists': checklists,
        'sector_users': sector_users,
        'total_checklists': total_checklists,
        'completed_checklists': completed_checklists,
        'completion_rate': completion_rate,
        'date_filter': date_filter,
        'user_filter': user_filter,
        'status_filter': status_filter,
        'sector_name': sector_name,
        'is_superadmin': is_superadmin,
    }
    
    return render(request, 'checklist/sector_management.html', context)


@login_required
def checklist_detail_view(request, checklist_id):
    """Detalhes de um checklist específico"""
    from core.models import DailyChecklist
    from django.db import models
    from django.http import Http404
    from django.contrib import messages
    
    user = request.user
    
    try:
        # Primeiro verifica se o checklist existe
        checklist_exists = DailyChecklist.objects.filter(id=checklist_id).exists()
        if not checklist_exists:
            raise Http404("Checklist não encontrado")
        
        # Lógica de permissões por hierarquia
        if user.hierarchy in ['SUPERADMIN', 'ADMIN']:
            # SuperAdmins e Admins podem ver qualquer checklist
            checklist = DailyChecklist.objects.get(id=checklist_id)
        
        elif user.hierarchy in ['SUPERVISOR', 'ADMINISTRATIVO'] or user.is_staff:
            # Supervisores podem ver:
            # 1. Seus próprios checklists
            # 2. Checklists que criaram
            # 3. Checklists de usuários do seu setor
            checklist_query = DailyChecklist.objects.filter(
                id=checklist_id
            ).filter(
                models.Q(user=user) |                          # Próprios checklists
                models.Q(created_by=user) |                    # Checklists que criou
                models.Q(user__sector=user.sector)             # Checklists do seu setor
            )
            
            if not checklist_query.exists():
                messages.error(request, 'Você não tem permissão para visualizar este checklist.')
                return redirect('checklist_dashboard')
            
            checklist = checklist_query.first()
        
        else:
            # Usuários comuns podem ver:
            # 1. Seus próprios checklists
            # 2. Checklists que criaram para outros usuários padrões do mesmo setor
            checklist_query = DailyChecklist.objects.filter(
                id=checklist_id
            ).filter(
                models.Q(user=user) |                          # Próprios checklists
                models.Q(created_by=user)                      # Checklists que criou
            )
            
            if not checklist_query.exists():
                messages.error(request, 'Você não tem permissão para visualizar este checklist.')
                return redirect('checklist_dashboard')
            
            checklist = checklist_query.first()
                
    except DailyChecklist.DoesNotExist:
        raise Http404("Checklist não encontrado")
    
    # Carregar itens com evidências
    items = checklist.items.prefetch_related('evidences').all().order_by('order', 'title')
    
    # Verificar se pode editar (dono do checklist ou quem criou)
    can_edit = (
        user == checklist.user or
        user == checklist.created_by or
        user.hierarchy in ['SUPERADMIN', 'ADMIN'] or
        (user.hierarchy in ['SUPERVISOR', 'ADMINISTRATIVO'] and checklist.user.sector == user.sector)
    )
    
    context = {
        'checklist': checklist,
        'items': items,
        'completion_percentage': checklist.get_completion_percentage(),
        'can_edit': can_edit,
    }
    return render(request, 'checklist/detail.html', context)


@login_required
@require_POST
def update_checklist_item_status(request, item_id):
    """Atualizar status de um item do checklist"""
    from core.models import ChecklistItem
    
    try:
        item = get_object_or_404(
            ChecklistItem,
            id=item_id,
            checklist__user=request.user
        )
        
        # Verificar se é atualização de "Não se aplica"
        is_not_applicable = request.POST.get('is_not_applicable')
        if is_not_applicable is not None:
            item.is_not_applicable = is_not_applicable == 'true'
            if item.is_not_applicable:
                item.status = 'NOT_APPLICABLE'
                item.custom_status = ''
            item.save()
            
            message = f'Item "{item.title}" marcado como {"não aplicável" if item.is_not_applicable else "aplicável"}!'
            
            return JsonResponse({
                'success': True,
                'is_not_applicable': item.is_not_applicable,
                'message': message
            })
        
        # Atualização de status
        new_status = request.POST.get('status')
        
        # Verificar se é status personalizado ou padrão
        if new_status.startswith('CUSTOM_'):
            # Status personalizado
            item.custom_status = new_status
            item.status = 'DONE'  # Marcar como concluído para cálculos
            item.is_not_applicable = False
        else:
            # Status padrão
            if new_status not in ['PENDING', 'DOING', 'DONE']:
                return JsonResponse({'success': False, 'error': 'Status inválido'})
            
            item.status = new_status
            item.custom_status = ''
            item.is_not_applicable = False
        
        item.save()
        
        # Recalcular porcentagem de conclusão
        completion_percentage = item.checklist.get_completion_percentage()
        is_completed = item.checklist.completed_at is not None
        
        return JsonResponse({
            'success': True,
            'new_status': new_status,
            'completion_percentage': completion_percentage,
            'is_completed': is_completed,
            'message': f'✅ Item "{item.title}" atualizado!'
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
@require_POST
def delete_checklist(request, checklist_id):
    """Excluir um checklist (apenas o próprio usuário ou superiores)"""
    from core.models import DailyChecklist
    from django.db import models
    
    user = request.user
    
    try:
        # Lógica de permissões por hierarquia
        user_hierarchy = getattr(user, 'hierarchy', None)
        
        if user_hierarchy in ['SUPERADMIN', 'ADMIN']:
            # SuperAdmins e Admins podem excluir qualquer checklist
            checklist = get_object_or_404(DailyChecklist, id=checklist_id)
        
        elif user_hierarchy in ['SUPERVISOR', 'ADMINISTRATIVO'] or user.is_staff:
            # Supervisores podem excluir:
            # 1. Seus próprios checklists
            # 2. Checklists que criaram
            # 3. Checklists de usuários do seu setor
            checklist_query = DailyChecklist.objects.filter(
                id=checklist_id
            ).filter(
                models.Q(user=user) |                          # Próprios checklists
                models.Q(created_by=user) |                    # Checklists que criou
                models.Q(user__sector=user.sector)             # Checklists do seu setor
            )
            
            if not checklist_query.exists():
                messages.error(request, 'Você não tem permissão para excluir este checklist.')
                return redirect('checklist_dashboard')
            
            checklist = checklist_query.first()
        
        else:
            # Usuários comuns só podem excluir seus próprios checklists
            checklist = get_object_or_404(DailyChecklist, id=checklist_id, user=user)
        
        # Armazenar informações antes de excluir
        checklist_title = checklist.title
        checklist_date = checklist.date.strftime('%d/%m/%Y')
        
        # Excluir o checklist (cascade irá excluir os itens automaticamente)
        checklist.delete()
        
        messages.success(request, f'✅ Checklist "{checklist_title}" de {checklist_date} excluído com sucesso!')
        return redirect('checklist_dashboard')
        
    except DailyChecklist.DoesNotExist:
        messages.error(request, 'Checklist não encontrado.')
        return redirect('checklist_dashboard')
    except Exception as e:
        messages.error(request, f'Erro ao excluir checklist: {str(e)}')
        return redirect('checklist_dashboard')


@login_required
@require_POST
def api_upload_checklist_item_evidence(request, item_id):
    """API para upload de evidências (imagens, vídeos, documentos) de itens de checklist"""
    from core.models import ChecklistItem, ChecklistItemEvidence
    from django.db import models
    
    try:
        item = ChecklistItem.objects.select_related('checklist', 'checklist__user', 'checklist__created_by').get(id=item_id)
    except ChecklistItem.DoesNotExist:
        return JsonResponse({'error': 'Item não encontrado'}, status=404)
    
    user = request.user
    checklist = item.checklist
    
    # Verificar permissão
    can_upload = (
        user == checklist.user or  # É o dono do checklist
        user == checklist.created_by or  # É quem criou/atribuiu
        user.hierarchy in ['SUPERADMIN', 'ADMIN'] or
        (user.hierarchy in ['SUPERVISOR', 'ADMINISTRATIVO'] and checklist.user.sector == user.sector)
    )
    
    if not can_upload:
        return JsonResponse({'error': 'Você não tem permissão para enviar evidências neste item.'}, status=403)
    
    files = request.FILES.getlist('files')
    
    if not files:
        return JsonResponse({'error': 'Nenhum arquivo enviado.'}, status=400)
    
    uploaded = []
    
    for file in files:
        # Determinar o tipo de evidência
        content_type = file.content_type.lower()
        
        if content_type.startswith('image/'):
            evidence_type = 'image'
        elif content_type.startswith('video/'):
            evidence_type = 'video'
        else:
            evidence_type = 'document'
        
        # Criar a evidência
        evidence = ChecklistItemEvidence.objects.create(
            item=item,
            evidence_type=evidence_type,
            file=file,
            original_filename=file.name,
            uploaded_by=user,
            order=item.evidences.count()
        )
        
        uploaded.append({
            'id': evidence.id,
            'type': evidence_type,
            'filename': evidence.original_filename or file.name,
            'url': evidence.file.url if evidence.file else None,
            'icon': evidence.get_file_icon()
        })
    
    return JsonResponse({
        'success': True,
        'message': f'{len(uploaded)} arquivo(s) enviado(s) com sucesso!',
        'files': uploaded
    })


@login_required
@require_POST
def api_delete_checklist_item_evidence(request, evidence_id):
    """API para deletar uma evidência de item de checklist"""
    from core.models import ChecklistItemEvidence
    from django.db import models
    
    try:
        evidence = ChecklistItemEvidence.objects.select_related(
            'item', 'item__checklist', 'item__checklist__user', 'item__checklist__created_by'
        ).get(id=evidence_id)
    except ChecklistItemEvidence.DoesNotExist:
        return JsonResponse({'error': 'Evidência não encontrada'}, status=404)
    
    user = request.user
    checklist = evidence.item.checklist
    
    # Verificar permissão
    can_delete = (
        user == checklist.user or  # É o dono do checklist
        user == checklist.created_by or  # É quem criou/atribuiu
        user == evidence.uploaded_by or  # É quem enviou a evidência
        user.hierarchy in ['SUPERADMIN', 'ADMIN'] or
        (user.hierarchy in ['SUPERVISOR', 'ADMINISTRATIVO'] and checklist.user.sector == user.sector)
    )
    
    if not can_delete:
        return JsonResponse({'error': 'Você não tem permissão para excluir esta evidência.'}, status=403)
    
    # Deletar o arquivo do S3/storage
    if evidence.file:
        evidence.file.delete(save=False)
    
    # Deletar o registro
    evidence.delete()
    
    return JsonResponse({
        'success': True,
        'message': 'Evidência excluída com sucesso!'
    })


# ===== ATIVIDADES/TAREFAS VIEWS =====
@login_required
def tasks_dashboard_view(request):
    """Dashboard de tarefas do usuário"""
    from core.models import TaskActivity
    from django.utils import timezone
    from datetime import date, timedelta
    
    user = request.user
    today = timezone.now()
    
    # Tarefas do usuário
    user_tasks = TaskActivity.objects.filter(assigned_to=user)
    
    # Separar por status
    pending_tasks = user_tasks.filter(status='PENDING').order_by('due_date')
    doing_tasks = user_tasks.filter(status='DOING').order_by('due_date')
    done_tasks = user_tasks.filter(status='DONE').order_by('-completed_at')[:10]
    
    # Tarefas em atraso
    overdue_tasks = user_tasks.filter(
        status__in=['PENDING', 'DOING'],
        due_date__lt=today
    ).order_by('due_date')
    
    # Estatísticas
    total_tasks = user_tasks.count()
    completed_tasks = user_tasks.filter(status='DONE').count()
    completion_rate = 0
    if total_tasks > 0:
        completion_rate = round((completed_tasks / total_tasks) * 100)
    
    context = {
        'pending_tasks': pending_tasks,
        'doing_tasks': doing_tasks,
        'done_tasks': done_tasks,
        'overdue_tasks': overdue_tasks,
        'completion_rate': completion_rate,
        'total_tasks': total_tasks,
        'completed_tasks': completed_tasks,
    }
    return render(request, 'tasks/dashboard.html', context)


@login_required
@require_POST
def update_task_status(request, task_id):
    """Atualizar status de uma tarefa"""
    from core.models import TaskActivity
    
    try:
        task = get_object_or_404(
            TaskActivity,
            id=task_id,
            assigned_to=request.user
        )
        
        new_status = request.POST.get('status')
        if new_status not in ['PENDING', 'DOING', 'DONE']:
            return JsonResponse({'success': False, 'error': 'Status inválido'})
        
        task.status = new_status
        task.save()
        
        return JsonResponse({
            'success': True,
            'new_status': new_status,
            'message': f'Tarefa "{task.title}" atualizada!'
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


# ===== VIEWS ADMINISTRATIVAS PARA SUPERVISORES =====
@login_required
def manage_checklists_view(request):
    """Gerenciar templates de checklist"""
    # Supervisores podem ver todos os templates, usuários comuns só os seus próprios
    is_supervisor = request.user.hierarchy in ['SUPERVISOR', 'ADMIN', 'SUPERADMIN', 'ADMINISTRATIVO'] or request.user.is_staff
    """Gerenciar templates de checklist (apenas supervisores)"""
    from core.models import ChecklistTemplate, ChecklistTemplateItem
    from django.db.models import Q
    
    if request.method == 'POST' and is_supervisor:
        # Criar novo template de checklist (apenas supervisores)
        title = request.POST.get('title')
        description = request.POST.get('description')
        use_custom_statuses = request.POST.get('use_custom_statuses') == 'on'
        
        if title:
            template = ChecklistTemplate.objects.create(
                title=title,
                description=description,
                created_by=request.user,
                use_custom_statuses=use_custom_statuses
            )
            
            # Processar status personalizados se habilitado
            if use_custom_statuses:
                custom_statuses = []
                status_counter = 1
                
                while True:
                    label = request.POST.get(f'custom_status_label_{status_counter}')
                    if not label:
                        break
                    
                    icon = request.POST.get(f'custom_status_icon_{status_counter}', 'fas fa-star')
                    color = request.POST.get(f'custom_status_color_{status_counter}', 'blue')
                    
                    custom_statuses.append({
                        'value': f'CUSTOM_{status_counter}',
                        'label': label,
                        'icon': icon,
                        'color': color
                    })
                    
                    status_counter += 1
                
                template.custom_statuses = custom_statuses
                template.save()
            
            # Adicionar itens se fornecidos
            items = request.POST.getlist('items[]')
            for i, item_title in enumerate(items):
                if item_title.strip():
                    ChecklistTemplateItem.objects.create(
                        template=template,
                        title=item_title.strip(),
                        order=i + 1
                    )
            
            messages.success(request, f'✅ Template "{title}" criado com sucesso!')
            return redirect('manage_checklists')
        else:
            messages.error(request, 'Nome do template é obrigatório!')
    
    # SUPERADMIN vê todos os templates
    # Outros usuários veem templates de todos os setores onde participam
    if request.user.hierarchy == 'SUPERADMIN' or request.user.is_staff:
        templates = ChecklistTemplate.objects.all().prefetch_related('items').order_by('-created_at')
    else:
        # Usuários veem templates de todos os setores onde participam
        user_sectors = []
        if request.user.sector:
            user_sectors.append(request.user.sector)
        user_sectors.extend(request.user.sectors.all())
        
        templates = ChecklistTemplate.objects.filter(
            Q(created_by=request.user) | 
            Q(created_by__sector__in=user_sectors) |
            Q(created_by__sectors__in=user_sectors)
        ).distinct().prefetch_related('items').order_by('-created_at')
    
    context = {
        'templates': templates,
        'is_supervisor': is_supervisor,
    }
    return render(request, 'admin_panel/manage_checklists.html', context)


def create_recurring_checklists(parent_checklist, template=None):
    """Cria checklists recorrentes baseado no tipo de repetição"""
    from core.models import ChecklistItem, DailyChecklist
    from datetime import date, timedelta
    
    created_count = 0
    start_date = parent_checklist.date + timedelta(days=1)  # Começa no dia seguinte
    
    if parent_checklist.repeat_type == 'DAILY':
        # Repetir todos os dias por 30 dias
        end_date = parent_checklist.repeat_end_date or (start_date + timedelta(days=29))
        current_date = start_date
        
        while current_date <= end_date:
            recurring_checklist = DailyChecklist.objects.create(
                user=parent_checklist.user,
                template=parent_checklist.template,
                title=parent_checklist.title,
                date=current_date,
                repeat_type='NONE',  # Instâncias não repetem
                is_recurring_instance=True,
                parent_checklist=parent_checklist,
                created_by=parent_checklist.created_by
            )
            
            # Copiar itens
            for item in parent_checklist.items.all():
                ChecklistItem.objects.create(
                    checklist=recurring_checklist,
                    title=item.title,
                    description=item.description,
                    order=item.order,
                    is_required=item.is_required
                )
            
            created_count += 1
            current_date += timedelta(days=1)
    
    elif parent_checklist.repeat_type == 'WEEKDAYS':
        # Repetir por 30 dias exceto finais de semana
        end_date = parent_checklist.repeat_end_date or (start_date + timedelta(days=42))  # Mais dias para compensar fins de semana
        current_date = start_date
        weekdays_created = 0
        
        while current_date <= end_date and weekdays_created < 30:
            # Pular fins de semana (5=sábado, 6=domingo)
            if current_date.weekday() < 5:
                recurring_checklist = DailyChecklist.objects.create(
                    user=parent_checklist.user,
                    template=parent_checklist.template,
                    title=parent_checklist.title,
                    date=current_date,
                    repeat_type='NONE',
                    is_recurring_instance=True,
                    parent_checklist=parent_checklist,
                    created_by=parent_checklist.created_by
                )
                
                # Copiar itens
                for item in parent_checklist.items.all():
                    ChecklistItem.objects.create(
                        checklist=recurring_checklist,
                        title=item.title,
                        description=item.description,
                        order=item.order,
                        is_required=item.is_required
                    )
                
                created_count += 1
                weekdays_created += 1
            
            current_date += timedelta(days=1)
    
    elif parent_checklist.repeat_type == 'CUSTOM_DAYS':
        # Repetir em dias específicos da semana
        end_date = parent_checklist.repeat_end_date or (start_date + timedelta(days=90))  # 3 meses por padrão
        current_date = start_date
        
        # Converter dias selecionados para inteiros
        selected_days = [int(day) for day in parent_checklist.repeat_days if day.isdigit()]
        
        while current_date <= end_date:
            if current_date.weekday() in selected_days:
                recurring_checklist = DailyChecklist.objects.create(
                    user=parent_checklist.user,
                    template=parent_checklist.template,
                    title=parent_checklist.title,
                    date=current_date,
                    repeat_type='NONE',
                    is_recurring_instance=True,
                    parent_checklist=parent_checklist,
                    created_by=parent_checklist.created_by
                )
                
                # Copiar itens
                for item in parent_checklist.items.all():
                    ChecklistItem.objects.create(
                        checklist=recurring_checklist,
                        title=item.title,
                        description=item.description,
                        order=item.order,
                        is_required=item.is_required
                    )
                
                created_count += 1
            
            current_date += timedelta(days=1)
    
    return created_count


@login_required
def create_daily_checklist(request):
    """Criar checklist diário para usuários"""
    from core.models import ChecklistTemplate, DailyChecklist, ChecklistItem
    from datetime import date, timedelta
    
    user = request.user
    is_supervisor = user.hierarchy in ['SUPERVISOR', 'ADMIN', 'SUPERADMIN', 'ADMINISTRATIVO'] or user.is_staff
    # Usuários padrões também podem criar checklists para outros usuários padrões do mesmo setor
    can_create_checklist = is_supervisor or (user.hierarchy == 'PADRAO' and (user.sector or user.sectors.exists()))
    
    if not can_create_checklist:
        messages.error(request, 'Você não tem permissão para acessar esta página.')
        return redirect('dashboard')
    
    if request.method == 'POST':
        # Verificar se é criação baseada em template ou customizada
        creation_type = request.POST.get('creation_type', 'template')
        user_ids = request.POST.getlist('user_ids')
        checklist_date = request.POST.get('date', date.today().isoformat())
        repeat_type = request.POST.get('repeat_type', 'NONE')
        repeat_days = request.POST.getlist('repeat_days[]')  # Dias da semana selecionados
        repeat_end_date = request.POST.get('repeat_end_date')
        
        if not user_ids:
            messages.error(request, 'Selecione pelo menos um usuário!')
            return redirect('create_daily_checklist')
        
        try:
            target_date = date.fromisoformat(checklist_date)
            created_count = 0
            
            if creation_type == 'template':
                template_id = request.POST.get('template_id')
                if not template_id:
                    messages.error(request, 'Selecione um template!')
                    return redirect('create_daily_checklist')
                
                template = ChecklistTemplate.objects.get(id=template_id)
                
                for user_id in user_ids:
                    user = User.objects.get(id=user_id)
                    
                    # Verificar se já existe checklist para esta data
                    if DailyChecklist.objects.filter(user=user, template=template, date=target_date).exists():
                        continue
                    
                    # Criar checklist principal
                    daily_checklist = DailyChecklist.objects.create(
                        user=user,
                        template=template,
                        title=template.title,
                        date=target_date,
                        repeat_daily=(repeat_type != 'NONE'),
                        repeat_type=repeat_type,
                        repeat_days=repeat_days if repeat_type == 'CUSTOM_DAYS' else [],
                        repeat_end_date=date.fromisoformat(repeat_end_date) if repeat_end_date else None,
                        created_by=request.user
                    )
                    
                    # Criar itens baseados no template
                    for template_item in template.items.all():
                        ChecklistItem.objects.create(
                            checklist=daily_checklist,
                            title=template_item.title,
                            description=template_item.description,
                            order=template_item.order
                        )
                    
                    # Criar checklists recorrentes se necessário
                    if repeat_type != 'NONE':
                        created_count += create_recurring_checklists(daily_checklist, template)
                    else:
                        created_count += 1
            
            else:  # creation_type == 'custom'
                custom_title = request.POST.get('custom_title')
                custom_items = request.POST.getlist('custom_items[]')
                
                if not custom_title or not custom_items:
                    messages.error(request, 'Título e itens são obrigatórios para checklist customizado!')
                    return redirect('create_daily_checklist')
                
                for user_id in user_ids:
                    user = User.objects.get(id=user_id)
                    
                    # Verificar se já existe checklist com este título para esta data
                    if DailyChecklist.objects.filter(user=user, title=custom_title, date=target_date).exists():
                        continue
                    
                    # Criar checklist customizado
                    daily_checklist = DailyChecklist.objects.create(
                        user=user,
                        title=custom_title,
                        date=target_date,
                        repeat_daily=(repeat_type != 'NONE'),
                        repeat_type=repeat_type,
                        repeat_days=repeat_days if repeat_type == 'CUSTOM_DAYS' else [],
                        repeat_end_date=date.fromisoformat(repeat_end_date) if repeat_end_date else None,
                        created_by=request.user
                    )
                    
                    # Criar itens customizados
                    for i, item_title in enumerate(custom_items):
                        if item_title.strip():
                            ChecklistItem.objects.create(
                                checklist=daily_checklist,
                                title=item_title.strip(),
                                order=i + 1
                            )
                    
                    # Criar checklists recorrentes se necessário
                    if repeat_type != 'NONE':
                        created_count += create_recurring_checklists(daily_checklist)
                    else:
                        created_count += 1
            
            # Implementar lógica de repetição
            if repeat_type != 'none' and created_count > 0:
                repeat_count = 0
                
                # Mapear dias da semana
                weekday_map = {
                    'monday': 0, 'tuesday': 1, 'wednesday': 2, 'thursday': 3,
                    'friday': 4, 'saturday': 5, 'sunday': 6
                }
                
                # Gerar datas para repetição
                future_dates = []
                current_date = target_date + timedelta(days=1)
                end_date = target_date + timedelta(days=30)
                
                if repeat_type == 'daily':
                    # Repetir todos os dias
                    while current_date <= end_date:
                        future_dates.append(current_date)
                        current_date += timedelta(days=1)
                elif repeat_type in weekday_map:
                    # Repetir em dia específico da semana
                    target_weekday = weekday_map[repeat_type]
                    while current_date <= end_date:
                        if current_date.weekday() == target_weekday:
                            future_dates.append(current_date)
                        current_date += timedelta(days=1)
                
                # Criar checklists para as datas futuras
                for future_date in future_dates:
                    for user_id in user_ids:
                        user = User.objects.get(id=user_id)
                        
                        # Verificar se já existe checklist para esta data
                        if creation_type == 'template':
                            if DailyChecklist.objects.filter(user=user, template=template, date=future_date).exists():
                                continue
                            
                            # Criar checklist repetido
                            future_checklist = DailyChecklist.objects.create(
                                user=user,
                                template=template,
                                title=template.title,
                                date=future_date,
                                repeat_daily=True,
                                created_by=request.user
                            )
                            
                            # Criar itens baseados no template
                            for template_item in template.items.all():
                                ChecklistItem.objects.create(
                                    checklist=future_checklist,
                                    title=template_item.title,
                                    description=template_item.description,
                                    order=template_item.order
                                )
                        else:
                            if DailyChecklist.objects.filter(user=user, title=custom_title, date=future_date).exists():
                                continue
                                
                            # Criar checklist customizado repetido
                            future_checklist = DailyChecklist.objects.create(
                                user=user,
                                title=custom_title,
                                date=future_date,
                                repeat_daily=True,
                                created_by=request.user
                            )
                            
                            # Criar itens customizados
                            for i, item_title in enumerate(custom_items):
                                if item_title.strip():
                                    ChecklistItem.objects.create(
                                        checklist=future_checklist,
                                        title=item_title.strip(),
                                        order=i + 1
                                    )
                        
                        repeat_count += 1
                
                # Mensagem específica para o tipo de repetição
                repeat_labels = {
                    'daily': 'repetição diária',
                    'monday': 'repetição nas segundas-feiras',
                    'tuesday': 'repetição nas terças-feiras', 
                    'wednesday': 'repetição nas quartas-feiras',
                    'thursday': 'repetição nas quintas-feiras',
                    'friday': 'repetição nas sextas-feiras',
                    'saturday': 'repetição nos sábados',
                    'sunday': 'repetição nos domingos'
                }
                repeat_label = repeat_labels.get(repeat_type, 'repetição')
                messages.success(request, f'{created_count} checklist(s) criado(s) com sucesso! {repeat_count} checklist(s) adicionais criados para {repeat_label}.')
            else:
                messages.success(request, f'{created_count} checklist(s) criado(s) com sucesso!')
            
            return redirect('manage_checklists')
            
        except Exception as e:
            messages.error(request, f'Erro ao criar checklists: {str(e)}')
    
    # GET - mostrar formulário
    templates = ChecklistTemplate.objects.all().order_by('title')
    
    # Filtrar usuários por setor para SUPERVISOR, ADMINISTRATIVO e PADRAO
    if user.hierarchy in ['SUPERVISOR', 'ADMINISTRATIVO']:
        # Pegar setores do usuário atual (tanto ManyToMany quanto ForeignKey)
        user_sectors = list(user.sectors.all())
        if user.sector:
            user_sectors.append(user.sector)
        
        if user_sectors:
            # Filtrar usuários que pertencem aos mesmos setores (considerando ambos os campos)
            from django.db.models import Q
            users = User.objects.filter(
                is_staff=False
            ).filter(
                Q(sectors__in=user_sectors) | Q(sector__in=user_sectors)
            ).exclude(
                id=user.id  # Excluir o próprio usuário
            ).distinct().order_by('first_name', 'username')
        else:
            # Se não tem setor definido, não pode criar para ninguém
            users = User.objects.none()
    elif user.hierarchy == 'PADRAO':
        # Usuários padrões só podem criar para outros usuários padrões do mesmo setor
        user_sectors = list(user.sectors.all())
        if user.sector:
            user_sectors.append(user.sector)
        
        if user_sectors:
            from django.db.models import Q
            # Filtrar apenas usuários PADRAO que pertencem aos mesmos setores
            users = User.objects.filter(
                is_staff=False,
                hierarchy='PADRAO'
            ).filter(
                Q(sectors__in=user_sectors) | Q(sector__in=user_sectors)
            ).exclude(
                id=user.id  # Excluir o próprio usuário
            ).distinct().order_by('first_name', 'username')
        else:
            users = User.objects.none()
    else:
        # ADMIN e SUPERADMIN podem criar para todos (exceto eles mesmos)
        users = User.objects.filter(is_staff=False).exclude(id=user.id).order_by('first_name', 'username')
    
    context = {
        'templates': templates,
        'users': users,
        'today': date.today(),
        'is_supervisor': is_supervisor,
    }
    return render(request, 'admin_panel/create_daily_checklist.html', context)


@login_required
def manage_tasks_view(request):
    """Gerenciar tarefas/atividades (apenas supervisores)"""
    if not (request.user.hierarchy in ['SUPERVISOR', 'ADMIN', 'SUPERADMIN', 'ADMINISTRATIVO'] or request.user.is_staff):
        messages.error(request, 'Você não tem permissão para acessar esta página.')
        return redirect('dashboard')
    """Gerenciar tarefas/atividades (apenas supervisores)"""
    from core.models import TaskActivity, TaskCategory
    from django.db.models import Q
    
    # Filtros
    status_filter = request.GET.get('status', '')
    user_filter = request.GET.get('user', '')
    category_filter = request.GET.get('category', '')
    sector_filter = request.GET.get('sector', '')
    
    # Segmentação por setor
    if request.user.hierarchy == 'SUPERADMIN':
        # SUPERADMIN vê todas as tarefas
        tasks = TaskActivity.objects.all().select_related('assigned_to', 'created_by', 'category')
        all_tasks = TaskActivity.objects.all()
        
        # SUPERADMIN pode filtrar por setor específico
        if sector_filter:
            tasks = tasks.filter(
                Q(assigned_to__sector_id=sector_filter) |
                Q(assigned_to__sectors__id=sector_filter)
            ).distinct()
            all_tasks = all_tasks.filter(
                Q(assigned_to__sector_id=sector_filter) |
                Q(assigned_to__sectors__id=sector_filter)
            ).distinct()
    else:
        # Outros usuários só veem tarefas do seu setor ou que criaram
        user_sectors = []
        if request.user.sector:
            user_sectors.append(request.user.sector)
        user_sectors.extend(request.user.sectors.all())
        
        # Filtrar tarefas por setor do usuário atribuído ou criador
        tasks = TaskActivity.objects.filter(
            Q(assigned_to__sector__in=user_sectors) |
            Q(assigned_to__sectors__in=user_sectors) |
            Q(created_by=request.user)
        ).distinct().select_related('assigned_to', 'created_by', 'category')
        
        all_tasks = TaskActivity.objects.filter(
            Q(assigned_to__sector__in=user_sectors) |
            Q(assigned_to__sectors__in=user_sectors) |
            Q(created_by=request.user)
        ).distinct()
    
    # Aplicar filtros
    if status_filter:
        tasks = tasks.filter(status=status_filter)
    
    if user_filter:
        tasks = tasks.filter(assigned_to_id=user_filter)
        
    if category_filter:
        tasks = tasks.filter(category_id=category_filter)
    
    tasks = tasks.order_by('-created_at')
    
    # Para estatísticas
    stats = {
        'total': all_tasks.count(),
        'pending': all_tasks.filter(status='PENDING').count(),
        'doing': all_tasks.filter(status='DOING').count(),
        'done': all_tasks.filter(status='DONE').count(),
    }
    
    # Categorias disponíveis para o usuário
    if request.user.hierarchy == 'SUPERADMIN':
        categories = TaskCategory.objects.filter(is_active=True)
    else:
        user_sectors = []
        if request.user.sector:
            user_sectors.append(request.user.sector)
        user_sectors.extend(request.user.sectors.all())
        
        categories = TaskCategory.objects.filter(
            is_active=True,
            sectors__in=user_sectors
        ).distinct()
        
        if not categories.exists():
            categories = TaskCategory.objects.filter(is_active=True, sectors__isnull=True)
    
    # Usuários disponíveis para filtro
    if request.user.hierarchy == 'SUPERADMIN':
        users = User.objects.filter(is_staff=False).order_by('first_name', 'username')
    else:
        # Apenas usuários do mesmo setor
        user_sectors = []
        if request.user.sector:
            user_sectors.append(request.user.sector)
        user_sectors.extend(request.user.sectors.all())
        
        users = User.objects.filter(
            Q(sector__in=user_sectors) |
            Q(sectors__in=user_sectors)
        ).distinct().order_by('first_name', 'username')
    
    # Setores disponíveis (apenas para SUPERADMIN)
    from users.models import Sector
    sectors = Sector.objects.all().order_by('name') if request.user.hierarchy == 'SUPERADMIN' else None
    
    context = {
        'tasks': tasks,
        'users': users,
        'categories': categories,
        'sectors': sectors,
        'stats': stats,
        'status_filter': status_filter,
        'user_filter': user_filter,
        'category_filter': category_filter,
        'sector_filter': sector_filter,
        'STATUS_CHOICES': TaskActivity.STATUS_CHOICES,
        'is_superadmin': request.user.hierarchy == 'SUPERADMIN',
    }
    return render(request, 'admin_panel/manage_tasks.html', context)


@login_required
def task_detail_view(request, task_id):
    """Visualizar detalhes da tarefa com chat"""
    from core.models import TaskActivity, TaskMessage, TaskAttachment, TaskSubtask
    
    task = get_object_or_404(TaskActivity, id=task_id)
    
    # Mensagens da tarefa
    task_messages = task.messages.select_related('user').order_by('created_at')
    
    # Anexos da tarefa
    task_attachments = task.attachments.select_related('uploaded_by').order_by('-uploaded_at')
    
    # Subtarefas
    subtasks = task.subtasks.select_related('completed_by', 'created_by').order_by('order', 'created_at')
    
    # Calcular progresso das subtarefas
    total_subtasks = subtasks.count()
    completed_subtasks = subtasks.filter(is_completed=True).count()
    subtask_progress = round((completed_subtasks / total_subtasks * 100)) if total_subtasks > 0 else 0
    
    # Marcar mensagens como lidas para o usuário atual
    task_messages.filter(is_read=False).exclude(user=request.user).update(is_read=True)
    
    context = {
        'task': task,
        'messages': task_messages,
        'attachments': task_attachments,
        'subtasks': subtasks,
        'subtask_progress': subtask_progress,
        'total_subtasks': total_subtasks,
        'completed_subtasks': completed_subtasks,
        'can_manage': task.can_be_managed_by(request.user),
    }
    
    return render(request, 'tasks/task_detail.html', context)


@login_required
def task_messages_api(request, task_id):
    """API para buscar mensagens da tarefa"""
    from core.models import TaskActivity, TaskMessage
    
    task = get_object_or_404(TaskActivity, id=task_id)
    
    messages = task.messages.select_related('user').order_by('created_at')
    
    messages_data = []
    for message in messages:
        messages_data.append({
            'id': message.id,
            'user': {
                'name': message.user.get_full_name(),
                'id': message.user.id,
            },
            'message': message.message,
            'message_type': message.message_type,
            'created_at': message.created_at.isoformat(),
            'is_own': message.user == request.user,
        })
    
    return JsonResponse({
        'success': True,
        'messages': messages_data
    })


@login_required
@require_POST
def send_task_message(request, task_id):
    """Enviar mensagem para o chat da tarefa"""
    from core.models import TaskActivity, TaskMessage
    
    task = get_object_or_404(TaskActivity, id=task_id)
    
    # Permitir qualquer usuário logado enviar mensagem
    message_text = request.POST.get('message', '').strip()
    if not message_text:
        return JsonResponse({'success': False, 'error': 'Mensagem não pode estar vazia'})
    
    # Criar mensagem
    message = TaskMessage.objects.create(
        task=task,
        user=request.user,
        message=message_text,
        message_type='MESSAGE'
    )
    
    return JsonResponse({
        'success': True,
        'message': {
            'id': message.id,
            'user': {
                'name': request.user.get_full_name(),
                'id': request.user.id,
            },
            'message': message.message,
            'message_type': message.message_type,
            'created_at': message.created_at.isoformat(),
            'is_own': True,
        }
    })


@login_required
@require_POST
def add_task_attachment(request, task_id):
    """Adicionar anexo a uma tarefa existente"""
    from core.models import TaskActivity, TaskAttachment
    
    task = get_object_or_404(TaskActivity, id=task_id)
    
    # Permitir qualquer usuário logado adicionar anexo
    # Verificar se há arquivo
    if 'file' not in request.FILES:
        return JsonResponse({'success': False, 'error': 'Nenhum arquivo enviado'})
    
    file = request.FILES['file']
    
    # Validar tamanho (máximo 50MB)
    if file.size > 50 * 1024 * 1024:
        return JsonResponse({'success': False, 'error': 'Arquivo muito grande. Máximo 50MB.'})
    
    try:
        # Criar anexo
        attachment = TaskAttachment.objects.create(
            task=task,
            file=file,
            uploaded_by=request.user,
            file_name=file.name,
            file_size=file.size
        )
        
        return JsonResponse({
            'success': True,
            'attachment': {
                'id': attachment.id,
                'file_name': attachment.file_name,
                'file_size': attachment.file_size_formatted,
                'file_url': attachment.file.url,
                'uploaded_by': request.user.get_full_name(),
                'uploaded_at': attachment.uploaded_at.strftime('%d/%m/%Y %H:%M'),
                'is_image': attachment.is_image,
                'is_video': attachment.is_video,
                'is_document': attachment.is_document,
            }
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': f'Erro ao salvar arquivo: {str(e)}'})


@login_required
@require_POST
def delete_task_attachment(request, attachment_id):
    """Deletar anexo de uma tarefa"""
    from core.models import TaskAttachment
    
    attachment = get_object_or_404(TaskAttachment, id=attachment_id)
    task = attachment.task
    
    # Verificar permissão (apenas quem enviou ou quem pode gerenciar a tarefa)
    if not (attachment.uploaded_by == request.user or 
            task.can_be_managed_by(request.user)):
        return JsonResponse({'success': False, 'error': 'Sem permissão'}, status=403)
    
    try:
        # Deletar arquivo físico
        if attachment.file:
            attachment.file.delete()
        
        # Deletar registro
        attachment.delete()
        
        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'success': False, 'error': f'Erro ao deletar arquivo: {str(e)}'})


# ===== SUBTAREFAS API =====
@login_required
def get_task_subtasks(request, task_id):
    """API para buscar subtarefas de uma tarefa"""
    from core.models import TaskActivity, TaskSubtask
    
    task = get_object_or_404(TaskActivity, id=task_id)
    
    subtasks = task.subtasks.select_related('completed_by', 'created_by').order_by('order', 'created_at')
    
    subtasks_data = []
    for subtask in subtasks:
        subtasks_data.append({
            'id': subtask.id,
            'title': subtask.title,
            'is_completed': subtask.is_completed,
            'completed_at': subtask.completed_at.strftime('%d/%m/%Y %H:%M') if subtask.completed_at else None,
            'completed_by': subtask.completed_by.get_full_name() if subtask.completed_by else None,
            'created_by': subtask.created_by.get_full_name(),
            'created_at': subtask.created_at.strftime('%d/%m/%Y %H:%M'),
        })
    
    # Calcular progresso
    total = len(subtasks_data)
    completed = sum(1 for s in subtasks_data if s['is_completed'])
    progress = round((completed / total * 100)) if total > 0 else 0
    
    return JsonResponse({
        'success': True,
        'subtasks': subtasks_data,
        'progress': progress,
        'total': total,
        'completed': completed
    })


@login_required
@require_POST
def add_task_subtask(request, task_id):
    """Adicionar subtarefa a uma tarefa"""
    from core.models import TaskActivity, TaskSubtask
    
    task = get_object_or_404(TaskActivity, id=task_id)
    
    # Verificar permissão (criador, responsável ou gerenciador)
    if not (request.user == task.created_by or 
            request.user == task.assigned_to or 
            task.can_be_managed_by(request.user)):
        return JsonResponse({'success': False, 'error': 'Sem permissão'}, status=403)
    
    title = request.POST.get('title', '').strip()
    if not title:
        return JsonResponse({'success': False, 'error': 'Título da subtarefa é obrigatório'})
    
    # Obter próxima ordem
    max_order = task.subtasks.aggregate(max_order=Max('order'))['max_order'] or 0
    
    subtask = TaskSubtask.objects.create(
        task=task,
        title=title,
        order=max_order + 1,
        created_by=request.user
    )
    
    return JsonResponse({
        'success': True,
        'subtask': {
            'id': subtask.id,
            'title': subtask.title,
            'is_completed': subtask.is_completed,
            'created_by': request.user.get_full_name(),
            'created_at': subtask.created_at.strftime('%d/%m/%Y %H:%M'),
        }
    })


@login_required
@require_POST
def toggle_task_subtask(request, subtask_id):
    """Alternar status de conclusão de uma subtarefa"""
    from core.models import TaskSubtask
    
    subtask = get_object_or_404(TaskSubtask, id=subtask_id)
    task = subtask.task
    
    # Verificar permissão (criador, responsável ou gerenciador)
    if not (request.user == task.created_by or 
            request.user == task.assigned_to or 
            task.can_be_managed_by(request.user)):
        return JsonResponse({'success': False, 'error': 'Sem permissão'}, status=403)
    
    if subtask.is_completed:
        subtask.mark_incomplete()
    else:
        subtask.mark_completed(request.user)
    
    # Calcular novo progresso
    total = task.subtasks.count()
    completed = task.subtasks.filter(is_completed=True).count()
    progress = round((completed / total * 100)) if total > 0 else 0
    
    return JsonResponse({
        'success': True,
        'subtask': {
            'id': subtask.id,
            'is_completed': subtask.is_completed,
            'completed_at': subtask.completed_at.strftime('%d/%m/%Y %H:%M') if subtask.completed_at else None,
            'completed_by': subtask.completed_by.get_full_name() if subtask.completed_by else None,
        },
        'progress': progress,
        'total': total,
        'completed': completed
    })


@login_required
@require_POST
def delete_task_subtask(request, subtask_id):
    """Deletar uma subtarefa"""
    from core.models import TaskSubtask
    
    subtask = get_object_or_404(TaskSubtask, id=subtask_id)
    task = subtask.task
    
    # Verificar permissão (criador da subtarefa, criador da tarefa ou gerenciador)
    if not (request.user == subtask.created_by or 
            request.user == task.created_by or 
            task.can_be_managed_by(request.user)):
        return JsonResponse({'success': False, 'error': 'Sem permissão'}, status=403)
    
    subtask.delete()
    
    # Calcular novo progresso
    total = task.subtasks.count()
    completed = task.subtasks.filter(is_completed=True).count()
    progress = round((completed / total * 100)) if total > 0 else 0
    
    return JsonResponse({
        'success': True,
        'progress': progress,
        'total': total,
        'completed': completed
    })


@login_required
def create_task_view(request):
    """Criar nova tarefa/atividade"""
    if not (request.user.hierarchy in ['SUPERVISOR', 'ADMIN', 'SUPERADMIN', 'ADMINISTRATIVO'] or request.user.is_staff):
        messages.error(request, 'Você não tem permissão para acessar esta página.')
        return redirect('dashboard')
    """Criar nova tarefa/atividade"""
    from core.models import TaskActivity, TaskAttachment
    from datetime import datetime
    
    if request.method == 'POST':
        title = request.POST.get('title')
        description = request.POST.get('description')
        assigned_to_id = request.POST.get('assigned_to')
        due_date = request.POST.get('due_date')
        priority = request.POST.get('priority', 'MEDIUM')
        
        if not title or not assigned_to_id:
            messages.error(request, 'Título e usuário são obrigatórios!')
            return redirect('create_task')
        
        try:
            assigned_user = User.objects.get(id=assigned_to_id)
            
            task = TaskActivity.objects.create(
                title=title,
                description=description,
                assigned_to=assigned_user,
                created_by=request.user,
                priority=priority,
                due_date=datetime.fromisoformat(due_date) if due_date else None
            )
            
            # Processar anexos
            attachments = request.FILES.getlist('attachments')
            for attachment in attachments:
                TaskAttachment.objects.create(
                    task=task,
                    file=attachment,
                    uploaded_by=request.user,
                    file_name=attachment.name,
                    file_size=attachment.size
                )
            
            messages.success(request, f'Tarefa "{title}" criada com sucesso!')
            return redirect('manage_tasks')
            
        except Exception as e:
            messages.error(request, f'Erro ao criar tarefa: {str(e)}')
    
    # Apenas SUPERADMINs podem atribuir tarefas para qualquer pessoa
    if request.user.hierarchy == 'SUPERADMIN':
        # SUPERADMIN pode criar para todos (exceto ele mesmo)
        users_qs = User.objects.filter(is_staff=False).exclude(id=request.user.id).order_by('first_name', 'username')
    else:
        # Outros usuários só podem criar para pessoas do seu setor
        # Pegar setores do usuário atual (tanto ManyToMany quanto ForeignKey)
        user_sectors = list(request.user.sectors.all())
        if request.user.sector:
            user_sectors.append(request.user.sector)
        
        if user_sectors:
            # Filtrar usuários que pertencem aos mesmos setores (considerando ambos os campos)
            from django.db.models import Q
            users_qs = User.objects.filter(
                is_staff=False
            ).filter(
                Q(sectors__in=user_sectors) | Q(sector__in=user_sectors)
            ).exclude(
                id=request.user.id  # Excluir o próprio usuário
            ).distinct().order_by('first_name', 'username')
        else:
            # Se não tem setor definido, não pode criar para ninguém
            users_qs = User.objects.none()
    
    # Permitir que o usuário atribua a tarefa para si mesmo (sempre no topo da lista)
    users = [request.user] + list(users_qs)
    
    context = {
        'users': users,
        'PRIORITY_CHOICES': TaskActivity.PRIORITY_CHOICES,
    }
    return render(request, 'admin_panel/create_task.html', context)


@login_required
def checklist_template_detail(request, template_id):
    """Ver detalhes de um template de checklist"""
    from core.models import ChecklistTemplate
    
    template = get_object_or_404(ChecklistTemplate, id=template_id)
    
    # Verificar se o usuário pode ver este template
    is_supervisor = request.user.hierarchy in ['SUPERVISOR', 'ADMIN', 'SUPERADMIN', 'ADMINISTRATIVO'] or request.user.is_staff
    is_superadmin = request.user.hierarchy == 'SUPERADMIN' or request.user.is_staff
    
    # SUPERADMIN pode ver todos os templates
    # Supervisores podem ver templates do mesmo setor
    # Usuários comuns apenas os próprios
    can_view = (
        is_superadmin or 
        template.created_by == request.user or
        (is_supervisor and template.created_by.sector == request.user.sector)
    )
    
    if not can_view:
        messages.error(request, 'Você não tem permissão para ver este template.')
        return redirect('manage_checklists')
    
    # Buscar todos os checklists criados a partir deste template
    from core.models import DailyChecklist
    related_checklists = DailyChecklist.objects.filter(template=template).select_related('user').order_by('-date')[:10]
    
    context = {
        'template': template,
        'related_checklists': related_checklists,
        'is_supervisor': is_supervisor,
        'can_edit': is_supervisor or template.created_by == request.user,
    }
    
    return render(request, 'admin_panel/checklist_template_detail.html', context)


@login_required
def edit_checklist_template(request, template_id):
    """Editar template de checklist"""
    from core.models import ChecklistTemplate, ChecklistTemplateItem
    
    template = get_object_or_404(ChecklistTemplate, id=template_id)
    
    # Verificar se o usuário pode editar este template
    is_supervisor = request.user.hierarchy in ['SUPERVISOR', 'ADMIN', 'SUPERADMIN', 'ADMINISTRATIVO'] or request.user.is_staff
    is_superadmin = request.user.hierarchy == 'SUPERADMIN' or request.user.is_staff
    
    # SUPERADMIN pode editar todos os templates
    # Supervisores podem editar templates do mesmo setor
    # Usuários comuns apenas os próprios
    can_edit = (
        is_superadmin or 
        template.created_by == request.user or
        (is_supervisor and template.created_by.sector == request.user.sector)
    )
    
    if not can_edit:
        messages.error(request, 'Você não tem permissão para editar este template.')
        return redirect('manage_checklists')
    
    if request.method == 'POST':
        try:
            # Atualizar dados básicos do template
            template.title = request.POST.get('name', '').strip()
            template.description = request.POST.get('description', '').strip()
            template.is_active = request.POST.get('is_active') == 'on'
            
            if not template.title:
                messages.error(request, 'Nome do template é obrigatório.')
                return redirect('edit_checklist_template', template_id=template.id)
            
            template.save()
            
            # Processar itens do checklist
            # Primeiro, marcar todos os itens existentes como "para deletar"
            existing_items = list(template.items.all())
            items_to_keep = []
            
            # Processar itens enviados no formulário
            item_descriptions = request.POST.getlist('item_description[]')
            item_observations = request.POST.getlist('item_observations[]')
            item_required = request.POST.getlist('item_required[]')
            item_ids = request.POST.getlist('item_id[]')
            
            for i, description in enumerate(item_descriptions):
                description = description.strip()
                if not description:
                    continue
                
                observations = item_observations[i] if i < len(item_observations) else ''
                is_required = str(i) in item_required
                item_id = item_ids[i] if i < len(item_ids) else ''
                
                if item_id and item_id.isdigit():
                    # Item existente - atualizar
                    try:
                        item = ChecklistTemplateItem.objects.get(id=int(item_id), template=template)
                        item.title = description
                        item.description = observations
                        item.is_required = is_required
                        item.order = i + 1
                        item.save()
                        items_to_keep.append(item.id)
                    except ChecklistTemplateItem.DoesNotExist:
                        # Item não encontrado, criar novo
                        ChecklistTemplateItem.objects.create(
                            template=template,
                            title=description,
                            description=observations,
                            is_required=is_required,
                            order=i + 1
                        )
                else:
                    # Item novo - criar
                    ChecklistTemplateItem.objects.create(
                        template=template,
                        title=description,
                        description=observations,
                        is_required=is_required,
                        order=i + 1
                    )
            
            # Deletar itens que não estão mais no formulário
            template.items.exclude(id__in=items_to_keep).delete()
            
            messages.success(request, 'Template atualizado com sucesso!')
            return redirect('checklist_template_detail', template_id=template.id)
            
        except Exception as e:
            messages.error(request, f'Erro ao atualizar template: {str(e)}')
            return redirect('edit_checklist_template', template_id=template.id)
    
    context = {
        'template': template,
        'is_supervisor': is_supervisor,
    }
    
    return render(request, 'admin_panel/edit_checklist_template.html', context)


@login_required
@require_POST
def delete_checklist_template(request, template_id):
    """Deletar template de checklist"""
    from core.models import ChecklistTemplate
    
    try:
        template = get_object_or_404(ChecklistTemplate, id=template_id)
        
        # Verificar se o usuário tem permissão
        if not (request.user.hierarchy in ['SUPERVISOR', 'ADMIN', 'SUPERADMIN', 'ADMINISTRATIVO'] or request.user.is_staff):
            return JsonResponse({'success': False, 'error': 'Permissão negada'})
        
        template_name = template.title
        template.delete()
        
        return JsonResponse({
            'success': True,
            'message': f'Template "{template_name}" removido com sucesso!'
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
@require_POST
def delete_task(request, task_id):
    """Deletar tarefa"""
    from core.models import TaskActivity
    
    try:
        task = get_object_or_404(TaskActivity, id=task_id)
        
        # Verificar se o usuário tem permissão (supervisor ou criador da tarefa)
        if not (request.user.hierarchy in ['SUPERVISOR', 'ADMIN', 'SUPERADMIN', 'ADMINISTRATIVO'] or 
                request.user.is_staff or task.created_by == request.user):
            return JsonResponse({'success': False, 'error': 'Permissão negada'})
        
        task_title = task.title
        task.delete()
        
        return JsonResponse({
            'success': True,
            'message': f'Tarefa "{task_title}" removida com sucesso!'
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@require_http_methods(["GET"])
@csrf_exempt
def daily_automation_api(request):
    """
    API para automação diária - retorna dados do usuário e chamados
    Usado para sistemas externos de automação
    """
    try:
        # Verificar parâmetros
        user_id = request.GET.get('user_id')
        api_token = request.GET.get('token')
        
        if not user_id:
            return JsonResponse({
                'success': False,
                'error': 'user_id é obrigatório'
            }, status=400)
        
        # Para ambiente de produção, remover verificação de token por enquanto
        # Implementar autenticação mais robusta posteriormente se necessário
        
        # Buscar usuário
        user = get_object_or_404(User, id=user_id)
        
        # Importar models necessários
        from tickets.models import Ticket
        
        # Chamados abertos do usuário
        user_open_tickets = Ticket.objects.filter(
            assigned_to=user,
            status__in=['open', 'in_progress', 'waiting']
        ).select_related('category', 'sector', 'created_by')
        
        # Chamados abertos do setor do usuário
        sector_open_tickets = Ticket.objects.filter(
            sector=user.sector,
            status__in=['open', 'in_progress', 'waiting']
        ).exclude(
            assigned_to=user  # Excluir os que já estão na lista do usuário
        ).select_related('category', 'sector', 'created_by', 'assigned_to')
        
        # Preparar dados dos chamados do usuário
        user_tickets_data = []
        for ticket in user_open_tickets:
            user_tickets_data.append({
                'id': ticket.id,
                'title': ticket.title,
                'description': ticket.description,
                'status': ticket.status,
                'priority': ticket.priority,
                'category': ticket.category.name if ticket.category else None,
                'created_by': {
                    'id': ticket.created_by.id,
                    'name': ticket.created_by.get_full_name() or ticket.created_by.username,
                    'email': ticket.created_by.email
                },
                'created_at': ticket.created_at.isoformat(),
                'updated_at': ticket.updated_at.isoformat()
            })
        
        # Preparar dados dos chamados do setor
        sector_tickets_data = []
        for ticket in sector_open_tickets:
            assigned_to_data = None
            if ticket.assigned_to:
                assigned_to_data = {
                    'id': ticket.assigned_to.id,
                    'name': ticket.assigned_to.get_full_name() or ticket.assigned_to.username,
                    'email': ticket.assigned_to.email
                }
            
            sector_tickets_data.append({
                'id': ticket.id,
                'title': ticket.title,
                'description': ticket.description,
                'status': ticket.status,
                'priority': ticket.priority,
                'category': ticket.category.name if ticket.category else None,
                'created_by': {
                    'id': ticket.created_by.id,
                    'name': ticket.created_by.get_full_name() or ticket.created_by.username,
                    'email': ticket.created_by.email
                },
                'assigned_to': assigned_to_data,
                'created_at': ticket.created_at.isoformat(),
                'updated_at': ticket.updated_at.isoformat()
            })
        
        # Dados do usuário
        user_data = {
            'id': user.id,
            'username': user.username,
            'email': user.email,
            'full_name': user.get_full_name(),
            'first_name': user.first_name,
            'last_name': user.last_name,
            'phone': getattr(user, 'phone', '') or '',
            'hierarchy': user.hierarchy,
            'is_active': user.is_active,
            'sector': {
                'id': user.sector.id,
                'name': user.sector.name
            } if user.sector else None,
            'last_login': user.last_login.isoformat() if user.last_login else None
        }
        
        # Resposta da API
        response_data = {
            'success': True,
            'timestamp': timezone.now().isoformat(),
            'user': user_data,
            'user_open_tickets': {
                'count': len(user_tickets_data),
                'tickets': user_tickets_data
            },
            'sector_open_tickets': {
                'count': len(sector_tickets_data),
                'tickets': sector_tickets_data
            },
            'summary': {
                'total_user_tickets': len(user_tickets_data),
                'total_sector_tickets': len(sector_tickets_data),
                'total_tickets': len(user_tickets_data) + len(sector_tickets_data)
            }
        }
        
        return JsonResponse(response_data)
        
    except User.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Usuário não encontrado'
        }, status=404)
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Erro interno: {str(e)}'
        }, status=500)


@login_required
def debug_webhooks_view(request):
    """View de debug para testar webhooks"""
    if not request.user.can_manage_users():
        return JsonResponse({'error': 'Permissão negada'}, status=403)
    
    from tickets.models import Webhook
    from communications.models import Communication
    
    # Buscar webhooks configurados
    webhooks = Webhook.objects.all()
    
    # Estatísticas
    stats = {
        'total_webhooks': webhooks.count(),
        'active_webhooks': webhooks.filter(is_active=True).count(),
        'communication_webhooks': webhooks.filter(
            event__in=['COMMUNICATION_CREATED', 'COMMUNICATION_UPDATED'],
            is_active=True
        ).count(),
        'approval_webhooks': webhooks.filter(
            event='APPROVAL_REQUEST',
            is_active=True
        ).count()
    }
    
    # Últimos comunicados
    recent_communications = Communication.objects.all()[:5]
    
    # Se é um POST, testar disparar um webhook
    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'test_communication':
            # Criar comunicado de teste
            test_comm = Communication(
                id=9999,  # ID fictício
                title="[TESTE] Comunicado de Teste - Debug",
                message="Este é um teste de webhook para comunicados.",
                sender=request.user,
                send_to_all=True
            )
            
            # Disparar webhooks manualmente
            try:
                test_comm.trigger_webhooks('COMMUNICATION_CREATED')
                result = {'success': True, 'message': 'Webhook testado com sucesso'}
            except Exception as e:
                result = {'success': False, 'error': str(e)}
            
            return JsonResponse(result)
        
        elif action == 'test_approval':
            # Testar webhook de aprovação
            from tickets.models import PurchaseOrderApproval
            
            # Simulação de aprovação de teste
            try:
                webhooks_approval = Webhook.objects.filter(
                    event='APPROVAL_REQUEST',
                    is_active=True
                )
                
                for webhook in webhooks_approval:
                    test_payload = {
                        'event': 'approval_request',
                        'test': True,
                        'purchase_approval': {
                            'id': 9999,
                            'ticket_id': 9999,
                            'amount': 100.00,
                            'approval_step': 1,
                            'status': 'PENDING',
                            'created_at': timezone.now().isoformat()
                        },
                        'approver_user': {
                            'id': request.user.id,
                            'name': request.user.get_full_name(),
                            'email': request.user.email,
                            'phone': getattr(request.user, 'phone', '') or ''
                        }
                    }
                    
                    webhook._send_webhook(test_payload)
                
                result = {'success': True, 'message': f'{webhooks_approval.count()} webhook(s) testado(s)'}
            except Exception as e:
                result = {'success': False, 'error': str(e)}
        
        elif action == 'test_communication_real':
            # Criar um comunicado real de teste
            try:
                from communications.models import Communication
                
                # Criar comunicado de teste
                test_communication = Communication.objects.create(
                    title="[TESTE DEBUG] Comunicado de Teste",
                    message="Este é um comunicado de teste criado pela funcionalidade de debug de webhooks. Pode ser ignorado.",
                    sender=request.user,
                    send_to_all=False  # Não enviar para todos para evitar spam
                )
                
                # Adicionar apenas o usuário atual como destinatário
                test_communication.recipients.add(request.user)
                
                result = {
                    'success': True, 
                    'message': f'Comunicado de teste criado (ID: {test_communication.id}). Verifique os logs no console do servidor para detalhes dos webhooks.'
                }
            except Exception as e:
                result = {'success': False, 'error': str(e)}
            
            return JsonResponse(result)
        
        elif action == 'list_webhooks':
            # Listar todos os webhooks detalhadamente
            try:
                all_webhooks = []
                for webhook in webhooks:
                    all_webhooks.append({
                        'id': webhook.id,
                        'name': webhook.name,
                        'url': webhook.url,
                        'event': webhook.event,
                        'is_active': webhook.is_active,
                        'category': webhook.category.name if webhook.category else None,
                        'sector': webhook.sector.name if webhook.sector else None
                    })
                
                result = {
                    'success': True, 
                    'webhooks': all_webhooks,
                    'count': len(all_webhooks)
                }
            except Exception as e:
                result = {'success': False, 'error': str(e)}
            
            return JsonResponse(result)
    
    context = {
        'webhooks': webhooks,
        'stats': stats,
        'recent_communications': recent_communications,
        'user': request.user,
    }
    
    return render(request, 'admin/debug_webhooks.html', context)


def _build_cs_statement_context(request, statement_user, is_admin_view=False):
    from django.core.paginator import Paginator
    from prizes.models import CSTransaction

    # Buscar transações do usuário (todas, independente do status, para mostrar histórico completo)
    transactions = CSTransaction.objects.filter(user=statement_user).order_by('-created_at')

    # Filtros opcionais
    transaction_type = request.GET.get('type', '')
    status_filter = request.GET.get('status', '')

    if transaction_type:
        transactions = transactions.filter(transaction_type=transaction_type)

    if status_filter:
        transactions = transactions.filter(status=status_filter)

    # Paginação
    paginator = Paginator(transactions, 20)  # 20 transações por página
    page_number = request.GET.get('page')
    transactions_page = paginator.get_page(page_number)

    # Calcular totais (apenas transações aprovadas)
    approved_transactions = CSTransaction.objects.filter(user=statement_user, status='APPROVED')

    total_credits = approved_transactions.filter(
        transaction_type__in=['CREDIT', 'ADJUSTMENT', 'REFUND']
    ).aggregate(total=Sum('amount'))['total'] or Decimal('0')

    total_debits = approved_transactions.filter(
        transaction_type__in=['DEBIT', 'REDEMPTION']
    ).aggregate(total=Sum('amount'))['total'] or Decimal('0')

    # Transações pendentes
    pending_count = CSTransaction.objects.filter(user=statement_user, status='PENDING').count()

    context = {
        'transactions': transactions_page,
        'total_credits': total_credits,
        'total_debits': total_debits,
        'current_balance': statement_user.balance_cs,
        'pending_count': pending_count,
        'transaction_type': transaction_type,
        'status_filter': status_filter,
        'user': request.user,
        'statement_user': statement_user,
        'is_admin_view': is_admin_view,
    }

    return context


@login_required
def my_cs_statement_view(request):
    """Extrato de Confianças C$ do usuário logado"""
    context = _build_cs_statement_context(request, request.user, is_admin_view=False)
    return render(request, 'users/my_cs_statement.html', context)


@login_required
def user_cs_statement_view(request, user_id):
    """Extrato de Confianças C$ de outro usuário (Administração+)"""
    if not request.user.can_access_admin_panel():
        messages.error(request, 'Você não tem permissão para acessar este extrato.')
        return redirect('manage_cs')

    statement_user = get_object_or_404(User, id=user_id)
    context = _build_cs_statement_context(request, statement_user, is_admin_view=True)
    return render(request, 'users/my_cs_statement.html', context)
