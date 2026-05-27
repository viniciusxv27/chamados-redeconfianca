from __future__ import annotations

from decimal import Decimal, InvalidOperation

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db.models import Count, Sum
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from django.views.decorators.http import require_POST

from .models import VendaD1, VendaD1Contestacao, VendaD1ChatMessage, VendaD1ChatAttachment
from .services import expire_deadlines, is_ilha, sync_d1, vendas_for_user


def _apply_common_filters(qs, request):
    """Filtros compartilhados entre a Lista e o Kanban."""
    status = (request.GET.get('status') or '').strip()
    if status:
        qs = qs.filter(status=status)
    pdv = (request.GET.get('pdv') or '').strip()
    if pdv:
        qs = qs.filter(pdv__icontains=pdv)
    vendedor = (request.GET.get('vendedor') or '').strip()
    if vendedor:
        qs = qs.filter(vendedor__icontains=vendedor)
    pilar = (request.GET.get('pilar') or '').strip()
    if pilar:
        qs = qs.filter(pilar__iexact=pilar)
    return qs, {
        'filtro_status': status,
        'filtro_pdv': pdv,
        'filtro_vendedor': vendedor,
        'filtro_pilar': pilar,
    }


def _available_pilares(qs):
    """Lista distinta de pilares para o filtro (sem vazios, 'Erro' e 'Simcard')."""
    seen: set[str] = set()
    result: list[str] = []
    excluded = {'ERRO', 'SIMCARD'}
    for p in qs.exclude(pilar='').values_list('pilar', flat=True).distinct():
        if not p:
            continue
        norm = p.strip()
        if not norm or norm.upper() in excluded:
            continue
        key = norm.upper()
        if key in seen:
            continue
        seen.add(key)
        result.append(norm)
    result.sort(key=str.upper)
    return result


@login_required
def lista(request):
    # Sempre que carrega a lista, expira deadlines vencidas (cheap).
    expire_deadlines()

    qs = vendas_for_user(request.user)

    show_expired = request.GET.get('expirados') == '1'
    if not show_expired:
        qs = qs.exclude(acordo_status=VendaD1.ACORDO_EXPIRADO)

    pilares = _available_pilares(qs)
    qs, filtros = _apply_common_filters(qs, request)

    return render(request, 'validad1/lista.html', {
        'vendas': qs.order_by('-data_da_venda', '-id'),
        'is_ilha': is_ilha(request.user),
        'status_choices': VendaD1.STATUS_CHOICES,
        'pilares': pilares,
        'show_expired': show_expired,
        **filtros,
    })


@login_required
def kanban(request):
    """Visão Kanban dos 3 status principais."""
    expire_deadlines()
    qs = vendas_for_user(request.user)
    show_expired = request.GET.get('expirados') == '1'
    if not show_expired:
        qs = qs.exclude(acordo_status=VendaD1.ACORDO_EXPIRADO)

    pilares = _available_pilares(qs)
    qs, filtros = _apply_common_filters(qs, request)

    colunas = []
    for key, label in VendaD1.STATUS_CHOICES:
        col_qs = qs.filter(status=key).order_by('-data_da_venda', '-id')
        colunas.append({
            'key': key,
            'label': label,
            'items': list(col_qs),
            'count': col_qs.count(),
            'total': col_qs.aggregate(t=Sum('valor'))['t'] or Decimal('0'),
        })

    return render(request, 'validad1/kanban.html', {
        'colunas': colunas,
        'is_ilha': is_ilha(request.user),
        'status_choices': VendaD1.STATUS_CHOICES,
        'pilares': pilares,
        'show_expired': show_expired,
        **filtros,
    })


@login_required
def detail(request, pk: int):
    venda = get_object_or_404(VendaD1, pk=pk)
    contestacoes = venda.contestacoes.all()
    return render(request, 'validad1/detail.html', {
        'venda': venda,
        'contestacoes': contestacoes,
        'is_ilha': is_ilha(request.user),
        'tipo_divergencia_choices': VendaD1.TIPO_DIVERGENCIA_CHOICES,
        'penalidade_choices': VendaD1.PENALIDADE_CHOICES,
    })


@login_required
@require_POST
def editar_venda(request, pk: int):
    """Permite à Ilha (quem confere/valida) corrigir telefone (nº acesso) e valor."""
    if not is_ilha(request.user):
        messages.error(request, 'Apenas a Ilha pode editar a venda.')
        return redirect('validad1:detail', pk=pk)

    venda = get_object_or_404(VendaD1, pk=pk)
    novo_acesso = (request.POST.get('numero_acesso') or '').strip()
    novo_valor_raw = (request.POST.get('valor') or '').strip().replace('.', '').replace(',', '.')

    alterados = []
    if novo_acesso != venda.numero_acesso:
        venda.numero_acesso = novo_acesso
        alterados.append('número de acesso')

    if novo_valor_raw:
        try:
            novo_valor = Decimal(novo_valor_raw)
            if novo_valor != venda.valor:
                venda.valor = novo_valor
                alterados.append('valor')
        except (InvalidOperation, ValueError):
            messages.error(request, 'Valor informado é inválido.')
            return redirect('validad1:detail', pk=pk)

    if alterados:
        venda.save(update_fields=['numero_acesso', 'valor', 'last_synced_at'])
        messages.success(request, 'Atualizado: ' + ', '.join(alterados) + '.')
    else:
        messages.info(request, 'Nada para alterar.')
    return redirect('validad1:detail', pk=pk)


@login_required
@require_POST
def sinalizar(request, pk: int):
    if not is_ilha(request.user):
        messages.error(request, 'Apenas a Ilha pode sinalizar vendas.')
        return redirect('validad1:detail', pk=pk)

    venda = get_object_or_404(VendaD1, pk=pk)
    acao = request.POST.get('acao')
    observacao = (request.POST.get('observacao') or '').strip()

    if acao == 'conformidade':
        venda.set_conformidade(por_usuario=request.user)
        messages.success(request, 'Venda marcada como em conformidade.')
    elif acao == 'divergente':
        tipo = request.POST.get('tipo_divergencia') or ''
        penalidade = request.POST.get('penalidade') or VendaD1.PEN_NENHUMA
        penalidade_detalhe = (request.POST.get('penalidade_detalhe') or '').strip()
        venda.acao_realizada_no_go = bool(request.POST.get('acao_realizada_no_go'))
        venda.set_divergente(
            tipo=tipo, penalidade=penalidade, observacao=observacao,
            por_usuario=request.user, penalidade_detalhe=penalidade_detalhe,
        )
        venda.save()
        messages.success(request, 'Venda sinalizada como divergente. Gerente tem 48h para responder.')
    else:
        messages.error(request, 'Ação inválida.')

    return redirect('validad1:detail', pk=pk)


@login_required
@require_POST
def de_acordo(request, pk: int):
    venda = get_object_or_404(VendaD1, pk=pk)
    if venda.acordo_status != VendaD1.ACORDO_PENDENTE:
        messages.warning(request, 'Esta venda já teve seu acordo respondido.')
        return redirect('validad1:detail', pk=pk)
    venda.acordo_status = VendaD1.ACORDO_DE_ACORDO
    venda.acordo_respondido_por = request.user
    venda.acordo_respondido_em = timezone.now()
    venda.save()
    messages.success(request, 'De acordo registrado.')
    return redirect('validad1:detail', pk=pk)


@login_required
@require_POST
def contestar(request, pk: int):
    venda = get_object_or_404(VendaD1, pk=pk)
    if venda.acordo_status != VendaD1.ACORDO_PENDENTE:
        messages.warning(request, 'Esta venda já teve seu acordo respondido.')
        return redirect('validad1:detail', pk=pk)

    motivo = (request.POST.get('motivo') or '').strip()
    if not motivo:
        messages.error(request, 'Descreva o motivo da contestação.')
        return redirect('validad1:detail', pk=pk)

    contestacao = VendaD1Contestacao.objects.create(
        venda=venda, aberto_por=request.user, motivo=motivo,
    )
    venda.acordo_status = VendaD1.ACORDO_CONTESTADO
    venda.acordo_respondido_por = request.user
    venda.acordo_respondido_em = timezone.now()
    venda.save()
    messages.success(request, 'Contestação aberta — Ilha de Qualidade notificada.')
    return redirect('validad1:contestacao_detail', pk=contestacao.pk)


@login_required
def contestacao_detail(request, pk: int):
    contestacao = get_object_or_404(VendaD1Contestacao, pk=pk)
    # Marca como visualizada quando o autor abre a tratativa (para destacar
    # respostas novas da Ilha na lista/kanban).
    if contestacao.aberto_por_id == request.user.id:
        contestacao.last_opener_view_at = timezone.now()
        contestacao.save(update_fields=['last_opener_view_at'])
    return render(request, 'validad1/contestacao.html', {
        'contestacao': contestacao,
        'mensagens_chat': contestacao.mensagens.select_related('autor').prefetch_related('anexos'),
        'is_ilha': is_ilha(request.user),
    })


@login_required
@require_POST
def contestacao_post(request, pk: int):
    contestacao = get_object_or_404(VendaD1Contestacao, pk=pk)
    texto = (request.POST.get('texto') or '').strip()
    arquivos = request.FILES.getlist('anexos')
    if not texto and not arquivos:
        return redirect('validad1:contestacao_detail', pk=pk)
    msg = VendaD1ChatMessage.objects.create(
        contestacao=contestacao, autor=request.user, texto=texto,
    )
    for f in arquivos:
        VendaD1ChatAttachment.objects.create(
            mensagem=msg,
            arquivo=f,
            nome_original=getattr(f, 'name', '')[:255],
            content_type=getattr(f, 'content_type', '') or '',
            tamanho=getattr(f, 'size', 0) or 0,
        )
    return redirect('validad1:contestacao_detail', pk=pk)


@login_required
@require_POST
def contestacao_resolver(request, pk: int):
    if not is_ilha(request.user):
        messages.error(request, 'Sem permissão.')
        return redirect('validad1:contestacao_detail', pk=pk)
    contestacao = get_object_or_404(VendaD1Contestacao, pk=pk)
    decisao = request.POST.get('decisao')
    resposta = (request.POST.get('resposta') or '').strip()

    if decisao == 'procedente':
        contestacao.status = VendaD1Contestacao.STATUS_PROCEDENTE
        # Se a Ilha concorda com o gerente, deve ajustar no Vivo GO.
        contestacao.venda.acao_realizada_no_go = True
        contestacao.venda.status = VendaD1.STATUS_CONFORMIDADE
        contestacao.venda.save()
    elif decisao == 'improcedente':
        contestacao.status = VendaD1Contestacao.STATUS_IMPROCEDENTE
    else:
        messages.error(request, 'Decisão inválida.')
        return redirect('validad1:contestacao_detail', pk=pk)

    contestacao.resposta = resposta
    contestacao.respondido_por = request.user
    contestacao.save()
    messages.success(request, 'Contestação resolvida.')
    return redirect('validad1:contestacao_detail', pk=pk)


@login_required
@require_POST
def sync_now(request):
    if not is_ilha(request.user):
        messages.error(request, 'Sem permissão.')
        return redirect('validad1:lista')
    stats = sync_d1()
    messages.success(
        request,
        f"Sync D-1 ({stats['target_date']}): {stats['created']} importadas, "
        f"{stats['expired']} expiradas (de {stats['total_in_source']} na fonte).",
    )
    return redirect('validad1:lista')


@login_required
def relatorio(request):
    qs = vendas_for_user(request.user)
    total = qs.count() or 1

    por_status = []
    for k, l in VendaD1.STATUS_CHOICES:
        sub = qs.filter(status=k)
        cnt = sub.count()
        por_status.append({
            'key': k, 'label': l, 'qtd': cnt,
            'percent': round((cnt / total) * 100, 1),
            'receita': sub.aggregate(t=Sum('valor'))['t'] or Decimal('0'),
        })

    por_divergencia_raw = list(
        qs.filter(status=VendaD1.STATUS_DIVERGENTE)
        .values('tipo_divergencia')
        .annotate(n=Count('id'), receita=Sum('valor'))
        .order_by('-n')
    )
    _tipo_map = dict(VendaD1.TIPO_DIVERGENCIA_CHOICES)
    por_divergencia = [
        {**d, 'tipo_label': _tipo_map.get(d['tipo_divergencia'], d['tipo_divergencia'] or 'Não classificado')}
        for d in por_divergencia_raw
    ]

    # Acordo (gerentes)
    por_acordo = []
    for k, l in VendaD1.ACORDO_CHOICES:
        cnt = qs.filter(acordo_status=k).count()
        por_acordo.append({'key': k, 'label': l, 'qtd': cnt})

    # Penalidades
    por_penalidade = []
    for k, l in VendaD1.PENALIDADE_CHOICES:
        cnt = qs.filter(penalidade=k, status=VendaD1.STATUS_DIVERGENTE).count()
        if cnt:
            por_penalidade.append({'key': k, 'label': l, 'qtd': cnt})

    receita_total = qs.aggregate(t=Sum('valor'))['t'] or Decimal('0')

    # Datasets para Chart.js
    tipo_labels_map = dict(VendaD1.TIPO_DIVERGENCIA_CHOICES)
    div_labels = [tipo_labels_map.get(d['tipo_divergencia'], d['tipo_divergencia'] or 'Não classificado') for d in por_divergencia]
    div_data = [d['n'] for d in por_divergencia]

    status_colors_map = {
        VendaD1.STATUS_PENDENTE: '#6B7280',
        VendaD1.STATUS_CONFORMIDADE: '#10B981',
        VendaD1.STATUS_DIVERGENTE: '#EF4444',
    }
    status_chart_labels = [s['label'] for s in por_status]
    status_chart_data = [s['qtd'] for s in por_status]
    status_chart_colors = [status_colors_map.get(s['key'], '#6B7280') for s in por_status]

    # Empilhado por PDV: cada barra = 1 PDV; segmentos = contagem por status.
    pdv_rows = (
        qs.exclude(pdv='')
        .values('pdv', 'status')
        .annotate(n=Count('id'))
    )
    pdv_totais: dict[str, dict[str, int]] = {}
    for row in pdv_rows:
        pdv = (row['pdv'] or '').strip() or '—'
        bucket = pdv_totais.setdefault(pdv, {k: 0 for k, _ in VendaD1.STATUS_CHOICES})
        bucket[row['status']] = bucket.get(row['status'], 0) + row['n']

    # Ordena por total decrescente para a barra mais cheia aparecer primeiro.
    pdv_sorted = sorted(
        pdv_totais.items(), key=lambda it: sum(it[1].values()), reverse=True,
    )
    pdv_labels = [p for p, _ in pdv_sorted]
    pdv_datasets = [
        {
            'label': label,
            'data': [pdv_totais[p].get(key, 0) for p in pdv_labels],
            'backgroundColor': status_colors_map.get(key, '#6B7280'),
        }
        for key, label in VendaD1.STATUS_CHOICES
    ]

    return render(request, 'validad1/relatorio.html', {
        'total': qs.count(),
        'receita_total': receita_total,
        'por_status': por_status,
        'por_divergencia': por_divergencia,
        'por_acordo': por_acordo,
        'por_penalidade': por_penalidade,
        'tipo_labels': tipo_labels_map,
        'status_chart_labels': status_chart_labels,
        'status_chart_data': status_chart_data,
        'status_chart_colors': status_chart_colors,
        'div_chart_labels': div_labels,
        'div_chart_data': div_data,
        'pdv_chart_labels': pdv_labels,
        'pdv_chart_datasets': pdv_datasets,
    })


@login_required
def export_excel(request):
    """Exporta as vendas D-1 do mês corrente para um arquivo .xlsx."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    qs = vendas_for_user(request.user).order_by('-data_da_venda', '-id')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Valida D-1'

    headers = [
        'Nº Venda', 'CPF', 'Nº Acesso', 'Produto', 'Valor',
        'PDV', 'Vendedor', 'Data da Venda', 'Pilar', 'Serviços',
        'Status', 'Tipo Divergência', 'Penalidade',
        'Ação no Vivo GO', 'Observação Ilha',
        'Acordo Gerente', 'Acordo Deadline',
    ]
    ws.append(headers)

    header_fill = PatternFill('solid', fgColor='660099')
    header_font = Font(bold=True, color='FFFFFF', name='Calibri', size=11)
    for col_idx, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    for v in qs:
        ws.append([
            v.numero_da_venda, v.cpf, v.numero_acesso, v.produto,
            float(v.valor or 0), v.pdv, v.vendedor,
            v.data_da_venda.strftime('%d/%m/%Y') if v.data_da_venda else '',
            v.pilar, v.servicos,
            v.get_status_display(),
            v.get_tipo_divergencia_display() if v.tipo_divergencia else '',
            v.get_penalidade_display(),
            'Sim' if v.acao_realizada_no_go else 'Não',
            v.observacao,
            v.get_acordo_status_display(),
            timezone.localtime(v.acordo_deadline).strftime('%d/%m/%Y %H:%M') if v.acordo_deadline else '',
        ])

    widths = [16, 16, 16, 28, 12, 22, 24, 14, 12, 22, 18, 22, 18, 14, 40, 18, 20]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = 'A2'

    filename = f'validad1_{timezone.now().strftime("%Y%m%d_%H%M")}.xlsx'
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response
