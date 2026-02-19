from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Q, Count, Sum
from django.core.paginator import Paginator
from django.http import HttpResponse, JsonResponse
from django.utils import timezone
from django.views.decorators.http import require_POST
from functools import wraps
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from datetime import timedelta

from .models import (
    Asset, Product, ProductMedia, InventoryItem, 
    StockMovement, InventoryCategory, InventoryManager, ItemRequest, ItemRequestItem
)
from .forms import (
    AssetForm, ProductForm, ProductMediaForm, InventoryItemForm,
    StockEntryForm, StockExitForm, InventoryCategoryForm,
    InventoryManagerForm, BulkInventoryItemForm,
    ItemRequestForm, ItemRequestItemForm, ItemRequestReviewForm, ItemRequestDeliveryForm,
    ItemRequestCounterProposalForm, ItemRequestCounterProposalResponseForm
)
from users.models import User, Sector
from core.middleware import log_action


# ============================================================================
# DECORADORES E PERMISSÕES
# ============================================================================

def _get_inventory_manager(user):
    """Retorna o perfil de gestor de inventário do usuário, ou None"""
    try:
        return user.inventory_manager_profile
    except InventoryManager.DoesNotExist:
        return None


def is_inventory_manager(user):
    """Verifica se o usuário é gestor de inventário ou superadmin"""
    if user.hierarchy in ['SUPERADMIN', 'ADMIN']:
        return True
    manager = _get_inventory_manager(user)
    return manager is not None and manager.is_active


def inventory_permission_required(permission=None):
    """Decorator para verificar permissões de inventário"""
    def decorator(view_func):
        @wraps(view_func)
        def _wrapped_view(request, *args, **kwargs):
            user = request.user
            
            # Superadmin e Admin têm acesso total
            if user.hierarchy in ['SUPERADMIN', 'ADMIN']:
                return view_func(request, *args, **kwargs)
            
            # Verificar se é gestor de inventário
            manager = _get_inventory_manager(user)
            if manager is None:
                messages.error(request, 'Você não tem permissão para acessar esta área.')
                return redirect('assets:inventory_dashboard')
            
            if not manager.is_active:
                messages.error(request, 'Seu acesso ao inventário está desativado.')
                return redirect('assets:inventory_dashboard')
            
            # Verificar permissão específica
            if permission:
                if not getattr(manager, permission, False):
                    messages.error(request, 'Você não tem permissão para esta ação.')
                    return redirect('assets:inventory_dashboard')
            
            return view_func(request, *args, **kwargs)
        return _wrapped_view
    return decorator


# ============================================================================
# DASHBOARD DE INVENTÁRIO
# ============================================================================

@login_required
def inventory_dashboard(request):
    """Dashboard principal do inventário"""
    # Estatísticas gerais
    total_products = Product.objects.filter(is_active=True).count()
    total_items = InventoryItem.objects.count()
    available_items = InventoryItem.objects.filter(status='available').count()
    in_use_items = InventoryItem.objects.filter(status='in_use').count()
    
    # Produtos com estoque baixo
    low_stock_products = []
    for product in Product.objects.filter(is_active=True, min_stock__gt=0):
        if product.is_low_stock:
            low_stock_products.append(product)
    
    # Últimas movimentações
    recent_movements = StockMovement.objects.select_related(
        'inventory_item', 'inventory_item__product', 'created_by'
    ).order_by('-created_at')[:10]
    
    # Estatísticas por categoria
    categories_stats = InventoryCategory.objects.filter(is_active=True).annotate(
        product_count=Count('products', filter=Q(products__is_active=True)),
        item_count=Count('products__inventory_items')
    )
    
    # Items por status
    status_stats = InventoryItem.objects.values('status').annotate(
        count=Count('id')
    ).order_by('status')
    
    # Verificar se é gestor
    is_manager = is_inventory_manager(request.user)
    
    # Solicitações pendentes (para gestores)
    pending_requests_count = ItemRequest.objects.filter(status='pending').count()
    
    # Minhas solicitações recentes
    my_recent_requests = ItemRequest.objects.filter(
        requested_by=request.user
    ).order_by('-requested_at')[:5]
    
    context = {
        'total_products': total_products,
        'total_items': total_items,
        'available_items': available_items,
        'in_use_items': in_use_items,
        'low_stock_products': low_stock_products[:5],
        'recent_movements': recent_movements,
        'categories_stats': categories_stats,
        'status_stats': status_stats,
        'is_manager': is_manager,
        'pending_requests_count': pending_requests_count,
        'my_recent_requests': my_recent_requests,
        'is_approver': can_approve_requests(request.user),
    }
    
    return render(request, 'assets/inventory/dashboard.html', context)


# ============================================================================
# CATEGORIAS
# ============================================================================

@login_required
@inventory_permission_required('can_manage_products')
def category_list(request):
    """Lista todas as categorias"""
    categories = InventoryCategory.objects.annotate(
        product_count=Count('products', filter=Q(products__is_active=True))
    ).order_by('name')
    
    context = {
        'categories': categories,
    }
    return render(request, 'assets/inventory/category_list.html', context)


@login_required
@inventory_permission_required('can_manage_products')
def category_create(request):
    """Criar nova categoria"""
    if request.method == 'POST':
        form = InventoryCategoryForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Categoria criada com sucesso!')
            return redirect('assets:category_list')
    else:
        form = InventoryCategoryForm()
    
    context = {
        'form': form,
        'title': 'Nova Categoria',
    }
    return render(request, 'assets/inventory/category_form.html', context)


@login_required
@inventory_permission_required('can_manage_products')
def category_edit(request, pk):
    """Editar categoria existente"""
    category = get_object_or_404(InventoryCategory, pk=pk)
    
    if request.method == 'POST':
        form = InventoryCategoryForm(request.POST, instance=category)
        if form.is_valid():
            form.save()
            messages.success(request, 'Categoria atualizada com sucesso!')
            return redirect('assets:category_list')
    else:
        form = InventoryCategoryForm(instance=category)
    
    context = {
        'form': form,
        'category': category,
        'title': f'Editar Categoria - {category.name}',
    }
    return render(request, 'assets/inventory/category_form.html', context)


@login_required
@inventory_permission_required('can_manage_products')
def category_delete(request, pk):
    """Deletar categoria"""
    category = get_object_or_404(InventoryCategory, pk=pk)
    
    if request.method == 'POST':
        name = category.name
        # Verificar se há produtos vinculados
        if category.products.exists():
            messages.error(request, f'Não é possível excluir a categoria "{name}". Existem produtos vinculados.')
            return redirect('assets:category_list')
        
        category.delete()
        messages.success(request, f'Categoria "{name}" excluída com sucesso!')
        return redirect('assets:category_list')
    
    context = {
        'category': category,
    }
    return render(request, 'assets/inventory/category_delete.html', context)


# ============================================================================
# PRODUTOS
# ============================================================================

@login_required
def product_list(request):
    """Lista todos os produtos"""
    query = request.GET.get('q', '')
    category_filter = request.GET.get('category', '')
    status_filter = request.GET.get('status', '')
    
    products = Product.objects.filter(is_active=True).select_related('category')
    
    if query:
        products = products.filter(
            Q(name__icontains=query) |
            Q(sku__icontains=query) |
            Q(brand__icontains=query) |
            Q(model__icontains=query)
        )
    
    if category_filter:
        products = products.filter(category_id=category_filter)
    
    if status_filter == 'low_stock':
        # Filtrar produtos com estoque baixo (feito em Python pois é propriedade)
        products = [p for p in products if p.is_low_stock]
    
    # Paginação
    paginator = Paginator(products, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    categories = InventoryCategory.objects.filter(is_active=True)
    
    context = {
        'page_obj': page_obj,
        'query': query,
        'category_filter': category_filter,
        'status_filter': status_filter,
        'categories': categories,
        'total_products': len(products) if isinstance(products, list) else products.count(),
        'is_manager': is_inventory_manager(request.user),
    }
    
    return render(request, 'assets/inventory/product_list.html', context)


@login_required
def product_detail(request, pk):
    """Detalhes de um produto"""
    product = get_object_or_404(Product.objects.select_related('category', 'created_by'), pk=pk)
    
    # Itens deste produto
    items = product.inventory_items.select_related('assigned_to', 'assigned_sector').order_by('inventory_number')
    
    # Estatísticas
    items_by_status = items.values('status').annotate(count=Count('id'))
    
    # Mídias
    media_items = product.media.all()
    
    context = {
        'product': product,
        'items': items,
        'items_by_status': items_by_status,
        'media_items': media_items,
        'is_manager': is_inventory_manager(request.user),
    }
    
    return render(request, 'assets/inventory/product_detail.html', context)


@login_required
@inventory_permission_required('can_manage_products')
def product_create(request):
    """Criar novo produto"""
    if request.method == 'POST':
        form = ProductForm(request.POST)
        if form.is_valid():
            product = form.save(commit=False)
            product.created_by = request.user
            product.save()
            messages.success(request, f'Produto "{product.name}" criado com sucesso!')
            return redirect('assets:product_detail', pk=product.pk)
    else:
        form = ProductForm()
    
    context = {
        'form': form,
        'title': 'Novo Produto',
    }
    return render(request, 'assets/inventory/product_form.html', context)


@login_required
@inventory_permission_required('can_manage_products')
def product_edit(request, pk):
    """Editar produto existente"""
    product = get_object_or_404(Product, pk=pk)
    
    if request.method == 'POST':
        form = ProductForm(request.POST, instance=product)
        if form.is_valid():
            form.save()
            messages.success(request, f'Produto "{product.name}" atualizado com sucesso!')
            return redirect('assets:product_detail', pk=product.pk)
    else:
        form = ProductForm(instance=product)
    
    context = {
        'form': form,
        'product': product,
        'title': f'Editar Produto - {product.name}',
    }
    return render(request, 'assets/inventory/product_form.html', context)


@login_required
@inventory_permission_required('can_manage_products')
def product_delete(request, pk):
    """Deletar produto"""
    product = get_object_or_404(Product, pk=pk)
    
    if request.method == 'POST':
        name = product.name
        # Verificar se há itens vinculados
        if product.inventory_items.exists():
            messages.error(request, f'Não é possível excluir o produto "{name}". Existem itens de inventário vinculados.')
            return redirect('assets:product_detail', pk=pk)
        
        product.delete()
        messages.success(request, f'Produto "{name}" excluído com sucesso!')
        return redirect('assets:product_list')
    
    context = {
        'product': product,
    }
    return render(request, 'assets/inventory/product_delete.html', context)


@login_required
@inventory_permission_required('can_manage_products')
def product_media_upload(request, pk):
    """Upload de mídia para produto"""
    product = get_object_or_404(Product, pk=pk)
    
    if request.method == 'POST':
        form = ProductMediaForm(request.POST, request.FILES)
        if form.is_valid():
            media = form.save(commit=False)
            media.product = product
            media.save()
            messages.success(request, 'Mídia adicionada com sucesso!')
            return redirect('assets:product_detail', pk=pk)
    else:
        form = ProductMediaForm()
    
    context = {
        'form': form,
        'product': product,
        'title': f'Adicionar Mídia - {product.name}',
    }
    return render(request, 'assets/inventory/product_media_form.html', context)


@login_required
@inventory_permission_required('can_manage_products')
@require_POST
def product_media_delete(request, pk, media_pk):
    """Deletar mídia do produto"""
    media = get_object_or_404(ProductMedia, pk=media_pk, product_id=pk)
    media.delete()
    messages.success(request, 'Mídia removida com sucesso!')
    return redirect('assets:product_detail', pk=pk)


# ============================================================================
# ITENS DE INVENTÁRIO
# ============================================================================

@login_required
def item_list(request):
    """Lista todos os itens de inventário"""
    query = request.GET.get('q', '')
    product_filter = request.GET.get('product', '')
    status_filter = request.GET.get('status', '')
    sector_filter = request.GET.get('sector', '')
    
    items = InventoryItem.objects.select_related(
        'product', 'product__category', 'assigned_to', 'assigned_sector'
    )
    
    if query:
        items = items.filter(
            Q(inventory_number__icontains=query) |
            Q(serial_number__icontains=query) |
            Q(product__name__icontains=query) |
            Q(product__sku__icontains=query)
        )
    
    if product_filter:
        items = items.filter(product_id=product_filter)
    
    if status_filter:
        items = items.filter(status=status_filter)
    
    if sector_filter:
        items = items.filter(assigned_sector_id=sector_filter)
    
    items = items.order_by('inventory_number')
    
    # Paginação
    paginator = Paginator(items, 25)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    products = Product.objects.filter(is_active=True)
    sectors = Sector.objects.all()
    
    context = {
        'page_obj': page_obj,
        'query': query,
        'product_filter': product_filter,
        'status_filter': status_filter,
        'sector_filter': sector_filter,
        'products': products,
        'sectors': sectors,
        'status_choices': InventoryItem.STATUS_CHOICES,
        'total_items': items.count(),
        'is_manager': is_inventory_manager(request.user),
    }
    
    return render(request, 'assets/inventory/item_list.html', context)


@login_required
def item_detail(request, pk):
    """Detalhes de um item de inventário"""
    item = get_object_or_404(
        InventoryItem.objects.select_related(
            'product', 'product__category', 'assigned_to', 
            'assigned_sector', 'created_by'
        ),
        pk=pk
    )
    
    # Histórico de movimentações
    movements = item.movements.select_related(
        'created_by', 'to_user', 'from_user', 'to_sector', 'from_sector'
    ).order_by('-created_at')
    
    context = {
        'item': item,
        'movements': movements,
        'is_manager': is_inventory_manager(request.user),
    }
    
    return render(request, 'assets/inventory/item_detail.html', context)


@login_required
@inventory_permission_required('can_manage_items')
def item_create(request):
    """Criar novo item de inventário"""
    product_id = request.GET.get('product')
    
    if request.method == 'POST':
        form = InventoryItemForm(request.POST, request.FILES)
        if form.is_valid():
            item = form.save(commit=False)
            item.created_by = request.user
            item.save()
            
            # Registrar movimentação de entrada
            StockMovement.objects.create(
                inventory_item=item,
                movement_type='entry',
                reason='initial_stock',
                to_location=item.location,
                notes='Cadastro inicial do item',
                created_by=request.user
            )
            
            messages.success(request, f'Item "{item.inventory_number}" cadastrado com sucesso!')
            return redirect('assets:item_detail', pk=item.pk)
    else:
        initial = {}
        if product_id:
            initial['product'] = product_id
        form = InventoryItemForm(initial=initial)
    
    context = {
        'form': form,
        'title': 'Novo Item de Inventário',
    }
    return render(request, 'assets/inventory/item_form.html', context)


@login_required
@inventory_permission_required('can_manage_items')
def item_edit(request, pk):
    """Editar item de inventário"""
    item = get_object_or_404(InventoryItem, pk=pk)
    
    if request.method == 'POST':
        form = InventoryItemForm(request.POST, request.FILES, instance=item)
        if form.is_valid():
            form.save()
            messages.success(request, f'Item "{item.inventory_number}" atualizado com sucesso!')
            return redirect('assets:item_detail', pk=item.pk)
    else:
        form = InventoryItemForm(instance=item)
    
    context = {
        'form': form,
        'item': item,
        'title': f'Editar Item - {item.inventory_number}',
    }
    return render(request, 'assets/inventory/item_form.html', context)


@login_required
@inventory_permission_required('can_manage_items')
def item_bulk_create(request):
    """Cadastro em lote de itens"""
    if request.method == 'POST':
        form = BulkInventoryItemForm(request.POST)
        if form.is_valid():
            product = form.cleaned_data['product']
            quantity = form.cleaned_data['quantity']
            prefix = form.cleaned_data['prefix']
            start_number = form.cleaned_data['start_number']
            condition = form.cleaned_data['condition']
            location = form.cleaned_data['location']
            purchase_date = form.cleaned_data['purchase_date']
            purchase_price = form.cleaned_data['purchase_price']
            
            created_items = []
            errors = []
            
            for i in range(quantity):
                inv_number = f"{prefix}{start_number + i:04d}"
                
                # Verificar se já existe
                if InventoryItem.objects.filter(inventory_number=inv_number).exists():
                    errors.append(f"Número {inv_number} já existe")
                    continue
                
                item = InventoryItem.objects.create(
                    product=product,
                    inventory_number=inv_number,
                    condition=condition,
                    location=location,
                    purchase_date=purchase_date,
                    purchase_price=purchase_price,
                    status='available',
                    created_by=request.user
                )
                created_items.append(item)
                
                # Registrar movimentação
                StockMovement.objects.create(
                    inventory_item=item,
                    movement_type='entry',
                    reason='initial_stock',
                    to_location=location,
                    notes='Cadastro em lote',
                    created_by=request.user
                )
            
            if created_items:
                messages.success(request, f'{len(created_items)} itens cadastrados com sucesso!')
            if errors:
                messages.warning(request, f'{len(errors)} erros: {", ".join(errors[:5])}')
            
            return redirect('assets:product_detail', pk=product.pk)
    else:
        form = BulkInventoryItemForm()
    
    context = {
        'form': form,
        'title': 'Cadastro em Lote',
    }
    return render(request, 'assets/inventory/item_bulk_form.html', context)


# ============================================================================
# MOVIMENTAÇÕES DE ESTOQUE
# ============================================================================

@login_required
@inventory_permission_required('can_register_entries')
def stock_entry(request):
    """Registrar entrada de estoque"""
    if request.method == 'POST':
        form = StockEntryForm(request.POST)
        if form.is_valid():
            movement = form.save(commit=False)
            movement.created_by = request.user
            movement.save()
            
            # Atualizar status do item para disponível
            item = movement.inventory_item
            item.status = 'available'
            item.assigned_to = None
            item.assigned_sector = None
            item.assigned_date = None
            item.save()
            
            messages.success(request, f'Entrada registrada para o item {item.inventory_number}!')
            return redirect('assets:item_detail', pk=item.pk)
    else:
        item_id = request.GET.get('item')
        initial = {}
        if item_id:
            initial['inventory_item'] = item_id
        form = StockEntryForm(initial=initial)
    
    context = {
        'form': form,
        'title': 'Registrar Entrada',
    }
    return render(request, 'assets/inventory/stock_entry_form.html', context)


@login_required
@inventory_permission_required('can_register_exits')
def stock_exit(request):
    """Registrar saída de estoque"""
    if request.method == 'POST':
        form = StockExitForm(request.POST)
        if form.is_valid():
            movement = form.save(commit=False)
            movement.created_by = request.user
            movement.save()  # form.save já atualiza o status do item
            
            messages.success(request, f'Saída registrada para o item {movement.inventory_item.inventory_number}!')
            return redirect('assets:item_detail', pk=movement.inventory_item.pk)
    else:
        item_id = request.GET.get('item')
        initial = {}
        if item_id:
            initial['inventory_item'] = item_id
        form = StockExitForm(initial=initial)
    
    context = {
        'form': form,
        'title': 'Registrar Saída',
    }
    return render(request, 'assets/inventory/stock_exit_form.html', context)


@login_required
def movement_list(request):
    """Lista todas as movimentações"""
    query = request.GET.get('q', '')
    type_filter = request.GET.get('type', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    
    movements = StockMovement.objects.select_related(
        'inventory_item', 'inventory_item__product', 
        'created_by', 'to_user', 'from_user', 'to_sector', 'from_sector'
    )
    
    if query:
        movements = movements.filter(
            Q(inventory_item__inventory_number__icontains=query) |
            Q(inventory_item__product__name__icontains=query) |
            Q(document_reference__icontains=query)
        )
    
    if type_filter:
        movements = movements.filter(movement_type=type_filter)
    
    if date_from:
        movements = movements.filter(created_at__date__gte=date_from)
    
    if date_to:
        movements = movements.filter(created_at__date__lte=date_to)
    
    movements = movements.order_by('-created_at')
    
    # Paginação
    paginator = Paginator(movements, 25)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'query': query,
        'type_filter': type_filter,
        'date_from': date_from,
        'date_to': date_to,
        'movement_types': StockMovement.MOVEMENT_TYPE_CHOICES,
        'total_movements': movements.count(),
    }
    
    return render(request, 'assets/inventory/movement_list.html', context)


# ============================================================================
# RELATÓRIOS
# ============================================================================

@login_required
@inventory_permission_required('can_view_reports')
def report_dashboard(request):
    """Dashboard de relatórios"""
    context = {
        'total_products': Product.objects.filter(is_active=True).count(),
        'total_items': InventoryItem.objects.count(),
        'total_movements': StockMovement.objects.count(),
    }
    return render(request, 'assets/inventory/report_dashboard.html', context)


@login_required
@inventory_permission_required('can_view_reports')
def report_stock(request):
    """Relatório de estoque atual"""
    products = Product.objects.filter(is_active=True).select_related('category').annotate(
        available_count=Count('inventory_items', filter=Q(inventory_items__status='available')),
        in_use_count=Count('inventory_items', filter=Q(inventory_items__status='in_use')),
        maintenance_count=Count('inventory_items', filter=Q(inventory_items__status='maintenance')),
        total_count=Count('inventory_items')
    ).order_by('category__name', 'name')
    
    context = {
        'products': products,
    }
    return render(request, 'assets/inventory/report_stock.html', context)


@login_required
@inventory_permission_required('can_view_reports')
def report_entries(request):
    """Relatório de entradas"""
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    
    entries = StockMovement.objects.filter(
        movement_type='entry'
    ).select_related(
        'inventory_item', 'inventory_item__product', 
        'created_by', 'from_user', 'from_sector'
    )
    
    if date_from:
        entries = entries.filter(created_at__date__gte=date_from)
    if date_to:
        entries = entries.filter(created_at__date__lte=date_to)
    
    entries = entries.order_by('-created_at')
    
    # Paginação
    paginator = Paginator(entries, 25)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'date_from': date_from,
        'date_to': date_to,
        'total_entries': entries.count(),
    }
    return render(request, 'assets/inventory/report_entries.html', context)


@login_required
@inventory_permission_required('can_view_reports')
def report_exits(request):
    """Relatório de saídas"""
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    
    exits = StockMovement.objects.filter(
        movement_type='exit'
    ).select_related(
        'inventory_item', 'inventory_item__product', 
        'created_by', 'to_user', 'to_sector'
    )
    
    if date_from:
        exits = exits.filter(created_at__date__gte=date_from)
    if date_to:
        exits = exits.filter(created_at__date__lte=date_to)
    
    exits = exits.order_by('-created_at')
    
    # Paginação
    paginator = Paginator(exits, 25)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'date_from': date_from,
        'date_to': date_to,
        'total_exits': exits.count(),
    }
    return render(request, 'assets/inventory/report_exits.html', context)


@login_required
@inventory_permission_required('can_view_reports')
def export_stock_excel(request):
    """Exportar estoque atual para Excel"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Estoque Atual"
    
    headers = [
        'Código', 'Produto', 'Categoria', 'Nº Inventário', 'Status', 
        'Condição', 'Responsável', 'Setor', 'Localização', 'Data Compra', 'Valor'
    ]
    
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
    
    items = InventoryItem.objects.select_related(
        'product', 'product__category', 'assigned_to', 'assigned_sector'
    ).order_by('product__name', 'inventory_number')
    
    for row, item in enumerate(items, 2):
        ws.cell(row=row, column=1, value=item.product.sku)
        ws.cell(row=row, column=2, value=item.product.name)
        ws.cell(row=row, column=3, value=item.product.category.name if item.product.category else '')
        ws.cell(row=row, column=4, value=item.inventory_number)
        ws.cell(row=row, column=5, value=item.get_status_display())
        ws.cell(row=row, column=6, value=item.get_condition_display())
        ws.cell(row=row, column=7, value=item.assigned_to.get_full_name() if item.assigned_to else '')
        ws.cell(row=row, column=8, value=item.assigned_sector.name if item.assigned_sector else '')
        ws.cell(row=row, column=9, value=item.location)
        ws.cell(row=row, column=10, value=item.purchase_date.strftime("%d/%m/%Y") if item.purchase_date else '')
        ws.cell(row=row, column=11, value=float(item.purchase_price) if item.purchase_price else 0)
    
    # Ajustar largura
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 15
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="estoque_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    wb.save(response)
    
    return response


@login_required
@inventory_permission_required('can_view_reports')
def export_movements_excel(request):
    """Exportar movimentações para Excel"""
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    movement_type = request.GET.get('type', '')
    
    movements = StockMovement.objects.select_related(
        'inventory_item', 'inventory_item__product', 
        'created_by', 'to_user', 'from_user', 'to_sector', 'from_sector'
    )
    
    if movement_type:
        movements = movements.filter(movement_type=movement_type)
    if date_from:
        movements = movements.filter(created_at__date__gte=date_from)
    if date_to:
        movements = movements.filter(created_at__date__lte=date_to)
    
    movements = movements.order_by('-created_at')
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Movimentações"
    
    headers = [
        'Data/Hora', 'Tipo', 'Motivo', 'Nº Inventário', 'Produto',
        'De (Usuário)', 'Para (Usuário)', 'De (Setor)', 'Para (Setor)',
        'Observações', 'Registrado por'
    ]
    
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
    
    for row, mov in enumerate(movements, 2):
        ws.cell(row=row, column=1, value=mov.created_at.strftime("%d/%m/%Y %H:%M"))
        ws.cell(row=row, column=2, value=mov.get_movement_type_display())
        ws.cell(row=row, column=3, value=mov.get_reason_display())
        ws.cell(row=row, column=4, value=mov.inventory_item.inventory_number)
        ws.cell(row=row, column=5, value=mov.inventory_item.product.name)
        ws.cell(row=row, column=6, value=mov.from_user.get_full_name() if mov.from_user else '')
        ws.cell(row=row, column=7, value=mov.to_user.get_full_name() if mov.to_user else '')
        ws.cell(row=row, column=8, value=mov.from_sector.name if mov.from_sector else '')
        ws.cell(row=row, column=9, value=mov.to_sector.name if mov.to_sector else '')
        ws.cell(row=row, column=10, value=mov.notes)
        ws.cell(row=row, column=11, value=mov.created_by.get_full_name() if mov.created_by else '')
    
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="movimentacoes_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    wb.save(response)
    
    return response


# ============================================================================
# GESTORES DE INVENTÁRIO
# ============================================================================

@login_required
def manager_list(request):
    """Lista gestores de inventário"""
    # Apenas superadmin e admin podem ver esta página
    if request.user.hierarchy not in ['SUPERADMIN', 'ADMIN']:
        # Ou gestores com permissão
        manager = _get_inventory_manager(request.user)
        if not manager or not manager.can_manage_managers:
            messages.error(request, 'Você não tem permissão para acessar esta área.')
            return redirect('assets:inventory_dashboard')
    
    managers = InventoryManager.objects.select_related('user', 'created_by').order_by(
        '-is_active', 'user__first_name'
    )
    
    context = {
        'managers': managers,
    }
    return render(request, 'assets/inventory/manager_list.html', context)


@login_required
def manager_create(request):
    """Criar novo gestor de inventário"""
    if request.user.hierarchy not in ['SUPERADMIN', 'ADMIN']:
        manager = _get_inventory_manager(request.user)
        if not manager or not manager.can_manage_managers:
            messages.error(request, 'Você não tem permissão para esta ação.')
            return redirect('assets:inventory_dashboard')
    
    if request.method == 'POST':
        form = InventoryManagerForm(request.POST)
        if form.is_valid():
            manager = form.save(commit=False)
            manager.created_by = request.user
            manager.save()
            messages.success(request, f'Gestor "{manager.user.get_full_name()}" adicionado com sucesso!')
            return redirect('assets:manager_list')
    else:
        form = InventoryManagerForm()
    
    context = {
        'form': form,
        'title': 'Novo Gestor de Inventário',
    }
    return render(request, 'assets/inventory/manager_form.html', context)


@login_required
def manager_edit(request, pk):
    """Editar gestor de inventário"""
    if request.user.hierarchy not in ['SUPERADMIN', 'ADMIN']:
        manager = _get_inventory_manager(request.user)
        if not manager or not manager.can_manage_managers:
            messages.error(request, 'Você não tem permissão para esta ação.')
            return redirect('assets:inventory_dashboard')
    
    manager = get_object_or_404(InventoryManager, pk=pk)
    
    if request.method == 'POST':
        form = InventoryManagerForm(request.POST, instance=manager)
        if form.is_valid():
            form.save()
            messages.success(request, f'Permissões de "{manager.user.get_full_name()}" atualizadas!')
            return redirect('assets:manager_list')
    else:
        form = InventoryManagerForm(instance=manager)
    
    context = {
        'form': form,
        'manager': manager,
        'title': f'Editar Gestor - {manager.user.get_full_name()}',
    }
    return render(request, 'assets/inventory/manager_form.html', context)


@login_required
def manager_delete(request, pk):
    """Remover gestor de inventário"""
    if request.user.hierarchy not in ['SUPERADMIN', 'ADMIN']:
        manager = _get_inventory_manager(request.user)
        if not manager or not manager.can_manage_managers:
            messages.error(request, 'Você não tem permissão para esta ação.')
            return redirect('assets:inventory_dashboard')
    
    manager = get_object_or_404(InventoryManager, pk=pk)
    
    if request.method == 'POST':
        name = manager.user.get_full_name()
        manager.delete()
        messages.success(request, f'Gestor "{name}" removido com sucesso!')
        return redirect('assets:manager_list')
    
    context = {
        'manager': manager,
        'title': f'Remover Gestor - {manager.user.get_full_name()}',
    }
    return render(request, 'assets/inventory/manager_delete.html', context)


@login_required
@require_POST
def manager_toggle(request, pk):
    """Ativar/desativar gestor de inventário"""
    if request.user.hierarchy not in ['SUPERADMIN', 'ADMIN']:
        manager = _get_inventory_manager(request.user)
        if not manager or not manager.can_manage_managers:
            return JsonResponse({'error': 'Sem permissão'}, status=403)
    
    manager = get_object_or_404(InventoryManager, pk=pk)
    manager.is_active = not manager.is_active
    manager.save()
    
    status = 'ativado' if manager.is_active else 'desativado'
    messages.success(request, f'Gestor "{manager.user.get_full_name()}" {status}!')
    return redirect('assets:manager_list')


# ============================================================================
# SOLICITAÇÕES DE ITENS
# ============================================================================

def can_approve_requests(user):
    """Verifica se o usuário pode aprovar/reprovar solicitações"""
    if user.hierarchy in ['SUPERADMIN', 'ADMIN']:
        return True
    manager = _get_inventory_manager(user)
    if manager:
        return manager.is_active and manager.can_approve_requests
    return False


@login_required
def item_request_create(request):
    """Criar nova solicitação de item do almoxarifado (múltiplos itens)"""
    products = Product.objects.filter(is_active=True).order_by('name')
    
    if request.method == 'POST':
        form = ItemRequestForm(request.POST)
        
        # Extrair itens do POST (enviados via JS)
        item_products = request.POST.getlist('item_product')
        item_quantities = request.POST.getlist('item_quantity')
        
        items_valid = True
        items_data = []
        errors = []
        
        if not item_products or len(item_products) == 0:
            items_valid = False
            errors.append('Adicione pelo menos um item à solicitação.')
        else:
            for i, (prod_id, qty) in enumerate(zip(item_products, item_quantities)):
                try:
                    product = Product.objects.get(pk=prod_id, is_active=True)
                    quantity = int(qty)
                    if quantity < 1:
                        raise ValueError
                    if product.current_stock < quantity:
                        errors.append(f'Estoque insuficiente para "{product.name}". Disponível: {product.current_stock}.')
                        items_valid = False
                    items_data.append({'product': product, 'quantity': quantity})
                except (Product.DoesNotExist, ValueError, TypeError):
                    errors.append(f'Item {i+1}: produto ou quantidade inválidos.')
                    items_valid = False
            
            # Verificar produtos duplicados
            product_ids = [d['product'].pk for d in items_data]
            if len(product_ids) != len(set(product_ids)):
                errors.append('Não é possível adicionar o mesmo produto mais de uma vez. Ajuste a quantidade.')
                items_valid = False
        
        if form.is_valid() and items_valid:
            item_request = form.save(commit=False)
            item_request.requested_by = request.user
            item_request.save()
            
            for item_data in items_data:
                ItemRequestItem.objects.create(
                    request=item_request,
                    product=item_data['product'],
                    quantity=item_data['quantity']
                )
            
            messages.success(request, 'Solicitação enviada com sucesso! Aguarde a aprovação.')
            return redirect('assets:item_request_list')
        else:
            for error in errors:
                messages.error(request, error)
    else:
        form = ItemRequestForm()
    
    return render(request, 'assets/inventory/item_request_form.html', {
        'form': form,
        'title': 'Solicitar Itens do Almoxarifado',
        'products': products,
    })


@login_required
def item_request_list(request):
    """Lista solicitações de itens"""
    user = request.user
    is_approver = can_approve_requests(user)
    
    # Filtros
    status_filter = request.GET.get('status', '')
    view_mode = request.GET.get('view', 'mine')  # mine ou all
    
    if is_approver and view_mode == 'all':
        requests_qs = ItemRequest.objects.all()
    else:
        requests_qs = ItemRequest.objects.filter(requested_by=user)
        view_mode = 'mine'
    
    if status_filter:
        requests_qs = requests_qs.filter(status=status_filter)
    
    requests_qs = requests_qs.select_related(
        'requested_by', 'reviewed_by', 'delivered_by'
    ).prefetch_related(
        'items', 'items__product', 'items__product__category'
    ).order_by('-requested_at')
    
    # Contadores
    pending_count = ItemRequest.objects.filter(status='pending').count()
    
    paginator = Paginator(requests_qs, 20)
    page_obj = paginator.get_page(request.GET.get('page'))
    
    context = {
        'page_obj': page_obj,
        'status_filter': status_filter,
        'view_mode': view_mode,
        'is_approver': is_approver,
        'pending_count': pending_count,
        'status_choices': ItemRequest.STATUS_CHOICES,
        'total_requests': requests_qs.count(),
    }
    return render(request, 'assets/inventory/item_request_list.html', context)


@login_required
def item_request_detail(request, pk):
    """Detalhes de uma solicitação"""
    item_request = get_object_or_404(
        ItemRequest.objects.select_related(
            'requested_by', 'reviewed_by', 'delivered_by'
        ).prefetch_related(
            'items', 'items__product', 'items__product__category'
        ), pk=pk
    )
    
    user = request.user
    is_approver = can_approve_requests(user)
    
    # Apenas o solicitante ou gestores podem ver
    if item_request.requested_by != user and not is_approver:
        messages.error(request, 'Você não tem permissão para ver esta solicitação.')
        return redirect('assets:item_request_list')
    
    context = {
        'item_request': item_request,
        'is_approver': is_approver,
    }
    return render(request, 'assets/inventory/item_request_detail.html', context)


@login_required
def item_request_approve(request, pk):
    """Aprovar uma solicitação"""
    if not can_approve_requests(request.user):
        messages.error(request, 'Você não tem permissão para aprovar solicitações.')
        return redirect('assets:item_request_list')
    
    item_request = get_object_or_404(ItemRequest, pk=pk)
    
    if not item_request.can_approve:
        messages.error(request, 'Esta solicitação não pode ser aprovada.')
        return redirect('assets:item_request_detail', pk=pk)
    
    if request.method == 'POST':
        form = ItemRequestReviewForm(request.POST)
        if form.is_valid():
            item_request.status = 'approved'
            item_request.reviewed_by = request.user
            item_request.reviewed_at = timezone.now()
            item_request.review_notes = form.cleaned_data.get('review_notes', '')
            item_request.save()
            messages.success(request, f'Solicitação #{item_request.pk} aprovada com sucesso!')
            return redirect('assets:item_request_detail', pk=pk)
    
    return redirect('assets:item_request_detail', pk=pk)


@login_required
def item_request_reject(request, pk):
    """Reprovar uma solicitação"""
    if not can_approve_requests(request.user):
        messages.error(request, 'Você não tem permissão para reprovar solicitações.')
        return redirect('assets:item_request_list')
    
    item_request = get_object_or_404(ItemRequest, pk=pk)
    
    if not item_request.can_approve:
        messages.error(request, 'Esta solicitação não pode ser reprovada.')
        return redirect('assets:item_request_detail', pk=pk)
    
    if request.method == 'POST':
        form = ItemRequestReviewForm(request.POST)
        if form.is_valid():
            item_request.status = 'rejected'
            item_request.reviewed_by = request.user
            item_request.reviewed_at = timezone.now()
            item_request.review_notes = form.cleaned_data.get('review_notes', '')
            item_request.save()
            messages.success(request, f'Solicitação #{item_request.pk} rejeitada.')
            return redirect('assets:item_request_detail', pk=pk)
    
    return redirect('assets:item_request_detail', pk=pk)


@login_required
def item_request_deliver(request, pk):
    """Marcar solicitação como entregue e remover itens do estoque"""
    if not can_approve_requests(request.user):
        messages.error(request, 'Você não tem permissão para registrar entregas.')
        return redirect('assets:item_request_list')
    
    item_request = get_object_or_404(ItemRequest, pk=pk)
    
    if not item_request.can_deliver:
        messages.error(request, 'Esta solicitação não pode ser entregue (precisa estar aprovada).')
        return redirect('assets:item_request_detail', pk=pk)
    
    if request.method == 'POST':
        form = ItemRequestDeliveryForm(request.POST)
        if form.is_valid():
            # Verificar estoque de cada item da solicitação
            request_items = item_request.items.select_related('product').all()
            stock_errors = []
            
            for req_item in request_items:
                available_count = InventoryItem.objects.filter(
                    product=req_item.product, status='available'
                ).count()
                if available_count < req_item.quantity:
                    stock_errors.append(
                        f'"{req_item.product.name}": disponível {available_count}, solicitado {req_item.quantity}'
                    )
            
            if stock_errors:
                messages.error(
                    request,
                    'Estoque insuficiente para: ' + '; '.join(stock_errors)
                )
                return redirect('assets:item_request_detail', pk=pk)
            
            # Registrar saída de cada item
            total_delivered = 0
            for req_item in request_items:
                qty = req_item.effective_quantity
                if qty == 0:
                    continue  # Item removido na contraproposta
                available_items = InventoryItem.objects.filter(
                    product=req_item.product, status='available'
                ).order_by('created_at')[:qty]
                
                for inv_item in available_items:
                    inv_item.status = 'in_use'
                    inv_item.assigned_to = item_request.requested_by
                    inv_item.assigned_date = timezone.now()
                    inv_item.save()
                    
                    StockMovement.objects.create(
                        inventory_item=inv_item,
                        movement_type='exit',
                        reason='assigned_to_user',
                        to_user=item_request.requested_by,
                        notes=f'Solicitação #{item_request.pk}: {item_request.reason}',
                        created_by=request.user
                    )
                total_delivered += qty
            
            # Atualizar a solicitação
            item_request.status = 'delivered'
            item_request.delivered_by = request.user
            item_request.delivered_at = timezone.now()
            item_request.delivery_notes = form.cleaned_data.get('delivery_notes', '')
            item_request.save()
            
            messages.success(
                request, 
                f'Solicitação #{item_request.pk} entregue com sucesso! '
                f'{total_delivered} item(ns) removido(s) do estoque.'
            )
            return redirect('assets:item_request_detail', pk=pk)
    
    return redirect('assets:item_request_detail', pk=pk)


@login_required
def item_request_cancel(request, pk):
    """Cancelar uma solicitação"""
    item_request = get_object_or_404(ItemRequest, pk=pk)
    
    # Apenas o solicitante ou gestores podem cancelar
    if item_request.requested_by != request.user and not can_approve_requests(request.user):
        messages.error(request, 'Você não tem permissão para cancelar esta solicitação.')
        return redirect('assets:item_request_list')
    
    if not item_request.can_cancel:
        messages.error(request, 'Esta solicitação não pode ser cancelada.')
        return redirect('assets:item_request_detail', pk=pk)
    
    if request.method == 'POST':
        item_request.status = 'cancelled'
        item_request.save()
        messages.success(request, f'Solicitação #{item_request.pk} cancelada.')
        return redirect('assets:item_request_list')
    
    return redirect('assets:item_request_detail', pk=pk)


# ============================================================================
# CATÁLOGO / MERCADINHO DO ALMOXARIFADO
# ============================================================================

@login_required
def store_catalog(request):
    """Visão de catálogo estilo mercadinho para solicitar itens"""
    categories = InventoryCategory.objects.filter(is_active=True).annotate(
        product_count=Count('products', filter=Q(products__is_active=True))
    ).order_by('name')
    
    # Filtros
    category_filter = request.GET.get('category', '')
    search_query = request.GET.get('q', '')
    
    products = Product.objects.filter(is_active=True).select_related('category').prefetch_related('media')
    
    if category_filter:
        products = products.filter(category_id=category_filter)
    
    if search_query:
        products = products.filter(
            Q(name__icontains=search_query) | 
            Q(description__icontains=search_query) |
            Q(sku__icontains=search_query) |
            Q(brand__icontains=search_query)
        )
    
    products = products.order_by('name')
    
    # Paginação
    paginator = Paginator(products, 12)
    page_obj = paginator.get_page(request.GET.get('page'))
    
    # Minhas solicitações recentes
    my_recent_requests = ItemRequest.objects.filter(
        requested_by=request.user
    ).order_by('-requested_at')[:5]
    
    # Contagem de solicitações pendentes do usuário
    my_pending_count = ItemRequest.objects.filter(
        requested_by=request.user, 
        status__in=['pending', 'counterproposal']
    ).count()
    
    context = {
        'categories': categories,
        'page_obj': page_obj,
        'category_filter': category_filter,
        'search_query': search_query,
        'my_recent_requests': my_recent_requests,
        'my_pending_count': my_pending_count,
    }
    return render(request, 'assets/inventory/store_catalog.html', context)


# ============================================================================
# CONTRAPROPOSTA
# ============================================================================

@login_required
def item_request_counterproposal(request, pk):
    """Gestor faz contraproposta de quantidades"""
    if not can_approve_requests(request.user):
        messages.error(request, 'Você não tem permissão para fazer contrapropostas.')
        return redirect('assets:item_request_list')
    
    item_request = get_object_or_404(
        ItemRequest.objects.select_related(
            'requested_by', 'reviewed_by'
        ).prefetch_related(
            'items', 'items__product'
        ), pk=pk
    )
    
    if not item_request.can_counterpropose:
        messages.error(request, 'Esta solicitação não pode receber contraproposta.')
        return redirect('assets:item_request_detail', pk=pk)
    
    if request.method == 'POST':
        form = ItemRequestCounterProposalForm(request.POST)
        if form.is_valid():
            # Processar quantidades propostas para cada item
            request_items = item_request.items.all()
            has_changes = False
            
            for req_item in request_items:
                proposed_qty_key = f'proposed_qty_{req_item.pk}'
                proposed_qty = request.POST.get(proposed_qty_key)
                
                if proposed_qty is not None:
                    try:
                        proposed_qty = int(proposed_qty)
                        if proposed_qty < 0:
                            proposed_qty = 0
                        req_item.proposed_quantity = proposed_qty
                        req_item.save()
                        if proposed_qty != req_item.quantity:
                            has_changes = True
                    except (ValueError, TypeError):
                        pass
            
            if not has_changes:
                messages.warning(request, 'Nenhuma quantidade foi alterada. Use "Aprovar" para manter as quantidades originais.')
                return redirect('assets:item_request_detail', pk=pk)
            
            # Atualizar status da solicitação
            item_request.status = 'counterproposal'
            item_request.counterproposal_by = request.user
            item_request.counterproposal_at = timezone.now()
            item_request.counterproposal_notes = form.cleaned_data['counterproposal_notes']
            item_request.save()
            
            messages.success(request, f'Contraproposta enviada para solicitação #{item_request.pk}.')
            return redirect('assets:item_request_detail', pk=pk)
    else:
        form = ItemRequestCounterProposalForm()
    
    context = {
        'item_request': item_request,
        'form': form,
    }
    return render(request, 'assets/inventory/item_request_counterproposal.html', context)


@login_required
def item_request_accept_counterproposal(request, pk):
    """Solicitante aceita a contraproposta"""
    item_request = get_object_or_404(ItemRequest, pk=pk)
    
    if item_request.requested_by != request.user:
        messages.error(request, 'Apenas o solicitante pode responder à contraproposta.')
        return redirect('assets:item_request_list')
    
    if not item_request.can_respond_counterproposal:
        messages.error(request, 'Esta solicitação não tem contraproposta pendente.')
        return redirect('assets:item_request_detail', pk=pk)
    
    if request.method == 'POST':
        form = ItemRequestCounterProposalResponseForm(request.POST)
        if form.is_valid():
            item_request.status = 'accepted'
            item_request.counterproposal_response_notes = form.cleaned_data.get('response_notes', '')
            item_request.counterproposal_responded_at = timezone.now()
            item_request.save()
            
            messages.success(request, f'Você aceitou a contraproposta da solicitação #{item_request.pk}.')
            return redirect('assets:item_request_detail', pk=pk)
    
    return redirect('assets:item_request_detail', pk=pk)


@login_required
def item_request_reject_counterproposal(request, pk):
    """Solicitante recusa a contraproposta"""
    item_request = get_object_or_404(ItemRequest, pk=pk)
    
    if item_request.requested_by != request.user:
        messages.error(request, 'Apenas o solicitante pode responder à contraproposta.')
        return redirect('assets:item_request_list')
    
    if not item_request.can_respond_counterproposal:
        messages.error(request, 'Esta solicitação não tem contraproposta pendente.')
        return redirect('assets:item_request_detail', pk=pk)
    
    if request.method == 'POST':
        form = ItemRequestCounterProposalResponseForm(request.POST)
        if form.is_valid():
            item_request.status = 'cancelled'
            item_request.counterproposal_response_notes = form.cleaned_data.get('response_notes', '')
            item_request.counterproposal_responded_at = timezone.now()
            item_request.save()
            
            messages.success(request, f'Você recusou a contraproposta. A solicitação #{item_request.pk} foi cancelada.')
            return redirect('assets:item_request_detail', pk=pk)
    
    return redirect('assets:item_request_detail', pk=pk)


# ============================================================================
# VIEWS LEGADAS - MANTIDAS PARA COMPATIBILIDADE
# ============================================================================

@login_required
def asset_list(request):
    """Lista todos os ativos com funcionalidade de busca e paginação"""
    query = request.GET.get('q', '')
    estado_filter = request.GET.get('estado', '')
    setor_filter = request.GET.get('setor', '')
    
    assets = Asset.objects.all()
    
    # Filtros
    if query:
        assets = assets.filter(
            Q(patrimonio_numero__icontains=query) |
            Q(nome__icontains=query) |
            Q(localizado__icontains=query) |
            Q(setor__icontains=query) |
            Q(pdv__icontains=query)
        )
    
    if estado_filter:
        assets = assets.filter(estado_fisico=estado_filter)
    
    if setor_filter:
        assets = assets.filter(setor__icontains=setor_filter)
    
    # Paginação
    paginator = Paginator(assets, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Para os filtros
    estados = Asset.ESTADO_FISICO_CHOICES
    setores = Asset.objects.values_list('setor', flat=True).distinct().order_by('setor')
    
    context = {
        'page_obj': page_obj,
        'query': query,
        'estado_filter': estado_filter,
        'setor_filter': setor_filter,
        'estados': estados,
        'setores': setores,
        'total_assets': assets.count(),
    }
    
    return render(request, 'assets/list.html', context)


@login_required
def asset_detail(request, pk):
    """Exibe detalhes de um ativo específico"""
    asset = get_object_or_404(Asset, pk=pk)
    
    context = {
        'asset': asset,
    }
    
    return render(request, 'assets/detail.html', context)


@login_required
def asset_create(request):
    """Cria um novo ativo"""
    if request.method == 'POST':
        form = AssetForm(request.POST, request.FILES)
        if form.is_valid():
            asset = form.save(commit=False)
            asset.created_by = request.user
            asset.save()
            messages.success(request, f'Ativo {asset.patrimonio_numero} criado com sucesso!')
            return redirect('assets:detail', pk=asset.pk)
    else:
        form = AssetForm()
    
    context = {
        'form': form,
        'title': 'Cadastrar Novo Ativo',
    }
    
    return render(request, 'assets/form.html', context)


@login_required
def asset_edit(request, pk):
    """Edita um ativo existente"""
    asset = get_object_or_404(Asset, pk=pk)
    
    if request.method == 'POST':
        form = AssetForm(request.POST, request.FILES, instance=asset)
        if form.is_valid():
            form.save()
            messages.success(request, f'Ativo {asset.patrimonio_numero} atualizado com sucesso!')
            return redirect('assets:detail', pk=asset.pk)
    else:
        form = AssetForm(instance=asset)
    
    context = {
        'form': form,
        'asset': asset,
        'title': f'Editar Ativo - {asset.patrimonio_numero}',
    }
    
    return render(request, 'assets/form.html', context)


@login_required
def asset_delete(request, pk):
    """Deleta um ativo"""
    asset = get_object_or_404(Asset, pk=pk)
    
    if request.method == 'POST':
        patrimonio_numero = asset.patrimonio_numero
        asset.delete()
        messages.success(request, f'Ativo {patrimonio_numero} removido com sucesso!')
        return redirect('assets:list')
    
    context = {
        'asset': asset,
    }
    
    return render(request, 'assets/delete.html', context)


@login_required
def export_assets_excel(request):
    """Exportar dados de ativos em Excel (exceto fotos)"""
    # Criar workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ativos"
    
    # Definir cabeçalhos
    headers = [
        'Nº Patrimônio', 'Nome', 'Localizado', 'Setor', 'PDV', 
        'Estado Físico', 'Observações', 'Criado por', 'Data Criação', 'Última Atualização'
    ]
    
    # Estilizar cabeçalhos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Buscar dados dos ativos
    assets = Asset.objects.all().select_related('created_by').order_by('patrimonio_numero')
    
    # Preencher dados
    for row, asset in enumerate(assets, 2):
        ws.cell(row=row, column=1, value=asset.patrimonio_numero)
        ws.cell(row=row, column=2, value=asset.nome)
        ws.cell(row=row, column=3, value=asset.localizado)
        ws.cell(row=row, column=4, value=asset.setor)
        ws.cell(row=row, column=5, value=asset.pdv)
        ws.cell(row=row, column=6, value=asset.get_estado_fisico_display())
        ws.cell(row=row, column=7, value=asset.observacoes or "")
        ws.cell(row=row, column=8, value=asset.created_by.get_full_name() if asset.created_by else "")
        ws.cell(row=row, column=9, value=asset.created_at.strftime("%Y-%m-%d %H:%M:%S"))
        ws.cell(row=row, column=10, value=asset.updated_at.strftime("%Y-%m-%d %H:%M:%S"))
    
    # Ajustar largura das colunas
    for col in range(1, len(headers) + 1):
        column_letter = get_column_letter(col)
        max_length = 0
        for row in ws[column_letter]:
            try:
                if len(str(row.value)) > max_length:
                    max_length = len(str(row.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Preparar response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="ativos_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    
    # Salvar workbook na response
    wb.save(response)
    
    # Log da ação
    log_action(
        request.user,
        'ASSET_EXPORT',
        f'Exportação de dados de ativos realizada',
        request
    )
    
    return response


@login_required
def import_assets_excel(request):
    """Importar ativos de Excel"""
    if request.method == 'POST':
        if 'excel_file' not in request.FILES:
            messages.error(request, 'Nenhum arquivo foi enviado.')
            return redirect('assets:list')
        
        excel_file = request.FILES['excel_file']
        
        try:
            # Ler o arquivo Excel
            wb = openpyxl.load_workbook(excel_file)
            ws = wb.active
            
            created_count = 0
            updated_count = 0
            error_count = 0
            errors = []
            
            # Processar cada linha (pular cabeçalho)
            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
                try:
                    patrimonio_numero, nome, localizado, setor, pdv, estado_fisico, observacoes, created_by, date_created, last_updated = row
                    
                    if not patrimonio_numero:  # Nº Patrimônio é obrigatório
                        continue
                    
                    # Converter estado físico de display para valor
                    estado_fisico_value = estado_fisico
                    if estado_fisico:
                        for value, display in Asset.ESTADO_FISICO_CHOICES:
                            if display.lower() == estado_fisico.lower():
                                estado_fisico_value = value
                                break
                    
                    # Verificar se ativo já existe pelo número de patrimônio
                    asset, created = Asset.objects.get_or_create(
                        patrimonio_numero=patrimonio_numero,
                        defaults={
                            'nome': nome or '',
                            'localizado': localizado or '',
                            'setor': setor or '',
                            'pdv': pdv or '',
                            'estado_fisico': estado_fisico_value or 'bom',
                            'observacoes': observacoes or '',
                            'created_by': request.user,  # Sempre o usuário que está importando
                        }
                    )
                    
                    if created:
                        created_count += 1
                    else:
                        # Atualizar ativo existente
                        if nome:
                            asset.nome = nome
                        if localizado:
                            asset.localizado = localizado
                        if setor:
                            asset.setor = setor
                        if pdv:
                            asset.pdv = pdv
                        if estado_fisico_value:
                            asset.estado_fisico = estado_fisico_value
                        if observacoes is not None:
                            asset.observacoes = observacoes
                        
                        asset.save()
                        updated_count += 1
                        
                except Exception as e:
                    error_count += 1
                    errors.append(f'Linha {row_num}: {str(e)}')
                    continue
            
            # Mensagem de resultado
            if created_count > 0 or updated_count > 0:
                message = f'Importação concluída! {created_count} ativos criados, {updated_count} ativos atualizados.'
                if error_count > 0:
                    message += f' {error_count} erros encontrados.'
                messages.success(request, message)
            else:
                messages.warning(request, 'Nenhum ativo foi importado.')
            
            if errors:
                for error in errors[:5]:  # Mostrar apenas os primeiros 5 erros
                    messages.error(request, error)
            
            # Log da ação
            log_action(
                request.user,
                'ASSET_IMPORT',
                f'Importação de ativos: {created_count} criados, {updated_count} atualizados, {error_count} erros',
                request
            )
            
        except Exception as e:
            messages.error(request, f'Erro ao processar arquivo: {str(e)}')
    
    return redirect('assets:list')
